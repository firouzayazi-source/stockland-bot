"""
پارسر عمومی فایل‌های ورودی حسابداری/موجودی — TXT، CSV، Excel (.xlsx).

هدف: یک نقطهٔ واحد برای «تشخیص فرمت فایل + تبدیل به لیست ردیف‌ها» تا هم آپلود
موجودی محصول (feed) هم آپلود هزینه/درآمد از همین یه تابع استفاده کنن — به‌جای
پارسر جدا برای هر فرمت در هر مسیر. توسعه‌پذیر: افزودن فرمت تازه (مثلاً TSV)
فقط یعنی یه branch تازه در parse_uploaded_rows، نه بازنویسی هیچ مسیر مصرف‌کننده.
"""
from __future__ import annotations

import csv
import io


def _detect_kind(filename: str) -> str:
    name = (filename or "").lower()
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        return "xlsx"
    if name.endswith(".csv"):
        return "csv"
    return "txt"


def parse_uploaded_rows(filename: str, raw: bytes) -> list[dict]:
    """
    فایل رو بر اساس پسوند به لیست دیکشنری تبدیل می‌کنه:
    - .xlsx/.xlsm: با openpyxl، شیت اول، ردیف اول = هدر (نام ستون‌ها lower‌شده).
    - .csv: با csv.DictReader (هدر از ردیف اول، نام ستون‌ها lower‌شده).
    - بقیه (txt): بدون هدر — هر خط غیرخالی یک ردیف تک‌ستونه با کلید 'raw' برمی‌گرده
      (همون فرمت قدیمی موجودی: هر خط یک آیتم؛ عمداً حدسی برای «شاید CSV باشه» زده
      نمی‌شه، چون اولین خط ممکنه خودش کاما داشته باشه بدون این‌که هدر باشه).
    خروجی همیشه لیست dict با کلیدهای lower-case — مصرف‌کننده خودش matching
    اسم ستون (فارسی/انگلیسی، چندحالته) رو انجام می‌ده.
    """
    kind = _detect_kind(filename)

    if kind == "xlsx":
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration:
            return []
        headers = [str(h).strip().lower() if h is not None else f"col{i}" for i, h in enumerate(header)]
        out = []
        for row in rows_iter:
            if row is None or all(v is None or str(v).strip() == "" for v in row):
                continue
            d = {}
            for i, h in enumerate(headers):
                v = row[i] if i < len(row) else None
                d[h] = "" if v is None else str(v).strip()
            out.append(d)
        return out

    text = raw.decode("utf-8-sig", errors="ignore")

    if kind == "csv":
        reader = csv.DictReader(io.StringIO(text))
        out = []
        for row in reader:
            if not row or all((v or "").strip() == "" for v in row.values()):
                continue
            out.append({(k or "").strip().lower(): (v or "").strip() for k, v in row.items()})
        return out

    # txt — طبق قرارداد همیشگی پروژه، بدون هدر: هر خط خودش یه ردیف تک‌مقداره (کلید 'raw').
    # عمداً هیچ حدسی برای «شاید CSV با هدر باشه» زده نمی‌شه — چون فرمت خطی موجود (هر خط
    # یک آیتم، با کاماهای احتمالی به‌عنوان بخشی از خود مقدار) می‌تونه اولین خطش هم کاما
    # داشته باشه بدون اینکه هدر باشه؛ حدس اشتباه یعنی اولین ردیف واقعی بی‌صدا حذف بشه.
    lines = [ln for ln in text.splitlines() if ln.strip() and not ln.strip().startswith("#")]

    # فرمت خطی ساده — هر خط یک ردیف تک‌ستونه
    return [{"raw": ln.strip()} for ln in lines]


def parse_xlsx_labeled_items(raw: bytes) -> list[str]:
    """
    فرمت اختصاصی آپلود موجودی محصول از اکسل — با parse_uploaded_rows فرق داره:
    ردیف اول = عنوان ستون‌ها (برچسب هر فیلد)، هر ردیف بعدی = **یک محصول کامل** که
    اطلاعاتش به‌جای یک ستون، توی چند ستون (چپ به راست) پخش شده. هر ردیف به یک آیتم
    چندخطی «برچسب: مقدار» تبدیل می‌شه — دقیقاً همون فرمتی که ورودی متنی *** ازش
    پشتیبانی می‌کنه (هر آیتم می‌تونه چندخطی باشه)، طبق درخواست صریح مالک پروژه.

    برخلاف parse_uploaded_rows، این‌جا متن اصلی هدر (بدون lower/تغییر) و ترتیب دقیق
    ستون‌ها حفظ می‌شه — چون این متن مستقیم به خریدار تحویل داده می‌شه (کاربر داخل
    <code> می‌بینتش)، هر تغییر ظاهری‌ای (حروف بزرگ/کوچک، ترتیب) توش قابل مشاهده‌ست.

    ستون‌هایی که هدرشون خالیه، یا مقدارشون توی یه ردیف خاص خالیه، از اون خط رد می‌شن
    (نه این‌که با برچسب مبهم مثل «col3:» نشون داده بشن).
    """
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return []
    headers = [str(h).strip() if h is not None else "" for h in header_row]

    def cell_str(v):
        if v is None:
            return ""
        if isinstance(v, float) and v.is_integer():
            return str(int(v))
        return str(v).strip()

    items = []
    for row in rows_iter:
        if row is None or all(cell_str(v) == "" for v in row):
            continue
        lines = []
        for i, h in enumerate(headers):
            if not h:
                continue
            v = cell_str(row[i]) if i < len(row) else ""
            if v:
                lines.append(f"{h}: {v}")
        if lines:
            items.append("\n".join(lines))
    return items


def pick(row: dict, *keys: str, default: str = "") -> str:
    """اولین کلیدی که توی row مقدار غیرخالی داشت رو برمی‌گردونه — برای تطبیق نام
    ستون‌های مختلف (فارسی/انگلیسی) بدون نیاز به فرمت ثابت فایل ورودی."""
    for k in keys:
        v = row.get(k)
        if v not in (None, ""):
            return str(v).strip()
    return default
