"""
admin_panel.py — پنل مدیریت وب استوک لند (نسخه کامل)
────────────────────────────────────────────────────
ویژگی‌ها:
  - سیستم ادمین چندنفره با اختیارات مجزا
  - مدیریت کامل تنظیمات ربات
  - بکاپ / بازیابی / ریست دیتابیس
  - مدیریت ادمین‌ها از پنل وب
"""

# تقویم شمسی — در دسترس همه توابع پنل
try:
    from db import fa_date, fa_now  # noqa: F401
except Exception:
    def fa_date(d, with_time=False): return str(d or "—")[:16 if with_time else 10]
    def fa_now(with_time=True): import datetime; return datetime.datetime.now().strftime("%Y/%m/%d  %H:%M" if with_time else "%Y/%m/%d")

import hashlib
import hmac as _hmac
import html
import json
import os
import shutil
import sqlite3
import threading
import time
from datetime import datetime, timedelta
from urllib.parse import quote as _urlquote

import requests as _requests
from fastapi import APIRouter, BackgroundTasks, Form, Request, UploadFile
from fastapi.responses import FileResponse, HTMLResponse, RedirectResponse, JSONResponse
from starlette.concurrency import run_in_threadpool

router = APIRouter(prefix="/admin")

# ── migrations at startup ────────────────────────────────────────────────────
try:
    from db import ensure_product_support_schema, ensure_discount_table, ensure_subscription_table, ensure_referral_schema, ensure_user_extra_schema, ensure_partner_system_schema, ensure_partner_wallet_schema, ensure_admin_notes_schema, ensure_partner_tiers_extended
    ensure_product_support_schema()
    ensure_discount_table()
    ensure_subscription_table()
    ensure_referral_schema()
    ensure_user_extra_schema()
    ensure_partner_system_schema()
    ensure_partner_wallet_schema()
    ensure_admin_notes_schema()
    ensure_partner_tiers_extended()
except Exception:
    pass

# ─────────────────────────── Config ────────────────────────────────────────

def _env(k: str, default: str = "") -> str:
    return os.getenv(k) or default

# ⚠️ رفع‌شده (پاک‌سازی SQLite): except sqlite3.IntegrityError به‌تنهایی زیر
# Postgres هیچ‌وقت catch نمی‌شه (psycopg2.IntegrityError کلاس کاملاً جداست) —
# همون رفع db.py، اینجا هم لازم بود (بخش پاک‌سازی SQLite سند).
try:
    import psycopg2 as _psycopg2_for_errors
    _INTEGRITY_ERRORS = (sqlite3.IntegrityError, _psycopg2_for_errors.IntegrityError)
except ImportError:
    _INTEGRITY_ERRORS = (sqlite3.IntegrityError,)

def _db():
    """اتصال دیتابیس پنل.
    ⚠️ رفع‌شده (ممیزی کامل پروژه): قبلاً همیشه sqlite3.connect خام می‌زد، کاملاً
    مستقل از DB_DIALECT — یعنی روی سروری که واقعاً به Postgres مهاجرت کرده، کل
    پنل ادمین روی یک فایل SQLite فانتوم/جدا از دادهٔ واقعی کار می‌کرد. حالا از
    db_conn.get_connection() استفاده می‌کنه که بر اساس DB_DIALECT سوییچ می‌کنه؛
    روی SQLite دقیقاً همون PRAGMA های قبلی (WAL/busy_timeout/synchronous) رو
    داخل خودش تنظیم می‌کنه (db_conn._open_sqlite_connection)، بدون تغییر رفتار."""
    import db_conn
    return db_conn.get_connection(_env("DB_PATH"))

# ─────────────────────────── Permissions ───────────────────────────────────

# ─── رجیستری مرکزی دسترسی‌ها ─────────────────────────────────────────────
# قانون: هر قابلیت جدید = یک کلید اینجا؛ منو، صفحه ادمین‌ها و گاردها همه از همین می‌خوانند.
ALL_PERMISSIONS = {
    # ⚠️ «dashboard» عمداً هیچ‌جا با _require() گیت نمی‌شه — چون خود صفحهٔ اصلی
    # (/admin/) مقصد استاندارد ریدایرکت هر _require دیگه‌ای در کل پنله (پیام
    # «دسترسی کافی ندارید» دقیقاً همین‌جا نشون داده می‌شه)؛ گیت‌کردنش یعنی
    # ادمینی که این دسترسی رو نداره وارد یه حلقهٔ ریدایرکت بی‌نهایت به خودِ همین
    # صفحه می‌شه. کلید فقط برای کامل‌بودن لیست چک‌باکس‌ها نگه داشته شده.
    "dashboard":  "مشاهده داشبورد",
    "categories": "مدیریت دسته‌بندی‌ها",
    "products":   "مدیریت محصولات",
    "feed":       "مدیریت موجودی",
    "orders":     "مشاهده سفارش‌ها",
    "discounts":  "کدهای تخفیف",
    "growth":     "رشد و فروش",
    "tickets":    "مدیریت تیکت‌ها",
    "wallets":    "مدیریت کیف‌پول",
    "users":      "کاربران",
    "partners":   "مدیریت همکاران",
    "accounting": "حسابداری",
    "notes":      "یادداشت مدیران",
    "settings":   "تنظیمات ربات",
    "database":   "بکاپ و دیتابیس (نمای کلی)",
    "backup":     "پشتیبان‌گیری",
    "restore":    "بازیابی از بکاپ",
    "recovery":   "بازگردانی اضطراری",
    "admins":     "مدیریت ادمین‌ها و نقش‌ها",
    "logs":       "گزارش فعالیت (لاگ‌ها)",
    "broadcast":  "پیام همگانی",
    "payment":    "مدیریت پرداخت (رسیدهای کارت‌به‌کارت)",
    "payment_gateways": "درگاه‌های پرداخت آنلاین",
    "reports":    "گزارش‌های مالی",
    "news":       "اخبار و RSS",
    "articles":   "مقالات و آموزش‌ها",
    "ai_pricing": "هوش مصنوعی و قیمت‌گذاری آیفون",
    "mini_app":   "مینی اپ",
    "panel_appearance": "ظاهر و قالب پنل",
    "notifications": "اعلان‌ها و تعامل کاربر (چک‌این/امتیازها)",
    "cache_cleanup": "پاکسازی کش و فایل‌های موقت",
}

# سازگاری با ادمین‌های قدیمی: کلید جدید ← والد قدیمی — یعنی ادمینی که قبلاً «والد» رو
# داشته، بعد از این آپدیت خودکار به کلید تازهٔ دقیق‌تر هم دسترسی داره، بدون نیاز به
# تنظیم مجدد دستی. هر دسترسی تازه‌ای که در آینده به پروژه اضافه می‌شه، فقط کافیه یک
# سطر اینجا (اگه زیرمجموعهٔ منطقی یه دسترسی قدیمی‌تره) و یک سطر توی ALL_PERMISSIONS
# اضافه بشه — نیازی به تغییر ساختار اصلی (_has/_require/صفحهٔ مدیریت ادمین‌ها) نیست.
PERM_LEGACY = {
    "discounts": "orders", "growth": "orders",
    "users": "wallets", "accounting": "wallets", "notes": "wallets",
    "logs": "admins",
    "categories": "products",
    "backup": "database", "restore": "database", "recovery": "database",
    "payment": "wallets",
    "payment_gateways": "settings",
    "reports": "wallets",
    "news": "settings", "articles": "settings",
    "ai_pricing": "settings",
    "mini_app": "settings",
    "panel_appearance": "settings",
    "notifications": "settings",
    "cache_cleanup": "database",
}

# ─────────────────────────── DB Schema for Admins ──────────────────────────

_ADMINS_TABLE_READY = False

def ensure_admins_table() -> None:
    global _ADMINS_TABLE_READY
    if _ADMINS_TABLE_READY:
        return
    try:
        conn = _db()
        conn.execute("""
            CREATE TABLE IF NOT EXISTS admins (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                telegram_id INTEGER UNIQUE,
                name TEXT NOT NULL,
                web_username TEXT UNIQUE,
                web_password_hash TEXT,
                permissions TEXT DEFAULT '[]',
                is_active INTEGER DEFAULT 1,
                notes TEXT,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP
            );
        """)
        conn.commit()
        conn.close()
        _ADMINS_TABLE_READY = True
    except Exception:
        pass

# ─────────────────────────── Auth & Session ────────────────────────────────
# ⚠️ رفع امنیتی (بخش ۱۳ آیتم‌های ۴/۵ سند): پیش‌فرض SESSION_SECRET یکپارچه شد
# (قبلاً _hash_pw از "stockland" و _make_session/_get_admin از "stockland-panel"
# استفاده می‌کردن — دو رمز پیش‌فرض متفاوت). هش پسورد هم از SHA256 بدون نمک به
# PBKDF2-HMAC-SHA256 نمکی (۲۰۰هزار تکرار) تغییر کرد. سازگاری عقب‌رو کامل:
# _verify_pw هر دو فرمت (هش قدیمی خام + فرمت جدید pbkdf2$...) رو می‌پذیره و بعد
# از اولین ورود موفق با فرمت قدیمی، خودکار به فرمت جدید ارتقا می‌ده (login_post).
_SESSION_SECRET_DEFAULT = "stockland-panel"
_LEGACY_HASH_SECRET_DEFAULT = "stockland"  # فقط برای وریفای هش‌های قدیمی، دیگه برای هش تازه استفاده نمی‌شه
_PBKDF2_ITERATIONS = 200_000

def _pbkdf2_hash(password: str, salt: bytes | None = None) -> str:
    secret = _env("SESSION_SECRET", _SESSION_SECRET_DEFAULT)
    if salt is None:
        salt = os.urandom(16)
    dk = hashlib.pbkdf2_hmac("sha256", (secret + password).encode(), salt, _PBKDF2_ITERATIONS)
    return f"pbkdf2${salt.hex()}${dk.hex()}"

def _hash_pw(password: str) -> str:
    """هش تازه/تغییریافته — همیشه با فرمت جدید نمکی PBKDF2."""
    return _pbkdf2_hash(password)

def _verify_pw(password: str, stored: str) -> bool:
    """مثل _hmac.compare_digest ولی هر دو فرمت (جدید pbkdf2$salt$hash، قدیمی SHA256 خام
    بدون نمک) رو می‌پذیره — برای admins.web_password_hash."""
    if not stored:
        return False
    if stored.startswith("pbkdf2$"):
        try:
            _, salt_hex, hash_hex = stored.split("$", 2)
            salt = bytes.fromhex(salt_hex)
            secret = _env("SESSION_SECRET", _SESSION_SECRET_DEFAULT)
            dk = hashlib.pbkdf2_hmac("sha256", (secret + password).encode(), salt, _PBKDF2_ITERATIONS)
            return _hmac.compare_digest(dk.hex(), hash_hex)
        except Exception:
            return False
    # فرمت قدیمی (پیش از این رفع): SHA256(secret_قدیمی + password) بدون نمک
    legacy_secret = _env("SESSION_SECRET", _LEGACY_HASH_SECRET_DEFAULT)
    legacy_hash = hashlib.sha256((legacy_secret + password).encode()).hexdigest()
    return _hmac.compare_digest(stored, legacy_hash)

def _verify_super_pw(password: str, stored: str) -> bool:
    """برای ADMIN_WEB_PASSWORD در .env — فرمت قدیمی پسورد خام (plaintext) بود؛
    فرمت جدید (بعد از اولین تغییر از پنل) pbkdf2$salt$hash است (بخش ۱۳ آیتم ۲)."""
    if not stored:
        return False
    if stored.startswith("pbkdf2$"):
        return _verify_pw(password, stored)
    return _hmac.compare_digest(password, stored)

IDLE_TIMEOUT_SECONDS = 300  # ۵ دقیقه عدم فعالیت → logout

def _make_session(admin_id: str) -> str:
    """session شامل admin_id و timestamp، امضاشده با HMAC."""
    import time as _t
    ts = str(int(_t.time()))
    secret = _env("SESSION_SECRET", _SESSION_SECRET_DEFAULT)
    payload = f"{admin_id}|{ts}"
    token = _hmac.new(secret.encode(), payload.encode(), hashlib.sha256).hexdigest()
    return f"{token}:{admin_id}|{ts}"

def _verify_session_cookie(request: Request) -> str | None:
    """فقط HMAC + idle timeout رو چک می‌کنه و admin_id رو برمی‌گردونه — بدون هیچ کوئری
    دیتابیس. برای جاهایی که فقط لازمه بدونیم «این یه سشن معتبره یا نه» (مثلاً تجدید
    کوکی sliding-window)، نه واقعاً دسترسی‌ها/is_active که نیاز به DB داره."""
    import time as _t
    cookie = request.cookies.get("adm", "")
    if not cookie or ":" not in cookie:
        return None

    token, payload = cookie.rsplit(":", 1)

    if "|" in payload:
        admin_id, ts_str = payload.rsplit("|", 1)
    else:
        admin_id, ts_str = payload, None

    expected_token = _hmac.new(
        _env("SESSION_SECRET", _SESSION_SECRET_DEFAULT).encode(),
        payload.encode(),
        hashlib.sha256,
    ).hexdigest()

    if not _hmac.compare_digest(token, expected_token):
        return None

    if ts_str:
        try:
            age = int(_t.time()) - int(ts_str)
            if age > IDLE_TIMEOUT_SECONDS:
                return None
        except Exception:
            return None

    return admin_id


def _get_admin(request: Request):
    """Returns (admin_id, is_super, permissions_list) or None.
    اعتبارسنجی HMAC + بررسی idle timeout. Session لغزنده نیست — عمر ثابت ۳۰۰ ثانیه از آخرین صدور.
    هر response باید با _refresh_session() کوکی رو تجدید کنه تا مدیر فعال kick نشه."""
    ensure_admins_table()
    admin_id = _verify_session_cookie(request)
    if not admin_id:
        return None

    if admin_id == "super":
        return ("super", True, list(ALL_PERMISSIONS.keys()))

    try:
        conn = _db()
        row = conn.execute(
            "SELECT id, permissions, is_active FROM admins WHERE id=? LIMIT 1;",
            (int(admin_id),),
        ).fetchone()
        conn.close()
        if not row or not row["is_active"]:
            return None
        perms = json.loads(row["permissions"] or "[]")
        return (str(row["id"]), False, perms)
    except Exception:
        return None


def _refresh_session(response, admin_info) -> None:
    """Session را تجدید می‌کند تا مدیر فعال بعد از ۵ دقیقه kick نشود.
    باید در ابتدای هر GET handler صدا زده شود."""
    if not admin_info:
        return
    admin_id = admin_info[0]
    new_cookie = _make_session(str(admin_id))
    response.set_cookie(
        "adm", new_cookie,
        max_age=IDLE_TIMEOUT_SECONDS,
        httponly=True,
        samesite="lax",
        secure=True,  # تولید همیشه HTTPS است (panel.stland.ir) — مثل کوکی sl_sess در api.py
    )


def _admin_id_of(admin_info) -> str:
    return admin_info[0] if admin_info else ""

def _has(admin_info, perm: str) -> bool:
    if not admin_info:
        return False
    _, is_super, perms = admin_info
    if is_super:
        return True
    if perm not in ALL_PERMISSIONS:
        return False             # کلید ناشناخته = رد (امنیت پیش‌فرض)
    if perm in (perms or []):
        return True
    legacy = PERM_LEGACY.get(perm)
    return bool(legacy and legacy in (perms or []))

def _require(admin_info, perm: str):
    """Returns 403 redirect if admin lacks permission."""
    if not admin_info:
        return RedirectResponse("/admin/login", status_code=303)
    if not _has(admin_info, perm):
        return RedirectResponse("/admin/?err=noperm", status_code=303)
    return None


def _require_any(admin_info, perms: list):
    """مثل _require ولی کافیه فقط یکی از perms رو داشته باشه — برای صفحاتی که چند
    زیردسترسی مجزا رو با هم نشون می‌دن (مثلاً /admin/database که دکمه‌های بکاپ/بازیابی/
    بازگردانی هرکدوم permission جدای خودشون رو دارن، ولی خودِ صفحه باید برای هرکدوم که
    حداقل یکی از این‌ها رو داشته باشه باز بشه، نه فقط کسی که همه‌شون رو داره)."""
    if not admin_info:
        return RedirectResponse("/admin/login", status_code=303)
    if not any(_has(admin_info, p) for p in perms):
        return RedirectResponse("/admin/?err=noperm", status_code=303)
    return None

def _redir(path: str) -> RedirectResponse:
    return RedirectResponse(path, status_code=303)


def _stbak_pg_guard():
    """⚠️ گارد امنیت عملیاتی (ممیزی کامل پروژه، پاک‌سازی SQLite): مسیرهای
    پشتیبان‌گیری/بازیابی/ریست مبتنی بر stbak_engine فقط روی فایل SQLite کار
    می‌کنن — اگه پروژه واقعاً روی Postgres مهاجرت کرده، این فایل کاملاً جدا از
    دادهٔ واقعیه. اجرای بی‌قید این مسیرها یعنی ادمین فکر می‌کنه بکاپ/بازیابی
    موفق بوده، در حالی که هیچ ربطی به دادهٔ واقعی نداشته — خطرناک‌تر از شکست
    آشکار. این تابع None برمی‌گردونه اگه دیالوگ SQLite باشه (بدون تغییر رفتار)،
    وگرنه پیام خطای واضح."""
    import db_conn
    if db_conn.is_postgres():
        return {"error": "این پروژه به Postgres مهاجرت کرده — بکاپ/بازیابی SQLite (.stbak قدیمی) اینجا اعمال نمی‌شه. از بخش «بکاپ Postgres (pg_dump)» بالای همین صفحه استفاده کن."}
    return None

# ─────────────────────────── کش سبک درون‌پروسه‌ای ──────────────────────────
# چند تا کوئری (تم پنل، شمارنده‌های badge سایدبار) توی _layout() روی *هر* صفحهٔ
# پنل دوباره از دیتابیس خونده می‌شدن، با اینکه به‌ندرت تغییر می‌کنن. یه کش
# TTL کوتاه (چند ثانیه) اینجا رو به یه کوئری واحد کاهش می‌ده و تازگی داده رو هم
# عملاً حفظ می‌کنه (badge با چند ثانیه تأخیر آپدیت می‌شه، نه اینکه اشتباه باشه).
_PANEL_CACHE: dict = {}
_PANEL_CACHE_TTL = 5.0  # ثانیه — برای badgeهایی که هر ثانیه چندبار خونده می‌شن کافیه

def _cached(key: str, ttl: float, fn):
    now = time.time()
    hit = _PANEL_CACHE.get(key)
    if hit is not None and (now - hit[0]) < ttl:
        return hit[1]
    val = fn()
    _PANEL_CACHE[key] = (now, val)
    return val

def _cache_invalidate(key: str) -> None:
    _PANEL_CACHE.pop(key, None)

# ─────────────────────────── HTML helpers ──────────────────────────────────

def e(s) -> str:
    return html.escape(str(s or ""))

def _open_ticket_count_uncached() -> int:
    try:
        conn = _db()
        n = conn.execute("SELECT COUNT(*) FROM tickets WHERE status='waiting_admin';").fetchone()[0]
        conn.close()
        return int(n)
    except Exception:
        return 0


def _open_ticket_count() -> int:
    return _cached("open_ticket_count", _PANEL_CACHE_TTL, _open_ticket_count_uncached)


def _pending_payout_count() -> int:
    try:
        conn = _db()
        n = conn.execute("SELECT COUNT(*) FROM partner_payouts WHERE status='pending';").fetchone()[0]
        conn.close()
        return int(n or 0)
    except Exception:
        return 0


def _pending_card2card_count() -> int:
    try:
        from db import count_card_receipts
        return count_card_receipts("pending")
    except Exception:
        return 0


def _pending_sell_requests_count() -> int:
    try:
        import iphone_valuation.db as ivdb
        # ⚠️ رفع‌شده: iv_sell_requests.status پیش‌فرضش 'new' هست (نه 'pending') — هم
        # ستون جدول (DEFAULT 'new') هم create_sell_request در bot.py صریحاً 'new' رو
        # ثبت می‌کنن، و منطق تأیید ادمین هم به 'contacted' تغییرش می‌ده. چک قبلی
        # دنبال 'pending' می‌گشت که هیچ‌وقت مقداردهی نمی‌شد — یعنی درخواست تازه هیچ‌وقت
        # نوتیف بالای «خرید و بخش مالی» رو فعال نمی‌کرد.
        return ivdb.count_sell_requests("new")
    except Exception:
        return 0


def _pending_receipts_count_uncached() -> int:
    """تعداد کل موارد در انتظار صفحهٔ «خرید و بخش مالی» — کارت‌به‌کارت + تسویه همکار + درخواست فروش."""
    return _pending_payout_count() + _pending_card2card_count() + _pending_sell_requests_count()


def _pending_receipts_count() -> int:
    return _cached("pending_receipts_count", _PANEL_CACHE_TTL, _pending_receipts_count_uncached)


def _pending_partner_count_uncached() -> int:
    try:
        conn = _db()
        n = conn.execute("SELECT COUNT(*) FROM partners WHERE status='pending';").fetchone()[0]
        conn.close()
        return int(n)
    except Exception:
        return 0


def _pending_partner_count() -> int:
    return _cached("pending_partner_count", _PANEL_CACHE_TTL, _pending_partner_count_uncached)


# ─── Theme System ──────────────────────────────────────────────────────────

DEFAULT_THEME = {
    "sidebar_bg":     "#05070A",
    "sidebar_text":   "#9AA7B8",
    "sidebar_active": "#2EC4B6",
    "primary":        "#2EC4B6",
    "primary_text":   "#ffffff",
    "accent":         "#F59E0B",
    "page_bg":        "#F7F8FA",
    "card_bg":        "#FFFFFF",
    "text_main":      "#111827",
    "text_muted":     "#6B7280",
    "border":         "#E5E7EB",
}

def _get_theme_uncached() -> dict:
    theme = dict(DEFAULT_THEME)
    try:
        conn = _db()
        rows = conn.execute("SELECT key, value FROM panel_theme;").fetchall()
        conn.close()
        for r in rows:
            if r["key"] in theme:
                theme[r["key"]] = r["value"]
    except Exception:
        pass
    return theme


def _get_theme() -> dict:
    # تم پنل تقریباً هیچ‌وقت عوض نمی‌شه ولی روی *هر* صفحه (توسط _layout) خونده می‌شد —
    # TTL طولانی‌تر از badgeها چون تغییرش خیلی کمتر اتفاق می‌افته؛ محل ذخیره هم با
    # _cache_invalidate("panel_theme") توی روت‌های ذخیرهٔ تم کاملاً پاک می‌شه، پس
    # همون لحظه‌ای که ادمین تم رو عوض می‌کنه، خودش نتیجهٔ تازه رو می‌بینه.
    return _cached("panel_theme", 60.0, _get_theme_uncached)


def _get_admin_prefs_uncached(admin_id: str) -> dict:
    try:
        conn = _db()
        rows = conn.execute(
            "SELECT key, value FROM admin_preferences WHERE admin_id=? AND key IN ('dark_mode','classic_mode');",
            (admin_id,)
        ).fetchall()
        conn.close()
        return {r["key"]: r["value"] for r in rows}
    except Exception:
        return {}


def _get_admin_prefs(admin_id) -> dict:
    # این کوئری قبلاً بدون کش، مستقیم داخل _layout() اجرا می‌شد — یعنی روی هر
    # رندر هر صفحهٔ پنل (که _layout در ۷۶+ نقطه صدا زده می‌شه) یه I/O بلاک‌کنندهٔ
    # synchronous جدا. همون الگوی _get_theme، فقط per-admin (کلید کش شامل
    # admin_id) چون این مقدار برخلاف تم، بین ادمین‌های مختلف فرق می‌کنه.
    return _cached(f"admin_prefs:{admin_id}", 30.0, lambda: _get_admin_prefs_uncached(str(admin_id)))

def _ensure_theme_table():
    conn = _db()
    try:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS panel_theme (
                key TEXT PRIMARY KEY,
                value TEXT NOT NULL
            );
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS admin_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                admin_id   INTEGER,
                admin_name TEXT,
                action     TEXT NOT NULL,
                section    TEXT,
                details    TEXT,
                ip         TEXT,
                created_at TEXT NOT NULL DEFAULT (datetime('now','localtime'))
            );
        """)
        conn.commit()
    finally:
        conn.close()


_ADMIN_LOGS_SCHEMA_READY = False  # فلگ per-process مشترک بین _log() و _fetch_order_logs()

def _log(request: Request, action: str, section: str = "", details: str = "", admin_info=None, result: str = "ok"):
    """ثبت فعالیت ادمین — هیچ‌وقت exception نمی‌ده.
    ⚠️ قبلاً این تابع خودش هیچ‌وقت جدول admin_logs رو نمی‌ساخت — فقط به این تکیه
    می‌کرد که یکی از ۳ روت دیگه (که _ensure_theme_table رو صدا می‌زنن) قبلش اجرا
    شده باشه؛ روی نصب کاملاً تازه، اگه هیچ‌کدوم از اون ۳ روت هنوز باز نشده بودن،
    هر INSERT اینجا با «no such table» شکست می‌خورد و بی‌صدا قورت داده می‌شد —
    یعنی هیچ لاگی ثبت نمی‌شد، بدون هیچ نشونه‌ای. الان خودش صریح گارد می‌کنه."""
    global _ADMIN_LOGS_SCHEMA_READY
    try:
        adm = admin_info or _get_admin(request)
        if not adm:
            return
        ip   = request.headers.get("X-Forwarded-For","").split(",")[0].strip() or (request.client.host if request.client else "—")
        name = adm[3] if len(adm) > 3 else f"admin#{adm[0]}"
        conn = _db()
        # ⚠️ رفع کارایی: چون _log() پرتکرارترین نقطهٔ لمس admin_logs است (هر اکشن
        # ادمین)، CREATE TABLE/ALTER/INDEX قبلاً هر بار اجرا می‌شدن — ALTER هم بعد
        # از بار اول همیشه با استثنا (ستون از قبل موجوده) شکست می‌خورد. با همون فلگ
        # per-process که _fetch_order_logs استفاده می‌کنه یکی شد، فقط یک‌بار امتحان می‌شه.
        if not _ADMIN_LOGS_SCHEMA_READY:
            conn.execute("""
                CREATE TABLE IF NOT EXISTS admin_logs (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    admin_id   INTEGER,
                    admin_name TEXT,
                    action     TEXT NOT NULL,
                    section    TEXT,
                    details    TEXT,
                    ip         TEXT,
                    created_at TEXT NOT NULL DEFAULT (datetime('now','localtime'))
                );
            """)
            try:
                conn.execute("ALTER TABLE admin_logs ADD COLUMN result TEXT DEFAULT 'ok';")
            except Exception:
                pass
            conn.execute("CREATE INDEX IF NOT EXISTS idx_admin_logs_section ON admin_logs(section, id);")
            conn.commit()
            _ADMIN_LOGS_SCHEMA_READY = True
        conn.execute(
            "INSERT INTO admin_logs (admin_id,admin_name,action,section,details,ip,result) VALUES (?,?,?,?,?,?,?);",
            (adm[0], name, action, section, details[:500] if details else "", ip, result)
        )
        conn.commit()
        conn.close()
    except Exception:
        pass


# ─── Rate Limiting ─────────────────────────────────────────────────────────
import time as _time
_login_attempts: dict = {}  # ip → [timestamps]
_LOGIN_MAX = 5
_LOGIN_WINDOW = 900  # 15 دقیقه

def _is_rate_limited(ip: str) -> tuple:
    """(is_blocked, remaining_seconds)"""
    now = _time.time()
    _login_attempts.setdefault(ip, [])
    _login_attempts[ip] = [t for t in _login_attempts[ip] if now - t < _LOGIN_WINDOW]
    if len(_login_attempts[ip]) >= _LOGIN_MAX:
        return True, int(_LOGIN_WINDOW - (now - _login_attempts[ip][0]))
    return False, 0

def _record_fail(ip: str):
    _login_attempts.setdefault(ip, []).append(_time.time())

def _clear_attempts(ip: str):
    _login_attempts.pop(ip, None)


# ─── Low Stock Notification ────────────────────────────────────────────────
_notified_low: set = set()  # کلیدهایی که قبلاً نوتیف گرفتن

def _notify_low_stock():
    """بررسی و ارسال اطلاع‌رسانی موجودی کم — توسط scheduler صدا زده می‌شه."""
    try:
        bot_token = _env("BOT_TOKEN")
        admin_id  = _env("ADMIN_ID")
        if not bot_token or not admin_id:
            return
        threshold = int(_env("LOW_STOCK_THRESHOLD", "5"))
        conn = _db()
        rows = conn.execute("""
            SELECT p.id, p.title,
                   COUNT(CASE WHEN pf.delivered=0 THEN 1 END) AS avail
            FROM products p
            LEFT JOIN product_feed pf ON pf.product_id = p.id
            WHERE p.is_active = 1
            GROUP BY p.id, p.title
            HAVING COUNT(CASE WHEN pf.delivered=0 THEN 1 END) <= ?
            ORDER BY avail ASC LIMIT 20;
        """, (threshold,)).fetchall()
        conn.close()

        for r in rows:
            pid, title, avail = r["id"], r["title"], int(r["avail"] or 0)
            key = f"{pid}:{avail}"
            if key in _notified_low:
                continue
            # پاک کردن کلیدهای قدیمی همین محصول
            _notified_low.discard(next((k for k in list(_notified_low) if k.startswith(f"{pid}:")), None))
            icon = "🔴" if avail == 0 else "⚠️"
            status = "موجودی صفر شد" if avail == 0 else f"موجودی کم ({avail} عدد)"
            msg = (f"{icon} <b>هشدار موجودی</b>\n"
                   f"📦 محصول: {title}\n"
                   f"📉 وضعیت: {status}\n"
                   f"🔗 /admin/feed/{pid}")
            try:
                _requests.post(
                    f"https://api.telegram.org/bot{bot_token}/sendMessage",
                    json={"chat_id": int(admin_id), "text": msg, "parse_mode": "HTML"},
                    timeout=10
                )
                _notified_low.add(key)
            except Exception:
                pass
    except Exception:
        pass


def _layout(title: str, body: str, admin_info=None,
            flash: str = "", flash_ok: bool = True) -> HTMLResponse:

    theme = _get_theme()

    # بارگذاری تم ذخیره‌شده مدیر
    saved_dark = ""
    saved_classic = ""
    if admin_info:
        prefs = _get_admin_prefs(admin_info[0])
        saved_dark = prefs.get("dark_mode", "")
        saved_classic = prefs.get("classic_mode", "")

    flash_html = ""
    if flash:
        icon = "circle-check" if flash_ok else "circle-alert"
        flash_html = f"""
        <div class="flash-msg flash-{'ok' if flash_ok else 'err'}">
          <i data-lucide="{icon}"></i><span>{e(flash)}</span>
        </div>"""

    perms = admin_info[2] if admin_info else []
    is_super = admin_info[1] if admin_info else False

    open_tickets    = _open_ticket_count()    if admin_info else 0
    pending_partners = _pending_partner_count() if admin_info else 0
    try:
        pending_receipts_top = _pending_receipts_count() if admin_info else 0
    except Exception:
        pending_receipts_top = 0
    bell_count = open_tickets

    def has_perm(perm):
        return is_super or _has(admin_info, perm)

    def nav_item(href, icon, label, perm=None, badge=0, badge_id=None):
        if perm and not has_perm(perm):
            return ""
        if badge_id:
            badge_html = f'<span id="{badge_id}" class="nav-badge {"hidden" if badge == 0 else ""}">{badge}</span>'
        else:
            badge_html = f'<span class="nav-badge">{badge}</span>' if badge > 0 else ""
        return f'<a href="{href}" class="nav-item" data-href="{href}"><i data-lucide="{icon}" class="nav-icon"></i><span class="nav-label">{label}</span>{badge_html}</a>'

    sidebar = ""
    if admin_info:
        sidebar = f"""
        <aside id="sidebar" class="sidebar">
          <div class="sidebar-header">
            <a href="/admin/" class="brand-lockup" aria-label="استوک‌لند" id="sb-brand">
              <div class="brand-text-only">
                <span class="brand-word brand-word--stock">STOCK</span>
                <span class="brand-word brand-word--land"> LAND</span>
                <small class="brand-subtitle">مدیریت فروشگاه</small>
              </div>
            </a>
            <button onclick="sbCollapse()" id="sb-col-btn" class="sb-col-btn" title="جمع/باز">
              <span id="sb-toggle-icon" class="sb-toggle-icon">›</span>
            </button>
          </div>
          <div class="nav-caption">منو اصلی</div>
          <nav class="sidebar-nav">
            {nav_item("/admin/", "layout-dashboard", "داشبورد")}
            <div class="nav-divider"><span>فروشگاه</span></div>
            {nav_item("/admin/categories", "tag", "دسته‌بندی‌ها", "categories")}
            {nav_item("/admin/products", "package", "محصولات", "products")}
            {nav_item("/admin/feed", "layers", "موجودی", "feed")}
            <div class="nav-divider"><span>فروش و مالی</span></div>
            {nav_item("/admin/orders", "shopping-bag", "سفارش‌ها", "orders")}
            {nav_item("/admin/receipts", "receipt", "خرید و بخش مالی", "payment", pending_receipts_top, "receipts-badge-nav")}
            {nav_item("/admin/wallets", "wallet", "کیف‌پول", "wallets")}
            {nav_item("/admin/discounts", "percent", "کدهای تخفیف", "discounts")}
            {nav_item("/admin/growth", "rocket", "رشد و فروش", "growth")}
            {nav_item("/admin/accounting", "calculator", "حسابداری", "accounting")}
            {nav_item("/admin/payment-gateways", "credit-card", "درگاه‌های پرداخت", "payment_gateways")}
            <div class="nav-divider"><span>محتوا</span></div>
            {nav_item("/admin/news-feed", "rss", "اخبار تکنولوژی", "news")}
            {nav_item("/admin/tutorials", "graduation-cap", "آموزش", "articles")}
            {nav_item("/admin/broadcast", "megaphone", "پیام‌رسانی", "broadcast")}
            <div class="nav-divider"><span>کاربران</span></div>
            {nav_item("/admin/users", "users", "کاربران", "users")}
            {nav_item("/admin/partners", "handshake", "همکاران و معرفی", "partners", pending_partners, "partner-badge-nav")}
            {nav_item("/admin/tickets", "message-square", "تیکت‌ها", "tickets", open_tickets, "ticket-badge-nav")}
            <div class="nav-divider"><span>ابزارها</span></div>
            {nav_item("/admin/iphone", "smartphone", "کارشناسی آیفون", "ai_pricing")}
            {nav_item("/admin/engagement", "star", "پاداش و تعامل", "notifications")}
            {nav_item("/admin/stock-requests", "bell-ring", "درخواست‌های موجودی", "notifications")}
            <div class="nav-divider"><span>سیستم</span></div>
            {nav_item("/admin/settings/panel", "settings", "تنظیمات", "settings")}
            {nav_item("/admin/database", "database", "پشتیبان‌گیری", "database")}
            {nav_item("/admin/system-cache", "trash-2", "پاکسازی کش", "cache_cleanup")}
            {nav_item("/admin/admins", "shield-check", "ادمین‌ها", "admins")}
            {nav_item("/admin/logs", "activity", "گزارش فعالیت", "logs")}
            {nav_item("/admin/notes", "edit-3", "یادداشت مدیران", "notes")}
          </nav>
          <div class="sidebar-footer">
            <div class="sidebar-status"><span class="status-dot"></span><div><strong>سامانه فعال</strong><small>همه سرویس‌ها پایدارند</small></div></div>
            <a href="/admin/logout" class="sidebar-logout"><i data-lucide="log-out"></i><span>خروج از پنل</span></a>
          </div>
        </aside>
        <div id="overlay" class="overlay" onclick="toggleSidebar()"></div>"""

    topbar = ""
    if admin_info:
        admin_label = "مدیر ارشد" if is_super else f"مدیر #{e(admin_info[0])}"
        topbar = f"""
        <header class="topbar">
          <div class="topbar-context">
            <button class="icon-button topbar-menu" onclick="toggleSidebar()" aria-label="بازکردن منو"><i data-lucide="menu"></i></button>
            <div><span class="topbar-eyebrow">STOCKLAND ADMIN</span><h1 class="topbar-title">{e(title)}</h1></div>
          </div>
          <div class="global-search-wrap">
            <i data-lucide="search"></i>
            <input id="globalSearch" class="global-search" type="search" placeholder="جست‌وجو در پنل..." autocomplete="off">
            <kbd>⌘ K</kbd>
          </div>
          <div class="topbar-actions">

            <a class="icon-button notification-button" href="/admin/receipts" aria-label="خرید و بخش مالی"><i data-lucide="receipt"></i><span id="receipts-badge-top" class="notification-count notification-count--warn {'hidden' if pending_receipts_top == 0 else ''}">{pending_receipts_top}</span></a>
            <a class="icon-button notification-button" href="/admin/tickets" aria-label="تیکت‌ها"><i data-lucide="bell"></i><span id="ticket-badge-top" class="notification-count {'hidden' if bell_count == 0 else ''}">{bell_count}</span></a>
            <a class="icon-button notification-button" href="/admin/partners" aria-label="همکاران"><i data-lucide="handshake"></i><span id="partner-badge-top" class="notification-count notification-count--warn {'hidden' if pending_partners == 0 else ''}">{pending_partners}</span></a>
            <a class="icon-button notification-button" href="/admin/notes" aria-label="یادداشت‌ها"><i data-lucide="edit-3"></i><span id="notes-badge-top" class="notification-count notification-count--danger hidden"></span></a>
            <a href="/admin/account" class="profile-trigger">
              <span class="profile-avatar"><i data-lucide="user-round"></i></span>
              <span class="profile-copy"><strong>{admin_label}</strong><small>مدیریت فروشگاه</small></span>
              <i data-lucide="settings" class="profile-chevron profile-chevron--sm"></i>
            </a>
          </div>
        </header>"""

    css_vars = ";".join(f"--{k.replace('_','-')}:{v}" for k,v in theme.items())

    html_response = HTMLResponse(f"""<!DOCTYPE html>
<html lang="fa" dir="rtl" data-saved-dark="{saved_dark}" data-saved-classic="{saved_classic}">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width,initial-scale=1,maximum-scale=1,user-scalable=no">
  <script>
    document.addEventListener('gesturestart',function(e){{e.preventDefault();}});
    document.addEventListener('gesturechange',function(e){{e.preventDefault();}});
    document.addEventListener('gestureend',function(e){{e.preventDefault();}});
    document.addEventListener('touchmove',function(e){{if(e.touches.length>1)e.preventDefault();}},{{passive:false}});
  </script>
  <title>{e(title)} — استوک لند</title>
  <link rel="preload" href="/app/vendor/fonts/Vazirmatn-Regular.woff2" as="font" type="font/woff2" crossorigin>
  <script src="/app/vendor/admin-tailwind.js"></script>
  <script>window.tailwind || document.write('<script src="https://cdn.tailwindcss.com"><\\/script>');</script>
  <script src="/app/vendor/admin-lucide.min.js"></script>
  <script>window.lucide || document.write('<script src="https://unpkg.com/lucide@0.468.0/dist/umd/lucide.min.js"><\\/script>');</script>
  <style>
    /* ═══════════════════════════════════════════════════════════
       STOCKLAND ADMIN — DESIGN SYSTEM v2
       Mobile-First · RTL-Native · Apple-Inspired
    ═══════════════════════════════════════════════════════════ */

    /* ── فونت محلی (سلف‌هاست، بدون وابستگی به fonts.googleapis.com) ── */
    @font-face {{ font-family:'Vazirmatn'; font-weight:300; font-display:swap; src:url('/app/vendor/fonts/Vazirmatn-Light.woff2') format('woff2'); }}
    @font-face {{ font-family:'Vazirmatn'; font-weight:400; font-display:swap; src:url('/app/vendor/fonts/Vazirmatn-Regular.woff2') format('woff2'); }}
    @font-face {{ font-family:'Vazirmatn'; font-weight:500; font-display:swap; src:url('/app/vendor/fonts/Vazirmatn-Medium.woff2') format('woff2'); }}
    @font-face {{ font-family:'Vazirmatn'; font-weight:600; font-display:swap; src:url('/app/vendor/fonts/Vazirmatn-SemiBold.woff2') format('woff2'); }}
    @font-face {{ font-family:'Vazirmatn'; font-weight:700; font-display:swap; src:url('/app/vendor/fonts/Vazirmatn-Bold.woff2') format('woff2'); }}
    @font-face {{ font-family:'Vazirmatn'; font-weight:800; font-display:swap; src:url('/app/vendor/fonts/Vazirmatn-ExtraBold.woff2') format('woff2'); }}

    /* ── Design Tokens ────────────────────────────────────────── */
    :root {{
      {css_vars}
      /* Color Palette */
      --clr-primary:      #2EC4B6;
      --clr-primary-dim:  rgba(46,196,182,.10);
      --clr-success:      #22C55E;
      --clr-success-dim:  #DCFCE7;
      --clr-danger:       #EF4444;
      --clr-danger-dim:   #FEE2E2;
      --clr-warning:      #F59E0B;
      --clr-warning-dim:  #FEF3C7;
      --clr-info:         #3B82F6;
      --clr-info-dim:     #EFF6FF;
      --clr-neutral:      #6B7280;
      --clr-neutral-dim:  #F3F4F6;

      /* Semantic Text */
      --txt-primary:   #111827;
      --txt-secondary: #374151;
      --txt-muted:     #6B7280;
      --txt-xmuted:    #9CA3AF;

      /* Backgrounds */
      --bg-page:    #F7F8FA;
      --bg-card:    #FFFFFF;
      --bg-input:   #FFFFFF;
      --bg-subtle:  #F9FAFB;

      /* Borders */
      --bdr:        #E5E7EB;
      --bdr-input:  #D1D5DB;
      --bdr-focus:  var(--clr-primary);

      /* Typography */
      --font:       'Vazirmatn', Tahoma, 'Segoe UI', sans-serif;

      /* Spacing Scale */
      --sp-1: 4px;   --sp-2: 8px;   --sp-3: 12px;  --sp-4: 16px;
      --sp-5: 20px;  --sp-6: 24px;  --sp-7: 28px;  --sp-8: 32px;

      /* Border Radius */
      --r-sm: 8px;  --r-md: 12px;  --r-lg: 16px;  --r-xl: 20px;

      /* Shadows */
      --shadow-card:  0 1px 3px rgba(15,23,42,.06), 0 4px 12px rgba(15,23,42,.04);
      --shadow-hover: 0 4px 16px rgba(15,23,42,.10);
      --shadow-modal: 0 20px 60px rgba(15,23,42,.16);

      /* Layout */
      --sidebar-w:   272px;
      --topbar-h:    60px;
      --glass-level: 0;

      /* Legacy compat */
      --primary:     #2EC4B6;
      --page-bg:     #F7F8FA;
      --card-bg:     #FFFFFF;
      --text-main:   #111827;
      --text-muted:  #6B7280;
      --border:      #E5E7EB;
      --success:     #22C55E;
      --warning:     #F59E0B;
      --danger:      #EF4444;
      --card-shadow: var(--shadow-card);
    }}

    /* ── Reset ────────────────────────────────────────────────── */
    *, *::before, *::after {{ box-sizing:border-box; }}
    html {{ background:var(--bg-page); scroll-behavior:smooth; -webkit-text-size-adjust:100%; }}
    body {{
      margin:0; font-family:var(--font); background:var(--bg-page);
      color:var(--txt-primary); min-height:100vh;
      -webkit-font-smoothing:antialiased; direction:rtl;
    }}
    button,input,select,textarea {{ font-family:var(--font); }}
    a {{ color:inherit; text-decoration:none; }}
    svg {{ flex:0 0 auto; display:block; }}
    img {{ max-width:100%; }}
    .hidden {{ display:none !important; }}

    /* ── Typography Scale ─────────────────────────────────────── */
    .t-page   {{ font-size:24px; font-weight:700; line-height:1.2; color:var(--txt-primary); }}
    .t-section{{ font-size:18px; font-weight:700; line-height:1.3; color:var(--txt-primary); }}
    .t-card   {{ font-size:15px; font-weight:600; line-height:1.4; color:var(--txt-primary); }}
    .t-body   {{ font-size:14px; font-weight:400; line-height:1.6; color:var(--txt-secondary); }}
    .t-label  {{ font-size:12px; font-weight:500; line-height:1.4; color:var(--txt-muted); }}
    .t-mono   {{ font-family:'SF Mono','Fira Code',monospace; font-size:12px; }}

    /* ── Layout ───────────────────────────────────────────────── */
    .main-wrap.with-sidebar {{ margin-right:var(--sidebar-w); padding-top:var(--topbar-h); min-height:100vh; }}
    .main-content {{ padding:var(--sp-6) var(--sp-7); max-width:1600px; }}

    /* ── Sidebar ──────────────────────────────────────────────── */
    .sidebar {{
      position:fixed; top:0; right:0; width:var(--sidebar-w); height:100vh; z-index:300;
      display:flex; flex-direction:column; overflow:hidden;
      background:linear-gradient(180deg,#0B1320 0%,#05070A 100%);
      border-left:1px solid rgba(255,255,255,.06);
      box-shadow:-4px 0 24px rgba(2,6,23,.14);
      transition:transform .25s cubic-bezier(.4,0,.2,1), width .22s ease;
      color:#8896A8;
    }}
    @media (min-width:769px) {{
      .sidebar.sb-collapsed {{ width:54px !important; overflow:visible; }}
      .sidebar.sb-collapsed .nav-label {{ display:none !important; }}
      .sidebar.sb-collapsed .brand-text-only {{ display:none !important; }}
      .sidebar.sb-collapsed .nav-badge {{ display:none !important; }}
      .sidebar.sb-collapsed .sidebar-nav a {{
        justify-content:center; padding:11px 0; border-radius:10px; margin:1px 4px;
      }}
      .sidebar.sb-collapsed .sidebar-nav a .nav-icon {{ margin:0 !important; }}
      .sidebar.sb-collapsed .sidebar-header {{ justify-content:center !important; padding:10px 4px !important; }}
      .sidebar.sb-collapsed #sb-brand {{ display:none; }}
      .sidebar.sb-collapsed #sb-toggle-icon {{ transform:rotate(180deg); }}
      body.sb-collapsed-body .main-wrap.with-sidebar {{ margin-right:54px !important; }}
      body.sb-collapsed-body .topbar {{ right:54px !important; }}
    }}
    .sidebar-header {{
      padding:18px 18px 16px; border-bottom:1px solid rgba(255,255,255,.06);
      display:flex; align-items:center; gap:10px; flex-shrink:0; min-height:68px;
    }}
    .brand-text-only {{
      font-size:20px; font-weight:900; letter-spacing:2px; direction:ltr; color:#E8EDF2;
    }}
    .brand-text-only span {{ color:var(--clr-primary); }}
    .brand-text-only small {{ display:block; font-size:9.5px; font-weight:400; color:#364454; letter-spacing:.6px; margin-top:2px; direction:rtl; }}
    .sidebar-nav {{ flex:1; overflow-y:auto; overscroll-behavior:contain; -webkit-overflow-scrolling:touch; padding:10px 10px; display:flex; flex-direction:column; gap:1px; scrollbar-width:none; }}
    .sidebar-nav::-webkit-scrollbar {{ display:none; }}
    .nav-divider {{
      display:flex; align-items:center; gap:8px; padding:14px 16px 5px; opacity:.6;
    }}
    .nav-divider span {{ font-size:9.5px; font-weight:700; letter-spacing:1.4px; text-transform:uppercase; color:#2D3A4A; white-space:nowrap; }}
    .nav-divider::after {{ content:""; flex:1; height:1px; background:rgba(255,255,255,.05); }}
    .nav-item {{
      display:flex; align-items:center; gap:10px; padding:9px 14px;
      border-radius:var(--r-md); text-decoration:none; color:#8896A8;
      font-size:13px; font-weight:500; cursor:pointer; border:none;
      background:none; width:100%; text-align:right; transition:all .15s;
      white-space:nowrap; position:relative;
    }}
    .nav-item i {{ width:16px; height:16px; flex-shrink:0; }}
    .nav-item:hover {{ background:rgba(255,255,255,.05); color:#C5CDD8; }}
    .nav-item.active {{
      background:rgba(46,196,182,.10); color:var(--clr-primary); font-weight:600;
      box-shadow:inset 3px 0 0 var(--clr-primary);
    }}
    .nav-item.active i {{ filter:drop-shadow(0 0 4px rgba(46,196,182,.5)); }}
    .nav-badge {{
      margin-right:auto; background:var(--clr-danger); color:#fff;
      font-size:9px; font-weight:700; padding:1px 6px; border-radius:20px;
      min-width:18px; text-align:center; line-height:16px;
    }}
    .sidebar-footer {{
      padding:12px 10px; border-top:1px solid rgba(255,255,255,.06); flex-shrink:0;
    }}
    .sidebar-status {{
      display:flex; align-items:center; gap:10px; padding:8px 14px; margin-bottom:4px;
      font-size:12px; color:#394A5A;
    }}
    .status-dot {{
      width:7px; height:7px; border-radius:50%; background:var(--clr-success); flex-shrink:0;
    }}
    .sidebar-logout {{
      display:flex; align-items:center; gap:10px; padding:9px 14px;
      border-radius:var(--r-md); color:#4A5568; text-decoration:none;
      font-size:13px; font-weight:500; transition:.15s;
    }}
    .sidebar-logout:hover {{ background:rgba(239,68,68,.1); color:var(--clr-danger); }}
    .sidebar-logout i {{ width:16px; height:16px; }}
    .icon-button {{
      width:38px; height:38px; border-radius:var(--r-md); background:none;
      border:1.5px solid var(--bdr); display:flex; align-items:center;
      justify-content:center; cursor:pointer; color:var(--txt-muted);
      transition:.15s; position:relative; text-decoration:none; flex-shrink:0;
    }}
    .icon-button:hover {{ background:var(--bg-subtle); color:var(--txt-primary); border-color:var(--bdr-input); }}
    .icon-button i {{ width:17px; height:17px; }}
    .notification-button {{ position:relative; }}
    .notification-count {{
      position:absolute; top:-4px; left:-4px; min-width:17px; height:17px;
      background:var(--clr-danger); color:#fff; border-radius:20px;
      font-size:9px; font-weight:700; display:flex; align-items:center;
      justify-content:center; padding:0 4px; border:2px solid var(--bg-card);
    }}

    /* ── Overlay ──────────────────────────────────────────────── */
    .overlay {{ display:none; position:fixed; inset:0; background:rgba(0,0,0,.5); z-index:299; backdrop-filter:blur(2px); }}
    .overlay.open {{ display:block; }}

    /* ── Topbar ───────────────────────────────────────────────── */
    .topbar {{
      position:fixed; top:0; right:var(--sidebar-w); left:0; height:var(--topbar-h);
      background:rgba(255,255,255,.9); backdrop-filter:blur(20px);
      border-bottom:1px solid var(--bdr); z-index:200;
      display:grid; grid-template-columns:auto minmax(0,1fr) auto;
      align-items:center; gap:var(--sp-4); padding:0 24px;
    }}
    .topbar-context {{ display:flex; align-items:center; gap:12px; min-width:0; }}
    .topbar-menu {{ display:none; }}
    .topbar-eyebrow {{ font-size:9px; font-weight:700; letter-spacing:1.5px; color:var(--txt-xmuted); text-transform:uppercase; }}
    .topbar-title {{ font-size:15px; font-weight:700; color:var(--txt-primary); white-space:nowrap; overflow:hidden; text-overflow:ellipsis; }}
    .global-search-wrap {{
      position:relative; display:flex; align-items:center; justify-self:center; width:100%; max-width:480px;
    }}
    .global-search-wrap i {{ position:absolute; right:13px; top:50%; transform:translateY(-50%); width:15px; height:15px; color:var(--txt-muted); pointer-events:none; }}
    .global-search {{
      width:100%; height:38px; background:var(--bg-subtle); border:1.5px solid var(--bdr);
      border-radius:var(--r-md); padding:0 40px 0 36px; font-size:13px; color:var(--txt-primary);
      outline:none; transition:.2s; direction:rtl;
    }}
    .global-search:focus {{ border-color:var(--clr-primary); background:#fff; box-shadow:0 0 0 3px var(--clr-primary-dim); }}
    .global-search-wrap kbd {{
      position:absolute; left:10px; top:50%; transform:translateY(-50%);
      font-size:10px; color:var(--txt-xmuted); background:var(--bg-subtle);
      border:1px solid var(--bdr); border-radius:5px; padding:1px 5px; font-family:inherit;
    }}
    .topbar-actions {{ display:flex; align-items:center; gap:6px; flex-shrink:0; }}
    .profile-trigger {{
      display:flex; align-items:center; gap:8px; padding:5px 10px 5px 8px;
      border-radius:var(--r-md); border:1.5px solid var(--bdr); cursor:pointer;
      text-decoration:none; color:var(--txt-primary); transition:.15s;
    }}
    .profile-trigger:hover {{ background:var(--bg-subtle); }}
    .profile-avatar {{
      width:28px; height:28px; border-radius:var(--r-sm);
      background:linear-gradient(135deg,var(--clr-primary),#0066CC);
      display:flex; align-items:center; justify-content:center; flex-shrink:0;
    }}
    .profile-avatar i {{ width:15px; height:15px; color:#fff; }}
    .profile-copy {{ display:flex; flex-direction:column; gap:1px; }}
    .profile-copy strong {{ font-size:12px; font-weight:600; color:var(--txt-primary); }}
    .profile-copy small {{ font-size:10px; color:var(--txt-muted); }}
    .profile-chevron {{ width:13px !important; height:13px !important; color:var(--txt-muted); flex-shrink:0; }}

    /* ── Flash Messages ───────────────────────────────────────── */
    .flash-msg {{
      display:flex; align-items:center; gap:10px; padding:12px 16px; border-radius:var(--r-md);
      margin-bottom:var(--sp-5); font-size:13.5px; font-weight:500;
      animation:slideIn .25s ease;
    }}
    .flash-msg i {{ width:17px; height:17px; flex-shrink:0; }}
    .flash-ok  {{ background:#F0FDF4; border:1.5px solid #BBF7D0; color:#166534; }}
    .flash-err {{ background:#FEF2F2; border:1.5px solid #FECACA; color:#991B1B; }}
    @keyframes slideIn {{ from {{ opacity:0; transform:translateY(-8px); }} to {{ opacity:1; transform:translateY(0); }} }}

    /* ── Page Header ──────────────────────────────────────────── */
    .page-header {{ margin-bottom:var(--sp-5); }}
    .page-header h1 {{ font-size:20px; font-weight:800; color:var(--txt-primary); line-height:1.2; margin:0 0 4px; }}
    .page-header p  {{ font-size:13px; color:var(--txt-muted); margin:0; }}

    /* ── Cards ────────────────────────────────────────────────── */
    .card {{
      background:var(--bg-card); border-radius:var(--r-lg);
      box-shadow:var(--shadow-card); border:1px solid rgba(229,231,235,.6);
      transition:box-shadow .2s;
    }}
    .card:hover {{ box-shadow:var(--shadow-hover); }}
    .card-p {{ padding:var(--sp-6) var(--sp-7); }}
    .card-p h2, .card-p .t-card {{ margin-bottom:var(--sp-5); }}

    /* ── Stat/KPI Cards ───────────────────────────────────────── */
    .stat-grid {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(160px,1fr)); gap:var(--sp-4); margin-bottom:var(--sp-6); }}
    .stat-card {{
      background:var(--bg-card); border-radius:var(--r-lg);
      padding:var(--sp-5) var(--sp-6); box-shadow:var(--shadow-card);
      border:1px solid rgba(229,231,235,.6); transition:.2s;
    }}
    .stat-card:hover {{ transform:translateY(-2px); box-shadow:var(--shadow-hover); }}

    /* ── Tables ───────────────────────────────────────────────── */
    .table-card {{ background:var(--bg-card); border-radius:var(--r-lg); box-shadow:var(--shadow-card); border:1px solid rgba(229,231,235,.6); overflow:hidden; }}
    .table-head {{ display:flex; align-items:center; justify-content:space-between; padding:16px 20px; border-bottom:1px solid var(--bdr); gap:12px; }}
    .table-head h2 {{ font-size:14px; font-weight:700; color:var(--txt-primary); margin:0; }}
    .table-wrap {{ overflow-x:auto; -webkit-overflow-scrolling:touch; }}
    table {{ width:100%; border-collapse:collapse; min-width:560px; }}
    thead th {{
      padding:10px 16px; font-size:10.5px; color:var(--txt-muted); font-weight:700;
      text-align:right; background:var(--bg-subtle); border-bottom:1.5px solid var(--bdr);
      text-transform:uppercase; letter-spacing:.4px; white-space:nowrap;
    }}
    tbody td {{
      padding:11px 16px; font-size:13px; color:var(--txt-secondary); border-bottom:1px solid #F3F4F6;
      vertical-align:middle;
    }}
    tbody tr:last-child td {{ border-bottom:none; }}
    tbody tr:hover td {{ background:#FAFBFC; }}

    /* Mobile: table → horizontal scroll */
    @media (max-width:640px) {{
      table {{ min-width:500px; }}
      tbody td, thead th {{ padding:9px 12px; font-size:12px; }}
    }}

    /* ── Buttons ──────────────────────────────────────────────── */
    .btn {{
      display:inline-flex; align-items:center; justify-content:center; gap:7px;
      min-height:40px; padding:0 18px; border-radius:var(--r-md);
      font-size:13px; font-weight:600; font-family:var(--font);
      border:none; cursor:pointer; transition:.15s; text-decoration:none;
      white-space:nowrap; flex-shrink:0;
    }}
    .btn i {{ width:15px; height:15px; flex-shrink:0; }}
    .btn-sm {{ min-height:32px; padding:0 12px; font-size:12px; border-radius:var(--r-sm); gap:5px; }}
    .btn-sm i {{ width:13px; height:13px; }}
    .btn-primary {{ background:var(--clr-primary); color:#fff; }}
    .btn-primary:hover {{ opacity:.88; }}
    .btn-success, .btn-green {{ background:var(--clr-success-dim); color:#166534; border:1.5px solid #BBF7D0; }}
    .btn-success:hover, .btn-green:hover {{ background:#BBFBD0; }}
    .btn-danger, .btn-red {{ background:var(--clr-danger-dim); color:#991B1B; border:1.5px solid #FECACA; }}
    .btn-danger:hover, .btn-red:hover {{ background:#FDD0D0; }}
    .btn-indigo {{ background:var(--clr-info-dim); color:#3730A3; border:1.5px solid #C7D2FE; }}
    .btn-indigo:hover {{ background:#E0E7FF; }}
    .btn-slate {{ background:var(--bg-subtle); color:var(--txt-muted); border:1.5px solid var(--bdr); }}
    .btn-slate:hover {{ background:#F1F5F9; color:var(--txt-primary); }}
    .btn-warning {{ background:var(--clr-warning-dim); color:#92400E; border:1.5px solid #FDE68A; }}

    @media (max-width:640px) {{ .btn {{ min-height:44px; padding:0 16px; }} .btn-sm {{ min-height:36px; }} }}

    /* ── Badges ───────────────────────────────────────────────── */
    .badge {{
      display:inline-flex; align-items:center; gap:4px;
      padding:3px 10px; border-radius:20px;
      font-size:11px; font-weight:600; white-space:nowrap; line-height:1.4;
    }}
    .badge i {{ width:10px; height:10px; }}
    .badge-dot {{ width:6px; height:6px; border-radius:50%; flex-shrink:0; }}
    .badge-primary {{ background:var(--clr-primary-dim); color:#0891B2; }}
    .badge-success {{ background:var(--clr-success-dim); color:#15803D; }}
    .badge-danger   {{ background:var(--clr-danger-dim); color:#B91C1C; }}
    .badge-warning  {{ background:var(--clr-warning-dim); color:#B45309; }}
    .badge-info     {{ background:var(--clr-info-dim); color:#1D4ED8; }}
    .badge-gray, .badge-neutral {{ background:var(--clr-neutral-dim); color:#374151; }}

    /* Status badges (ticket) */
    .status-badge {{ display:inline-flex; align-items:center; gap:5px; padding:3px 10px; border-radius:20px; font-size:11px; font-weight:600; }}
    .status-badge span {{ width:6px; height:6px; border-radius:50%; flex-shrink:0; }}
    .status-danger {{ background:var(--clr-danger-dim); color:#B91C1C; }} .status-danger span {{ background:var(--clr-danger); }}
    .status-warning {{ background:var(--clr-warning-dim); color:#B45309; }} .status-warning span {{ background:var(--clr-warning); }}
    .status-success {{ background:var(--clr-success-dim); color:#15803D; }} .status-success span {{ background:var(--clr-success); }}
    .status-neutral {{ background:var(--clr-neutral-dim); color:#374151; }} .status-neutral span {{ background:var(--clr-neutral); }}
    .status-info {{ background:var(--clr-info-dim); color:#1D4ED8; }} .status-info span {{ background:var(--clr-info); }}

    /* ── Forms ────────────────────────────────────────────────── */
    label {{ font-size:12.5px; color:var(--txt-secondary); display:block; margin-bottom:var(--sp-2); font-weight:600; }}
    input:not([type=checkbox]):not([type=radio]):not([type=range]),
    textarea, select {{
      width:100%; min-height:44px; border:1.5px solid var(--bdr-input);
      border-radius:var(--r-md); padding:10px 16px 10px 14px;
      font-size:16px !important; background:var(--bg-input); color:var(--txt-primary);
      outline:none; transition:border .18s, box-shadow .18s;
      direction:rtl; text-align:right; font-family:var(--font);
      -webkit-appearance:none; appearance:none;
    }}
    input:not([type=checkbox]):not([type=radio]):not([type=range]):focus,
    textarea:focus, select:focus {{
      border-color:var(--clr-primary);
      box-shadow:0 0 0 3px rgba(46,196,182,.12);
      background:#fff;
    }}
    textarea {{ min-height:110px; resize:vertical; }}
    select {{ background-image:url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='12' height='8' viewBox='0 0 12 8'%3E%3Cpath d='M1 1l5 5 5-5' stroke='%236B7280' stroke-width='1.5' fill='none' stroke-linecap='round' stroke-linejoin='round'/%3E%3C/svg%3E"); background-repeat:no-repeat; background-position:left 14px center; padding-left:36px; }}
    input[type=checkbox], input[type=radio] {{
      width:16px !important; height:16px !important; min-height:16px !important;
      padding:0 !important; border-radius:5px !important; cursor:pointer; flex-shrink:0;
      -webkit-appearance:auto; appearance:auto;
    }}
    input[type=range] {{
      min-height:unset !important; padding:0 !important; border:none !important;
      background:none !important; border-radius:0 !important; -webkit-appearance:auto; appearance:auto;
    }}
    input::placeholder, textarea::placeholder {{ color:var(--txt-xmuted); opacity:1; }}
    .perm-grid {{ display:flex; flex-wrap:wrap; gap:6px 18px; padding:14px; background:var(--bg-subtle); border-radius:var(--r-md); }}
    .perm-label {{ display:inline-flex; align-items:center; gap:7px; font-size:13px; font-weight:500; color:var(--txt-primary); cursor:pointer; white-space:nowrap; }}
    .perm-label input {{ margin:0; }}

    /* ── Sidebar Classic Mode ─────────────────────────────────── */
    body.sl-classic .sidebar {{ background:#FFFFFF !important; border-left:1px solid var(--bdr) !important; box-shadow:none !important; }}
    body.sl-classic .nav-item {{ color:#374151 !important; }}
    body.sl-classic .nav-item:hover {{ background:#F3F4F6 !important; color:#111827 !important; }}
    body.sl-classic .nav-item.active {{ background:#EFF6FF !important; color:#1D4ED8 !important; box-shadow:inset 3px 0 0 #3B82F6 !important; }}
    body.sl-classic .nav-item.active i {{ filter:none !important; }}
    body.sl-classic .sidebar-header {{ border-bottom-color:var(--bdr) !important; }}
    body.sl-classic .brand-text-only {{ color:#111827 !important; }}
    body.sl-classic .brand-text-only small {{ color:#9CA3AF !important; }}
    body.sl-classic .nav-divider span {{ color:#CBD5E1 !important; }}
    body.sl-classic .nav-divider::after {{ background:var(--bdr) !important; }}
    body.sl-classic .sidebar-footer {{ border-top-color:var(--bdr) !important; }}
    body.sl-classic .sidebar-status {{ color:#6B7280 !important; }}
    body.sl-classic .sidebar-logout {{ color:#6B7280 !important; }}
    body.sl-classic .sidebar-logout:hover {{ background:#FEF2F2 !important; color:#DC2626 !important; }}

    /* فاز ۵: اعداد جدول‌ها چپ‌چین (خوانایی مالی) */
    table td.no-fa, table td[style*="direction:ltr"] {{ text-align:left; direction:ltr; font-variant-numeric:tabular-nums; }}

    /* ── Dark Mode ────────────────────────────────────────────── */
    body.sl-dark, body.dark-mode {{
      --bg-page:#0E1621; --bg-card:#17212B; --bg-input:#1B2530; --bg-subtle:#232E3C;
      --txt-primary:#F5F5F5; --txt-secondary:#D9E1EA; --txt-muted:#8A99AC;
      --bdr:#2B3A4C; --bdr-input:#2B3A4C;
      --shadow-card:0 1px 4px rgba(0,0,0,.35);
      /* نام‌های قدیمی — هنوز جاهای زیادی (مخصوصاً داشبورد) استفاده می‌شن */
      --text-main:#F5F5F5; --text-muted:#8A99AC; --card-bg:#17212B; --border:#2B3A4C; --primary:#2EC4B6;
      background:#0E1621; color:#F5F5F5;
    }}
    body.sl-dark .topbar, body.dark-mode .topbar {{ background:rgba(14,22,33,.92) !important; border-color:#2B3A4C !important; }}
    body.sl-dark .global-search, body.dark-mode .global-search {{ background:#17212B !important; border-color:#2B3A4C !important; color:#F5F5F5 !important; }}
    /* ── کنتراست حالت شب — کیفیت تلگرام ── */
    body.sl-dark .card, body.dark-mode .card {{ background:#17212B !important; border-color:#2B3A4C !important; }}
    body.sl-dark .bg-white, body.dark-mode .bg-white {{ background:#17212B !important; }}
    body.sl-dark .bg-gray-50, body.dark-mode .bg-gray-50 {{ background:#1B2530 !important; }}
    body.sl-dark .bg-gray-100, body.dark-mode .bg-gray-100 {{ background:#232E3C !important; }}
    body.sl-dark .text-gray-800, body.sl-dark .text-gray-900,
    body.dark-mode .text-gray-800, body.dark-mode .text-gray-900 {{ color:#F5F5F5 !important; }}
    body.sl-dark .text-gray-700, body.dark-mode .text-gray-700 {{ color:#E4EAF1 !important; }}
    body.sl-dark .text-gray-600, body.sl-dark .text-gray-500,
    body.dark-mode .text-gray-600, body.dark-mode .text-gray-500 {{ color:#A9B6C6 !important; }}
    body.sl-dark .text-gray-400, body.dark-mode .text-gray-400 {{ color:#8A99AC !important; }}
    body.sl-dark .border, body.sl-dark .border-b, body.sl-dark .border-t,
    body.sl-dark .border-gray-200, body.sl-dark .border-gray-300, body.sl-dark [class*="border-gray-1"],
    body.dark-mode .border, body.dark-mode .border-b, body.dark-mode .border-t,
    body.dark-mode .border-gray-200, body.dark-mode .border-gray-300 {{ border-color:#2B3A4C !important; }}
    body.sl-dark input, body.sl-dark select, body.sl-dark textarea,
    body.dark-mode input, body.dark-mode select, body.dark-mode textarea {{
      background:#1B2530 !important; color:#F0F4F8 !important; border-color:#2B3A4C !important;
    }}
    body.sl-dark input::placeholder, body.sl-dark textarea::placeholder,
    body.dark-mode input::placeholder, body.dark-mode textarea::placeholder {{ color:#6B7B8F !important; }}
    body.sl-dark tr:hover, body.sl-dark .hover\\:bg-gray-50:hover, body.sl-dark .hover\\:bg-gray-100:hover,
    body.dark-mode tr:hover, body.dark-mode .hover\\:bg-gray-50:hover, body.dark-mode .hover\\:bg-gray-100:hover {{ background:#1F2A38 !important; }}
    body.sl-dark thead tr, body.dark-mode thead tr {{ background:#1B2530 !important; }}
    body.sl-dark code, body.dark-mode code {{ background:#232E3C; color:#8FD3F4; padding:1px 5px; border-radius:5px; }}
    /* بج‌های رنگی: پس‌زمینه شفاف تیره + متن روشن‌تر */
    body.sl-dark [class*="bg-green-100"], body.sl-dark [class*="bg-green-50"] {{ background:rgba(34,197,94,.16) !important; }}
    body.sl-dark [class*="text-green-7"] {{ color:#5DDE8A !important; }}
    body.sl-dark [class*="bg-red-100"], body.sl-dark [class*="bg-red-50"] {{ background:rgba(239,68,68,.16) !important; }}
    body.sl-dark [class*="text-red-6"], body.sl-dark [class*="text-red-7"], body.sl-dark [class*="text-red-5"] {{ color:#FF7B7B !important; }}
    body.sl-dark [class*="bg-amber-100"], body.sl-dark [class*="bg-amber-50"],
    body.sl-dark [class*="bg-yellow-100"], body.sl-dark [class*="bg-yellow-50"] {{ background:rgba(245,158,11,.16) !important; }}
    body.sl-dark [class*="text-amber-7"], body.sl-dark [class*="text-amber-8"], body.sl-dark [class*="text-yellow-7"] {{ color:#FFC46B !important; }}
    body.sl-dark [class*="bg-indigo-100"], body.sl-dark [class*="bg-indigo-50"] {{ background:rgba(99,102,241,.18) !important; }}
    body.sl-dark [class*="text-indigo-7"], body.sl-dark [class*="text-indigo-6"] {{ color:#A5B4FF !important; }}
    body.sl-dark [class*="bg-blue-100"], body.sl-dark [class*="bg-blue-50"] {{ background:rgba(59,130,246,.16) !important; }}
    body.sl-dark [class*="text-blue-7"], body.sl-dark [class*="text-blue-6"] {{ color:#7DB8FF !important; }}
    body.sl-dark [class*="bg-teal-100"], body.sl-dark [class*="bg-teal-50"] {{ background:rgba(20,184,166,.16) !important; }}
    body.sl-dark [class*="text-teal-7"] {{ color:#5EEAD4 !important; }}
    body.sl-dark [class*="bg-purple-100"], body.sl-dark [class*="bg-purple-50"] {{ background:rgba(168,85,247,.16) !important; }}
    body.sl-dark [class*="text-purple-7"] {{ color:#D0A8FF !important; }}
    body.sl-dark [class*="bg-orange-100"], body.sl-dark [class*="bg-orange-50"] {{ background:rgba(249,115,22,.16) !important; }}
    body.sl-dark [class*="text-orange-7"] {{ color:#FFAD70 !important; }}
    body.sl-dark [class*="border-green-2"], body.sl-dark [class*="border-red-2"],
    body.sl-dark [class*="border-amber-2"], body.sl-dark [class*="border-indigo-2"],
    body.sl-dark [class*="border-blue-2"], body.sl-dark [class*="border-teal-2"] {{ border-color:#2B3A4C !important; }}
    /* فاز ۵: تکمیل Dark Mode — رنگ‌های جامانده + سازگاری dark-mode alias */
    body.sl-dark [class*="bg-emerald-50"], body.dark-mode [class*="bg-emerald-50"],
    body.sl-dark [class*="bg-emerald-100"], body.dark-mode [class*="bg-emerald-100"] {{ background:rgba(16,185,129,.14) !important; }}
    body.sl-dark [class*="text-emerald-6"], body.sl-dark [class*="text-emerald-7"],
    body.dark-mode [class*="text-emerald-6"], body.dark-mode [class*="text-emerald-7"] {{ color:#4EE0B0 !important; }}
    body.sl-dark [class*="bg-slate-50"], body.dark-mode [class*="bg-slate-50"],
    body.sl-dark [class*="bg-slate-100"], body.dark-mode [class*="bg-slate-100"] {{ background:#1B2530 !important; }}
    body.sl-dark [class*="text-slate-7"], body.dark-mode [class*="text-slate-7"],
    body.sl-dark [class*="text-slate-6"], body.dark-mode [class*="text-slate-6"] {{ color:#B5C3D4 !important; }}
    body.sl-dark [class*="bg-purple-50"], body.dark-mode [class*="bg-purple-50"] {{ background:rgba(168,85,247,.14) !important; }}
    body.dark-mode [class*="text-purple-7"], body.dark-mode [class*="text-purple-6"] {{ color:#D0A8FF !important; }}
    body.dark-mode [class*="bg-green-100"], body.dark-mode [class*="bg-green-50"] {{ background:rgba(34,197,94,.16) !important; }}
    body.dark-mode [class*="text-green-7"], body.dark-mode [class*="text-green-6"] {{ color:#5DDE8A !important; }}
    body.dark-mode [class*="bg-red-100"], body.dark-mode [class*="bg-red-50"] {{ background:rgba(239,68,68,.16) !important; }}
    body.dark-mode [class*="text-red-6"], body.dark-mode [class*="text-red-7"] {{ color:#FF7B7B !important; }}
    body.dark-mode [class*="bg-amber-100"], body.dark-mode [class*="bg-amber-50"] {{ background:rgba(245,158,11,.16) !important; }}
    body.dark-mode [class*="text-amber-7"] {{ color:#FFC46B !important; }}
    body.dark-mode [class*="bg-indigo-100"], body.dark-mode [class*="bg-indigo-50"] {{ background:rgba(99,102,241,.18) !important; }}
    body.dark-mode [class*="text-indigo-7"], body.dark-mode [class*="text-indigo-6"] {{ color:#A5B4FF !important; }}
    body.dark-mode [class*="bg-blue-100"], body.dark-mode [class*="bg-blue-50"] {{ background:rgba(59,130,246,.16) !important; }}
    body.dark-mode [class*="text-blue-7"], body.dark-mode [class*="text-blue-6"] {{ color:#7DB8FF !important; }}
    body.dark-mode [class*="bg-teal-50"] {{ background:rgba(20,184,166,.16) !important; }}
    body.dark-mode [class*="text-teal-7"] {{ color:#5EEAD4 !important; }}
    /* گرادیان‌ها در دارک */
    body.sl-dark [class*="bg-gradient"], body.dark-mode [class*="bg-gradient"] {{ background:#17212B !important; }}
    /* جداول: اعداد چپ‌چین حفظ خوانایی */
    body.sl-dark table td, body.dark-mode table td {{ color:#D9E1EA; }}
    body.sl-dark summary, body.dark-mode summary {{ color:#E4EAF1 !important; }}
    /* فیکس: اعداد و متون مشکی در dark */
    body.sl-dark input, body.sl-dark select, body.sl-dark textarea,
    body.dark-mode input, body.dark-mode select, body.dark-mode textarea {{
      color:#E4EAF1 !important; background:#1B2530 !important; border-color:#2B3A4C !important;
    }}
    body.sl-dark td, body.sl-dark th, body.dark-mode td, body.dark-mode th {{ color:#D9E1EA !important; }}
    body.sl-dark .text-gray-800, body.sl-dark .text-gray-700, body.sl-dark .text-gray-600,
    body.dark-mode .text-gray-800, body.dark-mode .text-gray-700, body.dark-mode .text-gray-600 {{ color:#D0DAE6 !important; }}
    body.sl-dark .text-gray-500, body.sl-dark .text-gray-400,
    body.dark-mode .text-gray-500, body.dark-mode .text-gray-400 {{ color:#8899AA !important; }}
    body.sl-dark .command-center::before, body.dark-mode .command-center::before {{ display:none !important; }}
    body.sl-dark .command-center, body.dark-mode .command-center {{ background:#0F1923 !important; }}
    body.sl-dark .dashboard, body.dark-mode .dashboard {{ background:#0F1923 !important; }}
    body.sl-dark details.acc, body.dark-mode details.acc {{ background:#17212B !important; }}

    /* ── Misc Helpers ─────────────────────────────────────────── */
    .section-title {{ font-size:14px; font-weight:700; color:var(--txt-primary); margin-bottom:16px; padding-bottom:12px; border-bottom:1px solid var(--bdr); }}
    .empty-state {{ text-align:center; padding:var(--sp-8) var(--sp-6); color:var(--txt-muted); font-size:14px; }}
    .truncate {{ overflow:hidden; text-overflow:ellipsis; white-space:nowrap; }}.no-wrap {{ white-space:nowrap; }}
    .gap-2 {{ gap:8px; }}.gap-3 {{ gap:12px; }}.gap-4 {{ gap:16px; }}
    .mb-4 {{ margin-bottom:16px; }}.mb-5 {{ margin-bottom:20px; }}.mb-6 {{ margin-bottom:24px; }}
    .legacy-lucide {{ width:1em; height:1em; display:inline-block; vertical-align:-.16em; stroke-width:1.9; }}

    /* ── Sidebar Brand ────────────────────────────────────────── */
    .sidebar-header {{ justify-content:center; position:relative; }}
    .brand-lockup {{ margin:0 auto; }}
    .brand-text-only {{ direction:ltr; text-align:center; }}
    .brand-word {{ font-weight:900; letter-spacing:1.5px; }}
    .brand-word--stock {{ color:#E8EDF2; }}
    .brand-word--land  {{ color:var(--clr-primary); }}
    .brand-subtitle {{ display:block; font-size:9.5px; color:#536075; letter-spacing:.8px; font-weight:400; margin-top:3px; direction:rtl; }}
    .sb-col-btn {{
      position:absolute; left:-11px; top:50%; transform:translateY(-50%);
      width:22px; height:22px; border-radius:50%; background:var(--clr-primary);
      border:none; cursor:pointer; color:#fff; font-size:12px;
      display:flex; align-items:center; justify-content:center;
      z-index:60; box-shadow:0 2px 8px rgba(0,0,0,.3);
    }}
    .sb-toggle-icon {{ transition:transform .2s; display:inline-block; }}

    /* ── Topbar badge variants + profile ─────────────────────── */
    .notification-count--warn {{ background:#F59E0B; }}
    .notification-count--danger {{ background:var(--clr-danger); }}
    .profile-trigger {{ text-decoration:none; }}
    .profile-chevron--sm {{ width:14px; height:14px; opacity:.5; }}

    /* ── Login Page ────────────────────────────────────────────── */
    .login-wrap {{ min-height:80vh; display:flex; align-items:center; justify-content:center; }}
    .login-card {{ background:var(--bg-card); border-radius:var(--r-xl); box-shadow:var(--shadow-modal); padding:40px 36px; width:100%; max-width:380px; }}
    .login-header {{ text-align:center; margin-bottom:28px; }}
    .login-brand {{ font-size:24px; font-weight:900; letter-spacing:1.5px; margin-bottom:6px; direction:ltr; }}
    .login-brand .brand-word--stock {{ color:var(--txt-secondary); }}
    .login-subtitle {{ font-size:12px; color:var(--txt-muted); }}
    .login-form {{ display:flex; flex-direction:column; gap:14px; }}
    .field-label {{ font-size:12px; font-weight:600; color:var(--txt-muted); display:block; margin-bottom:6px; }}
    .login-submit {{
      width:100%; padding:12px; background:var(--primary); color:#fff;
      font-weight:700; font-size:14px; border:none; border-radius:var(--r-md);
      cursor:pointer; margin-top:4px;
    }}
    .alert-error {{ background:var(--clr-danger-dim); border:1px solid #FECACA; color:#991B1B; padding:12px 16px; border-radius:10px; font-size:13px; margin-bottom:14px; }}
    .alert-warning {{ background:var(--clr-warning-dim); border:1px solid #FDE68A; color:#92400E; padding:12px 16px; border-radius:10px; font-size:13px; margin-bottom:14px; }}
    body.sl-dark .alert-error, body.dark-mode .alert-error {{ background:rgba(239,68,68,.15); border-color:rgba(239,68,68,.3); color:#FCA5A5; }}
    body.sl-dark .alert-warning, body.dark-mode .alert-warning {{ background:rgba(245,158,11,.15); border-color:rgba(245,158,11,.3); color:#FDE68A; }}
    body.sl-dark .login-card, body.dark-mode .login-card {{ background:var(--bg-card); }}
    body.sl-dark .login-brand .brand-word--stock, body.dark-mode .login-brand .brand-word--stock {{ color:#E8EDF2; }}

    /* ── Settings / Account Forms ──────────────────────────────── */
    .settings-header {{ display:flex; align-items:center; justify-content:space-between; flex-wrap:wrap; gap:10px; margin-bottom:22px; }}
    .form-section {{ display:flex; flex-direction:column; gap:14px; }}
    .form-card {{ padding:var(--sp-5); margin-bottom:var(--sp-4); }}
    .form-card-title {{ font-size:14px; font-weight:700; margin-bottom:20px; padding-bottom:14px; border-bottom:1px solid var(--bdr); }}
    .field-icon {{ font-size:1.1rem; width:22px; text-align:center; flex-shrink:0; }}
    .field-icon-sm {{ font-size:1rem; width:20px; text-align:center; flex-shrink:0; }}
    .btn-link-muted {{ font-size:.76rem; color:var(--txt-xmuted); background:none; border:none; cursor:pointer; transition:color .15s; padding:0; }}
    .btn-link-muted:hover {{ color:#ef4444; }}
    .disabled-look {{ background:var(--bg-page); color:var(--txt-muted); cursor:not-allowed; }}
    .flex-col {{ display:flex; flex-direction:column; }}
    .flex-col-6 {{ display:flex; flex-direction:column; gap:6px; }}
    .grid-2col {{ display:grid; grid-template-columns:1fr 1fr; gap:16px; }}
    @media(max-width:640px){{ .grid-2col {{ grid-template-columns:1fr; }} }}
    .settings-title {{ font-size:1.4rem; font-weight:700; color:var(--txt-primary); margin:0; }}
    .footer-bar {{ display:flex; align-items:center; justify-content:space-between; padding-bottom:32px; flex-wrap:wrap; gap:10px; }}
    .text-section-label {{ display:block; font-size:.82rem; font-weight:600; color:var(--txt-secondary); margin-bottom:6px; }}
    .input-full {{ width:100%; box-sizing:border-box; }}
    .textarea-full {{ width:100%; box-sizing:border-box; resize:vertical; min-height:110px; }}
    .card-reset-footer {{ margin-top:12px; padding-top:12px; border-top:1px solid var(--bdr); display:flex; justify-content:flex-end; }}
    .mb-18 {{ margin-bottom:18px; }}
    .mb-14 {{ margin-bottom:14px; }}
    .mt-14 {{ margin-top:14px; }}
    .flex-col-10 {{ display:flex; flex-direction:column; gap:10px; }}
    .max-w-640 {{ max-width:640px; }}
    .btn-row-12 {{ display:flex; gap:12px; }}
    .info-box {{ background:var(--page-bg); border-radius:10px; padding:10px 14px; font-size:12.5px; margin-bottom:14px; }}

    /* ── Backup Modal ──────────────────────────────────────────── */
    .bk-overlay {{ display:none; position:fixed; top:0; left:0; width:100%; height:100%; background:rgba(0,0,0,.6); z-index:99999; }}
    .bk-modal {{ position:absolute; top:50%; left:50%; transform:translate(-50%,-50%); background:var(--bg-card); border-radius:var(--r-xl); padding:36px 24px; width:calc(100% - 48px); max-width:340px; text-align:center; box-shadow:var(--shadow-modal); }}
    .bk-icon {{ font-size:52px; margin-bottom:12px; line-height:1; }}
    .bk-title {{ font-weight:700; color:var(--txt-primary); font-size:17px; margin:0 0 6px 0; font-family:inherit; }}
    .bk-msg {{ font-size:12px; color:var(--txt-muted); margin:0 0 18px 0; }}
    .bk-track {{ direction:ltr; background:var(--bdr); border-radius:999px; height:8px; overflow:hidden; margin-bottom:20px; }}
    .bk-bar {{ background:#6366f1; height:8px; border-radius:999px; width:5%; transition:width .5s ease; }}
    .bk-done-btn {{ display:none; padding:11px 36px; background:#6366f1; color:#fff; border:none; border-radius:var(--r-md); font-size:14px; font-weight:600; cursor:pointer; font-family:inherit; }}

    /* ── Chat Bubbles (Ticket) ─────────────────────────────────── */
    .chat-row {{ display:flex; margin-bottom:10px; }}
    .chat-row--admin {{ justify-content:flex-end; }}
    .chat-row--user {{ justify-content:flex-start; }}
    .chat-col {{ max-width:80%; }}
    .chat-bubble {{ padding:10px 14px; font-size:13.5px; word-break:break-word; white-space:pre-wrap; }}
    .chat-bubble--admin {{ background:var(--primary); color:#fff; border-radius:18px 4px 18px 18px; box-shadow:0 1px 2px rgba(0,0,0,.1); }}
    .chat-bubble--user {{ background:var(--bg-card); border:1px solid var(--bdr); border-radius:4px 18px 18px 18px; box-shadow:0 1px 2px rgba(0,0,0,.06); }}
    body.sl-dark .chat-bubble--user, body.dark-mode .chat-bubble--user {{ background:#1B2530; border-color:#2B3A4C; }}
    .chat-meta {{ font-size:10px; color:#aaa; margin-top:3px; }}
    .chat-meta--admin {{ text-align:left; }}
    .chat-media-img {{ max-width:260px; max-height:200px; border-radius:10px; display:block; }}
    .chat-media-video {{ max-width:280px; max-height:200px; border-radius:10px; }}
    .chat-media-audio {{ max-width:260px; }}
    .chat-media-file {{ display:inline-flex; align-items:center; gap:6px; padding:7px 12px; background:rgba(0,0,0,.06); border-radius:9px; text-decoration:none; color:inherit; font-size:12px; }}
    .chat-empty {{ opacity:.5; font-size:12px; }}
    .chat-caption {{ margin-top:6px; font-size:13px; }}
    .chat-icon-label {{ opacity:.6; font-size:12px; }}
    .chat-empty-msg {{ opacity:.6; }}
    .chat-empty-msg--user {{ opacity:.4; }}

    /* ── Ticket Detail ─────────────────────────────────────────── */
    .ticket-header {{ display:flex; align-items:center; gap:10px; margin-bottom:18px; flex-wrap:wrap; }}
    .ticket-title {{ font-size:17px; font-weight:800; }}
    .ticket-info-grid {{ display:grid; grid-template-columns:auto 1fr; gap:6px 14px; font-size:12px; }}
    .ticket-info-dt {{ color:var(--txt-muted); }}
    .ticket-info-dd-bold {{ font-weight:600; }}
    .ticket-info-code {{ background:var(--bg-subtle); padding:1px 6px; border-radius:5px; }}
    .ticket-setup-card {{ margin-bottom:12px; border:2px solid rgba(46,196,182,.13); }}
    .ticket-setup-title {{ font-size:13px; font-weight:700; margin-bottom:12px; color:#166534; }}
    body.sl-dark .ticket-setup-title, body.dark-mode .ticket-setup-title {{ color:#86EFAC; }}
    .ticket-direct-card {{ border:2px dashed var(--bdr); background:var(--bg-subtle); }}
    .ticket-direct-label {{ font-size:11px; color:var(--txt-muted); margin-bottom:8px; }}
    .ticket-direct-form {{ display:flex; gap:8px; }}
    .ticket-direct-input {{ flex:1; border:1px solid var(--bdr); border-radius:10px; padding:8px 12px; font-size:13px; resize:none; font-family:inherit; }}
    .ticket-direct-btn {{ background:var(--clr-neutral); color:#fff; border:none; border-radius:10px; padding:8px 14px; font-size:12px; cursor:pointer; align-self:flex-end; }}
    .card-title-sm {{ font-size:13px; font-weight:700; margin-bottom:12px; }}
    .card-title-sm-10 {{ font-size:13px; font-weight:700; margin-bottom:10px; }}
    .muted-xs-mb8 {{ font-size:11px; color:var(--txt-muted); margin-bottom:8px; }}
    .flex-wrap-gap6 {{ display:flex; flex-wrap:wrap; gap:6px; }}
    .btn-full-mt12 {{ width:100%; margin-top:12px; }}
    .icon-15 {{ width:15px; }}
    .sidebar-flex-col-12 {{ display:flex; flex-direction:column; gap:12px; }}
    .chat-box {{ min-height:280px; max-height:500px; }}
    .archive-btn {{ width:100%; padding:8px 12px; background:#F3F4F6; color:#6B7280; border:1px solid #E5E7EB; border-radius:10px; font-size:12.5px; font-weight:600; cursor:pointer; }}
    body.sl-dark .archive-btn, body.dark-mode .archive-btn {{ background:#1B2530; color:#8A99AC; border-color:#2B3A4C; }}
    .setup-status-btn--active {{ background:var(--primary); color:#000; font-weight:700; }}
    .ticket-type-label {{ font-size:12px; color:var(--txt-muted); }}
    .d-inline {{ display:inline; }}
    .mt-12 {{ margin-top:12px; }}
    .count-primary {{ font-weight:700; color:var(--clr-primary); }}
    .stat-card-value {{ color:var(--text-main); }}
    .stat-card-sub {{ font-size:12px; color:var(--text-muted); margin-top:3px; }}
    .empty-result {{ display:block; padding:12px; color:var(--txt-muted); font-size:12px; }}
    .max-w-560 {{ max-width:560px; }}
    .ltr-num {{ direction:ltr; text-align:left; }}
    .ltr-left {{ direction:ltr; text-align:left; }}
    .scroll-anchor {{ scroll-margin-top:80px; }}
    .bidi-plain {{ unicode-bidi:plaintext; }}
    .quill-editor-box {{ height:280px; background:#fff; }}
    body.sl-dark .quill-editor-box, body.dark-mode .quill-editor-box {{ background:#17212B; }}
    .min-h-300 {{ min-height:300px; }}
    .tree-overlay-fixed {{ position:fixed; inset:0; background:rgba(0,0,0,.35); z-index:70; }}
    .tree-drawer-fixed {{
      position:fixed; top:0; left:0; height:100vh; width:min(92vw,380px); z-index:400;
      transform:translateX(-105%); transition:transform .25s; overflow-y:auto;
      border-radius:0; padding:0; box-shadow:4px 0 24px rgba(0,0,0,.2);
    }}
    .tree-drawer-header {{ z-index:1; }}
    .disabled-visual {{ opacity:.5; cursor:not-allowed; }}
    .disabled-visual--noclick {{ opacity:.5; pointer-events:none; }}
    .sort-link {{ text-decoration:none; font-size:11px; color:var(--text-muted); }}
    .sort-link--active {{ color:var(--clr-primary); font-weight:700; }}
    .chat-js-bubble {{ max-width:85%; white-space:pre-wrap; }}
    .icon-18 {{ width:18px; }}
    .icon-amber {{ color:#f59e0b; }}
    .icon-cyan {{ color:#0891b2; }}
    .m-0 {{ margin:0; }}
    .flex-gap14 {{ display:flex; align-items:center; gap:14px; }}
    .field-row-sm {{ padding:8px 14px; }}
    .mb-16 {{ margin-bottom:16px; }}
    .mb-20 {{ margin-bottom:20px; }}
    .mb-0 {{ margin-bottom:0; }}

    /* ── Filter Tabs / Pills ───────────────────────────────────── */
    .filter-tabs {{ display:flex; gap:6px; flex-wrap:wrap; }}
    .filter-tab {{
      display:inline-flex; align-items:center; gap:5px; padding:5px 12px;
      border-radius:9px; border:1.5px solid var(--bdr);
      font-size:11px; font-weight:500; text-decoration:none;
      background:var(--bg-card); color:var(--txt-muted);
    }}
    .filter-tab--active {{ border-color:var(--txt-secondary); background:var(--txt-secondary); color:#fff; font-weight:700; }}
    .filter-tab-lg {{
      display:inline-flex; align-items:center; gap:6px; padding:6px 14px;
      border-radius:10px; border:1.5px solid var(--bdr);
      font-size:12px; font-weight:500; text-decoration:none;
      background:var(--bg-card); color:var(--txt-muted);
    }}
    .filter-tab-lg--active {{ border-color:var(--clr-primary); background:var(--clr-primary); color:#fff; font-weight:700; }}
    .filter-tab-count {{ font-size:10px; padding:1px 6px; border-radius:20px; background:var(--bg-subtle); }}
    .filter-tab--active .filter-tab-count,
    .filter-tab-lg--active .filter-tab-count {{ background:rgba(0,0,0,.15); }}

    /* ── Status Badges ─────────────────────────────────────────── */
    .status-pill {{ padding:3px 9px; border-radius:20px; font-size:11px; font-weight:600; }}
    .status-pill-sm {{ padding:2px 7px; border-radius:20px; font-size:10px; font-weight:600; }}

    /* ── Show-More Button ──────────────────────────────────────── */
    .show-more-btn {{
      padding:6px 16px; background:var(--clr-neutral-dim); color:var(--clr-neutral);
      border:1px solid var(--bdr); border-radius:var(--r-sm); font-size:12px;
      cursor:pointer; font-family:inherit;
    }}
    body.sl-dark .show-more-btn, body.dark-mode .show-more-btn {{ background:#1B2530; color:#8A99AC; border-color:#2B3A4C; }}

    /* ── Pagination ────────────────────────────────────────────── */
    .pagination {{ display:flex; gap:6px; justify-content:center; margin-top:16px; flex-wrap:wrap; }}
    .page-link {{ padding:5px 12px; border-radius:var(--r-sm); font-size:12px; text-decoration:none; border:1px solid var(--bdr); color:var(--txt-muted); background:var(--bg-card); }}
    .page-link--active {{ background:var(--clr-primary); color:#000; }}

    /* ── Radio/Checkbox Cards ──────────────────────────────────── */
    .option-card {{ padding:12px; background:var(--page-bg); border-radius:var(--r-md); font-size:13px; }}
    .option-card--success {{ background:#F0FDF4; }}
    body.sl-dark .option-card--success, body.dark-mode .option-card--success {{ background:rgba(34,197,94,.12); }}
    .option-card--danger {{ background:#FEF2F2; }}
    body.sl-dark .option-card--danger, body.dark-mode .option-card--danger {{ background:rgba(239,68,68,.12); }}
    .option-radio, .option-check {{ width:17px; height:17px; min-height:17px; cursor:pointer; }}
    .option-check-sm {{ width:15px; height:15px; min-height:15px; }}
    .option-hint {{ font-size:11.5px; color:var(--txt-muted); margin-top:2px; }}
    .option-toggle-box {{ padding:12px 16px; background:var(--page-bg); border-radius:var(--r-md); }}
    .option-toggle-label {{ font-size:13px; }}
    .option-check-16 {{ width:16px; height:16px; min-height:16px; cursor:pointer; }}
    .option-reveal {{ display:none; margin-top:12px; }}

    /* ── Broadcast Form ────────────────────────────────────────── */
    .broadcast-form {{ display:flex; flex-direction:column; gap:18px; }}
    .label-hint {{ font-size:11px; font-weight:400; color:var(--txt-muted); }}
    .warning-box {{ background:#FEF3C7; border:1px solid #FDE68A; border-radius:12px; padding:12px 16px; font-size:13px; color:#92400E; }}
    body.sl-dark .warning-box, body.dark-mode .warning-box {{ background:rgba(253,230,138,.12); border-color:rgba(253,230,138,.25); color:#FDE68A; }}

    /* ── Progress Bar (generic) ────────────────────────────────── */
    .progress-track {{ direction:ltr; border-radius:9999px; overflow:hidden; }}
    .progress-bar {{ border-radius:9999px; }}

    /* ── Responsive Breakpoints ───────────────────────────────── */
    @media (max-width:1100px) {{
      .topbar {{ grid-template-columns:auto minmax(180px,1fr) auto; padding:0 16px; }}
      .profile-copy, .profile-chevron {{ display:none; }}
    }}
    @media (max-width:820px) {{
      :root {{ --sidebar-w:0px; }}
      .sidebar {{ transform:translateX(100%); }}
      .sidebar.open {{ transform:translateX(0); --sidebar-w:272px; }}
      .overlay.open {{ display:block; }}
      .topbar {{ right:0; }}
      .topbar-menu {{ display:flex !important; align-items:center; justify-content:center; }}
      .main-wrap.with-sidebar {{ margin-right:0 !important; }}
      .main-content {{ padding:var(--sp-4) var(--sp-4) var(--sp-8); }}
      .stat-grid {{ grid-template-columns:repeat(2,1fr); }}
    }}
    @media (max-width:640px) {{
      .topbar {{ grid-template-columns:1fr auto; height:56px; padding:0 12px; }}
      .global-search-wrap {{ display:none; }}
      .topbar-eyebrow {{ display:none; }}
      .topbar-actions {{ gap:4px; }}
      .profile-trigger {{ padding:4px; border:none; background:none; }}
      .main-wrap.with-sidebar {{ padding-top:56px; }}
      .main-content {{ padding:12px 12px 32px; }}
      .sidebar {{ width:min(86vw,280px); }}
      .card {{ border-radius:var(--r-lg); }}
      .card-p {{ padding:var(--sp-4) var(--sp-5); }}
      .stat-grid {{ grid-template-columns:repeat(2,1fr); gap:10px; }}
      .page-header h1 {{ font-size:18px; }}
      .filter-tabs {{ gap:5px; }}
      .filter-tab {{ padding:4px 10px; font-size:10.5px; }}
      .filter-tab-lg {{ padding:5px 11px; font-size:11.5px; }}
      .btn {{ font-size:13px; padding:0 14px; }}
    }}
    @media (max-width:375px) {{
      .main-content {{ padding:10px 10px 28px; }}
      .stat-grid {{ grid-template-columns:1fr 1fr; gap:8px; }}
    }}

    /* ── Safe Area (iPhone notch / home indicator) ──────────── */
    .topbar {{
      padding-right: max(24px, env(safe-area-inset-right));
      padding-left:  max(16px, env(safe-area-inset-left));
    }}
    .sidebar {{
      padding-bottom: env(safe-area-inset-bottom);
    }}
    .main-content {{
      padding-bottom: max(32px, calc(env(safe-area-inset-bottom) + 24px));
    }}
    @media (max-width:820px) {{
      .main-wrap.with-sidebar {{
        padding-top: max(60px, calc(env(safe-area-inset-top) + 56px));
      }}
    }}

    /* ── No horizontal scroll on page ──────────────────────── */
    html, body {{ overflow-x:hidden; max-width:100vw; }}
    .main-wrap {{ overflow-x:hidden; }}
    .table-wrap, .overflow-x-auto {{ overflow-x:auto; -webkit-overflow-scrolling:touch; max-width:100%; }}

    /* ── Touch targets 44px ─────────────────────────────────── */
    @media (max-width:820px) {{
      .btn-sm {{ min-height:36px; padding:0 12px; }}
      a.btn-sm, button.btn-sm {{ display:inline-flex; align-items:center; justify-content:center; }}
      .nav-item {{ min-height:44px; }}
      .icon-button {{ width:44px; height:44px; }}
    }}

    /* ── Prevent text overflow on mobile ────────────────────── */
    @media (max-width:640px) {{
      .truncate-mobile {{ max-width:160px; overflow:hidden; text-overflow:ellipsis; white-space:nowrap; }}
      td {{ word-break:break-word; }}
    }}
  </style>
</head>
<body>
{sidebar}
{topbar}
<div class="main-wrap {'with-sidebar' if admin_info else 'auth-layout'}">
  <div class="main-content">
    {flash_html}
    {body}
  </div>
</div>

<script>
(function(){{
  function renderIcons(){{ if(window.lucide) lucide.createIcons({{attrs:{{'aria-hidden':'true'}}}}); }}

  // اعمال glass level از DB هنگام load
  (function(){{
    var gl = parseFloat('{theme.get("glass", "0") if admin_info else "0"}') || 0;
    if(gl > 0) document.documentElement.style.setProperty('--glass-level', gl);
  }})();

  // Active nav
  var path = location.pathname;
  document.querySelectorAll('.nav-item[data-href]').forEach(function(el){{
    if(el.dataset.href === path || (path !== '/admin/' && el.dataset.href !== '/admin/' && path.startsWith(el.dataset.href))){{
      el.classList.add('active');
    }}
  }});

  // Toggle sidebar
  window.toggleSidebar = function(){{
    document.getElementById('sidebar')?.classList.toggle('open');
    document.getElementById('overlay')?.classList.toggle('open');
  }};

  // ── Collapse sidebar (desktop only) ────────────────────────────────────
  window.sbCollapse = function(){{
    if(window.innerWidth < 769) return;
    var sb = document.querySelector('.sidebar');
    var on = sb.classList.toggle('sb-collapsed');
    document.body.classList.toggle('sb-collapsed-body', on);
    localStorage.setItem('sl-sb-collapsed', on ? '1' : '0');
    var ico = document.getElementById('sb-toggle-icon');
    if(ico) ico.style.transform = on ? 'rotate(180deg)' : '';
  }};
  (function(){{
    if(window.innerWidth < 769) return;
    if(localStorage.getItem('sl-sb-collapsed') === '1') {{
      document.querySelector('.sidebar').classList.add('sb-collapsed');
      document.body.classList.add('sb-collapsed-body');
      var ico = document.getElementById('sb-toggle-icon');
      if(ico) ico.style.transform = 'rotate(180deg)';
    }}
  }})();

  // ── Classic / Dark Mode (با پشتیبانی هماهنگی سیستم) ──────────
  function _prefersDark(){{
    return window.matchMedia && window.matchMedia('(prefers-color-scheme: dark)').matches;
  }}
  function _resolveDark(){{
    var m = localStorage.getItem('sl-dark');
    if(m==='auto' || m===null) return _prefersDark();
    return m==='1';
  }}
  function applyMode(){{
    var isClassic = localStorage.getItem('sl-classic')==='1';
    var isDark = _resolveDark();
    document.body.classList.toggle('sl-classic', isClassic && !isDark);
    document.body.classList.toggle('sl-dark', isDark);
    document.body.classList.toggle('dark-mode', isDark);
  }}
  window.applyMode = applyMode;
  // اگر روی حالت خودکار است، با تغییر تم سیستم همگام شو
  if(window.matchMedia){{
    try {{
      window.matchMedia('(prefers-color-scheme: dark)').addEventListener('change', function(){{
        if((localStorage.getItem('sl-dark')||'auto')==='auto') applyMode();
      }});
    }} catch(e){{}}
  }}
  window.toggleClassic = function(){{
    var on = localStorage.getItem('sl-classic')==='1';
    localStorage.setItem('sl-classic', on?'0':'1');
    applyMode(); renderIcons();
    fetch('/admin/settings/save-theme?dark='+encodeURIComponent(localStorage.getItem('sl-dark')||'auto')+'&classic='+(on?'0':'1'),{{method:'POST'}}).catch(function(){{}});
  }};
  window.toggleDark = function(){{
    var on = _resolveDark();
    localStorage.setItem('sl-dark', on?'0':'1');
    applyMode(); renderIcons();
    fetch('/admin/settings/save-theme?dark='+(on?'0':'1')+'&classic='+(localStorage.getItem('sl-classic')==='1'?'1':'0'),{{method:'POST'}}).catch(function(){{}});
  }};
  // بارگذاری تم ذخیره‌شده از سرور
  (function(){{
    var saved = document.documentElement.getAttribute('data-saved-dark');
    var savedC = document.documentElement.getAttribute('data-saved-classic');
    if(saved==='1'||saved==='0'||saved==='auto') localStorage.setItem('sl-dark', saved);
    if(savedC==='1'||savedC==='0') localStorage.setItem('sl-classic', savedC);
  }})();
  applyMode();

  var search=document.getElementById('globalSearch');
  function searchPanel(){{
    if(!search||!results)return; var q=search.value.trim().toLowerCase(); results.innerHTML='';
    if(!q){{results.classList.remove('open');return;}}
    Array.from(document.querySelectorAll('.sidebar-nav .nav-item')).filter(function(a){{return a.textContent.trim().toLowerCase().includes(q);}}).slice(0,7).forEach(function(a){{
      var item=document.createElement('a'); item.href=a.href; item.innerHTML='<span>'+a.textContent.trim()+'</span><small>'+new URL(a.href).pathname+'</small>'; results.appendChild(item);
    }});
    if(!results.children.length) results.innerHTML='<span class="empty-result">نتیجه‌ای پیدا نشد</span>';
    results.classList.add('open');
  }}
  search?.addEventListener('input',searchPanel);
  document.addEventListener('keydown',function(ev){{if((ev.metaKey||ev.ctrlKey)&&ev.key.toLowerCase()==='k'){{ev.preventDefault();search?.focus();}}if(ev.key==='Escape')results?.classList.remove('open');}});

  // Convert legacy decorative emoji to Lucide + Persian digits — ادغام‌شده در یک پیمایش DOM.
  // قبلاً این دو تبدیل (اینجا + بلوک بعدی) هرکدوم جدا کل document.body رو پیمایش می‌کردن؛
  // روی صفحات با جدول بزرگ (سفارش‌ها/محصولات/تیکت‌ها) یعنی دوبار پیمایش کامل قبل از
  // این‌که صفحه قابل‌استفاده بشه. الان یک پیمایش برای هر دو کار کافیه.
  var emojiIcons={{'✅':'circle-check','❌':'circle-x','⚠️':'triangle-alert','⚠':'triangle-alert','📊':'bar-chart-2','🛒':'shopping-bag','🛍':'shopping-bag','📦':'package','🗂':'folders','🗃':'archive','💼':'briefcase','🧾':'receipt-text','💰':'wallet','💳':'credit-card','👥':'users','🤝':'handshake','🎫':'ticket','📢':'megaphone','⚙️':'settings','⚙':'settings','👤':'user-round','💾':'hard-drive-download','📈':'trending-up','📱':'smartphone','🔑':'key-round','🔴':'circle-alert','🟡':'circle-dot','🟢':'circle-check','🔄':'refresh-cw','👨‍💻':'user-round','🧩':'blocks','←':'arrow-left','↩':'log-out','☰':'menu','✕':'x','▲':'trending-up','▼':'trending-down'}};
  var emojiRe=/(✅|❌|⚠️|⚠|📊|🛒|🛍|📦|🗂|🗃|💼|🧾|💰|💳|👥|🤝|🎫|📢|⚙️|⚙|👤|💾|📈|📱|🔑|🔴|🟡|🟢|🔄|👨‍💻|🧩|←|↩|☰|✕|▲|▼)/g;
  var FA_DIGITS=['۰','۱','۲','۳','۴','۵','۶','۷','۸','۹'];
  function faConv(s){{ return s.replace(/[0-9]/g,function(d){{return FA_DIGITS[+d];}}); }}
  var FA_SKIP_TAGS={{SCRIPT:1,STYLE:1,TEXTAREA:1,INPUT:1,SELECT:1,OPTION:1,CODE:1,PRE:1,KBD:1}};
  function faSkip(parent){{ return !parent||FA_SKIP_TAGS[parent.tagName]||(parent.closest&&parent.closest('code,pre,kbd,.no-fa')); }}
  var walker=document.createTreeWalker(document.body,NodeFilter.SHOW_TEXT); var nodes=[]; while(walker.nextNode()) nodes.push(walker.currentNode);
  nodes.forEach(function(node){{
    var parent=node.parentElement;
    var skipIcon=!parent||parent.closest('script,style,input,textarea,select,option,[data-no-icon]');
    var skipFa=faSkip(parent);
    var hasEmoji=!skipIcon&&emojiRe.test(node.nodeValue); emojiRe.lastIndex=0;
    if(!hasEmoji){{
      if(!skipFa&&/[0-9]/.test(node.nodeValue)) node.nodeValue=faConv(node.nodeValue);
      return;
    }}
    var frag=document.createDocumentFragment(),last=0,m;
    while((m=emojiRe.exec(node.nodeValue))){{
      var piece=node.nodeValue.slice(last,m.index);
      frag.append(document.createTextNode(skipFa?piece:faConv(piece)));
      var i=document.createElement('i');i.setAttribute('data-lucide',emojiIcons[m[0]]);i.className='legacy-lucide';frag.append(i);
      last=m.index+m[0].length;
    }}
    var tail=node.nodeValue.slice(last);
    frag.append(document.createTextNode(skipFa?tail:faConv(tail)));
    node.replaceWith(frag);
  }});

  {"" if not admin_info else """
  // Idle logout
  var IDLE = 300000;
  var timer;
  function reset(){ clearTimeout(timer); timer = setTimeout(function(){ location.href='/admin/login?flash='+encodeURIComponent('به دلیل عدم فعالیت خارج شدید'); }, IDLE); }
  ['mousemove','keydown','click','scroll','touchstart'].forEach(function(ev){ document.addEventListener(ev,reset,true); });
  reset();

  // Badge polling (every 12s)
  function updateBadge(id, count){
    var el = document.getElementById(id);
    if(!el) return;
    if(count>0){ el.textContent=count; el.classList.remove('hidden'); }
    else el.classList.add('hidden');
  }
  setInterval(function(){
    fetch('/admin/badges.json').then(function(r){return r.json();}).then(function(d){
      updateBadge('ticket-badge-top', d.tickets||0);
      updateBadge('partner-badge-top', d.partners||0);
      updateBadge('notes-badge-top', d.notes||0);
      updateBadge('receipts-badge-top', d.receipts||0);
      updateBadge('ticket-badge-nav', d.tickets||0);
      updateBadge('partner-badge-nav', d.partners||0);
      updateBadge('receipts-badge-nav', d.receipts||0);
    }).catch(function(){});
  }, 12000);
  """}
  renderIcons();
}})();
</script>
<script>
/* ─── اعداد فارسی سراسری پنل ─── */
(function(){{
  var FA=['۰','۱','۲','۳','۴','۵','۶','۷','۸','۹'];
  function conv(s){{ return s.replace(/[0-9]/g,function(d){{return FA[+d];}}); }}
  var SKIP={{SCRIPT:1,STYLE:1,TEXTAREA:1,INPUT:1,SELECT:1,OPTION:1,CODE:1,PRE:1,KBD:1}};
  function walk(node){{
    if(node.nodeType===3){{
      if(/[0-9]/.test(node.nodeValue)) node.nodeValue=conv(node.nodeValue);
      return;
    }}
    if(node.nodeType!==1) return;
    if(SKIP[node.tagName]) return;
    if(node.closest && node.closest('code,pre,kbd,.no-fa')) return;
    for(var i=0;i<node.childNodes.length;i++) walk(node.childNodes[i]);
  }}
  function run(){{ walk(document.body); }}
  // پیمایش اولیهٔ صفحه دیگه لازم نیست — بلوک بالا (تبدیل ایموجی) همین کار رو توی
  // همون یک پیمایش انجام می‌ده؛ این IIFE فقط برای محتوای تازه (MutationObserver) لازمه.
  var mo=new MutationObserver(function(muts){{
    mo.disconnect();
    muts.forEach(function(m){{
      m.addedNodes && m.addedNodes.forEach(function(n){{ walk(n); }});
      if(m.type==='characterData' && /[0-9]/.test(m.target.nodeValue||'')){{
        var p=m.target.parentElement;
        if(p && !SKIP[p.tagName] && !(p.closest&&p.closest('code,pre,kbd,.no-fa')))
          m.target.nodeValue=conv(m.target.nodeValue);
      }}
    }});
    mo.observe(document.body,{{childList:true,subtree:true,characterData:true}});
  }});
  mo.observe(document.body,{{childList:true,subtree:true,characterData:true}});
}})();
</script>
</body>
</html>""")

    # تجدید session cookie برای مدیر فعال (sliding window)
    if admin_info:
        try:
            _refresh_session(html_response, admin_info)
        except Exception:
            pass
    return html_response

def _card(title, value, sub="", color="indigo"):
    colors = {
        "indigo": "#4f46e5","green":"#16a34a","red":"#dc2626",
        "amber":"#d97706","blue":"#2563eb","slate":"#475569",
    }
    fg = colors.get(color, "#4f46e5")
    return (
        f'<div class="card p-5">'
        f'<div class="text-xs font-semibold mb-1" style="color:{fg}">{e(title)}</div>'
        f'<div class="text-2xl font-bold stat-card-value">{e(str(value))}</div>'
        f'{"<div class=\"stat-card-sub\">"+e(sub)+"</div>" if sub else ""}'
        f'</div>'
    )

def _btn(text, href="", color="indigo", small=False, danger=False):
    cls = "btn btn-sm " if small else "btn "
    if danger or color == "red":   cls += "btn-red"
    elif color == "green":         cls += "btn-green"
    elif color == "slate":         cls += "btn-slate"
    else:                          cls += "btn-indigo"
    if href:
        return f'<a href="{e(href)}" class="{cls}">{e(text)}</a>'
    return f'<button type="submit" class="{cls}">{e(text)}</button>'

def _input(name, placeholder="", value="", type_="text", required=False):
    req = "required" if required else ""
    return f'<input type="{type_}" name="{name}" value="{e(value)}" placeholder="{e(placeholder)}" {req}>'

def _textarea(name, placeholder="", value="", rows=4, ltr=False):
    dir_attr = ' dir="ltr" style="text-align:left"' if ltr else ''
    return f'<textarea name="{name}" rows="{rows}" placeholder="{e(placeholder)}"{dir_attr}>{e(value)}</textarea>'

# ─────────────────────────── Login / Logout ────────────────────────────────

@router.get("/login", response_class=HTMLResponse)
async def login_get(request: Request, err: str = "", flash: str = ""):
    adm = _get_admin(request)
    if adm:
        return _redir("/admin/")

    err_html = ""
    if err == "1":
        err_html = '<div class="alert-error">❌ نام کاربری یا رمز اشتباه است</div>'
    elif err == "rate":
        mins = request.query_params.get("mins", "15")
        err_html = f'<div class="alert-warning">🔒 دسترسی موقتاً مسدود شد. {mins} دقیقه دیگر تلاش کنید.</div>'
    if flash:
        err_html += f'<div class="alert-warning">⏱ {e(flash)}</div>'

    body = f"""
    <div class="login-wrap">
      <div class="login-card">
        <div class="login-header">
          <div class="login-brand">
            <span class="brand-word brand-word--stock">STOCK</span><span class="brand-word brand-word--land"> LAND</span>
          </div>
          <div class="login-subtitle">پنل مدیریت فروشگاه</div>
        </div>
        {err_html}
        <form method="post" action="/admin/login" class="login-form">
          <div>
            <label class="field-label">نام کاربری</label>
            {_input("username","نام کاربری",required=True)}
          </div>
          <div>
            <label class="field-label">رمز ورود</label>
            {_input("password","رمز ورود",type_="password",required=True)}
          </div>
          <button type="submit" class="login-submit">ورود به پنل ←</button>
        </form>
      </div>
    </div>"""
    return _layout("ورود", body)

@router.post("/login")
async def login_post(request: Request, username: str = Form(""), password: str = Form("")):
    ensure_admins_table()
    _ensure_theme_table()
    # پشت nginx (که خودش روی همون سرور به 127.0.0.1:8001 وصل می‌شه)، request.client.host
    # همیشه IP خودِ nginx است، نه IP واقعی کاربر — یعنی rate-limit روی این کلید عملاً
    # برای همهٔ کاربران اینترنت مشترکه و هرکسی می‌تونه با ۵ تلاش ناموفق، ورود پنل رو
    # برای مدیر واقعی هم قفل کنه. همون الگوی _log() که از قبل درست این کار رو می‌کنه.
    ip = request.headers.get("X-Forwarded-For","").split(",")[0].strip() or (request.client.host if request.client else "unknown")

    blocked, remaining = _is_rate_limited(ip)
    if blocked:
        return _redir(f"/admin/login?err=rate&mins={remaining // 60 + 1}")

    username = username.strip()
    password = password.strip()

    # ⚠️ رفع امنیتی (بخش ۱۳ آیتم ۳ سند): قبلاً "admin"/"super" همیشه به‌عنوان
    # نام کاربری معتبر پذیرفته می‌شدن، حتی اگه ادمین صریحاً یک ADMIN_WEB_USERNAME
    # سفارشی تنظیم کرده باشه — یعنی سفارشی‌کردن یوزرنیم برای سخت‌ترکردن ورود عملاً
    # بی‌اثر بود. حالا این دو نام مستعار فقط وقتی پذیرفته می‌شن که هیچ یوزرنیم
    # سفارشی‌ای در env تنظیم نشده باشه (رفتار پیش‌فرض قدیمی دست‌نخورده می‌مونه).
    _raw_super_un = os.environ.get("ADMIN_WEB_USERNAME", "").strip()
    super_un = _raw_super_un or "admin"
    allowed_super_usernames = {super_un.lower()} if _raw_super_un else {"admin", "super"}
    super_pw = _env("ADMIN_WEB_PASSWORD")
    if username.lower() in allowed_super_usernames and super_pw and _verify_super_pw(password, super_pw):
        _clear_attempts(ip)
        resp = _redir("/admin/")
        resp.set_cookie("adm", _make_session("super"), max_age=300, httponly=True, samesite="lax", secure=True)
        _log(request, "ورود", "احراز هویت", f"سوپرادمین از {ip}")
        return resp

    try:
        conn = _db()
        row = conn.execute(
            "SELECT id, web_password_hash, is_active FROM admins WHERE web_username=? LIMIT 1;",
            (username,),
        ).fetchone()
        if row and row["is_active"] and _verify_pw(password, row["web_password_hash"]):
            # ارتقای خودکار هش قدیمی (SHA256 بدون نمک) به فرمت جدید PBKDF2 نمکی —
            # فقط وقتی که پسورد درست تایپ شده باشه (یعنی همین لحظه در دسترس داریمش خام)
            if not row["web_password_hash"].startswith("pbkdf2$"):
                try:
                    conn.execute("UPDATE admins SET web_password_hash=? WHERE id=?;", (_hash_pw(password), row["id"]))
                    conn.commit()
                except Exception:
                    pass
            conn.close()
            _clear_attempts(ip)
            resp = _redir("/admin/")
            resp.set_cookie("adm", _make_session(str(row["id"])), max_age=300, httponly=True, samesite="lax", secure=True)
            _log(request, "ورود", "احراز هویت", f"ادمین #{row['id']}")
            return resp
        conn.close()
    except Exception:
        pass

    _record_fail(ip)
    _, remaining2 = _is_rate_limited(ip)
    _log(request, "ورود ناموفق", "احراز هویت", f"یوزرنیم: {username[:30]} از {ip}")
    return _redir("/admin/login?err=1")

@router.get("/logout")
async def logout(request: Request):
    adm = _get_admin(request)
    if adm:
        _log(request, "خروج", "احراز هویت", "", admin_info=adm)
    resp = _redir("/admin/login")
    resp.delete_cookie("adm")
    return resp

# ─────────────────────────── Dashboard ─────────────────────────────────────

def _dashboard_fetch() -> dict:
    """کوئری‌های داشبورد — روی ترد جدا (run_in_threadpool) اجرا می‌شود تا event loop مشترک بلاک نشود."""
    conn = _db()
    try:
        today = datetime.utcnow().date().isoformat()
        yesterday = (datetime.utcnow().date() - timedelta(days=1)).isoformat()

        today_o   = conn.execute("SELECT COUNT(*), COALESCE(SUM(price),0) FROM orders WHERE created_at LIKE ?;", (today+"%",)).fetchone()
        yest_o    = conn.execute("SELECT COUNT(*), COALESCE(SUM(price),0) FROM orders WHERE created_at LIKE ?;", (yesterday+"%",)).fetchone()
        total_o   = conn.execute("SELECT COUNT(*), COALESCE(SUM(price),0) FROM orders;").fetchone()
        feed_avail = conn.execute("SELECT COUNT(*) FROM product_feed WHERE delivered=0;").fetchone()[0]
        pending   = conn.execute("SELECT COUNT(*) FROM pending_deliveries WHERE status='pending';").fetchone()[0]
        partners_pend = conn.execute("SELECT COUNT(*) FROM partners WHERE status='pending';").fetchone()[0]
        open_tix  = _open_ticket_count()
        wallets   = conn.execute("SELECT COUNT(*), COALESCE(SUM(balance),0) FROM wallets;").fetchone()
        products_cnt = conn.execute("SELECT COUNT(*) FROM products WHERE is_active=1;").fetchone()[0]

        # نمودار ۳۰ روز اخیر
        chart_data = conn.execute("""
            SELECT substr(created_at,1,10) as day, COALESCE(SUM(price),0) as total
            FROM orders WHERE created_at >= date('now','-30 days')
            GROUP BY day ORDER BY day ASC;
        """).fetchall()

        # محصولات کم موجودی
        # ⚠️ Postgres اجازهٔ ارجاع به alias (`avail`) در HAVING رو نمی‌ده (برخلاف
        # SQLite) — باید عبارت کامل تکرار بشه. GROUP BY هم p.title اضافه شد تا
        # اگه products.id روی سرور PK اعلام‌نشده باشه، خطای grouping نده.
        low_stock = conn.execute("""
            SELECT p.id, p.title, COUNT(CASE WHEN pf.delivered=0 THEN 1 END) as avail
            FROM products p LEFT JOIN product_feed pf ON pf.product_id=p.id
            WHERE p.is_active=1 GROUP BY p.id, p.title
            HAVING COUNT(CASE WHEN pf.delivered=0 THEN 1 END)<=5 ORDER BY avail ASC LIMIT 5;
        """).fetchall()

        # سفارش‌های اخیر
        recent = conn.execute("""
            SELECT id, user_id, title, price, created_at, COALESCE(status,'active') as status
            FROM orders ORDER BY id DESC LIMIT 8;
        """).fetchall()

        # آخرین بکاپ خودکار
        import glob as _g, os as _o
        auto_backups = sorted(_g.glob("/tmp/stockland_backups/auto_*.sqlite"), reverse=True)
        last_backup = _o.path.basename(auto_backups[0]).replace("auto_","").replace(".sqlite","") if auto_backups else None

        return dict(today=today, today_o=today_o, yest_o=yest_o, total_o=total_o, feed_avail=feed_avail,
                    pending=pending, partners_pend=partners_pend, open_tix=open_tix, wallets=wallets,
                    products_cnt=products_cnt, chart_data=chart_data, low_stock=low_stock,
                    recent=recent, last_backup=last_backup)
    finally:
        conn.close()


@router.get("/", response_class=HTMLResponse)
async def dashboard(request: Request, err: str = ""):
    adm = _get_admin(request)
    if not adm:
        return _redir("/admin/login")

    flash = "دسترسی کافی ندارید." if err == "noperm" else ""

    d = await run_in_threadpool(_dashboard_fetch)
    today = d["today"]
    today_o, yest_o, total_o = d["today_o"], d["yest_o"], d["total_o"]
    feed_avail, pending, partners_pend = d["feed_avail"], d["pending"], d["partners_pend"]
    open_tix, wallets, products_cnt = d["open_tix"], d["wallets"], d["products_cnt"]
    chart_data, low_stock, recent, last_backup = d["chart_data"], d["low_stock"], d["recent"], d["last_backup"]

    # محاسبه درصد تغییر نسبت به دیروز (فقط برای نمایش KPI)
    rev_change = 0
    if yest_o[1] > 0:
        rev_change = round(((today_o[1] - yest_o[1]) / yest_o[1]) * 100, 1)
    cnt_change = today_o[0] - yest_o[0]

    def pct_badge(value, suffix="%"):
        if value > 0:
            return f'<span class="trend trend-up"><i data-lucide="trending-up"></i>{value}{suffix}</span>'
        if value < 0:
            return f'<span class="trend trend-down"><i data-lucide="trending-down"></i>{abs(value)}{suffix}</span>'
        return '<span class="trend trend-flat"><i data-lucide="minus"></i>بدون تغییر</span>'

    def status_badge(status):
        styles = {
            "active":   ("success", "ارسال شد"),
            "returned": ("danger", "برگشتی"),
            "pending":  ("warning", "در انتظار"),
        }
        tone, label = styles.get(status, ("neutral", status))
        return f'<span class="status-badge status-{tone}"><span></span>{e(label)}</span>'

    chart_labels = json.dumps([r["day"][5:] for r in chart_data], ensure_ascii=False)
    chart_values = json.dumps([int(r["total"]) for r in chart_data])

    low_rows = "".join(f"""
    <a href="/admin/feed/{r['id']}" class="compact-row">
      <span class="row-icon {'danger' if r['avail'] == 0 else 'warning'}"><i data-lucide="package-minus"></i></span>
      <span class="row-copy"><strong>{e(r['title'])}</strong><small>نیازمند تأمین موجودی</small></span>
      <span class="stock-count {'out' if r['avail'] == 0 else ''}">{r['avail']} عدد</span>
    </a>""" for r in low_stock)

    recent_rows = "".join(f"""
    <tr>
      <td><a class="order-id" href="/admin/orders/{o['id']}">#{o['id']}</a></td>
      <td><div class="table-product"><span><i data-lucide="package"></i></span><strong>{e(o['title'][:34])}</strong></div></td>
      <td><code class="no-fa text-xs">{o['user_id']}</code></td>
      <td><strong class="money-cell">{int(o['price']):,}</strong><small class="currency">تومان</small></td>
      <td>{status_badge(o['status'] or 'active')}</td>
      <td><a class="table-action" href="/admin/orders/{o['id']}" aria-label="مشاهده سفارش"><i data-lucide="arrow-up-left"></i></a></td>
    </tr>""" for o in recent)

    # مشتق‌شده از همان سفارش‌های اخیر؛ بدون query یا تغییر منطق داده.
    product_summary = {}
    for order in recent:
        product_title = str(order["title"] or "بدون عنوان")
        if product_title not in product_summary:
            product_summary[product_title] = {"count": 0, "revenue": 0}
        product_summary[product_title]["count"] += 1
        product_summary[product_title]["revenue"] += int(order["price"] or 0)

    top_product_rows = "".join(f"""
      <div class="rank-row"><span class="rank-number">{index}</span><span class="rank-copy"><strong>{e(name[:30])}</strong><small>{data['count']} سفارش اخیر</small></span><strong class="rank-value">{data['revenue']:,}</strong></div>
    """ for index, (name, data) in enumerate(sorted(product_summary.items(), key=lambda item: (item[1]["count"], item[1]["revenue"]), reverse=True)[:5], 1))

    activity_rows = "".join(f"""
      <div class="activity-row"><span class="activity-icon"><i data-lucide="shopping-bag"></i></span><span><strong>سفارش #{o['id']} ثبت شد</strong><small>{e(o['title'][:28])} · {fa_date(o['created_at'], with_time=True)}</small></span></div>
    """ for o in recent[:5])

    def command_item(icon, label, value, meta, tone="cyan", href="#"):
        return f"""
        <a href="{href}" class="command-item command-{tone}">
          <span class="command-icon"><i data-lucide="{icon}"></i></span>
          <span class="command-copy"><small>{label}</small><strong>{value}</strong><em>{meta}</em></span>
          <i data-lucide="arrow-up-left" class="command-arrow"></i>
        </a>"""

    command_items = "".join([
        command_item("clock-3", "سفارش‌های معلق", pending, "نیازمند پیگیری", "warning" if pending else "success", "/admin/orders"),
        command_item("package-search", "کمبود موجودی", len(low_stock), "محصول در آستانه اتمام", "danger" if low_stock else "success", "/admin/feed"),
        command_item("message-square", "تیکت‌های باز", open_tix, "در انتظار پاسخ مدیر", "danger" if open_tix else "success", "/admin/tickets"),
        command_item("handshake", "همکاران معلق", partners_pend, "درخواست بررسی‌نشده", "warning" if partners_pend else "success", "/admin/partners"),
        command_item("database-backup", "آخرین بکاپ", "ثبت شده" if last_backup else "ناموجود", last_backup or "هنوز بکاپی ثبت نشده", "success" if last_backup else "warning", "/admin/database"),
    ])

    tasks = "".join([
        f'<a href="/admin/feed" class="task-row"><span class="task-status danger"><i data-lucide="package-minus"></i></span><span><strong>تأمین موجودی محصولات</strong><small>{len(low_stock)} محصول نیازمند بررسی</small></span><i data-lucide="chevron-left"></i></a>',
        f'<a href="/admin/tickets" class="task-row"><span class="task-status warning"><i data-lucide="message-circle"></i></span><span><strong>پاسخ به تیکت‌ها</strong><small>{open_tix} تیکت منتظر پاسخ</small></span><i data-lucide="chevron-left"></i></a>',
        f'<a href="/admin/orders" class="task-row"><span class="task-status cyan"><i data-lucide="truck"></i></span><span><strong>تحویل‌های معلق</strong><small>{pending} سفارش در صف ارسال</small></span><i data-lucide="chevron-left"></i></a>',
        '<a href="/admin/database" class="task-row"><span class="task-status success"><i data-lucide="server"></i></span><span><strong>وضعیت سرور</strong><small>سرویس در دسترس است</small></span><i data-lucide="chevron-left"></i></a>',
        ('<a href="/admin/database" class="task-row">'
         f'<span class="task-status {"success" if last_backup else "warning"}">'
         '<i data-lucide="hard-drive-download"></i></span>'
         f'<span><strong>نسخه پشتیبان</strong><small>{e(last_backup or "هنوز ثبت نشده")}</small></span>'
         '<i data-lucide="chevron-left"></i></a>'),
        '<a href="/admin/broadcast" class="task-row"><span class="task-status success"><i data-lucide="bot"></i></span><span><strong>ربات استوک‌لند</strong><small>فعال و آماده پاسخ‌گویی</small></span><i data-lucide="chevron-left"></i></a>',
    ])

    body = f"""
    <style>
      .dashboard {{ display:flex; flex-direction:column; gap:28px; }}
      .dashboard-head {{ display:flex; align-items:flex-end; justify-content:space-between; gap:20px; }}
      .dashboard-head h2 {{ margin:0; color:var(--text-main); font-size:25px; font-weight:800; letter-spacing:-.02em; }}
      .dashboard-head p {{ margin:6px 0 0; color:var(--text-muted); font-size:12px; }}
      .date-chip {{ display:flex; align-items:center; gap:8px; padding:9px 13px; border:1px solid var(--border); border-radius:12px; background:var(--card-bg); color:var(--text-muted); font-size:11px; }}
      .date-chip svg {{ width:16px; color:#0891b2; }}
      .section-heading {{ display:flex; align-items:flex-end; justify-content:space-between; gap:16px; margin-bottom:14px; }}
      .section-heading h3 {{ margin:0; color:var(--text-main); font-size:16px; font-weight:750; }}
      .section-heading p {{ margin:4px 0 0; color:var(--text-muted); font-size:10.5px; }}
      .section-link {{ display:inline-flex; align-items:center; gap:5px; color:#0891b2; font-size:11px; font-weight:650; text-decoration:none; }} .section-link svg {{ width:14px; }}

      .command-center {{ position:relative; overflow:hidden; padding:22px; border-radius:26px; background:linear-gradient(130deg,#07111b 0%,#091827 58%,#0a2631 100%); box-shadow:0 24px 65px rgba(4,12,23,.18); }}
      .command-center::before {{ content:""; position:absolute; top:-130px; left:15%; width:380px; height:280px; border-radius:50%; background:rgba(0,215,255,.10); filter:blur(70px); }}
      .command-head {{ position:relative; display:flex; justify-content:space-between; align-items:center; margin-bottom:16px; color:#fff; }}
      .command-title {{ display:flex; align-items:center; gap:11px; }} .command-title>span {{ width:36px; height:36px; display:grid; place-items:center; border-radius:12px; background:rgba(0,215,255,.1); color:#42e2ff; border:1px solid rgba(0,215,255,.12); }}
      .command-title svg {{ width:18px; }} .command-title h3 {{ margin:0; font-size:15px; }} .command-title p {{ margin:3px 0 0; color:#74849a; font-size:9.5px; }}
      .live-pill {{ display:flex; align-items:center; gap:7px; padding:7px 10px; border-radius:999px; color:#9de9bd; background:rgba(34,197,94,.08); border:1px solid rgba(34,197,94,.12); font-size:9px; }}
      .live-pill span {{ width:6px; height:6px; border-radius:50%; background:#22c55e; box-shadow:0 0 10px #22c55e; }}
      .command-grid {{ position:relative; display:grid; grid-template-columns:repeat(auto-fit,minmax(145px,1fr)); gap:9px; overflow-x:auto; padding-bottom:2px; scrollbar-width:thin; }}
      .command-item {{ min-width:145px; min-height:142px; display:flex; flex-direction:column; position:relative; padding:15px; border:1px solid rgba(255,255,255,.07); border-radius:18px; background:rgba(255,255,255,.035); text-decoration:none; transition:200ms; }}
      .command-item:hover {{ background:rgba(255,255,255,.065); border-color:rgba(0,215,255,.18); transform:translateY(-2px); }}
      .command-icon {{ width:33px; height:33px; display:grid; place-items:center; border-radius:11px; background:rgba(0,215,255,.08); color:#42e2ff; }} .command-icon svg {{ width:17px; }}
      .command-warning .command-icon {{ background:rgba(245,158,11,.09); color:#fbbf24; }} .command-danger .command-icon {{ background:rgba(239,68,68,.09); color:#fb7185; }} .command-success .command-icon {{ background:rgba(34,197,94,.09); color:#4ade80; }}
      .command-copy {{ margin-top:auto; }} .command-copy small,.command-copy strong,.command-copy em {{ display:block; font-style:normal; }} .command-copy small {{ color:#8290a3; font-size:9.5px; }} .command-copy strong {{ color:#f8fafc; font-size:17px; margin-top:3px; }} .command-copy em {{ color:#65758a; font-size:8.5px; margin-top:3px; white-space:nowrap; overflow:hidden; text-overflow:ellipsis; }}
      .command-arrow {{ position:absolute; top:16px; left:14px; width:13px; color:#465568; }}

      .kpi-grid {{ display:grid; grid-template-columns:repeat(4,minmax(0,1fr)); gap:16px; }}
      .kpi-card {{ min-height:188px; padding:20px; overflow:hidden; position:relative; }} .kpi-card::after {{ content:""; position:absolute; top:-55px; left:-45px; width:130px; height:130px; border-radius:50%; background:rgba(0,215,255,.035); }}
      .kpi-top {{ display:flex; align-items:center; justify-content:space-between; }} .kpi-icon {{ width:40px; height:40px; display:grid; place-items:center; border-radius:13px; color:#0891b2; background:#ecfeff; }} .kpi-icon svg {{ width:19px; }}
      .kpi-label {{ color:var(--text-muted); font-size:10.5px; font-weight:600; }} .kpi-value {{ margin:15px 0 3px; color:var(--text-main); font-size:25px; font-weight:800; letter-spacing:-.025em; direction:ltr; text-align:left; }} .kpi-unit {{ color:var(--text-muted); font-size:9px; font-weight:500; margin-right:3px; }}
      .trend {{ display:inline-flex; align-items:center; gap:3px; font-size:9px; font-weight:700; direction:ltr; }} .trend svg {{ width:12px; }} .trend-up {{ color:#16a34a; }} .trend-down {{ color:#e11d48; }} .trend-flat {{ color:#94a3b8; }}
      .sparkline {{ position:absolute; right:15px; left:15px; bottom:11px; height:43px; opacity:.9; }} .sparkline path.line {{ fill:none; stroke:#06b6d4; stroke-width:2.2; stroke-linecap:round; stroke-linejoin:round; }} .sparkline path.area {{ fill:url(#sparkFill); opacity:.55; }}

      .chart-card {{ padding:22px 24px 18px; }} .chart-toolbar {{ display:flex; align-items:center; justify-content:space-between; margin-bottom:10px; }} .chart-legend {{ display:flex; align-items:center; gap:7px; color:var(--text-muted); font-size:10px; }} .chart-legend span {{ width:7px; height:7px; border-radius:50%; background:#00d7ff; box-shadow:0 0 0 4px rgba(0,215,255,.08); }} .chart-period {{ padding:7px 10px; border:1px solid var(--border); border-radius:10px; color:var(--text-muted); font-size:9px; background:var(--card-bg); }}
      .chart-shell {{ position:relative; height:400px; direction:ltr; }}
      .two-column {{ display:grid; grid-template-columns:minmax(0,3fr) minmax(320px,2fr); gap:18px; }} .panel-card {{ overflow:hidden; }} .panel-header {{ min-height:68px; display:flex; align-items:center; justify-content:space-between; padding:0 20px; border-bottom:1px solid var(--border); }} .panel-header h3 {{ margin:0; font-size:14px; }} .panel-header p {{ margin:3px 0 0; color:var(--text-muted); font-size:9.5px; }} .table-wrap {{ overflow:auto; max-height:500px; }}
      .order-id {{ color:#0891b2; font:700 11px/1 Inter,sans-serif !important; text-decoration:none; }} .table-product {{ display:flex; align-items:center; gap:9px; min-width:180px; }} .table-product>span {{ width:31px; height:31px; display:grid; place-items:center; border-radius:10px; background:#f1f5f9; color:#64748b; }} .table-product svg {{ width:15px; }} .table-product strong {{ font-size:11px; font-weight:600; }} .muted-cell {{ color:var(--text-muted); font-size:10px; }} .money-cell {{ display:block; font-size:11px; direction:ltr; text-align:left; }} .currency {{ color:var(--text-muted); font-size:8px; }}
      .status-badge {{ display:inline-flex; align-items:center; gap:5px; padding:5px 8px; border-radius:999px; font-size:9px; font-weight:650; white-space:nowrap; }} .status-badge>span {{ width:5px; height:5px; border-radius:50%; }} .status-success {{ color:#15803d; background:#ecfdf3; }} .status-success>span {{ background:#22c55e; }} .status-warning {{ color:#a16207; background:#fffbeb; }} .status-warning>span {{ background:#f59e0b; }} .status-danger {{ color:#be123c; background:#fff1f2; }} .status-danger>span {{ background:#ef4444; }} .status-neutral {{ color:#475569; background:#f1f5f9; }} .status-neutral>span {{ background:#94a3b8; }}
      .table-action {{ width:30px; height:30px; display:grid; place-items:center; border:1px solid var(--border); border-radius:9px; color:#64748b; }} .table-action svg {{ width:14px; }}
      .tasks-list {{ padding:8px 12px 12px; }} .task-row {{ min-height:67px; display:grid; grid-template-columns:38px 1fr 16px; align-items:center; gap:10px; padding:7px 8px; border-bottom:1px solid #eef0f3; text-decoration:none; transition:160ms; }} .task-row:last-child {{ border:0; }} .task-row:hover {{ background:var(--page-bg); border-radius:12px; }} .task-row>svg {{ width:14px; color:#b1b7c2; }} .task-row strong,.task-row small {{ display:block; }} .task-row strong {{ color:var(--text-main); font-size:10.5px; }} .task-row small {{ color:var(--text-muted); font-size:8.5px; margin-top:3px; }}
      .task-status {{ width:36px; height:36px; display:grid; place-items:center; border-radius:11px; }} .task-status svg {{ width:16px; }} .task-status.danger {{ color:#e11d48; background:#fff1f2; }} .task-status.warning {{ color:#d97706; background:#fffbeb; }} .task-status.cyan {{ color:#0891b2; background:#ecfeff; }} .task-status.success {{ color:#16a34a; background:#f0fdf4; }}

      .three-column {{ display:grid; grid-template-columns:repeat(3,minmax(0,1fr)); gap:18px; }} .mini-panel {{ padding-bottom:12px; overflow:hidden; }} .mini-panel .panel-header {{ min-height:64px; }} .rank-row {{ min-height:56px; display:grid; grid-template-columns:28px 1fr auto; align-items:center; gap:9px; padding:7px 18px; }} .rank-number {{ width:25px; height:25px; display:grid; place-items:center; border-radius:8px; background:#f1f5f9; color:#64748b; font:700 9px/1 Inter,sans-serif !important; }} .rank-copy strong,.rank-copy small {{ display:block; }} .rank-copy strong {{ font-size:10px; }} .rank-copy small {{ color:var(--text-muted); font-size:8px; margin-top:3px; }} .rank-value {{ color:#0891b2; font-size:9px; direction:ltr; }}
      .compact-row {{ min-height:58px; display:grid; grid-template-columns:36px 1fr auto; align-items:center; gap:9px; padding:7px 16px; text-decoration:none; }} .row-icon {{ width:34px; height:34px; display:grid; place-items:center; border-radius:10px; }} .row-icon svg {{ width:15px; }} .row-icon.warning {{ color:#d97706; background:#fffbeb; }} .row-icon.danger {{ color:#e11d48; background:#fff1f2; }} .row-copy strong,.row-copy small {{ display:block; }} .row-copy strong {{ font-size:9.5px; }} .row-copy small {{ color:var(--text-muted); font-size:8px; margin-top:3px; }} .stock-count {{ padding:4px 7px; border-radius:999px; color:#a16207; background:#fffbeb; font-size:8px; font-weight:700; }} .stock-count.out {{ color:#be123c; background:#fff1f2; }}
      .activity-row {{ min-height:58px; display:grid; grid-template-columns:35px 1fr; align-items:center; gap:9px; padding:7px 16px; }} .activity-icon {{ width:34px; height:34px; display:grid; place-items:center; border-radius:50%; color:#0891b2; background:#ecfeff; position:relative; }} .activity-icon svg {{ width:15px; }} .activity-row strong,.activity-row small {{ display:block; }} .activity-row strong {{ font-size:9.5px; }} .activity-row small {{ color:var(--text-muted); font-size:8px; margin-top:3px; }} .empty-state {{ padding:38px 18px; color:var(--text-muted); text-align:center; font-size:10px; }}
      body.dark-mode .kpi-icon,body.dark-mode .table-product>span,body.dark-mode .rank-number {{ background:#172330; }} body.dark-mode .task-row,body.dark-mode .panel-header {{ border-color:#273244; }} body.dark-mode .status-success {{ background:rgba(34,197,94,.1); }} body.dark-mode .status-warning {{ background:rgba(245,158,11,.1); }} body.dark-mode .status-danger {{ background:rgba(239,68,68,.1); }}
      @media(max-width:1200px) {{ .kpi-grid {{ grid-template-columns:repeat(2,1fr); }} .three-column {{ grid-template-columns:1fr 1fr; }} .three-column>*:last-child {{ grid-column:1/-1; }} }}
      @media(max-width:940px) {{ .two-column {{ grid-template-columns:1fr; }} .command-grid {{ grid-template-columns:none; grid-auto-flow:column; grid-auto-columns:155px; }} }}
      /* فاز ۵: حالت Classic — مرکز فرمان روشن و ساده */
      body.sl-classic .command-center {{ background:#FFFFFF !important; box-shadow:0 1px 4px rgba(0,0,0,.06) !important; border:1px solid #E5E7EB !important; }}
      body.sl-classic .command-center::before {{ display:none !important; }}
      body.sl-classic .command-title h3 {{ color:#111827 !important; }}
      body.sl-classic .command-title p {{ color:#6B7280 !important; }}
      body.sl-classic .command-title>span {{ background:#EFF6FF !important; color:#2563EB !important; border-color:#DBEAFE !important; }}
      body.sl-classic .command-item {{ background:#F9FAFB !important; border:1px solid #E5E7EB !important; }}
      body.sl-classic .command-item strong {{ color:#111827 !important; }}
      body.sl-classic .command-item small {{ color:#6B7280 !important; }}
      body.sl-classic .command-icon {{ background:#EFF6FF !important; color:#2563EB !important; }}
      body.sl-classic .live-pill {{ background:#ECFDF5 !important; color:#059669 !important; border-color:#D1FAE5 !important; }}
      body.sl-classic .kpi-card {{ background:#FFFFFF !important; border:1px solid #E5E7EB !important; box-shadow:0 1px 3px rgba(0,0,0,.05) !important; }}
      body.sl-classic .kpi-label, body.sl-classic .kpi-unit {{ color:#6B7280 !important; }}
      body.sl-classic .chart-card {{ background:#FFFFFF !important; border:1px solid #E5E7EB !important; }}
      /* Dark mode هم برای مرکز فرمان */
      body.sl-dark .command-item, body.dark-mode .command-item {{ background:#1B2530 !important; border-color:#2B3A4C !important; }}
      body.sl-dark .kpi-card, body.dark-mode .kpi-card {{ background:#17212B !important; border-color:#2B3A4C !important; }}
      body.sl-dark .chart-card, body.dark-mode .chart-card {{ background:#17212B !important; border-color:#2B3A4C !important; }}
      body.sl-dark .activity-icon, body.dark-mode .activity-icon {{ background:#1B2530 !important; }}
            /* موبایل: دکمه collapse sidebar مخفی + body fixed وقتی sidebar بازه */
    @media(max-width:768px) {{
      .sidebar-collapse-btn {{ display:none !important; }}
      body.sidebar-open {{ overflow:hidden !important; position:fixed !important; width:100% !important; }}
      .sidebar {{ overflow-y:auto !important; -webkit-overflow-scrolling:touch; }}
    }}
    @media(max-width:640px) {{ .dashboard {{ gap:22px; }} .dashboard-head {{ align-items:flex-start; }} .dashboard-head h2 {{ font-size:21px; }} .date-chip {{ display:none; }} .command-center {{ padding:18px 14px; border-radius:22px; }} .command-grid {{ margin-left:-14px; padding-left:14px; }} .kpi-grid,.three-column {{ grid-template-columns:1fr; }} .three-column>*:last-child {{ grid-column:auto; }} .kpi-card {{ min-height:174px; }} .chart-card {{ padding:18px 12px 12px; }} .chart-shell {{ height:320px; }} .panel-header {{ padding:0 14px; }} }}
    </style>

    <main class="dashboard">
      <div class="dashboard-head">
        <div><h2>مرکز مدیریت استوک‌لند</h2><p>نمای یکپارچه فروش، عملیات و سلامت سرویس‌های فروشگاه</p></div>
        <div class="date-chip"><i data-lucide="calendar-days"></i><span>{today}</span></div>
      </div>

      <section class="command-center" aria-labelledby="command-title">
        <div class="command-head"><div class="command-title"><span><i data-lucide="command"></i></span><div><h3 id="command-title">مرکز فرمان</h3><p>مواردی که همین حالا به توجه شما نیاز دارند</p></div></div><div class="live-pill"><span></span>به‌روزرسانی زنده</div></div>
        <div class="command-grid">{command_items}</div>
      </section>

      <section aria-labelledby="kpi-title">
        <div class="section-heading"><div><h3 id="kpi-title">شاخص‌های کلیدی</h3><p>خلاصه عملکرد امروز و وضعیت فروشگاه</p></div></div>
        <div class="kpi-grid">
          <article class="card kpi-card"><div class="kpi-top"><span class="kpi-icon"><i data-lucide="banknote"></i></span>{pct_badge(rev_change)}</div><div class="kpi-value">{int(today_o[1]):,}<span class="kpi-unit">تومان</span></div><div class="kpi-label">درآمد امروز</div><svg class="sparkline" viewBox="0 0 240 45" preserveAspectRatio="none"><defs><linearGradient id="sparkFill" x1="0" y1="0" x2="0" y2="1"><stop offset="0" stop-color="#00d7ff" stop-opacity=".24"/><stop offset="1" stop-color="#00d7ff" stop-opacity="0"/></linearGradient></defs><path class="area" d="M0,37 C24,34 35,18 58,24 S93,39 118,23 S154,11 178,18 S210,7 240,10 L240,45 L0,45Z"/><path class="line" d="M0,37 C24,34 35,18 58,24 S93,39 118,23 S154,11 178,18 S210,7 240,10"/></svg></article>
          <article class="card kpi-card"><div class="kpi-top"><span class="kpi-icon"><i data-lucide="shopping-cart"></i></span>{pct_badge(cnt_change, "")}</div><div class="kpi-value">{today_o[0]:,}<span class="kpi-unit">سفارش</span></div><div class="kpi-label">سفارش‌های امروز</div><svg class="sparkline" viewBox="0 0 240 45" preserveAspectRatio="none"><path class="area" d="M0,31 C25,14 48,38 72,26 S112,8 137,20 S175,32 199,17 S223,13 240,7 L240,45 L0,45Z"/><path class="line" d="M0,31 C25,14 48,38 72,26 S112,8 137,20 S175,32 199,17 S223,13 240,7"/></svg></article>
          <article class="card kpi-card"><div class="kpi-top"><span class="kpi-icon"><i data-lucide="boxes"></i></span><span class="trend trend-flat"><i data-lucide="package-check"></i>{products_cnt} محصول فعال</span></div><div class="kpi-value">{feed_avail:,}<span class="kpi-unit">آیتم</span></div><div class="kpi-label">موجودی قابل تحویل</div><svg class="sparkline" viewBox="0 0 240 45" preserveAspectRatio="none"><path class="area" d="M0,28 C22,24 42,27 65,19 S105,30 129,21 S162,12 187,17 S218,10 240,14 L240,45 L0,45Z"/><path class="line" d="M0,28 C22,24 42,27 65,19 S105,30 129,21 S162,12 187,17 S218,10 240,14"/></svg></article>
          <article class="card kpi-card"><div class="kpi-top"><span class="kpi-icon"><i data-lucide="trending-up"></i></span><span class="trend trend-flat"><i data-lucide="receipt-text"></i>{total_o[0]:,} سفارش</span></div><div class="kpi-value">{int(total_o[1]):,}<span class="kpi-unit">تومان</span></div><div class="kpi-label">کل درآمد</div><svg class="sparkline" viewBox="0 0 240 45" preserveAspectRatio="none"><path class="area" d="M0,39 C26,34 38,30 62,31 S101,20 124,23 S160,15 181,16 S216,5 240,9 L240,45 L0,45Z"/><path class="line" d="M0,39 C26,34 38,30 62,31 S101,20 124,23 S160,15 181,16 S216,5 240,9"/></svg></article>
        </div>
      </section>

      <section class="card chart-card" aria-labelledby="sales-title">
        <div class="chart-toolbar"><div class="section-heading m-0"><div><h3 id="sales-title">تحلیل فروش</h3><p>روند درآمد در ۳۰ روز اخیر</p></div></div><div class="flex-gap14"><span class="chart-legend"><span></span>فروش روزانه</span><span class="chart-period">۳۰ روز اخیر</span></div></div>
        <div class="chart-shell"><canvas id="salesChart"></canvas></div>
      </section>

      <section class="two-column">
        <article class="card panel-card"><div class="panel-header"><div><h3>سفارش‌های اخیر</h3><p>آخرین تراکنش‌های ثبت‌شده در فروشگاه</p></div><a href="/admin/orders" class="section-link">مشاهده همه<i data-lucide="arrow-left"></i></a></div><div class="table-wrap"><table><thead><tr><th>شناسه</th><th>محصول</th><th>کاربر</th><th>مبلغ</th><th>وضعیت</th><th></th></tr></thead><tbody>{recent_rows or '<tr><td colspan="6" class="empty-state">هنوز سفارشی ثبت نشده است</td></tr>'}</tbody></table></div></article>
        <article class="card panel-card"><div class="panel-header"><div><h3>وظایف مرکز فرمان</h3><p>اولویت‌های عملیاتی امروز</p></div><span class="status-badge status-warning"><span></span>نیازمند توجه</span></div><div class="tasks-list">{tasks}</div></article>
      </section>

      <section class="three-column">
        <article class="card mini-panel"><div class="panel-header"><div><h3>محصولات برتر</h3><p>بر اساس سفارش‌های اخیر</p></div><i data-lucide="trophy" class="icon-18 icon-amber"></i></div>{top_product_rows or '<div class="empty-state">داده‌ای برای نمایش وجود ندارد</div>'}</article>
        <article class="card mini-panel"><div class="panel-header"><div><h3>موجودی رو به اتمام</h3><p>محصولات نیازمند تأمین</p></div><a href="/admin/feed" class="section-link">مدیریت<i data-lucide="arrow-left"></i></a></div>{low_rows or '<div class="empty-state">موجودی همه محصولات کافی است</div>'}</article>
        <article class="card mini-panel"><div class="panel-header"><div><h3>فعالیت‌های اخیر</h3><p>رویدادهای تازه فروشگاه</p></div><i data-lucide="history" class="icon-18 icon-cyan"></i></div>{activity_rows or '<div class="empty-state">فعالیت تازه‌ای ثبت نشده است</div>'}</article>
      </section>
    </main>

    <script src="https://cdn.jsdelivr.net/npm/chart.js@4/dist/chart.umd.min.js"></script>
    <script>
    (function(){{
      var canvas=document.getElementById('salesChart'); if(!canvas||!window.Chart)return;
      var context=canvas.getContext('2d'); var gradient=context.createLinearGradient(0,0,0,400); gradient.addColorStop(0,'rgba(0,215,255,.22)'); gradient.addColorStop(1,'rgba(0,215,255,0)');
      new Chart(context, {{
        type: 'line',
        data: {{
          labels: {chart_labels},
          datasets: [{{
            label: 'فروش روزانه', data: {chart_values}, borderColor: '#2EC4B6',
            backgroundColor: gradient, borderWidth: 2.5, fill: true, tension: .42,
            pointRadius: 0, pointHoverRadius: 5, pointHoverBackgroundColor: '#2EC4B6',
            pointHoverBorderColor: '#fff', pointHoverBorderWidth: 3
          }}]
        }},
        options: {{
          responsive: true, maintainAspectRatio: false,
          interaction: {{ intersect: false, mode: 'index' }},
          plugins: {{
            legend: {{ display: false }},
            tooltip: {{
              rtl: true, titleFont: {{ family: 'Vazirmatn', size: 11 }},
              bodyFont: {{ family: 'Vazirmatn', size: 11 }}, backgroundColor: 'rgba(5,7,10,.94)',
              padding: 12, cornerRadius: 12, displayColors: false,
              callbacks: {{ label: function(ctx) {{ return Number(ctx.raw || 0).toLocaleString('fa-IR') + ' تومان'; }} }}
            }}
          }},
          scales: {{
            y: {{
              beginAtZero: true, border: {{ display: false }},
              grid: {{ color: 'rgba(148,163,184,.12)', drawTicks: false }},
              ticks: {{
                padding: 12, color: '#94A3B8', font: {{ family: 'Vazirmatn', size: 9 }},
                callback: function(v) {{ return v >= 1000000 ? (v / 1000000).toLocaleString('fa-IR') + ' م' : v.toLocaleString('fa-IR'); }}
              }}
            }},
            x: {{
              border: {{ display: false }}, grid: {{ display: false }},
              ticks: {{ color: '#94A3B8', font: {{ family: 'Vazirmatn', size: 9 }}, maxRotation: 0, autoSkip: true, maxTicksLimit: 10 }}
            }}
          }}
        }}
      }});
    }})();
    </script>"""

    return _layout("داشبورد", body, adm, flash=flash, flash_ok=False)

# ─────────────────────────── Settings ──────────────────────────────────────

DEFAULT_UI_TEXTS = {
    "MAIN_BTN_OTHER_PRODUCTS": "🛍 سایر محصولات فروشگاه",
    "MAIN_BTN_BUY_APPLE_ID":  "📱 سرویس اپل آیدی",
    "MAIN_BTN_MY_ORDERS":     "🧾 خریدهای من",
    "MAIN_BTN_WALLET":        "💰 کیف پول",
    "MAIN_BTN_PARTNER_REQUEST":"📝 درخواست نمایندگی",
    "MAIN_BTN_PARTNER_PANEL": "🤝 پنل همکار",
    "MAIN_BTN_GUIDE":         "🔑 راهنما",
    "MAIN_BTN_SUPPORT":       "👨‍💻 پشتیبانی",
    "SUPPORT_TEXT":           "متن پشتیبانی...",
    "HELP_TEXT":              "متن راهنما...",
    "TXT_MAIN_MENU_TITLE":    "منوی اصلی",
}

MAIN_BUTTONS = [
    "MAIN_BTN_OTHER_PRODUCTS",
    "MAIN_BTN_BUY_APPLE_ID",
    "MAIN_BTN_MY_ORDERS",
    "MAIN_BTN_WALLET",
    "MAIN_BTN_PARTNER_REQUEST",
    "MAIN_BTN_PARTNER_PANEL",
    "MAIN_BTN_GUIDE",
    "MAIN_BTN_SUPPORT",
]

def _get_ui(conn, key: str) -> str:
    try:
        row = conn.execute("SELECT value FROM ui_texts WHERE key=? LIMIT 1;", (key,)).fetchone()
        return row["value"] if row else DEFAULT_UI_TEXTS.get(key, "")
    except Exception:
        return DEFAULT_UI_TEXTS.get(key, "")

def _set_ui(conn, key: str, value: str) -> None:
    now = datetime.utcnow().isoformat()
    conn.execute(
        "INSERT INTO ui_texts(key,value,updated_at) VALUES(?,?,?) "
        "ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=excluded.updated_at;",
        (key, value, now),
    )

@router.post("/settings/save-theme")
async def save_theme_pref(request: Request, dark: str = "0", classic: str = "0"):
    adm = _get_admin(request)
    if not adm:
        return JSONResponse({"ok": False})
    try:
        conn = _db()
        conn.execute("""
            CREATE TABLE IF NOT EXISTS admin_preferences (
                admin_id TEXT NOT NULL, key TEXT NOT NULL, value TEXT,
                PRIMARY KEY (admin_id, key)
            );
        """)
        import db_conn as _dc
        _dc.ensure_unique_constraint(conn, "admin_preferences", ["admin_id", "key"])
        # ⚠️ INSERT OR REPLACE (خاص SQLite) به ON CONFLICT تبدیل شد — کلید ترکیبی
        # (admin_id,key)، پرتابل بین SQLite/Postgres
        conn.execute(
            "INSERT INTO admin_preferences (admin_id,key,value) VALUES (?,?,?) "
            "ON CONFLICT(admin_id,key) DO UPDATE SET value=excluded.value;",
            (str(adm[0]), "dark_mode", dark if dark in ("1", "0", "auto") else "0"))
        conn.execute(
            "INSERT INTO admin_preferences (admin_id,key,value) VALUES (?,?,?) "
            "ON CONFLICT(admin_id,key) DO UPDATE SET value=excluded.value;",
            (str(adm[0]), "classic_mode", "1" if classic == "1" else "0"))
        conn.commit(); conn.close()
        _cache_invalidate(f"admin_prefs:{adm[0]}")
    except Exception:
        pass
    return JSONResponse({"ok": True})


@router.get("/receipts", response_class=HTMLResponse)
async def card_receipts_page(request: Request, status: str = "pending", flash: str = "",
                             pay_q: str = "", pay_sort: str = "date_desc"):
    adm = _get_admin(request)
    guard = _require(adm, "payment")
    if guard: return guard
    from db import get_card_receipts, ensure_card_receipts_schema
    ensure_card_receipts_schema()
    receipts = get_card_receipts(status)

    tabs = ''.join(
        f'<a href="/admin/receipts?status={s}" class="px-3 py-1.5 rounded-lg text-xs border '
        f'{"bg-indigo-600 text-white" if status==s else "bg-white text-gray-500"}">{l}</a>'
        for s, l in [("pending","⏳ در انتظار"),("approved","✅ تأیید شده"),("rejected","❌ رد شده"),("","همه")]
    )

    def _method_cell(r):
        try:
            m = r["method"] if "method" in r.keys() else "card"
        except Exception:
            m = "card"
        if m and str(m).startswith("crypto"):
            net = "USDT" if "usdt" in str(m) else "TRX"
            try:
                tx = (r["txid"] or "")[:14] if "txid" in r.keys() else ""
            except Exception:
                tx = ""
            return (f'<span class="px-2 py-1 bg-amber-50 text-amber-700 border border-amber-200 rounded text-xs">₿ {net}</span>'
                    + (f'<div class="text-[10px] text-gray-400 mt-1" dir="ltr"><code>{e(tx)}…</code></div>' if tx else ""))
        return f'<a href="/admin/receipts/{r["id"]}/view" class="px-2 py-1 bg-indigo-50 text-indigo-700 rounded text-xs">💳 مشاهده رسید</a>'

    rows = "".join(f"""<tr class="border-b hover:bg-gray-50 text-sm">
      <td class="px-3 py-3 text-xs text-gray-400">#{r['id']}</td>
      <td class="px-3 py-3">{e(r['full_name'] or r['username'] or str(r['user_id']))}</td>
      <td class="px-3 py-3 font-bold text-green-600">{int(r['amount'] or 0):,}</td>
      <td class="px-3 py-3">{_method_cell(r)}</td>
      <td class="px-3 py-3"><span class="px-2 py-0.5 rounded text-xs
        {'bg-amber-100 text-amber-700' if r['status']=='pending' else 'bg-green-100 text-green-700' if r['status']=='approved' else 'bg-red-100 text-red-600'}">
        {'⏳' if r['status']=='pending' else '✅' if r['status']=='approved' else '❌'}
      </span></td>
      <td class="px-3 py-3 text-xs text-gray-400">{fa_date(r['created_at'] or '', with_time=True)}</td>
      {'<td class="px-3 py-3 flex gap-1"><form method="post" action="/admin/receipts/' + str(r["id"]) + '/approve"><button class="px-2 py-1 bg-green-600 text-white rounded text-xs">✅ تأیید</button></form><form method="post" action="/admin/receipts/' + str(r["id"]) + '/reject"><button class="px-2 py-1 bg-red-50 text-red-600 border border-red-200 rounded text-xs">❌ رد</button></form></td>' if r['status']=='pending' else '<td></td>'}
    </tr>""" for r in receipts) or "<tr><td colspan='7' class='text-center py-6 text-gray-400'>رسیدی یافت نشد</td></tr>"

    def _pay_link_fn(qq, srt):
        from urllib.parse import quote
        return f"/admin/receipts?status={status}&pay_q={quote(qq)}&pay_sort={srt}#payouts"
    payouts_html = ""
    try:
        payouts_html = _partner_payouts_section_html(pay_q, pay_sort, _pay_link_fn)
    except Exception:
        payouts_html = ""

    sell_section = ""
    if _has(adm, "ai_pricing"):
        sell_section = _iv_sell_requests_section_html()

    body = f"""
    <h1 class="text-2xl font-bold text-gray-800 mb-4">🧾 خرید و بخش مالی</h1>
    <h2 class="text-lg font-bold text-gray-700 mb-3">💳 رسیدهای پرداخت (کارت‌به‌کارت و رمزارز)</h2>
    <div class="flex gap-2 mb-4">{tabs}</div>
    <div class="card overflow-hidden"><div class="overflow-x-auto">
      <table class="w-full text-right min-w-max">
        <thead><tr class="text-xs text-gray-500 border-b bg-gray-50">
          <th class="px-3 py-2">#</th><th class="px-3 py-2">کاربر</th>
          <th class="px-3 py-2">مبلغ (ت)</th><th class="px-3 py-2">رسید</th>
          <th class="px-3 py-2">وضعیت</th><th class="px-3 py-2">تاریخ</th><th></th>
        </tr></thead>
        <tbody>{rows}</tbody>
      </table>
    </div></div>
    {payouts_html}
    {sell_section}"""
    return _layout("خرید و بخش مالی", body, adm, flash=flash)


def _iv_sell_requests_section_html() -> str:
    """درخواست‌های «می‌خوام بفروشم» از ویزارد کارشناسی آیفون — طبق درخواست صریح مالک
    پروژه، درست زیر لیست کارت‌به‌کارت روی همون صفحه (نه یک صفحهٔ کاملاً جدا)."""
    import iphone_valuation.db as ivdb
    reqs = ivdb.list_sell_requests()
    rows = "".join(f"""
      <tr class="border-b hover:bg-gray-50 text-sm {'opacity-50' if r['status']=='contacted' else ''}">
        <td class="px-3 py-3 text-xs text-gray-400">#{r['id']}</td>
        <td class="px-3 py-3">{e(r['user_name'] or str(r['user_id']))}</td>
        <td class="px-3 py-3" dir="ltr"><code>{e(r['phone'])}</code></td>
        <td class="px-3 py-3 text-xs">{e(r['city'] or '—')}</td>
        <td class="px-3 py-3 text-xs">{e(r['model_name'] or '—')}</td>
        <td class="px-3 py-3 font-bold text-green-600">{f"{r['estimated_price']:,}" if r['estimated_price'] else 'توافقی'}</td>
        <td class="px-3 py-3 text-xs text-gray-400">{fa_date(r['created_at'] or '', with_time=True)}</td>
        <td class="px-3 py-3">
          {'<span class="px-2 py-0.5 rounded text-xs bg-green-100 text-green-700">✅ تماس گرفته شد</span>' if r['status']=='contacted' else
           f'<form method="post" action="/admin/iphone/sell-requests/{r["id"]}/contacted"><button class="px-2 py-1 bg-indigo-600 text-white rounded text-xs">📞 تماس گرفتم</button></form>'}
        </td>
      </tr>""" for r in reqs) or "<tr><td colspan='8' class='text-center py-6 text-gray-400'>درخواستی ثبت نشده</td></tr>"
    return f"""
    <h2 class="text-xl font-bold text-gray-800 mb-4 mt-8">🤝 درخواست‌های «می‌خوام بفروشم» (کارشناسی آیفون)</h2>
    <div class="card overflow-hidden"><div class="overflow-x-auto">
      <table class="w-full text-right min-w-max">
        <thead><tr class="text-xs text-gray-500 border-b bg-gray-50">
          <th class="px-3 py-2">#</th><th class="px-3 py-2">کاربر</th><th class="px-3 py-2">شماره تماس</th>
          <th class="px-3 py-2">شهر</th><th class="px-3 py-2">مدل</th><th class="px-3 py-2">قیمت تخمینی</th>
          <th class="px-3 py-2">تاریخ</th><th class="px-3 py-2"></th>
        </tr></thead>
        <tbody>{rows}</tbody>
      </table>
    </div></div>"""


@router.post("/iphone/sell-requests/{rid}/contacted")
async def iphone_sell_request_contacted(request: Request, rid: int):
    adm = _get_admin(request)
    guard = _require(adm, "ai_pricing")
    if guard: return guard
    import iphone_valuation.db as ivdb
    ivdb.update_sell_request_status(rid, "contacted")
    _log(request, "علامت‌گذاری تماس با درخواست فروش", "کارشناسی آیفون", str(rid), admin_info=adm)
    return _redir("/admin/receipts")


@router.get("/receipts/{rid}/view", response_class=HTMLResponse)
async def receipt_view(request: Request, rid: int):
    adm = _get_admin(request)
    guard = _require(adm, "payment")
    if guard: return guard
    from db import get_card_receipts
    all_r = [r for r in get_card_receipts("") if r["id"] == rid]
    if not all_r:
        return _redir("/admin/receipts?flash=رسید+یافت+نشد")
    r = all_r[0]
    body = f"""
    <div class="flex items-center gap-3 mb-6">
      {_btn("← خرید و بخش مالی", "/admin/receipts", "slate", small=True)}
      <h1 class="text-xl font-bold text-gray-800">رسید #{rid}</h1>
    </div>
    <div class="grid md:grid-cols-2 gap-4">
      <div class="card p-5">
        <h2 class="font-bold text-gray-700 mb-3">اطلاعات پرداخت</h2>
        <div class="space-y-2 text-sm">
          <div class="flex justify-between"><span class="text-gray-400">کاربر</span><span>{e(r['full_name'] or str(r['user_id']))}</span></div>
          <div class="flex justify-between"><span class="text-gray-400">مبلغ</span><span class="font-bold text-green-600">{int(r['amount']):,} تومان</span></div>
          <div class="flex justify-between"><span class="text-gray-400">وضعیت</span><span>{r['status']}</span></div>
          <div class="flex justify-between"><span class="text-gray-400">تاریخ</span><span>{fa_date(r['created_at'] or '', with_time=True)}</span></div>
        </div>
        {'''<div class="mt-4 space-y-3">
          <label class="text-sm font-medium text-gray-700 block">مبلغ تأیید شده (تومان)</label>
          <form method="post" action="/admin/receipts/''' + str(rid) + '''/approve" class="flex gap-2">
            <input type="number" name="confirmed_amount" value="''' + str(int(r['amount'])) + '''" required
              class="flex-1 border border-gray-200 rounded-lg px-3 py-2 text-sm" placeholder="مبلغ واقعی واریز شده">
            <button class="px-4 py-2 bg-green-600 text-white rounded-lg text-sm font-medium whitespace-nowrap">✅ تأیید و شارژ</button>
          </form>
          <form method="post" action="/admin/receipts/''' + str(rid) + '''/reject">
            <button class="px-4 py-2 bg-red-50 text-red-600 border border-red-200 rounded-lg text-sm w-full">❌ رد درخواست</button>
          </form>
        </div>''' if r['status']=='pending' else '<p class="mt-3 text-sm text-gray-400">این رسید قبلاً بررسی شده است.</p>'}
        <div class="mt-4 pt-4 border-t border-gray-100">
          <form method="post" action="/admin/receipts/{rid}/delete" onsubmit="return confirm('⚠️ این رسید برای همیشه حذف می‌شود. ادامه؟')">
            <button class="px-4 py-2 bg-red-50 text-red-500 border border-red-200 rounded-lg text-sm w-full">🗑 حذف این رسید</button>
          </form>
        </div>
      </div>
      <div class="card p-5 text-center">
        <h2 class="font-bold text-gray-700 mb-3">تصویر رسید</h2>
        <img src="/admin/receipts/{rid}/image" alt="رسید" class="max-w-full rounded-xl border border-gray-200">
      </div>
    </div>"""
    return _layout(f"رسید #{rid}", body, adm)


@router.post("/receipts/delete-all")
async def receipts_delete_all(request: Request):
    adm = _get_admin(request)
    guard = _require(adm, "payment")
    if guard: return guard
    from db import delete_all_card_receipts
    delete_all_card_receipts()
    _log(request, "حذف همه رسیدهای کارت‌به‌کارت", "کیف‌پول", "bulk delete")
    return _redir("/admin/receipts?flash=همه+رسیدها+حذف+شدند")


@router.post("/receipts/{rid}/delete")
async def receipt_delete(request: Request, rid: int):
    adm = _get_admin(request)
    guard = _require(adm, "payment")
    if guard: return guard
    from db import delete_card_receipt
    delete_card_receipt(rid)
    _log(request, f"حذف رسید #{rid}", "کیف‌پول", f"receipt:{rid}")
    return _redir("/admin/receipts?flash=رسید+حذف+شد")


@router.get("/receipts/{rid}/image")
async def receipt_image(request: Request, rid: int):
    adm = _get_admin(request)
    guard = _require(adm, "payment")
    if guard: return guard
    from db import get_card_receipts
    all_r = [r for r in get_card_receipts("") if r["id"] == rid]
    if not all_r:
        from fastapi.responses import Response
        return Response("not found", 404)
    r = all_r[0]
    try:
        import requests as _req
        token = _env("BOT_TOKEN","")
        file_info = _req.get(f"https://api.telegram.org/bot{token}/getFile?file_id={r['file_id']}").json()
        file_path = file_info["result"]["file_path"]
        img = _req.get(f"https://api.telegram.org/file/bot{token}/{file_path}").content
        from fastapi.responses import Response
        return Response(img, media_type="image/jpeg")
    except Exception:
        from fastapi.responses import Response
        return Response("error", 500)


@router.post("/receipts/{rid}/approve")
async def receipt_approve(request: Request, rid: int):
    adm = _get_admin(request)
    guard = _require(adm, "payment")
    if guard: return guard
    form = await request.form()
    confirmed_amount = int(form.get("confirmed_amount") or 0)
    from db import get_card_receipts, update_card_receipt, add_wallet_balance
    all_r = [r for r in get_card_receipts("pending") if r["id"] == rid]
    if not all_r:
        return _redir("/admin/receipts?flash=رسید+یافت+نشد")
    r = all_r[0]
    # از مبلغ وارد شده استفاده کن، اگه نبود از مبلغ اصلی
    amount = confirmed_amount if confirmed_amount > 0 else int(r["amount"] or 0)
    update_card_receipt(rid, "approved", f"تأیید ادمین — مبلغ: {amount:,}", amount=amount)
    add_wallet_balance(r["user_id"], amount)
    _log(request, f"تأیید رسید #{rid}", "کیف‌پول", f"user:{r['user_id']} amount:{amount:,}")
    try:
        await run_in_threadpool(_tg_send, r["user_id"],
            f"✅ پرداخت شما تأیید شد!\n"
            f"مبلغ <b>{amount:,}</b> تومان به کیف پول شما اضافه شد.")
    except Exception:
        pass
    return _redir(f"/admin/receipts?flash=رسید+{rid}+تأیید+شد")


@router.post("/receipts/{rid}/reject")
async def receipt_reject(request: Request, rid: int):
    adm = _get_admin(request)
    guard = _require(adm, "payment")
    if guard: return guard
    from db import get_card_receipts, update_card_receipt
    all_r = [r for r in get_card_receipts("pending") if r["id"] == rid]
    if not all_r:
        return _redir("/admin/receipts?flash=رسید+یافت+نشد")
    r = all_r[0]
    update_card_receipt(rid, "rejected", "رد ادمین")
    _log(request, f"رد رسید #{rid}", "کیف‌پول", f"user:{r['user_id']}")
    try:
        await run_in_threadpool(_tg_send, r["user_id"],
            "❌ متأسفانه رسید پرداخت شما تأیید نشد.\n"
            "لطفاً با پشتیبانی تماس بگیرید.")
    except Exception: pass
    return _redir(f"/admin/receipts?flash=رسید+{rid}+رد+شد")


def _webhook_status_snippet() -> str:
    """خلاصه وضعیت اتصال ربات برای نمایش در صفحه تنظیمات."""
    try:
        from db import get_cfg
        mode = (get_cfg("bot_run_mode", "") or "").strip().lower()
    except Exception:
        mode = ""
    if mode == "webhook":
        return ('<div class="mb-3 px-3 py-2 bg-green-50 border border-green-200 rounded-lg text-sm text-green-700">'
                '🚀 Webhook فعال — سریع‌ترین حالت</div>')
    if mode == "polling":
        return ('<div class="mb-3 px-3 py-2 bg-blue-50 border border-blue-200 rounded-lg text-sm text-blue-700">'
                '🔄 Polling فعال</div>')
    if mode == "stopped":
        return ('<div class="mb-3 px-3 py-2 bg-gray-100 border border-gray-300 rounded-lg text-sm text-gray-700">'
                '⏸ ربات متوقف است</div>')
    return ('<div class="mb-3 px-3 py-2 bg-amber-50 border border-amber-200 rounded-lg text-sm text-amber-700">'
            '❓ حالت اتصال نامشخص — یک بار وارد صفحه مدیریت شوید</div>')


@router.get("/settings/panel", response_class=HTMLResponse)
async def settings_hub(request: Request, flash: str = ""):
    adm = _get_admin(request)
    guard = _require(adm, "settings")
    if guard: return guard

    # تم فعلی
    saved_dark = "auto"; saved_classic = "0"
    try:
        _prefs = _get_admin_prefs(adm[0])
        saved_dark = _prefs.get("dark_mode") or "auto"
        saved_classic = _prefs.get("classic_mode") or "0"
    except Exception:
        pass

    body = f"""
    <h1 class="text-2xl font-bold text-gray-800 mb-6">⚙️ تنظیمات</h1>
    <div class="grid md:grid-cols-2 gap-4">
      <div class="card p-6">
        <h2 class="font-bold text-gray-700 mb-3">🌗 حالت نمایش</h2>
        <p class="text-sm text-gray-500 mb-4">انتخاب شما ذخیره می‌شود و پس از خروج و ورود مجدد هم باقی می‌ماند.</p>

        <button id="btn-daynight" onclick="hubToggleDark()"
          class="w-full py-3 mb-2 rounded-xl text-sm font-semibold border transition flex items-center justify-center gap-2"></button>

        <button id="btn-classic" onclick="hubToggleClassic()"
          class="w-full py-3 mb-2 rounded-xl text-sm font-semibold border transition flex items-center justify-center gap-2"></button>

        <button id="btn-auto" onclick="hubSetAuto()"
          class="w-full py-3 rounded-xl text-sm font-semibold border transition flex items-center justify-center gap-2"></button>

        <p class="text-xs text-gray-400 mt-3">🖥 در حالت «هماهنگ با سیستم»، پنل به‌صورت خودکار با تم روز/شب دستگاه شما هماهنگ می‌شود.</p>

        <script>
        function _hubMode(){{ return localStorage.getItem('sl-dark') || '{saved_dark}' || 'auto'; }}
        function _hubIsDark(){{
          var m=_hubMode();
          if(m==='auto') return window.matchMedia&&window.matchMedia('(prefers-color-scheme: dark)').matches;
          return m==='1';
        }}
        function _hubSave(){{
          fetch('/admin/settings/save-theme?dark='+encodeURIComponent(localStorage.getItem('sl-dark')||'auto')
            +'&classic='+(localStorage.getItem('sl-classic')==='1'?'1':'0'), {{method:'POST'}}).catch(function(){{}});
        }}
        function hubRender(){{
          var m=_hubMode(), dark=_hubIsDark(), classic=localStorage.getItem('sl-classic')==='1';
          var dn=document.getElementById('btn-daynight');
          if(dark){{
            dn.textContent='☀️ رفتن به حالت روز';
            dn.className='w-full py-3 mb-2 rounded-xl text-sm font-semibold border transition flex items-center justify-center gap-2 bg-amber-50 text-amber-700 border-amber-200 hover:bg-amber-100';
          }} else {{
            dn.textContent='🌙 رفتن به حالت شب';
            dn.className='w-full py-3 mb-2 rounded-xl text-sm font-semibold border transition flex items-center justify-center gap-2 bg-gray-800 text-gray-100 border-gray-600 hover:bg-gray-700';
          }}
          var bc=document.getElementById('btn-classic');
          bc.textContent = classic ? '🎨 حالت کلاسیک: روشن — بازگشت به پیش‌فرض' : '🎨 فعال‌کردن حالت کلاسیک';
          bc.className='w-full py-3 mb-2 rounded-xl text-sm font-semibold border transition flex items-center justify-center gap-2 '
            +(classic?'bg-blue-600 text-white border-blue-600 hover:bg-blue-700':'bg-blue-50 text-blue-700 border-blue-200 hover:bg-blue-100');
          var ba=document.getElementById('btn-auto');
          ba.textContent = m==='auto' ? '🖥 هماهنگ با سیستم: فعال ✓' : '🖥 هماهنگ با سیستم';
          ba.className='w-full py-3 rounded-xl text-sm font-semibold border transition flex items-center justify-center gap-2 '
            +(m==='auto'?'bg-teal-600 text-white border-teal-600':'bg-teal-50 text-teal-700 border-teal-200 hover:bg-teal-100');
        }}
        function hubToggleDark(){{
          localStorage.setItem('sl-dark', _hubIsDark()?'0':'1');
          window.applyMode&&window.applyMode(); _hubSave(); hubRender();
        }}
        function hubToggleClassic(){{
          var on=localStorage.getItem('sl-classic')==='1';
          localStorage.setItem('sl-classic', on?'0':'1');
          window.applyMode&&window.applyMode(); _hubSave(); hubRender();
        }}
        function hubSetAuto(){{
          localStorage.setItem('sl-dark','auto');
          window.applyMode&&window.applyMode(); _hubSave(); hubRender();
        }}
        hubRender();
        </script>
      </div>
      <div class="card p-6">
        <h2 class="font-bold text-gray-700 mb-3">🔘 مدیریت دکمه‌ها</h2>
        <p class="text-sm text-gray-500 mb-4">ویرایش برچسب دکمه‌ها، فعال/غیرفعال‌سازی و تنظیمات متنی مهم.</p>
        <a href="/admin/settings" class="block w-full py-2.5 bg-indigo-50 text-indigo-700 border border-indigo-200 rounded-lg text-sm font-medium text-center hover:bg-indigo-100 transition">
          رفتن به تنظیمات دکمه‌ها ←
        </a>
      </div>
      <!-- حالت تعمیرات -->
      <div class="card p-6">
        <h2 class="font-bold text-gray-700 mb-3">🚧 حالت تعمیرات</h2>
        <p class="text-sm text-gray-500 mb-4">وقتی فعاله، فقط ادمین می‌تونه ربات رو استفاده کنه.</p>
        {'<div class="mb-3 px-3 py-2 bg-red-50 border border-red-200 rounded-lg text-sm text-red-700 font-medium">⚠️ ربات الان در حالت تعمیرات است</div>' if __import__("db").get_maintenance_mode() else '<div class="mb-3 px-3 py-2 bg-green-50 border border-green-200 rounded-lg text-sm text-green-700">✅ ربات فعال است</div>'}
        <div class="flex gap-3">
          <form method="post" action="/admin/settings/maintenance?enable=1">
            <button class="px-4 py-2 bg-red-600 text-white rounded-lg text-sm font-medium hover:bg-red-700">🚧 فعال‌کردن تعمیرات</button>
          </form>
          <form method="post" action="/admin/settings/maintenance?enable=0">
            <button class="px-4 py-2 bg-green-600 text-white rounded-lg text-sm font-medium hover:bg-green-700">✅ بازگشایی ربات</button>
          </form>
        </div>
      </div>
      <!-- Webhook -->
      <div class="card p-6">
        <h2 class="font-bold text-gray-700 mb-3">🔗 اتصال ربات (Webhook)</h2>
        <p class="text-sm text-gray-500 mb-4">مدیریت نحوه اتصال ربات به تلگرام — Webhook (سریع‌تر) یا Polling.</p>
        {_webhook_status_snippet()}
        <a href="/admin/webhook" class="inline-block px-4 py-2 bg-indigo-600 text-white rounded-lg text-sm font-medium hover:bg-indigo-700">
          🔧 مدیریت کامل Webhook</a>
      </div>
    </div>"""
    return _layout("تنظیمات", body, adm, flash=flash)


# ══════════════════════════════════════════════════════════════════════════
# ─── تأیید قوانین قبل از خرید (بخش ۷ سند مینی‌اپ) ─────────────────────────────
# ══════════════════════════════════════════════════════════════════════════

_PURCHASE_TERMS_CFG_KEY = "PURCHASE_TERMS_TEXT"

@router.get("/settings/purchase-terms", response_class=HTMLResponse)
async def purchase_terms_page(request: Request, flash: str = ""):
    adm = _get_admin(request)
    guard = _require(adm, "settings")
    if guard: return guard
    from db import get_cfg
    text = get_cfg(_PURCHASE_TERMS_CFG_KEY, "")
    body = f"""
    <div class="flex items-center gap-3 mb-6">
      {_btn("← تنظیمات","/admin/settings/panel","slate",small=True)}
      <h1 class="text-2xl font-bold text-gray-800">📜 قوانین خرید</h1>
    </div>
    <div class="card p-6 max-w-2xl">
      <p class="text-sm text-gray-500 mb-4">این متن، پیش‌فرض عمومیه — وقتی برای یک محصول «نیاز به تأیید قوانین خرید» فعال باشه ولی متن اختصاصی خودِ اون محصول خالی گذاشته شده باشه، همین متن قبل از پرداخت به کاربر نشون داده می‌شه (هم در ربات، هم در مینی‌اپ). هر محصول می‌تونه از فرم ویرایش خودش، متن قوانین اختصاصی جدا داشته باشه — چون ممکنه محصولات مختلف قوانین متفاوتی داشته باشن. کاربر فقط با تیک‌زدن چک‌باکس تأیید می‌تونه ادامه بده.</p>
      <form method="post" action="/admin/settings/purchase-terms">
        {_textarea("text", "مثلاً: کالای دیجیتال پس از تحویل، قابل استرداد نیست مگر در صورت خرابی...", value=text, rows=10)}
        <div class="mt-3">{_btn("💾 ذخیره", color="green")}</div>
      </form>
    </div>"""
    return _layout("قوانین خرید", body, adm, flash=flash)


@router.post("/settings/purchase-terms")
async def purchase_terms_save(request: Request, text: str = Form("")):
    adm = _get_admin(request)
    guard = _require(adm, "settings")
    if guard: return guard
    from db import set_cfg
    set_cfg(_PURCHASE_TERMS_CFG_KEY, text.strip())
    _log(request, "ویرایش متن قوانین خرید", "قوانین خرید", "", admin_info=adm)
    return _redir(f"/admin/settings/purchase-terms?flash={e('✅ ذخیره شد')}")


@router.get("/settings", response_class=HTMLResponse)
async def settings_get(request: Request, group: str = "", flash: str = ""):
    adm = _get_admin(request)
    guard = _require(adm, "settings")
    if guard: return guard

    try:
        from ui_texts import (
            DEFAULT_UI_TEXTS as _D,
            EDITABLE_BUTTON_GROUPS as _BG,
            CRITICAL_TEXT_KEYS as _CTK,
            CRITICAL_TEXT_LABELS as _CTL,
            BUTTON_ICONS as _BICONS,
            MAIN_BUTTON_KEYS as _BK,
        )
    except ImportError:
        _D = {}; _BG = {}; _CTK = []; _CTL = {}; _BICONS = {}; _BK = []

    conn = _db()
    try:
        db_texts = {r["key"]: r["value"] for r in conn.execute("SELECT key, value FROM ui_texts;").fetchall()}
        btn_states = {k: db_texts.get(f"MAIN_BTN_ENABLED_{k}", "1") not in ("0","false","off","no") for k in _BK}
    finally:
        conn.close()

    def gv(k): return db_texts.get(k, _D.get(k, ""))

    # ─── Toggle switch CSS ─────────────────────────────────────────────────
    tog_css = """<style>
.tog{display:inline-flex;align-items:center;cursor:pointer;flex-shrink:0}
.tog input{display:none}
.tog-track{width:44px;height:24px;background:#d1d5db;border-radius:12px;position:relative;transition:background .22s;flex-shrink:0}
.tog-track::after{content:'';position:absolute;width:18px;height:18px;border-radius:50%;background:#fff;top:3px;left:3px;transition:transform .22s;box-shadow:0 1px 3px rgba(0,0,0,.25)}
.tog input:checked~.tog-track{background:#6366f1}
.tog input:checked~.tog-track::after{transform:translateX(20px)}
.tog-off .tog-track{background:#ef4444}
.field-row{display:flex;align-items:center;gap:10px;padding:10px 14px;border:1px solid #f1f5f9;border-radius:12px;transition:background .15s}
.field-row:hover{background:#f8fafc}
.field-inp{flex:1;border:1px solid #e2e8f0;border-radius:8px;padding:7px 12px;font-size:.85rem;background:#fff;outline:none;transition:border-color .15s,box-shadow .15s;direction:rtl}
.field-inp:focus{border-color:#6366f1;box-shadow:0 0 0 2px rgba(99,102,241,.15)}
.save-btn{display:inline-flex;align-items:center;gap:6px;padding:9px 22px;background:#6366f1;color:#fff;border:none;border-radius:10px;font-size:.9rem;font-weight:600;cursor:pointer;transition:opacity .2s,transform .1s}
.save-btn:disabled{opacity:.35;cursor:not-allowed}
.save-btn:not(:disabled):hover{opacity:.9}
.save-btn:not(:disabled):active{transform:scale(.97)}
.sec-hdr{font-weight:700;color:#374151;font-size:1rem;display:flex;align-items:center;gap:8px;margin-bottom:4px}
.sec-sub{font-size:.78rem;color:#9ca3af;margin-bottom:14px}
</style>"""

    # ─── Section 1: دکمه‌های منوی اصلی (با toggle + برچسب) ──────────────
    main_rows = ""
    for key in _BG.get("دکمه‌های منوی اصلی", []):
        en = btn_states.get(key, True)
        val = gv(key)
        icon = _BICONS.get(key, "")
        main_rows += f"""
      <div class="field-row">
        <label class="tog" title="{'فعال' if en else 'غیرفعال'}">
          <input type="checkbox" name="enable_{e(key)}" {"checked" if en else ""} onchange="markDirty();this.closest('.tog').classList.toggle('tog-off',!this.checked)">
          <span class="tog-track"></span>
        </label>
        <span class="field-icon">{icon}</span>
        <input type="text" class="field-inp" name="field_{e(key)}" value="{e(val)}" oninput="markDirty()" placeholder="برچسب دکمه">
      </div>"""

    # ─── Section 2: دکمه‌های پنل همکار (فقط برچسب) ──────────────────────
    partner_rows = ""
    for key in _BG.get("دکمه‌های پنل همکار", []):
        val = gv(key)
        icon = _BICONS.get(key, "")
        partner_rows += f"""
      <div class="field-row field-row-sm">
        <span class="field-icon-sm">{icon}</span>
        <input type="text" class="field-inp" name="field_{e(key)}" value="{e(val)}" oninput="markDirty()">
      </div>"""

    # ─── Section 3: دکمه‌های کیف‌پول (فقط برچسب) ────────────────────────
    wallet_rows = ""
    for key in _BG.get("دکمه‌های کیف‌پول و پرداخت", []):
        val = gv(key)
        icon = _BICONS.get(key, "")
        wallet_rows += f"""
      <div class="field-row field-row-sm">
        <span class="field-icon-sm">{icon}</span>
        <input type="text" class="field-inp" name="field_{e(key)}" value="{e(val)}" oninput="markDirty()">
      </div>"""

    # ─── Section 4: تنظیمات متنی مهم ─────────────────────────────────────
    text_fields = ""
    for key in _CTK:
        val = gv(key)
        lbl = _CTL.get(key, key)
        if key == "WALLET_QUICK_AMOUNTS":
            fld = f'<input type="text" class="field-inp input-full" name="field_{e(key)}" value="{e(val)}" oninput="markDirty()" placeholder="10000,50000,100000,500000">'
        else:
            fld = f'<textarea class="field-inp textarea-full" name="field_{e(key)}" oninput="markDirty()" dir="rtl">{e(val)}</textarea>'
        text_fields += f"""
      <div class="mb-18">
        <label class="text-section-label">{e(lbl)}</label>
        {fld}
      </div>"""

    body = f"""
    {tog_css}

    <div class="settings-header">
      <h1 class="settings-title">⚙️ تنظیمات ربات</h1>
      <button id="sbtn-top" form="sf" type="submit" class="save-btn" disabled>💾 ذخیره تغییرات</button>
    </div>

    <form id="sf" method="post" action="/admin/settings/save-all">
      <input type="hidden" name="is_combined" value="1">

      <!-- ─── دکمه‌های منوی اصلی ─────────────────────────────────────── -->
      <div class="card form-card">
        <div class="sec-hdr">🔘 دکمه‌های منوی اصلی</div>
        <div class="sec-sub">دکمه‌های Reply Keyboard در منوی کاربران — می‌توانید نمایش هر دکمه را فعال یا غیرفعال کنید.</div>
        <div class="flex-col-6">
          {main_rows}
        </div>
        <div class="card-reset-footer">
          <button type="button" onclick="resetSection('main')" class="btn-link-muted">
            🔄 بازگردانی این بخش به پیش‌فرض
          </button>
        </div>
      </div>

      <!-- ─── دکمه‌های پنل همکار + کیف‌پول (۲ ستون) ─────────────────── -->
      <div class="grid-2col mb-16">
        <div class="card form-card mb-0">
          <div class="sec-hdr">🤝 دکمه‌های پنل همکار</div>
          <div class="sec-sub">Inline Keyboard — داشبورد همکاران</div>
          <div class="flex-col-6">
            {partner_rows}
          </div>
        </div>
        <div class="card form-card mb-0">
          <div class="sec-hdr">💰 دکمه‌های کیف‌پول</div>
          <div class="sec-sub">Inline Keyboard — بخش کیف‌پول</div>
          <div class="flex-col-6">
            {wallet_rows}
          </div>
        </div>
      </div>

      <!-- ─── تنظیمات متنی مهم ─────────────────────────────────────────── -->
      <div class="card form-card mb-20">
        <div class="sec-hdr">📝 تنظیمات متنی</div>
        <div class="sec-sub">این متن‌ها مستقیماً در ربات نمایش داده می‌شوند — بقیه متن‌ها ثابت و پیش‌فرض هستند.</div>
        {text_fields}
      </div>

      <!-- ─── دکمه‌های پایین صفحه ──────────────────────────────────────── -->
      <div class="footer-bar">
        <button type="button" onclick="confirmResetAll()" class="btn-link-muted">
          🔄 بازگردانی همه به پیش‌فرض
        </button>
        <button id="sbtn-bot" form="sf" type="submit" class="save-btn" disabled>💾 ذخیره همه تغییرات</button>
      </div>
    </form>

    <script>
    var _dirty = false;
    function markDirty() {{
      if (_dirty) return;
      _dirty = true;
      ['sbtn-top','sbtn-bot'].forEach(function(id){{
        var b=document.getElementById(id);
        if(b){{b.disabled=false;}}
      }});
    }}
    window.addEventListener('beforeunload', function(e){{
      if(_dirty){{e.preventDefault();e.returnValue='';}}
    }});
    document.getElementById('sf').addEventListener('submit',function(){{_dirty=false;}});

    function resetSection(sec) {{
      if(!confirm('تنظیمات این بخش به پیش‌فرض برگردانده شود؟')) return;
      var fd = new FormData();
      fd.append('section', sec);
      fetch('/admin/settings/reset-section', {{method:'POST', body:fd}})
        .then(function(){{location.reload();}});
    }}
    function confirmResetAll() {{
      if(!confirm('همه تنظیمات دکمه‌ها و متن‌ها به حالت اولیه بازگردانده شوند؟')) return;
      fetch('/admin/settings/reset-all', {{method:'POST'}})
        .then(function(){{location.reload();}});
    }}
    </script>"""

    return _layout("تنظیمات", body, adm, flash=flash)


@router.post("/settings/maintenance")
async def settings_maintenance(request: Request, enable: str = "0"):
    adm = _get_admin(request)
    guard = _require(adm, "settings")
    if guard: return guard
    from db import set_maintenance_mode
    on = enable == "1"
    set_maintenance_mode(on)
    _log(request, "حالت تعمیرات", "تنظیمات", "فعال" if on else "غیرفعال")
    msg = "تعمیرات+فعال+شد" if on else "ربات+فعال+شد"
    return _redir(f"/admin/settings/panel?flash={msg}")


def _settings_action_bar():
    return """
      <div class="flex items-center gap-3 pt-4 border-t">
        <button type="submit" id="save-btn" disabled
          class="px-6 py-2.5 bg-indigo-600 text-white rounded-xl text-sm font-semibold transition disabled-visual">
          💾 ذخیره تغییرات
        </button>
        <button type="submit" formaction="/admin/settings/reset-group" onclick="return confirmReset()"
          class="px-5 py-2.5 bg-gray-100 text-gray-600 rounded-xl text-sm font-medium hover:bg-gray-200 transition">
          🔄 بازگردانی پیش‌فرض
        </button>
      </div>"""


@router.post("/settings/theme")
async def settings_theme_save(request: Request):
    adm = _get_admin(request)
    guard = _require(adm, "panel_appearance")
    if guard: return guard
    _ensure_theme_table()
    form = await request.form()
    conn = _db()
    try:
        for key in ("primary", "glass"):
            val = str(form.get(key) or "").strip()
            if val:
                conn.execute(
                    "INSERT INTO panel_theme (key, value) VALUES (?,?) ON CONFLICT(key) DO UPDATE SET value=excluded.value;",
                    (key, val)
                )
        conn.commit()
    finally:
        conn.close()
    _cache_invalidate("panel_theme")
    return _redir("/admin/settings?flash=تنظیمات+رنگ+ذخیره+شد")


@router.post("/settings/theme/reset")
async def settings_theme_reset(request: Request):
    adm = _get_admin(request)
    guard = _require(adm, "panel_appearance")
    if guard: return guard
    conn = _db()
    try:
        conn.execute("DELETE FROM panel_theme;")
        conn.commit()
    finally:
        conn.close()
    _cache_invalidate("panel_theme")
    return _redir("/admin/settings?flash=رنگ‌های+پیش‌فرض+بازگردانده+شد")


def _clear_ui_cache():
    try:
        from ui_texts import ui_cache_clear
        ui_cache_clear()
    except Exception:
        pass


@router.post("/settings/save-all")
async def settings_save_all(request: Request):
    adm = _get_admin(request)
    guard = _require(adm, "settings")
    if guard: return guard
    form = await request.form()
    is_combined = form.get("is_combined") == "1"

    try:
        from ui_texts import (
            DEFAULT_UI_TEXTS as _D,
            EDITABLE_BUTTON_GROUPS as _BG,
            CRITICAL_TEXT_KEYS as _CTK,
            MAIN_BUTTON_KEYS as _BK,
        )
    except ImportError:
        _D = {}; _BG = {}; _CTK = []; _BK = []

    conn = _db()
    try:
        if is_combined:
            # ─── ذخیره برچسب همه دکمه‌ها ───────────────────────────────
            all_btn_keys = [k for keys in _BG.values() for k in keys]
            for key in all_btn_keys:
                new_val = form.get(f"field_{key}")
                if new_val is None:
                    continue
                new_val = str(new_val).strip()
                default = _D.get(key, "")
                if new_val == "" or new_val == default:
                    conn.execute("DELETE FROM ui_texts WHERE key=?;", (key,))
                else:
                    conn.execute(
                        "INSERT INTO ui_texts (key,value,updated_at) VALUES (?,?,datetime('now')) "
                        "ON CONFLICT(key) DO UPDATE SET value=excluded.value,updated_at=datetime('now');",
                        (key, new_val)
                    )

            # ─── ذخیره وضعیت فعال/غیرفعال دکمه‌های منو ─────────────────
            for key in _BK:
                enabled = form.get(f"enable_{key}") is not None  # checkbox: present=True
                conn.execute(
                    "INSERT INTO ui_texts (key,value,updated_at) VALUES (?,?,datetime('now')) "
                    "ON CONFLICT(key) DO UPDATE SET value=excluded.value,updated_at=datetime('now');",
                    (f"MAIN_BTN_ENABLED_{key}", "1" if enabled else "0")
                )

            # ─── ذخیره متن‌های مهم ──────────────────────────────────────
            for key in _CTK:
                new_val = form.get(f"field_{key}")
                if new_val is None:
                    continue
                new_val = str(new_val).strip()
                default = _D.get(key, "")
                if new_val == "" or new_val == default:
                    conn.execute("DELETE FROM ui_texts WHERE key=?;", (key,))
                else:
                    conn.execute(
                        "INSERT INTO ui_texts (key,value,updated_at) VALUES (?,?,datetime('now')) "
                        "ON CONFLICT(key) DO UPDATE SET value=excluded.value,updated_at=datetime('now');",
                        (key, new_val)
                    )
        conn.commit()
    finally:
        conn.close()
    _clear_ui_cache()
    _log(request, "ذخیره تنظیمات", "تنظیمات", "combined")
    return _redir("/admin/settings?flash=✅+تغییرات+ذخیره+شد")


@router.post("/settings/reset-section")
async def settings_reset_section(request: Request):
    """بازگردانی یک بخش خاص به پیش‌فرض"""
    adm = _get_admin(request)
    guard = _require(adm, "settings")
    if guard: return guard
    form = await request.form()
    section = str(form.get("section", ""))
    try:
        from ui_texts import EDITABLE_BUTTON_GROUPS as _BG, MAIN_BUTTON_KEYS as _BK
    except ImportError:
        _BG = {}; _BK = []

    conn = _db()
    try:
        if section == "main":
            for k in _BG.get("دکمه‌های منوی اصلی", []):
                conn.execute("DELETE FROM ui_texts WHERE key=?;", (k,))
            for k in _BK:
                conn.execute("DELETE FROM ui_texts WHERE key=?;", (f"MAIN_BTN_ENABLED_{k}",))
        conn.commit()
    finally:
        conn.close()
    _clear_ui_cache()
    _log(request, "بازگردانی بخش", "تنظیمات", section)
    return _redir("/admin/settings?flash=🔄+به+پیش‌فرض+بازگردانده+شد")


@router.post("/settings/reset-all")
async def settings_reset_all(request: Request):
    """بازگردانی همه تنظیمات دکمه‌ها و متن‌ها به پیش‌فرض"""
    adm = _get_admin(request)
    guard = _require(adm, "settings")
    if guard: return guard
    try:
        from ui_texts import EDITABLE_BUTTON_GROUPS as _BG, CRITICAL_TEXT_KEYS as _CTK, MAIN_BUTTON_KEYS as _BK
    except ImportError:
        _BG = {}; _CTK = []; _BK = []

    conn = _db()
    try:
        all_keys = [k for keys in _BG.values() for k in keys] + _CTK
        for k in all_keys:
            conn.execute("DELETE FROM ui_texts WHERE key=?;", (k,))
        for k in _BK:
            conn.execute("DELETE FROM ui_texts WHERE key=?;", (f"MAIN_BTN_ENABLED_{k}",))
        conn.commit()
    finally:
        conn.close()
    _clear_ui_cache()
    _log(request, "بازگردانی همه", "تنظیمات", "all")
    return _redir("/admin/settings?flash=🔄+همه+تنظیمات+به+پیش‌فرض+بازگردانده+شد")


@router.post("/settings/reset-group")
async def settings_reset_group(request: Request):
    """backward compat — همه را به پیش‌فرض برمی‌گرداند"""
    return await settings_reset_all(request)


@router.post("/settings/save-field")
async def settings_save_field(request: Request, key: str = Form(""), value: str = Form(""),
                               group: str = Form("")):
    adm = _get_admin(request)
    guard = _require(adm, "settings")
    if guard: return guard

    if key:
        conn = _db()
        try:
            now = datetime.now().isoformat()
            conn.execute("INSERT INTO ui_texts(key,value,updated_at) VALUES(?,?,?) "
                        "ON CONFLICT(key) DO UPDATE SET value=excluded.value,updated_at=excluded.updated_at;",
                        (key, value.strip(), now))
            conn.commit()
        finally:
            conn.close()
        try:
            from ui_texts import ui_cache_clear; ui_cache_clear()
        except Exception: pass

    return _redir(f"/admin/settings?group={e(group)}&flash=ذخیره+شد")


@router.post("/settings/save-group")
async def settings_save_group(request: Request, group: str = Form("")):
    adm = _get_admin(request)
    guard = _require(adm, "settings")
    if guard: return guard

    try:
        from ui_texts import TEXT_GROUPS as _GROUPS
    except ImportError:
        _GROUPS = {}

    form = await request.form()
    keys = _GROUPS.get(group, [])
    conn = _db()
    try:
        now = datetime.now().isoformat()
        for key in keys:
            val = (form.get(key) or "").strip()
            if val:
                conn.execute("INSERT INTO ui_texts(key,value,updated_at) VALUES(?,?,?) "
                            "ON CONFLICT(key) DO UPDATE SET value=excluded.value,updated_at=excluded.updated_at;",
                            (key, val, now))
        conn.commit()
    finally:
        conn.close()
    try:
        from ui_texts import ui_cache_clear; ui_cache_clear()
    except Exception: pass
    return _redir(f"/admin/settings?group={e(group)}&flash=همه+فیلدهای+این+بخش+ذخیره+شدند")


@router.post("/settings/reset-field")
async def settings_reset_field(request: Request, key: str = Form(""), group: str = Form("")):
    adm = _get_admin(request)
    guard = _require(adm, "settings")
    if guard: return guard

    if key:
        conn = _db()
        try:
            conn.execute("DELETE FROM ui_texts WHERE key=?;", (key,))
            conn.commit()
        finally:
            conn.close()
        try:
            from ui_texts import ui_cache_clear; ui_cache_clear()
        except Exception: pass

    return _redir(f"/admin/settings?group={e(group)}&flash={e(key)}+به+پیش‌فرض+بازگشت")


@router.post("/settings/toggle-btn")
async def settings_toggle_btn(request: Request, key: str = Form("")):
    adm = _get_admin(request)
    guard = _require(adm, "settings")
    if guard: return guard

    if key not in MAIN_BUTTONS:
        return _redir("/admin/settings?tab=buttons")
    flag_key = f"MAIN_BTN_ENABLED_{key}"
    conn = _db()
    try:
        cur_val = _get_ui(conn, flag_key)
        is_on = cur_val not in ("0", "false", "off", "no")
        # prevent disabling last button
        if is_on:
            enabled_count = sum(
                1 for k in MAIN_BUTTONS
                if _get_ui(conn, f"MAIN_BTN_ENABLED_{k}") not in ("0", "false", "off", "no")
            )
            if enabled_count <= 1:
                conn.close()
                return _redir("/admin/settings?tab=buttons&flash=حداقل+یک+دکمه+باید+فعال+بماند")
        _set_ui(conn, flag_key, "0" if is_on else "1")
        conn.commit()
    finally:
        conn.close()
    return _redir("/admin/settings?tab=buttons&flash=وضعیت+دکمه+تغییر+کرد")

@router.post("/settings/add-svc")
async def settings_add_svc(request: Request, title: str = Form(""), emoji: str = Form("🧩")):
    adm = _get_admin(request)
    guard = _require(adm, "settings")
    if guard: return guard

    title = title.strip()
    if not title:
        return _redir("/admin/settings?tab=services")
    key = title.replace(" ", "_")[:32]
    conn = _db()
    try:
        now = datetime.utcnow().isoformat()
        conn.execute(
            "INSERT OR IGNORE INTO other_services (service_key, title, emoji, is_active, created_at) VALUES (?,?,?,1,?);",
            (key, title, emoji.strip() or "🧩", now),
        )
        conn.commit()
    finally:
        conn.close()
    return _redir("/admin/settings?tab=services&flash=دسته+اضافه+شد")

@router.post("/settings/toggle-svc")
async def settings_toggle_svc(request: Request, key: str = Form("")):
    adm = _get_admin(request)
    guard = _require(adm, "settings")
    if guard: return guard
    conn = _db()
    try:
        conn.execute("UPDATE other_services SET is_active=CASE WHEN is_active=1 THEN 0 ELSE 1 END WHERE service_key=?;", (key,))
        conn.commit()
    finally:
        conn.close()
    return _redir("/admin/settings?tab=services&flash=وضعیت+تغییر+کرد")

@router.post("/settings/delete-svc")
async def settings_delete_svc(request: Request, key: str = Form("")):
    adm = _get_admin(request)
    guard = _require(adm, "settings")
    if guard: return guard
    if key == "general":
        return _redir("/admin/settings?tab=services")
    conn = _db()
    try:
        conn.execute("DELETE FROM product_feed WHERE product_id IN (SELECT id FROM products WHERE category=?);", (key,))
        conn.execute("DELETE FROM products WHERE category=?;", (key,))
        conn.execute("DELETE FROM other_services WHERE service_key=?;", (key,))
        conn.commit()
    finally:
        conn.close()
    return _redir("/admin/settings?tab=services&flash=دسته+حذف+شد")

# ─────────────────────────── Database ──────────────────────────────────────

# ══════════════════════════════════════════════════════════════════════════════
# ─── بکاپ / ریستور / ریست (فرمت اختصاصی .stbak) ─────────────────────────────
# ══════════════════════════════════════════════════════════════════════════════

# ══════════════════════════════════════════════════════════════════════════════
# ─── Backup / Restore / Reset ────────────────────────────────────────────────
# ══════════════════════════════════════════════════════════════════════════════

import threading as _threading, uuid as _uuid, tempfile as _tempfile

_JOB_DIR = "/tmp/stbak_jobs"
try:
    os.makedirs(_JOB_DIR, exist_ok=True)
except Exception:
    _JOB_DIR = "/tmp"


def _job_write(job_id: str, data: dict):
    try:
        import json as _j
        path = f"{_JOB_DIR}/{job_id}.json"
        tmp  = path + ".tmp"
        with open(tmp, "w", encoding="utf-8") as f:
            _j.dump(data, f)
        os.replace(tmp, path)
    except Exception:
        pass  # بی‌صدا fail کن


def _job_read(job_id: str) -> dict:
    try:
        import json as _j
        path = f"{_JOB_DIR}/{job_id}.json"
        with open(path, encoding="utf-8") as f:
            return _j.load(f)
    except Exception:
        return {"status": "not_found", "progress": 0}


def _job_start(fn, *args, **kwargs) -> str:
    job_id = str(_uuid.uuid4())[:8]
    _job_write(job_id, {"status": "running", "progress": 5, "message": "", "result": None})

    def _worker():
        try:
            def _cb(pct):
                _job_write(job_id, {"status": "running", "progress": int(pct), "message": "", "result": None})
            result = fn(*args, progress_cb=_cb, **kwargs)
            if isinstance(result, bytes):
                fpath = f"{_JOB_DIR}/{job_id}.stbak"
                with open(fpath, "wb") as f:
                    f.write(result)
                _job_write(job_id, {"status": "done", "progress": 100, "message": "", "result": {"file": fpath}})
            else:
                _job_write(job_id, {"status": "done", "progress": 100, "message": "", "result": result})
        except Exception as ex:
            _job_write(job_id, {"status": "error", "progress": 0, "message": str(ex), "result": None})

    _threading.Thread(target=_worker, daemon=True).start()
    return job_id


@router.get("/database", response_class=HTMLResponse)
async def database_page(request: Request, flash: str = ""):
    adm = _get_admin(request)
    guard = _require_any(adm, ["database", "backup", "restore", "recovery"])
    if guard: return guard

    # ☁️ داده‌های بکاپ ابری
    from backup_uploader import get_cloud_settings as _gcs
    from db import get_cfg as _gk
    _cb = _gcs()
    _cb_last_ok = _gk("cloudbk_last_ok", "")
    _gdrive_env_ok = bool(os.getenv("GDRIVE_CLIENT_ID","").strip() and os.getenv("GDRIVE_FOLDER_ID","").strip())
    _gdrive_connected = bool(_cb.get("gdrive_refresh_token",""))
    _cb_report_html = ""
    try:
        import json as _j
        _rep = _j.loads(_gk("cloudbk_last_report", "") or "{}")
        _errs = [r for r in _rep.get("results", []) if not r.get("ok")]
        if _errs:
            _cb_report_html = '<span class="text-red-500 bidi-plain">آخرین خطا: ' + e(str(_errs[0].get("driver")) + " — " + str(_errs[0].get("error"))[:100]) + "</span>"
    except Exception:
        pass

    from stbak_engine import MODULES, SECTION_LABELS
    import glob as _gl, os as _os

    # لیست بکاپ‌های موجود — بر اساس DB_DIALECT سوییچ می‌کنه (فایل SQLite/.stbak
    # یا pg_dump/.stbak، هرکدوم که واقعاً دادهٔ زندهٔ پروژه رو پوشش می‌ده).
    import db_conn as _dbc_page
    _is_pg = _dbc_page.is_postgres()
    try:
        if _is_pg:
            import pg_backup
            _auto_list = pg_backup.list_local_backups()[:_MAX_BACKUPS]
        else:
            from stbak_engine import list_local_backups as _list_stbak
            _auto_list = _list_stbak(_BACKUP_DIR)[:_MAX_BACKUPS]
    except Exception:
        _auto_list = []

    # آخرین ۵ بکاپ برای بخش «بازیابی» (کشویی) — هم برای نمایش هم برای دکمهٔ
    # بازیابی مستقیم هر ردیف.
    _last5 = _auto_list[:5]
    _restore_rows = ""
    for _b in _last5:
        _fn = _b["name"]
        _sz = _b["size"]
        _szs = f"{_sz//1024} KB" if _sz < 1024*1024 else f"{_sz/1024/1024:.1f} MB"
        try:
            from datetime import datetime as _dt2
            from db import fa_date as _fad2
            _iso = _dt2.fromtimestamp(_b["mtime"]).strftime("%Y-%m-%d %H:%M:%S")
            _label = _fad2(_iso, with_time=True)  # شمسی + اعداد فارسی
        except Exception:
            _label = _fn
        _restore_rows += f"""
          <div class="flex items-center justify-between gap-2 py-2 px-3 border-b border-gray-100 last:border-0">
            <div class="text-sm text-gray-700">{_label}<span class="text-xs text-gray-400 mr-2">({_szs})</span></div>
            <div class="flex gap-1.5 shrink-0">
              <a href="/admin/database/download/{e(_fn)}"
                 class="px-2 py-1 text-xs bg-gray-50 text-gray-600 border border-gray-200 rounded-lg">⬇</a>
              <button onclick="doRestoreFile('{e(_fn)}')"
                class="px-2 py-1 text-xs bg-green-50 text-green-700 border border-green-200 rounded-lg">♻️ بازیابی</button>
            </div>
          </div>"""
    if not _restore_rows:
        _restore_rows = '<div class="px-3 py-4 text-center text-sm text-gray-400">هنوز هیچ بکاپی ساخته نشده</div>'

    def _mod_checks(name, checked=True):
        """چک‌باکس‌های ماژول برای بکاپ سفارشی — فقط ماژول‌های واقعی، بدون
        چک‌باکس تکراری/جدا برای «حسابداری» (که قبلاً هیچ اثری هم روی بکاپ
        نداشت، فقط روی ریست — ماژول «accounting» خودش همون داده‌ها رو پوشش
        می‌ده)."""
        return "".join(
            f'<label class="flex items-center gap-2 text-sm cursor-pointer py-1.5 px-2 rounded-lg hover:bg-gray-50 transition">'
            f'<input type="checkbox" name="{name}" value="{k}"'
            f'{" checked" if checked else ""} class="w-4 h-4 rounded">'
            f'<span>{v["label"]}</span></label>'
            for k, v in MODULES.items()
        )

    def _chk(name, checked=True):
        acc_chk = (f'<label class="flex items-center gap-2 text-sm cursor-pointer py-1.5 px-2 rounded-lg hover:bg-red-50 transition col-span-2 border-t border-dashed border-red-100 mt-1">'
                   f'<input type="checkbox" name="{name}" value="__accounting__"'
                   f'{" checked" if checked else ""} class="w-4 h-4 rounded text-red-500">'
                   f'<span class="text-red-600 font-medium">💰 داده‌های حسابداری (هزینه‌ها + قیمت خرید)</span></label>')
        return _mod_checks(name, checked) + acc_chk

    backup_checks = _mod_checks("sections", checked=True)
    reset_checks  = _chk("reset_sections", checked=False)

    _js = """
    function toggle(id,chk){document.getElementById(id).classList.toggle('hidden',!chk.checked);}
    var _busy=false;
    function ovShow(t,sub){
      document.getElementById('bk-overlay').style.display='block';
      document.getElementById('ov-icon').textContent='\u23f3';
      document.getElementById('ov-title').textContent=t;
      document.getElementById('ov-msg').textContent=sub||'';
      document.getElementById('ov-close').style.display='none';
      var b=document.getElementById('ov-bar');
      b.style.transition='width .3s ease';b.style.width='3%';b.style.background='#6366f1';
    }
    // \u067e\u06cc\u0634\u0631\u0641\u062a \u0648\u0627\u0642\u0639\u06cc (\u0646\u0647 \u0627\u0646\u06cc\u0645\u06cc\u0634\u0646 \u0633\u0627\u062e\u062a\u06af\u06cc) \u2014 \u0627\u0632 polling \u0648\u0636\u0639\u06cc\u062a job \u067e\u0631 \u0645\u06cc\u200c\u0634\u0647
    function ovProgress(pct,msg){
      var b=document.getElementById('ov-bar');
      b.style.width=Math.max(3,Math.min(99,pct))+'%';
      if(msg) document.getElementById('ov-msg').textContent=msg;
    }
    async function pollJob(jobId){
      for(let i=0;i<300;i++){ // \u062d\u062f\u0627\u06a9\u062b\u0631 ~\u06f5 \u062f\u0642\u06cc\u0642\u0647 (300\u00d71s)
        await new Promise(r=>setTimeout(r,1000));
        var r=await fetch('/admin/database/job/'+jobId);
        var d=await r.json();
        if(d.status==='running'){ovProgress(d.progress||3,'\u062f\u0631 \u062d\u0627\u0644 \u0627\u0646\u062c\u0627\u0645... '+(d.progress||0)+'\u066a');continue;}
        return d;
      }
      throw new Error('\u0632\u0645\u0627\u0646\u200c\u0628\u0646\u062f\u06cc \u0639\u0645\u0644\u06cc\u0627\u062a \u062a\u0645\u0627\u0645 \u0634\u062f');
    }
    function ovResult(ok,t,msg){
      document.getElementById('ov-icon').textContent=ok?'\u2705':'\u274c';
      document.getElementById('ov-title').textContent=t;
      document.getElementById('ov-msg').textContent=msg;
      var b=document.getElementById('ov-bar');b.style.transition='width .3s';
      b.style.width='100%';b.style.background=ok?'#22c55e':'#ef4444';
      document.getElementById('ov-close').style.display='inline-block';
    }
    function getSelected(n){
      return Array.from(document.querySelectorAll('input[name="'+n+'"]:checked')).map(function(i){return i.value;});
    }
    async function runJob(type){
      if(_busy)return;_busy=true;
      try{
        if(type==='backup'){
          // الگوی استاندارد بکاپ‌گیری سنگین: شروع job در پس‌زمینه + poll پیشرفت واقعی
          // (نه یه درخواست مسدودکننده با نوار پیشرفت ساختگی — قبلاً همین باعث می‌شد
          // بکاپ‌های بزرگ حس «هنگ» بدن، حتی وقتی خودِ سرور داشت درست کار می‌کرد).
          ovShow('در حال ساخت بکاپ...','شروع...');
          var fd=new FormData();
          if(document.getElementById('b-toggle').checked) getSelected('sections').forEach(function(v){fd.append('sections',v);});
          else fd.append('full','1');
          var r=await fetch('/admin/database/backup/start',{method:'POST',body:fd});
          var start=await r.json();
          if(start.error) throw new Error(start.error);
          var jobResult=await pollJob(start.job_id);
          if(jobResult.status!=='done') throw new Error(jobResult.message||'خطا در ساخت بکاپ');
          ovProgress(100,'در حال دانلود...');
          window.location.href='/admin/database/backup/download/'+start.job_id;
          ovResult(true,'بکاپ آماده شد','فایل دانلود شد و یک نسخه هم روی سرور ذخیره شد');
        }else{
          ovShow('در حال ریست...','لطفاً صبر کنید');
          var fd2=new FormData();
          if(document.getElementById('r-toggle').checked) getSelected('reset_sections').forEach(function(v){fd2.append('reset_sections',v);});
          else fd2.append('full','1');
          var r2=await fetch('/admin/database/reset/sync',{method:'POST',body:fd2});
          var d2=await r2.json();
          if(d2.error) throw new Error(d2.error);
          ovResult(true,'ریست انجام شد',(d2.total_deleted||0)+' رکورد حذف شد');
        }
      }catch(err){ovResult(false,'عملیات ناموفق',err.message||'خطا');}
      finally{_busy=false;}
    }
    async function doRestoreFile(filename){
      if(!confirm('⚠️ این عملیات داده‌های فعلی را با این بکاپ جایگزین می‌کند. ادامه؟'))return;
      if(_busy)return;_busy=true;
      ovShow('در حال بازیابی...','شروع...');
      try{
        var fd=new FormData(); fd.append('filename',filename);
        var r=await fetch(RESTORE_URL,{method:'POST',body:fd});
        var start=await r.json();
        if(start.error) throw new Error(start.error);
        if(start.job_id){
          var job=await pollJob(start.job_id);
          if(job.status!=='done') throw new Error(job.message||'خطا در بازیابی');
          var d=job.result||{};
          if(d.errors&&d.errors.length) ovResult(true,'بازیابی با هشدار',(d.total||0)+' رکورد — '+d.errors.length+' خطای جدولی');
          else ovResult(true,'بازیابی موفق','بکاپ با موفقیت بازیابی شد');
        }else{
          ovResult(true,'بازیابی موفق','بکاپ با موفقیت بازیابی شد');
        }
      }catch(err){ovResult(false,'بازیابی ناموفق',err.message||'خطا');}
      finally{_busy=false;}
    }
    async function runRestore(){
      if(_busy)return;
      var file=document.getElementById('restore-file').files[0];
      if(!file){alert('فایل انتخاب نشده');return;}
      if(!file.name.endsWith('.stbak')){alert('فقط فایل بکاپ (.stbak) مجاز است');return;}
      _busy=true;ovShow('در حال بازیابی...','شروع...');
      try{
        var fd=new FormData();fd.append('backup_file',file);
        var r=await fetch(RESTORE_UPLOAD_URL,{method:'POST',body:fd});
        var start=await r.json();
        if(start.error) throw new Error(start.error);
        if(start.job_id){
          var job=await pollJob(start.job_id);
          if(job.status!=='done') throw new Error(job.message||'خطا در بازیابی');
          var d=job.result||{};
          if(d.errors&&d.errors.length) ovResult(true,'بازیابی با هشدار',(d.total||0)+' رکورد بازیابی شد — '+d.errors.length+' خطای جدولی');
          else ovResult(true,'بازیابی موفق',(d.total||0)+' رکورد بازیابی شد');
        }else{
          ovResult(true,'بازیابی موفق','بکاپ با موفقیت بازیابی شد');
        }
      }catch(err){ovResult(false,'بازیابی ناموفق',err.message||'خطا');}
      finally{_busy=false;}
    }
    """
    _restore_urls_js = f"""
    var RESTORE_URL = {"'/admin/database/pg-backup/restore'" if _is_pg else "'/admin/database/restore-auto'"};
    var RESTORE_UPLOAD_URL = {"'/admin/database/pg-backup/restore-upload'" if _is_pg else "'/admin/database/restore/start'"};
    """
    _js = _restore_urls_js + _js

    body = f"""
    <div class="flex items-center justify-between mb-6">
      <h1 class="text-2xl font-bold text-gray-800">💾 پشتیبان‌گیری و بازیابی</h1>
    </div>

    <!-- بکاپ -->
    <div class="card p-6 mb-4">
      <div class="flex items-center gap-3 mb-5">
        <span class="w-10 h-10 bg-indigo-100 text-indigo-700 rounded-xl flex items-center justify-center text-xl">📦</span>
        <h2 class="font-bold text-gray-800 text-lg">بکاپ</h2>
      </div>
      <label class="flex items-center gap-2 cursor-pointer mb-4 select-none p-3 bg-gray-50 rounded-xl">
        <input type="checkbox" id="b-toggle" onchange="toggle('b-secs',this)"
          class="w-4 h-4 rounded text-indigo-600">
        <div>
          <span class="text-sm font-medium text-gray-700">انتخاب سفارشی</span>
          <span class="text-xs text-gray-400 block">پیش‌فرض: بکاپ کامل از همه بخش‌ها</span>
        </div>
      </label>
      <div id="b-secs" class="hidden grid grid-cols-2 gap-0.5 bg-white border border-gray-100 rounded-xl p-3 mb-4 max-h-56 overflow-y-auto">
        {backup_checks}
      </div>
      <button onclick="runJob('backup')"
        class="w-full py-3 bg-indigo-600 hover:bg-indigo-700 text-white rounded-xl text-sm font-semibold transition">
        ⬇ دریافت فایل بکاپ
      </button>
    </div>

    <!-- بازیابی -->
    <div class="card p-6 mb-4">
      <div class="flex items-center gap-3 mb-5">
        <span class="w-10 h-10 bg-green-100 text-green-700 rounded-xl flex items-center justify-center text-xl">♻️</span>
        <h2 class="font-bold text-gray-800 text-lg">بازیابی</h2>
      </div>

      <details class="mb-4 group">
        <summary class="cursor-pointer select-none p-3 bg-gray-50 rounded-xl text-sm font-medium text-gray-700 flex items-center justify-between">
          <span>🕐 ۵ بکاپ آخر</span>
          <span class="text-xs text-gray-400 group-open:rotate-180 transition">▾</span>
        </summary>
        <div class="mt-2 border border-gray-100 rounded-xl overflow-hidden">
          {_restore_rows}
        </div>
      </details>

      <div class="border-t border-gray-100 my-4"><p class="text-xs text-gray-400 text-center mt-3">یا بارگذاری فایل بکاپ از روی سیستم</p></div>
      <div class="border-2 border-dashed border-gray-200 rounded-xl p-5 text-center mb-4">
        <input type="file" id="restore-file" accept=".stbak"
          class="text-sm text-gray-600 file:ml-2 file:py-2 file:px-4 file:rounded-lg file:border-0 file:text-sm file:bg-green-50 file:text-green-700">
      </div>
      <button onclick="runRestore()"
        class="w-full py-3 bg-green-600 hover:bg-green-700 text-white rounded-xl text-sm font-semibold transition">
        ♻️ بازیابی از فایل
      </button>
    </div>

    <!-- ریست -->
    <div class="card p-6 border-2 border-red-100">
      <div class="flex items-center gap-3 mb-5">
        <span class="w-10 h-10 bg-red-100 text-red-700 rounded-xl flex items-center justify-center text-xl">🗑</span>
        <div><h2 class="font-bold text-red-700 text-lg">ریست سیستم</h2>
             <p class="text-xs text-red-400">این عملیات برگشت‌ناپذیر است</p></div>
      </div>
      <label class="flex items-center gap-2 cursor-pointer mb-4 select-none p-3 bg-red-50 rounded-xl">
        <input type="checkbox" id="r-toggle" onchange="toggle('r-secs',this)"
          class="w-4 h-4 rounded text-red-500">
        <div>
          <span class="text-sm font-medium text-gray-700">انتخاب سفارشی</span>
          <span class="text-xs text-gray-400 block">پیش‌فرض: ریست کامل سیستم</span>
        </div>
      </label>
      <div id="r-secs" class="hidden grid grid-cols-2 gap-0.5 bg-white border border-red-100 rounded-xl p-3 mb-4 max-h-56 overflow-y-auto">
        {reset_checks}
      </div>
      <button onclick="if(confirm('⚠️ این عملیات برگشت‌ناپذیر است. ادامه می‌دهید؟')) runJob('reset')"
        class="w-full py-3 bg-red-600 hover:bg-red-700 text-white rounded-xl text-sm font-semibold transition">
        🗑 اجرای ریست
      </button>
    </div>

    <!-- Progress overlay -->
    <div id="bk-overlay" class="bk-overlay">
      <div id="ov-modal" class="bk-modal">
        <div class="bk-icon" id="ov-icon">⏳</div>
        <h3 class="bk-title" id="ov-title">در حال انجام...</h3>
        <p class="bk-msg" id="ov-msg">لطفاً صبر کنید</p>
        <div class="bk-track">
          <div id="ov-bar" class="bk-bar"></div>
        </div>
        <button id="ov-close" onclick="document.getElementById('bk-overlay').style.display='none'"
          class="bk-done-btn">
          بستن
        </button>
      </div>
    </div>

    <script>
    {_js}
    </script>

    <!-- 💾 تنظیمات پشتیبان‌گیری -->
    <div class="card p-5 mt-6 mb-8" id="cloudbk">
      <h2 class="font-bold text-gray-800 text-lg mb-1">💾 پشتیبان‌گیری</h2>
      <p class="text-xs text-gray-400 mb-4">مدیریت بکاپ خودکار و مقاصد ابری — هر بخش مستقل فعال/غیرفعال می‌شود.</p>

      <form method="post" action="/admin/database/cloud-save" class="space-y-3">

        <!-- ⏰ زمان‌بندی -->
        <div class="p-4 bg-gray-50 border border-gray-200 rounded-xl">
          <div class="flex items-center justify-between flex-wrap gap-2 mb-2">
            <div class="flex items-center gap-2">
              <span class="text-lg">⏰</span>
              <span class="font-semibold text-gray-700 text-sm">بکاپ خودکار</span>
            </div>
            <label class="flex items-center gap-1.5 text-xs">
              <input type="checkbox" name="enabled" {('checked' if int(_cb.get('enabled') or 0) else '')}
                class="w-4 h-4 rounded border-gray-300 text-indigo-600">
              <span class="text-gray-600">فعال</span>
            </label>
          </div>
          <div class="flex gap-3">
            <div class="flex-1">
              <label class="text-[10px] text-gray-400 block mb-1">ساعت اجرا (۰-۲۳)</label>
              <input type="number" name="hour" value="{int(_cb.get('hour') or 4)}" min="0" max="23"
                class="w-full border border-gray-200 rounded-lg px-3 py-2 text-sm text-center">
            </div>
          </div>
        </div>

        <!-- 💾 محلی -->
        <div class="p-4 bg-emerald-50/50 border border-emerald-200 rounded-xl">
          <div class="flex items-center justify-between flex-wrap gap-2 mb-2">
            <div class="flex items-center gap-2">
              <span class="text-lg">💾</span>
              <span class="font-semibold text-gray-700 text-sm">بکاپ محلی</span>
            </div>
            <label class="flex items-center gap-1.5 text-xs">
              <input type="checkbox" name="local_enabled" checked disabled
                class="w-4 h-4 rounded border-gray-300 text-emerald-600">
              <span class="text-gray-500">همیشه فعال</span>
            </label>
          </div>
          <div class="flex gap-3">
            <div class="flex-1">
              <label class="text-[10px] text-gray-400 block mb-1">نگهداری (تعداد بکاپ)</label>
              <input type="number" name="retention" value="{int(_cb.get('retention') or 3)}" min="1" max="30"
                class="w-full border border-gray-200 rounded-lg px-3 py-2 text-sm text-center">
            </div>
          </div>
          <p class="text-[10px] text-gray-400 mt-2">فایل‌ها در مسیر /opt/stockland/data/backups ذخیره می‌شوند. قدیمی‌ها خودکار حذف می‌شوند.</p>
        </div>

        <!-- 📢 کانال تلگرام -->
        <div class="p-4 bg-blue-50/50 border border-blue-200 rounded-xl">
          <div class="flex items-center justify-between flex-wrap gap-2 mb-2">
            <div class="flex items-center gap-2">
              <span class="text-lg">📢</span>
              <span class="font-semibold text-gray-700 text-sm">کانال تلگرام</span>
            </div>
            <label class="flex items-center gap-1.5 text-xs">
              <input type="checkbox" name="tg_enabled" {('checked' if int(_cb.get('tg_enabled') or 0) else '')}
                class="w-4 h-4 rounded border-gray-300 text-blue-600">
              <span class="text-gray-600">فعال</span>
            </label>
          </div>
          <div>
            <label class="text-[10px] text-gray-400 block mb-1">Chat ID کانال</label>
            <input type="text" name="tg_channel" value="{e(_cb.get('tg_channel',''))}"
              placeholder="@channel یا -1001234567890"
              class="w-full border border-gray-200 rounded-lg px-3 py-2 text-sm" dir="ltr">
          </div>
          <p class="text-[10px] text-gray-400 mt-2">ربات باید ادمین کانال باشد. بکاپ‌ها در تلگرام بدون محدودیت نگه‌داری می‌شوند.</p>
        </div>

        <!-- 🗂 Google Drive (OAuth) -->
        <div class="p-4 bg-amber-50/50 border border-amber-200 rounded-xl">
          <div class="flex items-center justify-between flex-wrap gap-2 mb-2">
            <div class="flex items-center gap-2">
              <span class="text-lg">🗂</span>
              <span class="font-semibold text-gray-700 text-sm">Google Drive</span>
            </div>
            <label class="flex items-center gap-1.5 text-xs">
              <input type="checkbox" name="gdrive_enabled" {('checked' if int(_cb.get('gdrive_enabled') or 0) else '')}
                class="w-4 h-4 rounded border-gray-300 text-amber-600">
              <span class="text-gray-600">فعال</span>
            </label>
          </div>
          <div class="text-[11px] mb-2">
            {('<span class="text-green-600">✅ متصل به Google Drive</span>' if _gdrive_connected else '<span class="text-amber-600">⚠️ هنوز متصل نشده — از دکمه زیر اتصال بزنید</span>')}
          </div>
          {('' if _gdrive_connected else '<button type="button" onclick="startGdriveConnect()" class="mb-2 px-4 py-2 bg-blue-600 hover:bg-blue-700 text-white rounded-lg text-xs font-semibold">🔗 اتصال به Google Drive</button>')}
          <div id="gdrive_connect_box" class="hidden mb-2 p-3 bg-white border border-blue-200 rounded-lg text-sm"></div>
          <p class="text-[10px] text-gray-400">فقط ۳۰ بکاپ آخر در Drive نگهداری. اتصال یک‌بار انجام می‌شود و برای همیشه کار می‌کند.</p>
        </div>

        <!-- دکمه‌ها -->
        <div class="flex gap-2 pt-1">
          <button type="submit" class="flex-1 py-2.5 bg-indigo-600 hover:bg-indigo-700 text-white rounded-xl text-sm font-semibold">💾 ذخیره تنظیمات</button>
        </div>
      </form>
      <form method="post" action="/admin/database/cloud-run" class="mt-2">
        <button class="w-full py-2.5 bg-green-50 hover:bg-green-100 text-green-700 border border-green-200 rounded-xl text-sm font-semibold">▶ بکاپ فوری (همه مقاصد فعال)</button>
      </form>

      <script>
      async function startGdriveConnect(){{
        const box=document.getElementById('gdrive_connect_box');
        box.classList.remove('hidden');
        box.innerHTML='<span class="text-gray-500">⏳ در حال دریافت کد...</span>';
        try{{
          const r=await fetch('/admin/database/gdrive/start',{{method:'POST'}});
          const d=await r.json();
          if(!d.ok){{box.innerHTML='<span class="text-red-600 bidi-plain">❌ '+d.error+'</span>';return;}}
          box.innerHTML=`
            <div class="text-center space-y-3">
              <p class="text-sm text-gray-700">لینک زیر را باز کنید و کد را وارد کنید:</p>
              <a href="${{d.url}}" target="_blank" class="block px-4 py-2 bg-blue-600 text-white rounded-lg text-sm">🔗 باز کردن لینک تأیید</a>
              <div class="text-2xl font-bold tracking-widest text-indigo-700 bg-indigo-50 rounded-lg py-3 no-fa" dir="ltr">${{d.user_code}}</div>
              <p class="text-xs text-gray-400">پس از تأیید در گوگل، دکمه زیر را بزنید</p>
              <button onclick="pollGdrive('${{d.device_code}}')" class="px-6 py-2 bg-green-600 text-white rounded-lg text-sm">✅ تأیید کردم</button>
            </div>`;
        }}catch(e){{box.innerHTML='<span class="text-red-600 bidi-plain">خطا: '+e.message+'</span>';}}
      }}
      async function pollGdrive(dc){{
        const box=document.getElementById('gdrive_connect_box');
        box.innerHTML='<span class="text-gray-500">⏳ بررسی تأیید...</span>';
        for(let i=0;i<24;i++){{
          await new Promise(r=>setTimeout(r,5000));
          try{{
            const r=await fetch('/admin/database/gdrive/poll',{{method:'POST',headers:{{'Content-Type':'application/json'}},body:JSON.stringify({{device_code:dc}})}});
            const d=await r.json();
            if(d.ok){{box.innerHTML='<span class="text-green-600 font-bold">✅ اتصال موفق! صفحه را رفرش کنید.</span>';return;}}
            if(!d.pending){{box.innerHTML='<span class="text-red-600 bidi-plain">❌ '+(d.error||'خطا')+'</span>';return;}}
          }}catch(e){{}}
        }}
        box.innerHTML='<span class="text-red-600">⏱ زمان منقضی شد. دوباره تلاش کنید.</span>';
      }}
      </script>
    </div>
    </div>"""

    return _layout("پشتیبان‌گیری", body, adm, flash=flash)


@router.post("/database/pg-backup/restore")
async def pg_backup_restore(request: Request):
    """بازیابی از بکاپ Postgres محلی (فایل .stbak ساخته‌شده با pg_dump/pg_backup.py)."""
    from fastapi.responses import JSONResponse
    adm = _get_admin(request)
    guard = _require(adm, "restore")
    if guard: return JSONResponse({"error": "unauthorized"})
    import db_conn
    if not db_conn.is_postgres():
        return JSONResponse({"error": "این مسیر فقط برای حالت Postgres است"})
    form = await request.form()
    filename = str(form.get("filename", "")).strip()
    if not filename or ".." in filename or "/" in filename:
        return JSONResponse({"error": "فایل نامعتبر"})

    def _run():
        import pg_backup
        import os as _os
        path = _os.path.join(pg_backup.BACKUP_DIR, filename)
        if not _os.path.exists(path):
            return {"ok": False, "error": "فایل یافت نشد"}
        return pg_backup.restore_backup(path)

    try:
        rep = await run_in_threadpool(_run)
        if not rep.get("ok"):
            return JSONResponse({"error": rep.get("error", "خطای نامشخص")})
        _log(request, "بازیابی بکاپ Postgres", "دیتابیس", filename, admin_info=adm)
        return JSONResponse({"ok": True})
    except Exception as ex:
        return JSONResponse({"error": str(ex)[:200]})


@router.post("/database/pg-backup/restore-upload")
async def pg_backup_restore_upload(request: Request, backup_file: UploadFile = None):
    """بازیابی Postgres از یک فایل بکاپ آپلودشده از روی سیستم ادمین (نه از
    لیست بکاپ‌های محلی سرور) — معادل Postgres همون /database/restore/start."""
    from fastapi.responses import JSONResponse
    adm = _get_admin(request)
    guard = _require(adm, "restore")
    if guard: return JSONResponse({"error": "unauthorized"})
    import db_conn
    if not db_conn.is_postgres():
        return JSONResponse({"error": "این مسیر فقط برای حالت Postgres است"})
    if not backup_file or not (backup_file.filename or "").endswith(".stbak"):
        return JSONResponse({"error": "فقط فایل بکاپ (.stbak) مجاز است"})
    raw = await backup_file.read()

    def _run():
        import pg_backup
        import tempfile as _tf
        tmp_path = None
        try:
            with _tf.NamedTemporaryFile(suffix=".stbak", delete=False) as tf:
                tf.write(raw)
                tmp_path = tf.name
            return pg_backup.restore_backup(tmp_path)
        finally:
            if tmp_path:
                try: os.remove(tmp_path)
                except Exception: pass

    try:
        rep = await run_in_threadpool(_run)
        if not rep.get("ok"):
            return JSONResponse({"error": rep.get("error", "خطای نامشخص")})
        _log(request, "بازیابی از فایل آپلودی", "دیتابیس", backup_file.filename, admin_info=adm)
        return JSONResponse({"ok": True})
    except Exception as ex:
        return JSONResponse({"error": str(ex)[:200]})


@router.post("/database/backup/full-sync")
async def backup_full_sync(request: Request):
    adm = _get_admin(request)
    guard = _require(adm, "backup")
    if guard: return guard
    from fastapi.responses import Response as FResponse, PlainTextResponse
    _pgg = _stbak_pg_guard()
    if _pgg: return PlainTextResponse(_pgg["error"], status_code=409)
    form = await request.form()
    is_full = form.get("full") == "1"
    sections = None if is_full else (form.getlist("sections") or None)
    def _build_and_save():
        from stbak_engine import create_stbak, stbak_filename, _rotate_local
        raw = create_stbak(_DB_PATH(), modules=sections)
        fname = stbak_filename("full" if is_full else "custom", _BACKUP_DIR)
        # ذخیرهٔ یک نسخه محلی هم — قبلاً این بکاپ فقط مستقیم دانلود می‌شد و هیچ ردی از
        # خودش توی _BACKUP_DIR نمی‌ذاشت؛ یعنی نه توی لیست «بکاپ‌های خودکار» پنل دیده
        # می‌شد نه «بازگردانی اضطراری» (که فقط _BACKUP_DIR رو می‌گرده) پیداش می‌کرد.
        try:
            os.makedirs(_BACKUP_DIR, exist_ok=True)
            with open(os.path.join(_BACKUP_DIR, fname), "wb") as _f:
                _f.write(raw)
            _rotate_local(_BACKUP_DIR, _MAX_BACKUPS)
        except Exception:
            pass  # اگه ذخیرهٔ محلی شکست خورد، حداقل دانلود مستقیم برای ادمین کار کنه
        return raw, fname

    try:
        # ساخت بکاپ (زیپ‌کردن مدیا + dry-run اعتبارسنجی) کار سنگین synchronous است؛
        # روی ترد جدا اجرا می‌شه تا در طول این مدت کل اپ (ربات + بقیهٔ ادمین‌ها) قفل نشه.
        raw, fname = await run_in_threadpool(_build_and_save)
        _log(request, "بکاپ کامل دستی", "دیتابیس", fname, admin_info=adm)
        return FResponse(content=raw, media_type="application/octet-stream",
                          headers={"Content-Disposition": f'attachment; filename="{fname}"'})
    except Exception as ex:
        return PlainTextResponse(f"خطا در بکاپ: {str(ex)[:150]}", status_code=500)


@router.get("/database/download/{filename}")
async def database_download_auto(request: Request, filename: str):
    """دانلود یک بکاپ خودکار (فایل .stbak)."""
    from fastapi.responses import FileResponse, PlainTextResponse
    adm = _get_admin(request)
    guard = _require(adm, "backup")
    if guard: return guard
    import os
    if ".." in filename or "/" in filename:
        return PlainTextResponse("نام فایل نامعتبر", status_code=400)
    # ⚠️ رفع‌شده: pg_backup فایل‌هاش رو در pg_backup.BACKUP_DIR می‌سازه (مسیر
    # جدا از _BACKUP_DIR که مخصوص stbak_engine/SQLite است) — بدون این چک،
    # لینک دانلود بکاپ‌های Postgres همیشه ۴۰۴ می‌داد.
    candidates = [_BACKUP_DIR]
    try:
        import pg_backup
        candidates.append(pg_backup.BACKUP_DIR)
    except Exception:
        pass
    path = None
    for d in candidates:
        p = os.path.join(d, filename)
        if os.path.exists(p):
            path = p
            break
    if not path:
        return PlainTextResponse("فایل یافت نشد", status_code=404)
    _log(request, "دانلود بکاپ خودکار", "دیتابیس", filename, admin_info=adm)
    return FileResponse(path, filename=filename, media_type="application/octet-stream")


@router.post("/database/restore-auto")
async def restore_auto(request: Request):
    """job-based — پاسخ فوری، پیشرفت واقعی از /database/job/{id} (همون الگوی backup/start)."""
    from fastapi.responses import JSONResponse
    adm = _get_admin(request)
    guard = _require(adm, "restore")
    if guard: return JSONResponse({"error": "unauthorized"})
    _pgg = _stbak_pg_guard()
    if _pgg: return JSONResponse(_pgg)
    form = await request.form()
    filename = str(form.get("filename","")).strip()
    if not filename or ".." in filename or "/" in filename:
        return JSONResponse({"error": "فایل نامعتبر"})
    path = os.path.join(_BACKUP_DIR, filename)
    if not os.path.exists(path):
        return JSONResponse({"error": "فایل یافت نشد"})
    with open(path, "rb") as f:
        raw = f.read()
    from stbak_engine import restore_stbak
    job_id = _job_start(restore_stbak, raw, _DB_PATH(), safety_backup_dir=_BACKUP_DIR)
    _log(request, "شروع بازیابی بکاپ خودکار", "دیتابیس", f"job:{job_id} — {filename}", admin_info=adm)
    return JSONResponse({"job_id": job_id})


@router.post("/database/recover-latest")
async def recover_latest(request: Request):
    """🆘 بازگردانی اضطراری یک‌کلیکی — بخش Recovery: بدون نیاز به انتخاب دستی فایل،
    بین بکاپ‌های محلی (جدیدترین اول) هرکدوم که واقعاً سالم بود (validate_stbak پاس بشه)
    رو پیدا می‌کنه و همون رو بازیابی می‌کنه — اگه جدیدترین فایل خراب بود، خودکار سراغ
    بکاپ قبلی می‌ره، نه اینکه کل Recovery متوقف بشه."""
    from fastapi.responses import JSONResponse
    adm = _get_admin(request)
    guard = _require(adm, "recovery")
    if guard: return JSONResponse({"error": "unauthorized"})
    _pgg = _stbak_pg_guard()
    if _pgg: return JSONResponse(_pgg)
    from stbak_engine import list_local_backups, validate_stbak, restore_stbak, StbakError

    candidates = list_local_backups(_BACKUP_DIR)
    if not candidates:
        return JSONResponse({"error": "هیچ بکاپ محلی‌ای برای بازگردانی یافت نشد"})

    def _find_and_restore():
        tried = []
        for cand in candidates:
            try:
                with open(cand["path"], "rb") as f:
                    raw = f.read()
                validate_stbak(raw)  # فقط سالم‌بودن رو چک می‌کنه، هنوز چیزی رو دست نمی‌زنه
            except StbakError as ex:
                tried.append(f"{cand['name']}: {ex}")
                continue
            except Exception as ex:
                tried.append(f"{cand['name']}: {ex}")
                continue
            try:
                res = restore_stbak(raw, _DB_PATH(), safety_backup_dir=_BACKUP_DIR)
                return cand["name"], res, tried
            except Exception as ex:
                tried.append(f"{cand['name']}: خطای بازیابی — {ex}")
                continue
        return None, None, tried

    # اعتبارسنجی + بازیابی هر دو سنگین‌ان (خواندن/باز کردن zip + درج کامل داده) —
    # روی ترد جدا تا کل اپ در طول این مدت قفل نشه.
    restored_name, res, tried = await run_in_threadpool(_find_and_restore)
    if restored_name is None:
        return JSONResponse({"error": "هیچ‌کدام از بکاپ‌های محلی سالم نبودند", "details": tried})

    _log(request, "بازگردانی اضطراری (Recovery)", "دیتابیس",
         f"{restored_name} — {len(res['errors'])} خطای جدولی", admin_info=adm)
    return JSONResponse({
        "ok": True, "restored_from": restored_name,
        "total": res["total"], "warnings": res["errors"],
        "skipped": tried,
    })


@router.post("/database/restore/sync")
async def restore_sync(request: Request, backup_file: UploadFile = None):
    from fastapi.responses import JSONResponse
    adm = _get_admin(request)
    guard = _require(adm, "restore")
    if guard: return JSONResponse({"error": "unauthorized"})
    _pgg = _stbak_pg_guard()
    if _pgg: return JSONResponse(_pgg)
    if not backup_file or not (backup_file.filename or "").endswith(".stbak"):
        return JSONResponse({"error": "فقط فایل .stbak مجاز است"})
    raw = await backup_file.read()
    from stbak_engine import restore_stbak, StbakError
    try:
        res = await run_in_threadpool(restore_stbak, raw, _DB_PATH(), safety_backup_dir=_BACKUP_DIR)
        if res["errors"]:
            _log(request, "بازیابی از فایل (با هشدار)", "دیتابیس",
                 f"{backup_file.filename} — {len(res['errors'])} خطا", admin_info=adm)
            return JSONResponse({"ok": True, "warnings": res["errors"], "total": res["total"]})
        _log(request, "بازیابی از فایل", "دیتابیس", backup_file.filename, admin_info=adm)
        return JSONResponse({"ok": True, "total": res["total"]})
    except StbakError as ex:
        return JSONResponse({"error": str(ex)})
    except Exception as ex:
        return JSONResponse({"error": str(ex)[:150]})


def _generic_factory_reset(modules: list = None) -> dict:
    """معادل دیالوگ‌آگاه stbak_engine.factory_reset — همون منطق (DELETE FROM
    روی جدول‌های ماژول‌های انتخاب‌شده) ولی با db_conn.get_connection() که خودش
    بین SQLite/Postgres مسیریابی می‌کنه، به‌جای sqlite3.connect() خام. برخلاف
    نسخهٔ SQLite، دست به PRAGMA/sqlite_sequence نمی‌زنه (خاص SQLite و برای
    درستی ریست ضروری نیست)."""
    from stbak_engine import MODULES, resolve_reset
    selected = list(MODULES.keys()) if modules is None else resolve_reset(modules)
    tables = []
    for mod in selected:
        tables.extend(MODULES.get(mod, {}).get("tables", []))
    tables = list(dict.fromkeys(tables))

    import db_conn
    conn = db_conn.get_connection(_DB_PATH())
    cleared, errors = {}, []
    try:
        for t in tables:
            try:
                cnt = conn.execute(f'SELECT COUNT(*) FROM "{t}";').fetchone()[0]
                conn.execute(f'DELETE FROM "{t}";')
                conn.commit()
                cleared[t] = cnt
            except Exception as ex:
                try: conn.rollback()
                except Exception: pass
                errors.append(f"{t}: {ex}")
    finally:
        conn.close()
    return {"cleared": cleared, "errors": errors, "total_deleted": sum(cleared.values())}


@router.post("/database/reset/sync")
async def reset_sync(request: Request):
    from fastapi.responses import JSONResponse
    adm = _get_admin(request)
    guard = _require(adm, "database")
    if guard: return JSONResponse({"error": "unauthorized"})
    form = await request.form()
    is_full = form.get("full") == "1"
    all_secs = form.getlist("reset_sections") or []

    # جداسازی حسابداری از بقیه
    reset_accounting = is_full or "__accounting__" in all_secs
    secs = None if is_full else ([s for s in all_secs if s != "__accounting__"] or None)

    import db_conn
    is_pg = db_conn.is_postgres()

    def _do_reset():
        total_deleted = 0
        if is_pg:
            result = _generic_factory_reset(modules=secs)
        else:
            from stbak_engine import factory_reset
            result = factory_reset(_DB_PATH(), modules=secs)
        total_deleted += result["total_deleted"]
        # ریست حسابداری
        if reset_accounting:
            conn = _db()
            try:
                conn.execute("DELETE FROM expenses;")
                try: conn.execute("DELETE FROM feed_batches;")
                except Exception: pass
                conn.commit()
                total_deleted += 1
            finally:
                conn.close()
        return total_deleted

    try:
        total_deleted = await run_in_threadpool(_do_reset)
        _log(request, "ریست", "دیتابیس", f"{total_deleted} رکورد")
        return JSONResponse({"ok": True, "total_deleted": total_deleted})
    except Exception as ex:
        return JSONResponse({"error": str(ex)[:100]})


@router.get("/database/job/{job_id}")
async def job_status(request: Request, job_id: str):
    from fastapi.responses import JSONResponse
    return JSONResponse(_job_read(job_id))


@router.post("/database/backup/start")
async def backup_start(request: Request):
    from fastapi.responses import JSONResponse
    adm = _get_admin(request)
    if not adm: return JSONResponse({"error": "unauthorized"})
    if not _has(adm, "backup"): return JSONResponse({"error": "unauthorized"})
    form = await request.form()
    is_full = form.get("full") == "1"
    all_secs = form.getlist("sections") or []

    import db_conn
    if db_conn.is_postgres():
        # «انتخاب سفارشی» زیر Postgres: کلیدهای ماژول به لیست جدول‌های واقعی‌شون
        # (stbak_engine.MODULES) ترجمه می‌شن و مستقیم به pg_dump -t پاس داده
        # می‌شن — معادل کامل همون قابلیتی که قبلاً فقط برای SQLite بود.
        tables = None
        if not is_full and all_secs:
            from stbak_engine import MODULES
            tset = []
            for s in all_secs:
                tset.extend(MODULES.get(s, {}).get("tables", []))
            tables = list(dict.fromkeys(tset)) or None

        def _pg_job(progress_cb=None):
            import pg_backup
            if progress_cb: progress_cb(40)
            fpath = pg_backup.create_backup(tables=tables)
            if progress_cb: progress_cb(100)
            return {"file": fpath}

        job_id = _job_start(_pg_job)
    else:
        sections = None if is_full else (all_secs or None)
        db = _DB_PATH()
        from stbak_engine import create_stbak
        job_id = _job_start(create_stbak, db, modules=sections)

    _log(request, "شروع بکاپ", "دیتابیس", f"job:{job_id} mode:{'full' if is_full else 'custom'}")
    return JSONResponse({"job_id": job_id})


@router.get("/database/backup/download/{job_id}")
async def backup_download_job(request: Request, job_id: str):
    from fastapi.responses import Response as FResponse
    adm = _get_admin(request)
    guard = _require(adm, "backup")
    if guard: return guard
    job = _job_read(job_id)
    if job.get("status") != "done":
        return _redir("/admin/database?flash=بکاپ+هنوز+آماده+نشده")
    fpath = job.get("result", {}).get("file")
    if not fpath or not os.path.exists(fpath):
        return _redir("/admin/database?flash=فایل+بکاپ+یافت+نشد")
    raw = open(fpath, "rb").read()
    import db_conn
    if db_conn.is_postgres():
        # pg_backup.create_backup() از قبل فایل رو مستقیم توی BACKUP_DIR خودش
        # ساخته و چرخش (rotation) رو هم انجام داده — نیازی به ذخیرهٔ دوبارهٔ
        # دستی با نام‌گذاری دیگه نیست (که قبلاً باعث می‌شد یه نسخهٔ اضافه توی
        # مسیر SQLite ذخیره بشه، جایی که اصلاً خونده نمی‌شه).
        fname = os.path.basename(fpath)
    else:
        from stbak_engine import stbak_filename, _rotate_local
        fname = stbak_filename("full", _BACKUP_DIR)
        # همون رفع مشترک با backup_full_sync — یه نسخه هم روی سرور ذخیره بشه تا توی
        # لیست پنل دیده بشه و بشه از بخش «بازیابی» انتخابش کرد.
        try:
            os.makedirs(_BACKUP_DIR, exist_ok=True)
            with open(os.path.join(_BACKUP_DIR, fname), "wb") as _f:
                _f.write(raw)
            _rotate_local(_BACKUP_DIR, _MAX_BACKUPS)
        except Exception:
            pass
    return FResponse(content=raw, media_type="application/octet-stream",
                     headers={"Content-Disposition": f'attachment; filename="{fname}"'})


@router.post("/database/restore/start")
async def restore_start(request: Request, backup_file: UploadFile = None):
    """نسخهٔ job-based بازیابی از فایل — برخلاف restore/sync، پاسخ فوری می‌ده و
    پیشرفت واقعی از /database/job/{id} قابل poll کردنه (همون الگوی backup/start)."""
    from fastapi.responses import JSONResponse
    adm = _get_admin(request)
    if not adm: return JSONResponse({"error": "unauthorized"})
    if not _has(adm, "restore"): return JSONResponse({"error": "unauthorized"})
    _pgg = _stbak_pg_guard()
    if _pgg: return JSONResponse(_pgg)
    if not backup_file or not (backup_file.filename or "").endswith(".stbak"):
        return JSONResponse({"error": "فقط فایل .stbak مجاز است"})
    raw = await backup_file.read()
    from stbak_engine import restore_stbak
    job_id = _job_start(restore_stbak, raw, _DB_PATH(), safety_backup_dir=_BACKUP_DIR)
    _log(request, "شروع بازیابی از فایل", "دیتابیس", f"job:{job_id} — {backup_file.filename}", admin_info=adm)
    return JSONResponse({"job_id": job_id})


@router.post("/database/reset/start")
async def reset_start(request: Request):
    from fastapi.responses import JSONResponse
    adm = _get_admin(request)
    if not adm: return JSONResponse({"error": "unauthorized"})
    _pgg = _stbak_pg_guard()
    if _pgg: return JSONResponse(_pgg)
    form = await request.form()
    is_full = form.get("full") == "1"
    sections = None if is_full else form.getlist("reset_sections") or None
    db = _DB_PATH()
    from stbak_engine import factory_reset
    job_id = _job_start(factory_reset, db, modules=sections)
    _log(request, "شروع ریست", "دیتابیس",
         f"job:{job_id} mode:{'full' if is_full else 'custom'}")
    return JSONResponse({"job_id": job_id})


@router.get("/admins", response_class=HTMLResponse)
async def admins_list(request: Request, flash: str = ""):
    adm = _get_admin(request)
    guard = _require(adm, "admins")
    if guard: return guard

    ensure_admins_table()
    conn = _db()
    try:
        admins = conn.execute("SELECT * FROM admins ORDER BY id DESC;").fetchall()
    finally:
        conn.close()

    rows = ""
    for a in admins:
        perms_list = json.loads(a["permissions"] or "[]")
        badges = " ".join(f'<span class="px-2 py-0.5 text-xs bg-blue-100 text-blue-700 rounded-full">{ALL_PERMISSIONS.get(p,p)}</span>' for p in perms_list) or '<span class="text-xs text-gray-400">بدون اختیار</span>'
        status_b = '<span class="px-2 py-0.5 text-xs bg-green-100 text-green-700 rounded-full">فعال</span>' if a["is_active"] else '<span class="px-2 py-0.5 text-xs bg-red-100 text-red-700 rounded-full">غیرفعال</span>'
        rows += f"""<tr class="border-b hover:bg-gray-50">
          <td class="px-4 py-3 text-sm font-medium text-gray-800">{e(a["name"])}</td>
          <td class="px-4 py-3 text-xs text-gray-400">{e(a["telegram_id"] or "—")}</td>
          <td class="px-4 py-3 text-xs text-gray-500">{e(a["web_username"] or "—")}</td>
          <td class="px-4 py-3"><div class="flex flex-wrap gap-1">{badges}</div></td>
          <td class="px-4 py-3">{status_b}</td>
          <td class="px-4 py-3">
            <div class="flex gap-1">
              <a href="/admin/admins/{a['id']}/edit" class="btn-sm bg-indigo-50 text-indigo-700 border border-indigo-200 rounded px-2 py-1 text-xs">ویرایش</a>
              <form method="post" action="/admin/admins/{a['id']}/toggle" class="inline">
                <button class="btn-sm {"bg-red-50 text-red-600 border border-red-200" if a["is_active"] else "bg-green-50 text-green-600 border border-green-200"} rounded px-2 py-1 text-xs">{"غیرفعال" if a["is_active"] else "فعال"}</button>
              </form>
              <form method="post" action="/admin/admins/{a['id']}/delete" class="inline" onsubmit="return confirm('حذف شود؟')">
                <button class="btn-sm bg-red-50 text-red-600 border border-red-200 rounded px-2 py-1 text-xs">حذف</button>
              </form>
            </div>
          </td>
        </tr>"""

    perm_checks = '<div class="flex flex-wrap gap-3 p-4 bg-gray-50 rounded-lg">' + "".join(
        f'<label class="flex items-center gap-2 text-sm cursor-pointer"><input type="checkbox" name="perm_{k}" value="1" class="rounded">{v}</label>'
        for k, v in ALL_PERMISSIONS.items()
    ) + '</div>'

    body = f"""
    <div class="flex items-center justify-between mb-6">
      <h1 class="text-2xl font-bold text-gray-800">👥 مدیریت ادمین‌ها</h1>
    </div>
    <div class="card p-6 mb-6">
      <h2 class="font-bold text-gray-700 mb-4">➕ افزودن ادمین جدید</h2>
      <form method="post" action="/admin/admins/add" class="space-y-4">
        <div class="grid grid-cols-2 gap-4">
          <div><label class="text-sm font-medium text-gray-700 block mb-1">نام نمایشی *</label>{_input("name","مثلاً: پشتیبانی",required=True)}</div>
          <div><label class="text-sm font-medium text-gray-700 block mb-1">آیدی تلگرام</label>{_input("telegram_id","مثلاً: 123456789",type_="number")}</div>
          <div><label class="text-sm font-medium text-gray-700 block mb-1">یوزرنیم *</label>{_input("web_username","مثلاً: support1",required=True)}</div>
          <div><label class="text-sm font-medium text-gray-700 block mb-1">رمز *</label>{_input("web_password","رمز قوی",type_="password",required=True)}</div>
        </div>
        <div><label class="text-sm font-medium text-gray-700 block mb-1">اختیارات دسترسی</label>{perm_checks}</div>
        {_btn("افزودن ادمین","",color="green")}
      </form>
    </div>
    <div class="card overflow-hidden">
      <div class="px-5 py-3 border-b bg-gray-50 flex items-center justify-between">
        <span class="font-medium text-gray-700">ادمین‌های فعلی ({len(admins)})</span>
      </div>
      <div class="overflow-x-auto">
        <table class="w-full text-right min-w-max">
          <thead><tr class="text-xs text-gray-500 border-b bg-gray-50">
            <th class="px-4 py-3">نام</th><th class="px-4 py-3">تلگرام</th>
            <th class="px-4 py-3">یوزرنیم</th><th class="px-4 py-3">اختیارات</th>
            <th class="px-4 py-3">وضعیت</th><th class="px-4 py-3">عملیات</th>
          </tr></thead>
          <tbody>{rows or "<tr><td colspan='6' class='text-center py-8 text-gray-400 text-sm'>ادمینی اضافه نشده</td></tr>"}</tbody>
        </table>
      </div>
    </div>"""
    return _layout("ادمین‌ها", body, adm, flash=flash)


@router.get("/account", response_class=HTMLResponse)
async def account_page(request: Request, flash: str = ""):
    adm = _get_admin(request)
    if not adm or not adm[1]:
        return _redir("/admin/")
    current_username = _env("ADMIN_WEB_USERNAME", "admin")
    body = f"""
    <div class="max-w-560">
      <div class="page-header"><h1>تنظیمات حساب</h1><p>اطلاعات امنیتی مدیر ارشد</p></div>
      <div class="card card-p form-card">
        <h2 class="form-card-title">تغییر نام کاربری</h2>
        <form method="post" action="/admin/account/username" class="form-section">
          <div><label>نام کاربری فعلی</label>
            <input type="text" value="{e(current_username)}" disabled class="disabled-look">
          </div>
          <div><label>نام کاربری جدید</label>{_input("new_username","فقط a-z, 0-9, _ (حداقل ۳ کاراکتر)",required=True)}</div>
          {_btn("ذخیره نام کاربری","",color="indigo")}
        </form>
      </div>
      <div class="card card-p form-card">
        <h2 class="form-card-title">تغییر رمز پنل</h2>
        <form method="post" action="/admin/admins/super/password" class="form-section">
          <div><label>رمز جدید</label>{_input("new_password","رمز قوی",type_="password",required=True)}</div>
          <div><label>تکرار رمز</label>{_input("confirm_password","تکرار رمز",type_="password",required=True)}</div>
          {_btn("ذخیره رمز","",color="green")}
        </form>
      </div>
      <div class="card card-p">
        <h2 class="form-card-title">تغییر آیدی تلگرام</h2>
        <form method="post" action="/admin/admins/super/telegram_id" class="form-section">
          <div><label>آیدی عددی تلگرام</label>{_input("new_telegram_id","مثلاً: 638469407",type_="number",required=True)}</div>
          {_btn("ذخیره آیدی","",color="green")}
        </form>
      </div>
    </div>"""
    return _layout("تنظیمات حساب", body, adm, flash=flash)


@router.post("/account/username")
async def account_change_username(request: Request, new_username: str = Form("")):
    adm = _get_admin(request)
    if not adm or not adm[1]:
        return _redir("/admin/login")
    new_username = new_username.strip().lower()
    import re as _re
    if not new_username or not _re.match(r'^[a-z0-9_]{3,32}$', new_username):
        return _redir("/admin/account?flash=نام+کاربری+نامعتبر+است+(فقط+حروف+انگلیسی+کوچک+و+عدد)")
    env_path = os.path.join(os.path.dirname(__file__), ".env")
    try:
        lines = open(env_path, encoding="utf-8").readlines() if os.path.exists(env_path) else []
        found = False
        new_lines = []
        for line in lines:
            if line.startswith("ADMIN_WEB_USERNAME="):
                new_lines.append(f"ADMIN_WEB_USERNAME={new_username}\n")
                found = True
            else:
                new_lines.append(line)
        if not found:
            new_lines.append(f"ADMIN_WEB_USERNAME={new_username}\n")
        open(env_path, "w", encoding="utf-8").writelines(new_lines)
        os.environ["ADMIN_WEB_USERNAME"] = new_username
    except Exception as ex:
        return _redir(f"/admin/account?flash=خطا:+{str(ex)[:40]}")
    return _redir("/admin/account?flash=نام+کاربری+تغییر+کرد")


@router.post("/admins/super/password")
async def super_change_password(request: Request, new_password: str = Form(""), confirm_password: str = Form("")):
    adm = _get_admin(request)
    if not adm or not adm[1]:  # فقط سوپرادمین
        return _redir("/admin/login")
    if not new_password or new_password != confirm_password:
        return _redir("/admin/admins?flash=رمزها+یکسان+نیستند+یا+خالی+است")
    # ⚠️ رفع امنیتی (بخش ۱۳ آیتم ۲ سند): قبلاً رمز جدید عیناً plaintext در .env
    # نوشته می‌شد. حالا هش نمکی PBKDF2 (همون فرمت admins.web_password_hash) ذخیره
    # می‌شه؛ _verify_super_pw هم فرمت جدید هم plaintext قدیمی (برای مقدار bootstrap
    # دستی‌ای که هنوز از پنل تغییر نکرده) رو می‌پذیره.
    hashed_password = _hash_pw(new_password)
    env_path = os.path.join(os.path.dirname(__file__), ".env")
    try:
        lines = open(env_path, encoding="utf-8").readlines()
        new_lines = []
        found = False
        for line in lines:
            if line.startswith("ADMIN_WEB_PASSWORD="):
                new_lines.append(f"ADMIN_WEB_PASSWORD={hashed_password}\n")
                found = True
            else:
                new_lines.append(line)
        if not found:
            new_lines.append(f"ADMIN_WEB_PASSWORD={hashed_password}\n")
        open(env_path, "w", encoding="utf-8").writelines(new_lines)
        os.environ["ADMIN_WEB_PASSWORD"] = hashed_password
    except Exception as ex:
        return _redir(f"/admin/admins?flash=خطا:+{str(ex)[:40]}")
    return _redir("/admin/admins?flash=رمز+سوپرادمین+تغییر+کرد")


@router.post("/admins/super/telegram_id")
async def super_change_telegram_id(request: Request, new_telegram_id: str = Form("")):
    adm = _get_admin(request)
    if not adm or not adm[1]:
        return _redir("/admin/login")
    try:
        int(new_telegram_id)
    except ValueError:
        return _redir("/admin/admins?flash=آیدی+نامعتبر")
    env_path = os.path.join(os.path.dirname(__file__), ".env")
    try:
        lines = open(env_path, encoding="utf-8").readlines()
        new_lines = []
        found = False
        for line in lines:
            if line.startswith("ADMIN_ID="):
                new_lines.append(f"ADMIN_ID={new_telegram_id}\n")
                found = True
            else:
                new_lines.append(line)
        if not found:
            new_lines.append(f"ADMIN_ID={new_telegram_id}\n")
        open(env_path, "w", encoding="utf-8").writelines(new_lines)
        os.environ["ADMIN_ID"] = new_telegram_id
    except Exception as ex:
        return _redir(f"/admin/admins?flash=خطا:+{str(ex)[:40]}")
    return _redir("/admin/admins?flash=آیدی+تلگرام+تغییر+کرد+—+ربات+را+ریستارت+کنید")


@router.post("/admins/add")
async def admins_add(request: Request):
    adm = _get_admin(request)
    guard = _require(adm, "admins")
    if guard: return guard

    form = await request.form()
    name         = (form.get("name") or "").strip()
    web_username = (form.get("web_username") or "").strip()
    web_password = (form.get("web_password") or "").strip()
    telegram_id  = form.get("telegram_id") or None
    notes        = (form.get("notes") or "").strip()

    if not name or not web_username or not web_password:
        return _redir("/admin/admins?flash=فیلدهای+اجباری+را+پر+کنید")

    perms = [p.replace("perm_", "") for p in form.keys() if p.startswith("perm_")]

    try:
        tg_id = int(telegram_id) if telegram_id else None
    except ValueError:
        tg_id = None

    ensure_admins_table()
    conn = _db()
    try:
        conn.execute(
            """INSERT INTO admins (telegram_id, name, web_username, web_password_hash, permissions, notes, created_at)
               VALUES (?,?,?,?,?,?,?);""",
            (tg_id, name, web_username, _hash_pw(web_password), json.dumps(perms), notes, datetime.utcnow().isoformat()),
        )
        conn.commit()
    except _INTEGRITY_ERRORS:
        conn.close()
        return _redir("/admin/admins?flash=یوزرنیم+یا+تلگرام+تکراری+است")
    finally:
        try: conn.close()
        except: pass

    _log(request, "ایجاد ادمین", "ادمین‌ها", f"یوزرنیم: {web_username}")
    return _redir("/admin/admins?flash=ادمین+جدید+اضافه+شد")

@router.get("/admins/{aid}/edit", response_class=HTMLResponse)
async def admins_edit_get(request: Request, aid: int, flash: str = ""):
    adm = _get_admin(request)
    guard = _require(adm, "admins")
    if guard: return guard

    ensure_admins_table()
    conn = _db()
    try:
        a = conn.execute("SELECT * FROM admins WHERE id=? LIMIT 1;", (aid,)).fetchone()
    finally:
        conn.close()

    if not a:
        return _redir("/admin/admins")

    cur_perms = json.loads(a["permissions"] or "[]")
    perm_checks = ""
    for perm_key, perm_label in ALL_PERMISSIONS.items():
        checked = "checked" if perm_key in cur_perms else ""
        perm_checks += f"""
        <label class="flex items-center gap-2 text-sm cursor-pointer">
          <input type="checkbox" name="perm_{perm_key}" {checked}
            class="rounded border-gray-300 text-indigo-600">
          {e(perm_label)}
        </label>"""

    body = f"""
    <a href="/admin/admins" class="text-indigo-600 text-sm mb-4 inline-block">← بازگشت به ادمین‌ها</a>
    <h1 class="text-2xl font-bold text-gray-800 mb-6">✏️ ویرایش ادمین: {e(a["name"])}</h1>
    <div class="bg-white rounded-xl shadow p-6 max-w-2xl">
      <form method="post" action="/admin/admins/{aid}/edit" class="space-y-4">
        <div class="grid md:grid-cols-2 gap-4">
          <div>
            <label class="text-xs text-gray-500 block mb-1">نام نمایشی</label>
            {_input("name", "", str(a["name"] or ""), required=True)}
          </div>
          <div>
            <label class="text-xs text-gray-500 block mb-1">آیدی تلگرام</label>
            {_input("telegram_id", "", str(a["telegram_id"] or ""), type_="number")}
          </div>
          <div>
            <label class="text-xs text-gray-500 block mb-1">یوزرنیم پنل</label>
            {_input("web_username", "", str(a["web_username"] or ""))}
          </div>
          <div>
            <label class="text-xs text-gray-500 block mb-1">رمز جدید (خالی = بدون تغییر)</label>
            {_input("web_password", "رمز جدید (اختیاری)", type_="password")}
          </div>
        </div>
        <div>
          <label class="text-xs text-gray-500 block mb-2">اختیارات</label>
          <div class="grid grid-cols-2 md:grid-cols-3 gap-2 p-4 bg-gray-50 rounded-lg">
            {perm_checks}
          </div>
        </div>
        <div>
          <label class="text-xs text-gray-500 block mb-1">یادداشت</label>
          {_input("notes", "", str(a["notes"] or ""))}
        </div>
        {_btn("ذخیره تغییرات", color="green")}
      </form>
    </div>"""

    return _layout(f"ویرایش ادمین #{aid}", body, adm, flash=flash)

@router.post("/admins/{aid}/edit")
async def admins_edit_post(request: Request, aid: int):
    adm = _get_admin(request)
    guard = _require(adm, "admins")
    if guard: return guard

    form = await request.form()
    name         = (form.get("name") or "").strip()
    web_username = (form.get("web_username") or "").strip()
    web_password = (form.get("web_password") or "").strip()
    telegram_id  = form.get("telegram_id") or None
    notes        = (form.get("notes") or "").strip()
    perms        = [p.replace("perm_", "") for p in form.keys() if p.startswith("perm_")]

    try:
        tg_id = int(telegram_id) if telegram_id else None
    except ValueError:
        tg_id = None

    ensure_admins_table()
    conn = _db()
    try:
        if web_password:
            conn.execute(
                "UPDATE admins SET name=?,telegram_id=?,web_username=?,web_password_hash=?,permissions=?,notes=? WHERE id=?;",
                (name, tg_id, web_username, _hash_pw(web_password), json.dumps(perms), notes, aid),
            )
        else:
            conn.execute(
                "UPDATE admins SET name=?,telegram_id=?,web_username=?,permissions=?,notes=? WHERE id=?;",
                (name, tg_id, web_username, json.dumps(perms), notes, aid),
            )
        conn.commit()
    finally:
        conn.close()

    detail = f"id:{aid} — دسترسی‌های تازه: {', '.join(perms) or '—'}" + (" — رمز عوض شد" if web_password else "")
    _log(request, "ویرایش ادمین", "ادمین‌ها", detail, admin_info=adm)
    return _redir(f"/admin/admins/{aid}/edit?flash=ذخیره+شد")

@router.post("/admins/{aid}/toggle")
async def admins_toggle(request: Request, aid: int):
    adm = _get_admin(request)
    guard = _require(adm, "admins")
    if guard: return guard
    ensure_admins_table()
    conn = _db()
    try:
        conn.execute("UPDATE admins SET is_active=CASE WHEN is_active=1 THEN 0 ELSE 1 END WHERE id=?;", (aid,))
        conn.commit()
        row = conn.execute("SELECT is_active FROM admins WHERE id=?;", (aid,)).fetchone()
    finally:
        conn.close()
    new_state = "فعال" if (row and row["is_active"]) else "غیرفعال"
    _log(request, "تغییر وضعیت ادمین", "ادمین‌ها", f"id:{aid} — وضعیت جدید: {new_state}", admin_info=adm)
    return _redir("/admin/admins?flash=وضعیت+تغییر+کرد")

@router.post("/admins/{aid}/delete")
async def admins_delete(request: Request, aid: int):
    adm = _get_admin(request)
    guard = _require(adm, "admins")
    if guard: return guard
    ensure_admins_table()
    conn = _db()
    try:
        conn.execute("DELETE FROM admins WHERE id=?;", (aid,))
        conn.commit()
    finally:
        conn.close()
    _log(request, "حذف ادمین", "ادمین‌ها", f"id: {aid}")
    return _redir("/admin/admins?flash=ادمین+حذف+شد")

# ─────────────────────────── Categories ────────────────────────────────────

def _render_cat_tree(cats_all: list, parent_id=None, depth=0) -> str:
    """رندر درختی دسته‌بندی‌ها"""
    rows = ""
    children = [c for c in cats_all if c["parent_id"] == parent_id]
    for cat in children:
        indent = "　" * depth
        emoji = (cat["emoji"] or "").strip()
        label = f"{emoji} {cat['name']}".strip() if emoji else cat["name"]
        active_badge = '<span class="text-xs text-green-600">فعال</span>' if cat["is_active"] else '<span class="text-xs text-red-500">غیرفعال</span>'
        rows += f"""
        <tr class="border-b hover:bg-gray-50">
          <td class="px-4 py-2 text-sm">{indent}{'└ ' if depth else ''}{e(label)}</td>
          <td class="px-4 py-2">{active_badge}</td>
          <td class="px-4 py-2 text-xs text-gray-400">{cat['sort_order']}</td>
          <td class="px-4 py-2 flex gap-1 flex-wrap">
            {_btn("ویرایش", f"/admin/categories/{cat['id']}/edit", "indigo", small=True)}
            <form method="post" action="/admin/categories/{cat['id']}/toggle" class="inline">
              <button class="btn-sm {"bg-red-100 text-red-700" if cat["is_active"] else "bg-green-100 text-green-700"} rounded">{"غیرفعال" if cat["is_active"] else "فعال"}</button>
            </form>
            <form method="post" action="/admin/categories/{cat['id']}/delete" onsubmit="return confirm('حذف شود؟ همه زیردسته‌ها و محصولات هم حذف می‌شوند.')" class="inline">
              <button class="btn-sm bg-red-100 text-red-700 rounded">حذف</button>
            </form>
          </td>
        </tr>"""
        rows += _render_cat_tree(cats_all, parent_id=cat["id"], depth=depth + 1)
    return rows


def _cat_select_options(cats_all: list, selected_id=None, exclude_id=None, parent_id=None, depth=0) -> str:
    opts = ""
    children = [c for c in cats_all if c["parent_id"] == parent_id]
    for cat in children:
        if cat["id"] == exclude_id:
            continue
        indent = "── " * depth
        sel = "selected" if cat["id"] == selected_id else ""
        opts += f'<option value="{cat["id"]}" {sel}>{indent}{e(cat["name"])}</option>'
        opts += _cat_select_options(cats_all, selected_id, exclude_id, parent_id=cat["id"], depth=depth + 1)
    return opts


@router.get("/categories", response_class=HTMLResponse)
async def categories_list(request: Request, flash: str = ""):
    adm = _get_admin(request)
    guard = _require(adm, "categories")
    if guard: return guard

    conn = _db()
    try:
        cats = conn.execute("SELECT * FROM categories ORDER BY parent_id NULLS FIRST, sort_order, name;").fetchall()
    finally:
        conn.close()

    tree_rows = _render_cat_tree(cats)
    cat_opts = '<option value="">— بدون والد (دسته ریشه) —</option>' + _cat_select_options(cats)

    body = f"""
    <div class="flex items-center justify-between mb-6">
      <h1 class="text-2xl font-bold text-gray-800">🗂 دسته‌بندی‌ها</h1>
    </div>

    <div class="card p-6 mb-6">
      <h2 class="font-bold text-gray-700 mb-4">➕ افزودن دسته جدید</h2>
      <form method="post" action="/admin/categories/add" class="grid md:grid-cols-4 gap-4 items-end">
        <div>
          <label class="text-xs text-gray-500 block mb-1">نام دسته *</label>
          {_input("name", "مثلاً: هوش مصنوعی", required=True)}
        </div>
        <div>
          <label class="text-xs text-gray-500 block mb-1">ایموجی</label>
          {_input("emoji", "🧩")}
        </div>
        <div>
          <label class="text-xs text-gray-500 block mb-1">والد (زیردسته‌ی چه چیزی؟)</label>
          <select name="parent_id">
            {cat_opts}
          </select>
        </div>
        <div>
          <label class="text-xs text-gray-500 block mb-1">ترتیب نمایش</label>
          {_input("sort_order", "0", "0", "number")}
        </div>
        <div class="md:col-span-4">{_btn("➕ افزودن دسته", color="green")}</div>
      </form>
    </div>

    <div class="card overflow-hidden">
      <div class="px-5 py-3 border-b bg-gray-50 text-sm font-medium text-gray-700">
        ساختار دسته‌بندی‌ها ({len(cats)} دسته)
        <span class="text-xs text-gray-400 mr-2">دسته‌های ریشه در منوی اصلی ربات نمایش داده می‌شوند</span>
      </div>
      <div class="overflow-x-auto">
        <table class="w-full text-right min-w-max">
          <thead><tr class="text-xs text-gray-500 border-b bg-gray-50">
            <th class="px-4 py-3">نام</th><th class="px-4 py-3">وضعیت</th>
            <th class="px-4 py-3">ترتیب</th><th class="px-4 py-3">عملیات</th>
          </tr></thead>
          <tbody>{tree_rows or "<tr><td colspan='4' class='text-center py-8 text-gray-400'>هنوز دسته‌ای اضافه نشده</td></tr>"}</tbody>
        </table>
      </div>
    </div>"""

    return _layout("دسته‌بندی‌ها", body, adm, flash=flash)


@router.post("/categories/add")
async def categories_add(request: Request, name: str = Form(""), emoji: str = Form(""),
                          parent_id: str = Form(""), sort_order: str = Form("0")):
    adm = _get_admin(request)
    guard = _require(adm, "categories")
    if guard: return guard

    name = name.strip()
    if not name:
        return _redir("/admin/categories?flash=نام+دسته+الزامی+است")

    pid = int(parent_id) if parent_id.strip().isdigit() else None
    slug = "".join(c if c.isalnum() else "_" for c in name).lower()[:40]
    now = datetime.now().isoformat()

    conn = _db()
    try:
        conn.execute(
            "INSERT INTO categories (name, slug, parent_id, emoji, sort_order, is_active, created_at) VALUES (?,?,?,?,?,1,?);",
            (name, slug, pid, emoji.strip() or "", int(sort_order or 0), now)
        )
        conn.commit()
    except Exception as ex:
        return _redir(f"/admin/categories?flash=خطا: {str(ex)[:50]}")
    finally:
        conn.close()
    return _redir("/admin/categories?flash=دسته+اضافه+شد")


@router.get("/categories/{cid}/edit", response_class=HTMLResponse)
async def categories_edit_get(request: Request, cid: int, flash: str = ""):
    adm = _get_admin(request)
    guard = _require(adm, "categories")
    if guard: return guard

    conn = _db()
    try:
        cat = conn.execute("SELECT * FROM categories WHERE id=? LIMIT 1;", (cid,)).fetchone()
        cats_all = conn.execute("SELECT * FROM categories ORDER BY parent_id NULLS FIRST, sort_order, name;").fetchall()
    finally:
        conn.close()

    if not cat:
        return _redir("/admin/categories")

    cat_opts = '<option value="">— بدون والد (دسته ریشه) —</option>' + _cat_select_options(
        cats_all, selected_id=cat["parent_id"], exclude_id=cid
    )

    body = f"""
    <a href="/admin/categories" class="text-indigo-600 text-sm mb-4 inline-block">← بازگشت به دسته‌بندی‌ها</a>
    <h1 class="text-2xl font-bold text-gray-800 mb-6">✏️ ویرایش: {e(cat["name"])}</h1>
    <div class="card p-6 max-w-xl">
      <form method="post" action="/admin/categories/{cid}/edit" class="space-y-4">
        <div>
          <label class="text-sm font-medium text-gray-700 block mb-1">نام دسته</label>
          {_input("name", "", str(cat["name"]), required=True)}
        </div>
        <div class="grid grid-cols-2 gap-4">
          <div>
            <label class="text-sm font-medium text-gray-700 block mb-1">ایموجی</label>
            {_input("emoji", "🧩", str(cat["emoji"] or ""))}
          </div>
          <div>
            <label class="text-sm font-medium text-gray-700 block mb-1">ترتیب نمایش</label>
            {_input("sort_order", "0", str(cat["sort_order"] or 0), "number")}
          </div>
        </div>
        <div>
          <label class="text-sm font-medium text-gray-700 block mb-1">والد</label>
          <select name="parent_id" class="w-full border border-gray-300 rounded-lg px-3 py-2">{cat_opts}</select>
        </div>
        <div class="flex items-center gap-3">
          <label class="text-sm font-medium text-gray-700">فعال</label>
          <input type="checkbox" name="is_active" value="1" {"checked" if cat["is_active"] else ""} class="rounded">
        </div>
        {_btn("ذخیره تغییرات", color="green")}
      </form>
    </div>"""

    return _layout(f"ویرایش دسته #{cid}", body, adm, flash=flash)


@router.post("/categories/{cid}/edit")
async def categories_edit_post(request: Request, cid: int,
    name: str = Form(""), emoji: str = Form(""), parent_id: str = Form(""),
    sort_order: str = Form("0"), is_active: str = Form("")):
    adm = _get_admin(request)
    guard = _require(adm, "categories")
    if guard: return guard

    pid = int(parent_id) if parent_id.strip().isdigit() else None
    active = 1 if is_active == "1" else 0

    conn = _db()
    try:
        conn.execute(
            "UPDATE categories SET name=?, emoji=?, parent_id=?, sort_order=?, is_active=? WHERE id=?;",
            (name.strip(), emoji.strip(), pid, int(sort_order or 0), active, cid)
        )
        conn.commit()
    finally:
        conn.close()
    return _redir(f"/admin/categories/{cid}/edit?flash=ذخیره+شد")


@router.post("/categories/{cid}/toggle")
async def categories_toggle(request: Request, cid: int):
    adm = _get_admin(request)
    guard = _require(adm, "categories")
    if guard: return guard
    conn = _db()
    try:
        conn.execute("UPDATE categories SET is_active=CASE WHEN is_active=1 THEN 0 ELSE 1 END WHERE id=?;", (cid,))
        conn.commit()
    finally:
        conn.close()
    return _redir("/admin/categories?flash=وضعیت+تغییر+کرد")


@router.post("/categories/{cid}/delete")
async def categories_delete(request: Request, cid: int):
    adm = _get_admin(request)
    guard = _require(adm, "categories")
    if guard: return guard

    def collect_ids(conn, cat_id):
        ids = [cat_id]
        children = conn.execute("SELECT id FROM categories WHERE parent_id=?;", (cat_id,)).fetchall()
        for ch in children:
            ids.extend(collect_ids(conn, ch[0]))
        return ids

    conn = _db()
    try:
        all_ids = collect_ids(conn, cid)
        placeholders = ",".join("?" * len(all_ids))
        conn.execute(f"DELETE FROM product_feed WHERE product_id IN (SELECT id FROM products WHERE category_id IN ({placeholders}));", all_ids)
        conn.execute(f"DELETE FROM products WHERE category_id IN ({placeholders});", all_ids)
        conn.execute(f"DELETE FROM categories WHERE id IN ({placeholders});", all_ids)
        conn.commit()
    finally:
        conn.close()
    return _redir("/admin/categories?flash=دسته+و+زیردسته‌ها+حذف+شدند")


# ─────────────────────────── Products ──────────────────────────────────────

@router.get("/products", response_class=HTMLResponse)
async def products_list(request: Request, page: int = 0, q: str = "", flash: str = ""):
    adm = _get_admin(request)
    guard = _require(adm, "products")
    if guard: return guard

    try:
        from db import ensure_product_support_schema
        ensure_product_support_schema()
    except Exception:
        pass

    PAGE = 50
    conn = _db()
    try:
        where = "WHERE (p.title LIKE ? OR p.category LIKE ?)" if q else ""
        params_q = (f"%{q}%", f"%{q}%") if q else ()
        total = conn.execute(f"SELECT COUNT(*) FROM products p {where};", params_q).fetchone()[0]
        products = conn.execute(f"""
            SELECT p.*, COUNT(CASE WHEN pf.delivered=0 THEN 1 END) as feed_avail,
                   COUNT(pf.id) as feed_total
            FROM products p LEFT JOIN product_feed pf ON pf.product_id=p.id
            {where}
            GROUP BY p.id ORDER BY p.category, p.id
            LIMIT ? OFFSET ?;
        """, params_q+(PAGE, page*PAGE)).fetchall()
    finally:
        conn.close()

    pages = max((total+PAGE-1)//PAGE, 1)

    rows = ""
    for p in products:
        avail = int(p["feed_avail"] or 0)
        ac = "red" if avail==0 else ("yellow" if avail<5 else "green")
        status_badge = '<span class="px-2 py-0.5 text-xs bg-green-100 text-green-700 rounded-full">فعال</span>' if p["is_active"] else '<span class="px-2 py-0.5 text-xs bg-red-100 text-red-700 rounded-full">غیرفعال</span>'
        img_url = p["image_url"] if "image_url" in p.keys() else ""
        thumb = f'<img src="{e(img_url)}" class="w-8 h-8 rounded object-cover">' if img_url else '<div class="w-8 h-8 rounded bg-gray-100 flex items-center justify-center text-xs text-gray-300">—</div>'
        rows += f"""
        <tr class="border-b hover:bg-gray-50">
          <td class="px-4 py-3">{thumb}</td>
          <td class="px-4 py-3 text-sm font-medium text-gray-800">{e(p["title"])}</td>
          <td class="px-4 py-3 text-xs text-gray-400">{e(p["category"])}</td>
          <td class="px-4 py-3 text-sm font-medium text-indigo-700">{int(p["price"]):,}</td>
          <td class="px-4 py-3">{status_badge}</td>
          <td class="px-4 py-3">
            <span class="px-2 py-0.5 text-xs rounded-full bg-{ac}-100 text-{ac}-700">{avail}/{int(p["feed_total"] or 0)}</span>
          </td>
          <td class="px-4 py-3">
            <div class="flex gap-1">
              <a href="/admin/products/{p['id']}" class="btn-sm bg-indigo-50 text-indigo-700 border border-indigo-200 rounded px-2 py-1 text-xs">✏️</a>
              <a href="/admin/feed/{p['id']}" class="btn-sm bg-teal-50 text-teal-700 border border-teal-200 rounded px-2 py-1 text-xs">📦</a>
              <form method="post" action="/admin/products/{p['id']}/toggle" class="inline">
                <button class="btn-sm {"bg-red-50 text-red-600 border border-red-200" if p["is_active"] else "bg-green-50 text-green-600 border border-green-200"} rounded px-2 py-1 text-xs">
                  {"⊘" if p["is_active"] else "✓"}
                </button>
              </form>
              <form method="post" action="/admin/products/{p['id']}/delete" class="inline"
                onsubmit="return confirm('حذف شود؟')">
                <button class="btn-sm bg-red-50 text-red-600 border border-red-200 rounded px-2 py-1 text-xs">🗑</button>
              </form>
            </div>
          </td>
        </tr>"""

    pager = '<div class="flex gap-2 mt-4 justify-center">' + "".join(
        f'<a href="/admin/products?page={i}&q={e(q)}" class="px-3 py-1 rounded border text-sm {"bg-indigo-600 text-white" if i==page else "bg-white"}">{i+1}</a>'
        for i in range(min(pages, 10))
    ) + "</div>" if pages > 1 else ""

    body = f"""
    <div class="flex items-center justify-between mb-6 flex-wrap gap-3">
      <h1 class="text-2xl font-bold text-gray-800">📦 محصولات ({total:,})</h1>
      <div class="flex items-center gap-2 flex-wrap">
        <form method="get" class="flex gap-2">
          {_input("q","جستجو عنوان/دسته...",q)} {_btn("جستجو","","slate",True)}
        </form>
        {_btn("➕ محصول جدید", "/admin/products/new", "green")}
      </div>
    </div>
    <div class="card overflow-hidden">
      <div class="overflow-x-auto">
        <table class="w-full text-right min-w-max">
          <thead><tr class="text-xs text-gray-500 border-b bg-gray-50">
            <th class="px-4 py-3">عکس</th><th class="px-4 py-3">عنوان</th><th class="px-4 py-3">دسته</th>
            <th class="px-4 py-3">قیمت</th><th class="px-4 py-3">وضعیت</th>
            <th class="px-4 py-3">موجودی</th><th class="px-4 py-3">عملیات</th>
          </tr></thead>
          <tbody>{rows or "<tr><td colspan='7' class='text-center py-8 text-gray-400'>محصولی ثبت نشده</td></tr>"}</tbody>
        </table>
      </div>
      {pager}
    </div>"""

    return _layout("محصولات", body, adm, flash=flash)

@router.get("/products/new", response_class=HTMLResponse)
async def product_new_get(request: Request):
    adm = _get_admin(request)
    guard = _require(adm, "products")
    if guard: return guard

    conn = _db()
    try:
        cats_all = conn.execute("SELECT * FROM categories WHERE is_active=1 ORDER BY parent_id NULLS FIRST, sort_order, name;").fetchall()
    finally:
        conn.close()

    if not cats_all:
        body = f"""
        <a href="/admin/products" class="text-indigo-600 text-sm mb-4 inline-block">← بازگشت به محصولات</a>
        <h1 class="text-2xl font-bold text-gray-800 mb-6">➕ محصول جدید</h1>
        <div class="card p-6">
          <p class="text-amber-600">⚠️ ابتدا باید دسته‌بندی بسازید.</p>
          <div class="mt-4">{_btn("← ساخت دسته‌بندی", "/admin/categories", "indigo")}</div>
        </div>"""
        return _layout("محصول جدید", body, adm)

    cat_opts = _cat_select_options(cats_all)

    body = f"""
    <a href="/admin/products" class="text-indigo-600 text-sm mb-4 inline-block">← بازگشت به محصولات</a>
    <h1 class="text-2xl font-bold text-gray-800 mb-6">➕ محصول جدید</h1>
    <form method="post" action="/admin/products/new" enctype="multipart/form-data" class="card p-6 max-w-2xl space-y-4">
      <div>
        <label class="text-sm font-medium text-gray-700 block mb-1">دسته‌بندی *</label>
        <select name="category_id" required class="w-full border border-gray-300 rounded-lg px-3 py-2">
          <option value="">انتخاب کنید...</option>
          {cat_opts}
        </select>
      </div>
      <div><label class="text-sm font-medium text-gray-700 block mb-1">عنوان محصول *</label>
        {_input("title", "عنوان محصول", required=True)}</div>
      <div class="grid grid-cols-2 gap-4">
        <div><label class="text-sm font-medium text-gray-700 block mb-1">قیمت (تومان) *</label>
          {_input("price", "250000", type_="number", required=True)}</div>
        <div><label class="text-sm font-medium text-gray-700 block mb-1">قیمت همکار (0=یکسان)</label>
          {_input("partner_price", "0", type_="number")}</div>
      </div>
      <div class="grid grid-cols-2 gap-4">
        <div><label class="text-sm font-medium text-gray-700 block mb-1">سقف روزانه مشتری</label>
          {_input("limit_c", "0", type_="number")}</div>
        <div><label class="text-sm font-medium text-gray-700 block mb-1">سقف روزانه همکار</label>
          {_input("limit_p", "0", type_="number")}</div>
      </div>
      <div><label class="text-sm font-medium text-gray-700 block mb-1">توضیحات</label>
        {_textarea("description", "توضیحات محصول...", rows=3)}</div>
      <div>
        <label class="text-sm font-medium text-gray-700 block mb-1">تصویر محصول (اختیاری)</label>
        <input type="file" name="image" accept="image/*" class="w-full border border-gray-300 rounded-lg px-3 py-2 text-sm">
        <div class="text-xs text-gray-400 mt-1">اگه عکسی انتخاب نشه، آیکون پیش‌فرض دسته‌بندی نمایش داده می‌شه.</div>
      </div>
      <div class="option-toggle-box">
        <label class="perm-label option-toggle-label">
          <input type="checkbox" name="support_after_purchase" value="1"
            id="setup_chk_new" onchange="document.getElementById('setup_msg_new').style.display=this.checked?'block':'none'"
            class="option-check-16">
          <div>
            <strong>نیاز به راه‌اندازی / دریافت اطلاعات مشتری</strong>
            <div class="option-hint">پس از خرید، به جای تحویل مستقیم، گفتگوی راه‌اندازی باز می‌شود</div>
          </div>
        </label>
        <div id="setup_msg_new" class="option-reveal">
          <label class="field-label">
            متن راهنما برای مشتری — چه اطلاعاتی باید بفرستد؟
          </label>
          {_textarea("setup_message","مثلاً: لطفاً ایمیل اپل، شماره موبایل و کد تأیید دو مرحله‌ای را ارسال کنید.",rows=3)}
        </div>
      </div>
      <div class="option-toggle-box">
        <label class="perm-label option-toggle-label">
          <input type="checkbox" name="notify_on_restock" value="1"
            class="option-check-16">
          <div>
            <strong>اطلاع‌رسانی موجود شدن مجدد</strong>
            <div class="option-hint">وقتی موجودی این محصول صفر بشه، به‌جای دکمهٔ خرید، دکمهٔ «موجود شد اطلاع بده» به کاربر نشون داده می‌شه (بدون امکان خرید/پیش‌خرید). اگه خاموش باشه، فقط پیام «ناموجود» دیده می‌شه.</div>
          </div>
        </label>
      </div>
      <div class="option-toggle-box">
        <label class="perm-label option-toggle-label">
          <input type="checkbox" name="require_terms" value="1"
            id="terms_chk_new" onchange="document.getElementById('terms_text_new').style.display=this.checked?'block':'none'"
            class="option-check-16">
          <div>
            <strong>نیاز به تأیید قوانین خرید</strong>
            <div class="option-hint">اگه فعال باشه، کاربر قبل از پرداخت باید متن قوانین خرید رو ببینه و با تیک‌زدن یک چک‌باکس تأیید کنه (هم در ربات، هم در مینی‌اپ). پیش‌فرض خاموش.</div>
          </div>
        </label>
        <div id="terms_text_new" class="option-reveal">
          <label class="field-label">
            متن قوانین اختصاصی این محصول (اختیاری)
          </label>
          {_textarea("terms_text","اگه خالی بذارید، از متن پیش‌فرض عمومی استفاده می‌شه (قابل ویرایش از تنظیمات → قوانین خرید).",rows=4)}
        </div>
      </div>
      <div class="flex gap-3">{_btn("ذخیره محصول", color="green")} {_btn("انصراف", "/admin/products", "slate")}</div>
    </form>"""

    return _layout("محصول جدید", body, adm)

@router.post("/products/new")
async def product_new_post(request: Request,
    category_id: str=Form(""), title: str=Form(""), price: str=Form("0"),
    partner_price: str=Form("0"), limit_c: str=Form("0"), limit_p: str=Form("0"),
    description: str=Form("")):
    adm = _get_admin(request)
    guard = _require(adm, "products")
    if guard: return guard
    form = await request.form()
    support_after = 1 if form.get("support_after_purchase") == "1" else 0
    notify_on_restock = 1 if form.get("notify_on_restock") == "1" else 0
    require_terms = 1 if form.get("require_terms") == "1" else 0

    if not category_id.strip().isdigit():
        return _redir("/admin/products/new?flash=دسته‌بندی+انتخاب+کنید")

    # migration: اطمینان از وجود ستون قبل از ذخیره
    try:
        from db import ensure_product_support_schema
        ensure_product_support_schema()
    except Exception:
        pass

    cat_id = int(category_id)
    pp = int(partner_price or 0)
    slug = "".join(c if c.isalnum() else "_" for c in title).lower()[:40] or "product"

    image_file = form.get("image")
    image_url = ""
    if image_file is not None and getattr(image_file, "filename", ""):
        image_url = await _save_tutorial_file(image_file, "products", (".jpg", ".jpeg", ".png", ".webp", ".gif"), "pr")

    conn = _db()
    try:
        cat = conn.execute("SELECT slug, name FROM categories WHERE id=?;", (cat_id,)).fetchone()
        cat_slug = cat["slug"] if cat else str(cat_id)
        conn.execute("""
            INSERT INTO products (category, category_id, product_key, title, price, partner_price,
                daily_limit_customer, daily_limit_partner, description, is_active, support_after_purchase, setup_message, image_url, notify_on_restock, require_terms, terms_text, created_by, created_at)
            VALUES (?,?,?,?,?,?,?,?,?,1,?,?,?,?,?,?,?,datetime('now'));""",
            (cat_slug, cat_id, slug, title.strip(), int(price or 0), pp if pp > 0 else None,
             int(limit_c or 0), int(limit_p or 0), description.strip(), support_after,
             str(form.get("setup_message","")).strip(), image_url, notify_on_restock, require_terms,
             str(form.get("terms_text","")).strip(), _admin_id_of(adm)))
        conn.commit()
    finally:
        conn.close()
    return _redir("/admin/products?flash=محصول+اضافه+شد")

@router.get("/products/duplicates", response_class=HTMLResponse)
async def duplicate_products_page(request: Request, flash: str = ""):
    adm = _get_admin(request)
    guard = _require(adm, "feed")
    if guard: return guard
    import duplicate_products as dup
    groups = dup.find_duplicate_groups()

    def _pcard(p, is_original):
        badge = ('<span class="px-2 py-0.5 bg-emerald-100 text-emerald-700 rounded text-[10px] font-bold">✅ اصل</span>' if is_original
                 else '<span class="px-2 py-0.5 bg-red-100 text-red-700 rounded text-[10px] font-bold">⚠️ تکراری</span>')
        del_btn = "" if is_original else f'''
          <form method="post" action="/admin/products/duplicates/{p['id']}/delete" onsubmit="return confirm('این نسخهٔ تکراری حذف بشه؟ موجودی/دسته‌بندی قیمت‌خرید و FAQهای این نسخه پاک می‌شن (نسخهٔ اصل و سفارش‌های قبلی دست‌نخورده می‌مونن).')">
            <button class="text-xs text-red-600 border border-red-200 rounded-lg px-2 py-1 mt-2">🗑 حذف این تکراری</button>
          </form>'''
        return f"""
        <div class="border rounded-lg p-3 {'bg-emerald-50/40' if is_original else 'bg-red-50/40'}">
          <div class="flex items-center justify-between gap-2 mb-1">
            <span class="text-xs text-gray-400">#{p['id']}</span>{badge}
          </div>
          <div class="text-xs text-gray-500">دسته: {e(p['category'])}</div>
          <div class="text-xs text-gray-500">قیمت: {int(p['price'] or 0):,} ت</div>
          <div class="text-xs text-gray-500">موجودی: {p['stock']} عدد</div>
          <div class="text-xs text-gray-500">بچ خرید: {p['feed_batches_count']} ({int(p['feed_batches_cost']):,} ت)</div>
          <div class="text-xs text-gray-500">سفارش‌های ثبت‌شده: {p['orders_count']}</div>
          {del_btn}
        </div>"""

    groups_html = "".join(f"""
      <div class="card p-4 mb-4">
        <h3 class="font-bold text-gray-700 text-sm mb-3">📦 {e(g['title'])} <span class="text-xs text-gray-400 font-normal">({len(g['products'])} نسخه)</span></h3>
        <div class="grid md:grid-cols-3 gap-3">
          {"".join(_pcard(p, p['is_original']) for p in g['products'])}
        </div>
      </div>""" for g in groups) or '<div class="card p-10 text-center text-gray-400 text-sm">محصول تکراری‌ای (با تطابق دقیق عنوان) پیدا نشد. ✅</div>'

    body = f"""
    <div class="flex items-center gap-3 mb-6">
      {_btn("← موجودی","/admin/feed","slate",small=True)}
      <h1 class="text-2xl font-bold text-gray-800">🧬 بررسی محصولات تکراری</h1>
    </div>
    <div class="text-xs text-gray-400 mb-4">معیار تشخیص: تطابق دقیق عنوان محصول، مستقل از دسته‌بندی. در هر گروه، اولین محصول ثبت‌شده «اصل» در نظر گرفته می‌شه و بقیه «تکراری»‌ان. حذف هر تکراری فقط داده‌های همون نسخه (موجودی، بچ خرید، FAQ، امتیاز) رو پاک می‌کنه — نسخهٔ اصل و سفارش‌های قبلی دست‌نخورده می‌مونن.</div>
    {groups_html}"""
    return _layout("محصولات تکراری", body, adm, flash=flash)


@router.get("/products/{pid}/faqs", response_class=HTMLResponse)
async def product_faqs_page(request: Request, pid: int, flash: str = ""):
    adm = _get_admin(request)
    guard = _require(adm, "products")
    if guard: return guard
    from db import get_product_faqs, get_product_ratings_list, get_product_rating, ensure_faq_schema, ensure_ratings_schema
    ensure_faq_schema(); ensure_ratings_schema()
    conn = _db()
    try:
        prod = conn.execute("SELECT title FROM products WHERE id=?;", (pid,)).fetchone()
    finally:
        conn.close()
    if not prod:
        return _redir("/admin/products")
    faqs = get_product_faqs(pid)
    ratings = get_product_ratings_list(pid, 20)
    rstat = get_product_rating(pid)

    faq_rows = "".join(f"""<tr class="border-b hover:bg-gray-50">
      <td class="px-4 py-3 text-sm font-medium">{e(f['question'])}</td>
      <td class="px-4 py-3 text-sm text-gray-500">{e(f['answer'][:60])}...</td>
      <td class="px-4 py-3">
        <form method="post" action="/admin/products/{pid}/faqs/{f['id']}/delete" onsubmit="return confirm('حذف؟')">
          <button class="text-xs text-red-400 hover:text-red-600">حذف</button>
        </form>
      </td>
    </tr>""" for f in faqs) or "<tr><td colspan='3' class='text-center py-4 text-gray-400'>سوالی ثبت نشده</td></tr>"

    rating_rows = "".join(f"""<tr class="border-b hover:bg-gray-50 text-sm">
      <td class="px-3 py-2">{"⭐️"*r['rating']}</td>
      <td class="px-3 py-2">{e(r['full_name'] or '—')}</td>
      <td class="px-3 py-2 text-gray-500">{e(r['comment'] or '—')}</td>
      <td class="px-3 py-2 text-xs text-gray-400">{fa_date(r['created_at'])}</td>
    </tr>""" for r in ratings) or "<tr><td colspan='4' class='text-center py-4 text-gray-400'>نظری ثبت نشده</td></tr>"

    body = f"""
    <div class="flex items-center gap-3 mb-6">
      {_btn("← محصولات", "/admin/products", "slate", small=True)}
      <h1 class="text-xl font-bold text-gray-800">📦 {e(prod[0])}</h1>
    </div>

    <div class="grid md:grid-cols-2 gap-4 mb-6">
      <!-- FAQ -->
      <div class="card p-6">
        <h2 class="font-bold text-gray-700 mb-4">❓ سوالات متداول</h2>
        <form method="post" action="/admin/products/{pid}/faqs/new" class="space-y-3 mb-4">
          <input type="text" name="question" placeholder="سوال..." required
            class="w-full border border-gray-200 rounded-lg px-3 py-2 text-sm">
          {_textarea("answer", "جواب...", rows=3)}
          {_btn("افزودن سوال","",color="indigo",small=True)}
        </form>
        <div class="overflow-x-auto"><table class="w-full text-right text-sm">
          <thead><tr class="text-xs text-gray-500 border-b">
            <th class="px-4 py-2">سوال</th><th class="px-4 py-2">جواب</th><th></th>
          </tr></thead>
          <tbody>{faq_rows}</tbody>
        </table></div>
      </div>

      <!-- امتیازها -->
      <div class="card p-6">
        <h2 class="font-bold text-gray-700 mb-2">⭐️ نظرات کاربران</h2>
        <div class="text-3xl font-bold text-amber-500 mb-1">{rstat['avg']}/5</div>
        <div class="text-xs text-gray-400 mb-4">{rstat['count']} نظر ثبت‌شده</div>
        <div class="overflow-x-auto"><table class="w-full text-right">
          <thead><tr class="text-xs text-gray-500 border-b">
            <th class="px-3 py-2">امتیاز</th><th class="px-3 py-2">کاربر</th>
            <th class="px-3 py-2">نظر</th><th class="px-3 py-2">تاریخ</th>
          </tr></thead>
          <tbody>{rating_rows}</tbody>
        </table></div>
      </div>
    </div>"""
    return _layout(f"FAQ — {prod[0]}", body, adm, flash=flash)


@router.post("/products/{pid}/faqs/new")
async def product_faq_new(request: Request, pid: int):
    adm = _get_admin(request)
    guard = _require(adm, "products")
    if guard: return guard
    from db import add_product_faq, ensure_faq_schema
    ensure_faq_schema()
    form = await request.form()
    q = str(form.get("question","")).strip()
    a = str(form.get("answer","")).strip()
    if q and a:
        add_product_faq(pid, q, a)
        _log(request, "افزودن FAQ", "محصولات", f"product:{pid}")
    return _redir(f"/admin/products/{pid}/faqs?flash=سوال+اضافه+شد")


@router.post("/products/{pid}/faqs/{fid}/delete")
async def product_faq_delete(request: Request, pid: int, fid: int):
    adm = _get_admin(request)
    guard = _require(adm, "products")
    if guard: return guard
    from db import delete_product_faq
    delete_product_faq(fid)
    return _redir(f"/admin/products/{pid}/faqs?flash=حذف+شد")


@router.get("/products/{pid}", response_class=HTMLResponse)
async def product_edit_get(request: Request, pid: int, flash: str = ""):
    adm = _get_admin(request)
    guard = _require(adm, "products")
    if guard: return guard

    try:
        from db import ensure_product_support_schema
        ensure_product_support_schema()
    except Exception:
        pass

    conn = _db()
    try:
        p = conn.execute("SELECT * FROM products WHERE id=?;", (pid,)).fetchone()
        services = conn.execute("SELECT service_key, title FROM other_services ORDER BY title;").fetchall()
        feed = conn.execute("SELECT COUNT(*) as t, COUNT(CASE WHEN delivered=0 THEN 1 END) as a FROM product_feed WHERE product_id=?;", (pid,)).fetchone()
    finally:
        conn.close()

    if not p:
        return _redir("/admin/products")

    from db import get_price_history
    price_history = get_price_history(pid, limit=10)

    cats = ""
    for s in [("apple","سرویس‌های اپل آیدی")] + [(r["service_key"],r["title"]) for r in services]:
        sel = "selected" if s[0]==p["category"] else ""
        cats += f'<option value="{e(s[0])}" {sel}>{e(s[1])}</option>'

    body = f"""
    <a href="/admin/products" class="text-indigo-600 text-sm mb-4 inline-block">← بازگشت به محصولات</a>
    <h1 class="text-2xl font-bold text-gray-800 mb-6">✏️ ویرایش محصول #{pid}</h1>
    <div class="grid md:grid-cols-3 gap-6">
      <div class="md:col-span-2">
        <form method="post" action="/admin/products/{pid}/edit" enctype="multipart/form-data" class="bg-white rounded-xl shadow p-6 space-y-4">
          <div><label class="text-sm font-medium text-gray-700 block mb-1">دسته</label>
            <select name="category" class="w-full border border-gray-300 rounded-lg px-3 py-2">{cats}</select></div>
          <div><label class="text-sm font-medium text-gray-700 block mb-1">عنوان</label>
            {_input("title","",str(p["title"] or ""),required=True)}</div>
          <div class="grid grid-cols-2 gap-4">
            <div><label class="text-sm font-medium text-gray-700 block mb-1">قیمت</label>
              {_input("price","",str(p["price"] or 0),"number",True)}</div>
            <div><label class="text-sm font-medium text-gray-700 block mb-1">قیمت همکار</label>
              {_input("partner_price","",str(p["partner_price"] or 0),"number")}</div>
          </div>
          <div class="grid grid-cols-2 gap-4">
            <div><label class="text-sm font-medium text-gray-700 block mb-1">سقف مشتری</label>
              {_input("limit_c","",str(p["daily_limit_customer"] or 0),"number")}</div>
            <div><label class="text-sm font-medium text-gray-700 block mb-1">سقف همکار</label>
              {_input("limit_p","",str(p["daily_limit_partner"] or 0),"number")}</div>
          </div>
          <div><label class="text-sm font-medium text-gray-700 block mb-1">توضیحات</label>
            {_textarea("description","",str(p["description"] or ""),rows=3)}</div>
          <div>
            <label class="text-sm font-medium text-gray-700 block mb-1">تصویر محصول {f'<img src="{e(p["image_url"])}" class="inline w-8 h-8 rounded object-cover align-middle mr-1">' if (p["image_url"] if "image_url" in p.keys() else "") else ""}</label>
            <input type="file" name="image" accept="image/*" class="w-full border border-gray-300 rounded-lg px-3 py-2 text-sm">
            {f'<label class="flex items-center gap-1 text-xs text-red-500 mt-1"><input type="checkbox" name="image_remove" value="1"> 🗑 حذف تصویر فعلی (برگشت به آیکون پیش‌فرض)</label>' if (p["image_url"] if "image_url" in p.keys() else "") else ""}
          </div>
          <div class="p-4 bg-gray-50 rounded-lg">
            <label class="flex items-center gap-3 cursor-pointer mb-3">
              <input type="checkbox" name="support_after_purchase" value="1"
                id="setup_chk_edit"
                onchange="document.getElementById('setup_msg_edit').style.display=this.checked?'block':'none'"
                {"checked" if int(p["support_after_purchase"] if "support_after_purchase" in p.keys() else 0) else ""}>
              <div>
                <div class="text-sm font-medium text-gray-800">نیاز به راه‌اندازی / دریافت اطلاعات مشتری</div>
                <div class="text-xs text-gray-400 mt-0.5">پس از خرید، محصول مستقیم ارسال نمی‌شود — گفتگوی راه‌اندازی باز می‌شود</div>
              </div>
            </label>
            <div id="setup_msg_edit" style="display:{"block" if int(p["support_after_purchase"] if "support_after_purchase" in p.keys() else 0) else "none"}">
              <label class="text-xs font-medium text-gray-600 block mb-1">
                متن راهنما برای مشتری — چه اطلاعاتی باید بفرستد؟
              </label>
              {_textarea("setup_message","مثلاً: ایمیل اپل، شماره موبایل و کد تأیید را ارسال کنید.",
                         value=str(p["setup_message"] if "setup_message" in p.keys() else ""),rows=3)}
            </div>
          </div>
          <div class="p-4 bg-gray-50 rounded-lg">
            <label class="flex items-center gap-3 cursor-pointer">
              <input type="checkbox" name="notify_on_restock" value="1"
                {"checked" if int(p["notify_on_restock"] if "notify_on_restock" in p.keys() else 0) else ""}>
              <div>
                <div class="text-sm font-medium text-gray-800">اطلاع‌رسانی موجود شدن مجدد</div>
                <div class="text-xs text-gray-400 mt-0.5">وقتی موجودی صفر بشه، دکمهٔ خرید به «🔔 موجود شد اطلاع بده» تبدیل می‌شه (بدون خرید/پیش‌خرید). اگه خاموش باشه، فقط پیام «ناموجود» دیده می‌شه.</div>
              </div>
            </label>
          </div>
          <div class="p-4 bg-gray-50 rounded-lg">
            <label class="flex items-center gap-3 cursor-pointer mb-3">
              <input type="checkbox" name="require_terms" value="1"
                id="terms_chk_edit"
                onchange="document.getElementById('terms_text_edit').style.display=this.checked?'block':'none'"
                {"checked" if int(p["require_terms"] if "require_terms" in p.keys() else 0) else ""}>
              <div>
                <div class="text-sm font-medium text-gray-800">نیاز به تأیید قوانین خرید</div>
                <div class="text-xs text-gray-400 mt-0.5">قبل از پرداخت، کاربر باید متن قوانین رو تأیید کنه (هم ربات، هم مینی‌اپ).</div>
              </div>
            </label>
            <div id="terms_text_edit" style="display:{"block" if int(p["require_terms"] if "require_terms" in p.keys() else 0) else "none"}">
              <label class="text-xs font-medium text-gray-600 block mb-1">
                متن قوانین اختصاصی این محصول (اختیاری)
              </label>
              {_textarea("terms_text","اگه خالی بذارید، از متن پیش‌فرض عمومی استفاده می‌شه.",
                         value=str(p["terms_text"] if "terms_text" in p.keys() else ""),rows=4)}
              <div class="text-xs text-gray-400 mt-1">متن پیش‌فرض عمومی از <a href="/admin/settings/purchase-terms" target="_blank" class="text-indigo-600 underline">اینجا</a> قابل ویرایشه.</div>
            </div>
          </div>
          {_btn("ذخیره", color="green")}
        </form>
      </div>
      <div class="space-y-4">
        <div class="bg-white rounded-xl shadow p-5">
          <h3 class="font-bold text-gray-700 mb-2">📦 موجودی</h3>
          <div class="text-3xl font-bold text-indigo-700">{int(feed["a"] or 0)}</div>
          <div class="text-xs text-gray-400 mb-3">از {int(feed["t"] or 0)} کل</div>
          {_btn("مدیریت موجودی →", f"/admin/feed/{pid}", "teal")}
          {_btn("❓ FAQ و نظرات", f"/admin/products/{pid}/faqs", "amber")}
        </div>
        <div class="bg-white rounded-xl shadow p-5">
          <h3 class="font-bold text-gray-700 mb-2 text-sm">🗓 اطلاعات ثبت</h3>
          <div class="text-xs text-gray-500">تاریخ ثبت: {fa_date(p["created_at"]) if ("created_at" in p.keys() and p["created_at"]) else "نامشخص (قبل از این نسخه)"}</div>
          <div class="text-xs text-gray-500 mt-1">ثبت‌کننده: {e(str(p["created_by"])) if ("created_by" in p.keys() and p["created_by"]) else "نامشخص"}</div>
        </div>
        {f'''<div class="bg-white rounded-xl shadow p-5">
          <h3 class="font-bold text-gray-700 mb-2 text-sm">💹 تاریخچهٔ تغییر قیمت</h3>
          <div class="space-y-2 max-h-64 overflow-y-auto">
            {"".join(f'<div class="text-xs border-b pb-1.5"><div class="text-gray-400">{fa_date(h["changed_at"], with_time=True)}</div><div class="text-gray-700">{int(h["old_price"]):,} ← <b>{int(h["new_price"]):,}</b> تومان</div></div>' for h in price_history)}
          </div>
        </div>''' if price_history else ""}
        <div class="bg-white rounded-xl shadow p-5 space-y-2">
          <form method="post" action="/admin/products/{pid}/toggle">
            <button type="submit" class="w-full py-2 text-sm rounded-lg border-2 border-{"red" if p["is_active"] else "green"}-300 text-{"red" if p["is_active"] else "green"}-700 hover:bg-{"red" if p["is_active"] else "green"}-50">
              {"🔴 غیرفعال کردن" if p["is_active"] else "🟢 فعال کردن"}
            </button>
          </form>
          <form method="post" action="/admin/products/{pid}/delete" onsubmit="return confirm('حذف شود؟')">
            <button type="submit" class="w-full py-2 text-sm rounded-lg border-2 border-red-200 text-red-600 hover:bg-red-50">🗑 حذف</button>
          </form>
        </div>
      </div>
    </div>"""

    return _layout(f"محصول #{pid}", body, adm, flash=flash)

@router.post("/products/{pid}/edit")
async def product_edit_post(request: Request, pid: int,
    category: str=Form(""), title: str=Form(""), price: str=Form("0"),
    partner_price: str=Form("0"), limit_c: str=Form("0"), limit_p: str=Form("0"),
    description: str=Form("")):
    adm = _get_admin(request)
    guard = _require(adm, "products")
    if guard: return guard
    form = await request.form()
    support_after = 1 if form.get("support_after_purchase") == "1" else 0
    notify_on_restock = 1 if form.get("notify_on_restock") == "1" else 0
    require_terms = 1 if form.get("require_terms") == "1" else 0
    pp = int(partner_price or 0)
    # migration: اطمینان از وجود ستون
    try:
        from db import ensure_product_support_schema
        ensure_product_support_schema()
    except Exception:
        pass

    image_file = form.get("image")
    if image_file is not None and getattr(image_file, "filename", ""):
        image_url = await _save_tutorial_file(image_file, "products", (".jpg", ".jpeg", ".png", ".webp", ".gif"), "pr")
    elif form.get("image_remove") == "1":
        image_url = ""
    else:
        image_url = None  # یعنی دست نخوره — همون مقدار قبلی بمونه

    new_price = int(price or 0)
    new_pp = pp if pp > 0 else 0
    conn = _db()
    try:
        old_row = conn.execute("SELECT price, COALESCE(partner_price,0) FROM products WHERE id=?;", (pid,)).fetchone()
        if image_url is None:
            conn.execute("""UPDATE products SET category=?,title=?,price=?,partner_price=?,
                daily_limit_customer=?,daily_limit_partner=?,description=?,support_after_purchase=?,setup_message=?,notify_on_restock=?,require_terms=?,terms_text=? WHERE id=?;""",
                (category,title.strip(),int(price or 0),pp if pp>0 else None,
                 int(limit_c or 0),int(limit_p or 0),description.strip(),support_after,
                 str(form.get("setup_message","")).strip(),notify_on_restock,require_terms,
                 str(form.get("terms_text","")).strip(),pid))
        else:
            conn.execute("""UPDATE products SET category=?,title=?,price=?,partner_price=?,
                daily_limit_customer=?,daily_limit_partner=?,description=?,support_after_purchase=?,setup_message=?,image_url=?,notify_on_restock=?,require_terms=?,terms_text=? WHERE id=?;""",
                (category,title.strip(),int(price or 0),pp if pp>0 else None,
                 int(limit_c or 0),int(limit_p or 0),description.strip(),support_after,
                 str(form.get("setup_message","")).strip(),image_url,notify_on_restock,require_terms,
                 str(form.get("terms_text","")).strip(),pid))
        conn.commit()
    finally:
        conn.close()

    if old_row:
        from db import log_price_change
        log_price_change(pid, old_row[0], new_price, old_row[1], new_pp, changed_by=_admin_id_of(adm))

    return _redir(f"/admin/products/{pid}?flash=ذخیره+شد")

@router.post("/products/{pid}/toggle")
async def product_toggle(request: Request, pid: int):
    adm = _get_admin(request)
    guard = _require(adm, "products")
    if guard: return guard
    conn = _db()
    try:
        conn.execute("UPDATE products SET is_active=CASE WHEN is_active=1 THEN 0 ELSE 1 END WHERE id=?;", (pid,))
        conn.commit()
    finally:
        conn.close()
    return _redir(f"/admin/products/{pid}?flash=وضعیت+تغییر+کرد")

@router.post("/products/{pid}/delete")
async def product_delete(request: Request, pid: int):
    adm = _get_admin(request)
    guard = _require(adm, "products")
    if guard: return guard

    conn = _db()
    try:
        # چک feed — اگه موجودی داشت اجازه حذف نده
        avail = conn.execute(
            "SELECT COUNT(*) FROM product_feed WHERE product_id=? AND delivered=0;", (pid,)
        ).fetchone()[0]
        if avail > 0:
            conn.close()
            return _redir(f"/admin/products/{pid}?flash=⚠️+محصول+{avail}+موجودی+دارد.+ابتدا+موجودی+را+از+بخش+فید+پاک+کنید")

        conn.execute("DELETE FROM product_feed WHERE product_id=?;", (pid,))
        conn.execute("DELETE FROM products WHERE id=?;", (pid,))
        conn.commit()
    finally:
        conn.close()
    _log(request, "حذف محصول", "محصولات", f"id: {pid}")
    return _redir("/admin/products?flash=محصول+حذف+شد")

# ─────────────────────────── Feed ──────────────────────────────────────────

@router.get("/feed", response_class=HTMLResponse)
async def feed_overview(request: Request):
    adm = _get_admin(request)
    guard = _require(adm, "feed")
    if guard: return guard

    conn = _db()
    try:
        products = conn.execute("""
            SELECT p.id, p.title, p.category,
                   COUNT(CASE WHEN pf.delivered=0 THEN 1 END) as avail,
                   COUNT(pf.id) as total,
                   COALESCE(fas.threshold, 5) as threshold
            FROM products p
            LEFT JOIN product_feed pf ON pf.product_id=p.id
            LEFT JOIN feed_alert_settings fas ON fas.product_id=p.id
            GROUP BY p.id, p.title, p.category, fas.threshold ORDER BY avail ASC, p.title;
        """).fetchall()
    finally:
        conn.close()

    rows = ""
    for p in products:
        avail = int(p["avail"] or 0)
        total = int(p["total"] or 0)
        pct = int(avail/max(total,1)*100)
        c = "red" if avail==0 else ("yellow" if avail<=int(p["threshold"]) else "green")
        rows += f"""
        <tr class="border-b hover:bg-gray-50">
          <td class="px-4 py-3 font-medium text-sm">{e(p["title"])}</td>
          <td class="px-4 py-3 text-xs text-gray-400">{e(p["category"])}</td>
          <td class="px-4 py-3">
            <div class="flex items-center gap-2">
              <div class="flex-1 bg-gray-100 rounded-full h-2 ltr-num">
                <div class="bg-{c}-500 h-2 rounded-full" style="width:{pct}%"></div>
              </div>
              <span class="text-sm font-medium text-{c}-700 w-16">{avail}/{total}</span>
            </div>
          </td>
          <td class="px-4 py-3">{_btn("مدیریت →", f"/admin/feed/{p['id']}", "indigo", small=True)}</td>
        </tr>"""

    body = f"""
    <div class="flex items-center justify-between mb-6 flex-wrap gap-2">
      <h1 class="text-2xl font-bold text-gray-800">🗃 مدیریت موجودی</h1>
      {_btn("🧬 بررسی محصولات تکراری", "/admin/products/duplicates", "amber", small=True)}
    </div>
    <div class="card overflow-hidden">
      <div class="overflow-x-auto">
        <table class="w-full text-right min-w-max">
          <thead><tr class="text-xs text-gray-500 border-b bg-gray-50">
            <th class="px-4 py-3">محصول</th><th class="px-4 py-3">دسته</th>
            <th class="px-4 py-3">موجودی</th><th class="px-4 py-3"></th>
          </tr></thead>
          <tbody>{rows or "<tr><td colspan='4' class='text-center py-8 text-gray-400'>محصولی ثبت نشده</td></tr>"}</tbody>
        </table>
      </div>
    </div>"""

    return _layout("موجودی", body, adm)


# ══════════════════════════════════════════════════════════════════════════
# ─── بررسی و حذف محصولات تکراری (بخش ۸ سند مینی‌اپ) ──────────────────────────
# ══════════════════════════════════════════════════════════════════════════


@router.post("/products/duplicates/{pid}/delete")
async def duplicate_product_delete(request: Request, pid: int):
    adm = _get_admin(request)
    guard = _require(adm, "feed")
    if guard: return guard
    import duplicate_products as dup
    try:
        result = dup.delete_duplicate_product(pid)
        _log(request, "حذف محصول تکراری", "محصولات تکراری",
             f"product:{pid} — feed:{result.get('feed_items',0)} batches:{result.get('feed_batches',0)} faqs:{result.get('faqs',0)} ratings:{result.get('ratings',0)}",
             admin_info=adm)
        msg = e("✅ حذف شد")
    except ValueError as ex:
        msg = e(f"❌ {ex}")
    return _redir(f"/admin/products/duplicates?flash={msg}")


@router.get("/feed/{pid}", response_class=HTMLResponse)
async def feed_detail(request: Request, pid: int, page: int=0, flash: str=""):
    adm = _get_admin(request)
    guard = _require(adm, "feed")
    if guard: return guard

    PAGE = 20
    conn = _db()
    try:
        product = conn.execute("SELECT * FROM products WHERE id=?;", (pid,)).fetchone()
        if not product:
            return _redir("/admin/feed")
        total = conn.execute("SELECT COUNT(*) FROM product_feed WHERE product_id=?;", (pid,)).fetchone()[0]
        avail = conn.execute("SELECT COUNT(*) FROM product_feed WHERE product_id=? AND delivered=0;", (pid,)).fetchone()[0]
        items = conn.execute("""
            SELECT id, data, delivered, created_at FROM product_feed
            WHERE product_id=? ORDER BY id DESC LIMIT ? OFFSET ?;
        """, (pid, PAGE, page*PAGE)).fetchall()
        # تعداد برگشتی این محصول
        try:
            returned_cnt = conn.execute(
                "SELECT COUNT(*) FROM orders WHERE product_id=? AND status='returned';", (str(pid),)
            ).fetchone()[0]
        except Exception:
            returned_cnt = 0
    finally:
        conn.close()

    pages = max((total+PAGE-1)//PAGE, 1)
    # پیدا کردن تکراری‌ها برای نمایش برچسب
    conn2 = _db()
    try:
        dup_data = set(
            row[0] for row in conn2.execute("""
                SELECT data FROM product_feed WHERE product_id=? AND delivered=0
                GROUP BY data HAVING COUNT(*)>1;
            """, (pid,)).fetchall()
        )
    except Exception:
        dup_data = set()
    finally:
        conn2.close()

    items_html = ""
    for item in items:
        preview = str(item["data"] or "").splitlines()[0][:80] if item["data"] else "---"
        badge = '<span class="px-2 py-0.5 text-xs bg-green-100 text-green-700 rounded-full">تحویل‌شده</span>' if item["delivered"] else '<span class="px-2 py-0.5 text-xs bg-blue-100 text-blue-700 rounded-full">موجود</span>'
        dup_badge = ' <span class="px-1.5 py-0.5 text-xs bg-red-100 text-red-600 rounded-full font-bold">تکراری</span>' if (not item["delivered"] and item["data"] in dup_data) else ""
        items_html += f"""
        <tr class="border-b hover:bg-gray-50 text-sm">
          <td class="px-4 py-2 text-gray-400 font-mono">#{item["id"]}</td>
          <td class="px-4 py-2 font-mono text-xs truncate max-w-xs" dir="ltr" style="text-align:left">{e(preview)}{dup_badge}</td>
          <td class="px-4 py-2">{badge}</td>
          <td class="px-4 py-2 text-gray-400 text-xs">{fa_date(item["created_at"] or "")}</td>
          <td class="px-4 py-2 flex gap-1">
            {_btn("ویرایش", f"/admin/feed/item/{item['id']}/edit", "indigo", small=True)}
            <form method="post" action="/admin/feed/item/{item['id']}/delete" onsubmit="return confirm('حذف شود؟')" class="inline">
              <button class="btn-sm bg-red-100 text-red-600 rounded">حذف</button>
            </form>
          </td>
        </tr>"""

    pager = '<div class="flex gap-2 mt-4 justify-center">' + "".join(
        f'<a href="/admin/feed/{pid}?page={i}" class="px-3 py-1 rounded border text-sm {"bg-indigo-600 text-white" if i==page else "bg-white text-gray-600"}">{i+1}</a>'
        for i in range(min(pages, 10))
    ) + "</div>" if pages > 1 else ""

    body = f"""
    <a href="/admin/feed" class="text-indigo-600 text-sm mb-4 inline-block">← بازگشت به موجودی</a>
    <h1 class="text-2xl font-bold text-gray-800 mb-6">🗃 موجودی: {e(product["title"])}</h1>
    <div class="grid grid-cols-4 gap-4 mb-6">
      {_card("کل آیتم‌ها", str(total), "", "slate")}
      {_card("موجود", str(avail), "", "green")}
      {_card("تحویل‌شده", str(total-avail), "", "indigo")}
      {_card("برگشتی ↩️", str(returned_cnt), "بازگردانده‌شده", "red")}
    </div>

    <div class="card p-6 mb-6">
      <h2 class="font-bold text-gray-700 mb-4">➕ افزودن موجودی</h2>

      <!-- اطلاعات حسابداری (مشترک) -->
      <div class="bg-amber-50 border border-amber-200 rounded-xl p-4 mb-4">
        <h3 class="text-sm font-semibold text-amber-800 mb-3">📊 اطلاعات حسابداری (اختیاری)</h3>
        <div class="grid grid-cols-2 md:grid-cols-3 gap-3">
          <div><label class="text-xs text-gray-600 block mb-1">قیمت خرید هر واحد (ت) <span class="text-red-500">*</span></label>
            <input type="number" id="acc_purchase" name="purchase_price" value="0" min="0"
              class="w-full border border-amber-300 rounded-lg px-3 py-2 text-sm bg-amber-50" placeholder="اجباری"></div>
          <div><label class="text-xs text-gray-600 block mb-1">هزینه‌های جانبی (ت)</label>
            <input type="number" id="acc_side" name="side_cost" value="0"
              class="w-full border border-gray-200 rounded-lg px-3 py-2 text-sm"></div>
          <div><label class="text-xs text-gray-600 block mb-1">یادداشت</label>
            <input type="text" id="acc_notes" name="batch_notes" placeholder="مثلاً: خرید دوره‌ای"
              class="w-full border border-gray-200 rounded-lg px-3 py-2 text-sm"></div>
        </div>
      </div>

      <!-- افزودن متنی -->
      <form method="post" action="/admin/feed/{pid}/upload" class="mb-4">
        <input type="hidden" name="purchase_price" id="txt_pp" value="0">
        <input type="hidden" name="side_cost" id="txt_sc" value="0">
        <input type="hidden" name="batch_notes" id="txt_bn" value="">
        <div class="text-xs text-gray-500 bg-gray-50 p-3 rounded-lg mb-3">
          هر خط = یک آیتم | برای چندخطی: <code class="bg-gray-200 px-1 rounded">***</code> بین آیتم‌ها
        </div>
        {_textarea("items", "آیتم‌ها را اینجا paste کنید...", rows=6)}
        <div class="mt-3">{_btn("افزودن متنی", color="green")}</div>
      </form>

      <!-- آپلود فایل -->
      <form method="post" action="/admin/feed/{pid}/bulk-upload" enctype="multipart/form-data">
        <input type="hidden" name="purchase_price" id="file_pp" value="0">
        <input type="hidden" name="side_cost" id="file_sc" value="0">
        <input type="hidden" name="batch_notes" id="file_bn" value="">
        <div class="border-2 border-dashed border-gray-200 rounded-xl p-5 text-center">
          <div class="text-3xl mb-2">📁</div>
          <div class="text-sm font-semibold text-gray-700 mb-1">آپلود فایل (TXT / CSV / Excel)</div>
          <div class="text-xs text-gray-400 mb-4">TXT/CSV: هر خط یک آیتم (یا ستون data/item/account) — Excel: ردیف اول = عنوان ستون‌ها، هر ردیف بعدی یک محصول کامل (همهٔ ستون‌های اون ردیف با هم ترکیب می‌شن)</div>
          <div class="text-right mb-3">
            <label class="text-xs text-gray-500 block mb-1">📝 توضیحات مشترک (اختیاری) — به انتهای هر آیتم این آپلود اضافه می‌شه</label>
            {_textarea("common_suffix", "مثلاً: لطفاً پس از تحویل، Find My را از تنظیمات آیفون خاموش کنید.", rows=2)}
          </div>
          <input type="file" name="file" accept=".txt,.csv,.xlsx,.xlsm" required
            class="block w-full text-sm text-gray-600 mb-4 file:ml-2 file:py-2 file:px-4 file:rounded-lg file:border-0 file:text-sm file:bg-indigo-50 file:text-indigo-700 hover:file:bg-indigo-100">
          <button type="submit"
            class="w-full py-2.5 bg-indigo-600 hover:bg-indigo-700 text-white rounded-xl text-sm font-semibold transition">
            ⬆ آپلود و افزودن
          </button>
        </div>
      </form>
      <script>
      // sync accounting fields to hidden inputs
      function syncAcc(){{
        var pp=document.getElementById('acc_purchase').value;
        var sc=document.getElementById('acc_side').value;
        var bn=document.getElementById('acc_notes').value;
        ['txt_pp','file_pp'].forEach(id=>document.getElementById(id).value=pp);
        ['txt_sc','file_sc'].forEach(id=>document.getElementById(id).value=sc);
        ['txt_bn','file_bn'].forEach(id=>document.getElementById(id).value=bn);
      }}
      document.getElementById('acc_purchase').addEventListener('input',syncAcc);
      document.getElementById('acc_side').addEventListener('input',syncAcc);
      document.getElementById('acc_notes').addEventListener('input',syncAcc);
      </script>
    </div>

    <div class="card overflow-hidden">
      <div class="px-5 py-3 border-b bg-gray-50 flex flex-wrap justify-between items-center gap-2">
        <span class="text-sm font-medium">لیست آیتم‌ها ({total})</span>
        <div class="flex gap-2 overflow-x-auto pb-1">
          <form method="post" action="/admin/feed/{pid}/clear-delivered" onsubmit="return confirm('تحویل‌شده‌ها پاک شوند؟')">
            <button class="px-3 py-1.5 text-xs text-red-400 hover:text-red-600 border border-red-200 rounded-lg">🗑 پاک‌سازی تحویل‌شده‌ها</button>
          </form>
          <form method="post" action="/admin/feed/{pid}/delete-all" onsubmit="return confirm('⚠️ همه {total} آیتم حذف شوند؟')">
            <button class="px-3 py-1.5 text-xs text-red-600 hover:text-red-800 font-bold border border-red-300 rounded-lg">🗑🗑 حذف کل موجودی ({total})</button>
          </form>
        </div>
      </div>
      <div class="overflow-x-auto">
        <table class="w-full text-right min-w-max">
          <thead><tr class="text-xs text-gray-500 border-b bg-gray-50">
            <th class="px-4 py-3">ID</th>
            <th class="px-4 py-3">پیش‌نمایش</th>
            <th class="px-4 py-3">وضعیت</th>
            <th class="px-4 py-3">تاریخ</th>
            <th class="px-4 py-3"></th>
          </tr></thead>
          <tbody>{items_html or "<tr><td colspan='5' class='text-center py-8 text-gray-400'>آیتمی ثبت نشده</td></tr>"}</tbody>
        </table>
      </div>
      {pager}
    </div>"""

    return _layout(f"موجودی #{pid}", body, adm, flash=flash)

@router.post("/feed/{pid}/upload")
async def feed_upload(request: Request, pid: int, items: str=Form(""),
                      purchase_price: int=Form(0), side_cost: int=Form(0), batch_notes: str=Form("")):
    adm = _get_admin(request)
    guard = _require(adm, "feed")
    if guard: return guard

    import re as _re
    raw = items.strip()
    if _re.search(r"^\s*\*{3,}\s*$", raw, _re.MULTILINE):
        blocks = [b.strip() for b in _re.split(r"^\s*\*{3,}\s*$", raw, flags=_re.MULTILINE) if b.strip()]
    else:
        blocks = [ln.strip() for ln in raw.splitlines() if ln.strip()]

    if not blocks:
        return _redir(f"/admin/feed/{pid}?flash=آیتمی+یافت+نشد")

    now = datetime.utcnow().isoformat()
    conn = _db()
    try:
        # ثبت batch حسابداری اگه قیمت خرید داده شده
        batch_id = None
        if purchase_price > 0:
            from db import create_feed_batch, link_batch_to_feed, ensure_feed_batch_schema
            ensure_feed_batch_schema()
            batch_id = create_feed_batch(pid, purchase_price, side_cost, len(blocks), batch_notes)

        conn.executemany("INSERT INTO product_feed (product_id,data,delivered,created_at) VALUES (?,?,0,?);",
                         [(pid, b, now) for b in blocks])
        conn.execute("INSERT INTO feed_alert_settings (product_id,threshold,last_notified_remaining,updated_at) VALUES (?,5,NULL,?) "
                     "ON CONFLICT(product_id) DO UPDATE SET last_notified_remaining=NULL, updated_at=excluded.updated_at;", (pid, now))
        conn.commit()

        if batch_id:
            from db import link_batch_to_feed
            link_batch_to_feed(pid, batch_id, 0, len(blocks))
    finally:
        conn.close()

    # بررسی تکراری
    from db import add_feed_items as _check_dup
    dup_count = 0
    # (تکراری‌ها قبلاً ثبت شدن، فقط شمارش)
    conn2 = _db()
    try:
        dup_count = sum(1 for b in blocks if conn2.execute(
            "SELECT COUNT(*) FROM product_feed WHERE product_id=? AND data=? AND delivered=0;", (pid, b)
        ).fetchone()[0] > 1)
    except Exception:
        pass
    finally:
        conn2.close()

    dispatched = 0
    try:
        from bot import try_dispatch_pending_for_product
        dispatched = try_dispatch_pending_for_product(pid, limit=len(blocks))
    except Exception:
        pass

    msg = f"{len(blocks)}+آیتم+اضافه+شد"
    if dup_count:
        msg += f"+({dup_count}+تکراری)"
    if dispatched:
        msg += f"+و+{dispatched}+سفارش+معلق+تحویل+داده+شد"
    _log(request, f"افزودن {len(blocks)} آیتم به موجودی", "موجودی", f"product:{pid}")

    # اگر قیمت خرید ثبت نشده → صفحه قیمت‌گذاری batch باز شود (دقیقاً همون الزامی که
    # مسیر آپلود فایل از قبل داشت — این مسیر «افزودن متنی» تا این نشست همچین چکی
    # نداشت و آیتم بدون هیچ batch/قیمت خریدی بی‌صدا درج می‌شد، یعنی هزینهٔ خریدش
    # هیچ‌وقت در حسابداری لحاظ نمی‌شد)
    if purchase_price <= 0:
        return _redir(f"/admin/feed/{pid}/batch-pricing?n={len(blocks)}")
    return _redir(f"/admin/feed/{pid}?flash={msg}")

def _notify_restock_subscribers(pid: int, was_out_of_stock: bool) -> None:
    """اطلاع‌رسانی «موجود شد» به مشترکان (ربات) + علاقه‌مندی‌کننده‌ها (مینی‌اپ) — از
    feed_bulk_upload به‌عنوان BackgroundTask صدا زده می‌شه تا تماس‌های sync با تلگرام
    (که می‌تونن برای هر مشترک یه بار تکرار بشن) جلوی جواب‌دادن به خودِ ادمین رو نگیرن."""
    try:
        from db import get_stock_subscribers, mark_subscriptions_notified, reset_subscriptions_on_restock
        from db import get_product_by_id as _gpbi
        subs = get_stock_subscribers(pid)
        if subs:
            _prod = _gpbi(pid)
            _title = _prod[2] if _prod else f"محصول #{pid}"
            bot_token = _env("BOT_TOKEN")
            for sub_uid in subs:
                try:
                    _requests.post(
                        f"https://api.telegram.org/bot{bot_token}/sendMessage",
                        json={"chat_id": sub_uid,
                              "text": f"🔔 محصول <b>{_title}</b> موجود شد!\nهم‌اکنون می‌توانید خرید کنید.",
                              "parse_mode": "HTML"},
                        timeout=5
                    )
                except Exception:
                    pass
            mark_subscriptions_notified(pid)
            reset_subscriptions_on_restock(pid)
    except Exception:
        pass

    # اطلاع‌رسانی به علاقه‌مندی‌کننده‌های محصول (مینی‌اپ) — فقط وقتی واقعاً از ناموجود به موجود برگشته
    try:
        from db import get_product_favoriters, get_product_by_id as _gpbi2, add_notification
        favoriters = get_product_favoriters(pid) if was_out_of_stock else []
        if favoriters:
            _prod2 = _gpbi2(pid)
            _title2 = _prod2[2] if _prod2 else f"محصول #{pid}"
            bot_token2 = _env("BOT_TOKEN")
            for fav_uid in favoriters:
                try:
                    _requests.post(
                        f"https://api.telegram.org/bot{bot_token2}/sendMessage",
                        json={"chat_id": fav_uid,
                              "text": f"💚 محصولی که به علاقه‌مندی‌هاتون اضافه کرده بودید موجود شد!\n<b>{_title2}</b>",
                              "parse_mode": "HTML"},
                        timeout=5
                    )
                except Exception:
                    pass
                try:
                    add_notification(fav_uid, "موجود شد", f"«{_title2}» که به علاقه‌مندی‌هاتون اضافه کرده بودید موجود شد.", icon="💚")
                except Exception:
                    pass
    except Exception:
        pass


@router.post("/feed/{pid}/bulk-upload")
async def feed_bulk_upload(request: Request, pid: int, background_tasks: BackgroundTasks, file: UploadFile = None):
    adm = _get_admin(request)
    guard = _require(adm, "feed")
    if guard: return guard
    if not file or not file.filename:
        return _redir(f"/admin/feed/{pid}?flash=فایلی+انتخاب+نشد")
    # خواندن batch cost از form
    form = await request.form()
    purchase_price = int(form.get("purchase_price") or 0)
    side_cost      = int(form.get("side_cost") or 0)
    batch_notes    = str(form.get("batch_notes") or "").strip()

    try:
        raw = await file.read()
    except Exception:
        return _redir(f"/admin/feed/{pid}?flash=خطا+در+خواندن+فایل")

    fname_lower = (file.filename or "").lower()
    items = []
    if fname_lower.endswith(".xlsx") or fname_lower.endswith(".xlsm"):
        # مسیر اختصاصی اکسل — طبق درخواست صریح مالک پروژه، جدا از CSV/TXT: ردیف اول
        # هدر ستون‌هاست، هر ردیف بعدی یک محصول کامل که اطلاعاتش توی چند ستون (چپ به
        # راست) پخش شده، نه یک ستون. هر ردیف به یک آیتم چندخطی «برچسب: مقدار» تبدیل
        # می‌شه (دقیقاً فرمت *** ورودی متنی).
        try:
            from import_utils import parse_xlsx_labeled_items
            items = parse_xlsx_labeled_items(raw)
        except Exception:
            return _redir(f"/admin/feed/{pid}?flash=خطا+در+خواندن+فایل")
    elif fname_lower.endswith(".csv"):
        # مسیر CSV — از پارسر عمومی import_utils استفاده می‌کنه؛ ستون آیتم با چند اسم
        # رایج تطبیق داده می‌شه، وگرنه اولین ستون موجود در ردیف.
        try:
            from import_utils import parse_uploaded_rows, pick
            rows = parse_uploaded_rows(file.filename, raw)
        except Exception:
            return _redir(f"/admin/feed/{pid}?flash=خطا+در+خواندن+فایل")
        for row in rows:
            item = pick(row, "data", "item", "account", "content", "متن", "آیتم", "دیتا")
            if not item and row:
                item = next(iter(row.values()), "")
            item = (item or "").strip()
            if item:
                items.append(item)
    else:
        # مسیر قدیمی — TXT، دو فرمت پشتیبانی‌شده (بدون تغییر رفتار)
        text = raw.decode("utf-8", errors="ignore")
        if "***" in text:
            parts = text.split("***")
            for part in parts:
                item = part.strip()
                if item and item not in ("", "\n"):
                    items.append(item)
        else:
            for line in text.splitlines():
                line = line.strip()
                if not line or line.startswith("#"):
                    continue
                item = line.split(",")[0].strip()
                if item:
                    items.append(item)

    # توضیحات مشترک (اختیاری) — به انتهای هر آیتم همین آپلود اضافه می‌شه، مستقل از
    # فرمت فایل (TXT/CSV/Excel)؛ طبق درخواست صریح مالک پروژه — مثلاً یادآوری خاموش
    # کردن Find My برای هر اکانت اپل آیدی. فقط برای همین یک آپلود اعمال می‌شه، جایی
    # ذخیره نمی‌شه.
    common_suffix = str(form.get("common_suffix") or "").strip()
    if common_suffix and items:
        items = [f"{item}\n{common_suffix}" for item in items]

    if not items:
        return _redir(f"/admin/feed/{pid}?flash=فایل+خالی+است")

    # قبل از درج، موجودی فعلی رو چک می‌کنیم — اطلاع‌رسانی علاقه‌مندی‌ها فقط برای گذار
    # واقعی «ناموجود → موجود» باشه، نه هر بار که ادمین به یه محصول موجود موجودی اضافه می‌کنه.
    was_out_of_stock = False
    try:
        from db import get_feed_stats
        _t0, _remaining0, _d0 = get_feed_stats(pid)
        was_out_of_stock = (int(_remaining0 or 0) == 0)
    except Exception:
        pass

    conn = _db()
    try:
        conn.executemany(
            "INSERT INTO product_feed (product_id, data, delivered, created_at) VALUES (?,?,0,datetime('now'));",
            [(pid, item) for item in items]
        )
        conn.commit()
        # ثبت batch حسابداری
        if purchase_price > 0:
            from db import create_feed_batch, link_batch_to_feed, ensure_feed_batch_schema
            ensure_feed_batch_schema()
            batch_id = create_feed_batch(pid, purchase_price, side_cost, len(items), batch_notes)
            link_batch_to_feed(pid, batch_id, 0, len(items))
    finally:
        conn.close()

    _log(request, "آپلود موجودی", "موجودی", f"محصول #{pid} — {len(items)} آیتم", admin_info=adm)

    # ارسال خودکار به سفارشات معلق (FIFO)
    dispatched = 0
    try:
        import sys, os
        app_dir = os.path.dirname(__file__)
        if app_dir not in sys.path:
            sys.path.insert(0, app_dir)
        from bot import try_dispatch_pending_for_product
        dispatched = try_dispatch_pending_for_product(pid, limit=len(items))
        if dispatched > 0:
            _log(request, "ارسال خودکار سفارش معلق", "موجودی",
                 f"محصول #{pid} — {dispatched} سفارش تحویل داده شد")
    except Exception as _ex:
        _tg_logger.warning("pending dispatch error: %s", _ex)

    # اطلاع‌رسانی مشترکان/علاقه‌مندی‌ها در پس‌زمینه — قبلاً این حلقه‌ها (که می‌تونن به ازای
    # هر مشترک یه تماس sync با تلگرام بزنن) مستقیم توی درخواست ادمین اجرا می‌شدن و کل
    # event loop (یعنی کل ربات+پنل+API، چون همه یه پروسه‌ان) رو تا پایانشون قفل می‌کردن.
    background_tasks.add_task(_notify_restock_subscribers, pid, was_out_of_stock)

    flash_msg = f"✅+{len(items)}+آیتم+اضافه+شد"
    if dispatched > 0:
        flash_msg += f"+و+{dispatched}+سفارش+معلق+تحویل+داده+شد"

    # اگر قیمت خرید ثبت نشده → صفحه قیمت‌گذاری batch باز شود
    if purchase_price <= 0:
        return _redir(f"/admin/feed/{pid}/batch-pricing?n={len(items)}")
    return _redir(f"/admin/feed/{pid}?flash={flash_msg}")


@router.get("/feed/{pid}/batch-pricing", response_class=HTMLResponse)
async def feed_batch_pricing_get(request: Request, pid: int, n: int = 0):
    """صفحه قیمت‌گذاری بعد از آپلود فایل — قیمت خرید + هزینه جانبی + یادداشت"""
    adm = _get_admin(request)
    guard = _require(adm, "feed")
    if guard: return guard

    conn = _db()
    try:
        prod = conn.execute("SELECT title FROM products WHERE id=?;", (pid,)).fetchone()
    finally:
        conn.close()
    title = prod["title"] if prod else f"#{pid}"

    body = f"""
    <div class="max-w-lg mx-auto">
      <div class="card p-6">
        <div class="text-center mb-5">
          <div class="text-4xl mb-2">✅</div>
          <h1 class="text-lg font-bold text-gray-800">{n} آیتم به «{e(title)}» اضافه شد</h1>
          <p class="text-sm text-gray-500 mt-1">حالا اطلاعات حسابداری این محموله را ثبت کنید</p>
        </div>
        <form method="post" action="/admin/feed/{pid}/batch-pricing" class="space-y-4">
          <input type="hidden" name="n" value="{n}">
          <div>
            <label class="text-sm font-medium text-gray-700 block mb-1">قیمت خرید هر واحد (تومان) <span class="text-red-500">*</span></label>
            <input type="number" name="purchase_price" required min="1" placeholder="مثلاً 45000"
              class="w-full border border-gray-300 rounded-lg px-3 py-2.5 text-sm">
          </div>
          <div>
            <label class="text-sm font-medium text-gray-700 block mb-1">هزینه‌های جانبی (تومان)</label>
            <input type="number" name="side_cost" value="0" min="0"
              class="w-full border border-gray-300 rounded-lg px-3 py-2.5 text-sm">
          </div>
          <div>
            <label class="text-sm font-medium text-gray-700 block mb-1">یادداشت</label>
            <input type="text" name="batch_notes" placeholder="مثلاً: خرید دوره‌ای از تأمین‌کننده"
              class="w-full border border-gray-300 rounded-lg px-3 py-2.5 text-sm">
          </div>
          <div class="flex gap-3 pt-2">
            <button type="submit" class="flex-1 py-2.5 bg-green-600 hover:bg-green-700 text-white rounded-xl text-sm font-semibold transition">💾 ثبت و پایان</button>
            <a href="/admin/feed/{pid}" class="px-5 py-2.5 bg-gray-100 text-gray-500 rounded-xl text-sm text-center hover:bg-gray-200 transition">رد شدن</a>
          </div>
        </form>
      </div>
    </div>"""
    return _layout("قیمت‌گذاری محموله", body, adm)


@router.post("/feed/{pid}/batch-pricing")
async def feed_batch_pricing_post(request: Request, pid: int):
    adm = _get_admin(request)
    guard = _require(adm, "feed")
    if guard: return guard
    form = await request.form()
    purchase_price = int(form.get("purchase_price") or 0)
    side_cost      = int(form.get("side_cost") or 0)
    batch_notes    = str(form.get("batch_notes") or "").strip()
    n              = int(form.get("n") or 0)

    if purchase_price > 0 and n > 0:
        from db import create_feed_batch, link_batch_to_feed, ensure_feed_batch_schema
        ensure_feed_batch_schema()
        batch_id = create_feed_batch(pid, purchase_price, side_cost, n, batch_notes)
        link_batch_to_feed(pid, batch_id, 0, n)
        _log(request, "قیمت‌گذاری محموله", "موجودی", f"محصول #{pid} — {n} آیتم @ {purchase_price:,}ت", admin_info=adm)

    return _redir(f"/admin/feed/{pid}?flash=✅+اطلاعات+حسابداری+ثبت+شد")


@router.post("/feed/{pid}/delete-all")
async def feed_delete_all(request: Request, pid: int):
    adm = _get_admin(request)
    guard = _require(adm, "feed")
    if guard: return guard
    conn = _db()
    try:
        count = conn.execute("SELECT COUNT(*) FROM product_feed WHERE product_id=?;", (pid,)).fetchone()[0]
        conn.execute("DELETE FROM product_feed WHERE product_id=?;", (pid,))
        conn.commit()
    finally:
        conn.close()
    _log(request, "حذف کل موجودی", "موجودی", f"محصول #{pid} — {count} آیتم حذف شد")
    return _redir(f"/admin/feed/{pid}?flash=✅+{count}+آیتم+حذف+شد")


@router.post("/feed/{pid}/clear-delivered")
async def feed_clear(request: Request, pid: int, background_tasks: BackgroundTasks):
    adm = _get_admin(request)
    guard = _require(adm, "feed")
    if guard: return guard

    def _do_clear(product_id: int):
        conn = _db()
        try:
            # حذف دسته‌ای با LIMIT برای جلوگیری از قفل شدن DB
            while True:
                r = conn.execute(
                    "DELETE FROM product_feed WHERE rowid IN "
                    "(SELECT rowid FROM product_feed WHERE product_id=? AND delivered=1 LIMIT 500);",
                    (product_id,)
                )
                conn.commit()
                if r.rowcount == 0:
                    break
        finally:
            conn.close()

    # شمارش قبل از حذف
    conn = _db()
    try:
        n = conn.execute("SELECT COUNT(*) FROM product_feed WHERE product_id=? AND delivered=1;", (pid,)).fetchone()[0]
    finally:
        conn.close()

    background_tasks.add_task(_do_clear, pid)
    return _redir(f"/admin/feed/{pid}?flash={n}+آیتم+در+حال+پاکسازی+است")

@router.post("/feed/item/{fid}/delete")
async def feed_item_delete(request: Request, fid: int):
    adm = _get_admin(request)
    guard = _require(adm, "feed")
    if guard: return guard
    conn = _db()
    try:
        row = conn.execute("SELECT product_id FROM product_feed WHERE id=?;", (fid,)).fetchone()
        pid = row["product_id"] if row else 0
        conn.execute("DELETE FROM product_feed WHERE id=?;", (fid,))
        conn.commit()
    finally:
        conn.close()
    return _redir(f"/admin/feed/{pid}?flash=آیتم+حذف+شد")


@router.get("/feed/item/{fid}/edit", response_class=HTMLResponse)
async def feed_item_edit_get(request: Request, fid: int, flash: str = ""):
    adm = _get_admin(request)
    guard = _require(adm, "feed")
    if guard: return guard

    conn = _db()
    try:
        item = conn.execute("SELECT * FROM product_feed WHERE id=? LIMIT 1;", (fid,)).fetchone()
        if not item:
            return _redir("/admin/feed")
        product = conn.execute("SELECT title FROM products WHERE id=? LIMIT 1;", (item["product_id"],)).fetchone()
        product_title = product["title"] if product else f"#{item['product_id']}"
        # اطلاعات batch
        batch = None
        if item["batch_id"]:
            batch = conn.execute("SELECT * FROM feed_batches WHERE id=?;", (item["batch_id"],)).fetchone()
    finally:
        conn.close()

    batch_section = ""
    if batch:
        batch_section = f"""
        <div class="border-t pt-4 mt-2">
          <h3 class="text-sm font-semibold text-gray-700 mb-3">📊 اطلاعات حسابداری Batch #{batch['id']}</h3>
          <div class="grid grid-cols-2 gap-3">
            <div><label class="text-xs text-gray-500 block mb-1">قیمت خرید هر واحد (ت)</label>
              <input type="number" name="purchase_price" value="{batch['purchase_price']}"
                class="w-full border border-gray-200 rounded-lg px-3 py-2 text-sm"></div>
            <div><label class="text-xs text-gray-500 block mb-1">هزینه جانبی (ت)</label>
              <input type="number" name="side_cost" value="{batch['side_cost']}"
                class="w-full border border-gray-200 rounded-lg px-3 py-2 text-sm"></div>
          </div>
          <input type="hidden" name="batch_id" value="{batch['id']}">
        </div>"""

    body = f"""
    <a href="/admin/feed/{item['product_id']}" class="text-indigo-600 text-sm mb-4 inline-block">← بازگشت به موجودی</a>
    <h1 class="text-xl font-bold text-gray-800 mb-6">✏️ ویرایش آیتم فید #{fid} <span class="text-sm text-gray-400 font-normal">{e(product_title)}</span></h1>
    <div class="card p-6 max-w-2xl">
      <form method="post" action="/admin/feed/item/{fid}/edit" class="space-y-4">
        <div>
          <label class="text-sm font-medium text-gray-700 block mb-1">محتوای آیتم</label>
          {_textarea("data", "", str(item["data"] or ""), rows=6, ltr=True)}
        </div>
        <div class="flex items-center gap-3">
          <label class="text-sm font-medium text-gray-700">تحویل داده شده</label>
          <input type="checkbox" name="delivered" value="1" {"checked" if item["delivered"] else ""}>
        </div>
        {batch_section}
        <div class="flex gap-3">
          {_btn("ذخیره", color="green")}
          {_btn("انصراف", f"/admin/feed/{item['product_id']}", "slate")}
        </div>
      </form>
    </div>"""
    return _layout(f"ویرایش فید #{fid}", body, adm, flash=flash)


@router.post("/feed/item/{fid}/edit")
async def feed_item_edit_post(request: Request, fid: int):
    adm = _get_admin(request)
    guard = _require(adm, "feed")
    if guard: return guard
    form = await request.form()
    data      = str(form.get("data","")).strip()
    delivered = 1 if form.get("delivered") == "1" else 0
    batch_id  = form.get("batch_id")
    pp        = int(form.get("purchase_price") or 0)
    sc        = int(form.get("side_cost") or 0)

    conn = _db()
    try:
        row = conn.execute("SELECT product_id FROM product_feed WHERE id=?;", (fid,)).fetchone()
        pid = row["product_id"] if row else 0
        conn.execute("UPDATE product_feed SET data=?, delivered=? WHERE id=?;",
                     (data, delivered, fid))
        if batch_id:
            conn.execute("UPDATE feed_batches SET purchase_price=?, side_cost=? WHERE id=?;",
                         (pp, sc, int(batch_id)))
        conn.commit()
    finally:
        conn.close()
    return _redir(f"/admin/feed/{pid}?flash=آیتم+ویرایش+شد")

# ─────────────────────────── Orders ────────────────────────────────────────

@router.get("/discounts", response_class=HTMLResponse)
async def discounts_list(request: Request, flash: str = ""):
    adm = _get_admin(request)
    guard = _require(adm, "discounts")
    if guard: return guard
    from db import ensure_discount_table
    ensure_discount_table()
    conn = _db()
    try:
        try:
            codes = conn.execute("""
                SELECT dc.*,
                       (SELECT COUNT(*) FROM discount_usage du WHERE du.code_id=dc.id) as real_uses
                FROM discount_codes dc ORDER BY id DESC;
            """).fetchall()
        except Exception:
            codes = conn.execute("SELECT * FROM discount_codes ORDER BY id DESC;").fetchall()
        products   = conn.execute("SELECT id,title FROM products WHERE is_active=1 ORDER BY title;").fetchall()
        categories = conn.execute("SELECT id,name FROM categories WHERE is_active=1 ORDER BY name;").fetchall()
    finally:
        conn.close()

    prod_opts = '<option value="">— همه محصولات —</option>' + "".join(f'<option value="{p["id"]}">{e(p["title"])}</option>' for p in products)
    cat_opts  = '<option value="">— همه دسته‌ها —</option>'  + "".join(f'<option value="{c["id"]}">{e(c["name"])}</option>'  for c in categories)

    rows = ""
    for c in codes:
        type_fa  = {"percent":"درصد","fixed":"ثابت (تومان)","wallet":"اعتبار کیف‌پول"}.get(c["type"],"")
        status_b = '<span class="px-2 py-0.5 text-xs bg-green-100 text-green-700 rounded-full">فعال</span>' if c["is_active"] else '<span class="px-2 py-0.5 text-xs bg-red-100 text-red-700 rounded-full">غیرفعال</span>'
        flags    = []
        try:
            if c["first_buy_only"]: flags.append('<span class="px-1.5 py-0.5 text-xs bg-purple-100 text-purple-700 rounded">اولین خرید</span>')
            if c["vip_only"]:       flags.append('<span class="px-1.5 py-0.5 text-xs bg-yellow-100 text-yellow-700 rounded">VIP</span>')
        except Exception: pass
        try: max_v = c["max_value"]
        except: max_v = 0
        try: real_u = c["real_uses"]
        except: real_u = c["used_count"] or 0
        flag_html = " ".join(flags) or "—"
        rows += f"""<tr class="border-b hover:bg-gray-50">
          <td class="px-4 py-3"><code class="text-sm font-bold text-indigo-700">{e(c['code'])}</code></td>
          <td class="px-4 py-3 text-sm text-gray-700">{c['value']} {type_fa}{f" (سقف {max_v:,})" if max_v else ""}</td>
          <td class="px-4 py-3 text-xs text-gray-500">{real_u} / {c['max_uses'] or '∞'}</td>
          <td class="px-4 py-3 text-xs text-gray-400">{fa_date(c['expires_at']) if c['expires_at'] else '—'}</td>
          <td class="px-4 py-3">{flag_html}</td>
          <td class="px-4 py-3">{status_b}</td>
          <td class="px-4 py-3">
            <div class="flex gap-1">
              <form method="post" action="/admin/discounts/{c['id']}/toggle" class="inline">
                <button class="btn-sm {'bg-red-50 text-red-600 border border-red-200' if c['is_active'] else 'bg-green-50 text-green-600 border border-green-200'} rounded px-2 py-1 text-xs">{'غیرفعال' if c['is_active'] else 'فعال'}</button>
              </form>
              <form method="post" action="/admin/discounts/{c['id']}/delete" class="inline" onsubmit="return confirm('حذف شود؟')">
                <button class="btn-sm bg-red-50 text-red-600 border border-red-200 rounded px-2 py-1 text-xs">حذف</button>
              </form>
            </div>
          </td>
        </tr>"""

    body = f"""
    <div class="flex items-center justify-between mb-6">
      <h1 class="text-2xl font-bold text-gray-800">🏷 کدهای تخفیف</h1>
    </div>
    <div class="card p-6 mb-6">
      <h2 class="font-bold text-gray-700 mb-4">➕ کد جدید</h2>
      <form method="post" action="/admin/discounts/add" class="space-y-4">
        <div class="grid grid-cols-2 gap-4">
          <div><label class="text-sm font-medium text-gray-700 block mb-1">کد *</label>{_input("code","مثلاً: STLAND20",required=True)}</div>
          <div><label class="text-sm font-medium text-gray-700 block mb-1">نوع</label>
            <select name="type"><option value="percent">درصدی (%)</option><option value="fixed">ثابت (تومان)</option><option value="wallet">اعتبار کیف‌پول</option></select></div>
          <div><label class="text-sm font-medium text-gray-700 block mb-1">مقدار *</label>{_input("value","مثلاً: 20",type_="number",required=True)}</div>
          <div><label class="text-sm font-medium text-gray-700 block mb-1">سقف تخفیف (درصدی — ۰=نامحدود)</label>{_input("max_value","0",type_="number")}</div>
          <div><label class="text-sm font-medium text-gray-700 block mb-1">حداقل مبلغ سفارش</label>{_input("min_amount","0",type_="number")}</div>
          <div><label class="text-sm font-medium text-gray-700 block mb-1">سقف کل (۰=نامحدود)</label>{_input("max_uses","0",type_="number")}</div>
          <div><label class="text-sm font-medium text-gray-700 block mb-1">سقف هر کاربر</label>{_input("max_uses_per_user","0",type_="number")}</div>
          <div><label class="text-sm font-medium text-gray-700 block mb-1">تاریخ انقضا</label><input type="date" name="expires_at"></div>
          <div><label class="text-sm font-medium text-gray-700 block mb-1">محصول خاص</label><select name="product_id">{prod_opts}</select></div>
          <div><label class="text-sm font-medium text-gray-700 block mb-1">دسته خاص</label><select name="category_id">{cat_opts}</select></div>
        </div>
        <div class="flex flex-wrap gap-4 p-4 bg-gray-50 rounded-lg">
          <label class="flex items-center gap-2 text-sm cursor-pointer"><input type="checkbox" name="first_buy_only" value="1"> فقط اولین خرید</label>
          <label class="flex items-center gap-2 text-sm cursor-pointer"><input type="checkbox" name="vip_only" value="1"> فقط VIP</label>
        </div>
        <div><label class="text-sm font-medium text-gray-700 block mb-1">توضیحات</label>{_input("description","مناسبت، هدف...")}</div>
        {_btn("➕ افزودن","",color="green")}
      </form>
    </div>
    <div class="card overflow-hidden">
      <div class="overflow-x-auto">
        <table class="w-full text-right min-w-max">
          <thead><tr class="text-xs text-gray-500 border-b bg-gray-50">
            <th class="px-4 py-3">کد</th><th class="px-4 py-3">تخفیف</th>
            <th class="px-4 py-3">استفاده</th><th class="px-4 py-3">انقضا</th>
            <th class="px-4 py-3">ویژگی</th><th class="px-4 py-3">وضعیت</th><th class="px-4 py-3">عملیات</th>
          </tr></thead>
          <tbody>{rows or "<tr><td colspan='7' class='text-center py-8 text-gray-400'>کدی اضافه نشده</td></tr>"}</tbody>
        </table>
      </div>
    </div>"""
    return _layout("کدهای تخفیف", body, adm, flash=flash)


@router.post("/discounts/add")
async def discounts_add(request: Request):
    adm = _get_admin(request)
    guard = _require(adm, "discounts")
    if guard: return guard
    from db import ensure_discount_table
    ensure_discount_table()
    form     = await request.form()
    code     = str(form.get("code","")).strip().upper()
    dtype    = str(form.get("type","percent"))
    value    = int(form.get("value") or 0)
    max_val  = int(form.get("max_value") or 0)
    min_amt  = int(form.get("min_amount") or 0)
    max_uses = int(form.get("max_uses") or 0)
    max_per  = int(form.get("max_uses_per_user") or 0)
    exp      = str(form.get("expires_at","")).strip() or None
    pid      = form.get("product_id") or None
    cid      = form.get("category_id") or None
    fbo      = 1 if form.get("first_buy_only") == "1" else 0
    vip      = 1 if form.get("vip_only") == "1" else 0
    desc     = str(form.get("description","")).strip()
    if not code or not value:
        return _redir("/admin/discounts?flash=کد+و+مقدار+اجباری+است")
    conn = _db()
    try:
        try:
            conn.execute("""INSERT INTO discount_codes
                (code,type,value,max_value,min_amount,max_uses,max_uses_per_user,
                 product_id,category_id,first_buy_only,vip_only,expires_at,description)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?);""",
                (code,dtype,value,max_val,min_amt,max_uses,max_per,
                 pid or None, cid or None, fbo, vip, exp, desc))
        except Exception:
            # fallback برای schema قدیمی
            conn.execute("INSERT INTO discount_codes (code,type,value,max_uses,min_amount,expires_at) VALUES (?,?,?,?,?,?);",
                (code,dtype,value,max_uses,min_amt,exp))
        conn.commit()
    except Exception as ex:
        return _redir(f"/admin/discounts?flash=خطا:+{str(ex)}")
    finally:
        conn.close()
    _log(request, "ایجاد کد تخفیف", "تخفیف", f"کد: {code} | نوع: {dtype} | مقدار: {value}")
    return _redir("/admin/discounts?flash=کد+تخفیف+اضافه+شد")


@router.post("/discounts/{cid}/toggle")
async def discount_toggle(request: Request, cid: int):
    adm = _get_admin(request)
    guard = _require(adm, "discounts")
    if guard: return guard
    conn = _db()
    try:
        conn.execute("UPDATE discount_codes SET is_active=1-is_active WHERE id=?;", (cid,))
        conn.commit()
    finally:
        conn.close()
    return _redir("/admin/discounts?flash=وضعیت+تغییر+کرد")


@router.post("/discounts/{cid}/delete")
async def discount_delete(request: Request, cid: int):
    adm = _get_admin(request)
    guard = _require(adm, "discounts")
    if guard: return guard
    conn = _db()
    try:
        conn.execute("DELETE FROM discount_codes WHERE id=?;", (cid,))
        conn.commit()
    finally:
        conn.close()
    _log(request, "حذف کد تخفیف", "تخفیف", f"id:{cid}")
    return _redir("/admin/discounts?flash=کد+حذف+شد")


@router.get("/referrals", response_class=HTMLResponse)
async def referrals_page(request: Request, flash: str = ""):
    # ادغام شد در /admin/partners?tab=referrals
    return _redir("/admin/partners?tab=referrals")


async def _old_referrals_page_unused(request: Request, flash: str = ""):
    settings = get_referral_settings()
    conn = _db()
    try:
        refs = conn.execute("""
            SELECT r.*, u1.full_name as referrer_name, u2.full_name as referred_name
            FROM referrals r
            LEFT JOIN users u1 ON u1.user_id=r.referrer_id
            LEFT JOIN users u2 ON u2.user_id=r.referred_id
            ORDER BY r.id DESC LIMIT 200;
        """).fetchall()
        total     = conn.execute("SELECT COUNT(*) FROM referrals;").fetchone()[0]
        rewarded  = conn.execute("SELECT COUNT(*) FROM referrals WHERE rewarded=1;").fetchone()[0]
        total_pay = conn.execute("SELECT COALESCE(SUM(reward_amount),0) FROM referrals WHERE rewarded=1;").fetchone()[0]
    except Exception:
        refs = []; total = rewarded = total_pay = 0
    finally:
        conn.close()

    rows = "".join(f"""<tr class="border-b hover:bg-gray-50">
      <td class="px-4 py-3 text-xs text-gray-400">#{r['id']}</td>
      <td class="px-4 py-3 text-sm">{e(r['referrer_name'] or str(r['referrer_id']))}</td>
      <td class="px-4 py-3 text-sm">{e(r['referred_name'] or str(r['referred_id']))}</td>
      <td class="px-4 py-3">{'<span class="px-2 py-0.5 text-xs bg-green-100 text-green-700 rounded-full">✅ پرداخت شد</span>' if r['rewarded'] else '<span class="px-2 py-0.5 text-xs bg-gray-100 text-gray-600 rounded-full">در انتظار خرید</span>'}</td>
      <td class="px-4 py-3 text-sm font-medium text-green-600">{int(r['reward_amount'] or 0):,} ت</td>
      <td class="px-4 py-3 text-xs text-gray-400">{fa_date(r['created_at'])}</td>
    </tr>""" for r in refs)

    body = f"""
    <div class="flex items-center justify-between mb-6">
      <h1 class="text-2xl font-bold text-gray-800">👥 سیستم معرفی</h1>
    </div>
    <div class="grid grid-cols-3 gap-4 mb-6">
      <div class="card p-5 text-center"><div class="text-2xl font-bold text-indigo-600">{total}</div><div class="text-xs text-gray-400 mt-1">کل معرفی‌ها</div></div>
      <div class="card p-5 text-center"><div class="text-2xl font-bold text-green-600">{rewarded}</div><div class="text-xs text-gray-400 mt-1">پرداخت شده</div></div>
      <div class="card p-5 text-center"><div class="text-2xl font-bold text-amber-600">{int(total_pay):,}</div><div class="text-xs text-gray-400 mt-1">جمع پاداش (تومان)</div></div>
    </div>
    <div class="card p-6 mb-6">
      <h2 class="font-bold text-gray-700 mb-4">⚙️ تنظیمات</h2>
      <form method="post" action="/admin/referrals/settings" class="flex flex-wrap gap-4 items-end">
        <div><label class="text-sm font-medium text-gray-700 block mb-1">مبلغ پاداش (تومان)</label>
          {_input("reward_amount","",str(settings.get("reward_amount",5000)),"number",True)}</div>
        <div><label class="text-sm font-medium text-gray-700 block mb-1">وضعیت سیستم</label>
          <select name="is_active">
            <option value="1" {"selected" if settings.get("is_active") else ""}>فعال</option>
            <option value="0" {"" if settings.get("is_active") else "selected"}>غیرفعال</option>
          </select></div>
        {_btn("ذخیره","",color="green")}
      </form>
    </div>
    <div class="card overflow-hidden">
      <div class="overflow-x-auto">
        <table class="w-full text-right min-w-max">
          <thead><tr class="text-xs text-gray-500 border-b bg-gray-50">
            <th class="px-4 py-3">#</th><th class="px-4 py-3">معرف</th><th class="px-4 py-3">کاربر جدید</th>
            <th class="px-4 py-3">وضعیت</th><th class="px-4 py-3">پاداش</th><th class="px-4 py-3">تاریخ</th>
          </tr></thead>
          <tbody>{rows or "<tr><td colspan='6' class='text-center py-8 text-gray-400'>معرفی‌ای ثبت نشده</td></tr>"}</tbody>
        </table>
      </div>
    </div>"""
    return _layout("سیستم معرفی", body, adm, flash=flash)


@router.post("/referrals/settings")
async def referrals_settings(request: Request):
    adm = _get_admin(request)
    guard = _require(adm, "partners")
    if guard: return guard
    form = await request.form()
    amount = int(form.get("reward_amount") or 0)
    active = int(form.get("is_active") or 1)
    max_inv = int(form.get("max_invites") or 0)
    from db import ensure_referral_schema
    ensure_referral_schema()
    conn = _db()
    try:
        conn.execute(
            "UPDATE referral_settings SET reward_amount=?,is_active=?,max_invites=?,updated_at=datetime('now') WHERE id=1;",
            (amount, active, max_inv))
        conn.commit()
    finally:
        conn.close()
    # به‌روزرسانی bot_config برای check_invite_cap
    try:
        from db import set_cfg
        import json as _j
        set_cfg("referral_settings_ext", _j.dumps({"max_invites": max_inv, "cap_reset_on_purchase": 1}))
    except Exception:
        pass
    _log(request, "تنظیم هدیه دعوت", "همکاران", f"پاداش: {amount:,} | سقف: {max_inv}", admin_info=adm)
    return _redir("/admin/partners?tab=settings&flash=تنظیمات+هدیه+دعوت+ذخیره+شد")


@router.get("/orders/export.xlsx")
async def orders_export_excel(request: Request, q: str = "", status: str = ""):
    adm = _get_admin(request)
    guard = _require(adm, "orders")
    if guard: return guard

    conn = _db()
    try:
        wheres, params = [], []
        if q:
            wheres.append("(title LIKE ? OR CAST(user_id AS TEXT) LIKE ?)")
            params += [f"%{q}%", f"%{q}%"]
        if status:
            wheres.append("status=?"); params.append(status)
        where_sql = ("WHERE " + " AND ".join(wheres)) if wheres else ""
        orders = conn.execute(
            f"SELECT id,user_id,category,title,price,status,created_at FROM orders {where_sql} ORDER BY id DESC LIMIT 5000;",
            params
        ).fetchall()
    finally:
        conn.close()

    status_fa = {"active": "ارسال شد", "returned": "برگشتی", "pending": "در انتظار"}

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from fastapi.responses import Response
        import io

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "orders"
        ws.sheet_view.rightToLeft = True

        headers = ["#", "User ID", "دسته", "محصول", "مبلغ", "وضعیت", "تاریخ"]
        hfill = PatternFill("solid", fgColor="2EC4B6")
        hfont = Font(bold=True, color="FFFFFF", name="Calibri")
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.fill = hfill; c.font = hfont
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22

        for ri, o in enumerate(orders, 2):
            ws.cell(ri, 1, o["id"] or "")
            ws.cell(ri, 2, o["user_id"] or "")
            ws.cell(ri, 3, o["category"] or "")
            ws.cell(ri, 4, o["title"] or "")
            ws.cell(ri, 5, int(o["price"] or 0))
            ws.cell(ri, 6, status_fa.get(o["status"] or "", o["status"] or ""))
            ws.cell(ri, 7, fa_date(o["created_at"], with_time=True))
            if ri % 2 == 0:
                for ci in range(1, 8):
                    ws.cell(ri, ci).fill = PatternFill("solid", fgColor="F8FAFB")

        col_widths = [8, 12, 16, 28, 14, 14, 18]
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(1, ci).column_letter].width = w

        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)
        _log(request, "خروجی Excel", "سفارش‌ها", f"{len(orders)} ردیف")
        return Response(
            content=buf.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=orders_{len(orders)}.xlsx"}
        )

    except ImportError:
        # Fallback: CSV اگه openpyxl نصب نیست
        from fastapi.responses import Response
        lines = ["#,User ID,دسته,محصول,مبلغ,وضعیت,تاریخ"]
        for o in orders:
            lines.append(f'{o["id"]},{o["user_id"] or ""},{o["category"] or ""},'
                         f'"{(o["title"] or "").replace(chr(34), "")}",'
                         f'{int(o["price"] or 0)},{status_fa.get(o["status"] or "", "")},'
                         f'{fa_date(o["created_at"] or "", with_time=True)}')
        csv_content = "\ufeff" + "\n".join(lines)  # BOM برای UTF-8 در Excel
        _log(request, "خروجی CSV", "سفارش‌ها", f"{len(orders)} ردیف")
        return Response(
            content=csv_content.encode("utf-8"),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f"attachment; filename=orders_{len(orders)}.csv"}
        )


@router.get("/orders", response_class=HTMLResponse)
async def orders_list(request: Request, page: int=0, q: str="", flash: str=""):
    adm = _get_admin(request)
    guard = _require(adm, "orders")
    if guard: return guard

    PAGE = 30
    conn = _db()
    try:
        where = "WHERE user_id LIKE ?" if q else ""
        params_q = (f"%{q}%",) if q else ()
        total = conn.execute(f"SELECT COUNT(*) FROM orders {where};", params_q).fetchone()[0]
        orders = conn.execute(f"SELECT * FROM orders {where} ORDER BY id DESC LIMIT ? OFFSET ?;",
                              params_q+(PAGE, page*PAGE)).fetchall()
        # آمار برگشتی
        returned_total = conn.execute("SELECT COUNT(*) FROM orders WHERE status='returned';").fetchone()[0]
        recent_logs = _fetch_order_logs(conn, limit=50)
    finally:
        conn.close()

    pages = max((total+PAGE-1)//PAGE, 1)

    def order_status_badge(st):
        if st == "returned":
            return '<span class="px-2 py-0.5 text-xs rounded-full bg-red-100 text-red-700">برگشتی</span>'
        return '<span class="px-2 py-0.5 text-xs rounded-full bg-green-100 text-green-700">فعال</span>'

    rows = ""
    for o in orders:
        st = o["status"] if "status" in o.keys() and o["status"] else "active"
        is_returned = st == "returned"
        # سه دکمهٔ قدیمی (ویرایش/برگشت/تعویض) ادغام شدن توی یک دکمهٔ «وضعیت سفارش»
        # که همه‌شون رو یه‌جا نشون می‌ده (بخش ۴۲ CLAUDE.md). ستون «وضعیت» فقط بج
        # وضعیت خام رو نشون می‌ده — توضیح اضافه (اصلاح‌شده/تعویض‌شده و...) از اینجا
        # حذف شد و به‌جاش فقط در لاگ فعالیت پایین صفحه دیده می‌شه (بخش ۴۳ CLAUDE.md).
        action_btns = f'<a href="/admin/orders/{o["id"]}" class="px-2 py-1 text-xs bg-indigo-50 text-indigo-700 rounded hover:bg-indigo-100">🧾 وضعیت سفارش</a>'

        rows += f"""
        <tr class="border-b hover:bg-gray-50 text-sm {'bg-red-50/30' if is_returned else ''}">
          <td class="px-4 py-2 text-gray-400">#{o["id"]}</td>
          <td class="px-4 py-2 font-mono text-xs"><code>{e(o["user_id"])}</code></td>
          <td class="px-4 py-2">{e(o["title"])}</td>
          <td class="px-4 py-2 text-green-700 font-medium">{int(o["price"]):,} ت</td>
          <td class="px-4 py-2">{order_status_badge(st)}</td>
          <td class="px-4 py-2 text-gray-400 text-xs">{fa_date(o["created_at"] or "", with_time=True)}</td>
          <td class="px-4 py-2 flex gap-1 items-center">{action_btns}</td>
        </tr>"""

    pager = '<div class="flex gap-2 mt-4 justify-center">' + "".join(
        f'<a href="/admin/orders?page={i}" class="px-3 py-1 rounded border text-sm {"bg-indigo-600 text-white" if i==page else "bg-white"}">{i+1}</a>'
        for i in range(min(pages, 10))
    ) + "</div>" if pages > 1 else ""

    log_rows_html = "".join(_order_log_row_html(r) for r in recent_logs) or '<div class="text-xs text-gray-400 text-center py-6">فعالیتی ثبت نشده</div>'

    body = f"""
    <div class="flex items-center justify-between mb-6 flex-wrap gap-3">
      <h1 class="text-2xl font-bold text-gray-800">🧾 سفارش‌ها ({total:,})</h1>
      <div class="flex items-center gap-2 flex-wrap">
        <a href="/admin/orders/export.xlsx?q={e(q)}" class="btn-sm bg-green-50 text-green-700 border border-green-200 rounded px-3 py-1.5 text-xs">⬇ Excel</a>
        <span class="px-2 py-0.5 text-xs bg-red-100 text-red-700 rounded-full">↩️ برگشتی: {returned_total}</span>
        <form method="get" class="flex gap-2">
          {_input("q","جستجو User ID...",q)} {_btn("جستجو","","slate",True)}
        </form>
      </div>
    </div>
    <div class="card overflow-hidden">
      <div class="overflow-x-auto">
        <table class="w-full text-right min-w-max">
          <thead><tr class="text-xs text-gray-500 border-b bg-gray-50">
            <th class="px-4 py-3">#</th><th class="px-4 py-3">User ID</th>
            <th class="px-4 py-3">محصول</th><th class="px-4 py-3">مبلغ</th>
            <th class="px-4 py-3">وضعیت</th><th class="px-4 py-3">تاریخ</th><th class="px-4 py-3">عملیات</th>
          </tr></thead>
          <tbody>{rows or "<tr><td colspan='7' class='text-center py-8 text-gray-400'>سفارشی ثبت نشده</td></tr>"}</tbody>
        </table>
      </div>
      {pager}
    </div>

    <details open class="card mt-6 overflow-hidden">
      <summary class="p-6 pb-4 cursor-pointer select-none list-none flex items-center justify-between">
        <span class="font-bold text-gray-700">🕓 آخرین فعالیت‌های سفارش‌ها</span>
        <span class="text-xs text-gray-400">برگشت · تعویض · ارسال مجدد · ویرایش — برای بستن/باز کردن کلیک کنید</span>
      </summary>
      <div class="px-6 pb-6 -mt-2">
        <div class="border-t border-gray-100 max-h-[300px] overflow-y-auto">{log_rows_html}</div>
      </div>
    </details>"""

    return _layout("سفارش‌ها", body, adm, flash=flash)


_ORDER_LOG_ICONS = {
    "برگشت سفارش": "↩️", "تعویض کالا": "🔄", "ارسال مجدد": "📦", "ویرایش سفارش": "✏️",
}
_ORDER_LOG_COLORS = {
    "برگشت سفارش": "border-red-300", "تعویض کالا": "border-amber-300",
    "ارسال مجدد": "border-green-300", "ویرایش سفارش": "border-blue-300",
}
_ORDER_LOG_VALUE_LABELS = {
    "restore": "بازگشت به موجودی", "delete": "حذف دائم از فید",
    "none": "بدون تغییر", "full": "بازگشت کامل",
    "custom_add": "افزایش دلخواه", "custom_deduct": "کسر دلخواه", "auto": "خودکار (اختلاف قیمت)",
    "wrong_product": "ارسال محصول اشتباه", "replacement": "جایگزینی محصول",
    "order_fix": "اصلاح سفارش", "customer_request": "درخواست مشتری", "other": "سایر",
}


def _order_log_value_label(v: str) -> str:
    """کدهای داخلی (restore/custom_add/...) رو به برچسب فارسی خوانا ترجمه می‌کنه؛
    مقادیری مثل «custom_add(+1,500)» رو هم پوشش می‌ده (پیشوند شناسایی می‌شه، باقی
    رشته دست‌نخورده می‌مونه)."""
    import re as _re
    m = _re.match(r"^([a-zA-Z_]+)(.*)$", v)
    if m and m.group(1) in _ORDER_LOG_VALUE_LABELS:
        return _ORDER_LOG_VALUE_LABELS[m.group(1)] + m.group(2)
    return v


def _order_log_row_html(row) -> str:
    """رندر خوانا و حرفه‌ای یک ردیف لاگ فعالیت سفارش — رشتهٔ خام details (پایپ‌جدا،
    key:value) به چیپ‌های جدا با برچسب فارسی تبدیل می‌شه؛ شمارهٔ سفارش لینک مستقیم
    به هاب «وضعیت سفارش» می‌گیره (بخش ۴۳ CLAUDE.md)."""
    import re as _re
    action = row["action"]
    icon = _ORDER_LOG_ICONS.get(action, "🕓")
    border = _ORDER_LOG_COLORS.get(action, "border-gray-300")

    parts = [p.strip() for p in (row["details"] or "").split("|") if p.strip()]
    order_ref = ""
    if parts:
        m = _re.match(r"^سفارش\s*#(\d+)(?:→#(\d+))?", parts[0])
        if m:
            oid1, oid2 = m.group(1), m.group(2)
            if oid2:
                order_ref = (f'<a href="/admin/orders/{oid1}" class="font-mono text-indigo-600 hover:underline">#{oid1}</a>'
                             f' <span class="text-gray-300">→</span> '
                             f'<a href="/admin/orders/{oid2}" class="font-mono text-indigo-600 hover:underline">#{oid2}</a>')
            else:
                order_ref = f'<a href="/admin/orders/{oid1}" class="font-mono text-indigo-600 hover:underline">#{oid1}</a>'
            parts = parts[1:]

    chips = []
    for p in parts:
        if ":" in p:
            k, v = p.split(":", 1)
            k, v = k.strip(), v.strip()
            if not k or not v:
                continue
            chips.append(f'<span class="chip inline-block px-1.5 py-0.5 rounded bg-gray-50 border border-gray-200 text-[11px] text-gray-600">{e(k)}: {e(_order_log_value_label(v))}</span>')
        else:
            chips.append(f'<span class="chip inline-block px-1.5 py-0.5 rounded bg-gray-50 border border-gray-200 text-[11px] text-gray-600">{e(p)}</span>')
    chips_html = " ".join(chips)

    return f"""
    <div class="flex items-start gap-3 py-3 border-b border-gray-100 last:border-0 border-r-2 {border} pr-3">
      <span class="text-base leading-none mt-0.5">{icon}</span>
      <div class="flex-1 min-w-0">
        <div class="flex items-center gap-2 flex-wrap text-xs">
          <span class="font-semibold text-gray-700">{e(action)}</span>
          {order_ref}
          <span class="text-gray-300">·</span>
          <span class="text-gray-400">{fa_date(row["created_at"] or "", with_time=True)}</span>
          <span class="text-gray-300">·</span>
          <span class="text-gray-400">{e(row["admin_name"] or "")}</span>
        </div>
        {f'<div class="mt-1.5 flex flex-wrap gap-1">{chips_html}</div>' if chips_html else ""}
      </div>
    </div>"""


def _fetch_order_logs(conn, limit: int = 50):
    """جدول admin_logs قبلاً فقط از داخل ۳ روت خاص (نه اینجا) لیزی ساخته می‌شد — روی
    نصب کاملاً تازه‌ای که هنوز هیچ‌کدوم از اونا صدا زده نشده، این کوئری با «no such
    table» شکست می‌خورد (دقیقاً همون کلاس مشکلی که _log() هم داشت، فقط اونجا با
    try/except بیرونی خاموش قورت داده می‌شد). اینجا صریح گارد شده.

    ⚠️ طبق درخواست صریح مالک پروژه (بخش ۴۳ CLAUDE.md)، تاریخچهٔ فعالیت سفارش‌ها
    فقط در همین یک نقطه (پایین صفحهٔ لیست سفارش‌ها) نمایش داده می‌شه — نسخهٔ
    قبلی این تابع یک شاخهٔ per-order هم داشت (برای کارت تاریخچهٔ هاب «وضعیت
    سفارش») که عمداً حذف شد تا این اطلاعات دوبار/دو-جا نمایش داده نشه.

    ⚠️ رفع کارایی: CREATE TABLE/INDEX قبلاً هر بار (یعنی هر بار صفحهٔ لیست سفارش‌ها
    باز می‌شد) اجرا می‌شد. حالا با فلگ per-process فقط یک‌بار امتحان می‌شه — دقیقاً
    الگوی _INDEXES_DONE در db.py (بخش ۲۴ سند)."""
    global _ADMIN_LOGS_SCHEMA_READY
    try:
        if not _ADMIN_LOGS_SCHEMA_READY:
            conn.execute("""
                CREATE TABLE IF NOT EXISTS admin_logs (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    admin_id   INTEGER,
                    admin_name TEXT,
                    action     TEXT NOT NULL,
                    section    TEXT,
                    details    TEXT,
                    ip         TEXT,
                    created_at TEXT NOT NULL DEFAULT (datetime('now','localtime'))
                );
            """)
            # کوئری این تابع WHERE section=? ORDER BY id DESC می‌زنه — بدون ایندکس
            # با رشد admin_logs (هر اکشن ادمین یک ردیف) هر بار full-scan می‌شه.
            conn.execute("CREATE INDEX IF NOT EXISTS idx_admin_logs_section ON admin_logs(section, id);")
            conn.commit()
            _ADMIN_LOGS_SCHEMA_READY = True
        return conn.execute(
            "SELECT * FROM admin_logs WHERE section=? ORDER BY id DESC LIMIT ?;",
            ("سفارش‌ها", limit)
        ).fetchall()
    except Exception:
        return []


@router.get("/orders/{oid}", response_class=HTMLResponse)
async def order_status_page(request: Request, oid: int, flash: str = ""):
    """🧾 وضعیت سفارش — صفحهٔ یکپارچهٔ برگشت/تعویض/ارسال مجدد. قبلاً این قابلیت‌ها
    روی چند صفحهٔ جدا بودن (بدون هیچ صفحهٔ خلاصه‌ای که با /admin/orders/{oid}
    برسه بهش — همون آدرسی که دکمه‌های «انصراف» صفحات برگشت/تعویض بهش لینک می‌دادن
    ولی هیچ‌وقت وجود نداشت، یعنی همیشه ۴۰۴ می‌گرفتن). طبق درخواست صریح مالک پروژه
    ادغام شدن (بخش ۴۲ CLAUDE.md).

    ⚠️ بازطراحی دوم (بخش ۴۳ CLAUDE.md، طبق درخواست صریح): کارت «اطلاعات سفارش»
    (نمایش خام id/کاربر/مبلغ/موجودی/تاریخ) کاملاً حذف شد — این اطلاعات از قبل
    توی ردیف لیست سفارش‌ها دیده می‌شن، تکرارشون اینجا بی‌معنی بود. کارت تاریخچهٔ
    این سفارش هم حذف شد — لاگ فعالیت فقط یک‌جا (پایین لیست سفارش‌ها) نمایش داده
    می‌شه. «تعویض کالا» و «ارسال کالای جایگزین» در یک تب واحد ادغام شدن (هدف هر
    دو یکیه: دادن یه آیتم متفاوت به کاربر — یا از همون محصول، یا محصولی کاملاً
    دیگه)؛ اگه فقط یکی از این دو واقعاً در دسترس باشه (مثلاً سفارش برگشتی فقط
    ارسال‌جایگزین رو پشتیبانی می‌کنه، نه تعویض)، انتخابگر نوع عملیات اصلاً نشون
    داده نمی‌شه و مستقیم همون فرم می‌آد.

    ⚠️ بازطراحی سوم (بخش ۴۴ CLAUDE.md، طبق درخواست صریح): کارت «ویرایش اطلاعات
    پایه» (عنوان/قیمت سفارش) هم کاملاً حذف شد — اطلاعات پایهٔ محصول جای درستش
    بخش محصول/موجودیه، نه اینجا. روت POST /orders/{oid}/edit هم چون بدون هیچ
    فرمی که بهش لینک بده کد مرده می‌شد، حذف شد."""
    adm = _get_admin(request)
    guard = _require(adm, "orders")
    if guard: return guard

    conn = _db()
    conn.row_factory = sqlite3.Row
    try:
        order = conn.execute("SELECT * FROM orders WHERE id=?;", (oid,)).fetchone()
        if not order:
            return _redir("/admin/orders?flash=سفارش+یافت+نشد")

        resend_items = conn.execute("""
            SELECT pf.id, pf.data FROM product_feed pf
            JOIN products p ON p.id=pf.product_id
            WHERE p.title=? AND pf.delivered=0
            ORDER BY pf.id LIMIT 5;
        """, (order["title"],)).fetchall()

        exch_products = conn.execute("""
            SELECT p.id, p.title, p.price, COUNT(pf.id) AS stock
            FROM products p LEFT JOIN product_feed pf ON pf.product_id=p.id AND pf.delivered=0
            WHERE p.is_active=1
            GROUP BY p.id, p.title, p.price HAVING COUNT(pf.id)>0
            ORDER BY p.title;
        """).fetchall()
    finally:
        conn.close()

    st = order["status"] if "status" in order.keys() and order["status"] else "active"
    is_active = st == "active"
    is_returned = st == "returned"
    price = int(order["price"] or 0)

    status_badge = (
        '<span class="px-2 py-0.5 text-xs rounded-full bg-red-100 text-red-700">برگشتی</span>' if is_returned
        else '<span class="px-2 py-0.5 text-xs rounded-full bg-green-100 text-green-700">فعال</span>'
    )

    # ── تب برگشت ────────────────────────────────────────────────────────────
    return_panel = ""
    if is_active:
        return_panel = f"""
          <form method="post" action="/admin/orders/{oid}/return">
            <div class="mb-3">
              <div class="text-xs text-gray-500 mb-1">تکلیف محصول</div>
              <div class="flex-col-10">
                <label class="perm-label option-card">
                  <input type="radio" name="product_action" value="restore" checked class="option-radio">
                  <div><strong>بازگشت به موجودی</strong><div class="option-hint">محصول مجدداً قابل فروش می‌شود</div></div>
                </label>
                <label class="perm-label option-card">
                  <input type="radio" name="product_action" value="delete" class="option-radio">
                  <div><strong>حذف دائم از فید</strong><div class="option-hint">محصول از چرخه فروش خارج می‌شود</div></div>
                </label>
              </div>
            </div>
            <div class="mb-3">
              <div class="text-xs text-gray-500 mb-1">تکلیف کیف‌پول کاربر</div>
              <div class="flex-col-10">
                <label class="perm-label option-card">
                  <input type="radio" name="wallet_action" value="none" checked class="option-radio">
                  <div><strong>بدون تغییر کیف‌پول</strong></div>
                </label>
                <label class="perm-label option-card option-card--success">
                  <input type="radio" name="wallet_action" value="full" class="option-radio">
                  <div><strong>بازگشت کامل — {price:,} تومان</strong></div>
                </label>
                <label class="perm-label option-card">
                  <input type="radio" name="wallet_action" value="custom_add" class="option-radio">
                  <div><strong>افزایش مبلغ دلخواه</strong></div>
                </label>
                <label class="perm-label option-card option-card--danger">
                  <input type="radio" name="wallet_action" value="custom_deduct" class="option-radio">
                  <div><strong>کسر از کیف‌پول</strong></div>
                </label>
              </div>
              <div class="mt-2">
                <label class="text-xs text-gray-500">مبلغ (تومان) — فقط برای گزینه‌های دلخواه</label>
                {_input("custom_amount", f"مثلاً: {price}", type_="number")}
              </div>
            </div>
            <div class="mb-3">
              <label class="text-xs text-gray-500 block mb-1">علت برگشت</label>
              <select name="reason" class="w-full border border-gray-200 rounded-lg px-3 py-2 text-sm">
                <option value="wrong_product">ارسال محصول اشتباه</option>
                <option value="replacement">جایگزینی محصول</option>
                <option value="order_fix">اصلاح سفارش</option>
                <option value="customer_request">درخواست مشتری</option>
                <option value="other">سایر</option>
              </select>
            </div>
            <div class="mb-3">
              <label class="text-xs text-gray-500 block mb-1">توضیحات اضافه (اختیاری)</label>
              {_input("note", "توضیح بیشتر...")}
            </div>
            <label class="perm-label option-toggle-label mb-3">
              <input type="checkbox" name="notify_user" value="1" checked class="option-check-sm">
              ارسال نوتیف به کاربر
            </label>
            {_btn("ثبت برگشت","",color="red")}
          </form>"""
    else:
        return_panel = '<p class="text-gray-400 text-sm py-6 text-center">این سفارش قبلاً برگشت خورده — برای ارسال کالای جدید از تب «تعویض / ارسال جایگزین» استفاده کنید.</p>'

    # ── تب تعویض / ارسال جایگزین (ادغام‌شده) ────────────────────────────────
    resend_opts = "".join(
        f'<option value="{f["id"]}">{f["id"]} — {str(f["data"] or "")[:50]}</option>'
        for f in resend_items
    )
    has_resend = bool(resend_opts)

    exch_opts = "".join(
        f'<option value="{p["id"]}" data-price="{int(p["price"] or 0)}">{e(p["title"])} — {int(p["price"] or 0):,} تومان (موجودی: {p["stock"]})</option>'
        for p in exch_products
    )
    has_exchange = is_active and bool(exch_opts)

    resend_form = f"""
        <form method="post" action="/admin/orders/{oid}/resend">
          <div class="mb-3">
            <label class="text-xs text-gray-500 block mb-1">انتخاب آیتم از موجودی (همون محصول)</label>
            <select name="feed_id" required class="w-full border border-gray-200 rounded-lg px-3 py-2 text-sm">
              <option value="">انتخاب کنید...</option>
              {resend_opts}
            </select>
          </div>
          <div class="mb-3">
            <div class="text-xs text-gray-500 mb-1">تکلیف کیف‌پول کاربر</div>
            <div class="flex-col-10">
              <label class="perm-label option-card">
                <input type="radio" name="wallet_action" value="none" checked class="option-radio">
                <div><strong>بدون تغییر کیف‌پول</strong></div>
              </label>
              <label class="perm-label option-card">
                <input type="radio" name="wallet_action" value="custom_add" class="option-radio">
                <div><strong>افزایش مبلغ دلخواه</strong></div>
              </label>
              <label class="perm-label option-card option-card--danger">
                <input type="radio" name="wallet_action" value="custom_deduct" class="option-radio">
                <div><strong>کسر از کیف‌پول</strong></div>
              </label>
            </div>
            <div class="mt-2">
              <label class="text-xs text-gray-500">مبلغ (تومان) — فقط برای گزینه‌های دلخواه</label>
              {_input("custom_amount", "مثلاً: 5000", type_="number")}
            </div>
          </div>
          <label class="perm-label option-toggle-label mb-3">
            <input type="checkbox" name="notify_user" value="1" checked class="option-check-sm">
            اطلاع‌رسانی به کاربر
          </label>
          {_btn("ارسال آیتم جایگزین","",color="green")}
        </form>""" if has_resend else '<p class="text-red-500 text-sm">موجودی دیگری از همین محصول در دسترس نیست</p>'

    exchange_form = f"""
        <form method="post" action="/admin/orders/{oid}/exchange">
          <div class="mb-3">
            <div class="text-xs text-gray-500 mb-1">تکلیف محصول قدیم</div>
            <div class="flex-col-10">
              <label class="perm-label option-card">
                <input type="radio" name="old_product_action" value="restore" checked class="option-radio">
                <div><strong>بازگشت به موجودی</strong></div>
              </label>
              <label class="perm-label option-card">
                <input type="radio" name="old_product_action" value="delete" class="option-radio">
                <div><strong>حذف دائم از فید</strong></div>
              </label>
            </div>
          </div>
          <div class="mb-3">
            <div class="text-xs text-gray-500 mb-1">محصول جایگزین</div>
            <select name="new_product_id" id="exch-prod-{oid}" required class="w-full border rounded-lg px-3 py-2 text-sm mb-2" onchange="document.getElementById('exch-diff-{oid}').textContent=(parseInt(this.selectedOptions[0].dataset.price||0)-{price}).toLocaleString('en-US')">
              <option value="">انتخاب محصول جایگزین...</option>
              {exch_opts}
            </select>
            <div class="info-box">اختلاف قیمت (جدید − قدیم): <strong id="exch-diff-{oid}">۰</strong> تومان</div>
          </div>
          <div class="mb-3">
            <div class="text-xs text-gray-500 mb-1">تسویهٔ اختلاف قیمت روی کیف‌پول</div>
            <div class="flex-col-10">
              <label class="perm-label option-card">
                <input type="radio" name="wallet_mode" value="auto" checked class="option-radio">
                <div><strong>خودکار بر اساس اختلاف قیمت</strong></div>
              </label>
              <label class="perm-label option-card">
                <input type="radio" name="wallet_mode" value="none" class="option-radio">
                <div><strong>بدون تغییر کیف‌پول</strong></div>
              </label>
            </div>
          </div>
          {_btn("ثبت تعویض","",color="indigo")}
        </form>""" if has_exchange else ('<p class="text-red-500 text-sm">هیچ محصول دیگری با موجودی فعال نیست</p>' if is_active else "")

    if has_resend and has_exchange:
        swap_panel = f"""
        <div class="mb-4">
          <label class="text-xs text-gray-500 block mb-1">نوع عملیات</label>
          <select id="swap-mode-{oid}" class="w-full border border-gray-200 rounded-lg px-3 py-2 text-sm"
                  onchange="document.getElementById('swap-resend-{oid}').classList.toggle('hidden', this.value!=='resend'); document.getElementById('swap-exchange-{oid}').classList.toggle('hidden', this.value!=='exchange');">
            <option value="">انتخاب کنید...</option>
            <option value="resend">📦 ارسال آیتم دیگر از همین محصول</option>
            <option value="exchange">🔄 تعویض با محصول کاملاً متفاوت</option>
          </select>
        </div>
        <div id="swap-resend-{oid}" class="hidden">{resend_form}</div>
        <div id="swap-exchange-{oid}" class="hidden">{exchange_form}</div>"""
    elif has_resend:
        swap_panel = resend_form
    elif has_exchange:
        swap_panel = exchange_form
    else:
        swap_panel = '<p class="text-gray-400 text-sm py-6 text-center">در حال حاضر نه آیتم جایگزین از همین محصول و نه محصول دیگری با موجودی در دسترس است.</p>'

    return_tab_disabled = "" if is_active else "disabled"
    tabs_card = f"""
    <div class="card card-p mb-6">
      <div class="flex gap-1 border-b border-gray-200 mb-4">
        <button type="button" id="ord-tab-btn-return-{oid}" {return_tab_disabled}
                class="px-4 py-2 text-sm font-semibold border-b-2 {'text-indigo-600 border-indigo-600' if is_active else 'text-gray-300 border-transparent cursor-not-allowed'}"
                onclick="stlOrderTab({oid},'return')">↩️ برگشت</button>
        <button type="button" id="ord-tab-btn-swap-{oid}"
                class="px-4 py-2 text-sm font-semibold border-b-2 {'text-gray-500 border-transparent' if is_active else 'text-indigo-600 border-indigo-600'}"
                onclick="stlOrderTab({oid},'swap')">🔄 تعویض / ارسال جایگزین</button>
      </div>
      <div id="ord-tab-panel-return-{oid}" class="{'hidden' if not is_active else ''}">{return_panel}</div>
      <div id="ord-tab-panel-swap-{oid}" class="{'hidden' if is_active else ''}">{swap_panel}</div>
    </div>
    <script>
    function stlOrderTab(oid, tab) {{
      var tabs = ['return','swap'];
      tabs.forEach(function(t) {{
        var btn = document.getElementById('ord-tab-btn-'+t+'-'+oid);
        var panel = document.getElementById('ord-tab-panel-'+t+'-'+oid);
        if (t === tab) {{
          if (btn) {{ btn.classList.add('text-indigo-600','border-indigo-600'); btn.classList.remove('text-gray-500','text-gray-300','border-transparent'); }}
          if (panel) panel.classList.remove('hidden');
        }} else {{
          if (btn && !btn.disabled) {{ btn.classList.remove('text-indigo-600','border-indigo-600'); btn.classList.add('text-gray-500','border-transparent'); }}
          if (panel) panel.classList.add('hidden');
        }}
      }});
    }}
    </script>"""

    body = f"""
    <a href="/admin/orders" class="text-indigo-600 text-sm mb-4 inline-block">← بازگشت به سفارش‌ها</a>
    <div class="flex items-center gap-3 mb-6 flex-wrap">
      <h1 class="text-xl font-bold text-gray-800">🧾 وضعیت سفارش #{oid}</h1>
      {status_badge}
    </div>
    <div class="max-w-2xl">
      {tabs_card}
    </div>"""
    return _layout(f"وضعیت سفارش #{oid}", body, adm, flash=flash)


@router.post("/orders/{oid}/return")
async def order_return(request: Request, oid: int):
    adm = _get_admin(request)
    guard = _require(adm, "orders")
    if guard: return guard

    form = await request.form()
    product_action = str(form.get("product_action", "restore"))
    wallet_action  = str(form.get("wallet_action", "none"))
    custom_amount  = int(form.get("custom_amount") or 0)
    reason         = str(form.get("reason", "other"))
    note           = str(form.get("note", ""))
    notify_user    = form.get("notify_user") == "1"

    try:
        from db import order_mark_returned_advanced
        result = order_mark_returned_advanced(
            oid,
            product_action=product_action,
            wallet_action=wallet_action,
            custom_amount=custom_amount,
        )
    except Exception as ex:
        _tg_logger.error("order_return error: %s", ex)
        return _redir(f"/admin/orders/{oid}?flash=خطا:+{str(ex)[:50]}")

    if not result.get("ok"):
        return _redir(f"/admin/orders/{oid}?flash={result.get('error','خطا')}")

    # حذف پیام تحویل
    if result.get("chat_id") and result.get("message_id"):
        await run_in_threadpool(_tg_delete_message, result["chat_id"], result["message_id"])

    # نوتیف به کاربر — از wallet_delta واقعی result استفاده می‌کنه (نه بازسازی از
    # wallet_action/custom_amount ورودی فرم) تا هر سه حالت (بازگشت کامل/افزایش دلخواه/
    # کسر دلخواه) درست پوشش داده بشن؛ قبلاً custom_deduct هیچ پیامی نداشت.
    if notify_user and result.get("user_id"):
        delta = result.get("wallet_delta", 0)
        wallet_msg = ""
        if delta > 0:
            wallet_msg = f"\n💰 مبلغ {delta:,} تومان به کیف‌پول شما افزوده شد."
        elif delta < 0:
            wallet_msg = f"\n💳 مبلغ {abs(delta):,} تومان از کیف‌پول شما کسر شد."
        await run_in_threadpool(_tg_send, int(result["user_id"]),
            f"⚠️ سفارش #{oid} (<b>{html.escape(str(result.get('title') or ''))}</b>) "
            f"توسط پشتیبانی برگشت داده شد.{wallet_msg}\n"
            "در صورت سوال با پشتیبانی در تماس باشید.")

    # لاگ کامل
    _log(request, "برگشت سفارش", "سفارش‌ها",
         f"سفارش #{oid} | محصول: {product_action} | کیف‌پول: {wallet_action} | علت: {reason} | {note[:80]}")

    return _redir(f"/admin/orders/{oid}?flash=✅+برگشت+ثبت+شد+—+می‌توانید+کالای+جایگزین+ارسال+کنید")


@router.post("/orders/{oid}/resend")
async def order_resend_post(request: Request, oid: int):
    adm = _get_admin(request)
    guard = _require(adm, "orders")
    if guard: return guard
    form = await request.form()
    feed_id       = int(form.get("feed_id") or 0)
    notify        = form.get("notify_user") == "1"
    wallet_action = str(form.get("wallet_action", "none"))  # none | custom_add | custom_deduct
    custom_amount = int(form.get("custom_amount") or 0)
    if not feed_id:
        return _redir(f"/admin/orders/{oid}?flash=آیتم+انتخاب+نشد")

    conn = _db()
    try:
        order = conn.execute("SELECT * FROM orders WHERE id=?;", (oid,)).fetchone()
        feed  = conn.execute("SELECT * FROM product_feed WHERE id=? AND delivered=0;", (feed_id,)).fetchone()
        if not order or not feed:
            return _redir(f"/admin/orders/{oid}?flash=خطا:+داده+یافت+نشد")

        user_id = int(order["user_id"])
        title   = order["title"]
        data    = feed["data"]

        # علامت‌گذاری feed به عنوان تحویل‌شده — یعنی این آیتم از موجودی کم می‌شه (همون
        # چیزی که با گزینهٔ «بازگشت به موجودی» موقع برگشت قبلی برگردونده بودیم)
        conn.execute("UPDATE product_feed SET delivered=1, order_id=?, delivered_at=datetime('now') WHERE id=?;",
                     (oid, feed_id))
        # آپدیت order با feed_id جدید + علامت resent_at — تا در لیست سفارش‌ها روشن
        # باشه این یه اصلاح/جایگزینیه، نه یه تحویل تازهٔ مستقل (بخش ۴۱ CLAUDE.md)
        conn.execute("UPDATE orders SET feed_id=?, status='active', resent_at=datetime('now') WHERE id=?;", (feed_id, oid))

        # تکلیف کیف‌پول — طبق درخواست صریح مالک پروژه («باید بپرسه که از کاربر پول کم
        # بشه یا نه یا افزایش پیدا کنه یا نه»، بخش ۴۲ CLAUDE.md) — همون سه گزینهٔ فرم
        # برگشت (بدون «بازگشت کامل»، چون ارسال مجدد جایگزینی همون‌قیمته، نه لغو سفارش).
        wallet_delta = 0
        if wallet_action == "custom_add" and custom_amount:
            wallet_delta = abs(custom_amount)
        elif wallet_action == "custom_deduct" and custom_amount:
            wallet_delta = -abs(custom_amount)
        if wallet_delta != 0:
            wrow = conn.execute("SELECT balance FROM wallets WHERE user_id=?;", (user_id,)).fetchone()
            if wrow:
                new_bal = max(0, int(wrow["balance"]) + wallet_delta)
                conn.execute("UPDATE wallets SET balance=?, updated_at=datetime('now') WHERE user_id=?;", (new_bal, user_id))
            else:
                conn.execute("INSERT INTO wallets (user_id, balance, updated_at) VALUES (?,?,datetime('now'));",
                             (user_id, max(0, wallet_delta)))
        conn.commit()
    finally:
        conn.close()

    # ارسال به کاربر — از send_telegram_message_with_id استفاده می‌کنه (نه _tg_send)
    # تا message_id رو هم بگیریم و در delivery_messages ذخیره کنیم؛ وگرنه اگه این
    # سفارش بعداً دوباره برگشت بخوره، پیامِ همین ارسال مجدد قابل حذف از چت نیست
    # (بخش ۴۰ CLAUDE.md).
    if notify:
        try:
            import html as _html
            wallet_msg = ""
            if wallet_delta > 0:
                wallet_msg = f"\n💰 مبلغ {wallet_delta:,} تومان به کیف‌پول شما افزوده شد."
            elif wallet_delta < 0:
                wallet_msg = f"\n💳 مبلغ {abs(wallet_delta):,} تومان از کیف‌پول شما کسر شد."
            from tg_notify import send_telegram_message_with_id
            token = _env("BOT_TOKEN")
            ok, msg_id = await run_in_threadpool(
                send_telegram_message_with_id, token, user_id,
                f"📦 محصول جدید برای سفارش #{oid} ارسال شد:\n\n"
                f"<code>{_html.escape(str(data))}</code>{wallet_msg}",
                "HTML"
            )
            if ok and msg_id:
                conn2 = _db()
                try:
                    conn2.execute("""
                        CREATE TABLE IF NOT EXISTS delivery_messages (
                            feed_id INTEGER PRIMARY KEY, order_id INTEGER,
                            chat_id INTEGER NOT NULL, message_id INTEGER NOT NULL, created_at TEXT NOT NULL
                        );
                    """)
                    conn2.execute(
                        "INSERT INTO delivery_messages (feed_id, order_id, chat_id, message_id, created_at) "
                        "VALUES (?,?,?,?,datetime('now')) ON CONFLICT(feed_id) DO UPDATE SET "
                        "order_id=excluded.order_id, chat_id=excluded.chat_id, "
                        "message_id=excluded.message_id, created_at=excluded.created_at;",
                        (feed_id, oid, user_id, msg_id)
                    )
                    conn2.commit()
                finally:
                    conn2.close()
        except Exception:
            _tg_logger.exception("resend delivery_messages insert failed")

    _log(request, "ارسال مجدد", "سفارش‌ها",
         f"سفارش #{oid} | feed_item:{feed_id} | از موجودی کم شد | کیف‌پول:{wallet_action}({wallet_delta:+,})")
    wallet_note = ""
    if wallet_delta > 0:
        wallet_note = f"+{wallet_delta:,}+ت+به+کیف‌پول+اضافه+شد"
    elif wallet_delta < 0:
        wallet_note = f"+{abs(wallet_delta):,}+ت+از+کیف‌پول+کسر+شد"
    else:
        wallet_note = "+بدون+پرداخت+اضافی"
    return _redir(f"/admin/orders/{oid}?flash=✅+محصول+جدید+ارسال+شد+(آیتم+از+موجودی+کم+شد{wallet_note})")


@router.post("/orders/{oid}/exchange")
async def order_exchange_post(request: Request, oid: int):
    adm = _get_admin(request)
    guard = _require(adm, "orders")
    if guard: return guard
    form = await request.form()
    new_product_id = int(form.get("new_product_id") or 0)
    old_product_action = str(form.get("old_product_action", "restore"))
    wallet_mode = str(form.get("wallet_mode", "auto"))
    if not new_product_id:
        return _redir(f"/admin/orders/{oid}?flash=محصول+جایگزین+انتخاب+نشد")

    from db import exchange_order
    conn = _db()
    try:
        old_order = conn.execute("SELECT price FROM orders WHERE id=?;", (oid,)).fetchone()
        new_product = conn.execute("SELECT price FROM products WHERE id=?;", (new_product_id,)).fetchone()
    finally:
        conn.close()
    if not old_order or not new_product:
        return _redir(f"/admin/orders/{oid}?flash=خطا:+داده+یافت+نشد")

    wallet_delta = 0
    if wallet_mode == "auto":
        wallet_delta = int(old_order["price"] or 0) - int(new_product["price"] or 0)

    result = exchange_order(oid, new_product_id, old_product_action=old_product_action, wallet_delta=wallet_delta)
    if not result.get("ok"):
        return _redir(f"/admin/orders/{oid}?flash={result.get('error','خطا')}")

    if result.get("old_chat_id") and result.get("old_message_id"):
        await run_in_threadpool(_tg_delete_message, result["old_chat_id"], result["old_message_id"])

    if result.get("user_id"):
        delta = result.get("wallet_delta", 0)
        delta_msg = ""
        if delta > 0:
            delta_msg = f"\n💰 مبلغ {delta:,} تومان به کیف‌پول شما بازگردانده شد."
        elif delta < 0:
            delta_msg = f"\n💳 مبلغ {abs(delta):,} تومان بابت اختلاف قیمت از کیف‌پول شما کسر شد."
        await run_in_threadpool(_tg_send, int(result["user_id"]),
            f"🔄 سفارش شما تعویض شد.\n\n"
            f"محصول قبلی: <b>{html.escape(str(result.get('old_title') or ''))}</b>\n"
            f"محصول جدید: <b>{html.escape(str(result.get('new_title') or ''))}</b>\n\n"
            f"<code>{html.escape(str(result.get('new_feed_data') or ''))}</code>{delta_msg}")

    _log(request, "تعویض کالا", "سفارش‌ها",
         f"سفارش #{oid}→#{result.get('new_order_id')} | محصول قدیم:{old_product_action} | اختلاف کیف‌پول:{wallet_delta}",
         admin_info=adm)
    return _redir(f"/admin/orders/{result.get('new_order_id')}?flash=✅+تعویض+ثبت+شد+(سفارش+جدید+#{result.get('new_order_id')})")


# ─────────────────────────── Wallets ───────────────────────────────────────

@router.get("/users/export.xlsx")
async def users_export(request: Request):
    adm = _get_admin(request)
    guard = _require(adm, "users")
    if guard: return guard
    conn = _db()
    try:
        users = conn.execute("""
            SELECT u.user_id, u.username, u.full_name, u.first_seen, u.last_seen,
                   COALESCE(MAX(w.balance),0) AS balance, COUNT(DISTINCT o.id) AS orders
            FROM users u
            LEFT JOIN wallets w ON w.user_id=u.user_id
            LEFT JOIN orders o ON o.user_id = u.user_id
            GROUP BY u.user_id ORDER BY u.last_seen DESC LIMIT 10000;
        """).fetchall()
    finally:
        conn.close()
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from fastapi.responses import Response
        import io
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "users"
        ws.sheet_view.rightToLeft = True
        headers = ["User ID","یوزرنیم","نام","اولین ورود","آخرین فعالیت","خریدها","کیف‌پول"]
        hfill = PatternFill("solid", fgColor="2EC4B6")
        for ci,h in enumerate(headers,1):
            c = ws.cell(1,ci,h); c.fill=hfill; c.font=Font(bold=True,color="FFFFFF")
        for ri,u in enumerate(users,2):
            for ci,v in enumerate([u[0],u[1] or "",u[2] or "",str(u[3] or "")[:10],str(u[4] or "")[:10],u[6] or 0,u[5] or 0],1):
                ws.cell(ri,ci,v)
        buf=io.BytesIO(); wb.save(buf); buf.seek(0)
        return Response(content=buf.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition":"attachment; filename=users.xlsx"})
    except ImportError:
        from fastapi.responses import Response
        lines=["\ufeffUser ID,یوزرنیم,نام,اولین ورود,آخرین فعالیت,خریدها,کیف‌پول"]
        for u in users:
            lines.append(f'{u[0]},{u[1] or ""},{u[2] or ""},{str(u[3] or "")[:10]},{str(u[4] or "")[:10]},{u[6] or 0},{u[5] or 0}')
        return Response(content="\n".join(lines).encode("utf-8"),
            media_type="text/csv;charset=utf-8",
            headers={"Content-Disposition":"attachment; filename=users.csv"})


@router.get("/users/{uid}", response_class=HTMLResponse)
async def user_detail(request: Request, uid: int, flash: str = ""):
    adm = _get_admin(request)
    guard = _require(adm, "users")
    if guard: return guard
    from db import get_user_full, get_user_orders, get_user_tickets, ensure_user_extra_schema
    ensure_user_extra_schema()
    user = get_user_full(uid)
    if not user:
        return _redir("/admin/users?flash=کاربر+یافت+نشد")
    orders  = get_user_orders(uid, 10)
    tickets = get_user_tickets(uid, 10)

    is_partner = user.get("is_partner")
    is_blocked = user.get("is_blocked", 0)
    partner_badge = '<span class="px-2 py-0.5 text-xs bg-amber-100 text-amber-700 rounded-full">🤝 همکار</span>' if is_partner else ''
    status_badge  = '<span class="px-2 py-0.5 text-xs bg-red-100 text-red-700 rounded-full">مسدود</span>' if is_blocked else '<span class="px-2 py-0.5 text-xs bg-green-100 text-green-700 rounded-full">فعال</span>'

    # خریدها
    order_rows = "".join(f"""<tr class="border-b hover:bg-gray-50">
      <td class="px-4 py-2 text-xs text-gray-400">#{o['id']}</td>
      <td class="px-4 py-2 text-sm">{e(o['title'] or '—')}</td>
      <td class="px-4 py-2 text-sm font-medium text-green-600">{int(o['price'] or 0):,} ت</td>
      <td class="px-4 py-2 text-xs text-gray-400">{fa_date(o['created_at'], with_time=True)}</td>
    </tr>""" for o in orders)

    # تیکت‌ها
    ticket_rows = "".join(f"""<tr class="border-b hover:bg-gray-50">
      <td class="px-4 py-2 text-xs"><a href="/admin/tickets/{t['id']}" class="text-indigo-600">#{t['id']}</a></td>
      <td class="px-4 py-2 text-xs">{e((t['type'] if 'type' in t.keys() else '') or 'پشتیبانی')}</td>
      <td class="px-4 py-2 text-xs text-gray-400">{e((t['status'] or '')[:20])}</td>
      <td class="px-4 py-2 text-xs text-gray-400">{fa_date(t['updated_at'] or '', with_time=True)}</td>
    </tr>""" for t in tickets)

    note_val = e(user.get("admin_note", "") or "")
    tags_val = e(user.get("tags", "") or "")

    body = f"""
    <div class="flex items-center gap-3 mb-6">
      {_btn("← کاربران", "/admin/users", "slate", small=True)}
      <h1 class="text-2xl font-bold text-gray-800">{e(user['full_name'] or 'کاربر')}</h1>
      {partner_badge} {status_badge}
    </div>

    <div class="grid md:grid-cols-3 gap-4 mb-6">
      <div class="card p-5 text-center"><div class="text-2xl font-bold text-indigo-600">{user.get('order_count',0)}</div><div class="text-xs text-gray-400 mt-1">تعداد خرید</div></div>
      <div class="card p-5 text-center"><div class="text-2xl font-bold text-green-600">{int(user.get('balance',0)):,}</div><div class="text-xs text-gray-400 mt-1">کیف‌پول (ت)</div></div>
      <div class="card p-5 text-center"><div class="text-2xl font-bold text-gray-700">{len(tickets)}</div><div class="text-xs text-gray-400 mt-1">تیکت‌ها</div></div>
    </div>

    <div class="grid md:grid-cols-2 gap-4 mb-4">
      <!-- اطلاعات پایه -->
      <div class="card p-6">
        <h2 class="font-bold text-gray-700 mb-4">اطلاعات پایه</h2>
        <div class="space-y-2 text-sm">
          <div class="flex justify-between"><span class="text-gray-400">User ID</span><code class="text-xs bg-gray-100 px-2 rounded">{user['user_id']}</code></div>
          <div class="flex justify-between"><span class="text-gray-400">یوزرنیم</span><span>{"@"+e(user['username']) if user['username'] else "—"}</span></div>
          <div class="flex justify-between"><span class="text-gray-400">عضویت</span><span>{fa_date(user['first_seen'] or '')}</span></div>
          <div class="flex justify-between"><span class="text-gray-400">آخرین فعالیت</span><span>{fa_date(user['last_seen'] or '')}</span></div>
        </div>
        <div class="mt-4 pt-4 border-t flex gap-2">
          <a href="/admin/wallets?q={uid}" class="btn-sm bg-indigo-50 text-indigo-700 border border-indigo-200 rounded px-3 py-1.5 text-xs">مدیریت کیف‌پول</a>
          <form method="post" action="/admin/users/{uid}/toggle-block" class="inline">
            <button class="btn-sm {'bg-green-50 text-green-700 border-green-200' if is_blocked else 'bg-red-50 text-red-600 border-red-200'} border rounded px-3 py-1.5 text-xs">
              {'رفع مسدودی' if is_blocked else 'مسدود کردن'}
            </button>
          </form>
        </div>
      </div>

      <!-- یادداشت و برچسب -->
      <div class="card p-6">
        <h2 class="font-bold text-gray-700 mb-4">یادداشت خصوصی مدیر</h2>
        <form method="post" action="/admin/users/{uid}/note">
          <textarea name="admin_note" rows="3" placeholder="یادداشت خصوصی درباره این کاربر..."
            class="w-full border border-gray-200 rounded-lg px-3 py-2 text-sm mb-3">{note_val}</textarea>
          <label class="text-xs text-gray-500 block mb-1">برچسب‌ها (با کاما جدا کنید)</label>
          <input type="text" name="tags" value="{tags_val}" placeholder="VIP، مشکوک، ..."
            class="w-full border border-gray-200 rounded-lg px-3 py-2 text-sm mb-3">
          {_btn("ذخیره یادداشت", color="green")}
        </form>
      </div>
    </div>

    <div class="grid md:grid-cols-2 gap-4">
      <!-- خریدها -->
      <div class="card overflow-hidden">
        <div class="px-5 py-3 border-b bg-gray-50 font-medium text-sm text-gray-700">🛒 خریدهای اخیر</div>
        <div class="overflow-x-auto"><table class="w-full text-right min-w-max">
          <thead><tr class="text-xs text-gray-500 border-b"><th class="px-4 py-2">#</th><th class="px-4 py-2">محصول</th><th class="px-4 py-2">مبلغ</th><th class="px-4 py-2">تاریخ</th></tr></thead>
          <tbody>{order_rows or "<tr><td colspan='4' class='text-center py-4 text-gray-400 text-xs'>خریدی ندارد</td></tr>"}</tbody>
        </table></div>
      </div>

      <!-- تیکت‌ها -->
      <div class="card overflow-hidden">
        <div class="px-5 py-3 border-b bg-gray-50 font-medium text-sm text-gray-700">🎫 تیکت‌ها</div>
        <div class="overflow-x-auto"><table class="w-full text-right min-w-max">
          <thead><tr class="text-xs text-gray-500 border-b"><th class="px-4 py-2">#</th><th class="px-4 py-2">نوع</th><th class="px-4 py-2">وضعیت</th><th class="px-4 py-2">آپدیت</th></tr></thead>
          <tbody>{ticket_rows or "<tr><td colspan='4' class='text-center py-4 text-gray-400 text-xs'>تیکتی ندارد</td></tr>"}</tbody>
        </table></div>
      </div>
    </div>"""
    return _layout(f"کاربر {uid}", body, adm, flash=flash)


@router.post("/users/{uid}/note")
async def user_save_note(request: Request, uid: int):
    adm = _get_admin(request)
    guard = _require(adm, "users")
    if guard: return guard
    form = await request.form()
    note = str(form.get("admin_note", "")).strip()
    tags = str(form.get("tags", "")).strip()
    from db import update_user_note
    update_user_note(uid, note, tags)
    _log(request, "ویرایش یادداشت کاربر", "کاربران", f"user:{uid}")
    return _redir(f"/admin/users/{uid}?flash=یادداشت+ذخیره+شد")


@router.post("/users/{uid}/toggle-block")
async def user_toggle_block(request: Request, uid: int):
    adm = _get_admin(request)
    guard = _require(adm, "users")
    if guard: return guard
    from db import toggle_user_block
    blocked = toggle_user_block(uid)
    _log(request, "مسدود/رفع مسدودی کاربر", "کاربران", f"user:{uid} blocked:{blocked}")
    return _redir(f"/admin/users/{uid}?flash={'کاربر+مسدود+شد' if blocked else 'مسدودی+رفع+شد'}")


@router.get("/users", response_class=HTMLResponse)
async def users_list(request: Request, page: int = 0, q: str = "", sort: str = "last_seen", flash: str = ""):
    adm = _get_admin(request)
    guard = _require(adm, "users")
    if guard: return guard

    PAGE = 50
    conn = _db()
    try:
        sort_col = sort if sort in ("user_id","full_name","first_seen","last_seen","orders","balance") else "last_seen"
        where = "WHERE u.user_id=? OR u.username LIKE ? OR u.full_name LIKE ?" if q else ""
        params = (int(q) if q.isdigit() else 0, f"%{q}%", f"%{q}%") if q else ()
        total = conn.execute(f"SELECT COUNT(*) FROM users u {where};", params).fetchone()[0]
        users = conn.execute(f"""
            SELECT u.*,
                   COALESCE(MAX(w.balance),0) AS balance,
                   COUNT(DISTINCT o.id) AS orders,
                   (SELECT 1 FROM partners p WHERE p.tg_user_id=u.user_id AND p.status='approved' LIMIT 1) AS is_partner
            FROM users u
            LEFT JOIN wallets w ON w.user_id=u.user_id
            LEFT JOIN orders o ON o.user_id = u.user_id
            {where}
            GROUP BY u.user_id
            ORDER BY {sort_col} DESC
            LIMIT ? OFFSET ?;
        """, params+(PAGE, page*PAGE)).fetchall()
    finally:
        conn.close()

    pages = max((total+PAGE-1)//PAGE, 1)

    def sort_link(col, label):
        active = sort == col
        cls = "sort-link sort-link--active" if active else "sort-link"
        return f'<a href="?q={e(q)}&sort={col}" class="{cls}">{label} {"↓" if active else ""}</a>'

    rows = ""
    for u in users:
        try: is_partner = u["is_partner"]
        except: is_partner = None
        partner_badge = ' <span class="px-1.5 py-0.5 text-xs bg-amber-100 text-amber-700 rounded">همکار</span>' if is_partner else ""

        try: blocked = u["is_blocked"]
        except: blocked = 0
        if blocked:
            status_badge = '<span class="px-2 py-0.5 text-xs bg-red-100 text-red-700 rounded-full">مسدود</span>'
        else:
            status_badge = '<span class="px-2 py-0.5 text-xs bg-green-100 text-green-700 rounded-full">فعال</span>'

        rows += f"""<tr class="border-b hover:bg-gray-50">
          <td class="px-4 py-3"><code class="text-xs bg-gray-100 px-1.5 rounded">{u["user_id"]}</code></td>
          <td class="px-4 py-3 text-sm font-medium text-gray-800">{e(u["full_name"] or "—")}{partner_badge}</td>
          <td class="px-4 py-3 text-xs text-gray-400">{"@"+e(u["username"]) if u["username"] else "—"}</td>
          <td class="px-4 py-3 text-xs text-gray-400">{fa_date(u["first_seen"] or "")}</td>
          <td class="px-4 py-3 text-xs text-gray-400">{fa_date(u["last_seen"] or "")}</td>
          <td class="px-4 py-3 text-sm font-bold text-gray-700">{u["orders"] or 0}</td>
          <td class="px-4 py-3 text-sm font-bold text-green-600">{int(u["balance"] or 0):,}</td>
          <td class="px-4 py-3">{status_badge}</td>
          <td class="px-4 py-3"><a href="/admin/users/{u['user_id']}" class="btn-sm bg-indigo-50 text-indigo-700 border border-indigo-200 rounded px-2 py-1 text-xs">پروفایل</a></td>
        </tr>"""

    body = f"""
    <div class="flex items-center justify-between mb-6">
      <h1 class="text-2xl font-bold text-gray-800">👤 کاربران ({total:,})</h1>
      <a href="/admin/users/export.xlsx" class="btn-sm bg-green-50 text-green-700 border border-green-200 rounded px-3 py-1.5 text-xs">⬇ Excel</a>
    </div>
    <div class="card p-4 mb-4">
      <form method="get" class="flex gap-3">
        {_input("q","جستجو User ID یا نام...",value=q)}
        <button type="submit" class="btn-sm bg-indigo-600 text-white rounded px-4 py-2 text-sm shrink-0">جستجو</button>
        {"<a href='/admin/users' class='btn-sm bg-gray-100 text-gray-600 border border-gray-200 rounded px-3 py-2 text-sm'>پاک</a>" if q else ""}
      </form>
    </div>
    <div class="card overflow-hidden">
      <div class="overflow-x-auto">
        <table class="w-full text-right min-w-max">
          <thead><tr class="text-xs text-gray-500 border-b bg-gray-50">
            <th class="px-4 py-3">{sort_link("user_id","ID")}</th>
            <th class="px-4 py-3">نام</th>
            <th class="px-4 py-3">یوزرنیم</th>
            <th class="px-4 py-3">{sort_link("first_seen","عضویت")}</th>
            <th class="px-4 py-3">{sort_link("last_seen","آخرین فعالیت")}</th>
            <th class="px-4 py-3">{sort_link("orders","خریدها")}</th>
            <th class="px-4 py-3">{sort_link("balance","کیف‌پول")}</th>
            <th class="px-4 py-3">وضعیت</th>
            <th class="px-4 py-3"></th>
          </tr></thead>
          <tbody>{rows or "<tr><td colspan='9' class='text-center py-8 text-gray-400'>کاربری یافت نشد</td></tr>"}</tbody>
        </table>
      </div>
      {'<div class="flex gap-2 mt-4 justify-center p-4">' + "".join(
          f'<a href="/admin/users?page={i}&q={e(q)}&sort={sort}" class="px-3 py-1 rounded border text-sm {"bg-indigo-600 text-white" if i==page else "bg-white"}">{i+1}</a>'
          for i in range(min(pages, 10))
      ) + '</div>' if pages > 1 else ''}
    </div>"""
    return _layout("کاربران", body, adm, flash=flash)


@router.get("/wallets", response_class=HTMLResponse)
async def wallets_list(request: Request, q: str="", flash: str=""):
    adm = _get_admin(request)
    guard = _require(adm, "wallets")
    if guard: return guard

    conn = _db()
    try:
        where = "WHERE user_id=?" if (q and q.isdigit()) else ""
        params = (int(q),) if (q and q.isdigit()) else ()
        wallets = conn.execute(f"SELECT * FROM wallets {where} ORDER BY balance DESC LIMIT 50;", params).fetchall()
        totals  = conn.execute("SELECT COUNT(*), COALESCE(SUM(balance),0) FROM wallets;").fetchone()
    finally:
        conn.close()

    rows = "".join(f"""
        <tr class="border-b hover:bg-gray-50 text-sm">
          <td class="px-4 py-2 font-mono text-xs"><code>{w["user_id"]}</code></td>
          <td class="px-4 py-2 font-bold text-{"green" if int(w["balance"])>0 else "gray"}-700">{int(w["balance"]):,} ت</td>
          <td class="px-4 py-2 text-gray-400 text-xs">{fa_date(w["updated_at"] or "", with_time=True)}</td>
          <td class="px-4 py-2">
            <details class="inline-block">
              <summary class="cursor-pointer px-2 py-1 text-xs bg-indigo-50 text-indigo-700 rounded hover:bg-indigo-100 list-none">✏️ ویرایش موجودی</summary>
              <form method="post" action="/admin/wallets/adjust" class="flex gap-2 items-end mt-2 p-2 bg-gray-50 rounded-lg">
                <input type="hidden" name="uid" value="{w["user_id"]}">
                <input type="number" name="amount" placeholder="مبلغ" required class="w-24 border border-gray-300 rounded px-2 py-1">
                <select name="op" class="border border-gray-300 rounded px-2 py-1">
                  <option value="add">➕ افزودن</option>
                  <option value="sub">➖ کاهش</option>
                  <option value="set">✏️ تنظیم</option>
                </select>
                <button class="px-3 py-1 bg-indigo-600 text-white rounded text-xs">ثبت</button>
              </form>
            </details>
            <a href="/admin/wallets/{w["user_id"]}/history"
              class="px-2 py-1 text-xs bg-teal-50 text-teal-700 border border-teal-200 rounded hover:bg-teal-100 mr-1">🧾 تاریخچه شارژ</a>
          </td>
        </tr>""" for w in wallets)

    body = f"""
    <div class="flex items-center justify-between mb-6">
      <h1 class="text-2xl font-bold text-gray-800">💰 کیف‌پول‌ها</h1>
      <span class="text-sm text-gray-500">{int(totals[0])} کاربر — جمع: {int(totals[1]):,} تومان</span>
    </div>
    <div class="card p-6 mb-6">
      <h2 class="font-bold text-gray-700 mb-4">تنظیم موجودی</h2>
      <form method="post" action="/admin/wallets/adjust" class="flex gap-3 flex-wrap items-end">
        <div><label class="text-xs text-gray-500 block mb-1">User ID</label>{_input("uid","",type_="number",required=True)}</div>
        <div><label class="text-xs text-gray-500 block mb-1">مبلغ</label>{_input("amount","",type_="number",required=True)}</div>
        <div><label class="text-xs text-gray-500 block mb-1">عملیات</label>
          <select name="op">
            <option value="add">➕ افزودن</option>
            <option value="sub">➖ کاهش</option>
            <option value="set">✏️ تنظیم مستقیم</option>
          </select>
        </div>
        {_btn("اعمال")}
      </form>
    </div>
    <div class="card overflow-hidden">
      <div class="px-5 py-3 border-b bg-gray-50 flex gap-2">
        <form method="get" class="flex gap-2">
          {_input("q","جستجو User ID...",q)} {_btn("جستجو","","slate",True)}
        </form>
      </div>
      <div class="overflow-x-auto">
        <table class="w-full text-right min-w-max">
          <thead><tr class="text-xs text-gray-500 border-b bg-gray-50">
            <th class="px-4 py-3">User ID</th><th class="px-4 py-3">موجودی</th>
            <th class="px-4 py-3">آپدیت</th><th class="px-4 py-3">عملیات</th>
          </tr></thead>
          <tbody>{rows or "<tr><td colspan='4' class='text-center py-8 text-gray-400'>کاربری یافت نشد</td></tr>"}</tbody>
        </table>
      </div>
    </div>"""

    return _layout("کیف‌پول‌ها", body, adm, flash=flash)


@router.get("/wallets/{uid}/history", response_class=HTMLResponse)
async def wallet_charge_history(request: Request, uid: int):
    """تاریخچه شارژ کاربر — تاریخ، مبلغ، روش (درگاه / کارت به کارت)"""
    adm = _get_admin(request)
    guard = _require(adm, "wallets")
    if guard: return guard

    charges = []
    conn = _db()
    try:
        # شارژهای درگاه (زرین‌پال) — فقط پرداخت‌شده‌ها
        try:
            for r in conn.execute("""
                SELECT amount, COALESCE(paid_at, created_at) AS dt, ref_id
                FROM zarinpal_transactions
                WHERE user_id=? AND status IN ('paid','OK','success','verified')
                  AND COALESCE(payment_type,'wallet')='wallet'
                ORDER BY id DESC LIMIT 100;
            """, (uid,)).fetchall():
                charges.append({"dt": r["dt"] or "", "amount": int(r["amount"] or 0),
                                "method": "🌐 درگاه پرداخت", "ref": r["ref_id"] or "—"})
        except Exception:
            pass
        # شارژهای کارت به کارت — تأییدشده‌ها
        try:
            for r in conn.execute("""
                SELECT amount, COALESCE(updated_at, created_at) AS dt
                FROM card_receipts
                WHERE user_id=? AND status='approved'
                ORDER BY id DESC LIMIT 100;
            """, (uid,)).fetchall():
                charges.append({"dt": r["dt"] or "", "amount": int(r["amount"] or 0),
                                "method": "💳 کارت به کارت", "ref": "—"})
        except Exception:
            pass
        # موجودی فعلی
        w = conn.execute("SELECT balance FROM wallets WHERE user_id=?;", (uid,)).fetchone()
        balance = int(w["balance"]) if w else 0
    finally:
        conn.close()

    charges.sort(key=lambda c: c["dt"], reverse=True)
    total_charged = sum(c["amount"] for c in charges)

    rows = "".join(f"""<tr class="border-b hover:bg-gray-50 text-sm">
        <td class="px-4 py-2.5 text-xs text-gray-500">{fa_date(c['dt'], with_time=True)}</td>
        <td class="px-4 py-2.5 font-bold text-green-700">{c['amount']:,} ت</td>
        <td class="px-4 py-2.5">{c['method']}</td>
        <td class="px-4 py-2.5 text-xs text-gray-400"><code>{e(str(c['ref']))}</code></td>
      </tr>""" for c in charges)

    body = f"""
    <div class="flex items-center gap-3 mb-6 flex-wrap">
      {_btn("← کیف‌پول‌ها", "/admin/wallets", "slate", small=True)}
      <h1 class="text-xl font-bold text-gray-800">🧾 تاریخچه شارژ — کاربر <code>{uid}</code></h1>
    </div>
    <div class="grid grid-cols-3 gap-4 mb-6">
      {_card("موجودی فعلی", f"{balance:,} ت", "", "indigo")}
      {_card("تعداد شارژ", f"{len(charges)}", "", "teal")}
      {_card("جمع شارژها", f"{total_charged:,} ت", "", "green")}
    </div>
    <div class="card overflow-hidden">
      <div class="overflow-x-auto"><table class="w-full text-right min-w-max">
        <thead><tr class="text-xs text-gray-500 border-b bg-gray-50">
          <th class="px-4 py-3">تاریخ</th><th class="px-4 py-3">مبلغ</th>
          <th class="px-4 py-3">روش پرداخت</th><th class="px-4 py-3">کد پیگیری</th>
        </tr></thead>
        <tbody>{rows or "<tr><td colspan='4' class='text-center py-8 text-gray-400'>شارژی ثبت نشده</td></tr>"}</tbody>
      </table></div>
    </div>"""

    return _layout("تاریخچه شارژ", body, adm)

@router.post("/wallets/adjust")
async def wallet_adjust(request: Request, uid: str=Form(""), amount: str=Form("0"), op: str=Form("add")):
    adm = _get_admin(request)
    guard = _require(adm, "wallets")
    if guard: return guard
    try:
        user_id = int(uid); amt = int(amount)
    except ValueError:
        return _redir("/admin/wallets?flash=مقادیر+نامعتبر")
    now = datetime.utcnow().isoformat()
    conn = _db()
    try:
        # جدول لاگ تراکنش‌های دستی
        conn.execute("""
            CREATE TABLE IF NOT EXISTS wallet_admin_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER, op TEXT, amount INTEGER,
                old_balance INTEGER, new_balance INTEGER,
                admin_id TEXT, created_at TEXT
            );
        """)
        row = conn.execute("SELECT balance FROM wallets WHERE user_id=?;", (user_id,)).fetchone()
        cur = int(row["balance"] if row else 0)
        new_bal = cur+amt if op=="add" else max(0,cur-amt) if op=="sub" else amt
        conn.execute("INSERT INTO wallets (user_id,balance,updated_at) VALUES (?,?,?) "
                     "ON CONFLICT(user_id) DO UPDATE SET balance=excluded.balance, updated_at=excluded.updated_at;",
                     (user_id, new_bal, now))
        # ثبت تراکنش در سوابق
        admin_id = adm[0] if adm else "?"
        conn.execute(
            "INSERT INTO wallet_admin_log (user_id, op, amount, old_balance, new_balance, admin_id, created_at) "
            "VALUES (?,?,?,?,?,?,?);",
            (user_id, op, amt, cur, new_bal, str(admin_id), now)
        )
        conn.commit()
    finally:
        conn.close()

    # اطلاع به کاربر
    op_label = {"add": "افزایش", "sub": "کاهش", "set": "تنظیم"}.get(op, op)
    _log(request, "تعدیل دستی کیف‌پول", "کیف‌پول‌ها",
         f"کاربر {user_id}: {op_label} {amt:,} تومان ({cur:,} ← {new_bal:,})", admin_info=adm)
    try:
        await run_in_threadpool(
            _tg_send,
            user_id,
            f"💰 موجودی کیف‌پول شما توسط پشتیبانی {op_label} یافت.\n"
            f"موجودی فعلی: <b>{new_bal:,}</b> تومان"
        )
    except Exception:
        pass

    return _redir(f"/admin/wallets?flash=موجودی+{user_id}+به+{new_bal:,}+تومان+تنظیم+شد")

# ─────────────────────────── Telegram Helper ───────────────────────────────

import logging as _logging
_tg_logger = _logging.getLogger("admin_panel.tg")


def _tg_send(chat_id: int, text: str, parse_mode: str = "HTML",
              reply_markup: dict | None = None) -> bool:
    token = _env("BOT_TOKEN")
    if not token:
        _tg_logger.error("BOT_TOKEN not set — cannot send Telegram message")
        return False
    try:
        data: dict = {"chat_id": int(chat_id), "text": text, "parse_mode": parse_mode}
        if reply_markup:
            data["reply_markup"] = json.dumps(reply_markup)
        r = _requests.post(
            f"https://api.telegram.org/bot{token}/sendMessage",
            json=data, timeout=15
        )
        if not r.ok:
            _tg_logger.error("Telegram sendMessage failed: %s %s", r.status_code, r.text[:200])
        return r.ok
    except Exception as ex:
        _tg_logger.exception("_tg_send error: %s", ex)
        return False


def _tg_send_photo(chat_id: int, photo_url: str, caption: str = "",
                    reply_markup: dict | None = None) -> bool:
    token = _env("BOT_TOKEN")
    if not token:
        return False
    try:
        data: dict = {"chat_id": chat_id, "photo": photo_url, "caption": caption, "parse_mode": "HTML"}
        if reply_markup:
            data["reply_markup"] = json.dumps(reply_markup)
        r = _requests.post(
            f"https://api.telegram.org/bot{token}/sendPhoto",
            json=data, timeout=15
        )
        return r.ok
    except Exception:
        return False


def _tg_send_document(chat_id: int, file_bytes: bytes, filename: str, caption: str = "",
                       as_photo: bool = False, reply_markup: dict | None = None) -> dict:
    """آپلود مستقیم فایل/عکس (بایت خام، نه URL) به تلگرام — برای ارسال پیوست ادمین از پنل.
    برمی‌گردونه {"ok": bool, "file_id": str|None} — file_id برای ذخیره در ticket_messages لازمه
    تا بعداً همون الگوی پروکسی/دانلود موجود (بخش ۳۴ CLAUDE.md) روش کار کنه."""
    token = _env("BOT_TOKEN")
    if not token:
        return {"ok": False, "file_id": None}
    method = "sendPhoto" if as_photo else "sendDocument"
    field = "photo" if as_photo else "document"
    try:
        data = {"chat_id": chat_id, "caption": caption, "parse_mode": "HTML"}
        if reply_markup:
            data["reply_markup"] = json.dumps(reply_markup)
        r = _requests.post(
            f"https://api.telegram.org/bot{token}/{method}",
            data=data,
            files={field: (filename, file_bytes)},
            timeout=30,
        )
        j = r.json()
        if not j.get("ok"):
            _tg_logger.error("Telegram %s failed: %s", method, str(j)[:300])
            return {"ok": False, "file_id": None}
        result = j.get("result", {})
        if as_photo:
            photos = result.get("photo") or []
            fid = photos[-1]["file_id"] if photos else None
        else:
            fid = (result.get("document") or {}).get("file_id")
        return {"ok": True, "file_id": fid}
    except Exception as ex:
        _tg_logger.exception("_tg_send_document error: %s", ex)
        return {"ok": False, "file_id": None}


def _tg_delete_message(chat_id: int, message_id: int) -> bool:
    """حذف یک پیام از چت کاربر (برای برگشت محصول)."""
    token = _env("BOT_TOKEN")
    if not token or not chat_id or not message_id:
        return False
    try:
        r = _requests.post(
            f"https://api.telegram.org/bot{token}/deleteMessage",
            json={"chat_id": int(chat_id), "message_id": int(message_id)}, timeout=15
        )
        return r.ok
    except Exception as ex:
        _tg_logger.error("_tg_delete_message error: %s", ex)
        return False


# ─────────────────────────── Tickets ───────────────────────────────────────

def _ticket_status_badge(status: str) -> str:
    colors = {
        "waiting_admin": "red",
        "waiting_user":  "yellow",
        "closed":        "gray",
        # backward compat
        "open": "green", "in_progress": "yellow",
    }
    labels = {
        "waiting_admin": "منتظر ادمین",
        "waiting_user":  "منتظر کاربر",
        "closed":        "بسته",
        "open": "باز", "in_progress": "در بررسی",
    }
    c = colors.get(status, "slate")
    l = labels.get(status, status)
    return f'<span class="px-2 py-0.5 text-xs rounded-full bg-{c}-100 text-{c}-700">{l}</span>'


@router.get("/logs", response_class=HTMLResponse)
async def admin_logs_page(request: Request, q: str = "", section: str = "", admin_name: str = "", page: int = 0, flash: str = ""):
    adm = _get_admin(request)
    guard = _require(adm, "logs")
    if guard: return guard
    _ensure_theme_table()

    PER_PAGE = 50
    conn = _db()
    try:
        sections = [r[0] for r in conn.execute("SELECT DISTINCT section FROM admin_logs WHERE section!='' ORDER BY section;").fetchall()]
        admin_names = [r[0] for r in conn.execute("SELECT DISTINCT admin_name FROM admin_logs WHERE admin_name IS NOT NULL ORDER BY admin_name;").fetchall()]

        wheres, params = [], []
        if q:
            wheres.append("(admin_name LIKE ? OR action LIKE ? OR details LIKE ?)")
            params += [f"%{q}%", f"%{q}%", f"%{q}%"]
        if section:
            wheres.append("section=?"); params.append(section)
        if admin_name:
            wheres.append("admin_name=?"); params.append(admin_name)
        where_sql = ("WHERE " + " AND ".join(wheres)) if wheres else ""

        total = conn.execute(f"SELECT COUNT(*) FROM admin_logs {where_sql};", params).fetchone()[0]
        logs = conn.execute(
            f"SELECT * FROM admin_logs {where_sql} ORDER BY id DESC LIMIT ? OFFSET ?;",
            params + [PER_PAGE, page * PER_PAGE]
        ).fetchall()
    finally:
        conn.close()

    pages = max(1, (total + PER_PAGE - 1) // PER_PAGE)

    def action_badge(a):
        danger_words = ("حذف", "ناموفق", "ریست", "خروج", "بازیابی")
        warn_words = ("تغییر", "ویرایش", "غیرفعال")
        if any(w in a for w in danger_words): return "badge badge-danger"
        if any(w in a for w in warn_words): return "badge badge-warning"
        return "badge badge-success"

    def log_badge(a):
        if any(w in a for w in ("حذف","ناموفق","ریست")): return f'<span class="px-2 py-0.5 text-xs bg-red-100 text-red-700 rounded-full">{e(a)}</span>'
        if any(w in a for w in ("تغییر","ویرایش","غیرفعال")): return f'<span class="px-2 py-0.5 text-xs bg-yellow-100 text-yellow-700 rounded-full">{e(a)}</span>'
        return f'<span class="px-2 py-0.5 text-xs bg-green-100 text-green-700 rounded-full">{e(a)}</span>'
    rows = "".join(f"""<tr class="border-b hover:bg-gray-50">
      <td class="px-4 py-3 text-xs text-gray-400">#{l["id"]}</td>
      <td class="px-4 py-3 text-sm font-medium text-gray-800">{e(l["admin_name"] or "—")}</td>
      <td class="px-4 py-3">{log_badge(l["action"])}</td>
      <td class="px-4 py-3 text-xs text-gray-500">{e(l["section"] or "—")}</td>
      <td class="px-4 py-3 text-xs text-gray-500 max-w-xs truncate" title="{e(l['details'] or '')}">{e((l["details"] or "")[:60])}</td>
      <td class="px-4 py-3 text-xs font-mono text-gray-400">{e(l["ip"] or "—")}</td>
      <td class="px-4 py-3 text-xs text-gray-400">{fa_date(l["created_at"] or "", with_time=True)}</td>
    </tr>""" for l in logs)

    section_opts = "<option value=''>همه بخش‌ها</option>" + "".join(
        f'<option value="{e(s)}" {"selected" if section==s else ""}>{e(s)}</option>' for s in sections
    )
    admin_opts = "<option value=''>همه ادمین‌ها</option>" + "".join(
        f'<option value="{e(n)}" {"selected" if admin_name==n else ""}>{e(n)}</option>' for n in admin_names
    )

    pagination = ""
    if pages > 1:
        pagination = '<div class="pagination">'
        for i in range(pages):
            active = i == page
            cls = "page-link page-link--active" if active else "page-link"
            pagination += f'<a href="?q={e(q)}&section={e(section)}&admin_name={e(admin_name)}&page={i}" class="{cls}">{i+1}</a>'
        pagination += "</div>"

    body = f"""
    <div class="flex items-center justify-between mb-6">
      <div>
        <h1 class="text-2xl font-bold text-gray-800">📋 گزارش فعالیت</h1>
        <p class="text-xs text-gray-400 mt-1">{total:,} رویداد ثبت‌شده</p>
      </div>
    </div>
    <div class="card p-4 mb-4">
      <form method="get" class="flex flex-wrap gap-3 items-end">
        <div class="flex-1 min-w-40"><label class="text-xs text-gray-500 block mb-1">جستجو</label>{_input("q","عملیات، جزئیات...",value=q)}</div>
        <div class="min-w-36"><label class="text-xs text-gray-500 block mb-1">ادمین</label><select name="admin_name">{admin_opts}</select></div>
        <div class="min-w-36"><label class="text-xs text-gray-500 block mb-1">بخش</label><select name="section">{section_opts}</select></div>
        <button type="submit" class="btn-sm bg-indigo-600 text-white rounded px-4 py-2 text-sm">فیلتر</button>
        <a href="/admin/logs" class="btn-sm bg-gray-100 text-gray-600 border border-gray-200 rounded px-3 py-2 text-sm">پاک</a>
      </form>
    </div>
    <div class="card overflow-hidden">
      <div class="overflow-x-auto">
        <table class="w-full text-right min-w-max">
          <thead><tr class="text-xs text-gray-500 border-b bg-gray-50">
            <th class="px-4 py-3">#</th><th class="px-4 py-3">ادمین</th>
            <th class="px-4 py-3">عملیات</th><th class="px-4 py-3">بخش</th>
            <th class="px-4 py-3">جزئیات</th><th class="px-4 py-3">IP</th>
            <th class="px-4 py-3">زمان</th>
          </tr></thead>
          <tbody>{rows or "<tr><td colspan='7' class='text-center py-8 text-gray-400'>رویدادی ثبت نشده</td></tr>"}</tbody>
        </table>
      </div>
    </div>
    {pagination}"""
    return _layout("گزارش فعالیت", body, adm, flash=flash)


def _partner_payouts_section_html(q: str, sort: str, link_fn) -> str:
    """بخش «تسویه همکاران» — ادغام‌شده در صفحهٔ «خرید و بخش مالی» (طبق درخواست صریح
    مالک پروژه: مرکز مالی که قبلاً زیر تیکت‌ها Embed می‌شد حذف شد و بخش تسویهٔ همکار
    (تنها قسمتی که به‌جز کارت‌به‌کارت داشت) اینجا منتقل شد — بدون از دست رفتن قابلیت.
    link_fn(q, sort) -> URL برای لینک‌های مرتب‌سازی/جستجو
    """
    conn = _db(); conn.row_factory = sqlite3.Row
    try:
        payouts = conn.execute("""
            SELECT p.id, p.user_id, p.amount, p.status, p.created_at,
                   u.full_name, u.username
            FROM partner_payouts p LEFT JOIN users u ON u.user_id = p.user_id
            ORDER BY p.id DESC;
        """).fetchall()
    except Exception:
        payouts = []
    finally:
        conn.close()

    rows = [{
        "id": p["id"], "user_id": p["user_id"],
        "user_name": p["full_name"] or p["username"] or str(p["user_id"]),
        "amount": int(p["amount"] or 0), "status": p["status"],
        "created_at": p["created_at"] or "",
    } for p in payouts]

    if q:
        ql = q.strip().lower()
        rows = [r for r in rows if
                ql in str(r["user_name"]).lower() or ql in str(r["user_id"])
                or ql in str(r["id"]) or ql in str(r["amount"])
                or ql in str(r["status"]).lower()]

    if sort == "date_asc":
        rows.sort(key=lambda r: r["created_at"])
    elif sort == "amount_desc":
        rows.sort(key=lambda r: -r["amount"])
    elif sort == "amount_asc":
        rows.sort(key=lambda r: r["amount"])
    else:
        rows.sort(key=lambda r: r["created_at"], reverse=True)

    status_map = {
        "pending":  ("⏳ جدید",     "bg-amber-100 text-amber-700"),
        "approved": ("✅ تأیید شد", "bg-green-100 text-green-700"),
        "rejected": ("❌ رد شد",    "bg-red-100 text-red-600"),
    }
    pending_count = sum(1 for r in rows if r["status"] == "pending")

    def sort_link(key, label):
        active = sort == key
        return f'<a href="{link_fn(q, key)}" class="text-xs {"text-indigo-600 font-bold" if active else "text-gray-400"}">{label}</a>'

    rows_html = "".join(f"""<tr class="border-b hover:bg-gray-50 text-sm">
      <td class="px-3 py-3 font-medium">{e(str(r['user_name']))}</td>
      <td class="px-3 py-3 text-xs text-gray-400"><code>{r['user_id']}</code></td>
      <td class="px-3 py-3 font-bold text-green-600">{r['amount']:,}</td>
      <td class="px-3 py-3"><span class="px-2 py-0.5 rounded text-xs {status_map.get(r['status'], (r['status'], 'bg-gray-100 text-gray-600'))[1]}">{status_map.get(r['status'], (r['status'], ''))[0]}</span></td>
      <td class="px-3 py-3 text-xs text-gray-400">{fa_date(r['created_at'], with_time=True)}</td>
      <td class="px-3 py-3"><a href="/admin/partners/payout/{r['id']}" class="px-2 py-1 bg-indigo-50 text-indigo-700 rounded text-xs">مشاهده و رسیدگی</a></td>
    </tr>""" for r in rows) or "<tr><td colspan='6' class='text-center py-8 text-gray-400'>درخواست تسویه‌ای یافت نشد</td></tr>"

    return f"""
    <div id="payouts" class="scroll-anchor mt-8">
      <div class="flex items-center justify-between mb-4 flex-wrap gap-2">
        <h2 class="text-xl font-bold text-gray-800">💰 تسویه همکاران</h2>
        {f'<span class="px-3 py-1 bg-amber-100 text-amber-700 rounded-full text-xs font-bold">{pending_count} در انتظار رسیدگی</span>' if pending_count else ''}
      </div>
      <form method="get" class="flex gap-2 mb-4">
        <input type="hidden" name="pay_sort" value="{sort}">
        <input type="text" name="pay_q" value="{e(q)}" placeholder="جستجو: نام، آیدی، مبلغ، وضعیت..."
          class="flex-1 border border-gray-200 rounded-lg px-3 py-2 text-sm">
        <button class="px-4 py-2 bg-indigo-600 text-white rounded-lg text-sm">جستجو</button>
      </form>
      <div class="card overflow-hidden"><div class="overflow-x-auto">
        <table class="w-full text-right min-w-max">
          <thead><tr class="text-xs text-gray-500 border-b bg-gray-50">
            <th class="px-3 py-2">کاربر</th>
            <th class="px-3 py-2">ID</th>
            <th class="px-3 py-2">{sort_link('amount_desc' if sort!='amount_desc' else 'amount_asc','مبلغ ↕')}</th>
            <th class="px-3 py-2">وضعیت</th>
            <th class="px-3 py-2">{sort_link('date_asc' if sort=='date_desc' else 'date_desc','تاریخ ↕')}</th>
            <th class="px-3 py-2">عملیات</th>
          </tr></thead>
          <tbody>{rows_html}</tbody>
        </table>
      </div></div>
    </div>"""


@router.get("/financial", response_class=HTMLResponse)
async def financial_queue(request: Request):
    """آدرس قدیمی مرکز مالی — طبق درخواست صریح مالک پروژه محتوایش با «خرید و بخش
    مالی» ادغام شد؛ این مسیر فقط برای لینک/بوکمارک‌های قدیمی نگه داشته شده."""
    return _redir("/admin/receipts")


@router.get("/tickets", response_class=HTMLResponse)
async def tickets_list(request: Request, status_filter: str = "", type_filter: str = "",
                       show_archived: str = "0", page: int = 0, flash: str = ""):
    adm = _get_admin(request)
    if not adm:
        return _redir("/admin/login")
    guard = _require(adm, "tickets")
    if guard: return guard

    from db import ensure_ticket_archive_schema
    ensure_ticket_archive_schema()

    PAGE = 100
    conn = _db()
    try:
        wheres, params = [], []
        if status_filter:
            wheres.append("t.status=?"); params.append(status_filter)
        if type_filter:
            wheres.append("t.type=?"); params.append(type_filter)
        if show_archived == "1":
            wheres.append("t.archived=1")
        else:
            wheres.append("(t.archived IS NULL OR t.archived=0)")
        where_sql = ("WHERE " + " AND ".join(wheres)) if wheres else ""
        total = conn.execute(f"SELECT COUNT(*) FROM tickets t {where_sql};", params).fetchone()[0]
        page_params = params + [PAGE, page * PAGE]
        try:
            tickets = conn.execute(f"""
                SELECT t.*,
                       (SELECT COUNT(*) FROM ticket_messages m WHERE m.ticket_id=t.id) AS msg_count,
                       (SELECT sender FROM ticket_messages m WHERE m.ticket_id=t.id ORDER BY m.id DESC LIMIT 1) AS last_sender,
                       p.title AS product_title
                FROM tickets t
                LEFT JOIN products p ON p.id=t.product_id
                {where_sql}
                ORDER BY t.updated_at DESC, t.id DESC LIMIT ? OFFSET ?;
            """, page_params).fetchall()
        except Exception:
            tickets = conn.execute(f"""
                SELECT t.*, (SELECT COUNT(*) FROM ticket_messages m WHERE m.ticket_id=t.id) AS msg_count,
                NULL AS last_sender, NULL AS product_title
                FROM tickets t {where_sql}
                ORDER BY id DESC LIMIT ? OFFSET ?;
            """, page_params).fetchall()
        # قبلاً ۹ کوئری COUNT جدا (۶ status + ۳ type) — با GROUP BY به ۲ کوئری کاهش
        # پیدا کرد؛ با ایندکس‌های تازهٔ tickets(status)/tickets(type) هم سریع‌تره.
        stats = {s: 0 for s in ("waiting_admin","waiting_user","closed","waiting_info","reviewing","ready_delivery")}
        for row in conn.execute("SELECT status, COUNT(*) AS c FROM tickets GROUP BY status;").fetchall():
            if row["status"] in stats:
                stats[row["status"]] = row["c"]
        type_counts = {t: 0 for t in ("support","product_setup","partner_support")}
        for row in conn.execute("SELECT type, COUNT(*) AS c FROM tickets GROUP BY type;").fetchall():
            if row["type"] in type_counts:
                type_counts[row["type"]] = row["c"]
    finally:
        conn.close()

    pages = max((total + PAGE - 1) // PAGE, 1)

    def tq(sf="", tf=""):
        return f"?status_filter={sf}&type_filter={tf}"

    type_tabs = '<div class="filter-tabs mb-3">'
    for lbl, val, cnt in [
        ("همه", "", sum(type_counts.values())),
        ("🔵 پشتیبانی", "support", type_counts["support"]),
        ("🟢 راه‌اندازی", "product_setup", type_counts["product_setup"]),
        ("🤝 همکاران", "partner_support", type_counts["partner_support"]),
    ]:
        active = type_filter == val
        cls = "filter-tab-lg filter-tab-lg--active" if active else "filter-tab-lg"
        type_tabs += f'<a href="{tq(status_filter, val)}" class="{cls}">{lbl} <span class="filter-tab-count">{cnt}</span></a>'
    type_tabs += "</div>"

    status_tabs = '<div class="filter-tabs mb-4">'
    for lbl, val, cnt in [("همه","",sum(stats.values())),("منتظر اطلاعات","waiting_info",stats.get("waiting_info",0)),
                          ("نیاز به پاسخ","waiting_admin",stats.get("waiting_admin",0)),("در بررسی","reviewing",stats.get("reviewing",0)),
                          ("آماده تحویل","ready_delivery",stats.get("ready_delivery",0)),("منتظر کاربر","waiting_user",stats.get("waiting_user",0)),
                          ("بسته","closed",stats.get("closed",0))]:
        active = status_filter == val
        cls = "filter-tab filter-tab--active" if active else "filter-tab"
        status_tabs += f'<a href="{tq(val, type_filter)}" class="{cls}">{lbl} {cnt}</a>'
    status_tabs += "</div>"

    def sbadge(s):
        defs = {"waiting_info":("🔴","منتظر اطلاعات","status-danger"),
                "waiting_admin":("🔴","نیاز به پاسخ","status-danger"),
                "reviewing":("🟡","در بررسی","status-warning"),
                "waiting_user":("🟢","پاسخ داده شد","status-success"),
                "ready_delivery":("🟢","آماده تحویل","status-success"),
                "closed":("⚫","بسته","status-neutral")}
        icon,label,cls = defs.get(s,("⚪",s,"status-neutral"))
        return f'<span class="status-badge {cls}">{icon} {label}</span>'

    def tbadge(t):
        defs = {"product_setup":("🟢","راه‌اندازی","status-success"),
                "partner_support":("🤝","همکاران","status-success"),
                "support":("🔵","پشتیبانی","status-info")}
        icon,label,cls = defs.get(t,("🔵","پشتیبانی","status-info"))
        return f'<span class="status-badge status-pill-sm {cls}">{icon} {label}</span>'

    ticket_rows_list = []
    for t in tickets:
        try: ls = t["last_sender"]
        except: ls = None
        try: ptitle = e((t["product_title"] or "")[:24])
        except: ptitle = ""
        try: tid_type = t["type"] or "support"
        except: tid_type = "support"

        # type badge
        type_colors = {"product_setup":("green","راه‌اندازی"),"partner_support":("teal","همکاران"),"support":("blue","پشتیبانی")}
        tc, tl = type_colors.get(tid_type,("blue","پشتیبانی"))
        type_b = f'<span class="px-2 py-0.5 text-xs bg-{tc}-100 text-{tc}-700 rounded-full">{tl}</span>'

        # status badge
        status_colors = {"waiting_info":("red","منتظر اطلاعات"),"waiting_admin":("red","نیاز به پاسخ"),
                         "reviewing":("yellow","در بررسی"),"waiting_user":("green","پاسخ داده شد"),
                         "ready_delivery":("green","آماده تحویل"),"closed":("gray","بسته")}
        sc, sl = status_colors.get(t["status"],("gray",t["status"]))
        status_b = f'<span class="px-2 py-0.5 text-xs bg-{sc}-100 text-{sc}-700 rounded-full">{sl}</span>'

        last_col = ptitle or ("↗ کاربر" if ls=="user" else ("↙ ادمین" if ls=="admin" else ""))

        if show_archived == "1":
            action_btns = f"""<a href="/admin/tickets/{t['id']}" onclick="event.stopPropagation()"
              class="px-2 py-1 text-xs bg-indigo-50 text-indigo-700 border border-indigo-200 rounded">👁 مشاهده</a>
              <button type="button" onclick="event.stopPropagation();unarchiveTicket({t['id']})"
              class="px-2 py-1 text-xs bg-blue-50 text-blue-600 border border-blue-200 rounded mr-1">بازگردانی</button>"""
        else:
            action_btns = f"""<a href="/admin/tickets/{t['id']}" onclick="event.stopPropagation()"
              class="px-2 py-1 text-xs bg-indigo-50 text-indigo-700 border border-indigo-200 rounded">👁 مشاهده</a>
              <button type="button" onclick="event.stopPropagation();deleteTicket({t['id']})"
              class="px-2 py-1 text-xs bg-red-50 text-red-500 border border-red-200 rounded mr-1">🗑 حذف</button>"""

        row_html = f"""<tr class="border-b hover:bg-gray-50 cursor-pointer" onclick="location.href='/admin/tickets/{t['id']}'">
          <td class="px-4 py-3" data-label="#"><a href="/admin/tickets/{t['id']}" class="text-xs font-bold text-indigo-600">#{t['id']}</a></td>
          <td class="px-4 py-3" data-label="نوع">{type_b}</td>
          <td class="px-4 py-3" data-label="کاربر"><code class="text-xs bg-gray-100 px-1.5 py-0.5 rounded">{e(str(t['user_id']))}</code></td>
          <td class="px-4 py-3" data-label="وضعیت">{status_b}</td>
          <td class="px-4 py-3 text-xs text-gray-400" data-label="محصول">{last_col}</td>
          <td class="px-4 py-3 text-xs text-gray-400" data-label="پیام‌ها">{int(t['msg_count'] or 0)} پیام</td>
          <td class="px-4 py-3 text-xs text-gray-400" data-label="آپدیت">{fa_date(t['updated_at'] or '', with_time=True)}</td>
          <td class="px-4 py-3 whitespace-nowrap" data-label="">{action_btns}</td>
        </tr>"""
        ticket_rows_list.append(row_html)

    recent_rows = ticket_rows_list[:3]
    older_rows  = ticket_rows_list[3:]
    rows = "".join(recent_rows)
    older_rows_html = "".join(older_rows)

    tickets_toggle_btn = ""
    if older_rows:
        tickets_toggle_btn = f"""
        <div class="text-center p-3">
          <button type="button" id="toggle-older-tickets-btn" data-older-count="{len(older_rows)}" class="show-more-btn">
            🔽 نمایش {len(older_rows)} تیکت قدیمی‌تر
          </button>
        </div>"""

    # صفحه‌بندی سرور-محور — قبلاً سقف ثابت ۲۰۰ ردیف بدون هیچ صفحهٔ بعدی بود؛
    # تیکت‌های قدیمی‌تر از سقف اصلاً از دید مدیر محو می‌شدن. الگوی /admin/products.
    pager = ('<div class="flex gap-2 mt-4 justify-center">' + "".join(
        f'<a href="?status_filter={status_filter}&type_filter={type_filter}&show_archived={show_archived}&page={i}" '
        f'class="px-3 py-1 rounded border text-sm {"bg-indigo-600 text-white" if i==page else "bg-white"}">{i+1}</a>'
        for i in range(min(pages, 10))
    ) + "</div>") if pages > 1 else ""

    body = f"""
    <div class="flex items-center justify-between mb-4 flex-wrap gap-2">
      <h1 class="text-2xl font-bold text-gray-800">🎫 تیکت‌های پشتیبانی ({total:,})</h1>
      <a href="?status_filter={status_filter}&type_filter={type_filter}&show_archived={'0' if show_archived=='1' else '1'}"
         class="px-3 py-1.5 text-xs rounded-lg border {'bg-gray-700 text-white' if show_archived=='1' else 'bg-gray-50 text-gray-500 border-gray-200'}">
        {'🔙 بازگشت به لیست فعال' if show_archived=='1' else '📦 آرشیو شده‌ها'}
      </a>
    </div>
    {type_tabs}
    {status_tabs}
    <div class="card overflow-hidden">
      <div class="overflow-x-auto ticket-table-wrap">
        <table class="w-full text-right min-w-max">
          <thead><tr class="text-xs text-gray-500 border-b bg-gray-50">
            <th class="px-4 py-3">#</th><th class="px-4 py-3">نوع</th>
            <th class="px-4 py-3">کاربر</th><th class="px-4 py-3">وضعیت</th>
            <th class="px-4 py-3">محصول/پیام</th><th class="px-4 py-3">تعداد</th>
            <th class="px-4 py-3">آپدیت</th><th class="px-4 py-3"></th>
          </tr></thead>
          <tbody>{rows or "<tr><td colspan='8' class='text-center py-8 text-gray-400'>تیکتی یافت نشد</td></tr>"}</tbody>
          <tbody id="older-tickets-block" style="display:none">{older_rows_html}</tbody>
        </table>
      </div>
      {tickets_toggle_btn}
      {pager}
    </div>

    <script>
      window.archiveTicket = function(id){{
        if(!confirm('این تیکت آرشیو شود؟')) return;
        fetch('/admin/tickets/'+id+'/archive', {{method:'POST'}}).then(function(){{ location.reload(); }});
      }};
      window.unarchiveTicket = function(id){{
        fetch('/admin/tickets/'+id+'/unarchive', {{method:'POST'}}).then(function(){{ location.reload(); }});
      }};
      window.deleteTicket = function(id){{
        if(!confirm('⚠️ این تیکت برای همیشه حذف می‌شود. ادامه؟')) return;
        fetch('/admin/tickets/'+id+'/delete', {{method:'POST'}}).then(function(){{ location.reload(); }});
      }};
      (function(){{
        var btn = document.getElementById('toggle-older-tickets-btn');
        if(btn){{
          btn.addEventListener('click', function(){{
            var el = document.getElementById('older-tickets-block');
            var cnt = btn.getAttribute('data-older-count');
            if(el.style.display === 'none'){{
              el.style.display = 'table-row-group';
              btn.textContent = '🔼 بستن تیکت‌های قدیمی‌تر';
            }} else {{
              el.style.display = 'none';
              btn.textContent = '🔽 نمایش ' + cnt + ' تیکت قدیمی‌تر';
            }}
          }});
        }}
      }})();
    </script>"""

    return _layout("تیکت‌ها", body, adm, flash=flash)


@router.post("/tickets/{tid}/archive")
async def ticket_archive(request: Request, tid: int):
    adm = _get_admin(request)
    if not adm: return JSONResponse({"ok": False})
    if not _has(adm, "tickets"): return JSONResponse({"ok": False})
    from db import archive_ticket
    archive_ticket(tid)
    _log(request, "آرشیو تیکت", "تیکت‌ها", f"#{tid}")
    return JSONResponse({"ok": True})


@router.post("/tickets/{tid}/unarchive")
async def ticket_unarchive(request: Request, tid: int):
    adm = _get_admin(request)
    if not adm: return JSONResponse({"ok": False})
    if not _has(adm, "tickets"): return JSONResponse({"ok": False})
    from db import unarchive_ticket
    unarchive_ticket(tid)
    _log(request, "بازگردانی تیکت از آرشیو", "تیکت‌ها", f"#{tid}")
    return JSONResponse({"ok": True})


@router.post("/tickets/{tid}/delete")
async def ticket_delete(request: Request, tid: int):
    adm = _get_admin(request)
    if not adm: return JSONResponse({"ok": False})
    if not _has(adm, "tickets"): return JSONResponse({"ok": False})
    from db import delete_ticket
    delete_ticket(tid)
    _log(request, "حذف تیکت", "تیکت‌ها", f"#{tid}")
    return JSONResponse({"ok": True})


@router.get("/tickets/{tid}", response_class=HTMLResponse)
async def ticket_detail(request: Request, tid: int, flash: str = ""):
    adm = _get_admin(request)
    if not adm:
        return _redir("/admin/login")
    guard = _require(adm, "tickets")
    if guard: return guard

    conn = _db()
    try:
        ticket = conn.execute("""
            SELECT t.*, p.title as product_title
            FROM tickets t LEFT JOIN products p ON t.product_id=p.id
            WHERE t.id=? LIMIT 1;
        """, (tid,)).fetchone()

        if not ticket:
            return _redir("/admin/tickets")

        messages = conn.execute(
            "SELECT * FROM ticket_messages WHERE ticket_id=? ORDER BY id ASC;", (tid,)
        ).fetchall()

        # وقتی ادمین تیکت رو باز می‌کنه → status به in_progress تغییر کنه (badge پاک بشه)
        if ticket["status"] == "open":
            conn.execute("UPDATE tickets SET status='in_progress' WHERE id=?;", (tid,))
            conn.commit()

    finally:
        conn.close()

    is_general = (not ticket["product_id"] or int(ticket["product_id"] or 0) == 0)
    is_closed = (ticket["status"] == "closed")
    user_id_val = int(ticket["user_id"])

    # ── مکالمه با نمایش کامل رسانه ──────────────────────────────────────────
    bot_token = _env("BOT_TOKEN")
    chat_html = ""
    last_msg_id = 0

    def _render_media(msg) -> str:
        """رندر امن رسانه — هرگز crash نمی‌دهد."""
        try:
            # sqlite3.Row از .get() پشتیبانی نمی‌کند، از try/except استفاده می‌کنیم
            try: mt = (msg["media_type"] or "").strip().lower()
            except: mt = ""
            try: fid = msg["media_file_id"] or ""
            except: fid = ""
            try: txt = (msg["text"] or "").strip()
            except: txt = ""
            try: fname = (msg["file_name"] or "").strip()
            except: fname = ""

            proxy_qs = f"?fn={_urlquote(fname, safe='')}" if fname else ""
            caption = f'<div class="chat-caption">{e(txt)}</div>' if txt and not txt.startswith("[") else ""
            proxy = f"/admin/tickets/media/{e(fid)}{proxy_qs}" if fid else ""

            if mt == "photo" and proxy:
                return (
                    f'<a href="{proxy}" target="_blank">'
                    f'<img src="{proxy}" class="chat-media-img" '
                    f'onerror="this.parentElement.innerHTML=\'📷 خطا در بارگذاری\'"></a>'
                    + caption
                )
            elif mt == "voice" and proxy:
                return f'<audio controls class="chat-media-audio"><source src="{proxy}"></audio>{caption}'
            elif mt == "video" and proxy:
                return f'<video controls class="chat-media-video"><source src="{proxy}"></video>{caption}'
            elif mt in ("document", "audio") and proxy:
                icon = "🎵" if mt == "audio" else "📎"
                label = txt if txt and not txt.startswith("[") else "دانلود فایل"
                return f'<a href="{proxy}" download target="_blank" class="chat-media-file">{icon} {e(label)}</a>'
            elif mt and mt not in ("text", ""):
                icons = {"sticker": "🎭", "animation": "🎬", "video_note": "📹"}
                return f'{icons.get(mt, "📁")} <em class="chat-icon-label">[{e(mt)}]</em>{caption}'
            else:
                return e(txt) if txt else ""
        except Exception:
            return '<em class="chat-empty">[خطا]</em>'

    older_messages = messages[:-3] if len(messages) > 3 else []
    recent_messages = messages[-3:] if len(messages) > 3 else messages

    def _render_msg(msg) -> str:
        is_adm = msg["sender"] == "admin"
        content_html = _render_media(msg)
        try: src = msg["source"] or ""
        except: src = ""
        src_icon = "🖥" if src not in ("telegram", "") else "📱"
        time_str = fa_date(msg["created_at"] or "", with_time=True)
        if is_adm:
            return f"""
        <div class="chat-row chat-row--admin" data-msg-id="{msg['id']}">
          <div class="chat-col">
            <div class="chat-bubble chat-bubble--admin">
              {content_html or '<em class="chat-empty-msg">پیام خالی</em>'}
            </div>
            <div class="chat-meta chat-meta--admin">{src_icon} ادمین · {time_str}</div>
          </div>
        </div>"""
        else:
            return f"""
        <div class="chat-row chat-row--user" data-msg-id="{msg['id']}">
          <div class="chat-col">
            <div class="chat-bubble chat-bubble--user">
              {content_html or '<em class="chat-empty-msg--user">پیام خالی</em>'}
            </div>
            <div class="chat-meta">📱 کاربر {user_id_val} · {time_str}</div>
          </div>
        </div>"""

    for msg in messages:
        last_msg_id = max(last_msg_id, int(msg["id"] or 0))

    older_html = "".join(_render_msg(m) for m in older_messages)
    recent_html = "".join(_render_msg(m) for m in recent_messages)

    toggle_btn = ""
    if older_messages:
        toggle_btn = f"""
        <div class="text-center mb-3">
          <button type="button" id="toggle-older-btn" class="show-more-btn" onclick="
            var el=document.getElementById('older-messages-block');
            var btn=document.getElementById('toggle-older-btn');
            if(el.style.display==='none'){{ el.style.display='block'; btn.textContent='🔼 بستن پیام‌های قبلی'; }}
            else {{ el.style.display='none'; btn.textContent='🔽 نمایش {len(older_messages)} پیام قبلی'; }}
          ">
            🔽 نمایش {len(older_messages)} پیام قبلی
          </button>
        </div>
        <div id="older-messages-block" style="display:none">{older_html}</div>"""

    chat_html = toggle_btn + recent_html

    if not chat_html.strip() or not messages:
        chat_html = '<div class="text-center py-8 text-gray-400 text-sm" id="no-msgs">پیامی ثبت نشده</div>'

    # ── فرم پاسخ ─────────────────────────────────────────────────────────────
    reply_form = ""
    if not is_closed:
        reply_form = f"""
        <div class="card p-4 mt-4">
          <form method="post" action="/admin/tickets/{tid}/reply" id="reply-form" enctype="multipart/form-data">
            <div class="mb-3">
              <label class="text-xs text-gray-500 block mb-1">
                پاسخ به کاربر <code class="bg-gray-100 px-1 rounded">{user_id_val}</code>
              </label>
              <textarea name="text" id="reply-text" rows="3"
                placeholder="متن پاسخ را بنویسید (در صورت پیوست فایل، اختیاری است)..."
                class="w-full border border-gray-200 rounded-xl px-3 py-2 text-sm resize-none focus:ring-2 focus:ring-indigo-300"></textarea>
            </div>
            <div class="mb-3">
              <label class="text-xs text-gray-500 block mb-1">📎 پیوست عکس/فایل (اختیاری)</label>
              <input type="file" name="attachment" id="reply-attachment"
                accept="image/*,.pdf,.doc,.docx,.zip,.rar,.txt,.xlsx,.xls"
                class="w-full border border-gray-200 rounded-xl px-3 py-1.5 text-xs">
            </div>
            <div class="flex justify-between items-center">
              <span class="text-xs text-gray-300">Ctrl+Enter برای ارسال سریع</span>
              <button type="submit"
                class="bg-indigo-600 hover:bg-indigo-700 text-white px-5 py-2 rounded-xl text-sm font-medium">
                📤 ارسال پاسخ
              </button>
            </div>
          </form>
        </div>"""

    # ── وضعیت دکمه‌ها (با status های جدید v2) ─────────────────────────────
    status_btns = ""
    cur_status = ticket["status"]
    for lbl2, val2, cls2 in [
        ("🔓 بازکردن", "waiting_admin", "bg-green-50 text-green-700 border-green-200"),
        ("🔒 بستن تیکت", "closed", "bg-gray-100 text-gray-600 border-gray-200"),
    ]:
        if val2 != cur_status:
            status_btns += f"""
            <form method="post" action="/admin/tickets/{tid}/status" class="inline-block mr-1 mb-1">
              <input type="hidden" name="status" value="{val2}">
              <button class="btn-sm border rounded-lg px-3 py-1.5 text-xs {cls2}">{lbl2}</button>
            </form>"""

    # ── ساختار صفحه ─────────────────────────────────────────────────────────
    ticket_type_str = "پشتیبانی عمومی" if is_general else e(ticket["product_title"] or "-")

    try: t_type = ticket["type"] or "support"
    except: t_type = "support"
    try: t_order_id = ticket["order_id"] or 0
    except: t_order_id = 0
    try: t_feed_id = ticket["feed_id"]
    except: t_feed_id = None
    try: t_setup_status = ticket["setup_status"] or ticket["status"]
    except: t_setup_status = ticket["status"]

    type_labels = {
        "product_setup": "🟢 راه‌اندازی محصول",
        "partner": "🟠 همکاری",
        "support": "🔵 پشتیبانی عمومی",
    }
    type_label = type_labels.get(t_type, "🔵 پشتیبانی عمومی")

    # بخش اختصاصی product_setup
    setup_panel = ""
    if t_type == "product_setup" and not is_closed:
        setup_status_opts = [
            ("waiting_info", "🔴 منتظر اطلاعات"),
            ("reviewing", "🟡 در حال بررسی"),
            ("ready_delivery", "🟢 آماده تحویل"),
        ]
        setup_status_btns = "".join(
            f'<form method="post" action="/admin/tickets/{tid}/setup-status" class="d-inline">'
            f'<input type="hidden" name="status" value="{sv}">'
            f'<button class="btn btn-slate btn-sm {"setup-status-btn--active" if t_setup_status==sv else ""}">{sl}</button></form>'
            for sv, sl in setup_status_opts
        )
        deliver_btn = ""
        if t_feed_id and t_setup_status in ("ready_delivery", "reviewing"):
            deliver_btn = f"""
            <form method="post" action="/admin/tickets/{tid}/deliver"
                  onsubmit="return confirm('محصول به کاربر تحویل داده شود و گفتگو بسته شود؟')">
              <button class="btn btn-primary btn-full-mt12">
                <i data-lucide="send" class="icon-15"></i> تحویل محصول و بستن گفتگو
              </button>
            </form>"""
        setup_panel = f"""
        <div class="card card-p ticket-setup-card">
          <h3 class="ticket-setup-title">🟢 اطلاعات راه‌اندازی محصول</h3>
          <dl class="ticket-info-grid">
            <dt class="ticket-info-dt">محصول</dt><dd class="ticket-info-dd-bold">{e(ticket["product_title"] or "—")}</dd>
            <dt class="ticket-info-dt">سفارش</dt><dd><code class="ticket-info-code">#{t_order_id}</code></dd>
            <dt class="ticket-info-dt">فید</dt><dd><code class="ticket-info-code">#{t_feed_id or "—"}</code></dd>
          </dl>
          <div class="mt-12">
            <div class="muted-xs-mb8">وضعیت راه‌اندازی</div>
            <div class="flex-wrap-gap6">{setup_status_btns}</div>
          </div>
          {deliver_btn}
        </div>"""

    direct_form_html = "" if is_closed else f"""
        <div class="card p-4 mt-3 ticket-direct-card">
          <p class="ticket-direct-label">📩 پیام مستقیم</p>
          <form method="post" action="/admin/tickets/{tid}/direct" class="ticket-direct-form">
            <textarea name="direct_msg" rows="2" placeholder="پیام آزاد..." class="ticket-direct-input"></textarea>
            <button type="submit" class="ticket-direct-btn">ارسال</button>
          </form>
        </div>"""

    body = f"""
    <div class="ticket-header">
      {_btn("← تیکت‌ها", "/admin/tickets", "slate", small=True)}
      <h1 class="ticket-title">تیکت #{tid}</h1>
      {_ticket_status_badge(ticket["status"])}
      <span class="ticket-type-label">{type_label}</span>
    </div>

    <div class="grid lg:grid-cols-3 gap-4">
      <div class="lg:col-span-2">
        <div class="card p-4 overflow-y-auto chat-box" id="chat-box">
          {chat_html}
        </div>
        {reply_form}
        {direct_form_html}
      </div>

      <div class="sidebar-flex-col-12">
        {setup_panel}
        <div class="card card-p">
          <h3 class="card-title-sm">اطلاعات تیکت</h3>
          <dl class="ticket-info-grid">
            <dt class="ticket-info-dt">User ID</dt><dd><code class="ticket-info-code">{user_id_val}</code></dd>
            <dt class="ticket-info-dt">نوع</dt><dd class="ticket-info-dd-bold">{type_label}</dd>
            <dt class="ticket-info-dt">پیام‌ها</dt><dd class="count-primary">{len(messages)}</dd>
            <dt class="ticket-info-dt">تاریخ</dt><dd class="ticket-info-dt">{fa_date(ticket["created_at"] or "", with_time=True)}</dd>
          </dl>
        </div>
        <div class="card card-p">
          <h3 class="card-title-sm-10">تغییر وضعیت</h3>
          {status_btns}
        </div>
        <div class="card card-p">
          <button type="button" onclick="archiveThisTicket({tid})" class="archive-btn">
            📦 آرشیو این تیکت
          </button>
        </div>
      </div>
    </div>
    <script>
      window.archiveThisTicket = function(id){{
        if(!confirm('این تیکت آرشیو شود؟ از لیست اصلی پنهان می‌شود.')) return;
        fetch('/admin/tickets/'+id+'/archive', {{method:'POST'}}).then(function(){{ location.href='/admin/tickets'; }});
      }};
      (function(){{
        var b = document.getElementById('chat-box');
        if(b) b.scrollTop = b.scrollHeight;

        // Ctrl+Enter
        var ta = document.getElementById('reply-text');
        if(ta) ta.addEventListener('keydown', function(e){{
          if(e.ctrlKey && e.key === 'Enter') document.getElementById('reply-form') && document.getElementById('reply-form').submit();
        }});

        // Auto-refresh هر ۱۰ ثانیه
        var lastId = {last_msg_id};
        var ticketId = {tid};
        function fetchNew() {{
          fetch('/admin/tickets/' + ticketId + '/messages.json?after=' + lastId)
            .then(function(r){{ return r.json(); }})
            .then(function(data){{
              if(!data.messages || data.messages.length === 0) return;
              var nm = document.getElementById('no-msgs');
              if(nm) nm.remove();
              data.messages.forEach(function(msg){{
                lastId = Math.max(lastId, msg.id);
                var d = document.createElement('div');
                var isA = msg.sender === 'admin';
                d.className = 'flex flex-col ' + (isA ? 'items-end' : 'items-start') + ' mb-3';
                var bbl = isA ? 'bg-indigo-600 text-white' : 'bg-white border border-gray-200 text-gray-800';
                var lbl = isA ? 'ادمین 👤' : 'کاربر';
                var txt = (msg.text || '').replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;');
                if(msg.media_type) txt = '[' + msg.media_type + '] ' + txt;
                d.innerHTML = '<div class="text-xs text-gray-400 mb-1">' + lbl + ' · ' + (msg.created_at||'').slice(0,16) + '</div>' +
                  '<div class="' + bbl + ' rounded-2xl px-4 py-2 text-sm chat-js-bubble">' + txt + '</div>';
                if(b) b.appendChild(d);
              }});
              if(b) b.scrollTop = b.scrollHeight;
            }}).catch(function(){{}});
        }}
        setInterval(fetchNew, 10000);
      }})();
    </script>"""

    return _layout(f"تیکت #{tid}", body, adm, flash=flash)


@router.post("/tickets/{tid}/setup-status")
async def ticket_setup_status(request: Request, tid: int):
    adm = _get_admin(request)
    if not adm: return _redir("/admin/login")
    guard = _require(adm, "tickets")
    if guard: return guard
    form = await request.form()
    new_status = str(form.get("status", "reviewing"))
    conn = _db()
    try:
        conn.execute("UPDATE tickets SET setup_status=?, status=?, updated_at=datetime('now') WHERE id=?;",
                     (new_status, new_status, tid))
        conn.commit()
    finally:
        conn.close()
    _log(request, "تغییر وضعیت راه‌اندازی", "تیکت‌ها", f"تیکت #{tid} → {new_status}")
    return _redir(f"/admin/tickets/{tid}?flash=وضعیت+تغییر+کرد")


@router.post("/tickets/{tid}/deliver")
async def ticket_deliver_product(request: Request, tid: int):
    adm = _get_admin(request)
    if not adm: return _redir("/admin/login")
    guard = _require(adm, "tickets")
    if guard: return guard
    conn = _db()
    try:
        ticket = conn.execute("SELECT * FROM tickets WHERE id=? LIMIT 1;", (tid,)).fetchone()
        if not ticket:
            return _redir(f"/admin/tickets?flash=تیکت+یافت+نشد")
        try: feed_id = ticket["feed_id"]
        except: feed_id = None
        try: feed_data = ticket["feed_data"]
        except: feed_data = None
        try: order_id = ticket["order_id"] or 0
        except: order_id = 0
        user_id = int(ticket["user_id"])
        try: ptitle = ticket["product_title"] if "product_title" in ticket.keys() else "محصول"
        except: ptitle = "محصول"

        if not feed_data:
            return _redir(f"/admin/tickets/{tid}?flash=اطلاعات+محصول+یافت+نشد")

        # ارسال به کاربر
        bot_token = _env("BOT_TOKEN")
        msg_text = (f"✅ <b>سفارش شما تکمیل شد</b>\n\n"
                    f"سفارش: #{order_id}\n"
                    f"محصول: {e(ptitle)}\n\n"
                    f"<code>{html.escape(str(feed_data))}</code>")
        try:
            await run_in_threadpool(
                _requests.post,
                f"https://api.telegram.org/bot{bot_token}/sendMessage",
                json={"chat_id": user_id, "text": msg_text, "parse_mode": "HTML"},
                timeout=10
            )
        except Exception as ex:
            _tg_logger.error("ticket deliver send error: %s", ex)
            return _redir(f"/admin/tickets/{tid}?flash=خطا+در+ارسال+به+کاربر")

        # بستن تیکت
        conn.execute("UPDATE tickets SET status='closed', setup_status='closed', closed_at=datetime('now') WHERE id=?;", (tid,))
        conn.commit()

        _log(request, "تحویل محصول از تیکت", "تیکت‌ها",
             f"تیکت #{tid} سفارش #{order_id} به کاربر {user_id} تحویل داده شد")
    finally:
        conn.close()
    return _redir(f"/admin/tickets/{tid}?flash=✅+محصول+تحویل+داده+شد+و+گفتگو+بسته+شد")


@router.post("/tickets/{tid}/reply")
async def ticket_reply(request: Request, tid: int, text: str = Form(""),
                        attachment: UploadFile = None):
    adm = _get_admin(request)
    if not adm:
        return _redir("/admin/login")
    guard = _require(adm, "tickets")
    if guard: return guard

    text = text.strip()
    has_attachment = bool(attachment and attachment.filename)
    if not text and not has_attachment:
        return _redir(f"/admin/tickets/{tid}?flash=متن+یا+پیوست+لازم+است")

    attach_bytes = None
    attach_name = None
    as_photo = False
    if has_attachment:
        try:
            attach_bytes = await attachment.read()
        except Exception:
            return _redir(f"/admin/tickets/{tid}?flash=خطا+در+خواندن+فایل+پیوست")
        if not attach_bytes:
            return _redir(f"/admin/tickets/{tid}?flash=فایل+پیوست+خالی+است")
        attach_name = attachment.filename
        as_photo = (attachment.content_type or "").startswith("image/")

    conn = _db()
    try:
        ticket = conn.execute(
            "SELECT user_id, status FROM tickets WHERE id=? LIMIT 1;", (tid,)
        ).fetchone()
        if not ticket:
            conn.close()
            return _redir("/admin/tickets?flash=تیکت+یافت+نشد")
        if ticket["status"] == "closed":
            conn.close()
            return _redir(f"/admin/tickets/{tid}?flash=تیکت+بسته+است")
        user_id = int(ticket["user_id"])

        # ─── migration اطمینان از وجود ستون‌های لازم ─────────────────────
        for col, typedef in [("source", "TEXT"), ("media_file_id", "TEXT"), ("updated_at", "TEXT"), ("user_msg_count", "INTEGER DEFAULT 0")]:
            try:
                conn.execute(f"ALTER TABLE tickets ADD COLUMN {col} {typedef};")
            except Exception:
                pass
            try:
                conn.execute(f"ALTER TABLE ticket_messages ADD COLUMN {col} {typedef};")
            except Exception:
                pass
        try:
            conn.execute("ALTER TABLE ticket_messages ADD COLUMN file_name TEXT;")
        except Exception:
            pass
    except Exception as ex:
        conn.close()
        _tg_logger.error("ticket_reply migration error: %s", ex)
        return _redir(f"/admin/tickets/{tid}?flash=خطای+پایگاه+داده:+{str(ex)}")

    # ─── ارسال به کاربر از طریق Telegram API — با دکمه «ادامه گفتگو» ────────
    continue_kb = {
        "inline_keyboard": [[
            {"text": "💬 ادامه گفتگو", "callback_data": f"ticket_v2_continue_{tid}"}
        ]]
    }
    token = _env("BOT_TOKEN", "")
    ok = False
    tg_file_id = None
    media_type = None

    if has_attachment:
        caption = f"💬 <b>پاسخ پشتیبانی</b> (تیکت #{tid}):\n\n{html.escape(text)}" if text else f"💬 <b>پیوست پشتیبانی</b> (تیکت #{tid})"
        res = await run_in_threadpool(
            _tg_send_document, user_id, attach_bytes, attach_name, caption, as_photo, continue_kb
        )
        ok = res.get("ok", False)
        tg_file_id = res.get("file_id")
        media_type = "photo" if as_photo else "document"
    else:
        msg_text = f"💬 <b>پاسخ پشتیبانی</b> (تیکت #{tid}):\n\n{html.escape(text)}"
        try:
            r = await run_in_threadpool(
                _requests.post,
                f"https://api.telegram.org/bot{token}/sendMessage",
                json={"chat_id": user_id, "text": msg_text,
                      "parse_mode": "HTML", "reply_markup": continue_kb},
                timeout=8
            )
            ok = r.json().get("ok", False)
        except Exception:
            ok = await run_in_threadpool(_tg_send, user_id, msg_text)

    # ─── ذخیره در ticket_messages (بعد از تلاش ارسال، تا media_file_id واقعی رو داشته باشیم) ──
    try:
        now = datetime.now().isoformat()
        stored_text = text if text else f"[{media_type or 'text'}]"
        conn.execute(
            "INSERT INTO ticket_messages (ticket_id, sender, text, media_type, media_file_id, file_name, source, created_at) "
            "VALUES (?,?,?,?,?,?,?,?);",
            (tid, "admin", stored_text, media_type, tg_file_id, attach_name, "panel", now)
        )

        # آپدیت وضعیت تیکت
        ticket_cols = [r[1] for r in conn.execute("PRAGMA table_info(tickets);").fetchall()]
        if "user_msg_count" in ticket_cols and "updated_at" in ticket_cols:
            conn.execute(
                "UPDATE tickets SET status='waiting_user', user_msg_count=0, updated_at=? WHERE id=?;",
                (now, tid)
            )
        elif "updated_at" in ticket_cols:
            conn.execute("UPDATE tickets SET status='waiting_user', updated_at=? WHERE id=?;", (now, tid))
        else:
            conn.execute("UPDATE tickets SET status='waiting_user' WHERE id=?;", (tid,))
        conn.commit()
    except Exception as ex:
        _tg_logger.error("ticket_reply DB error: %s", ex)
        return _redir(f"/admin/tickets/{tid}?flash=خطای+پایگاه+داده:+{str(ex)}")
    finally:
        conn.close()

    if ok:
        _log(request, "پاسخ تیکت", "تیکت‌ها", f"ticket #{tid}")
        return _redir(f"/admin/tickets/{tid}?flash=پاسخ+ارسال+شد")
    else:
        return _redir(f"/admin/tickets/{tid}?flash=ذخیره+شد+اما+ارسال+تلگرام+ناموفق")


@router.post("/tickets/{tid}/direct")
async def ticket_direct(request: Request, tid: int, direct_msg: str = Form("")):
    adm = _get_admin(request)
    if not adm:
        return _redir("/admin/login")
    guard = _require(adm, "tickets")
    if guard: return guard

    direct_msg = direct_msg.strip()
    if not direct_msg:
        return _redir(f"/admin/tickets/{tid}")

    conn = _db()
    try:
        ticket = conn.execute("SELECT user_id FROM tickets WHERE id=? LIMIT 1;", (tid,)).fetchone()
        user_id = ticket["user_id"] if ticket else None
    finally:
        conn.close()

    if user_id:
        await run_in_threadpool(_tg_send, user_id, f"📩 <b>پیام مستقیم از پشتیبانی:</b>\n\n{html.escape(direct_msg)}")

    return _redir(f"/admin/tickets/{tid}?flash=پیام+مستقیم+ارسال+شد")


@router.get("/badges.json")
async def badges_json(request: Request):
    """شمارنده‌های real-time برای navbar."""
    from fastapi.responses import JSONResponse
    adm = _get_admin(request)
    if not adm:
        return JSONResponse({"tickets": 0, "partners": 0, "notes": 0}, status_code=401)
    try:
        conn = _db()
        open_notes = conn.execute("SELECT COUNT(*) FROM admin_notes WHERE status='open';").fetchone()[0]
        conn.close()
    except Exception:
        open_notes = 0
    try:
        pending_receipts = _pending_receipts_count()
    except Exception:
        pending_receipts = 0
    return JSONResponse({
        "tickets": _open_ticket_count(),
        "partners": _pending_partner_count(),
        "notes": int(open_notes),
        "receipts": pending_receipts,
    })


@router.get("/tickets/{tid}/messages.json")
async def ticket_messages_json(request: Request, tid: int, after: int = 0):
    from fastapi.responses import JSONResponse
    adm = _get_admin(request)
    if not adm:
        return JSONResponse({"messages": []}, status_code=401)
    if not _has(adm, "tickets"):
        return JSONResponse({"messages": []}, status_code=403)
    conn = _db()
    conn.row_factory = sqlite3.Row
    try:
        rows = conn.execute(
            "SELECT id, sender, text, media_type, media_file_id, source, created_at "
            "FROM ticket_messages WHERE ticket_id=? AND id>? ORDER BY id ASC;",
            (tid, after)
        ).fetchall()
        return JSONResponse({"messages": [dict(r) for r in rows]})
    finally:
        conn.close()


@router.get("/tickets/media/{file_id}")
async def ticket_media(request: Request, file_id: str, fn: str = ""):
    """Proxy عکس‌های تیکت از Telegram."""
    from fastapi.responses import Response
    adm = _get_admin(request)
    if not adm:
        return Response(status_code=403)
    if not _has(adm, "tickets"):
        return Response(status_code=403)
    token = _env("BOT_TOKEN")
    if not token:
        return Response(status_code=404)
    try:
        # گرفتن file_path از Telegram — این مسیر پرتکرارترین مصرف‌کنندهٔ تماس sync با
        # تلگرام توی کل پنله (هر عکس ضمیمهٔ هر تیکت)، پس هر دو تماس حتماً باید توی
        # threadpool باشن، وگرنه یه تلگرام کند = کل پنل/API/webhook برای همه یخ می‌زنه.
        r1 = await run_in_threadpool(
            _requests.get, f"https://api.telegram.org/bot{token}/getFile?file_id={file_id}", timeout=10)
        fp = r1.json().get("result", {}).get("file_path", "")
        if not fp:
            return Response(status_code=404)
        r2 = await run_in_threadpool(
            _requests.get, f"https://api.telegram.org/file/bot{token}/{fp}", timeout=15)
        ct = r2.headers.get("content-type", "image/jpeg")
        headers = {}
        if fn:
            # نام فایل اصلی (پسوند واقعی) — بدونش مرورگر بعد از دانلود اسم/پسوند تلگرام
            # (مثل هش file_path) رو جایگزین می‌کنه، نه اسم واقعی که کاربر آپلود کرده بود.
            safe_fn = fn.replace('"', "").replace("\\", "").replace("\r", "").replace("\n", "")
            headers["Content-Disposition"] = f'inline; filename="{safe_fn}"'
        return Response(content=r2.content, media_type=ct, headers=headers)
    except Exception:
        return Response(status_code=502)


@router.post("/tickets/{tid}/status")
async def ticket_status(request: Request, tid: int, status: str = Form("")):
    adm = _get_admin(request)
    if not adm:
        return _redir("/admin/login")
    guard = _require(adm, "tickets")
    if guard: return guard

    valid = {"waiting_admin", "waiting_user", "closed", "open", "in_progress"}
    if status not in valid:
        return _redir(f"/admin/tickets/{tid}?flash=وضعیت+نامعتبر")

    # نرمال‌سازی: اگه status قدیمی بود، به جدید تبدیل کن
    status_map = {"open": "waiting_admin", "in_progress": "waiting_user"}
    status = status_map.get(status, status)

    conn = _db()
    try:
        now = datetime.now().isoformat()
        if status == "closed":
            conn.execute(
                "UPDATE tickets SET status='closed', closed_at=?, updated_at=? WHERE id=?;",
                (now, now, tid)
            )
        else:
            conn.execute("UPDATE tickets SET status=?, updated_at=? WHERE id=?;", (status, now, tid))
        conn.commit()
    except Exception as ex:
        _tg_logger.error("ticket_status error: %s", ex)
        return _redir(f"/admin/tickets/{tid}?flash=خطا+در+تغییر+وضعیت")
    finally:
        conn.close()

    return _redir(f"/admin/tickets/{tid}?flash=وضعیت+تیکت+تغییر+کرد")
    adm = _get_admin(request)
    if not adm:
        return _redir("/admin/login")

    valid = {"open", "in_progress", "closed"}
    if status not in valid:
        return _redir(f"/admin/tickets/{tid}")

    conn = _db()
    try:
        now = datetime.now().isoformat()
        if status == "closed":
            conn.execute("UPDATE tickets SET status=?, closed_at=?, closed_by='admin' WHERE id=?;", (status, now, tid))
        else:
            conn.execute("UPDATE tickets SET status=? WHERE id=?;", (status, tid))
        conn.commit()
    finally:
        conn.close()

    return _redir(f"/admin/tickets/{tid}?flash=وضعیت+تغییر+کرد")


# ─────────────────────────── Broadcast ─────────────────────────────────────

# وضعیت broadcast جاری
_broadcast_state: dict = {"running": False, "total": 0, "sent": 0, "failed": 0, "done": False}
_broadcast_lock = threading.Lock()


def _do_broadcast(user_ids: list[int], text: str, photo_url: str,
                   inline_buttons: list[dict], token: str) -> None:
    global _broadcast_state
    with _broadcast_lock:
        _broadcast_state.update({"running": True, "total": len(user_ids), "sent": 0, "failed": 0, "done": False})

    markup = None
    if inline_buttons:
        markup = {"inline_keyboard": [inline_buttons]}

    for uid in user_ids:
        try:
            if photo_url:
                _tg_send_photo(uid, photo_url, caption=text, reply_markup=markup)
            else:
                _tg_send(uid, text, reply_markup=markup)
            with _broadcast_lock:
                _broadcast_state["sent"] += 1
        except Exception:
            with _broadcast_lock:
                _broadcast_state["failed"] += 1
        time.sleep(0.05)  # rate limit safety

    with _broadcast_lock:
        _broadcast_state["running"] = False
        _broadcast_state["done"] = True


@router.get("/broadcast", response_class=HTMLResponse)
async def broadcast_page(request: Request, flash: str = ""):
    adm = _get_admin(request)
    guard = _require(adm, "broadcast")
    if guard: return guard

    conn = _db()
    try:
        # محافظت در برابر جداول ناموجود (علت صفحه سفید)
        try:
            total_users = conn.execute("SELECT COUNT(*) FROM users;").fetchone()[0]
        except Exception:
            total_users = 0
        try:
            total_buyers = conn.execute("SELECT COUNT(DISTINCT user_id) FROM orders;").fetchone()[0]
        except Exception:
            total_buyers = 0
        non_buyers = max(total_users - total_buyers, 0)
        try:
            products = conn.execute("SELECT id, title FROM products WHERE is_active=1 ORDER BY title;").fetchall()
        except Exception:
            products = []
        try:
            categories = conn.execute("SELECT id, name FROM categories WHERE is_active=1 ORDER BY name;").fetchall()
        except Exception:
            categories = []
    finally:
        conn.close()

    prod_opts = "".join(f'<option value="{p["id"]}">{e(p["title"])}</option>' for p in products)
    cat_opts = "".join(f'<option value="{c["id"]}">{e(c["name"])}</option>' for c in categories)

    # وضعیت broadcast جاری
    status_html = ""
    with _broadcast_lock:
        st = dict(_broadcast_state)
    if st["total"] > 0:
        pct = int(st["sent"] / max(st["total"], 1) * 100)
        status_color = "green" if st["done"] else "indigo"
        status_html = f"""
        <div class="card p-5 mb-6 border-r-4 border-{status_color}-500">
          <h3 class="font-bold text-{status_color}-700 mb-2">{"✅ ارسال تمام شد" if st["done"] else "🔄 در حال ارسال..."}</h3>
          <div class="bg-gray-100 rounded-full h-3 mb-2 ltr-num">
            <div class="bg-{status_color}-500 h-3 rounded-full" style="width:{pct}%"></div>
          </div>
          <div class="text-sm text-gray-600">
            ارسال‌شده: {st["sent"]} | ناموفق: {st["failed"]} | کل: {st["total"]}
          </div>
        </div>"""

    body = f"""
    <h1 class="text-2xl font-bold text-gray-800 mb-6">📢 پیام‌رسان</h1>

    {status_html}

    <div class="grid md:grid-cols-3 gap-4 mb-6">
      {_card("کل کاربران", str(total_users), "در سیستم", "indigo")}
      {_card("خریداران", str(total_buyers), "حداقل یک خرید", "green")}
      {_card("بدون خرید", str(non_buyers), "عضو بدون خرید", "orange")}
    </div>

    <div class="card card-p">
      <h2 class="section-title">ارسال پیام جدید</h2>
      <form method="post" action="/admin/broadcast/send" class="broadcast-form">

        <div>
          <label>مخاطبان *</label>
          <select name="target" id="target-select" onchange="toggleTargetOptions(this.value)">
            <option value="all">همه کاربران ({total_users} نفر)</option>
            <option value="buyers">خریداران ({total_buyers} نفر)</option>
            <option value="non_buyers">بدون خرید ({non_buyers} نفر)</option>
            <option value="product">خریداران یک محصول خاص</option>
            <option value="category">خریداران یک دسته خاص</option>
          </select>
        </div>

        <div id="product-select" style="display:none">
          <label>محصول</label>
          <select name="product_id">{prod_opts}</select>
        </div>

        <div id="category-select" style="display:none">
          <label>دسته‌بندی</label>
          <select name="category_id">{cat_opts}</select>
        </div>

        <div>
          <label>متن پیام * <span class="label-hint">(HTML پشتیبانی می‌شود: &lt;b&gt;، &lt;i&gt;، &lt;code&gt;)</span></label>
          {_textarea("text", "متن پیام را اینجا بنویسید...", rows=6)}
        </div>

        <div>
          <label>آدرس عکس <span class="label-hint">(اختیاری)</span></label>
          {_input("photo_url", "https://example.com/image.jpg")}
        </div>

        <div>
          <label>دکمه‌های Inline <span class="label-hint">(اختیاری — فرمت: متن|لینک — هر دکمه یک خط)</span></label>
          {_textarea("buttons", "دکمه اول|https://t.me/yourbot\nدکمه دوم|https://site.com", rows=3)}
        </div>

        <div class="warning-box">
          ⚠️ بعد از ارسال، عملیات در پس‌زمینه اجرا می‌شود. صفحه را ببندید و بعداً وضعیت را بررسی کنید.
        </div>

        {_btn("📢 شروع ارسال", color="green")}
      </form>
    </div>

    <script>
    function toggleTargetOptions(val) {{
      document.getElementById('product-select').style.display = val === 'product' ? 'block' : 'none';
      document.getElementById('category-select').style.display = val === 'category' ? 'block' : 'none';
    }}
    </script>"""

    return _layout("پیام‌رسانی", body, adm, flash=flash)


@router.post("/broadcast/send")
async def broadcast_send(request: Request, background_tasks: BackgroundTasks,
    target: str = Form("all"), text: str = Form(""), photo_url: str = Form(""),
    buttons: str = Form(""), product_id: str = Form(""), category_id: str = Form("")):
    adm = _get_admin(request)
    guard = _require(adm, "broadcast")
    if guard: return guard

    with _broadcast_lock:
        if _broadcast_state["running"]:
            return _redir("/admin/broadcast?flash=یک+broadcast+در+حال+اجرا+است")

    text = text.strip()
    if not text:
        return _redir("/admin/broadcast?flash=متن+پیام+الزامی+است")

    # parse inline buttons
    inline_buttons = []
    for line in (buttons or "").strip().splitlines():
        line = line.strip()
        if "|" in line:
            parts = line.split("|", 1)
            if len(parts) == 2 and parts[1].strip().startswith("http"):
                inline_buttons.append({"text": parts[0].strip(), "url": parts[1].strip()})

    # get target users
    pid = int(product_id) if product_id.strip().isdigit() else None
    cid = int(category_id) if category_id.strip().isdigit() else None

    conn = _db()
    try:
        if target == "all":
            rows = conn.execute("SELECT user_id FROM users;").fetchall()
        elif target == "buyers":
            rows = conn.execute("SELECT DISTINCT user_id FROM orders;").fetchall()
        elif target == "non_buyers":
            rows = conn.execute("""
                SELECT u.user_id FROM users u
                LEFT JOIN orders o ON u.user_id=o.user_id WHERE o.user_id IS NULL;
            """).fetchall()
        elif target == "product" and pid:
            rows = conn.execute("SELECT DISTINCT user_id FROM orders WHERE product_id=?;", (str(pid),)).fetchall()
        elif target == "category" and cid:
            rows = conn.execute("""
                SELECT DISTINCT o.user_id FROM orders o
                JOIN products p ON CAST(o.product_id AS INTEGER)=p.id WHERE p.category_id=?;
            """, (cid,)).fetchall()
        else:
            rows = []
        user_ids = [int(r[0]) for r in rows]
    finally:
        conn.close()

    if not user_ids:
        return _redir("/admin/broadcast?flash=هیچ+کاربری+یافت+نشد")

    token = _env("BOT_TOKEN")
    background_tasks.add_task(_do_broadcast, user_ids, text, photo_url.strip(), inline_buttons, token)

    return _redir(f"/admin/broadcast?flash=ارسال+به+{len(user_ids)}+کاربر+آغاز+شد")


@router.get("/broadcast/status")
async def broadcast_status(request: Request):
    adm = _get_admin(request)
    if not adm:
        return _redir("/admin/login")
    with _broadcast_lock:
        st = dict(_broadcast_state)
    from fastapi.responses import JSONResponse
    return JSONResponse(st)


# ─────────────────────────── Auto Daily Backup ────────────────────────────

_BACKUP_DIR = "/tmp/stockland_backups"

def _DB_PATH():
    return _env("DB_PATH", "/opt/stockland/app/stockland.db")
_MAX_BACKUPS = 6
_auto_backup_started = False


# ─────────────────────────── پاکسازی خودکار کش و فایل‌های موقت ────────────────
# طبق درخواست صریح مالک پروژه: فقط داده‌های غیرضروری/حجیم پاک بشن، هرگز
# دیتابیس/بکاپ‌ها (_BACKUP_DIR بالا)/.env/تنظیمات/کاربران. بررسی مستقیم کد نشون
# داد این پروژه تقریباً هیچ کش/فایل‌موقت واقعی نداره (نه پوشهٔ staging آپلود، نه
# فایل لاگ روی دیسک، `_tempfile` وارد‌شده در فایل حتی هیچ‌جا استفاده نمی‌شه) — پس
# اسکوپ واقع‌بینانه فقط دو هدف واقعاً بی‌خطره:
#   ۱) __pycache__ — کش بایت‌کد پایتون، پایتون خودش دوباره می‌سازتش، صفر ریسک.
#   ۲) آواتار یتیم (app_media/avatars) — در عمل نادر (آپلود آواتار خودش موقع
#      جایگزینی، فایل قدیمی رو با پسوند دیگه پاک می‌کنه؛ تنها سناریوی واقعی
#      باقی‌موندن یتیم یعنی رایتِ ناقص دیتابیس بعد از نوشتن فایل روی دیسک) —
#      همچنان به‌عنوان یه شبکهٔ ایمنی نگه داشته شده، نه چون انتظار می‌ره چیز
#      زیادی پیدا کنه.
# بقیهٔ زیرپوشه‌های app_media (تصویر محصول/کاور آموزش/گالری) عمداً اسکن نمی‌شن —
# چون DB reference هرکدوم شکل متفاوتی داره (بعضی JSON، بعضی ستون ساده) و ریسک
# حذف اشتباه یک فایل زندهٔ نمایش داده‌شده به کاربر واقعی، بیشتر از فایدهٔ آزاد
# کردن چند مگابایته — اگه لازم شد، باید جدا و با بررسی دقیق‌تر طراحی بشه.

_cache_cleanup_started = False


def _dir_size_and_count(path: str) -> tuple:
    total = 0
    count = 0
    for root, _dirs, files in os.walk(path):
        for f in files:
            try:
                total += os.path.getsize(os.path.join(root, f))
                count += 1
            except OSError:
                pass
    return total, count


def run_cache_cleanup() -> dict:
    """پاکسازی واقعی — هم از دکمهٔ «پاکسازی فوری» هم از ترد زمان‌بندی‌شده صدا زده می‌شه."""
    base_dir = os.path.dirname(os.path.abspath(__file__))
    freed_bytes = 0
    details = {}

    # ۱) __pycache__
    py_freed = 0
    py_count = 0
    try:
        for root, dirs, _files in os.walk(base_dir):
            dirs[:] = [d for d in dirs if d != ".git"]
            if "__pycache__" in dirs:
                pc_path = os.path.join(root, "__pycache__")
                sz, cnt = _dir_size_and_count(pc_path)
                try:
                    shutil.rmtree(pc_path, ignore_errors=True)
                    py_freed += sz
                    py_count += cnt
                except Exception:
                    pass
                dirs.remove("__pycache__")
    except Exception as ex:
        _tg_logger.error("cache cleanup __pycache__ error: %s", ex)
    freed_bytes += py_freed
    details["pycache"] = {"freed_bytes": py_freed, "files": py_count}

    # ۲) آواتارهای یتیم
    av_freed = 0
    av_count = 0
    try:
        avatars_dir = os.path.join(base_dir, "app_media", "avatars")
        if os.path.isdir(avatars_dir):
            conn = _db()
            try:
                kept = set()
                for row in conn.execute(
                    "SELECT avatar_url FROM users WHERE avatar_url IS NOT NULL AND avatar_url != '';"
                ):
                    url = row[0] or ""
                    kept.add(os.path.basename(url.split("?")[0]))
            finally:
                conn.close()
            cutoff = time.time() - 48 * 3600  # حاشیهٔ ایمنی در برابر race با آپلود تازه
            for fname in os.listdir(avatars_dir):
                fpath = os.path.join(avatars_dir, fname)
                if not os.path.isfile(fpath) or fname in kept:
                    continue
                try:
                    if os.path.getmtime(fpath) > cutoff:
                        continue
                    sz = os.path.getsize(fpath)
                    os.remove(fpath)
                    av_freed += sz
                    av_count += 1
                except OSError:
                    pass
    except Exception as ex:
        _tg_logger.error("cache cleanup avatars error: %s", ex)
    freed_bytes += av_freed
    details["orphan_avatars"] = {"freed_bytes": av_freed, "files": av_count}

    now_iso = datetime.now().isoformat()
    report = {"ran_at": now_iso, "freed_bytes": freed_bytes, "details": details}
    try:
        from db import set_cfg
        set_cfg("CACHE_CLEANUP_LAST_RUN", now_iso)
        set_cfg("CACHE_CLEANUP_LAST_REPORT", json.dumps(report, ensure_ascii=False))
    except Exception:
        pass
    _tg_logger.info("Cache cleanup done: %s bytes freed (%s)", freed_bytes, details)
    return report


def _start_cache_cleanup_thread() -> None:
    global _cache_cleanup_started
    if _cache_cleanup_started:
        return
    _cache_cleanup_started = True

    def _runner():
        import datetime as _dt
        from db import get_cfg
        while True:
            try:
                enabled = get_cfg("CACHE_CLEANUP_ENABLED", "0") == "1"
            except Exception:
                enabled = False
            if not enabled:
                _time.sleep(3600)  # هر ساعت چک کن که از پنل فعال شده یا نه
                continue

            mode = get_cfg("CACHE_CLEANUP_MODE", "daily")

            if mode == "interval":
                try:
                    hours = max(1, int(get_cfg("CACHE_CLEANUP_INTERVAL_HOURS", "24") or "24"))
                except Exception:
                    hours = 24
                _time.sleep(hours * 3600)
                try:
                    if get_cfg("CACHE_CLEANUP_ENABLED", "0") == "1":
                        run_cache_cleanup()
                except Exception as ex:
                    _tg_logger.error("cache cleanup scheduled run error: %s", ex)
                continue

            now = _dt.datetime.now()
            try:
                hour = max(0, min(23, int(get_cfg("CACHE_CLEANUP_HOUR", "4") or "4")))
            except Exception:
                hour = 4
            target = now.replace(hour=hour, minute=0, second=0, microsecond=0)

            if mode == "weekly":
                try:
                    weekday = max(0, min(6, int(get_cfg("CACHE_CLEANUP_WEEKDAY", "0") or "0")))
                except Exception:
                    weekday = 0
                days_ahead = (weekday - now.weekday()) % 7
                target += _dt.timedelta(days=days_ahead)
                if target <= now:
                    target += _dt.timedelta(days=7)
            else:  # daily
                if target <= now:
                    target += _dt.timedelta(days=1)

            sleep_secs = max(60, (target - now).total_seconds())
            _time.sleep(sleep_secs)
            try:
                if get_cfg("CACHE_CLEANUP_ENABLED", "0") == "1":
                    run_cache_cleanup()
            except Exception as ex:
                _tg_logger.error("cache cleanup scheduled run error: %s", ex)

    _threading.Thread(target=_runner, name="cache-cleanup", daemon=True).start()


def _fmt_bytes(n: int) -> str:
    n = int(n or 0)
    if n < 1024:
        return f"{n} بایت"
    if n < 1024 * 1024:
        return f"{n/1024:.1f} KB"
    return f"{n/1024/1024:.1f} MB"


@router.get("/system-cache", response_class=HTMLResponse)
async def system_cache_page(request: Request, flash: str = ""):
    adm = _get_admin(request)
    guard = _require(adm, "cache_cleanup")
    if guard: return guard
    from db import get_cfg

    enabled = get_cfg("CACHE_CLEANUP_ENABLED", "0") == "1"
    mode = get_cfg("CACHE_CLEANUP_MODE", "daily")
    hour = get_cfg("CACHE_CLEANUP_HOUR", "4")
    weekday = get_cfg("CACHE_CLEANUP_WEEKDAY", "0")
    interval_hours = get_cfg("CACHE_CLEANUP_INTERVAL_HOURS", "24")
    last_run = get_cfg("CACHE_CLEANUP_LAST_RUN", "")

    last_report_html = '<div class="text-sm text-gray-400">هنوز اجرا نشده.</div>'
    try:
        raw = get_cfg("CACHE_CLEANUP_LAST_REPORT", "")
        if raw:
            rep = json.loads(raw)
            freed = _fmt_bytes(rep.get("freed_bytes", 0))
            det = rep.get("details", {})
            py = det.get("pycache", {})
            av = det.get("orphan_avatars", {})
            last_report_html = f"""
              <div class="text-sm space-y-1">
                <div>🕐 آخرین اجرا: <b class="ltr-num">{e(fa_date(rep.get('ran_at',''), with_time=True))}</b></div>
                <div>💾 فضای آزادشده: <b class="text-emerald-600 ltr-num">{e(freed)}</b></div>
                <div class="text-xs text-gray-400">کش پایتون: {py.get('files',0)} فایل ({_fmt_bytes(py.get('freed_bytes',0))}) —
                  آواتار یتیم: {av.get('files',0)} فایل ({_fmt_bytes(av.get('freed_bytes',0))})</div>
              </div>"""
    except Exception:
        pass

    # Python weekday(): دوشنبه=۰ ... یکشنبه=۶ (میلادی) — چون _cache_cleanup_runner از
    # همین datetime.weekday() استفاده می‌کنه، ترتیب برچسب‌ها باید دقیقاً همین باشه.
    weekday_opts = ["دوشنبه", "سه‌شنبه", "چهارشنبه", "پنج‌شنبه", "جمعه", "شنبه", "یکشنبه"]
    weekday_select = "".join(
        f'<option value="{i}" {"selected" if str(i)==weekday else ""}>{lbl}</option>'
        for i, lbl in enumerate(weekday_opts)
    )
    hour_select = "".join(
        f'<option value="{h}" {"selected" if str(h)==hour else ""}>ساعت {h}:۰۰</option>'
        for h in range(24)
    )

    body = f"""
    <div class="text-xs text-gray-400 mb-4">
      پاکسازی خودکار فقط داده‌های غیرضروری/حجیم (کش بایت‌کد پایتون + آواتار یتیم) رو حذف می‌کنه —
      دیتابیس، بکاپ‌ها، تنظیمات و اطلاعات کاربران هرگز دست‌نخورده می‌مونن. این پروژه در عمل کش/فایل
      موقت زیادی تولید نمی‌کنه، پس انتظار نداشته باشید عدد بزرگی آزاد بشه — این ابزار برای نظافت
      دوره‌ایه، نه رفع مشکل کمبود فضا.
    </div>

    <div class="card p-5 mb-4">
      <h3 class="font-bold text-sm mb-3">📊 آخرین اجرا</h3>
      {last_report_html}
      <form method="post" action="/admin/system-cache/run-now" class="mt-4">
        <button class="bg-indigo-600 hover:bg-indigo-700 text-white px-4 py-2 rounded-xl text-sm font-medium">
          🧹 پاکسازی فوری کش
        </button>
      </form>
    </div>

    <div class="card p-5">
      <h3 class="font-bold text-sm mb-3">⚙️ زمان‌بندی خودکار</h3>
      <form method="post" action="/admin/system-cache/save-settings" class="space-y-4">
        <label class="flex items-center gap-2 text-sm">
          <input type="checkbox" name="enabled" value="1" {"checked" if enabled else ""}>
          فعال‌سازی پاکسازی خودکار
        </label>

        <div>
          <label class="text-xs text-gray-500 block mb-1">بازهٔ اجرا</label>
          <select name="mode" id="cc-mode" onchange="
            document.getElementById('cc-daily-hour').style.display = this.value!=='interval' ? 'block':'none';
            document.getElementById('cc-weekday').style.display = this.value==='weekly' ? 'block':'none';
            document.getElementById('cc-interval').style.display = this.value==='interval' ? 'block':'none';
          " class="w-full border border-gray-200 rounded-xl px-3 py-2 text-sm">
            <option value="daily" {"selected" if mode=="daily" else ""}>روزانه</option>
            <option value="weekly" {"selected" if mode=="weekly" else ""}>هفتگی</option>
            <option value="interval" {"selected" if mode=="interval" else ""}>بازهٔ سفارشی (هر N ساعت)</option>
          </select>
        </div>

        <div id="cc-daily-hour" style="display:{'block' if mode!='interval' else 'none'}">
          <label class="text-xs text-gray-500 block mb-1">ساعت اجرا</label>
          <select name="hour" class="w-full border border-gray-200 rounded-xl px-3 py-2 text-sm">{hour_select}</select>
        </div>

        <div id="cc-weekday" style="display:{'block' if mode=='weekly' else 'none'}">
          <label class="text-xs text-gray-500 block mb-1">روز هفته (فقط برای حالت هفتگی)</label>
          <select name="weekday" class="w-full border border-gray-200 rounded-xl px-3 py-2 text-sm">{weekday_select}</select>
        </div>

        <div id="cc-interval" style="display:{'block' if mode=='interval' else 'none'}">
          <label class="text-xs text-gray-500 block mb-1">هر چند ساعت یک‌بار</label>
          <input type="text" inputmode="numeric" name="interval_hours" value="{e(interval_hours)}"
            class="w-full border border-gray-200 rounded-xl px-3 py-2 text-sm ltr-num" style="direction:ltr;text-align:left">
        </div>

        <button type="submit" class="bg-emerald-600 hover:bg-emerald-700 text-white px-4 py-2 rounded-xl text-sm font-medium">
          💾 ذخیرهٔ تنظیمات
        </button>
      </form>
    </div>"""
    return _layout("پاکسازی کش و فایل‌های موقت", body, adm, flash=flash)


@router.post("/system-cache/save-settings")
async def system_cache_save_settings(request: Request):
    adm = _get_admin(request)
    guard = _require(adm, "cache_cleanup")
    if guard: return guard
    from db import set_cfg

    form = await request.form()
    enabled = "1" if form.get("enabled") == "1" else "0"
    mode = str(form.get("mode") or "daily")
    if mode not in ("daily", "weekly", "interval"):
        mode = "daily"
    try:
        hour = max(0, min(23, int(form.get("hour") or 4)))
    except Exception:
        hour = 4
    try:
        weekday = max(0, min(6, int(form.get("weekday") or 0)))
    except Exception:
        weekday = 0
    try:
        interval_hours = max(1, int(form.get("interval_hours") or 24))
    except Exception:
        interval_hours = 24

    set_cfg("CACHE_CLEANUP_ENABLED", enabled)
    set_cfg("CACHE_CLEANUP_MODE", mode)
    set_cfg("CACHE_CLEANUP_HOUR", str(hour))
    set_cfg("CACHE_CLEANUP_WEEKDAY", str(weekday))
    set_cfg("CACHE_CLEANUP_INTERVAL_HOURS", str(interval_hours))

    _log(request, "ذخیرهٔ تنظیمات پاکسازی کش", "سیستم", f"enabled={enabled} mode={mode}", admin_info=adm)
    return _redir(f"/admin/system-cache?flash={e('✅ تنظیمات ذخیره شد')}")


@router.post("/system-cache/run-now")
async def system_cache_run_now(request: Request):
    adm = _get_admin(request)
    guard = _require(adm, "cache_cleanup")
    if guard: return guard

    report = await run_in_threadpool(run_cache_cleanup)
    _log(request, "پاکسازی فوری کش", "سیستم", f"freed={report.get('freed_bytes',0)} bytes", admin_info=adm)
    freed_str = _fmt_bytes(report.get("freed_bytes", 0)).replace(" ", "+")
    return _redir(f"/admin/system-cache?flash=✅+پاکسازی+انجام+شد+—+{freed_str}+آزاد+شد")


def _do_auto_backup() -> None:
    """بکاپ خودکار روزانه — بر اساس DB_DIALECT سوییچ می‌کنه.
    ⚠️ تاریخچه (رفت‌وبرگشت مستند در CLAUDE.md): این تابع قبلاً pg_backup صدا
    می‌زد، بعد چون تولید SQLite بود به stbak_engine تغییر کرد. حالا که پروژه
    واقعاً به Postgres مهاجرت کرده، دوباره باید pg_backup صدا بزنه — وگرنه
    stbak_engine روی فایل SQLite فانتوم/غیرمرتبط با دادهٔ واقعی بکاپ می‌گیره
    (بکاپ ظاهراً موفق، ولی از دادهٔ واقعی خالیه — خطرناک‌تر از شکست آشکار)."""
    import db_conn as _dbc
    if _dbc.is_postgres():
        try:
            import pg_backup
            rep = pg_backup.run_full_backup()
            if not rep.get("ok"):
                _tg_logger.error("Auto-backup (Postgres) failed: %s", rep.get("error"))
                return
            _tg_logger.info("Auto-backup (Postgres) done: %s", rep.get("file"))
        except Exception as ex:
            _tg_logger.error("Auto-backup (Postgres) failed: %s", ex)
        return

    try:
        from stbak_engine import save_local_backup
        try:
            from backup_uploader import get_cloud_settings
            _cs = get_cloud_settings()
            retention = max(1, int(_cs.get("retention") or _MAX_BACKUPS))
        except Exception:
            retention = _MAX_BACKUPS
        dst = save_local_backup(_DB_PATH(), _BACKUP_DIR, modules=None, retention=retention)
    except Exception as ex:
        _tg_logger.error("Auto-backup failed: %s", ex)
        return
    _tg_logger.info("Auto-backup done: %s", dst)

    # ANALYZE هفتگی — برنامه‌ریز کوئری SQLite بدون آمار به‌روز، با رشد دیتابیس
    # تصمیمات کمتر بهینه می‌گیره؛ همراه همین جاب روزانهٔ بکاپ (فقط دوشنبه‌ها) اجرا می‌شه.
    # ⚠️ فقط زیر SQLite معنی داره — Postgres خودش autovacuum/analyze خودکار داره،
    # نیازی به کرون دستی نیست (بخش پاک‌سازی SQLite سند).
    try:
        import db_conn as _dbc
        if not _dbc.is_postgres():
            import datetime as _dt
            if _dt.datetime.now().weekday() == 0:
                _aconn = sqlite3.connect(_DB_PATH(), timeout=30)
                _aconn.execute("ANALYZE;")
                _aconn.close()
                _tg_logger.info("Weekly ANALYZE done")
    except Exception as ex:
        _tg_logger.error("Weekly ANALYZE failed: %s", ex)

    # ☁️ آپلود ابری (کانال تلگرام + گوگل درایو) — غیرهمزمان
    try:
        from backup_uploader import get_cloud_settings, upload_backup
        cs = get_cloud_settings()
        if int(cs.get("tg_enabled") or 0) or int(cs.get("gdrive_enabled") or 0):
            rep = upload_backup(dst)
            _tg_logger.info("Cloud backup queued: %s", rep)
    except Exception as ex:
        _tg_logger.error("Cloud backup failed: %s", ex)


def _start_auto_backup_thread() -> None:
    global _auto_backup_started
    if _auto_backup_started:
        return
    _auto_backup_started = True

    def _runner():
        import datetime as _dt
        while True:
            now = _dt.datetime.now()
            try:
                from backup_uploader import get_cloud_settings
                _h = int(get_cloud_settings().get("hour") or 4)
            except Exception:
                _h = 4
            target = now.replace(hour=max(0, min(23, _h)), minute=0, second=0, microsecond=0)
            if now >= target:
                target += _dt.timedelta(days=1)
            sleep_secs = (target - now).total_seconds()
            _time.sleep(sleep_secs)
            try:
                _do_auto_backup()
            except Exception as ex:
                _tg_logger.error("auto-backup error: %s", ex)

    _threading.Thread(target=_runner, name="auto-backup", daemon=True).start()

    # ⚠️ نگهبان بکاپ ابری — ساعتی چک، در صورت مشکل هشدار به ادمین
    def _cloud_watchdog():
        while True:
            _time.sleep(3600)
            try:
                from backup_uploader import watchdog_check
                msg = watchdog_check()
                if msg:
                    try:
                        from config import ADMIN_ID as _AID
                    except Exception:
                        _AID = int(_env("ADMIN_ID", "0") or 0)
                    if _AID:
                        _tg_api_send(_AID, msg)
            except Exception as ex:
                _tg_logger.error("cloud watchdog: %s", ex)

    _threading.Thread(target=_cloud_watchdog, name="cloud-watchdog", daemon=True).start()

    # thread جداگانه برای بررسی موجودی (هر ۲ ساعت)
    def _stock_runner():
        while True:
            _time.sleep(7200)
            try:
                _notify_low_stock()
            except Exception as ex:
                _tg_logger.error("low-stock check error: %s", ex)

    _threading.Thread(target=_stock_runner, name="stock-check", daemon=True).start()

    _start_cache_cleanup_thread()
    _tg_logger.info("Scheduler started (backup:24h, low-stock:2h, cache-cleanup:configurable)")


@router.get("/reports", response_class=HTMLResponse)
async def financial_report(request: Request):
    """آدرس قدیمی «گزارش‌های مالی» — طبق درخواست صریح مالک پروژه محتوایش با صفحهٔ
    حسابداری (که فرمول‌های دقیق‌تری داره، مثلاً هزینه خرید فقط روی کالای واقعاً
    تحویل‌شده) ادغام شد؛ این مسیر فقط برای لینک/بوکمارک‌های قدیمی نگه داشته شده."""
    return _redir("/admin/accounting")


# ══════════════════════════════════════════════════════════════════════════════
# ─── حسابداری (Light Accounting) ─────────────────────────────────────────────
# ══════════════════════════════════════════════════════════════════════════════

def _to_jalali(greg_str: str) -> str:
    """تبدیل تاریخ میلادی به شمسی."""
    try:
        from datetime import date as _d
        y, m, d = map(int, greg_str.split('-'))
        # الگوریتم ساده تبدیل
        g_y = y - 1600; g_m = m - 1; g_d = d - 1
        g_d_no = 365*g_y + (g_y+3)//4 - (g_y+99)//100 + (g_y+399)//400
        for i in range(g_m):
            g_d_no += [31,28+1 if (y%4==0 and y%100!=0) or y%400==0 else 28,31,30,31,30,31,31,30,31,30,31][i]
        g_d_no += g_d
        j_d_no = g_d_no - 79
        j_np = j_d_no // 12053; j_d_no %= 12053
        jy = 979 + 33*j_np + 4*(j_d_no//1461); j_d_no %= 1461
        if j_d_no >= 366:
            jy += (j_d_no-1)//365; j_d_no = (j_d_no-1)%365
        for i,v in enumerate([31,31,31,31,31,31,30,30,30,30,30,29]):
            if j_d_no >= v: j_d_no -= v
            else: jm = i+1; jd = j_d_no+1; break
        return f"{jy}/{jm:02d}/{jd:02d}"
    except Exception:
        return greg_str


def _month_start() -> str:
    from datetime import date, timedelta
    d = date.today(); return d.replace(day=1).isoformat()


def _week_start() -> str:
    from datetime import date, timedelta
    d = date.today(); return (d - timedelta(days=d.weekday())).isoformat()


@router.get("/accounting", response_class=HTMLResponse)
async def accounting_dashboard(request: Request, df: str = "", dt: str = "", df_fa: str = "", dt_fa: str = "", flash: str = ""):
    # df_fa/dt_fa (شمسی، از فرم فیلتر) در سمت سرور به میلادی تبدیل می‌شن؛ df/dt هم مستقیم می‌تونن میلادی باشن (لینک‌های میان‌بر)
    from db import jalali_str_to_gregorian_iso
    if df_fa: df = jalali_str_to_gregorian_iso(df_fa) or df
    if dt_fa: dt = jalali_str_to_gregorian_iso(dt_fa) or dt
    adm = _get_admin(request)
    guard = _require(adm, "accounting")
    if guard: return guard
    from db import get_accounting_kpis, ensure_accounting_schema
    ensure_accounting_schema()
    kpis = get_accounting_kpis(df, dt)
    def _m(n): return f"{int(n):,}"
    today_g = __import__('datetime').date.today().isoformat()
    df = df or today_g; dt = dt or today_g
    filter_html = f"""
    <form method="get" class="flex gap-2 items-center mb-6 flex-wrap">
      <div class="flex items-center gap-1 bg-white border border-gray-200 rounded-lg px-2 py-1">
        <span class="text-xs text-gray-400 whitespace-nowrap">از:</span>
        <input type="text" name="df_fa" value="{fa_date(df)}" placeholder="۱۴۰۴/۰۱/۰۱"
          class="w-28 text-sm outline-none" autocomplete="off">
        <input type="hidden" name="df" value="{df}">
      </div>
      <div class="flex items-center gap-1 bg-white border border-gray-200 rounded-lg px-2 py-1">
        <span class="text-xs text-gray-400 whitespace-nowrap">تا:</span>
        <input type="text" name="dt_fa" value="{fa_date(dt)}" placeholder="۱۴۰۴/۰۱/۳۱"
          class="w-28 text-sm outline-none" autocomplete="off">
        <input type="hidden" name="dt" value="{dt}">
      </div>
      <div class="flex gap-1">
        <button type="submit" class="px-3 py-1.5 bg-indigo-600 text-white rounded-lg text-xs font-medium">فیلتر</button>
        <a href="/admin/accounting" class="px-3 py-1.5 bg-gray-100 text-gray-600 rounded-lg text-xs">امروز</a>
        <a href="/admin/accounting?df=&dt=" class="px-3 py-1.5 bg-gray-50 text-gray-400 rounded-lg text-xs">همه</a>
      </div>
      <div class="flex gap-1">
        <a href="/admin/accounting?df={_month_start()}&dt={today_g}" class="px-2 py-1.5 bg-blue-50 text-blue-600 rounded text-xs">این ماه</a>
        <a href="/admin/accounting?df={_week_start()}&dt={today_g}" class="px-2 py-1.5 bg-blue-50 text-blue-600 rounded text-xs">این هفته</a>
        <a href="/admin/accounting?df=&dt=" class="px-2 py-1.5 bg-blue-50 text-blue-600 rounded text-xs">کل</a>
      </div>
    </form>"""
    body = f"""
    <div class="flex items-center justify-between mb-4 flex-wrap gap-2">
      <h1 class="text-2xl font-bold text-gray-800">💰 حسابداری</h1>
      <div class="flex gap-1 flex-wrap">
        <a href="/admin/accounting/expenses" class="px-3 py-1.5 text-xs bg-amber-50 text-amber-700 border border-amber-200 rounded-lg">📋 ثبت هزینه/پرداخت</a>
        <a href="/admin/accounting/cashflow" class="px-3 py-1.5 text-xs bg-indigo-50 text-indigo-700 border border-indigo-200 rounded-lg">🔄 گردش مالی</a>
        <a href="/admin/accounting/products" class="px-3 py-1.5 text-xs bg-green-50 text-green-700 border border-green-200 rounded-lg">📦 محصولات</a>
        <a href="/admin/accounting/partners" class="px-3 py-1.5 text-xs bg-purple-50 text-purple-700 border border-purple-200 rounded-lg">🤝 همکاران</a>
        <a href="/admin/accounting/exchanges" class="px-3 py-1.5 text-xs bg-orange-50 text-orange-700 border border-orange-200 rounded-lg">🔄 تعویض‌ها</a>
      </div>
    </div>
    {{filter_html}}

    <!-- مانده صندوق — برجسته -->
    <div class="card p-5 mb-4 bg-gradient-to-l from-emerald-50 to-transparent border-r-4 border-emerald-400">
      <div class="flex items-center justify-between flex-wrap gap-2">
        <div>
          <div class="text-xs text-gray-500 mb-1">مانده صندوق</div>
          <div class="text-3xl font-bold text-emerald-700 no-fa ltr-left">{_m(max(0,kpis["net_profit"]-kpis.get("total_payouts",0)))} <span class="text-sm font-normal text-gray-400">تومان</span></div>
        </div>
        <div class="text-xs text-gray-500 leading-6 text-left">
          <div>سود خالص: <b class="text-gray-700 no-fa">{_m(kpis["net_profit"])}</b></div>
          <div>پرداخت‌شده: <b class="text-gray-700 no-fa">-{_m(kpis.get("total_payouts",0))}</b></div>
        </div>
      </div>
    </div>

    <!-- فروش مستقیم در برابر همکاری -->
    <div class="grid grid-cols-2 gap-3 mb-4">
      <div class="card p-3">
        <div class="text-[10px] text-gray-400 mb-1">🔵 فروش مستقیم</div>
        <div class="font-bold text-blue-700 no-fa ltr-left">{_m(kpis.get("direct_sales",0))} <span class="text-[10px] font-normal text-gray-400">ت</span></div>
      </div>
      <div class="card p-3">
        <div class="text-[10px] text-gray-400 mb-1">🤝 فروش همکاری</div>
        <div class="font-bold text-purple-700 no-fa ltr-left">{_m(kpis.get("partner_sales",0))} <span class="text-[10px] font-normal text-gray-400">ت</span></div>
      </div>
    </div>

    <!-- جدول خلاصه سود و زیان — سبک، مثل صورت‌مالی -->
    <div class="card overflow-hidden mb-4">
      <div class="px-4 py-3 border-b bg-gray-50 flex items-center justify-between">
        <h2 class="font-bold text-gray-700 text-sm">📊 صورت سود و زیان</h2>
        <span class="text-xs text-gray-400">حاشیه سود: <b class="text-emerald-600">٪{kpis["margin_pct"]}</b></span>
      </div>
      <table class="w-full text-sm">
        <tbody>
          <tr class="border-b hover:bg-gray-50"><td class="px-4 py-2.5 text-gray-600">فروش کل ({_m(kpis["total_orders"])} سفارش)</td><td class="px-4 py-2.5 text-left font-semibold text-blue-700 no-fa">{_m(kpis["total_sales"])}</td></tr>
          <tr class="border-b hover:bg-gray-50 bg-red-50/30"><td class="px-4 py-2.5 text-gray-600">− هزینه خرید کالا</td><td class="px-4 py-2.5 text-left text-red-600 no-fa">-{_m(kpis["total_cost"])}</td></tr>
          <tr class="border-b hover:bg-gray-50 bg-red-50/30"><td class="px-4 py-2.5 text-gray-600">− پورسانت همکاران</td><td class="px-4 py-2.5 text-left text-red-600 no-fa">-{_m(kpis["total_commission"])}</td></tr>
          <tr class="border-b hover:bg-gray-50 bg-red-50/30"><td class="px-4 py-2.5 text-gray-600">− هزینه‌های عمومی</td><td class="px-4 py-2.5 text-left text-red-600 no-fa">-{_m(kpis["total_expenses"])}</td></tr>
          <tr class="border-b-2 border-emerald-200 bg-emerald-50"><td class="px-4 py-3 font-bold text-emerald-700">= سود خالص</td><td class="px-4 py-3 text-left font-bold text-emerald-700 text-lg no-fa">{_m(kpis["net_profit"])}</td></tr>
        </tbody>
      </table>
    </div>

    <!-- گزیده‌ی سریع — ۴ ستون کوچک جمع‌وجور -->
    <div class="card p-3 mb-4">
      <div class="grid grid-cols-2 md:grid-cols-4 gap-2 text-sm">
        <div class="p-2 border-l border-gray-100">
          <div class="text-[10px] text-gray-400 mb-1">فروش امروز</div>
          <div class="font-bold text-blue-700 no-fa ltr-left">{_m(kpis["today_sales"])} <span class="text-[10px] font-normal text-gray-400">ت</span></div>
        </div>
        <div class="p-2 border-l border-gray-100">
          <div class="text-[10px] text-gray-400 mb-1">فروش این ماه</div>
          <div class="font-bold text-indigo-700 no-fa ltr-left">{_m(kpis["month_sales"])} <span class="text-[10px] font-normal text-gray-400">ت</span></div>
        </div>
        <div class="p-2 border-l border-gray-100">
          <div class="text-[10px] text-gray-400 mb-1">میانگین سود/سفارش</div>
          <div class="font-bold text-teal-700 no-fa ltr-left">{_m(kpis["avg_profit"])} <span class="text-[10px] font-normal text-gray-400">ت</span></div>
        </div>
        <div class="p-2">
          <div class="text-[10px] text-gray-400 mb-1">موجودی انبار</div>
          <div class="font-bold text-slate-700 no-fa ltr-left">{_m(kpis["stock_count"])} <span class="text-[10px] font-normal text-gray-400">آیتم</span></div>
          <div class="text-[10px] text-gray-400 mt-0.5 no-fa ltr-left">ارزش: {_m(kpis.get("stock_value",0))} ت</div>
        </div>
      </div>
    </div>

    <!-- پرداخت‌های تسویه‌شده -->
    <div class="card p-3 flex items-center justify-between text-sm bg-purple-50/40">
      <div class="flex items-center gap-2">
        <span class="text-lg">💸</span>
        <div>
          <div class="text-xs text-gray-500">مجموع تسویه‌های پرداخت‌شده به همکاران</div>
          <div class="font-bold text-purple-700 no-fa ltr-left">{_m(kpis.get("total_payouts",0))} تومان</div>
        </div>
      </div>
      <a href="/admin/partners?tab=payouts" class="text-xs text-purple-600 hover:underline">مشاهده تسویه‌ها →</a>
    </div>"""
    return _layout("حسابداری", body.replace("{filter_html}", filter_html), adm, flash=flash)


def _acbar(label, value, total, color):
    pct = max(0, min(100, int(value/total*100) if total>0 else 0))
    return f"""<div><div class="flex justify-between text-xs text-gray-500 mb-1">
      <span>{label}</span><span>{int(value):,} ت ({pct}٪)</span></div>
      <div class="h-2 bg-gray-100 rounded-full ltr-num"><div class="{color} h-2 rounded-full" style="width:{pct}%"></div></div></div>"""


@router.get("/accounting/expenses", response_class=HTMLResponse)
async def accounting_expenses(request: Request, cat: str="", df: str="", dt: str="", flash: str=""):
    adm = _get_admin(request)
    guard = _require(adm, "accounting")
    if guard: return guard
    from db import get_expenses, get_expense_categories, ensure_accounting_schema
    ensure_accounting_schema()
    cats = get_expense_categories()
    expenses = get_expenses(df, dt, cat)
    total = sum(ex["amount"] for ex in expenses)
    cat_opts = "".join(f'<option value="{c}" {"selected" if cat==c else ""}>{c}</option>' for c in cats)
    _PT_BADGE = {
        "salary": '<span class="px-2 py-0.5 bg-blue-50 text-blue-700 rounded text-[10px]">👤 حقوق</span>',
        "partner_payout": '<span class="px-2 py-0.5 bg-purple-50 text-purple-700 rounded text-[10px]">🤝 همکار</span>',
        "other": '<span class="px-2 py-0.5 bg-gray-100 text-gray-600 rounded text-[10px]">📌 سایر</span>',
    }
    def _ptb(ex):
        try: return _PT_BADGE.get(ex["payment_type"] or "expense",
            '<span class="px-2 py-0.5 bg-amber-50 text-amber-700 rounded text-[10px]">💳 هزینه</span>')
        except Exception:
            return '<span class="px-2 py-0.5 bg-amber-50 text-amber-700 rounded text-[10px]">💳 هزینه</span>'
    def _payee(ex):
        try: return e((ex["payee_name"] or "")[:20])
        except Exception: return ""
    rows = "".join(f'''<tr class="border-b hover:bg-gray-50 text-sm">
      <td class="px-3 py-2 text-xs text-gray-400">{fa_date(ex["expense_date"])}</td>
      <td class="px-3 py-2">{_ptb(ex)}</td>
      <td class="px-3 py-2 font-medium">{e(ex["title"])}</td>
      <td class="px-3 py-2 text-xs text-gray-500">{_payee(ex)}</td>
      <td class="px-3 py-2"><span class="px-2 py-0.5 bg-gray-100 text-gray-600 rounded text-xs">{e(ex["category"])}</span></td>
      <td class="px-3 py-2 font-bold text-red-600 no-fa ltr-left">{int(ex["amount"]):,}</td>
      <td class="px-3 py-2 text-xs text-gray-400">{e((ex["description"] or "")[:30])}</td>
      <td class="px-3 py-2"><form method="post" action="/admin/accounting/expenses/{ex["id"]}/delete" onsubmit="return confirm(\'حذف؟\')"><button class="text-xs text-red-400 hover:text-red-600">حذف</button></form></td>
    </tr>''' for ex in expenses) or "<tr><td colspan='8' class='text-center py-6 text-gray-400'>پرداختی ثبت نشده</td></tr>"
    body = f"""
    <div class="flex items-center gap-3 mb-6">
      {_btn("← حسابداری","/admin/accounting","slate",small=True)}
      <h1 class="text-2xl font-bold text-gray-800">📋 ثبت هزینه و پرداخت‌ها</h1>
    </div>
    <div class="grid md:grid-cols-2 gap-4 mb-6">
      <div class="card p-6">
        <h2 class="font-bold text-gray-700 mb-4">+ ثبت پرداخت جدید</h2>
        <form method="post" action="/admin/accounting/expenses/new" class="space-y-3">
          <div><label class="text-xs block mb-1">نوع پرداخت</label>
            <select name="payment_type" class="w-full border rounded-lg px-3 py-2 text-sm">
              <option value="expense">💳 هزینه عمومی</option>
              <option value="salary">👤 حقوق پرسنل</option>
              <option value="partner_payout">🤝 پرداخت همکار</option>
              <option value="other">📌 سایر پرداخت‌ها</option>
            </select></div>
          <div class="grid grid-cols-2 gap-3">
            <div><label class="text-xs block mb-1">عنوان</label><input type="text" name="title" required class="w-full border rounded-lg px-3 py-2 text-sm"></div>
            <div><label class="text-xs block mb-1">مبلغ (تومان)</label><input type="number" name="amount" required class="w-full border rounded-lg px-3 py-2 text-sm"></div>
            <div><label class="text-xs block mb-1">دسته</label><select name="category" class="w-full border rounded-lg px-3 py-2 text-sm"><option value="">انتخاب...</option>{cat_opts}</select></div>
            <div><label class="text-xs block mb-1">تاریخ</label><input type="date" name="expense_date" class="w-full border rounded-lg px-3 py-2 text-sm"></div>
          </div>
          <input type="text" name="payee_name" placeholder="نام گیرنده (پرسنل/همکار — اختیاری)" class="w-full border rounded-lg px-3 py-2 text-sm">
          <input type="text" name="description" placeholder="توضیحات (اختیاری)" class="w-full border rounded-lg px-3 py-2 text-sm">
          {_btn("💾 ثبت پرداخت","",color="red",small=True)}
        </form>
      </div>
      <div class="card p-6">
        <h2 class="font-bold text-gray-700 mb-3">+ دسته‌بندی جدید</h2>
        <form method="post" action="/admin/accounting/categories/new" class="flex gap-2 mb-4">
          <input type="text" name="name" placeholder="نام دسته" class="flex-1 border rounded-lg px-3 py-2 text-sm">
          {_btn("اضافه","",color="indigo",small=True)}
        </form>
        <div class="flex flex-wrap gap-1">{" ".join(f'<span class="px-2 py-1 bg-gray-100 text-gray-600 rounded text-xs">{c}</span>' for c in cats)}</div>
      </div>
    </div>
    <div class="card p-6 mb-6">
      <h2 class="font-bold text-gray-700 mb-1">📁 وارد کردن دسته‌ای از فایل</h2>
      <p class="text-xs text-gray-400 mb-4">فایل TXT/CSV/Excel با ستون‌های عنوان، مبلغ، دسته (اختیاری)، تاریخ (اختیاری)، توضیحات (اختیاری) — هر ردیف یک هزینه/پرداخت ثبت می‌کنه.</p>
      <form method="post" action="/admin/accounting/expenses/import" enctype="multipart/form-data" class="flex flex-wrap items-center gap-3">
        <input type="file" name="file" accept=".txt,.csv,.xlsx,.xlsm" required
          class="flex-1 min-w-[200px] text-sm text-gray-600 file:ml-2 file:py-2 file:px-4 file:rounded-lg file:border-0 file:text-sm file:bg-indigo-50 file:text-indigo-700 hover:file:bg-indigo-100">
        <select name="default_payment_type" class="border rounded-lg px-3 py-2 text-sm">
          <option value="expense">💳 پیش‌فرض: هزینه عمومی</option>
          <option value="salary">👤 پیش‌فرض: حقوق پرسنل</option>
          <option value="partner_payout">🤝 پیش‌فرض: پرداخت همکار</option>
          <option value="other">📌 پیش‌فرض: سایر</option>
        </select>
        {_btn("⬆ وارد کردن","",color="indigo",small=True)}
      </form>
    </div>
    <div class="card overflow-hidden">
      <div class="px-4 py-3 bg-gray-50 border-b flex flex-wrap justify-between items-center gap-2">
        <form method="get" class="flex gap-2 flex-wrap">
          <input type="date" name="df" value="{df}" class="border rounded px-2 py-1 text-xs">
          <input type="date" name="dt" value="{dt}" class="border rounded px-2 py-1 text-xs">
          <select name="cat" class="border rounded px-2 py-1 text-xs"><option value="">همه</option>{cat_opts}</select>
          <button class="px-3 py-1 bg-indigo-600 text-white rounded text-xs">فیلتر</button>
        </form>
        <div class="flex gap-2 items-center">
          <span class="text-sm font-bold text-red-600">جمع: {total:,} ت</span>
          <a href="/admin/accounting/expenses/export?df={df}&dt={dt}&cat={cat}" class="px-3 py-1 bg-green-50 text-green-700 border border-green-200 rounded text-xs">⬇ Excel</a>
        </div>
      </div>
      <div class="overflow-x-auto"><table class="w-full text-right min-w-max">
        <thead><tr class="text-xs text-gray-500 border-b bg-gray-50">
          <th class="px-3 py-2">تاریخ</th><th class="px-3 py-2">نوع</th><th class="px-3 py-2">عنوان</th>
          <th class="px-3 py-2">گیرنده</th><th class="px-3 py-2">دسته</th>
          <th class="px-3 py-2">مبلغ</th><th class="px-3 py-2">توضیح</th><th></th>
        </tr></thead><tbody>{rows}</tbody>
      </table></div>
    </div>"""
    return _layout("هزینه‌ها", body, adm, flash=flash)


@router.post("/accounting/expenses/new")
async def accounting_expense_new(request: Request):
    adm = _get_admin(request); guard = _require(adm, "accounting")
    if guard: return guard
    from db import create_expense, ensure_accounting_schema
    ensure_accounting_schema()
    form = await request.form()
    ptype = str(form.get("payment_type","expense")).strip()
    if ptype not in ("expense","salary","partner_payout","other"): ptype = "expense"
    eid = create_expense(str(form.get("title","")).strip(), str(form.get("category","سایر")),
                         int(form.get("amount") or 0), str(form.get("expense_date","")),
                         str(form.get("description","")),
                         payment_type=ptype,
                         payee_name=str(form.get("payee_name","")).strip())
    _log(request, "ثبت پرداخت", "حسابداری", f"id:{eid} نوع:{ptype}", admin_info=adm)
    return _redir("/admin/accounting/expenses?flash=✅+پرداخت+ثبت+شد")


@router.post("/accounting/expenses/import")
async def accounting_expenses_import(request: Request, file: UploadFile = None):
    """وارد کردن دسته‌ای هزینه/پرداخت از فایل TXT/CSV/Excel — با import_utils.parse_uploaded_rows
    (همون پارسر مشترکی که آپلود موجودی محصول هم استفاده می‌کنه). ستون‌های عنوان/مبلغ
    اجباری‌ان (فارسی یا انگلیسی، هرکدوم بود)؛ بقیه اختیاری با پیش‌فرض منطقی."""
    adm = _get_admin(request); guard = _require(adm, "accounting")
    if guard: return guard
    if not file or not file.filename:
        return _redir("/admin/accounting/expenses?flash=فایلی+انتخاب+نشد")
    form = await request.form()
    default_ptype = str(form.get("default_payment_type", "expense"))
    if default_ptype not in ("expense", "salary", "partner_payout", "other"):
        default_ptype = "expense"

    try:
        raw = await file.read()
        from import_utils import parse_uploaded_rows, pick
        rows = parse_uploaded_rows(file.filename, raw)
    except Exception as ex:
        return _redir(f"/admin/accounting/expenses?flash=خطا+در+خواندن+فایل:+{str(ex)[:60]}")

    from db import create_expense, ensure_accounting_schema
    ensure_accounting_schema()
    created, skipped = 0, 0
    for row in rows:
        title = pick(row, "title", "عنوان", "raw")
        amount_raw = pick(row, "amount", "مبلغ", "price", "قیمت")
        try:
            amount = int(float(str(amount_raw).replace(",", "").strip())) if amount_raw else 0
        except Exception:
            amount = 0
        if not title or amount <= 0:
            skipped += 1
            continue
        category = pick(row, "category", "دسته", "دسته‌بندی", default="سایر")
        date_ = pick(row, "date", "تاریخ", "expense_date")
        desc = pick(row, "description", "توضیحات", "توضیح")
        payee = pick(row, "payee", "payee_name", "گیرنده", "نام گیرنده")
        ptype = pick(row, "payment_type", "نوع") or default_ptype
        if ptype not in ("expense", "salary", "partner_payout", "other"):
            ptype = default_ptype
        create_expense(title, category, amount, date_, desc, payment_type=ptype, payee_name=payee)
        created += 1

    _log(request, "وارد کردن دسته‌ای هزینه", "حسابداری", f"فایل:{file.filename} | ثبت:{created} | رد:{skipped}", admin_info=adm)
    return _redir(f"/admin/accounting/expenses?flash=✅+{created}+ردیف+ثبت+شد" + (f"+—+{skipped}+ردیف+ناقص+رد+شد" if skipped else ""))


@router.post("/accounting/expenses/{eid}/delete")
async def accounting_expense_delete(request: Request, eid: int):
    adm = _get_admin(request); guard = _require(adm, "accounting")
    if guard: return guard
    from db import delete_expense; delete_expense(eid)
    _log(request, "حذف هزینه", "حسابداری", f"id:{eid}")
    return _redir("/admin/accounting/expenses?flash=حذف+شد")


@router.post("/accounting/categories/new")
async def accounting_category_new(request: Request):
    adm = _get_admin(request); guard = _require(adm, "accounting")
    if guard: return guard
    from db import add_expense_category
    form = await request.form(); name = str(form.get("name","")).strip()
    if name: add_expense_category(name)
    return _redir("/admin/accounting/expenses?flash=دسته+اضافه+شد")


@router.get("/accounting/cashflow", response_class=HTMLResponse)
async def accounting_cashflow(request: Request, df: str="", dt: str=""):
    adm = _get_admin(request); guard = _require(adm, "accounting")
    if guard: return guard
    from db import get_cashflow, ensure_accounting_schema; ensure_accounting_schema()
    rows_data = get_cashflow(df, dt, 200)
    tc = {"فروش":"green","شارژ کیف‌پول":"blue","هزینه":"red","پورسانت":"amber"}
    rows = "".join(f'''<tr class="border-b hover:bg-gray-50 text-sm">
      <td class="px-3 py-2 text-xs text-gray-400">{fa_date(r["created_at"], with_time=True)}</td>
      <td class="px-3 py-2"><span class="px-2 py-0.5 rounded text-xs bg-{tc.get(r["type"],"gray")}-100 text-{tc.get(r["type"],"gray")}-700">{r["type"]}</span></td>
      <td class="px-3 py-2 text-xs">{str(r["description"] or "")[:40]}</td>
      <td class="px-3 py-2 font-bold {"text-green-600" if r["direction"]=="income" else "text-red-500"}">{"+" if r["direction"]=="income" else "-"}{int(r["amount"] or 0):,}</td>
    </tr>''' for r in rows_data)
    body = f"""<div class="flex items-center gap-3 mb-6">
      {_btn("← حسابداری","/admin/accounting","slate",small=True)}
      <h1 class="text-2xl font-bold text-gray-800">🔄 گردش مالی</h1>
    </div>
    <form method="get" class="flex gap-2 mb-4 flex-wrap">
      <input type="date" name="df" value="{df}" class="border rounded-lg px-3 py-2 text-sm">
      <input type="date" name="dt" value="{dt}" class="border rounded-lg px-3 py-2 text-sm">
      <button class="px-4 py-2 bg-indigo-600 text-white rounded-lg text-sm">فیلتر</button>
      <a href="/admin/accounting/cashflow/export?df={df}&dt={dt}" class="px-3 py-2 bg-green-50 text-green-700 border border-green-200 rounded-lg text-sm">⬇ Excel</a>
    </form>
    <div class="card overflow-hidden"><div class="overflow-x-auto"><table class="w-full text-right min-w-max">
      <thead><tr class="text-xs text-gray-500 border-b bg-gray-50">
        <th class="px-3 py-2">تاریخ</th><th class="px-3 py-2">نوع</th><th class="px-3 py-2">توضیح</th><th class="px-3 py-2">مبلغ</th>
      </tr></thead><tbody>{rows or "<tr><td colspan='4' class='text-center py-6 text-gray-400'>رکوردی یافت نشد</td></tr>"}</tbody>
    </table></div></div>"""
    return _layout("گردش مالی", body, adm)


@router.get("/accounting/products", response_class=HTMLResponse)
async def accounting_products(request: Request):
    adm = _get_admin(request); guard = _require(adm, "accounting")
    if guard: return guard
    from db import get_product_accounting
    prods = get_product_accounting(50)
    rows = "".join(f'''<tr class="border-b hover:bg-gray-50 text-sm">
      <td class="px-3 py-2 font-medium">{e(p["title"])}</td>
      <td class="px-3 py-2 text-center">{p["sale_count"]}</td>
      <td class="px-3 py-2 text-green-600">{int(p["total_revenue"] or 0):,}</td>
      <td class="px-3 py-2">{int(p["avg_cost"] or 0):,}</td>
      <td class="px-3 py-2">{int(p["last_cost"] or 0):,}</td>
      <td class="px-3 py-2 font-bold text-emerald-600">{int(p["total_revenue"] or 0) - int(p["avg_cost"] or 0)*max(int(p["sale_count"] or 1),1):,}</td>
      <td class="px-3 py-2 text-center">{p["stock"]}</td>
    </tr>''' for p in prods)
    body = f"""<div class="flex items-center gap-3 mb-6">
      {_btn("← حسابداری","/admin/accounting","slate",small=True)}
      <h1 class="text-2xl font-bold text-gray-800">📦 گزارش محصولات</h1>
      <a href="/admin/accounting/products/export" class="px-3 py-1.5 text-sm bg-green-50 text-green-700 border border-green-200 rounded-lg mr-auto">⬇ Excel</a>
    </div>
    <div class="card overflow-hidden"><div class="overflow-x-auto"><table class="w-full text-right min-w-max">
      <thead><tr class="text-xs text-gray-500 border-b bg-gray-50">
        <th class="px-3 py-2">محصول</th><th class="px-3 py-2 text-center">فروش</th><th class="px-3 py-2">درآمد</th>
        <th class="px-3 py-2">میانگین خرید</th><th class="px-3 py-2">آخرین خرید</th><th class="px-3 py-2">سود</th><th class="px-3 py-2">موجودی</th>
      </tr></thead><tbody>{rows or "<tr><td colspan='7' class='text-center py-6 text-gray-400'>داده‌ای یافت نشد</td></tr>"}</tbody>
    </table></div></div>"""
    return _layout("گزارش محصولات", body, adm)


@router.get("/accounting/partners", response_class=HTMLResponse)
async def accounting_partners_report(request: Request):
    adm = _get_admin(request); guard = _require(adm, "accounting")
    if guard: return guard
    from db import get_partner_accounting
    partners = get_partner_accounting(50)
    rows = "".join(f'''<tr class="border-b hover:bg-gray-50 text-sm">
      <td class="px-3 py-2 font-medium">{e(p["full_name"] or p["username"] or str(p["user_id"]))}</td>
      <td class="px-3 py-2 text-center">{p["sale_count"]}</td>
      <td class="px-3 py-2 text-green-600">{int(p["total_sales"] or 0):,}</td>
      <td class="px-3 py-2 text-amber-600">{int(p["commission_paid"] or 0):,}</td>
      <td class="px-3 py-2 font-bold text-emerald-600">{int(p["total_sales"] or 0)-int(p["commission_paid"] or 0):,}</td>
    </tr>''' for p in partners)
    body = f"""<div class="flex items-center gap-3 mb-6">
      {_btn("← حسابداری","/admin/accounting","slate",small=True)}
      <h1 class="text-2xl font-bold text-gray-800">🤝 گزارش همکاران</h1>
      <a href="/admin/accounting/partners/export" class="px-3 py-1.5 text-sm bg-green-50 text-green-700 border border-green-200 rounded-lg mr-auto">⬇ Excel</a>
    </div>
    <div class="card overflow-hidden"><div class="overflow-x-auto"><table class="w-full text-right min-w-max">
      <thead><tr class="text-xs text-gray-500 border-b bg-gray-50">
        <th class="px-3 py-2">همکار</th><th class="px-3 py-2 text-center">فروش</th>
        <th class="px-3 py-2">مجموع فروش</th><th class="px-3 py-2">پورسانت</th><th class="px-3 py-2">سود فروشگاه</th>
      </tr></thead><tbody>{rows or "<tr><td colspan='5' class='text-center py-6 text-gray-400'>داده‌ای یافت نشد</td></tr>"}</tbody>
    </table></div></div>"""
    return _layout("گزارش همکاران", body, adm)


@router.get("/accounting/exchanges", response_class=HTMLResponse)
async def accounting_exchanges_report(request: Request):
    adm = _get_admin(request); guard = _require(adm, "accounting")
    if guard: return guard
    from db import list_exchanges
    exchanges = list_exchanges(200)
    rows = "".join(f'''<tr class="border-b hover:bg-gray-50 text-sm">
      <td class="px-3 py-2 text-xs text-gray-400">{fa_date(x["exchanged_at"] or "", with_time=True)}</td>
      <td class="px-3 py-2 font-mono text-xs"><code>{e(x["user_id"])}</code></td>
      <td class="px-3 py-2">{e(x["old_title"])} <span class="text-xs text-gray-400 no-fa">({int(x["old_price"] or 0):,} ت)</span></td>
      <td class="px-3 py-2 text-gray-400">←</td>
      <td class="px-3 py-2">{e(x["new_title"])} <span class="text-xs text-gray-400 no-fa">({int(x["new_price"] or 0):,} ت)</span></td>
      <td class="px-3 py-2 font-bold no-fa {"text-red-600" if int(x["new_price"] or 0) > int(x["old_price"] or 0) else "text-emerald-600" if int(x["new_price"] or 0) < int(x["old_price"] or 0) else "text-gray-400"}">{int(x["new_price"] or 0) - int(x["old_price"] or 0):,}</td>
      <td class="px-3 py-2"><a href="/admin/orders/{x["new_order_id"]}" class="text-xs text-indigo-600 hover:underline">سفارش جدید #{x["new_order_id"]}</a></td>
    </tr>''' for x in exchanges)
    body = f"""<div class="flex items-center gap-3 mb-6">
      {_btn("← حسابداری","/admin/accounting","slate",small=True)}
      <h1 class="text-2xl font-bold text-gray-800">🔄 گزارش تعویض‌ها</h1>
      <a href="/admin/accounting/exchanges/export" class="px-3 py-1.5 text-sm bg-green-50 text-green-700 border border-green-200 rounded-lg mr-auto">⬇ Excel</a>
    </div>
    <div class="card overflow-hidden"><div class="overflow-x-auto"><table class="w-full text-right min-w-max">
      <thead><tr class="text-xs text-gray-500 border-b bg-gray-50">
        <th class="px-3 py-2">تاریخ</th><th class="px-3 py-2">کاربر</th><th class="px-3 py-2">محصول قبلی</th><th></th>
        <th class="px-3 py-2">محصول جدید</th><th class="px-3 py-2">اختلاف قیمت</th><th class="px-3 py-2">سفارش جدید</th>
      </tr></thead><tbody>{rows or "<tr><td colspan='7' class='text-center py-6 text-gray-400'>تعویضی ثبت نشده</td></tr>"}</tbody>
    </table></div></div>"""
    return _layout("گزارش تعویض‌ها", body, adm)


@router.get("/accounting/exchanges/export")
async def export_exchanges(request: Request):
    adm = _get_admin(request); guard = _require(adm, "accounting")
    if guard: return guard
    from db import list_exchanges
    exchanges = list_exchanges(1000)
    try:
        import io, openpyxl
        from fastapi.responses import StreamingResponse
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "تعویض‌ها"
        ws.append(["تاریخ", "کاربر", "محصول قبلی", "قیمت قبلی", "محصول جدید", "قیمت جدید", "اختلاف", "سفارش جدید"])
        for x in exchanges:
            ws.append([str(x["exchanged_at"] or ""), x["user_id"], x["old_title"], int(x["old_price"] or 0),
                       x["new_title"], int(x["new_price"] or 0), int(x["new_price"] or 0)-int(x["old_price"] or 0), x["new_order_id"]])
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=exchanges.xlsx"})
    except ImportError:
        return _redir("/admin/accounting/exchanges?flash=openpyxl+نصب+نیست")


@router.get("/accounting/expenses/export")
async def export_expenses(request: Request, df: str="", dt: str="", cat: str=""):
    adm = _get_admin(request); guard = _require(adm, "accounting")
    if guard: return guard
    from db import get_expenses; import io as _io
    data = get_expenses(df, dt, cat, 10000)
    try:
        import openpyxl; wb = openpyxl.Workbook(); ws = wb.active; ws.title="هزینه‌ها"
        ws.append(["تاریخ","عنوان","دسته","مبلغ","توضیح"])
        for r in data: ws.append([fa_date(r["expense_date"]),r["title"],r["category"],r["amount"],r["description"]])
        buf = _io.BytesIO(); wb.save(buf); buf.seek(0)
        from fastapi.responses import StreamingResponse
        return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition":"attachment; filename=expenses.xlsx"})
    except Exception:
        from fastapi.responses import PlainTextResponse
        csv = "تاریخ,عنوان,دسته,مبلغ\n"+"\n".join(f"{fa_date(r['expense_date'])},{r['title']},{r['category']},{r['amount']}" for r in data)
        return PlainTextResponse(csv, headers={"Content-Disposition":"attachment; filename=expenses.csv"})


@router.get("/accounting/cashflow/export")
async def export_cashflow(request: Request, df: str="", dt: str=""):
    adm = _get_admin(request); guard = _require(adm, "accounting")
    if guard: return guard
    from db import get_cashflow; import io as _io; data = get_cashflow(df, dt, 10000)
    try:
        import openpyxl; wb = openpyxl.Workbook(); ws = wb.active; ws.title="گردش مالی"
        ws.append(["تاریخ","نوع","توضیح","مبلغ","جهت"])
        for r in data: ws.append([fa_date(r["created_at"], with_time=True),r["type"],r["description"],r["amount"],r["direction"]])
        buf = _io.BytesIO(); wb.save(buf); buf.seek(0)
        from fastapi.responses import StreamingResponse
        return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition":"attachment; filename=cashflow.xlsx"})
    except Exception:
        from fastapi.responses import PlainTextResponse
        return PlainTextResponse("تاریخ,نوع,مبلغ\n", headers={"Content-Disposition":"attachment; filename=cashflow.csv"})


@router.get("/accounting/products/export")
async def export_products_report(request: Request):
    adm = _get_admin(request); guard = _require(adm, "accounting")
    if guard: return guard
    from db import get_product_accounting; import io as _io; data = get_product_accounting(1000)
    try:
        import openpyxl; wb = openpyxl.Workbook(); ws = wb.active; ws.title="محصولات"
        ws.append(["محصول","تعداد فروش","درآمد","میانگین خرید","آخرین خرید","موجودی"])
        for p in data: ws.append([p["title"],p["sale_count"],p["total_revenue"],p["avg_cost"],p["last_cost"],p["stock"]])
        buf = _io.BytesIO(); wb.save(buf); buf.seek(0)
        from fastapi.responses import StreamingResponse
        return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition":"attachment; filename=products.xlsx"})
    except Exception:
        from fastapi.responses import PlainTextResponse
        return PlainTextResponse("محصول,فروش,درآمد\n", headers={"Content-Disposition":"attachment; filename=products.csv"})


@router.get("/accounting/partners/export")
async def export_partners_report(request: Request):
    adm = _get_admin(request); guard = _require(adm, "accounting")
    if guard: return guard
    from db import get_partner_accounting; import io as _io; data = get_partner_accounting(1000)
    try:
        import openpyxl; wb = openpyxl.Workbook(); ws = wb.active; ws.title="همکاران"
        ws.append(["همکار","فروش","درآمد","پورسانت","سود فروشگاه"])
        for p in data:
            name = p["full_name"] or p["username"] or str(p["user_id"])
            ws.append([name,p["sale_count"],p["total_sales"],p["commission_paid"],int(p["total_sales"] or 0)-int(p["commission_paid"] or 0)])
        buf = _io.BytesIO(); wb.save(buf); buf.seek(0)
        from fastapi.responses import StreamingResponse
        return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition":"attachment; filename=partners.xlsx"})
    except Exception:
        from fastapi.responses import PlainTextResponse
        return PlainTextResponse("همکار,فروش\n", headers={"Content-Disposition":"attachment; filename=partners.csv"})


@router.get("/notes", response_class=HTMLResponse)
async def admin_notes_page(request: Request, status: str = "", flash: str = ""):
    adm = _get_admin(request)
    guard = _require(adm, "notes")
    if guard: return guard
    from db import get_admin_notes, ensure_admin_notes_schema
    ensure_admin_notes_schema()
    notes = get_admin_notes(status)

    filter_tabs = '<div class="flex gap-2 mb-4">' + "".join(
        f'<a href="/admin/notes?status={v}" class="px-3 py-1.5 rounded-lg border text-xs {"bg-indigo-600 text-white" if status==v else "bg-white text-gray-500"}">{l}</a>'
        for l,v in [("همه",""),("باز","open"),("انجام شد","done")]
    ) + '</div>'

    rows = ""
    for n in notes:
        sc = "green" if n["status"] == "done" else "amber"
        sl = "✅ انجام شد" if n["status"] == "done" else "🔵 باز"
        rows += f"""<tr class="border-b hover:bg-gray-50">
          <td class="px-4 py-3 text-xs text-gray-400">#{n['id']}</td>
          <td class="px-4 py-3 text-sm font-medium">{e(n['author'])}</td>
          <td class="px-4 py-3 text-sm">{e((n['text'] or '')[:60])}{'...' if len(n['text'] or '')>60 else ''}</td>
          <td class="px-4 py-3"><span class="px-2 py-0.5 text-xs bg-{sc}-100 text-{sc}-700 rounded-full">{sl}</span></td>
          <td class="px-4 py-3 text-xs text-gray-400">{fa_date(n['created_at'] or '', with_time=True)}</td>
          <td class="px-4 py-3 text-xs text-indigo-500">{n['reply_count']} پاسخ</td>
          <td class="px-4 py-3 flex gap-1">
            <a href="/admin/notes/{n['id']}" class="px-2 py-1 text-xs bg-indigo-50 text-indigo-700 rounded">مشاهده</a>
            <form method="post" action="/admin/notes/{n['id']}/toggle" class="inline">
              <button class="px-2 py-1 text-xs bg-gray-100 text-gray-600 rounded">{'رفع انجام' if n['status']=='done' else '✅ انجام شد'}</button>
            </form>
          </td>
        </tr>"""

    body = f"""
    <div class="flex items-center justify-between mb-6">
      <h1 class="text-2xl font-bold text-gray-800">📝 یادداشت مدیران</h1>
      <a href="/admin/notes/new" class="btn-sm bg-indigo-600 text-white rounded-lg px-4 py-2 text-sm">+ یادداشت جدید</a>
    </div>
    {filter_tabs}
    <div class="card overflow-hidden"><div class="overflow-x-auto">
      <table class="w-full text-right min-w-max">
        <thead><tr class="text-xs text-gray-500 border-b bg-gray-50">
          <th class="px-4 py-3">#</th><th class="px-4 py-3">نویسنده</th>
          <th class="px-4 py-3">متن</th><th class="px-4 py-3">وضعیت</th>
          <th class="px-4 py-3">تاریخ</th><th class="px-4 py-3">پاسخ</th>
          <th class="px-4 py-3">عملیات</th>
        </tr></thead>
        <tbody>{rows or "<tr><td colspan='7' class='text-center py-8 text-gray-400'>یادداشتی ثبت نشده</td></tr>"}</tbody>
      </table>
    </div></div>"""
    return _layout("یادداشت مدیران", body, adm, flash=flash)


@router.get("/notes/new", response_class=HTMLResponse)
async def admin_note_new_get(request: Request):
    adm = _get_admin(request)
    guard = _require(adm, "notes")
    if guard: return guard
    body = f"""
    <div class="flex items-center gap-3 mb-6">
      {_btn("← یادداشت‌ها", "/admin/notes", "slate", small=True)}
      <h1 class="text-2xl font-bold text-gray-800">📝 یادداشت جدید</h1>
    </div>
    <div class="card p-6 max-w-xl">
      <form method="post" action="/admin/notes/new" class="space-y-4">
        <div><label class="text-sm font-medium text-gray-700 block mb-1">متن یادداشت</label>
          {_textarea("text","یادداشت خود را بنویسید...",rows=6)}</div>
        {_btn("ثبت یادداشت","",color="indigo")}
      </form>
    </div>"""
    return _layout("یادداشت جدید", body, adm)


@router.post("/notes/new")
async def admin_note_new_post(request: Request):
    adm = _get_admin(request)
    guard = _require(adm, "notes")
    if guard: return guard
    form = await request.form()
    text = str(form.get("text","")).strip()
    if not text:
        return _redir("/admin/notes/new")
    from db import create_admin_note
    author = (adm[0] if adm else "مدیر")
    create_admin_note(author, text)
    _log(request, "ثبت یادداشت", "یادداشت‌ها", text[:40])
    return _redir("/admin/notes?flash=یادداشت+ثبت+شد")


@router.get("/notes/{nid}", response_class=HTMLResponse)
async def admin_note_detail(request: Request, nid: int, flash: str = ""):
    adm = _get_admin(request)
    guard = _require(adm, "notes")
    if guard: return guard
    from db import get_admin_note
    data = get_admin_note(nid)
    if not data:
        return _redir("/admin/notes")
    note = data["note"]
    replies = data["replies"]

    reply_rows = "".join(f"""<div class="p-4 bg-gray-50 rounded-lg mb-2">
      <div class="flex justify-between text-xs text-gray-400 mb-1">
        <span class="font-medium text-gray-700">{e(r['author'])}</span>
        <span>{fa_date(r['created_at'] or '', with_time=True)}</span>
      </div>
      <p class="text-sm text-gray-800">{e(r['text'])}</p>
    </div>""" for r in replies)

    body = f"""
    <div class="flex items-center gap-3 mb-6">
      {_btn("← یادداشت‌ها", "/admin/notes", "slate", small=True)}
      <h1 class="text-2xl font-bold text-gray-800">📝 یادداشت #{nid}</h1>
    </div>
    <div class="grid md:grid-cols-2 gap-4">
      <div class="card p-6">
        <div class="flex justify-between items-start mb-4">
          <div>
            <div class="text-sm font-medium text-gray-700">{e(note['author'])}</div>
            <div class="text-xs text-gray-400">{fa_date(note['created_at'] or '', with_time=True)}</div>
          </div>
          <span class="px-2 py-1 text-xs rounded-full {'bg-green-100 text-green-700' if note['status']=='done' else 'bg-amber-100 text-amber-700'}">
            {'✅ انجام شد' if note['status']=='done' else '🔵 باز'}
          </span>
        </div>
        <p class="text-sm text-gray-800 whitespace-pre-wrap mb-4">{e(note['text'])}</p>
        <form method="post" action="/admin/notes/{nid}/toggle">
          <button class="px-3 py-1.5 text-xs border rounded-lg {'bg-gray-100' if note['status']=='done' else 'bg-green-50 text-green-700 border-green-200'}">
            {'↩ بازگشایی' if note['status']=='done' else '✅ علامت انجام'}
          </button>
        </form>
      </div>
      <div class="card p-6">
        <h2 class="font-bold text-gray-700 mb-3">💬 پاسخ‌ها</h2>
        <div class="mb-4 max-h-64 overflow-y-auto">{reply_rows or '<p class="text-sm text-gray-400">پاسخی ثبت نشده</p>'}</div>
        <form method="post" action="/admin/notes/{nid}/reply">
          {_textarea("text","پاسخ شما...",rows=3)}
          <div class="mt-2">{_btn("ثبت پاسخ","",color="indigo",small=True)}</div>
        </form>
      </div>
    </div>"""
    return _layout(f"یادداشت #{nid}", body, adm, flash=flash)


@router.post("/notes/{nid}/reply")
async def admin_note_reply(request: Request, nid: int):
    adm = _get_admin(request)
    guard = _require(adm, "notes")
    if guard: return guard
    form = await request.form()
    text = str(form.get("text","")).strip()
    if text:
        from db import add_admin_note_reply
        author = (adm[0] if adm else "مدیر")
        add_admin_note_reply(nid, author, text)
    return _redir(f"/admin/notes/{nid}?flash=پاسخ+ثبت+شد")


@router.post("/notes/{nid}/toggle")
async def admin_note_toggle(request: Request, nid: int):
    adm = _get_admin(request)
    guard = _require(adm, "notes")
    if guard: return guard
    from db import toggle_admin_note_status
    new_status = toggle_admin_note_status(nid)
    _log(request, f"تغییر وضعیت یادداشت به {new_status}", "یادداشت‌ها", f"#{nid}")
    return _redir(f"/admin/notes/{nid}")

@router.get("/partners", response_class=HTMLResponse)
async def partners_list(request: Request, tab: str = "list", status_filter: str = "", flash: str = ""):
    adm = _get_admin(request)
    guard = _require(adm, "partners")
    if guard: return guard

    from db import fa_date, fa_now
    from db import (ensure_partner_system_schema, get_partner_tiers,
                    get_partner_commission, ensure_referral_schema, get_referral_settings)
    ensure_partner_system_schema()
    ensure_referral_schema()

    # تب‌های اصلی
    tab_defs = [("list","👥 لیست همکاران"),("tree","🌳 درخت همکاران"),("referrals","🔗 معرفی‌ها"),
                ("payouts","📤 تسویه‌ها"),("settings","⚙️ تنظیمات")]
    tabs_html = '<div class="flex gap-2 mb-6 overflow-x-auto pb-1">' + "".join(
        f'<a href="/admin/partners?tab={v}" class="px-4 py-2 rounded-lg border text-sm whitespace-nowrap {"bg-indigo-600 text-white" if tab==v else "bg-white text-gray-600"}">{l}</a>'
        for v, l in tab_defs
    ) + '</div>'

    content = ""

    # ─── تب لیست همکاران ─────────────────────────────────────────────────
    if tab == "list":
        conn = _db()
        try:
            where = "WHERE status=?" if status_filter else ""
            partners = conn.execute(
                f"SELECT * FROM partners {where} ORDER BY CASE status WHEN 'pending' THEN 0 ELSE 1 END, id DESC LIMIT 100;",
                (status_filter,) if status_filter else ()
            ).fetchall()
        finally:
            conn.close()

        sub_tabs = '<div class="flex gap-2 mb-4">' + "".join(
            f'<a href="/admin/partners?tab=list&status_filter={v}" class="px-3 py-1.5 rounded-lg border text-xs {"bg-amber-500 text-white" if status_filter==v else "bg-white text-gray-500"}">{l}</a>'
            for l, v in [("همه",""),("در انتظار","pending"),("تایید","approved"),("رد","rejected")]
        ) + '</div>'

        rows = ""
        for p in partners:
            st = p["status"] or "pending"
            bc = {"pending":"yellow","approved":"green","rejected":"red"}.get(st,"gray")
            bl = {"pending":"در انتظار","approved":"تایید","rejected":"رد"}.get(st,st)
            actions = f"""<a href="/admin/partners/{p['tg_user_id']}/profile"
                  class="px-2 py-1 text-xs bg-indigo-50 text-indigo-700 border border-indigo-200 rounded hover:bg-indigo-100">👤 پروفایل</a>"""
            if st == "pending":
                actions += f"""<form method="post" action="/admin/partners/{p['tg_user_id']}/approve" class="inline">
                  <button class="px-2 py-1 text-xs bg-green-100 text-green-700 rounded hover:bg-green-200">✅</button></form>
                  <form method="post" action="/admin/partners/{p['tg_user_id']}/reject" class="inline">
                  <button class="px-2 py-1 text-xs bg-red-100 text-red-700 rounded hover:bg-red-200">❌</button></form>"""
            rows += f"""<tr class="border-b hover:bg-gray-50 text-sm">
              <td class="px-4 py-3 font-mono text-xs"><code>{e(p['tg_user_id'])}</code></td>
              <td class="px-4 py-3">{e(p['full_name'])}</td>
              <td class="px-4 py-3 text-gray-500">{e(p['phone'])}</td>
              <td class="px-4 py-3 text-gray-400 text-xs">{e(p['city'])} | {e(p['shop_name'])}</td>
              <td class="px-4 py-3"><span class="px-2 py-0.5 text-xs rounded-full bg-{bc}-100 text-{bc}-700">{bl}</span></td>
              <td class="px-4 py-3 flex gap-1">{actions}</td>
            </tr>"""

        content = f"""{sub_tabs}
        <div class="card overflow-hidden"><div class="overflow-x-auto">
          <table class="w-full text-right min-w-max">
            <thead><tr class="text-xs text-gray-500 border-b bg-gray-50">
              <th class="px-4 py-3">User ID</th><th class="px-4 py-3">نام</th>
              <th class="px-4 py-3">شماره</th><th class="px-4 py-3">شهر/فروشگاه</th>
              <th class="px-4 py-3">وضعیت</th><th class="px-4 py-3">عملیات</th>
            </tr></thead>
            <tbody>{rows or "<tr><td colspan='6' class='text-center py-8 text-gray-400'>درخواستی یافت نشد</td></tr>"}</tbody>
          </table>
        </div></div>"""

    # ─── تب معرفی‌ها (ادغام‌شده) ─────────────────────────────────────────
    elif tab == "referrals":
        # ─── تب «معرفی‌ها، پاداش‌ها و درآمدها» ───────────────────────────
        manual_form = """
        <div class="card p-4 mb-4 border-r-4 border-indigo-400">
          <h2 class="font-bold text-gray-700 mb-1 text-sm">🔧 ثبت دستی معرفی</h2>
          <p class="text-xs text-gray-400 mb-3">اگر معرفی‌ای در ربات ثبت نشده، اینجا دستی وصل کنید.</p>
          <form method="post" action="/admin/partners/manual-referral" class="flex flex-wrap gap-2 items-end">
            <div><label class="text-[10px] text-gray-400 block mb-1">آیدی معرف</label>
              <input type="number" name="referrer_id" required placeholder="ID معرف"
                class="border border-gray-300 rounded-lg px-2.5 py-2 text-xs w-36" dir="ltr"></div>
            <div><label class="text-[10px] text-gray-400 block mb-1">آیدی دعوت‌شده</label>
              <input type="number" name="referred_id" required placeholder="ID دعوت‌شده"
                class="border border-gray-300 rounded-lg px-2.5 py-2 text-xs w-36" dir="ltr"></div>
            <label class="flex items-center gap-1.5 text-xs pb-1.5"><input type="checkbox" name="pay_reward" checked> پرداخت پاداش</label>
            <button class="px-4 py-2 bg-indigo-600 text-white rounded-lg text-xs font-semibold">➕ ثبت</button>
          </form>
        </div>"""
        ref_settings = get_referral_settings()
        conn = _db()
        try:
            refs = conn.execute("""
                SELECT r.*, u1.full_name as referrer_name, u2.full_name as referred_name
                FROM referrals r
                LEFT JOIN users u1 ON u1.user_id=r.referrer_id
                LEFT JOIN users u2 ON u2.user_id=r.referred_id
                ORDER BY r.id DESC LIMIT 200;
            """).fetchall()
            total     = conn.execute("SELECT COUNT(*) FROM referrals;").fetchone()[0]
            rewarded  = conn.execute("SELECT COUNT(*) FROM referrals WHERE rewarded=1;").fetchone()[0]
            total_pay = conn.execute("SELECT COALESCE(SUM(reward_amount),0) FROM referrals WHERE rewarded=1;").fetchone()[0]
            # درآمد کل پورسانت
            try:
                total_comm = conn.execute(
                    "SELECT COALESCE(SUM(amount),0) FROM partner_transactions WHERE type='credit';"
                ).fetchone()[0]
            except Exception:
                total_comm = 0
        except Exception:
            refs = []; total = rewarded = total_pay = total_comm = 0
        finally:
            conn.close()

        ref_rows = "".join(f"""<tr class="border-b hover:bg-gray-50">
          <td class="px-3 py-2 text-xs text-gray-400">#{r['id']}</td>
          <td class="px-3 py-2 text-sm font-medium">
            <a href="/admin/partners/{r['referrer_id']}/profile" class="text-indigo-600 hover:underline">
              {e(r['referrer_name'] or str(r['referrer_id']))}</a></td>
          <td class="px-3 py-2 text-sm">
            <a href="/admin/partners/{r['referred_id']}/profile" class="text-gray-600 hover:underline">
              {e(r['referred_name'] or str(r['referred_id']))}</a></td>
          <td class="px-3 py-2">{'<span class="px-2 py-0.5 text-xs bg-green-100 text-green-700 rounded-full">✅ پرداخت</span>' if r['rewarded'] else '<span class="px-2 py-0.5 text-xs bg-gray-100 text-gray-500 rounded-full">منتظر</span>'}</td>
          <td class="px-3 py-2 text-sm font-medium text-green-600">{int(r['reward_amount'] or 0):,} ت</td>
          <td class="px-3 py-2 text-xs text-gray-400">{fa_date(r['created_at'])}</td>
        </tr>""" for r in refs)

        # فاز ۲: نوار آماری سبک به‌جای کارت‌های بزرگ — کاربردی‌تر و کم‌فضاتر
        conversion_rate = round(rewarded*100/total) if total else 0
        avg_reward = int(total_pay/rewarded) if rewarded else 0
        stats_bar = f"""
        <div class="card p-3 mb-4">
          <div class="grid grid-cols-2 md:grid-cols-5 gap-2 text-sm">
            <div class="p-2 border-l border-gray-100">
              <div class="text-[10px] text-gray-400 mb-1">کل معرفی‌ها</div>
              <div class="font-bold text-indigo-700">{total}</div>
            </div>
            <div class="p-2 border-l border-gray-100">
              <div class="text-[10px] text-gray-400 mb-1">پاداش‌شده</div>
              <div class="font-bold text-green-700">{rewarded} <span class="text-[10px] font-normal text-gray-400">نفر</span></div>
            </div>
            <div class="p-2 border-l border-gray-100">
              <div class="text-[10px] text-gray-400 mb-1">جمع پاداش</div>
              <div class="font-bold text-green-700">{int(total_pay):,} <span class="text-[10px] font-normal text-gray-400">ت</span></div>
            </div>
            <div class="p-2 border-l border-gray-100">
              <div class="text-[10px] text-gray-400 mb-1">میانگین</div>
              <div class="font-bold text-amber-700">{avg_reward:,} <span class="text-[10px] font-normal text-gray-400">ت</span></div>
            </div>
            <div class="p-2">
              <div class="text-[10px] text-gray-400 mb-1">نرخ تبدیل</div>
              <div class="font-bold text-teal-700">٪{conversion_rate}</div>
            </div>
          </div>
        </div>"""
        content = manual_form + stats_bar + f"""
        <div class="card overflow-hidden">
          <div class="px-4 py-3 border-b flex items-center justify-between">
            <h2 class="font-bold text-gray-700 text-sm">📋 لیست معرفی‌ها</h2>
            <span class="text-xs text-gray-400">کلیک روی نام → پروفایل کامل</span>
          </div>
          <div class="overflow-x-auto"><table class="w-full text-right min-w-max">
            <thead><tr class="text-xs text-gray-500 border-b bg-gray-50">
              <th class="px-3 py-2">#</th><th class="px-3 py-2">معرف</th>
              <th class="px-3 py-2">کاربر جدید</th><th class="px-3 py-2">وضعیت</th>
              <th class="px-3 py-2">پاداش</th><th class="px-3 py-2">تاریخ</th>
            </tr></thead>
            <tbody>{ref_rows or "<tr><td colspan='6' class='text-center py-8 text-gray-400'>معرفی‌ای ثبت نشده</td></tr>"}</tbody>
          </table></div>
        </div>"""

    elif tab == "payouts":
        from db import get_partner_payouts, ensure_partner_wallet_schema, get_partner_payout_settings
        ensure_partner_wallet_schema()
        status_f = request.query_params.get("pstatus", "")
        payouts = get_partner_payouts(status=status_f, limit=100)
        pstats = '<div class="flex gap-2 mb-4 flex-wrap">' + "".join(
            f'<a href="/admin/partners?tab=payouts&pstatus={v}" class="px-3 py-1.5 rounded-lg border text-xs {"bg-indigo-600 text-white" if status_f==v else "bg-white text-gray-500"}">{l}</a>'
            for l,v in [("همه",""),("در انتظار","pending"),("تایید","approved"),("رد","rejected")]
        ) + '</div>'
        prows = ""
        for p in payouts:
            sc = {"pending":"yellow","approved":"green","rejected":"red"}.get(p["status"],"gray")
            sl = {"pending":"در انتظار","approved":"تایید شد","rejected":"رد شد"}.get(p["status"],p["status"])
            acts = ""
            if p["status"] == "pending":
                acts = f"""<form method="post" action="/admin/partners/payout/{p['id']}/approve" class="inline">
                  <button class="px-2 py-1 text-xs bg-green-100 text-green-700 rounded">✅ تایید</button></form>
                  <form method="post" action="/admin/partners/payout/{p['id']}/reject" class="inline ml-1">
                  <button class="px-2 py-1 text-xs bg-red-100 text-red-700 rounded">❌ رد</button></form>"""
            prows += f"""<tr class="border-b hover:bg-gray-50 text-sm">
              <td class="px-4 py-3 text-xs text-gray-400">#{p['id']}</td>
              <td class="px-4 py-3">{e(p['full_name'] or str(p['user_id']))}</td>
              <td class="px-4 py-3 font-bold text-green-600">{int(p['amount']):,} ت</td>
              <td class="px-4 py-3"><span class="px-2 py-0.5 text-xs bg-{sc}-100 text-{sc}-700 rounded-full">{sl}</span></td>
              <td class="px-4 py-3 text-xs text-gray-400">{fa_date(p['created_at'] or '')}</td>
              <td class="px-4 py-3">{acts}<a href="/admin/partners/payout/{p['id']}" class="px-2 py-1 text-xs bg-indigo-50 text-indigo-700 rounded mr-1">جزئیات</a></td>
            </tr>"""

        content = f"""{pstats}
        <div class="card overflow-hidden"><div class="overflow-x-auto">
          <table class="w-full text-right min-w-max">
            <thead><tr class="text-xs text-gray-500 border-b bg-gray-50">
              <th class="px-4 py-3">#</th><th class="px-4 py-3">همکار</th>
              <th class="px-4 py-3">مبلغ</th><th class="px-4 py-3">وضعیت</th>
              <th class="px-4 py-3">تاریخ</th><th class="px-4 py-3">عملیات</th>
            </tr></thead>
            <tbody>{prows or "<tr><td colspan='6' class='text-center py-8 text-gray-400'>درخواستی یافت نشد</td></tr>"}</tbody>
          </table>
        </div></div>"""

    # ─── تب تنظیمات پورسانت ──────────────────────────────────────────────
    elif tab == "settings":
        from db import (get_partner_payout_settings, ensure_partner_wallet_schema,
                        get_partner_tiers, ensure_partner_tiers_extended,
                        get_referral_settings)
        ensure_partner_wallet_schema()
        ensure_partner_tiers_extended()
        comm = get_partner_commission()
        ps   = get_partner_payout_settings()
        rs   = get_referral_settings()
        tiers = get_partner_tiers()

        # ── جدول سطوح ──
        tier_rows = ""
        for tr in tiers:
            try: commission = tr["commission_percent"] or 0
            except Exception: commission = 0
            try: cfixed = int(tr["commission_fixed"] or 0)
            except Exception: cfixed = 0
            try: ctype = tr["commission_type"] or ("fixed" if cfixed > 0 else "percent")
            except Exception: ctype = "fixed" if cfixed > 0 else "percent"
            try: photo_id = tr["photo_file_id"] or ""
            except Exception: photo_id = ""
            type_badge = ('<span class="px-2 py-0.5 bg-purple-100 text-purple-700 rounded text-[10px]">مبلغ ثابت</span>'
                          if ctype == "fixed" else
                          '<span class="px-2 py-0.5 bg-blue-100 text-blue-700 rounded text-[10px]">درصدی</span>')
            amount_label = f"{cfixed:,} ت" if ctype == "fixed" else f"٪{commission}"
            banner_badge = ('<span class="text-green-600 text-lg" title="بنر ذخیره شده">🖼</span>'
                            if photo_id else
                            '<span class="text-gray-300 text-lg" title="بدون بنر">🖼</span>')
            tier_rows += f"""<tr class="border-b hover:bg-gray-50">
              <td class="px-3 py-2 text-xl">{e(tr["icon"])}</td>
              <td class="px-3 py-2 font-medium">{e(tr["name"])}</td>
              <td class="px-3 py-2 text-gray-500 text-xs">{tr["min_orders"]}+ خرید</td>
              <td class="px-3 py-2">{type_badge}</td>
              <td class="px-3 py-2 font-semibold text-indigo-700">{amount_label}</td>
              <td class="px-3 py-2 text-center">{banner_badge}</td>
              <td class="px-3 py-2 whitespace-nowrap">
                <a href="/admin/partners/tier/{tr["id"]}/edit"
                  class="px-3 py-1 text-xs bg-indigo-50 text-indigo-700 rounded hover:bg-indigo-100">✏️ ویرایش</a>
                <button type="button" onclick="document.getElementById(\'bnr_{tr["id"]}\').click()"
                  class="px-3 py-1 text-xs bg-blue-50 text-blue-700 rounded hover:bg-blue-100 mr-1">📤 بنر</button>
                <form method="post" action="/admin/partners/tier/{tr["id"]}/upload-banner"
                  enctype="multipart/form-data" class="inline" id="bnrform_{tr["id"]}">
                  <input type="file" name="banner_file" id="bnr_{tr["id"]}" accept="image/*"
                    style="display:none" onchange="document.getElementById(\'bnrform_{tr["id"]}\').submit()">
                </form>
                <form method="post" action="/admin/partners/tier/{tr["id"]}/delete" class="inline"
                  onsubmit="return confirm(\'حذف سطح؟\')">
                  <button class="px-3 py-1 text-xs bg-red-50 text-red-600 rounded border border-red-100 mr-1">حذف</button>
                </form>
              </td>
            </tr>"""

        content = f"""
        <style>
        details.acc summary{{list-style:none;cursor:pointer;user-select:none}}
        details.acc summary::-webkit-details-marker{{display:none}}
        details.acc[open] .acc-arrow{{transform:rotate(90deg)}}
        .acc-arrow{{transition:transform .18s;display:inline-block;margin-inline-start:4px}}
        </style>

        <!-- ① هدیه دعوت -->
        <details class="acc card mb-3">
          <summary class="flex items-center gap-2 px-5 py-4 font-bold text-gray-700">
            🎁 هدیه دعوت<span class="acc-arrow">›</span>
            <span class="mr-auto text-xs font-normal text-gray-400">پاداش یک‌بارِ لحظه عضویت</span>
          </summary>
          <div class="px-5 pb-5 border-t pt-4">
            <form method="post" action="/admin/referrals/settings" class="grid grid-cols-1 md:grid-cols-3 gap-4 items-end">
              <div>
                <label class="text-sm font-medium text-gray-700 block mb-1">مبلغ هدیه (تومان)</label>
                <input type="number" name="reward_amount" value="{int(rs.get('reward_amount',5000))}" min="0"
                  class="w-full border border-gray-200 rounded-lg px-3 py-2 text-sm">
                <p class="text-xs text-gray-400 mt-1">۰ = غیرفعال</p>
              </div>
              <div>
                <label class="text-sm font-medium text-gray-700 block mb-1">سقف تعداد دعوت</label>
                <input type="number" name="max_invites" value="{int(rs.get('max_invites',0))}" min="0"
                  class="w-full border border-gray-200 rounded-lg px-3 py-2 text-sm">
                <p class="text-xs text-gray-400 mt-1">۰ = نامحدود</p>
              </div>
              <div>
                <label class="text-sm font-medium text-gray-700 block mb-1">وضعیت</label>
                <select name="is_active" class="w-full border border-gray-200 rounded-lg px-3 py-2 text-sm">
                  <option value="1" {"selected" if rs.get("is_active") else ""}>فعال</option>
                  <option value="0" {"" if rs.get("is_active") else "selected"}>غیرفعال</option>
                </select>
              </div>
              <div class="md:col-span-3">
                {_btn("💾 ذخیره هدیه دعوت","",color="green",small=True)}
              </div>
            </form>
          </div>
        </details>

        <!-- ② سطوح و پورسانت -->
        <details class="acc card mb-3">
          <summary class="flex items-center gap-2 px-5 py-4 font-bold text-gray-700">
            🏆 سطوح و پورسانت فروش<span class="acc-arrow">›</span>
            <span class="mr-auto text-xs font-normal text-gray-400">پورسانت هر خرید توسط دعوت‌شده‌ها</span>
          </summary>
          <div class="px-5 pb-5 border-t pt-4">
            <div class="flex items-center justify-between mb-4">
              <p class="text-xs text-gray-500 leading-6">
                هر سطح پورسانت اختصاصی خود را دارد. همکاران بر اساس تعداد خرید به سطوح بالاتر ارتقا می‌یابند.
              </p>
              <a href="/admin/partners/tier/new"
                class="shrink-0 px-4 py-2 bg-indigo-600 hover:bg-indigo-700 text-white rounded-lg text-sm font-semibold">
                ➕ سطح جدید</a>
            </div>
            <!-- جدول سطوح — کاربردی‌تر و خواناتر -->
            <div class="overflow-x-auto"><table class="w-full text-right min-w-max text-sm">
              <thead><tr class="text-xs text-gray-500 border-b bg-gray-50">
                <th class="px-3 py-2">آیکون</th>
                <th class="px-3 py-2">نام</th>
                <th class="px-3 py-2">شرط ارتقا</th>
                <th class="px-3 py-2">نوع پورسانت</th>
                <th class="px-3 py-2">مقدار</th>
                <th class="px-3 py-2">بنر</th>
                <th class="px-3 py-2">عملیات</th>
              </tr></thead>
              <tbody>{tier_rows or "<tr><td colspan='7' class='text-center py-8 text-gray-400'>سطحی تعریف نشده — روی «➕ سطح جدید» بزنید</td></tr>"}</tbody>
            </table></div>
            <div class="mt-4 p-3 bg-blue-50 border border-blue-200 rounded-lg text-xs text-blue-700">
              💡 محدودیت‌ها (حداقل خرید و سقف پورسانت) در ویرایش هر سطح تنظیم می‌شود، نه اینجا.
            </div>
          </div>
        </details>

        <!-- ③ تنظیمات تسویه -->
        <details class="acc card mb-3">
          <summary class="flex items-center gap-2 px-5 py-4 font-bold text-gray-700">
            📤 تنظیمات تسویه<span class="acc-arrow">›</span>
          </summary>
          <div class="px-5 pb-5 border-t pt-4">
            <form method="post" action="/admin/partners/payout-settings" class="space-y-4 max-w-2xl">
              <div class="grid grid-cols-2 gap-4">
                <div><label class="text-sm font-medium text-gray-700 block mb-1">حداقل مبلغ (ت)</label>
                  <input type="number" name="min_amount" value="{ps.get("min_amount",50000)}"
                    class="w-full border border-gray-200 rounded-lg px-3 py-2 text-sm"></div>
                <div><label class="text-sm font-medium text-gray-700 block mb-1">حداکثر مبلغ (ت)</label>
                  <input type="number" name="max_amount" value="{ps.get("max_amount",0)}"
                    class="w-full border border-gray-200 rounded-lg px-3 py-2 text-sm">
                  <p class="text-xs text-gray-400">۰ = نامحدود</p></div>
                <div><label class="text-sm font-medium text-gray-700 block mb-1">حداکثر درخواست در ماه</label>
                  <input type="number" name="max_per_month" value="{ps.get("max_per_month",2)}"
                    class="w-full border border-gray-200 rounded-lg px-3 py-2 text-sm"></div>
                <div><label class="text-sm font-medium text-gray-700 block mb-1">مدت بررسی (ساعت)</label>
                  <input type="number" name="review_hours" value="{ps.get("review_hours",48)}"
                    class="w-full border border-gray-200 rounded-lg px-3 py-2 text-sm"></div>
              </div>
              <div><label class="text-sm font-medium text-gray-700 block mb-1">وضعیت</label>
                <select name="is_active" class="border border-gray-200 rounded-lg px-3 py-2 text-sm">
                  <option value="1" {"selected" if ps.get("is_active") else ""}>فعال</option>
                  <option value="0" {"" if ps.get("is_active") else "selected"}>غیرفعال</option>
                </select></div>
              <div><label class="text-sm font-medium text-gray-700 block mb-1">راهنمای تسویه (به همکار)</label>
                <textarea name="guide_text" rows="6" class="w-full border border-gray-200 rounded-lg px-3 py-2 text-sm" dir="rtl">{e(ps.get("guide_text",""))}</textarea>
                <p class="text-xs text-gray-400 mt-1">این متن هنگام باز کردن صفحه تسویه به همکار نمایش داده می‌شود.</p></div>
              <div class="grid grid-cols-1 md:grid-cols-2 gap-4">
                <div><label class="text-xs text-gray-500 block mb-1">پیام تأیید تسویه</label>
                  <textarea name="approval_message" rows="4" class="w-full border border-gray-200 rounded-lg px-3 py-2 text-sm" dir="rtl">{e(ps.get("approval_message",""))}</textarea></div>
                <div><label class="text-xs text-gray-500 block mb-1">پیام رد تسویه</label>
                  <textarea name="rejection_message" rows="4" class="w-full border border-gray-200 rounded-lg px-3 py-2 text-sm" dir="rtl">{e(ps.get("rejection_message",""))}</textarea></div>
              </div>
              {_btn("💾 ذخیره تسویه","",color="indigo")}
            </form>
          </div>
        </details>"""
    elif tab == "tree":
        # ─── 🌳 درخت همکاران — رندر کامل سمت کلاینت با Lazy DOM ───────────
        content = """
        <!-- کارت‌های آماری -->
        <div id="tree-stats" class="grid grid-cols-2 md:grid-cols-3 lg:grid-cols-6 gap-3 mb-5"></div>

        <!-- جستجو -->
        <div class="card p-4 mb-4">
          <div class="flex gap-2">
            <input id="tree-q" type="text" placeholder="جستجو: نام، یوزرنیم یا آیدی تلگرام…"
              class="flex-1 border border-gray-300 rounded-lg px-3 py-2.5 text-sm" dir="rtl"
              onkeydown="if(event.key==='Enter')treeSearch()">
            <button onclick="treeSearch()" class="px-5 py-2.5 bg-indigo-600 hover:bg-indigo-700 text-white rounded-lg text-sm font-semibold transition">🔍 جستجو</button>
          </div>
          <div id="tree-results" class="mt-2 text-sm"></div>
        </div>

        <!-- درخت -->
        <div class="card p-4 overflow-x-auto min-h-300">
          <div id="tree-root" class="min-w-max"><div class="text-center text-gray-400 py-12">در حال بارگذاری درخت…</div></div>
        </div>

        <!-- پنل اطلاعات (سمت راست) -->
        <div id="tree-overlay" class="hidden tree-overlay-fixed" onclick="treeCloseDrawer()"></div>
        <div id="tree-drawer" class="card tree-drawer-fixed">
          <div class="flex items-center justify-between px-4 py-3 border-b sticky top-0 bg-white tree-drawer-header">
            <b class="text-gray-800">👤 اطلاعات همکار</b>
            <button onclick="treeCloseDrawer()" class="text-gray-400 hover:text-red-500 text-lg px-2">✕</button>
          </div>
          <div id="tree-drawer-body" class="p-4 text-sm"></div>
        </div>

        <style>
        .tn-row{display:flex;align-items:center;gap:8px;padding:7px 10px;border-radius:10px;cursor:pointer;transition:background .12s;position:relative}
        .tn-row:hover{background:#F5F7FB}
        body.sl-dark .tn-row:hover,body.dark-mode .tn-row:hover{background:#1F2A38}
        .tn-arrow{width:22px;height:22px;display:flex;align-items:center;justify-content:center;border-radius:6px;flex-shrink:0;
          color:#8A94A6;font-size:11px;transition:transform .18s;user-select:none}
        .tn-arrow.open{transform:rotate(-90deg)}
        .tn-arrow.leaf{visibility:hidden}
        .tn-tier{font-size:15px}
        .tn-kids{border-right:2px solid #E7EBF2;margin-right:20px;padding-right:6px}
        body.sl-dark .tn-kids,body.dark-mode .tn-kids{border-right-color:#2B3A4C}
        .tn-badge{font-size:11px;padding:2px 8px;border-radius:999px;white-space:nowrap}
        .tn-hl{outline:2px solid #6366F1;outline-offset:2px;border-radius:10px;animation:tnflash 1.6s ease 2}
        @keyframes tnflash{0%,100%{background:transparent}50%{background:rgba(99,102,241,.14)}}
        </style>

        <script>
        (function(){
          var T=null, EXP={}, CHUNK=100;
          var ST_LBL={approved:['فعال','bg-green-100 text-green-700'],
                      pending:['در انتظار','bg-amber-100 text-amber-700'],
                      rejected:['غیرفعال','bg-red-100 text-red-600'],
                      user:['کاربر','bg-gray-100 text-gray-500']};
          function fmt(n){return (n||0).toLocaleString('en-US');}
          function esc(s){var d=document.createElement('div');d.textContent=s==null?'':String(s);return d.innerHTML;}

          fetch('/admin/partners/tree-data').then(function(r){return r.json();}).then(function(d){
            T=d; renderStats(); renderRoots();
          }).catch(function(){document.getElementById('tree-root').innerHTML='<div class="text-center text-red-500 py-10">خطا در دریافت داده‌ها</div>';});

          function renderStats(){
            var s=T.stats, box=document.getElementById('tree-stats');
            var cards=[['🤝','کل همکاران',fmt(s.total_partners)],['✅','همکاران فعال',fmt(s.active)],
                       ['🕸','عمق شبکه',fmt(s.max_depth)+' لایه'],['👥','بیشترین زیرمجموعه',fmt(s.max_subs)],
                       ['🛒','فروش شبکه',fmt(s.net_sales)+' ت'],['💰','درآمد شبکه',fmt(s.net_income)+' ت']];
            box.innerHTML=cards.map(function(c){
              return '<div class="card p-3"><div class="text-xs text-gray-400 mb-1">'+c[0]+' '+c[1]+'</div>'
                    +'<div class="font-bold text-gray-800 text-sm">'+c[2]+'</div></div>';
            }).join('');
          }

          function nodeRow(id){
            var n=T.nodes[String(id)], st=ST_LBL[n.status]||ST_LBL.user;
            var row=document.createElement('div');
            row.className='tn-row'; row.id='tn-'+id;
            row.innerHTML=
              '<span class="tn-arrow'+(n.children.length?'':' leaf')+'" data-a>◀</span>'
             +'<span class="tn-tier">'+(n.tier==='—'?'👤':esc(n.tier.split(' ')[0]))+'</span>'
             +'<span class="font-medium text-gray-800">'+esc(n.name)+'</span>'
             +(n.username?'<span class="text-xs text-gray-400" dir="ltr">@'+esc(n.username)+'</span>':'')
             +'<code class="text-xs no-fa" dir="ltr">'+n.id+'</code>'
             +'<span class="tn-badge bg-indigo-50 text-indigo-600">🛒 '+fmt(n.sales)+'</span>'
             +'<span class="tn-badge bg-teal-50 text-teal-700">💰 '+fmt(n.income)+' ت</span>'
             +'<span class="tn-badge bg-gray-100 text-gray-500">👥 '+fmt(n.direct)+' / '+fmt(n.total)+'</span>'
             +'<span class="tn-badge '+st[1]+'">'+st[0]+'</span>';
            row.querySelector('[data-a]').onclick=function(ev){ev.stopPropagation();toggle(id);};
            row.onclick=function(){openDrawer(id);};
            return row;
          }

          function wrap(id){
            var w=document.createElement('div'); w.id='tw-'+id;
            w.appendChild(nodeRow(id));
            return w;
          }

          function renderRoots(){
            var box=document.getElementById('tree-root'); box.innerHTML='';
            if(!T.roots.length){box.innerHTML='<div class="text-center text-gray-400 py-12">هنوز معرفی‌ای ثبت نشده است</div>';return;}
            T.roots.forEach(function(r){box.appendChild(wrap(r));});
          }

          function toggle(id){ EXP[id]?collapse(id):expand(id); }

          function expand(id){
            var n=T.nodes[String(id)]; if(!n||!n.children.length)return;
            var w=document.getElementById('tw-'+id); if(!w)return;
            var kids=w.querySelector(':scope > .tn-kids');
            if(!kids){
              kids=document.createElement('div'); kids.className='tn-kids';
              kids.dataset.next='0'; w.appendChild(kids);
              fill(id,kids);
            }
            kids.style.display='';
            w.querySelector('[data-a]').classList.add('open');
            EXP[id]=true;
          }

          function fill(id,kids){
            var n=T.nodes[String(id)], from=+kids.dataset.next, to=Math.min(from+CHUNK,n.children.length);
            var more=kids.querySelector(':scope > .tn-more'); if(more)more.remove();
            for(var i=from;i<to;i++) kids.appendChild(wrap(n.children[i]));
            kids.dataset.next=to;
            if(to<n.children.length){
              var b=document.createElement('button');
              b.className='tn-more text-xs text-indigo-500 hover:text-indigo-700 py-2 pr-8 block';
              b.textContent='⬇ نمایش '+fmt(n.children.length-to)+' مورد دیگر…';
              b.onclick=function(){fill(id,kids);};
              kids.appendChild(b);
            }
          }

          function collapse(id){
            var w=document.getElementById('tw-'+id); if(!w)return;
            var kids=w.querySelector(':scope > .tn-kids');
            if(kids)kids.style.display='none';
            w.querySelector('[data-a]').classList.remove('open');
            EXP[id]=false;
          }

          function pathOf(id){
            var p=[],cur=id,g=0;
            while(cur!=null&&g++<500){p.unshift(cur);cur=T.nodes[String(cur)]?T.nodes[String(cur)].parent:null;}
            return p;
          }

          window.treeReveal=function(id){
            var p=pathOf(id);
            for(var i=0;i<p.length-1;i++){
              expand(p[i]);
              // اگر گره هدف در چانک‌های بعدی است، تا رسیدن به آن fill کن
              var kids=document.getElementById('tw-'+p[i]).querySelector(':scope > .tn-kids');
              var guard=0;
              while(kids&&!document.getElementById('tw-'+p[i+1])&&guard++<200){
                var mb=kids.querySelector(':scope > .tn-more'); if(!mb)break; mb.click();
              }
            }
            var el=document.getElementById('tn-'+id);
            if(el){el.scrollIntoView({behavior:'smooth',block:'center'});
              el.classList.add('tn-hl'); setTimeout(function(){el.classList.remove('tn-hl');},3500);}
          };

          window.treeSearch=function(){
            var q=(document.getElementById('tree-q').value||'').trim().toLowerCase();
            var out=document.getElementById('tree-results');
            if(!q){out.innerHTML='';return;}
            var hits=[];
            for(var k in T.nodes){var n=T.nodes[k];
              if(String(n.id).indexOf(q)>-1||(n.name||'').toLowerCase().indexOf(q)>-1||(n.username||'').toLowerCase().indexOf(q)>-1)
                {hits.push(n); if(hits.length>=15)break;}
            }
            if(!hits.length){out.innerHTML='<span class="text-gray-400">چیزی یافت نشد</span>';return;}
            out.innerHTML=hits.map(function(n){
              return '<button onclick="treeReveal('+n.id+')" class="ml-2 mb-1 px-2.5 py-1 bg-indigo-50 text-indigo-700 border border-indigo-200 rounded-lg text-xs hover:bg-indigo-100">'
                    +esc(n.name)+' <code>'+n.id+'</code></button>';
            }).join('');
            treeReveal(hits[0].id);
          };

          window.openDrawer=function(id){
            var n=T.nodes[String(id)], st=ST_LBL[n.status]||ST_LBL.user;
            var path=pathOf(id).map(function(pid){
              var pn=T.nodes[String(pid)];
              return '<button onclick="treeReveal('+pid+');treeCloseDrawer()" class="text-indigo-500 hover:underline">'+esc(pn?pn.name:pid)+'</button>';
            }).join(' <span class="text-gray-300">←</span> ');
            var ref=n.parent!=null?T.nodes[String(n.parent)]:null;
            function line(l,v){return '<div class="flex justify-between gap-3 py-2 border-b border-gray-100"><span class="text-gray-400 text-xs">'+l+'</span><span class="font-medium text-gray-700 text-left">'+v+'</span></div>';}
            document.getElementById('tree-drawer-body').innerHTML=
              '<div class="text-center mb-4"><div class="text-3xl mb-1">'+(n.tier==='—'?'👤':esc(n.tier.split(' ')[0]))+'</div>'
             +'<div class="font-bold text-gray-800">'+esc(n.name)+'</div>'
             +(n.username?'<div class="text-xs text-gray-400" dir="ltr">@'+esc(n.username)+'</div>':'')+'</div>'
             +line('آیدی تلگرام','<code>'+n.id+'</code>')
             +line('سطح همکاری',esc(n.tier))
             +line('وضعیت','<span class="tn-badge '+st[1]+'">'+st[0]+'</span>')
             +line('تاریخ عضویت',esc(n.joined||'—'))
             +line('معرف مستقیم',ref?esc(ref.name):'— (ریشه)')
             +line('زیرمجموعه مستقیم',fmt(n.direct))
             +line('کل زیرمجموعه‌ها',fmt(n.total))
             +line('تعداد خرید',fmt(n.sales))
             +line('مجموع خرید',fmt(n.spend)+' ت')
             +line('درآمد (پورسانت)',fmt(n.income)+' ت')
             +line('عمق در شبکه','لایه '+fmt(n.depth))
             +'<div class="mt-4"><div class="text-xs text-gray-400 mb-2">مسیر معرفی</div>'
             +'<div class="leading-7 text-xs">'+path+'</div></div>'
             +'<a href="/admin/partners/'+n.id+'/profile" class="block text-center mt-5 py-2.5 bg-indigo-600 hover:bg-indigo-700 text-white rounded-xl text-sm font-semibold transition">👤 پروفایل کامل</a>';
            document.getElementById('tree-overlay').classList.remove('hidden');
            document.getElementById('tree-drawer').style.transform='translateX(0)';
          };
          window.treeCloseDrawer=function(){
            document.getElementById('tree-overlay').classList.add('hidden');
            document.getElementById('tree-drawer').style.transform='translateX(-105%)';
          };
        })();
        </script>"""

    body = f"""
    <h1 class="text-2xl font-bold text-gray-800 mb-4">🤝 همکاران</h1>
    {tabs_html}
    {content}"""
    return _layout("همکاران", body, adm, flash=flash)


# ─── 🌳 درخت همکاران — موتور داده ─────────────────────────────────────────

def _build_partner_tree(conn) -> dict:
    """ساخت درخت کامل معرفی‌ها + آمار — O(N+E)، بدون محدودیت عمق، ضد حلقه."""
    from collections import deque

    edges = conn.execute(
        "SELECT referrer_id, referred_id, COALESCE(created_at,'') AS ca FROM referrals;"
    ).fetchall()
    users = {}
    for r in conn.execute("SELECT user_id, COALESCE(username,'') u, COALESCE(full_name,'') f, COALESCE(first_seen,'') fs FROM users;").fetchall():
        try: users[int(r["user_id"])] = r
        except Exception: pass
    partners = {}
    for r in conn.execute("SELECT tg_user_id, COALESCE(full_name,'') fn, COALESCE(status,'') st FROM partners;").fetchall():
        try: partners[int(r["tg_user_id"])] = r
        except Exception: pass
    tiers = conn.execute(
        "SELECT name, icon, min_orders, COALESCE(color,'#6B7280') color FROM partner_tiers ORDER BY min_orders ASC;"
    ).fetchall()
    sales = {}
    for r in conn.execute("""
        SELECT CAST(user_id AS INTEGER) u, COUNT(*) c, COALESCE(SUM(price),0) s,
               SUM(CASE WHEN buyer_type='partner' THEN 1 ELSE 0 END) pc
        FROM orders WHERE COALESCE(status,'active')!='returned'
        GROUP BY CAST(user_id AS INTEGER);""").fetchall():
        try: sales[int(r["u"])] = (int(r["c"]), int(r["s"]), int(r["pc"] or 0))
        except Exception: pass
    income = {}
    try:
        for r in conn.execute("SELECT user_id u, COALESCE(SUM(amount),0) s FROM partner_transactions WHERE type='credit' GROUP BY user_id;").fetchall():
            income[int(r["u"])] = int(r["s"])
    except Exception:
        pass

    # والد هر گره — اولین معرفی معتبر است؛ خودارجاعی رد می‌شود
    parent, joined, node_ids = {}, {}, set()
    for e in edges:
        try:
            a, b = int(e["referrer_id"]), int(e["referred_id"])
        except Exception:
            continue
        if a == b:
            continue
        node_ids.add(a); node_ids.add(b)
        if b not in parent:
            parent[b] = a
            joined[b] = (e["ca"] or "")[:10]

    # شکستن حلقه‌های احتمالی (داده خراب) — گرهِ حلقه‌ساز ریشه می‌شود
    for n in list(parent.keys()):
        seen, cur = set(), n
        while cur in parent:
            if cur in seen:
                parent.pop(n, None)
                break
            seen.add(cur)
            cur = parent[cur]

    children = {}
    for b, a in parent.items():
        children.setdefault(a, []).append(b)
    roots = sorted(n for n in node_ids if n not in parent)

    # عمق (BFS) + ترتیب برای پیمایش
    depth, order, dq = {}, [], deque((r, 1) for r in roots)
    while dq:
        n, d = dq.popleft()
        depth[n] = d
        order.append(n)
        for c in children.get(n, ()):
            dq.append((c, d + 1))

    # کل زیرمجموعه‌ها — post-order تکراری
    total_subs = {n: 0 for n in node_ids}
    for n in reversed(order):
        t = 0
        for c in children.get(n, ()):
            t += 1 + total_subs.get(c, 0)
        total_subs[n] = t

    def _tier_of(pc):
        cur = None
        for t in tiers:
            if pc >= int(t["min_orders"] or 0):
                cur = t
        return cur

    nodes = {}
    for n in node_ids:
        u, p = users.get(n), partners.get(n)
        c, s, pc = sales.get(n, (0, 0, 0))
        t = _tier_of(pc)
        name = (p["fn"] if p and p["fn"] else
                (u["f"] if u and u["f"] else
                 (u["u"] if u and u["u"] else f"کاربر {n}")))
        kids = children.get(n, [])
        kids.sort(key=lambda x: -total_subs.get(x, 0))
        nodes[str(n)] = {
            "id": n, "name": name,
            "username": (u["u"] if u else ""),
            "parent": parent.get(n),
            "children": kids,
            "tier": (f'{t["icon"]} {t["name"]}' if t else "—"),
            "tcolor": (t["color"] if t else "#6B7280"),
            "sales": c, "spend": s, "income": income.get(n, 0),
            "direct": len(kids), "total": total_subs.get(n, 0),
            "depth": depth.get(n, 1),
            "status": (p["st"] if p else "user"),
            "joined": fa_date(joined.get(n, "") or (u["fs"] if u else "")),
        }

    stats = {
        "total_partners": len(partners),
        "active": sum(1 for p in partners.values() if p["st"] == "approved"),
        "network": len(node_ids),
        "max_depth": max(depth.values()) if depth else 0,
        "max_subs": max((len(v) for v in children.values()), default=0),
        "net_sales": sum(sales.get(n, (0, 0, 0))[1] for n in node_ids),
        "net_income": sum(income.get(n, 0) for n in node_ids),
    }
    return {"stats": stats, "roots": roots, "nodes": nodes}


@router.get("/partners/tree-data")
async def partners_tree_data(request: Request):
    """JSON درخت همکاران — فقط ادمین"""
    adm = _get_admin(request)
    guard = _require(adm, "partners")
    if guard: return guard
    from db import ensure_referral_schema, ensure_partner_tiers_extended
    ensure_referral_schema()
    ensure_partner_tiers_extended()
    conn = _db()
    try:
        data = _build_partner_tree(conn)
    finally:
        conn.close()
    return JSONResponse(data)


@router.post("/partners/payout-settings")
async def partner_payout_settings_save(request: Request):
    adm = _get_admin(request)
    guard = _require(adm, "partners")
    if guard: return guard
    form = await request.form()
    from db import save_payout_settings_full, ensure_payout_settings_extended
    ensure_payout_settings_extended()
    save_payout_settings_full({
        "min_amount":         int(form.get("min_amount") or 50000),
        "max_amount":         int(form.get("max_amount") or 0),
        "max_per_month":      int(form.get("max_per_month") or 2),
        "is_active":          int(form.get("is_active") or 1),
        "review_hours":       int(form.get("review_hours") or 48),
        "guide_text":         str(form.get("guide_text","")).strip(),
        "approval_message":   str(form.get("approval_message","")).strip(),
        "rejection_message":  str(form.get("rejection_message","")).strip(),
    })
    _log(request, "تنظیمات تسویه", "همکاران", "updated")
    return _redir("/admin/partners?tab=settings&flash=تنظیمات+تسویه+ذخیره+شد")


@router.get("/partners/{uid}/profile", response_class=HTMLResponse)
async def partner_profile_admin(request: Request, uid: int):
    """پروفایل کامل همکار — مشخصات + خریدها + زیرمجموعه‌های مستقیم"""
    adm = _get_admin(request)
    guard = _require(adm, "partners")
    if guard: return guard

    conn = _db()
    try:
        p = conn.execute("SELECT * FROM partners WHERE tg_user_id=?;", (str(uid),)).fetchone()
        if not p:
            p = conn.execute("SELECT * FROM partners WHERE CAST(tg_user_id AS INTEGER)=?;", (uid,)).fetchone()

        orders = conn.execute("""
            SELECT * FROM orders WHERE CAST(user_id AS INTEGER)=?
            ORDER BY id DESC LIMIT 50;
        """, (uid,)).fetchall()
        o_total = conn.execute(
            "SELECT COUNT(*), COALESCE(SUM(price),0) FROM orders WHERE CAST(user_id AS INTEGER)=? AND COALESCE(status,'active')!='returned';",
            (uid,)).fetchone()

        subs = conn.execute("""
            SELECT r.referred_id,
                   COALESCE(u.full_name, u.username, 'کاربر ' || r.referred_id) AS name,
                   r.created_at,
                   COALESCE(o.cnt,0) AS ocnt, COALESCE(o.total,0) AS ototal
            FROM referrals r
            LEFT JOIN users u ON CAST(u.user_id AS INTEGER)=r.referred_id
            LEFT JOIN (SELECT CAST(user_id AS INTEGER) ouid, COUNT(*) cnt, SUM(price) total
                       FROM orders WHERE COALESCE(status,'active')!='returned'
                       GROUP BY CAST(user_id AS INTEGER)) o ON o.ouid=r.referred_id
            WHERE r.referrer_id=? ORDER BY ototal DESC;
        """, (uid,)).fetchall()

        try:
            pw = conn.execute("SELECT COALESCE(balance,0) FROM partner_wallets WHERE user_id=?;", (uid,)).fetchone()
            pw_bal = int(pw[0]) if pw else 0
        except Exception:
            pw_bal = 0
    finally:
        conn.close()

    name  = e((p["full_name"] if p else "") or f"کاربر {uid}")
    phone = e((p["phone"] if p else "") or "—")
    city  = e((p["city"] if p else "") or "—")
    shop  = e((p["shop_name"] if p else "") or "—")
    st    = (p["status"] if p else "—") or "—"
    st_bc = {"pending":"yellow","approved":"green","rejected":"red"}.get(st,"gray")
    st_bl = {"pending":"در انتظار","approved":"تایید شده","rejected":"رد شده"}.get(st,st)

    order_rows = "".join(f"""<tr class="border-b hover:bg-gray-50 text-sm">
        <td class="px-4 py-2.5 text-xs text-gray-400">#{o['id']}</td>
        <td class="px-4 py-2.5">{e(o['title'] or '—')}</td>
        <td class="px-4 py-2.5 font-medium">{int(o['price'] or 0):,} ت</td>
        <td class="px-4 py-2.5">{'<span class="px-2 py-0.5 text-xs bg-red-100 text-red-600 rounded-full">برگشتی</span>' if (o['status'] or 'active')=='returned' else '<span class="px-2 py-0.5 text-xs bg-green-100 text-green-700 rounded-full">فعال</span>'}</td>
        <td class="px-4 py-2.5 text-xs text-gray-400">{fa_date(o['created_at'] or '')}</td>
      </tr>""" for o in orders)

    sub_rows = "".join(f"""<tr class="border-b hover:bg-gray-50 text-sm">
        <td class="px-4 py-2.5 font-mono text-xs"><code>{s['referred_id']}</code></td>
        <td class="px-4 py-2.5">{e(s['name'])}</td>
        <td class="px-4 py-2.5">{int(s['ocnt'])} خرید</td>
        <td class="px-4 py-2.5 font-medium text-green-600">{int(s['ototal']):,} ت</td>
        <td class="px-4 py-2.5 text-xs text-gray-400">{fa_date(s['created_at'])}</td>
      </tr>""" for s in subs)

    body = f"""
    <div class="flex items-center gap-3 mb-6 flex-wrap">
      {_btn("← همکاران", "/admin/partners", "slate", small=True)}
      <h1 class="text-xl font-bold text-gray-800">👤 {name}</h1>
      <span class="px-2.5 py-1 text-xs rounded-full bg-{st_bc}-100 text-{st_bc}-700">{st_bl}</span>
    </div>

    <div class="grid grid-cols-2 md:grid-cols-4 gap-4 mb-6">
      {_card("خریدها", f"{int(o_total[0] or 0)}", "", "indigo")}
      {_card("جمع خرید", f"{int(o_total[1] or 0):,} ت", "", "green")}
      {_card("زیرمجموعه مستقیم", f"{len(subs)}", "", "amber")}
      {_card("کیف‌پول همکاری", f"{pw_bal:,} ت", "", "teal")}
    </div>

    <div class="card p-5 mb-6">
      <h2 class="font-bold text-gray-700 mb-3">📇 مشخصات</h2>
      <div class="grid grid-cols-2 md:grid-cols-4 gap-3 text-sm">
        <div><span class="text-xs text-gray-400 block">User ID</span><code>{uid}</code></div>
        <div><span class="text-xs text-gray-400 block">شماره</span>{phone}</div>
        <div><span class="text-xs text-gray-400 block">شهر</span>{city}</div>
        <div><span class="text-xs text-gray-400 block">فروشگاه</span>{shop}</div>
      </div>
    </div>

    <div class="grid lg:grid-cols-2 gap-6">
      <div class="card overflow-hidden">
        <div class="px-5 py-4 border-b"><h2 class="font-bold text-gray-700">🛒 خریدها (۵۰ مورد آخر)</h2></div>
        <div class="overflow-x-auto"><table class="w-full text-right min-w-max">
          <thead><tr class="text-xs text-gray-500 border-b bg-gray-50">
            <th class="px-4 py-2.5">#</th><th class="px-4 py-2.5">محصول</th>
            <th class="px-4 py-2.5">مبلغ</th><th class="px-4 py-2.5">وضعیت</th><th class="px-4 py-2.5">تاریخ</th>
          </tr></thead>
          <tbody>{order_rows or "<tr><td colspan='5' class='text-center py-8 text-gray-400'>خریدی ندارد</td></tr>"}</tbody>
        </table></div>
      </div>
      <div class="card overflow-hidden">
        <div class="px-5 py-4 border-b"><h2 class="font-bold text-gray-700">👥 زیرمجموعه‌های مستقیم</h2></div>
        <div class="overflow-x-auto"><table class="w-full text-right min-w-max">
          <thead><tr class="text-xs text-gray-500 border-b bg-gray-50">
            <th class="px-4 py-2.5">ID</th><th class="px-4 py-2.5">نام</th>
            <th class="px-4 py-2.5">خرید</th><th class="px-4 py-2.5">مبلغ</th><th class="px-4 py-2.5">عضویت</th>
          </tr></thead>
          <tbody>{sub_rows or "<tr><td colspan='5' class='text-center py-8 text-gray-400'>زیرمجموعه‌ای ندارد</td></tr>"}</tbody>
        </table></div>
      </div>
    </div>"""

    return _layout(f"پروفایل همکار", body, adm)


@router.get("/partners/payout/{pid}", response_class=HTMLResponse)
async def partner_payout_detail(request: Request, pid: int, flash: str = ""):
    adm = _get_admin(request)
    guard = _require(adm, "partners")
    if guard: return guard

    from db import (ensure_partner_wallet_schema, ensure_partner_bank_schema,
                    get_partner_bank_info, get_partner_wallet_balance,
                    get_partner_payouts, ensure_partner_tiers_extended,
                    get_partner_tier_for, get_partner_order_count)
    ensure_partner_wallet_schema(); ensure_partner_bank_schema(); ensure_partner_tiers_extended()

    conn = _db()
    import sqlite3 as _sq3; conn.row_factory = _sq3.Row
    try:
        pay = conn.execute("""
            SELECT p.*, u.full_name as u_name, u.username, u.first_seen, pr.phone, pr.shop_name, pr.city
            FROM partner_payouts p
            LEFT JOIN users u ON u.user_id=p.user_id
            LEFT JOIN partners pr ON pr.tg_user_id=p.user_id
            WHERE p.id=?;
        """, (pid,)).fetchone()
        if not pay:
            return _redir("/admin/partners?tab=payouts&flash=درخواست+یافت+نشد")

        uid = pay["user_id"]
        # تاریخچه تسویه‌ها
        prev_payouts = conn.execute("""
            SELECT COUNT(*) as cnt, COALESCE(SUM(amount),0) as total
            FROM partner_payouts WHERE user_id=? AND status='approved';
        """, (uid,)).fetchone()
        last_payout = conn.execute("""
            SELECT created_at, amount FROM partner_payouts
            WHERE user_id=? AND status='approved' ORDER BY id DESC LIMIT 1;
        """, (uid,)).fetchone()
        # مجموع پورسانت
        total_commission = conn.execute(
            "SELECT COALESCE(SUM(reward_amount),0) FROM referrals WHERE referrer_id=? AND rewarded=1;",
            (uid,)
        ).fetchone()[0]
    finally:
        conn.close()

    bank   = get_partner_bank_info(uid) or {}
    bal    = get_partner_wallet_balance(uid)
    order_cnt = get_partner_order_count(uid)
    tier   = get_partner_tier_for(order_cnt)

    sc = {"pending":"amber","approved":"green","rejected":"red"}.get(pay["status"],"gray")
    sl = {"pending":"در انتظار","approved":"تأیید شد","rejected":"رد شد"}.get(pay["status"], pay["status"])

    action_btns = ""
    if pay["status"] == "pending":
        action_btns = f"""
        <div class="card p-6">
          <h2 class="font-bold text-gray-700 mb-4">⚡ عملیات</h2>
          <div class="flex gap-3 flex-wrap">
            <form method="post" action="/admin/partners/payout/{pid}/approve">
              <button class="px-6 py-2.5 bg-green-600 text-white rounded-xl text-sm font-bold hover:bg-green-700">
                ✅ تأیید و پرداخت
              </button>
            </form>
            <button onclick="document.getElementById('reject-form').classList.toggle('hidden')"
              class="px-6 py-2.5 bg-red-50 text-red-600 border border-red-200 rounded-xl text-sm font-bold hover:bg-red-100">
              ❌ رد درخواست
            </button>
          </div>
          <form method="post" action="/admin/partners/payout/{pid}/reject" id="reject-form" class="hidden mt-4">
            <label class="text-sm font-medium text-gray-700 block mb-2">دلیل رد (نمایش به همکار):</label>
            {_textarea("note","توضیح اختیاری...",rows=3)}
            <div class="mt-3">
              <button type="submit" class="px-5 py-2 bg-red-600 text-white rounded-lg text-sm font-medium">
                ثبت رد درخواست
              </button>
            </div>
          </form>
        </div>"""

    body = f"""
    <div class="flex items-center gap-3 mb-6">
      {_btn("← تسویه‌ها", "/admin/partners?tab=payouts", "slate", small=True)}
      <h1 class="text-2xl font-bold text-gray-800">درخواست تسویه #{pid}</h1>
      <span class="px-3 py-1 text-sm bg-{sc}-100 text-{sc}-700 rounded-full font-medium">{sl}</span>
    </div>

    <div class="grid md:grid-cols-3 gap-4 mb-4">
      <div class="card p-5 text-center border-t-4 border-green-400">
        <div class="text-2xl font-bold text-green-600">{int(pay['amount']):,}</div>
        <div class="text-xs text-gray-400 mt-1">مبلغ درخواست (تومان)</div>
      </div>
      <div class="card p-5 text-center border-t-4 border-blue-400">
        <div class="text-2xl font-bold text-blue-600">{bal:,}</div>
        <div class="text-xs text-gray-400 mt-1">موجودی کیف‌پول (تومان)</div>
      </div>
      <div class="card p-5 text-center border-t-4 border-purple-400">
        <div class="text-2xl font-bold text-purple-600">{int(total_commission or 0):,}</div>
        <div class="text-xs text-gray-400 mt-1">مجموع پورسانت دریافتی</div>
      </div>
    </div>

    <div class="grid md:grid-cols-2 gap-4 mb-4">
      <!-- اطلاعات همکار -->
      <div class="card p-6">
        <h2 class="font-bold text-gray-700 mb-4">👤 اطلاعات همکار</h2>
        <div class="space-y-2 text-sm">
          <div class="flex justify-between"><span class="text-gray-400">نام</span><span class="font-medium">{e(pay['u_name'] or '—')}</span></div>
          <div class="flex justify-between"><span class="text-gray-400">فروشگاه</span><span>{e(pay['shop_name'] or '—')}</span></div>
          <div class="flex justify-between"><span class="text-gray-400">شهر</span><span>{e(pay['city'] or '—')}</span></div>
          <div class="flex justify-between"><span class="text-gray-400">موبایل</span><span>{e(pay['phone'] or '—')}</span></div>
          <div class="flex justify-between"><span class="text-gray-400">عضویت</span><span>{fa_date(pay['first_seen'] or '')}</span></div>
          <div class="flex justify-between"><span class="text-gray-400">سطح</span><span>{tier['icon']} {tier['name']}</span></div>
          <div class="flex justify-between"><span class="text-gray-400">User ID</span><code class="text-xs bg-gray-100 px-1.5 rounded">{uid}</code></div>
        </div>
        <div class="mt-3">
          <a href="/admin/users/{uid}" class="text-xs text-indigo-600 hover:underline">مشاهده پروفایل کامل ↗</a>
        </div>
      </div>

      <!-- اطلاعات بانکی -->
      <div class="card p-6">
        <h2 class="font-bold text-gray-700 mb-4">💳 اطلاعات بانکی</h2>
        <div class="space-y-2 text-sm">
          <div class="flex justify-between"><span class="text-gray-400">صاحب حساب</span><span class="font-medium">{e(bank.get('full_name','—'))}</span></div>
          <div class="flex justify-between"><span class="text-gray-400">شماره کارت</span><code class="text-xs bg-gray-100 px-2 py-0.5 rounded">{e(bank.get('card_number','—'))}</code></div>
          <div class="flex justify-between"><span class="text-gray-400">شماره شبا</span><code class="text-xs bg-gray-100 px-2 py-0.5 rounded">{e(bank.get('iban','—'))}</code></div>
        </div>
        <div class="mt-4 p-3 bg-amber-50 rounded-lg text-xs text-amber-700">
          ⚠️ قبل از تأیید، اطلاعات بانکی را تأیید کنید
        </div>
      </div>
    </div>

    <div class="card p-6 mb-4">
      <h2 class="font-bold text-gray-700 mb-4">📊 تاریخچه مالی</h2>
      <div class="grid grid-cols-2 md:grid-cols-4 gap-4 text-sm">
        <div class="text-center p-3 bg-gray-50 rounded-lg">
          <div class="font-bold text-gray-700">{int(prev_payouts['cnt'] or 0)}</div>
          <div class="text-xs text-gray-400">تسویه‌های موفق</div>
        </div>
        <div class="text-center p-3 bg-gray-50 rounded-lg">
          <div class="font-bold text-gray-700">{int(prev_payouts['total'] or 0):,}</div>
          <div class="text-xs text-gray-400">مجموع تسویه (ت)</div>
        </div>
        <div class="text-center p-3 bg-gray-50 rounded-lg">
          <div class="font-bold text-gray-700">{fa_date(last_payout['created_at'] or '—') if last_payout else '—'}</div>
          <div class="text-xs text-gray-400">آخرین تسویه</div>
        </div>
        <div class="text-center p-3 bg-gray-50 rounded-lg">
          <div class="font-bold text-gray-700">{order_cnt}</div>
          <div class="text-xs text-gray-400">تعداد خرید همکاری</div>
        </div>
      </div>
    </div>

    {action_btns}"""

    return _layout(f"تسویه #{pid}", body, adm, flash=flash)


@router.post("/partners/payout/{pid}/approve")
async def partner_payout_approve(request: Request, pid: int):
    adm = _get_admin(request)
    guard = _require(adm, "partners")
    if guard: return guard
    from db import process_partner_payout, get_payout_settings_full, ensure_payout_settings_extended
    ensure_payout_settings_extended()
    result = process_partner_payout(pid, approve=True)
    if result["ok"]:
        uid = result["user_id"]; amt = result["amount"]
        ps = get_payout_settings_full()
        hours = ps.get("review_hours", 48)
        msg = ps.get("approval_message","") or f"✅ درخواست تسویه {amt:,} تومان تأیید شد و ظرف {hours} ساعت پرداخت می‌شود."
        _log(request, "تأیید تسویه", "همکاران", f"payout:{pid} user:{uid} amount:{amt}")
        try: await run_in_threadpool(_tg_send, uid, msg)
        except Exception: pass
    return _redir("/admin/receipts?flash=تسویه+تأیید+شد#payouts")


@router.post("/partners/payout/{pid}/reject")
async def partner_payout_reject(request: Request, pid: int):
    adm = _get_admin(request)
    guard = _require(adm, "partners")
    if guard: return guard
    from db import process_partner_payout, get_payout_settings_full, ensure_payout_settings_extended
    ensure_payout_settings_extended()
    form = await request.form()
    note = str(form.get("note","")).strip()
    result = process_partner_payout(pid, approve=False, admin_note=note)
    if result["ok"]:
        uid = result["user_id"]; amt = result["amount"]
        ps = get_payout_settings_full()
        msg = ps.get("rejection_message","") or f"❌ درخواست تسویه {amt:,} تومان رد شد."
        if note: msg += f"\n\nدلیل: {note}"
        msg += f"\n\nمبلغ به کیف‌پول همکاری برگشت داده شد."
        _log(request, "رد تسویه", "همکاران", f"payout:{pid} user:{uid} amount:{amt}")
        try: await run_in_threadpool(_tg_send, uid, msg)
        except Exception: pass
    return _redir("/admin/receipts?flash=تسویه+رد+شد#payouts")


@router.post("/partners/tier/{tid}/delete-banner")
async def partner_tier_delete_banner(request: Request, tid: int):
    adm = _get_admin(request)
    guard = _require(adm, "partners")
    if guard: return guard
    conn = _db()
    try:
        conn.execute("UPDATE partner_tiers SET photo_file_id='' WHERE id=?;", (tid,))
        conn.commit()
    finally:
        conn.close()
    _log(request, f"حذف بنر سطح #{tid}", "همکاران")
    return _redir("/admin/partners?tab=settings&flash=بنر+حذف+شد")


@router.post("/partners/tier/{tid}/upload-banner")
async def partner_tier_upload_banner(request: Request, tid: int):
    adm = _get_admin(request)
    guard = _require(adm, "partners")
    if guard: return guard
    form = await request.form()
    file = form.get("banner_file")
    if not file or not file.filename:
        return _redir(f"/admin/partners?tab=settings&flash=فایلی+انتخاب+نشد")
    file_bytes = await file.read()
    # آپلود به تلگرام و دریافت file_id
    try:
        import requests as _req
        token = _env("BOT_TOKEN", "")
        admin_tg = _env("ADMIN_ID", "")
        if not token or not admin_tg:
            raise ValueError("BOT_TOKEN or ADMIN_ID not set")
        resp = _req.post(
            f"https://api.telegram.org/bot{token}/sendPhoto",
            data={"chat_id": admin_tg, "caption": f"بنر سطح #{tid}"},
            files={"photo": (file.filename, file_bytes, file.content_type or "image/jpeg")},
            timeout=15
        )
        result = resp.json()
        if not result.get("ok"):
            raise ValueError(result.get("description", "upload failed"))
        photo_id = result["result"]["photo"][-1]["file_id"]
        # ذخیره
        conn = _db()
        conn.execute("UPDATE partner_tiers SET photo_file_id=? WHERE id=?;", (photo_id, tid))
        conn.commit()
        conn.close()
        _log(request, f"آپلود بنر سطح #{tid}", "همکاران", photo_id[:20])
        return _redir(f"/admin/partners?tab=settings&flash=بنر+آپلود+شد")
    except Exception as ex:
        return _redir(f"/admin/partners?tab=settings&flash=خطا:+{str(ex)[:40]}")


@router.post("/partners/tier/save")
async def partner_tier_save(request: Request):
    adm = _get_admin(request)
    guard = _require(adm, "partners")
    if guard: return guard
    form = await request.form()
    tid  = form.get("tier_id")
    tid  = int(tid) if tid and str(tid).isdigit() else None
    from db import ensure_partner_tiers_extended
    ensure_partner_tiers_extended()
    conn = _db()
    try:
        name   = str(form.get("name","")).strip()
        icon   = str(form.get("icon","🥉")).strip()
        min_o  = int(form.get("min_orders") or 0)
        # فاز ۲: نوع پورسانت رادیویی — فقط یکی فعال
        ctype  = str(form.get("commission_type","percent")).strip()
        if ctype not in ("percent","fixed"): ctype = "percent"
        comm   = float(form.get("commission_percent") or 0) if ctype == "percent" else 0
        cfixed = int(form.get("commission_fixed") or 0) if ctype == "fixed" else 0
        min_amt = int(form.get("min_order_amount") or 0)
        max_pay = int(form.get("max_payout") or 0)
        color  = str(form.get("color","#6B7280")).strip()
        desc   = str(form.get("description","")).strip()
        photo  = str(form.get("photo_file_id","")).strip()
        levelup = str(form.get("levelup_message","")).strip()
        if tid:
            conn.execute("""UPDATE partner_tiers SET name=?,icon=?,min_orders=?,
                commission_percent=?,commission_fixed=?,color=?,description=?,photo_file_id=?,
                min_order_amount=?,max_payout=?,levelup_message=?,commission_type=? WHERE id=?;""",
                (name, icon, min_o, comm, cfixed, color, desc, photo,
                 min_amt, max_pay, levelup, ctype, tid))
        else:
            mx = conn.execute("SELECT COALESCE(MAX(sort_order),0)+1 FROM partner_tiers;").fetchone()[0]
            conn.execute("""INSERT INTO partner_tiers
                (name,icon,min_orders,commission_percent,commission_fixed,color,description,photo_file_id,
                 min_order_amount,max_payout,levelup_message,commission_type,sort_order)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?);""",
                (name, icon, min_o, comm, cfixed, color, desc, photo,
                 min_amt, max_pay, levelup, ctype, mx))
        conn.commit()
    finally:
        conn.close()
    _log(request, "ذخیره سطح همکاری", "همکاران", name, admin_info=adm)
    return _redir("/admin/partners?tab=settings&flash=✅+سطح+ذخیره+شد")


# فاز ۲: صفحه ادیت مستقل سطح (بازتر و کاربردی‌تر از فرم inline)
@router.get("/partners/tier/{tid}/edit", response_class=HTMLResponse)
@router.get("/partners/tier/new", response_class=HTMLResponse)
async def partner_tier_edit_page(request: Request, tid: int = 0, flash: str = ""):
    adm = _get_admin(request)
    guard = _require(adm, "partners")
    if guard: return guard
    from db import ensure_partner_tiers_extended
    ensure_partner_tiers_extended()

    # اگه tid داشت، بارگذاری از DB
    tr = None
    is_new = (tid == 0)
    if not is_new:
        conn = _db()
        try:
            r = conn.execute("SELECT * FROM partner_tiers WHERE id=?;", (tid,)).fetchone()
            tr = dict(r) if r else None
        finally:
            conn.close()
        if not tr:
            return _redir("/admin/partners?tab=settings&flash=❌+سطح+یافت+نشد")

    def _g(k, default=""):
        return tr.get(k, default) if tr else default

    ctype     = _g("commission_type","percent") or "percent"
    photo_id  = _g("photo_file_id","") or ""
    title = "سطح جدید" if is_new else f"ویرایش سطح: {e(_g('name',''))}"

    body = f"""
    <div class="max-w-3xl mx-auto p-4">
      <div class="mb-4 flex items-center gap-3">
        <a href="/admin/partners?tab=settings" class="text-gray-400 hover:text-gray-600 text-2xl">←</a>
        <h1 class="text-xl font-bold text-gray-800">🏆 {title}</h1>
      </div>

      <form method="post" action="/admin/partners/tier/save" class="space-y-4">
        <input type="hidden" name="tier_id" value="{tid or ''}">

        <!-- بخش ۱: هویت سطح -->
        <div class="card p-5">
          <h2 class="font-bold text-gray-700 text-sm mb-4">👑 هویت سطح</h2>
          <div class="grid grid-cols-2 md:grid-cols-4 gap-3">
            <div><label class="text-xs text-gray-500 block mb-1">آیکون</label>
              <input type="text" name="icon" value="{e(_g('icon','🥉'))}" required
                class="w-full border border-gray-200 rounded-lg px-3 py-2 text-lg text-center"></div>
            <div class="col-span-2"><label class="text-xs text-gray-500 block mb-1">نام سطح</label>
              <input type="text" name="name" value="{e(_g('name',''))}" required placeholder="مثلاً: طلا"
                class="w-full border border-gray-200 rounded-lg px-3 py-2 text-sm"></div>
            <div><label class="text-xs text-gray-500 block mb-1">رنگ</label>
              <input type="color" name="color" value="{_g('color','#6B7280')}"
                class="w-full h-10 border border-gray-200 rounded-lg"></div>
          </div>
          <div class="mt-3">
            <label class="text-xs text-gray-500 block mb-1">توضیح (نمایش در پنل همکار)</label>
            <input type="text" name="description" value="{e(_g('description',''))}" placeholder="مثلاً: دسترسی به تخفیف ویژه"
              class="w-full border border-gray-200 rounded-lg px-3 py-2 text-sm">
          </div>
        </div>

        <!-- بخش ۲: بنر اختصاصی -->
        <div class="card p-5">
          <h2 class="font-bold text-gray-700 text-sm mb-3">🖼 بنر اختصاصی سطح</h2>
          <p class="text-xs text-gray-400 mb-3">در پیام تبریک ارتقا به همکار نمایش داده می‌شود.</p>
          {f'<div class="mb-3 p-3 bg-green-50 border border-green-200 rounded-lg text-xs text-green-700">✅ بنر ذخیره شده — file_id: <code class="no-fa" dir="ltr">{photo_id[:40]}...</code></div>' if photo_id else '<div class="mb-3 p-3 bg-gray-50 border border-gray-200 rounded-lg text-xs text-gray-500">هنوز بنری آپلود نشده</div>'}
          <input type="hidden" name="photo_file_id" value="{e(photo_id)}">
          <p class="text-[10px] text-gray-400">💡 برای آپلود بنر جدید، ابتدا فرم را ذخیره کنید، سپس در جدول سطوح از دکمه «📤 آپلود بنر» استفاده کنید.</p>
        </div>

        <!-- بخش ۳: شرط ارتقا -->
        <div class="card p-5">
          <h2 class="font-bold text-gray-700 text-sm mb-4">📈 شرط ارتقا به این سطح</h2>
          <div><label class="text-xs text-gray-500 block mb-1">حداقل تعداد خرید موفق</label>
            <input type="number" name="min_orders" value="{_g('min_orders',0)}" min="0" required
              class="w-full border border-gray-200 rounded-lg px-3 py-2 text-sm">
            <p class="text-[10px] text-gray-400 mt-1">همکاری که این تعداد خرید موفق داشته باشد، خودکار به این سطح ارتقا می‌یابد.</p></div>
        </div>

        <!-- بخش ۴: نوع پورسانت -->
        <div class="card p-5">
          <h2 class="font-bold text-gray-700 text-sm mb-4">💰 نوع پورسانت</h2>
          <div class="grid grid-cols-2 gap-3 mb-4">
            <label class="p-3 border-2 border-gray-200 rounded-xl cursor-pointer hover:border-indigo-400 has-[:checked]:border-indigo-500 has-[:checked]:bg-indigo-50">
              <input type="radio" name="commission_type" value="percent" {"checked" if ctype=="percent" else ""} class="ml-2" onchange="toggleCommission()">
              <span class="font-semibold text-sm">درصدی</span>
              <p class="text-[10px] text-gray-400 mt-1">درصدی از مبلغ خرید</p>
            </label>
            <label class="p-3 border-2 border-gray-200 rounded-xl cursor-pointer hover:border-indigo-400 has-[:checked]:border-indigo-500 has-[:checked]:bg-indigo-50">
              <input type="radio" name="commission_type" value="fixed" {"checked" if ctype=="fixed" else ""} class="ml-2" onchange="toggleCommission()">
              <span class="font-semibold text-sm">مبلغ ثابت</span>
              <p class="text-[10px] text-gray-400 mt-1">مبلغ ثابت هر خرید</p>
            </label>
          </div>
          <div id="pct_wrap" class="grid grid-cols-1 gap-3">
            <label class="text-xs text-gray-500 block">درصد پورسانت (٪)</label>
            <input type="number" name="commission_percent" step="0.1" min="0" value="{_g('commission_percent',0)}"
              class="w-full border border-gray-200 rounded-lg px-3 py-2 text-sm">
          </div>
          <div id="fixed_wrap" class="grid grid-cols-1 gap-3" style="display:none">
            <label class="text-xs text-gray-500 block">مبلغ ثابت پورسانت (تومان)</label>
            <input type="number" name="commission_fixed" min="0" value="{_g('commission_fixed',0)}"
              class="w-full border border-gray-200 rounded-lg px-3 py-2 text-sm">
          </div>
          <div class="grid grid-cols-1 md:grid-cols-2 gap-3 mt-4">
            <div><label class="text-xs text-gray-500 block mb-1">حداقل مبلغ خرید (تومان)</label>
              <input type="number" name="min_order_amount" value="{_g('min_order_amount',0)}" min="0"
                class="w-full border border-gray-200 rounded-lg px-3 py-2 text-sm">
              <p class="text-[10px] text-gray-400 mt-1">۰ = بدون محدودیت</p></div>
            <div><label class="text-xs text-gray-500 block mb-1">سقف پورسانت هر خرید (تومان)</label>
              <input type="number" name="max_payout" value="{_g('max_payout',0)}" min="0"
                class="w-full border border-gray-200 rounded-lg px-3 py-2 text-sm">
              <p class="text-[10px] text-gray-400 mt-1">۰ = نامحدود</p></div>
          </div>
        </div>

        <!-- بخش ۵: پیام تبریک ارتقا -->
        <div class="card p-5">
          <h2 class="font-bold text-gray-700 text-sm mb-3">🎉 پیام تبریک ارتقا</h2>
          <p class="text-xs text-gray-400 mb-3">این پیام به صورت خودکار پس از ارتقای همکار به این سطح ارسال می‌شود.</p>
          <textarea name="levelup_message" rows="8" class="w-full border border-gray-200 rounded-lg px-3 py-2 text-sm font-mono" dir="rtl">{e(_g('levelup_message',''))}</textarea>
          <div class="mt-2 p-3 bg-blue-50 rounded-lg text-[11px] text-blue-700 leading-6">
            💡 <b>متغیرهای قابل استفاده در متن:</b><br>
            <code class="no-fa" dir="ltr">{{name}}</code> — نام همکار &nbsp;
            <code class="no-fa" dir="ltr">{{tier}}</code> — نام سطح جدید &nbsp;
            <code class="no-fa" dir="ltr">{{icon}}</code> — آیکون سطح<br>
            <code class="no-fa" dir="ltr">{{percent}}</code> — درصد پورسانت &nbsp;
            <code class="no-fa" dir="ltr">{{fixed}}</code> — مبلغ ثابت پورسانت<br>
            <code class="no-fa" dir="ltr">{{orders}}</code> — تعداد خرید &nbsp;
            <code class="no-fa" dir="ltr">{{min_amount}}</code> — حداقل خرید &nbsp;
            <code class="no-fa" dir="ltr">{{max_payout}}</code> — سقف
          </div>
        </div>

        <div class="flex gap-3">
          <button type="submit" class="flex-1 py-3 bg-green-600 hover:bg-green-700 text-white rounded-xl font-semibold">
            💾 ذخیره سطح</button>
          <a href="/admin/partners?tab=settings" class="flex-1 py-3 bg-gray-100 text-gray-600 rounded-xl font-semibold text-center">انصراف</a>
        </div>
      </form>
    </div>

    <script>
    function toggleCommission() {{
      const t = document.querySelector('input[name="commission_type"]:checked').value;
      document.getElementById('pct_wrap').style.display = (t === 'percent') ? 'grid' : 'none';
      document.getElementById('fixed_wrap').style.display = (t === 'fixed') ? 'grid' : 'none';
    }}
    toggleCommission();
    </script>"""
    return _layout(title, body, adm, flash=flash)


@router.post("/partners/tier/{tid}/delete")
async def partner_tier_delete(request: Request, tid: int):
    adm = _get_admin(request)
    guard = _require(adm, "partners")
    if guard: return guard
    from db import delete_partner_tier
    delete_partner_tier(tid)
    _log(request, "حذف سطح همکاری", "همکاران", f"tier:{tid}")
    return _redir("/admin/partners?tab=settings&flash=سطح+حذف+شد")


@router.post("/partners/commission")
async def partner_commission_save(request: Request):
    adm = _get_admin(request)
    guard = _require(adm, "partners")
    if guard: return guard
    form = await request.form()
    from db import save_partner_commission
    save_partner_commission(
        float(form.get("percent") or 5.0),
        int(form.get("min_order") or 0),
        int(form.get("max_payout") or 0),
        int(form.get("is_active") or 1)
    )
    _log(request, "تنظیم پورسانت همکاری", "همکاران", f"{form.get('percent')}%")
    return _redir("/admin/partners?tab=settings&flash=تنظیمات+ذخیره+شد")


@router.post("/partners/{uid}/approve")
async def partner_approve(request: Request, uid: int):
    adm = _get_admin(request)
    guard = _require(adm, "partners")
    if guard: return guard
    # از db.approve_partner() استفاده می‌کنه (نه UPDATE خام قبلی) — تا پاداش عضویت
    # معرفی‌های قبلی این کاربر (وقتی هنوز همکار نبود) هم خودکار حساب بشه (بخش ۳۹
    # CLAUDE.md). قبلاً این مسیر جدا از bot.py:approve_partner بود و این کچ‌آپ رو
    # نداشت — الان هر دو مسیر تأیید همکار (بات/پنل) دقیقاً یک منطق دارن.
    from db import approve_partner as _approve_partner
    _approve_partner(uid)
    try:
        import json as _json, requests as _rq
        token = _env("BOT_TOKEN","")
        if token:
            # منوی واقعی و کامل همکار — از همان سازنده ربات (نه هاردکد)
            markup = None
            try:
                from keyboards import main_menu as _mm
                markup = _mm(user_id=int(uid)).to_json()
            except Exception:
                markup = None
            payload = {
                "chat_id": int(uid),
                "text": "✅ <b>درخواست نمایندگی شما تایید شد!</b>\n\n"
                        "از این پس قیمت‌های ویژه همکار برای شما فعال است.\n"
                        "منوی کامل همکاری برای شما به‌روزرسانی شد 🤝",
                "parse_mode": "HTML",
            }
            if markup:
                payload["reply_markup"] = markup
            await run_in_threadpool(_rq.post, f"https://api.telegram.org/bot{token}/sendMessage", json=payload, timeout=8)
        else:
            await run_in_threadpool(_tg_send, int(uid), "✅ درخواست نمایندگی تایید شد! برای فعال‌سازی منو /start بزنید.")
    except Exception:
        pass
    return _redir("/admin/partners?flash=همکار+تایید+شد")


@router.post("/partners/{uid}/reject")
async def partner_reject(request: Request, uid: int):
    adm = _get_admin(request)
    guard = _require(adm, "partners")
    if guard: return guard
    conn = _db()
    try:
        conn.execute("UPDATE partners SET status='rejected' WHERE tg_user_id=?;", (uid,))
        conn.commit()
    finally:
        conn.close()
    try:
        await run_in_threadpool(_tg_send, int(uid),
            "❌ متأسفانه درخواست نمایندگی شما در این مرحله تأیید نشد.\n"
            "در صورت سوال با پشتیبانی در تماس باشید.")
    except Exception:
        pass
    return _redir("/admin/partners?flash=درخواست+رد+شد")

# ══════════════════════════════════════════════════════════════════════════════
# ─── 🚀 صفحه رشد و فروش ─────────────────────────────────────────────────────
# ══════════════════════════════════════════════════════════════════════════════

@router.get("/growth", response_class=HTMLResponse)
async def growth_page(request: Request, flash: str = ""):
    adm = _get_admin(request)
    guard = _require(adm, "growth")
    if guard: return guard

    from db import (ensure_growth_schema, list_flash_sales, get_winback_settings,
                    get_leaderboard_settings, get_social_settings, get_crypto_settings,
                    get_promo_settings, get_cfg, get_card2card_settings)
    ensure_growth_schema()
    wb, lb = get_winback_settings(), get_leaderboard_settings()
    soc, cr, pr = get_social_settings(), get_crypto_settings(), get_promo_settings()
    c2c = get_card2card_settings()
    webapp_url = get_cfg("webapp_url", "")

    conn = _db()
    try:
        products = conn.execute(
            "SELECT id, title FROM products WHERE COALESCE(is_active,1)=1 ORDER BY id DESC LIMIT 200;"
        ).fetchall()
    finally:
        conn.close()
    prod_opts = "".join(f'<option value="{p["id"]}">{e(p["title"])}</option>' for p in products)

    sales = list_flash_sales(20)
    sale_rows = "".join(f"""<tr class="border-b text-sm hover:bg-gray-50">
        <td class="px-3 py-2">{e(s['title'])}</td>
        <td class="px-3 py-2 font-bold text-red-600">{s['percent']}٪</td>
        <td class="px-3 py-2 text-xs text-gray-400">{fa_date(s['ends_at'], with_time=True)}</td>
        <td class="px-3 py-2">{'<span class="px-2 py-0.5 text-xs bg-green-100 text-green-700 rounded-full">🔴 زنده</span>' if s['live'] else '<span class="px-2 py-0.5 text-xs bg-gray-100 text-gray-400 rounded-full">پایان‌یافته</span>'}</td>
        <td class="px-3 py-2 whitespace-nowrap">
          {f'<form method="post" action="/admin/growth/flash/{s["id"]}/off" class="inline"><button class="text-xs text-amber-600 hover:underline ml-2">⏸ توقف</button></form>' if s['live'] else ''}
          <button type="button" class="text-xs text-indigo-500 hover:underline ml-2"
            onclick="flashEdit({s['product_id']},{s['percent']})">✏️ ویرایش</button>
          <form method="post" action="/admin/growth/flash/{s['id']}/delete" class="inline"
            onsubmit="return confirm('این فروش فوری حذف شود؟')">
            <button class="text-xs text-red-500 hover:underline">🗑 حذف</button></form>
        </td>
      </tr>""" for s in sales)

    def _chk(v): return "checked" if int(v or 0) else ""
    weekdays = ["دوشنبه","سه‌شنبه","چهارشنبه","پنجشنبه","جمعه","شنبه","یکشنبه"]
    wd_opts = "".join(f'<option value="{i}" {"selected" if int(lb.get("weekday") or 4)==i else ""}>{w}</option>'
                      for i, w in enumerate(weekdays))

    body = f"""
    <h1 class="text-2xl font-bold text-gray-800 mb-6">🚀 رشد و فروش</h1>

    <!-- ۳) فروش فوری -->
    <div class="card p-5 mb-5">
      <h2 class="font-bold text-gray-700 mb-1">🔥 فروش فوری (Flash Sale)</h2>
      <p class="text-xs text-gray-400 mb-4">تخفیف زمان‌دار با شمارش معکوس و بج «فقط N عدد مانده» روی صفحه محصول.</p>
      <form method="post" action="/admin/growth/flash/new" class="grid grid-cols-1 md:grid-cols-4 gap-3 mb-4">
        <select name="product_id" required class="border border-gray-300 rounded-lg px-3 py-2 text-sm md:col-span-2">
          <option value="">— انتخاب محصول —</option>{prod_opts}
        </select>
        <input type="number" name="percent" min="1" max="90" required placeholder="درصد تخفیف"
          class="border border-gray-300 rounded-lg px-3 py-2 text-sm">
        <div class="flex gap-2">
          <input type="number" name="hours" min="1" max="720" value="24" required placeholder="مدت (ساعت)"
            class="border border-gray-300 rounded-lg px-3 py-2 text-sm flex-1">
          <button class="px-4 py-2 bg-red-600 hover:bg-red-700 text-white rounded-lg text-sm font-semibold">🔥 شروع</button>
        </div>
      </form>
      <div class="overflow-x-auto"><table class="w-full text-right min-w-max">
        <thead><tr class="text-xs text-gray-500 border-b bg-gray-50">
          <th class="px-3 py-2">محصول</th><th class="px-3 py-2">تخفیف</th>
          <th class="px-3 py-2">پایان</th><th class="px-3 py-2">وضعیت</th><th class="px-3 py-2"></th></tr></thead>
        <tbody>{sale_rows or '<tr><td colspan="5" class="text-center py-6 text-gray-400">فروش فوری‌ای ثبت نشده</td></tr>'}</tbody>
      </table></div>
    </div>

    <form method="post" action="/admin/growth/save">

    <!-- ۱) بازگردانی -->
    <div class="card p-5 mb-5">
      <div class="flex items-center justify-between mb-1">
        <h2 class="font-bold text-gray-700">🎯 کمپین بازگردانی خودکار</h2>
        <label class="flex items-center gap-2 text-sm"><input type="checkbox" name="wb_enabled" {_chk(wb.get('enabled'))}> فعال</label>
      </div>
      <p class="text-xs text-gray-400 mb-4">به کاربرانی که مدتی خرید نکرده‌اند، خودکار کد تخفیف شخصی یک‌بارمصرف ارسال می‌شود.</p>
      <div class="grid grid-cols-2 md:grid-cols-5 gap-3 mb-3">
        <div><label class="text-xs text-gray-500 block mb-1">روزهای عدم خرید</label>
          <input type="number" name="wb_days" value="{int(wb.get('days_inactive') or 14)}" min="3" class="w-full border border-gray-300 rounded-lg px-3 py-2 text-sm"></div>
        <div><label class="text-xs text-gray-500 block mb-1">درصد تخفیف</label>
          <input type="number" name="wb_percent" value="{int(wb.get('percent') or 15)}" min="1" max="90" class="w-full border border-gray-300 rounded-lg px-3 py-2 text-sm"></div>
        <div><label class="text-xs text-gray-500 block mb-1">اعتبار کد (روز)</label>
          <input type="number" name="wb_expire" value="{int(wb.get('expire_days') or 3)}" min="1" class="w-full border border-gray-300 rounded-lg px-3 py-2 text-sm"></div>
        <div><label class="text-xs text-gray-500 block mb-1">فاصله تکرار (روز)</label>
          <input type="number" name="wb_cooldown" value="{int(wb.get('cooldown_days') or 30)}" min="7" class="w-full border border-gray-300 rounded-lg px-3 py-2 text-sm"></div>
        <div><label class="text-xs text-gray-500 block mb-1">ساعت ارسال</label>
          <input type="number" name="wb_hour" value="{int(wb.get('hour') or 11)}" min="0" max="23" class="w-full border border-gray-300 rounded-lg px-3 py-2 text-sm"></div>
      </div>
      <label class="text-xs text-gray-500 block mb-1">متن پیام — متغیرها: {{name}} {{code}} {{percent}} {{days}}</label>
      <textarea name="wb_message" rows="4" class="w-full border border-gray-300 rounded-lg px-3 py-2 text-sm" dir="rtl">{e(wb.get('message',''))}</textarea>
      <button type="submit" formaction="/admin/growth/run/winback"
        class="mt-2 text-xs text-indigo-500 hover:underline">▶ اجرای دستی همین حالا (حداکثر ۱۰ دقیقه دیگر)</button>
    </div>

    <!-- ۴) لیدربرد -->
    <div class="card p-5 mb-5">
      <div class="flex items-center justify-between mb-1">
        <h2 class="font-bold text-gray-700">🏆 لیدربرد هفتگی همکاران</h2>
        <label class="flex items-center gap-2 text-sm"><input type="checkbox" name="lb_enabled" {_chk(lb.get('enabled'))}> فعال</label>
      </div>
      <p class="text-xs text-gray-400 mb-4">هر هفته برترین همکاران بر اساس تعداد فروش، جایزه خودکار به کیف‌پول همکاری دریافت می‌کنند.</p>
      <div class="grid grid-cols-1 md:grid-cols-3 gap-3 mb-3">
        <div><label class="text-xs text-gray-500 block mb-1">روز اعلام نتایج</label>
          <select name="lb_weekday" class="w-full border border-gray-300 rounded-lg px-3 py-2 text-sm">{wd_opts}</select></div>
        <div class="md:col-span-2"><label class="text-xs text-gray-500 block mb-1">جوایز رتبه‌ها (تومان، با کاما — رتبه ۱ اول)</label>
          <input type="text" name="lb_rewards" value="{e(lb.get('rewards',''))}" class="w-full border border-gray-300 rounded-lg px-3 py-2 text-sm" dir="ltr"></div>
      </div>
      <label class="text-xs text-gray-500 block mb-1">پیام برنده — متغیرها: {{rank}} {{count}} {{reward}}</label>
      <textarea name="lb_message" rows="4" class="w-full border border-gray-300 rounded-lg px-3 py-2 text-sm" dir="rtl">{e(lb.get('message',''))}</textarea>
      <button type="submit" formaction="/admin/growth/run/leaderboard"
        class="mt-2 text-xs text-indigo-500 hover:underline">▶ اجرای دستی همین حالا (حداکثر ۱۰ دقیقه دیگر)</button>
    </div>

    <!-- ۶) کانال و نظرات + ۲) Upsell -->
    <div class="card p-5 mb-5">
      <h2 class="font-bold text-gray-700 mb-4">⭐ اعتمادسازی و پیشنهاد هوشمند</h2>
      <div class="grid grid-cols-1 md:grid-cols-3 gap-4 mb-3">
        <label class="flex items-center gap-2 text-sm"><input type="checkbox" name="soc_rating" {_chk(soc.get('rating'))}> درخواست امتیاز بعد از تحویل</label>
        <label class="flex items-center gap-2 text-sm"><input type="checkbox" name="soc_upsell" {_chk(soc.get('upsell'))}> پیشنهاد خرید بعدی (Upsell)</label>
        <label class="flex items-center gap-2 text-sm"><input type="checkbox" name="soc_salepost" {_chk(soc.get('sale_post'))}> پست خودکار فروش در کانال</label>
      </div>
      <div class="grid grid-cols-1 md:grid-cols-2 gap-3">
        <div><label class="text-xs text-gray-500 block mb-1">آیدی کانال (مثلاً @mychannel یا -100xxxx)</label>
          <input type="text" name="soc_channel" value="{e(soc.get('channel_id',''))}" class="w-full border border-gray-300 rounded-lg px-3 py-2 text-sm" dir="ltr"></div>
        <div><label class="text-xs text-gray-500 block mb-1">متن پست کانال — متغیر: {{title}}</label>
          <input type="text" name="soc_saletext" value="{e(soc.get('sale_post_text',''))}" class="w-full border border-gray-300 rounded-lg px-3 py-2 text-sm" dir="rtl"></div>
      </div>
      <p class="text-xs text-amber-500 mt-2">⚠️ ربات باید ادمین کانال باشد تا بتواند پست بگذارد.</p>
    </div>

    <!-- کارت‌به‌کارت -->
    <div class="card p-5 mb-5">
      <h2 class="font-bold text-gray-700 mb-4">💳 کارت مقصد کارت‌به‌کارت</h2>
      <div class="grid grid-cols-1 md:grid-cols-2 gap-3">
        <div><label class="text-xs text-gray-500 block mb-1">شماره کارت (۱۶ رقم)</label>
          <input type="text" name="c2c_number" value="{e(c2c.get('card_number',''))}" maxlength="19"
            class="w-full border border-gray-300 rounded-lg px-3 py-2 text-sm" dir="ltr"></div>
        <div><label class="text-xs text-gray-500 block mb-1">به نام</label>
          <input type="text" name="c2c_name" value="{e(c2c.get('card_name',''))}"
            class="w-full border border-gray-300 rounded-lg px-3 py-2 text-sm" dir="rtl"></div>
      </div>
      <p class="text-xs text-gray-400 mt-2">این کارت هم در ربات (شارژ کارت‌به‌کارت) هم در مینی‌اپ نمایش داده می‌شود — یک منبع واحد.</p>
    </div>

    <!-- ۷) رمزارز -->
    <div class="card p-5 mb-5">
      <div class="flex items-center justify-between mb-4">
        <h2 class="font-bold text-gray-700">₿ پرداخت رمزارز</h2>
        <label class="flex items-center gap-2 text-sm"><input type="checkbox" name="cr_enabled" {_chk(cr.get('enabled'))}> فعال</label>
      </div>
      <div class="grid grid-cols-1 md:grid-cols-2 gap-3 mb-3">
        <div><label class="text-xs text-gray-500 block mb-1">آدرس USDT (TRC20)</label>
          <input type="text" name="cr_usdt" value="{e(cr.get('usdt_trc20',''))}" class="w-full border border-gray-300 rounded-lg px-3 py-2 text-sm" dir="ltr"></div>
        <div><label class="text-xs text-gray-500 block mb-1">آدرس TRX</label>
          <input type="text" name="cr_trx" value="{e(cr.get('trx',''))}" class="w-full border border-gray-300 rounded-lg px-3 py-2 text-sm" dir="ltr"></div>
      </div>
      <label class="text-xs text-gray-500 block mb-1">راهنمای نمایش به کاربر</label>
      <input type="text" name="cr_note" value="{e(cr.get('note',''))}" class="w-full border border-gray-300 rounded-lg px-3 py-2 text-sm" dir="rtl">
      <p class="text-xs text-gray-400 mt-2">کاربر TXID را می‌فرستد؛ رسید در همان بخش رسیدهای مالی با برچسب ₿ ظاهر می‌شود و با همان دستور تأیید، کیف‌پول شارژ می‌شود.</p>
    </div>

    <!-- ۵) کیت تبلیغ + ۸) وب‌اپ -->
    <div class="grid md:grid-cols-2 gap-5 mb-5">
      <div class="card p-5">
        <h2 class="font-bold text-gray-700 mb-3">📣 کیت تبلیغاتی همکار</h2>
        <label class="text-xs text-gray-500 block mb-1">متن آماده — کپشن دعوت (لینک به‌صورت خودکار و جداگانه همراه پیام فرستاده می‌شه، نیازی به نوشتنش توی متن نیست)</label>
        <textarea name="pr_text" rows="6" class="w-full border border-gray-300 rounded-lg px-3 py-2 text-sm" dir="rtl">{e(pr.get('text',''))}</textarea>
        <p class="text-xs text-gray-400 mt-2">این متن هم داخل ربات هم در اشتراک‌گذاری تلگرام/مینی‌اپ استفاده می‌شه — سطح اشتراک‌گذاری تلگرام متن ساده‌ست (بدون بولد/تگ HTML)، پس از تگ‌هایی مثل &lt;b&gt; استفاده نکنید. متغیر اختیاری {{link}} هم در دسترسه اگه صراحتاً بخواید لینک رو داخل خود متن هم بیارید.</p>
      </div>
      <div class="card p-5">
        <h2 class="font-bold text-gray-700 mb-3">🛍 فروشگاه Mini App</h2>
        <label class="text-xs text-gray-500 block mb-1">آدرس عمومی وب‌اپ (HTTPS)</label>
        <input type="text" name="webapp_url" value="{e(webapp_url)}" placeholder="https://your-domain.com/admin/shop"
          class="w-full border border-gray-300 rounded-lg px-3 py-2 text-sm" dir="ltr">
        <p class="text-xs text-gray-400 mt-2">این پروژه صفحه فروشگاه را روی مسیر <code>/admin/shop</code> سرو می‌کند. دامنه HTTPS خودتان + این مسیر را اینجا وارد کنید تا دکمه «🛍 فروشگاه آنلاین» در منوی ربات ظاهر شود.</p>
      </div>
    </div>

    <div class="pb-10">
      <button type="submit" class="px-8 py-3 bg-indigo-600 hover:bg-indigo-700 text-white rounded-xl text-sm font-bold transition">💾 ذخیره همه تنظیمات رشد</button>
    </div>
    </form>
    <script>
    function flashEdit(pid, pct){{
      var s=document.querySelector('select[name=product_id]');
      var p=document.querySelector('input[name=percent]');
      if(s) s.value=String(pid);
      if(p) p.value=pct;
      window.scrollTo({{top:0,behavior:'smooth'}});
      if(s) s.focus();
    }}
    </script>"""
    return _layout("رشد و فروش", body, adm, flash=flash)


@router.post("/growth/save")
async def growth_save(request: Request):
    adm = _get_admin(request)
    guard = _require(adm, "growth")
    if guard: return guard
    form = await request.form()
    from db import set_cfg
    import json as _j

    def g(k, d=""): return str(form.get(k, d) or d).strip()
    def onoff(k): return 1 if form.get(k) is not None else 0

    set_cfg("winback", _j.dumps({
        "enabled": onoff("wb_enabled"), "days_inactive": int(g("wb_days", "14") or 14),
        "percent": int(g("wb_percent", "15") or 15), "expire_days": int(g("wb_expire", "3") or 3),
        "cooldown_days": int(g("wb_cooldown", "30") or 30), "hour": int(g("wb_hour", "11") or 11),
        "batch": 30, "message": g("wb_message"),
    }, ensure_ascii=False))
    set_cfg("leaderboard", _j.dumps({
        "enabled": onoff("lb_enabled"), "weekday": int(g("lb_weekday", "4") or 4),
        "rewards": g("lb_rewards"), "message": g("lb_message"),
    }, ensure_ascii=False))
    set_cfg("social", _j.dumps({
        "channel_id": g("soc_channel"), "sale_post": onoff("soc_salepost"),
        "rating": onoff("soc_rating"), "upsell": onoff("soc_upsell"),
        "sale_post_text": g("soc_saletext"),
    }, ensure_ascii=False))
    set_cfg("crypto", _j.dumps({
        "enabled": onoff("cr_enabled"), "usdt_trc20": g("cr_usdt"),
        "trx": g("cr_trx"), "note": g("cr_note"),
    }, ensure_ascii=False))
    set_cfg("card2card", _j.dumps({
        "card_number": g("c2c_number"), "card_name": g("c2c_name"),
    }, ensure_ascii=False))
    set_cfg("promo", _j.dumps({"text": g("pr_text")}, ensure_ascii=False))
    set_cfg("webapp_url", g("webapp_url"))
    _log(request, "ذخیره تنظیمات رشد", "رشد", "growth settings", admin_info=adm)
    return _redir("/admin/growth?flash=✅+تنظیمات+رشد+ذخیره+شد")


def _notify_flash_sale_favoriters(pid: int, percent: int) -> None:
    """اطلاع‌رسانی تخفیف فوری به علاقه‌مندی‌کننده‌های محصول — به‌عنوان BackgroundTask
    صدا زده می‌شه (همون دلیل _notify_restock_subscribers: حلقهٔ تماس‌های sync با
    تلگرام نباید جواب‌دادن به خودِ ادمین یا کل سرویس رو معطل کنه)."""
    try:
        from db import get_product_favoriters, get_product_by_id as _gpbi3, add_notification
        favoriters = get_product_favoriters(pid)
        if favoriters:
            _prod3 = _gpbi3(pid)
            _title3 = _prod3[2] if _prod3 else f"محصول #{pid}"
            bot_token3 = _env("BOT_TOKEN")
            for fav_uid in favoriters:
                try:
                    _requests.post(
                        f"https://api.telegram.org/bot{bot_token3}/sendMessage",
                        json={"chat_id": fav_uid,
                              "text": f"⚡️ محصولی که به علاقه‌مندی‌هاتون اضافه کرده بودید {percent}٪ تخفیف خورد!\n<b>{_title3}</b>",
                              "parse_mode": "HTML"},
                        timeout=5
                    )
                except Exception:
                    pass
                try:
                    add_notification(fav_uid, "تخفیف ویژه", f"«{_title3}» که به علاقه‌مندی‌هاتون اضافه کرده بودید {percent}٪ تخفیف خورد.", icon="⚡️")
                except Exception:
                    pass
    except Exception:
        pass


@router.post("/growth/flash/new")
async def growth_flash_new(request: Request, background_tasks: BackgroundTasks):
    adm = _get_admin(request)
    guard = _require(adm, "growth")
    if guard: return guard
    form = await request.form()
    from db import create_flash_sale
    pid = int(form.get("product_id") or 0)
    percent = int(form.get("percent") or 10)
    if pid:
        create_flash_sale(pid, percent, int(form.get("hours") or 24))
        _log(request, "فروش فوری", "رشد", f"محصول #{pid}", admin_info=adm)
        background_tasks.add_task(_notify_flash_sale_favoriters, pid, percent)
    return _redir("/admin/growth?flash=🔥+فروش+فوری+شروع+شد")


@router.post("/growth/flash/{sid}/delete")
async def growth_flash_delete(request: Request, sid: int):
    adm = _get_admin(request)
    guard = _require(adm, "growth")
    if guard: return guard
    conn = _db()
    try:
        conn.execute("DELETE FROM flash_sales WHERE id=?;", (sid,))
        conn.commit()
    finally:
        conn.close()
    try:
        from db import flash_map_invalidate
        flash_map_invalidate()
    except Exception:
        pass
    _log(request, "حذف فروش فوری", "رشد", f"#{sid}", admin_info=adm)
    return _redir("/admin/growth?flash=🗑+فروش+فوری+حذف+شد")


@router.post("/growth/flash/{sid}/off")
async def growth_flash_off(request: Request, sid: int):
    adm = _get_admin(request)
    guard = _require(adm, "growth")
    if guard: return guard
    from db import deactivate_flash_sale
    deactivate_flash_sale(sid)
    return _redir("/admin/growth?flash=⏹+فروش+فوری+متوقف+شد")


@router.post("/growth/run/winback")
async def growth_run_winback(request: Request):
    adm = _get_admin(request)
    guard = _require(adm, "growth")
    if guard: return guard
    from db import set_cfg
    set_cfg("winback_force", "1")
    return _redir("/admin/growth?flash=▶+بازگردانی+تا+۱۰+دقیقه+دیگر+اجرا+می‌شود")


@router.post("/growth/run/leaderboard")
async def growth_run_leaderboard(request: Request):
    adm = _get_admin(request)
    guard = _require(adm, "growth")
    if guard: return guard
    from db import set_cfg
    set_cfg("leaderboard_force", "1")
    return _redir("/admin/growth?flash=▶+لیدربرد+تا+۱۰+دقیقه+دیگر+اجرا+می‌شود")


# ══════════════════════════════════════════════════════════════════════════════
# ─── 🛍 فروشگاه Mini App تلگرام (عمومی — احراز با initData) ─────────────────
# ══════════════════════════════════════════════════════════════════════════════

def _tg_validate_init_data(init_data: str):
    """اعتبارسنجی رسمی Telegram WebApp initData — برمی‌گرداند user_id یا None."""
    import hmac as _hmac, hashlib as _hl, urllib.parse as _up, json as _j, time as _t
    try:
        from config import BOT_TOKEN
        pairs = dict(_up.parse_qsl(init_data, keep_blank_values=True))
        their_hash = pairs.pop("hash", "")
        if not their_hash:
            return None
        dcs = "\n".join(f"{k}={v}" for k, v in sorted(pairs.items()))
        secret = _hmac.new(b"WebAppData", BOT_TOKEN.encode(), _hl.sha256).digest()
        calc = _hmac.new(secret, dcs.encode(), _hl.sha256).hexdigest()
        if not _hmac.compare_digest(calc, their_hash):
            return None
        if _t.time() - int(pairs.get("auth_date", "0")) > 86400:
            return None
        return int(_j.loads(pairs.get("user", "{}")).get("id"))
    except Exception:
        return None


def _tg_api_send(chat_id, text):
    """ارسال پیام تلگرام از سمت پنل (بدون نیاز به نمونه ربات)."""
    try:
        import requests as _rq
        from config import BOT_TOKEN
        _rq.post(f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage",
                 json={"chat_id": chat_id, "text": text, "parse_mode": "HTML"}, timeout=8)
    except Exception:
        pass


@router.get("/shop", response_class=HTMLResponse)
async def shop_webapp(request: Request):
    """صفحه Mini App — عمومی؛ احراز داخل خودش با initData انجام می‌شود."""
    html_page = """<!DOCTYPE html>
<html lang="fa" dir="rtl"><head>
<meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1,maximum-scale=1,user-scalable=no">
<title>فروشگاه</title>
<script src="https://telegram.org/js/telegram-web-app.js"></script>
<style>
:root{--bg:#F6F7FB;--card:#fff;--txt:#1F2937;--mut:#8A94A6;--pri:#6366F1;--bdr:#E7EBF2}
@media(prefers-color-scheme:dark){:root{--bg:#0E1621;--card:#17212B;--txt:#F5F5F5;--mut:#8A99AC;--bdr:#2B3A4C}}
*{box-sizing:border-box;font-family:Vazirmatn,Tahoma,sans-serif}
body{margin:0;background:var(--bg);color:var(--txt)}
.top{position:sticky;top:0;background:var(--card);border-bottom:1px solid var(--bdr);padding:12px 16px;display:flex;justify-content:space-between;align-items:center;z-index:5}
.bal{font-size:13px;color:var(--mut)}
.grid{padding:14px;display:grid;gap:12px}
.p{background:var(--card);border:1px solid var(--bdr);border-radius:16px;padding:14px}
.pt{font-weight:700;margin-bottom:4px}
.pd{font-size:12px;color:var(--mut);margin-bottom:10px;line-height:1.8}
.row{display:flex;justify-content:space-between;align-items:center}
.pr{font-weight:800;color:var(--pri)}
.old{color:var(--mut);text-decoration:line-through;font-size:12px;margin-inline-start:6px}
.fl{font-size:11px;color:#EF4444;font-weight:700}
.st{font-size:11px;color:var(--mut)}
.btn{background:var(--pri);color:#fff;border:none;border-radius:12px;padding:9px 18px;font-size:13px;font-weight:700;font-family:inherit;cursor:pointer}
.btn:disabled{opacity:.4}
.msg{text-align:center;color:var(--mut);padding:48px 20px;font-size:14px}
.stars{font-size:11px;color:#F59E0B}
</style></head><body>
<div class="top"><b>🛍 فروشگاه</b><span class="bal" id="bal"></span></div>
<div id="list" class="grid"><div class="msg">در حال بارگذاری…</div></div>
<script>
var tg=window.Telegram&&Telegram.WebApp; if(tg){tg.ready();tg.expand();}
var initData=tg?tg.initData:'';
function fa(s){var F=['۰','۱','۲','۳','۴','۵','۶','۷','۸','۹'];return String(s).replace(/[0-9]/g,function(d){return F[+d];});}
function money(n){return fa((n||0).toLocaleString('en-US'))+' تومان';}
function api(path,body){return fetch(path,{method:'POST',headers:{'Content-Type':'application/json'},
  body:JSON.stringify(Object.assign({init_data:initData},body||{}))}).then(function(r){return r.json();});}
function load(){
  api('/admin/shop/api/catalog').then(function(d){
    if(!d.ok){document.getElementById('list').innerHTML='<div class="msg">'+(d.error||'خطای دسترسی — از داخل تلگرام باز کنید')+'</div>';return;}
    document.getElementById('bal').textContent='💰 '+money(d.balance);
    var h='';
    d.products.forEach(function(p){
      h+='<div class="p"><div class="pt">'+p.title+(p.rating?' <span class="stars">⭐ '+fa(p.rating)+' ('+fa(p.rcount)+')</span>':'')+'</div>'
        +(p.flash?'<div class="fl">🔥 فروش فوری '+fa(p.flash)+'٪ — تا '+p.flash_left+' دیگر</div>':'')
        +'<div class="pd">'+(p.description||'')+'</div>'
        +'<div class="row"><div><span class="pr">'+money(p.price)+'</span>'
        +(p.flash?'<span class="old">'+money(p.base)+'</span>':'')
        +'<div class="st">'+(p.stock>0?('موجودی: '+fa(p.stock)):'ناموجود')+'</div></div>'
        +'<button class="btn" '+(p.stock>0?'':'disabled')+' onclick="buy('+p.id+',this)">خرید</button></div></div>';
    });
    document.getElementById('list').innerHTML=h||'<div class="msg">محصولی موجود نیست</div>';
  }).catch(function(){document.getElementById('list').innerHTML='<div class="msg">خطا در ارتباط</div>';});
}
function buy(pid,btn){
  btn.disabled=true;btn.textContent='…';
  api('/admin/shop/api/buy',{product_id:pid}).then(function(d){
    if(d.ok){ if(tg)tg.showAlert('✅ خرید انجام شد!\\nمحصول به چت ربات ارسال شد.'); load(); }
    else { if(tg)tg.showAlert('❌ '+(d.error||'خطا')); btn.disabled=false;btn.textContent='خرید'; }
  }).catch(function(){btn.disabled=false;btn.textContent='خرید';});
}
load();
</script></body></html>"""
    return HTMLResponse(html_page)


@router.post("/shop/api/catalog")
async def shop_api_catalog(request: Request):
    import json as _j
    try:
        body = _j.loads((await request.body()) or b"{}")
    except Exception:
        body = {}
    uid = _tg_validate_init_data(str(body.get("init_data") or ""))
    if not uid:
        return JSONResponse({"ok": False, "error": "احراز هویت تلگرام ناموفق"})
    from db import get_wallet_balance, apply_flash_price, get_product_rating, get_feed_stats
    conn = _db()
    try:
        prods = conn.execute("""
            SELECT id, title, price, COALESCE(description,'') d
            FROM products WHERE COALESCE(is_active,1)=1 ORDER BY id DESC LIMIT 60;
        """).fetchall()
    finally:
        conn.close()
    out = []
    for p in prods:
        base = int(p["price"] or 0)
        price, fl = apply_flash_price(int(p["id"]), base)
        try:
            _t, rem, _d = get_feed_stats(int(p["id"]))
        except Exception:
            rem = 0
        try:
            r = get_product_rating(int(p["id"]))
        except Exception:
            r = {"avg": 0, "count": 0}
        out.append({"id": p["id"], "title": p["title"], "description": p["d"][:140],
                    "base": base, "price": price, "stock": int(rem or 0),
                    "flash": (fl["percent"] if fl else 0),
                    "flash_left": (fl["left_str"] if fl else ""),
                    "rating": (r["avg"] if r["count"] else 0), "rcount": r["count"]})
    return JSONResponse({"ok": True, "balance": get_wallet_balance(uid), "products": out})


@router.post("/shop/api/buy")
async def shop_api_buy(request: Request):
    """خرید از وب‌اپ — فقط با موجودی کیف‌پول؛ تحویل به چت ربات ارسال می‌شود."""
    import json as _j
    try:
        body = _j.loads((await request.body()) or b"{}")
    except Exception:
        body = {}
    uid = _tg_validate_init_data(str(body.get("init_data") or ""))
    if not uid:
        return JSONResponse({"ok": False, "error": "احراز هویت ناموفق"})
    pid = int(body.get("product_id") or 0)

    from db import (get_wallet_balance, subtract_wallet_balance, apply_flash_price,
                    create_order, claim_next_feed_item,
                    process_referral_commission)
    conn = _db()
    try:
        p = conn.execute(
            "SELECT id,title,price,category_id FROM products WHERE id=? AND COALESCE(is_active,1)=1;",
            (pid,)).fetchone()
    finally:
        conn.close()
    if not p:
        return JSONResponse({"ok": False, "error": "محصول یافت نشد"})

    price, _fl = apply_flash_price(pid, int(p["price"] or 0))
    if get_wallet_balance(uid) < price:
        return JSONResponse({"ok": False, "error": "موجودی کیف‌پول کافی نیست — از ربات شارژ کنید"})

    # ترتیب امن: اول کسر، بعد claim؛ اگر موجودی نبود → بازگشت وجه
    if not subtract_wallet_balance(uid, price):
        return JSONResponse({"ok": False, "error": "خطا در برداشت از کیف‌پول"})

    item = claim_next_feed_item(pid)
    if not item:
        from db import add_wallet_balance
        add_wallet_balance(uid, price)
        return JSONResponse({"ok": False, "error": "موجودی محصول تمام شده — مبلغ بازگشت داده شد"})

    order_id = create_order(uid, "webapp", p["title"], price, product_id=pid, buyer_type="customer")

    # تحویل در چت ربات
    await run_in_threadpool(_tg_api_send, uid,
        f"✅ <b>خرید از فروشگاه آنلاین</b>\n\n"
        f"🧾 سفارش #{order_id} — {p['title']}\n"
        f"💰 مبلغ: {price:,} تومان\n\n"
        f"📦 اطلاعات محصول:\n<code>{item[1]}</code>")

    # هوک پورسانت (پاداش عضویت جدا و در لحظه /start پرداخت می‌شود)
    try:
        cm = process_referral_commission(uid, order_id, price)
        if cm.get("paid"):
            _wl = "کیف‌پول همکاری" if cm.get("wallet") == "partner" else "کیف‌پول"
            await run_in_threadpool(_tg_api_send, cm["referrer_id"],
                f"💸 <b>پورسانت جدید!</b>\nیکی از دعوت‌شده‌های شما خرید کرد و "
                f"<b>{cm['amount']:,}</b> تومان (سطح {cm['tier_name']}) به {_wl} شما اضافه شد.")
    except Exception:
        pass
    # پست کانال
    try:
        from db import get_social_settings
        soc = get_social_settings()
        ch = str(soc.get("channel_id") or "").strip()
        if ch and int(soc.get("sale_post") or 0):
            await run_in_threadpool(_tg_api_send, ch, str(soc.get("sale_post_text") or "").format(title=p["title"]))
    except Exception:
        pass

    return JSONResponse({"ok": True, "order_id": order_id})


@router.post("/partners/manual-referral")
async def partners_manual_referral(request: Request):
    """🔧 ثبت دستی معرفی + پاداش عضویت اختیاری + اطلاع‌رسانی به معرف."""
    adm = _get_admin(request)
    guard = _require(adm, "partners")
    if guard: return guard
    form = await request.form()
    try:
        referrer_id = int(form.get("referrer_id") or 0)
        referred_id = int(form.get("referred_id") or 0)
    except Exception:
        return _redir("/admin/partners?tab=referrals&flash=❌+آیدی+نامعتبر")
    if not referrer_id or not referred_id or referrer_id == referred_id:
        return _redir("/admin/partners?tab=referrals&flash=❌+آیدی+نامعتبر")

    from db import register_referral, pay_signup_referral_reward, ensure_referral_schema
    ensure_referral_schema()
    ok = register_referral(referrer_id, referred_id)
    if not ok:
        return _redir("/admin/partners?tab=referrals&flash=⛔+این+کاربر+قبلاً+معرف+دارد")

    paid_txt = ""
    if form.get("pay_reward") is not None:
        pr = pay_signup_referral_reward(referrer_id, referred_id)
        if pr.get("paid"):
            paid_txt = f"+و+{pr['amount']:,}+تومان+پاداش+پرداخت+شد"
            try:
                _wl = "کیف‌پول همکاری" if pr.get("wallet") == "partner" else "کیف‌پول"
                await run_in_threadpool(_tg_api_send, referrer_id,
                    f"🎉 یک دعوت‌شده جدید برای شما ثبت شد!\n"
                    f"💰 پاداش عضویت: <b>{pr['amount']:,}</b> تومان به {_wl} شما اضافه شد.")
            except Exception:
                pass
    _log(request, "ثبت دستی معرفی", "همکاران", f"{referrer_id} → {referred_id}", admin_info=adm)
    return _redir(f"/admin/partners?tab=referrals&flash=✅+معرفی+ثبت+شد{paid_txt}")


# ══════════════════════════════════════════════════════════════════════════════
# ─── ☁️ روت‌های بکاپ ابری ────────────────────────────────────────────────────
# ══════════════════════════════════════════════════════════════════════════════

@router.post("/database/gdrive/start")
async def gdrive_oauth_start(request: Request):
    """شروع OAuth Device Flow برای Google Drive."""
    from fastapi.responses import JSONResponse
    adm = _get_admin(request)
    if not adm:
        return JSONResponse({"ok": False, "error": "unauthorized"})
    client_id = os.getenv("GDRIVE_CLIENT_ID", "").strip()
    if not client_id:
        return JSONResponse({"ok": False, "error": "GDRIVE_CLIENT_ID در env تنظیم نشده"})
    try:
        from backup_uploader import gdrive_device_start
        res = gdrive_device_start(client_id)
        if not res.get("ok"):
            return JSONResponse(res)
        return JSONResponse({
            "ok": True,
            "user_code": res["user_code"],
            "url": res["verification_url"],
            "device_code": res["device_code"],
        })
    except Exception as ex:
        return JSONResponse({"ok": False, "error": str(ex)[:120]})


@router.post("/database/gdrive/poll")
async def gdrive_oauth_poll(request: Request):
    """چک وضعیت تأیید OAuth — ذخیره refresh_token."""
    from fastapi.responses import JSONResponse
    adm = _get_admin(request)
    if not adm:
        return JSONResponse({"ok": False, "error": "unauthorized"})
    try:
        body = await request.json()
    except Exception:
        body = {}
    dc = str(body.get("device_code", "")).strip()
    if not dc:
        return JSONResponse({"ok": False, "error": "device_code خالی"})
    client_id = os.getenv("GDRIVE_CLIENT_ID", "").strip()
    client_secret = os.getenv("GDRIVE_CLIENT_SECRET", "").strip()
    if not client_id or not client_secret:
        return JSONResponse({"ok": False, "error": "GDRIVE_CLIENT_ID/SECRET تنظیم نشده"})
    try:
        from backup_uploader import gdrive_device_poll
        res = gdrive_device_poll(client_id, client_secret, dc)
        if res.get("ok"):
            _log(request, "اتصال Google Drive", "دیتابیس", "OAuth refresh token ذخیره شد", admin_info=adm)
        return JSONResponse(res)
    except Exception as ex:
        return JSONResponse({"ok": False, "error": str(ex)[:120]})


@router.post("/database/cloud-save")
async def database_cloud_save(request: Request):
    adm = _get_admin(request)
    guard = _require(adm, "database")
    if guard: return guard
    form = await request.form()
    from backup_uploader import get_cloud_settings, save_cloud_settings

    cfg = get_cloud_settings()  # حفظ od_refresh و مقادیر قبلی
    def g(k, d=""): return str(form.get(k, d) or d).strip()
    def onoff(k): return 1 if form.get(k) is not None else 0

    cfg.update({
        "enabled":        onoff("enabled"),
        "hour":           max(0, min(23, int(g("hour", "4") or 4))),
        "retention":      max(1, min(30, int(g("retention", "3") or 3))),
        "tg_enabled":     onoff("tg_enabled"),
        "tg_channel":     g("tg_channel"),
        "gdrive_enabled": onoff("gdrive_enabled"),
    })
    save_cloud_settings(cfg)
    _log(request, "تنظیمات بکاپ ابری", "دیتابیس", "cloud settings saved", admin_info=adm)
    return _redir("/admin/database?flash=✅+تنظیمات+بکاپ+ابری+ذخیره+شد#cloudbk")


@router.post("/database/cloud-run")
async def database_cloud_run(request: Request):
    """▶ بکاپ + آپلود فوری — در پس‌زمینه تا صفحه معطل نماند."""
    adm = _get_admin(request)
    guard = _require(adm, "backup")
    if guard: return guard
    try:
        import threading as _th
        _th.Thread(target=_do_auto_backup, name="cloud-run-now", daemon=True).start()
    except Exception as ex:
        return _redir(f"/admin/database?flash=❌+خطا:+{ex}#cloudbk")
    _log(request, "بکاپ ابری دستی", "دیتابیس", "run-now", admin_info=adm)
    return _redir("/admin/database?flash=▶+بکاپ+و+آپلود+شروع+شد+—+نتیجه+چند+لحظه+دیگر+در+وضعیت+همین+کارت#cloudbk")


# ══════════════════════════════════════════════════════════════════════════════
# ─── 🔗 مدیریت اتصال ربات (Webhook / Polling) ───────────────────────────────
# ══════════════════════════════════════════════════════════════════════════════

@router.get("/webhook", response_class=HTMLResponse)
async def webhook_management(request: Request, flash: str = ""):
    """صفحه مدیریت حالت اتصال ربات — Webhook / Polling / Stopped."""
    adm = _get_admin(request)
    guard = _require(adm, "settings")
    if guard: return guard

    from config import BOT_TOKEN, WEBHOOK_BASE_URL
    # حالت جاری از bot_config
    try:
        from db import get_cfg
        current_mode = (get_cfg("bot_run_mode", "") or "").strip().lower() or "unknown"
    except Exception:
        current_mode = "unknown"

    # وضعیت واقعی از تلگرام
    info = {}
    try:
        import requests as _req
        r = _req.get(f"https://api.telegram.org/bot{BOT_TOKEN}/getWebhookInfo", timeout=10)
        j = r.json()
        info = j.get("result", {}) if j.get("ok") else {}
    except Exception as ex:
        info = {"error": str(ex)[:120]}

    tg_url     = info.get("url") or ""
    tg_pending = info.get("pending_update_count", 0)
    tg_err     = info.get("last_error_message", "") or ""

    def _badge(mode, label, color):
        active = (current_mode == mode)
        return (f'<span class="px-3 py-1 rounded-full text-xs font-semibold '
                f'{"bg-"+color+"-100 text-"+color+"-700" if active else "bg-gray-100 text-gray-400"}">'
                f'{"● " if active else ""}{label}</span>')

    def _btn_mode(mode, label, icon, color, desc):
        is_current = (current_mode == mode)
        disabled = 'disabled' if is_current else ''
        disabled_cls = ' disabled-visual--noclick' if is_current else ''
        return f"""
        <form method="post" action="/admin/webhook/switch" class="flex-1">
          <input type="hidden" name="mode" value="{mode}">
          <button {disabled} class="w-full p-4 bg-{color}-50 border-2 border-{color}-200 hover:bg-{color}-100 rounded-xl text-right{disabled_cls}">
            <div class="text-2xl mb-1">{icon}</div>
            <div class="font-bold text-{color}-800 text-sm">{label}</div>
            <div class="text-[10px] text-{color}-600 mt-1">{desc}</div>
            {'<div class="text-[10px] text-'+color+'-500 mt-1">✓ حالت فعلی</div>' if is_current else ''}
          </button>
        </form>"""

    expected_url = f"{WEBHOOK_BASE_URL}/telegram/webhook/{BOT_TOKEN}"

    body = f"""
    <div class="max-w-3xl mx-auto p-4">
      <div class="mb-4">
        <h1 class="text-xl font-bold text-gray-800">🔗 اتصال ربات</h1>
        <p class="text-xs text-gray-400 mt-1">نحوه دریافت پیام‌های تلگرام — سوییچ نرم بدون restart</p>
      </div>

      <div class="card p-5 mb-4">
        <div class="flex items-center gap-2 mb-3 flex-wrap">
          <span class="text-sm font-semibold text-gray-600">حالت فعلی:</span>
          {_badge("webhook", "Webhook", "green")}
          {_badge("polling", "Polling", "blue")}
          {_badge("stopped", "متوقف", "gray")}
        </div>
        <div class="grid grid-cols-1 md:grid-cols-3 gap-3">
          {_btn_mode("webhook", "Webhook", "🚀", "green", "سریع، توصیه‌شده — تلگرام مستقیم پیام می‌فرستد")}
          {_btn_mode("polling", "Polling", "🔄", "blue", "پایدارتر — ربات هر چند ثانیه از تلگرام می‌پرسد")}
          {_btn_mode("stopped", "متوقف", "⏸", "gray", "ربات خاموش — برای تعمیر یا مواقع خاص")}
        </div>
      </div>

      <div class="card p-4 mb-4">
        <h2 class="font-bold text-gray-700 text-sm mb-3">📊 وضعیت واقعی از تلگرام</h2>
        <div class="grid grid-cols-1 md:grid-cols-3 gap-3 text-xs">
          <div class="p-3 bg-gray-50 rounded">
            <div class="text-gray-500 mb-1">Webhook ثبت‌شده:</div>
            <div class="font-mono break-all" dir="ltr">{e(tg_url) if tg_url else "— (بدون webhook)"}</div>
          </div>
          <div class="p-3 bg-gray-50 rounded">
            <div class="text-gray-500 mb-1">آپدیت‌های پندینگ:</div>
            <div class="font-bold {'text-red-600' if tg_pending>0 else 'text-green-600'}">{tg_pending}</div>
          </div>
          <div class="p-3 bg-gray-50 rounded">
            <div class="text-gray-500 mb-1">آخرین خطا:</div>
            <div class="text-red-600">{e(tg_err[:60]) if tg_err else "—"}</div>
          </div>
        </div>
      </div>

      <div class="card p-4">
        <h2 class="font-bold text-gray-700 text-sm mb-2">ℹ️ راهنما</h2>
        <ul class="text-xs text-gray-600 space-y-1 leading-6 list-disc pr-4">
          <li><b>Webhook</b> — تلگرام مستقیم به سرور شما پیام می‌فرستد. سریع‌ترین حالت. نیاز به دامنه با HTTPS دارد.</li>
          <li><b>Polling</b> — ربات هر چند ثانیه از تلگرام می‌پرسد. کندتر ولی روی هر سروری کار می‌کند.</li>
          <li><b>متوقف</b> — ربات هیچ پیامی دریافت نمی‌کند. مناسب مواقع تعمیر یا مهاجرت.</li>
          <li>تغییر حالت <b>بدون restart سرویس</b> انجام می‌شود و بلافاصله اعمال می‌گردد.</li>
        </ul>
        <div class="mt-3 p-2 bg-gray-50 rounded text-[10px] text-gray-500 font-mono" dir="ltr">
          آدرس Webhook: {e(expected_url)}
        </div>
      </div>
    </div>"""
    return _layout("اتصال ربات", body, adm, flash=flash)


@router.post("/webhook/switch")
async def webhook_switch(request: Request):
    """سوییچ نرم بین Webhook / Polling / Stopped بدون restart."""
    adm = _get_admin(request)
    guard = _require(adm, "settings")
    if guard: return guard
    form = await request.form()
    mode = (form.get("mode") or "").strip().lower()
    if mode not in ("webhook", "polling", "stopped"):
        return _redir("/admin/webhook?flash=❌+حالت+نامعتبر")
    try:
        from payment_service import switch_bot_mode
        ok, msg = switch_bot_mode(mode)
    except Exception as ex:
        ok, msg = False, f"خطا: {str(ex)[:80]}"
    _log(request, f"سوییچ حالت ربات → {mode}", "تنظیمات", msg, admin_info=adm)
    flag = "✅" if ok else "❌"
    return _redir(f"/admin/webhook?flash={flag}+{e(msg)}")


# نگه‌داشتن روت‌های قدیمی برای سازگاری — صرفاً redirect به نسخه جدید
@router.post("/webhook/set")
async def webhook_set(request: Request):
    return await webhook_switch_impl(request, "webhook")

@router.post("/webhook/remove")
async def webhook_remove(request: Request):
    return await webhook_switch_impl(request, "polling")

async def webhook_switch_impl(request: Request, mode: str):
    adm = _get_admin(request)
    guard = _require(adm, "settings")
    if guard: return guard
    try:
        from payment_service import switch_bot_mode
        ok, msg = switch_bot_mode(mode)
    except Exception as ex:
        ok, msg = False, f"خطا: {str(ex)[:80]}"
    _log(request, f"سوییچ حالت (سازگاری) → {mode}", "تنظیمات", msg, admin_info=adm)
    return _redir(f"/admin/webhook?flash={'✅' if ok else '❌'}+{e(msg)}")


# ══════════════════════════════════════════════════════════════════════════
# ─── محتوای اپ PWA (آموزش / اخبار / امکانات) ──────────────────────────────
# ══════════════════════════════════════════════════════════════════════════
APP_MEDIA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "app_media")

_KIND_LABELS = {"tutorial": "📚 آموزش", "feature": "✨ امکانات", "daily": "📋 لیست روزانه"}


@router.get("/app-content", response_class=HTMLResponse)
async def app_content_page(request: Request, kind: str = "", flash: str = ""):
    adm = _get_admin(request)
    guard = _require(adm, "news")
    if guard: return guard
    from db import get_app_content
    k = kind if kind in _KIND_LABELS else None
    items = get_app_content(kind=k, active_only=False, limit=100)

    tabs = ''.join(
        f'<a href="/admin/app-content?kind={s}" class="px-3 py-1.5 rounded-lg text-xs border '
        f'{"bg-indigo-600 text-white" if kind==s else "bg-white text-gray-500"}">{l}</a>'
        for s, l in [("", "همه"), ("tutorial", "📚 آموزش"), ("feature", "✨ امکانات")]
    )

    rows = ""
    for it in items:
        badge = _KIND_LABELS.get(it.get("kind"), it.get("kind"))
        active = '<span class="text-green-600 text-xs">فعال</span>' if int(it.get("is_active") or 0) \
                 else '<span class="text-gray-400 text-xs">غیرفعال</span>'
        img = f'<img src="{html.escape(it.get("image_url") or "")}" class="w-10 h-10 rounded-lg object-cover">' \
              if it.get("image_url") else '<div class="w-10 h-10 rounded-lg bg-gray-100"></div>'
        rows += f"""
        <tr class="border-b">
          <td class="p-2">{img}</td>
          <td class="p-2 text-sm font-medium">{html.escape(it.get('title') or '')}</td>
          <td class="p-2 text-xs">{badge}</td>
          <td class="p-2">{active}</td>
          <td class="p-2 text-xs text-gray-400">{html.escape(str(it.get('created_at') or '')[:16])}</td>
          <td class="p-2 whitespace-nowrap">
            <a href="/admin/app-content/{it['id']}/edit" class="text-indigo-600 text-xs ml-2">✏️ ویرایش</a>
            <form method="post" action="/admin/app-content/{it['id']}/delete" class="inline"
                  onsubmit="return confirm('حذف شود؟')">
              <button class="text-red-500 text-xs">🗑 حذف</button>
            </form>
          </td>
        </tr>"""
    if not rows:
        rows = '<tr><td colspan="6" class="p-6 text-center text-gray-400 text-sm">هنوز محتوایی ثبت نشده.</td></tr>'

    body = f"""
    <div class="flex items-center justify-between mb-4">
      <div class="flex gap-2">{tabs}</div>
      <a href="/admin/app-content/new" class="bg-indigo-600 text-white px-4 py-2 rounded-lg text-sm">＋ افزودن محتوا</a>
    </div>
    <div class="bg-white rounded-xl shadow-sm overflow-x-auto">
      <table class="w-full text-right">
        <thead><tr class="text-xs text-gray-400 border-b">
          <th class="p-2">تصویر</th><th class="p-2">عنوان</th><th class="p-2">نوع</th>
          <th class="p-2">وضعیت</th><th class="p-2">تاریخ</th><th class="p-2">عملیات</th>
        </tr></thead>
        <tbody>{rows}</tbody>
      </table>
    </div>"""
    return _layout("محتوای اپ", body, adm, flash=flash)


def _app_content_form(it=None):
    it = it or {}
    kind_opts = ''.join(
        f'<option value="{k}" {"selected" if it.get("kind")==k else ""}>{l}</option>'
        for k, l in _KIND_LABELS.items()
    )
    checked = "checked" if (it.get("is_active", 1) in (1, "1", True, None) and int(it.get("is_active", 1) or 0) == 1) or not it else ""
    return f"""
    <form method="post" action="/admin/app-content/save" enctype="multipart/form-data"
          class="bg-white rounded-xl shadow-sm p-5 max-w-2xl space-y-4">
      <input type="hidden" name="cid" value="{it.get('id','')}">
      <div>
        <label class="block text-xs text-gray-500 mb-1">نوع محتوا</label>
        <select name="kind" class="w-full border rounded-lg p-2 text-sm">{kind_opts}</select>
        <div class="text-xs text-amber-600 bg-amber-50 border border-amber-200 rounded-lg p-2 mt-2">
          ⚠️ «📚 آموزش» اینجا سیستم قدیمیه و دیگه هیچ‌جای مینی‌اپ نشون داده نمی‌شه — برای آموزش واقعی از
          <a href="/admin/tutorials" class="underline font-medium">صفحهٔ «آموزش»</a> زیر «مدیریت اپ» استفاده کن.
          «✨ امکانات» هم فقط با لینک مستقیم باز می‌شه، جایی توی اپ لیست/مرور نمی‌شه.
        </div>
      </div>
      <div>
        <label class="block text-xs text-gray-500 mb-1">عنوان</label>
        <input name="title" required value="{html.escape(str(it.get('title') or ''))}"
               class="w-full border rounded-lg p-2 text-sm">
      </div>
      <div>
        <label class="block text-xs text-gray-500 mb-1">متن</label>
        <textarea name="body" rows="8" class="w-full border rounded-lg p-2 text-sm">{html.escape(str(it.get('body') or ''))}</textarea>
      </div>
      <div>
        <label class="block text-xs text-gray-500 mb-1">تصویر (اختیاری — آپلود)</label>
        <input type="file" name="image" accept="image/*" class="w-full text-sm">
      </div>
      <div>
        <label class="block text-xs text-gray-500 mb-1">یا آدرس تصویر</label>
        <input name="image_url" value="{html.escape(str(it.get('image_url') or ''))}"
               class="w-full border rounded-lg p-2 text-sm" placeholder="https://…">
      </div>
      <div>
        <label class="block text-xs text-gray-500 mb-1">لینک بیرونی (اختیاری — تلگرام/اینستاگرام)</label>
        <input name="link_url" value="{html.escape(str(it.get('link_url') or ''))}" dir="ltr"
               class="w-full border rounded-lg p-2 text-sm" placeholder="https://t.me/... یا https://instagram.com/...">
        <div class="text-xs text-gray-400 mt-1">اگر پر بشه، پایین متن پست در اپ یک دکمهٔ «مشاهده در تلگرام/اینستاگرام» اضافه می‌شود.</div>
      </div>
      <label class="flex items-center gap-2 text-sm">
        <input type="checkbox" name="is_active" value="1" {checked}> فعال (نمایش در اپ)
      </label>
      <div class="flex gap-2">
        <button class="bg-indigo-600 text-white px-5 py-2 rounded-lg text-sm">ذخیره</button>
        <a href="/admin/app-content" class="px-5 py-2 rounded-lg text-sm border">انصراف</a>
      </div>
    </form>"""


@router.get("/app-content/new", response_class=HTMLResponse)
async def app_content_new(request: Request):
    adm = _get_admin(request)
    guard = _require(adm, "news")
    if guard: return guard
    return _layout("افزودن محتوا", _app_content_form(), adm)


@router.get("/app-content/{cid}/edit", response_class=HTMLResponse)
async def app_content_edit(request: Request, cid: int):
    adm = _get_admin(request)
    guard = _require(adm, "news")
    if guard: return guard
    from db import get_app_content_item
    it = get_app_content_item(cid)
    if not it:
        return _redir("/admin/app-content")
    return _layout("ویرایش محتوا", _app_content_form(it), adm)


@router.post("/app-content/save")
async def app_content_save(request: Request, cid: str = Form(""), kind: str = Form("news"),
                           title: str = Form(...), body: str = Form(""),
                           image_url: str = Form(""), link_url: str = Form(""), is_active: str = Form(""),
                           image: UploadFile = None):
    adm = _get_admin(request)
    guard = _require(adm, "news")
    if guard: return guard

    # آپلود تصویر (در صورت انتخاب فایل)
    final_image = (image_url or "").strip()
    try:
        if image is not None and getattr(image, "filename", ""):
            os.makedirs(APP_MEDIA_DIR, exist_ok=True)
            ext = os.path.splitext(image.filename)[1].lower() or ".jpg"
            if ext not in (".jpg", ".jpeg", ".png", ".webp", ".gif"):
                ext = ".jpg"
            fname = f"c{int(time.time())}{ext}"
            data = await image.read()
            if data:
                from image_utils import compress_image_bytes
                with open(os.path.join(APP_MEDIA_DIR, fname), "wb") as f:
                    f.write(compress_image_bytes(data, ext))
                final_image = f"/app-media/{fname}"
    except Exception:
        pass

    kind = kind if kind in _KIND_LABELS else "news"
    from db import add_app_content, update_app_content
    if cid.strip().isdigit():
        update_app_content(int(cid), kind, title.strip(), body, final_image,
                           1 if is_active == "1" else 0, link_url.strip())
    else:
        add_app_content(kind, title.strip(), body, final_image, link_url.strip())
    return _redir("/admin/app-content")


@router.get("/news-feed", response_class=HTMLResponse)
async def news_feed_page(request: Request, flash: str = ""):
    adm = _get_admin(request)
    guard = _require(adm, "news")
    if guard: return guard
    from db import get_cfg
    from core.news import get_feed
    url = get_cfg("NEWS_FEED_URL", "")
    d = get_feed(force=False)  # از کش می‌خونه، تا پنل رو کند نکنه
    status = d.get("status")
    status_badge = {
        "ok": '<span class="text-green-600 text-xs font-medium">🟢 متصل</span>',
        "error": '<span class="text-red-600 text-xs font-medium">🔴 خطا در آخرین تلاش</span>',
        "unset": '<span class="text-gray-400 text-xs font-medium">⚪️ هنوز تنظیم نشده</span>',
    }.get(status, status or "—")
    last_sync = d.get("last_sync")
    last_sync_str = datetime.fromtimestamp(int(last_sync)).strftime("%Y-%m-%d %H:%M") if last_sync else "—"
    items_count = len(d.get("items") or [])
    error_box = (f'<div class="text-xs text-red-500 mt-2">خطا: {html.escape(d.get("error") or "")}</div>'
                 if status == "error" else "")

    body = f"""
    <div class="bg-white rounded-xl shadow-sm p-5 max-w-2xl mb-4">
      <div class="text-sm font-medium mb-2">دامنهٔ وبلاگ (وردپرس)</div>
      <form method="post" action="/admin/news-feed/settings" class="flex gap-2 mb-3">
        <input name="url" value="{html.escape(url)}" dir="ltr"
               class="flex-1 border rounded-lg p-2 text-sm" placeholder="stland.ir">
        <button class="bg-indigo-600 text-white px-4 py-2 rounded-lg text-sm">ذخیره</button>
      </form>
      <div class="text-xs text-gray-400 -mt-2 mb-3">فقط دامنهٔ سایت رو وارد کنید (مثلاً stland.ir) — آدرس دقیق صفحهٔ آرشیو لازم نیست، از REST API رسمی وردپرس (wp-json) به‌صورت خودکار پست‌ها خونده می‌شه.</div>
      <div class="grid grid-cols-2 gap-3 text-sm">
        <div><span class="text-gray-400">وضعیت اتصال: </span>{status_badge}</div>
        <div><span class="text-gray-400">آخرین همگام‌سازی: </span>{last_sync_str}</div>
      </div>
      {error_box}
      <form method="post" action="/admin/news-feed/refresh" class="mt-3">
        <button class="border px-4 py-2 rounded-lg text-sm">🔄 بروزرسانی دستی</button>
      </form>
      <div class="text-xs text-gray-400 mt-2">تب «اخبار تکنولوژی» اپ همهٔ {items_count} مقالهٔ فچ‌شده رو مستقیم از همین فید نشون می‌ده — نیازی به ثبت دستی نیست.</div>
    </div>"""
    return _layout("اخبار تکنولوژی", body, adm, flash=flash)


@router.post("/news-feed/settings")
async def news_feed_settings_save(request: Request, url: str = Form("")):
    adm = _get_admin(request)
    guard = _require(adm, "news")
    if guard: return guard
    from db import set_cfg
    set_cfg("NEWS_FEED_URL", url.strip())
    _log(request, "تنظیم آدرس فید اخبار", "اخبار تکنولوژی", url.strip(), admin_info=adm)
    return _redir(f"/admin/news-feed?flash={e('✅ ذخیره شد')}")


@router.post("/news-feed/refresh")
async def news_feed_refresh(request: Request):
    adm = _get_admin(request)
    guard = _require(adm, "news")
    if guard: return guard
    from core.news import get_feed
    d = get_feed(force=True)
    ok = d.get("status") == "ok"
    msg = f"{len(d.get('items') or [])} خبر گرفته شد" if ok else f"خطا: {(d.get('error') or '')[:80]}"
    _log(request, "بروزرسانی دستی فید اخبار", "اخبار تکنولوژی", msg, admin_info=adm)
    return _redir(f"/admin/news-feed?flash={e(('✅ ' if ok else '❌ ') + msg)}")


@router.post("/app-content/{cid}/delete")
async def app_content_delete(request: Request, cid: int):
    adm = _get_admin(request)
    guard = _require(adm, "news")
    if guard: return guard
    from db import delete_app_content
    delete_app_content(cid)
    return _redir("/admin/app-content")


# ══════════════════════════════════════════════════════════════════════════
# ─── امتیازها و پاداش سرزدن روزانهٔ مینی‌اپ ──────────────────────────────────
# ══════════════════════════════════════════════════════════════════════════

@router.get("/engagement", response_class=HTMLResponse)
async def engagement_page(request: Request, flash: str = ""):
    adm = _get_admin(request)
    guard = _require(adm, "notifications")
    if guard: return guard
    from db import get_cfg, get_all_ratings
    reward = get_cfg("DAILY_CHECKIN_REWARD", "500")
    ratings = get_all_ratings(limit=100)

    rows = "".join(f"""
      <tr class="border-b">
        <td class="p-2 text-sm">{html.escape(r.get('product_title') or '—')}</td>
        <td class="p-2 text-sm">{'⭐️'*int(r.get('rating') or 0)}</td>
        <td class="p-2 text-sm text-gray-600">{html.escape((r.get('comment') or '').strip() or '—')}</td>
        <td class="p-2 text-xs text-gray-400">{html.escape((r.get('full_name') or '').strip() or ('کاربر '+str(r.get('user_id'))))}</td>
        <td class="p-2 text-xs text-gray-400">{fa_date(r.get('created_at') or '')}</td>
        <td class="p-2">
          <form method="post" action="/admin/engagement/ratings/{r['id']}/delete" onsubmit="return confirm('این نظر حذف بشه؟')">
            <button class="text-red-500 text-xs">🗑 حذف</button>
          </form>
        </td>
      </tr>""" for r in ratings) or '<tr><td colspan="6" class="p-6 text-center text-gray-400 text-sm">هنوز نظری ثبت نشده.</td></tr>'

    body = f"""
    <div class="bg-white rounded-xl shadow-sm p-5 max-w-2xl mb-4">
      <div class="text-sm font-medium mb-2">🎁 پاداش سرزدن روزانهٔ مینی‌اپ</div>
      <div class="text-xs text-gray-400 mb-3">مبلغی که هر کاربر با اولین بازکردن اپ در هر روز (یک‌بار در روز) به کیف‌پولش اضافه می‌شه.</div>
      <form method="post" action="/admin/engagement/settings" class="flex gap-2">
        <input type="number" name="reward" value="{e(reward)}" min="0" step="100"
               class="flex-1 border rounded-lg p-2 text-sm" placeholder="مثلاً ۵۰۰">
        <span class="self-center text-sm text-gray-400">تومان</span>
        <button class="bg-indigo-600 text-white px-4 py-2 rounded-lg text-sm">ذخیره</button>
      </form>
    </div>
    <div class="bg-white rounded-xl shadow-sm overflow-x-auto">
      <div class="px-4 py-3 border-b bg-gray-50 flex items-center justify-between">
        <h2 class="font-bold text-gray-700 text-sm">⭐️ نظرات و امتیازهای محصولات ({len(ratings)})</h2>
      </div>
      <table class="w-full text-right">
        <thead><tr class="text-xs text-gray-400 border-b">
          <th class="p-2">محصول</th><th class="p-2">امتیاز</th><th class="p-2">نظر</th>
          <th class="p-2">کاربر</th><th class="p-2">تاریخ</th><th class="p-2">عملیات</th>
        </tr></thead>
        <tbody>{rows}</tbody>
      </table>
    </div>"""
    return _layout("امتیازها و پاداش روزانه", body, adm, flash=flash)


@router.post("/engagement/settings")
async def engagement_settings_save(request: Request, reward: str = Form("0")):
    adm = _get_admin(request)
    guard = _require(adm, "notifications")
    if guard: return guard
    from db import set_cfg
    try:
        val = max(0, int(reward))
    except Exception:
        val = 0
    set_cfg("DAILY_CHECKIN_REWARD", str(val))
    _log(request, "تنظیم پاداش سرزدن روزانه", "امتیازها و پاداش روزانه", str(val), admin_info=adm)
    return _redir(f"/admin/engagement?flash={e('✅ ذخیره شد')}")


@router.post("/engagement/ratings/{rid}/delete")
async def engagement_rating_delete(request: Request, rid: int):
    adm = _get_admin(request)
    guard = _require(adm, "notifications")
    if guard: return guard
    from db import delete_rating
    delete_rating(rid)
    _log(request, "حذف نظر محصول", "امتیازها و پاداش روزانه", f"rating #{rid}", admin_info=adm)
    return _redir(f"/admin/engagement?flash={e('✅ حذف شد')}")


# ══════════════════════════════════════════════════════════════════════════
# ─── درخواست‌های «اطلاع‌رسانی موجود شدن» (بخش ۴ سند مینی‌اپ) ──────────────────
# ══════════════════════════════════════════════════════════════════════════

@router.get("/stock-requests", response_class=HTMLResponse)
async def stock_requests_page(request: Request, flash: str = ""):
    adm = _get_admin(request)
    guard = _require(adm, "notifications")
    if guard: return guard
    from db import list_stock_requests
    reqs = list_stock_requests()

    groups = {}
    order = []
    for r in reqs:
        pid = r["product_id"]
        if pid not in groups:
            groups[pid] = {"title": r.get("product_title") or f"محصول #{pid}",
                            "stock": r.get("product_stock") or 0, "rows": []}
            order.append(pid)
        groups[pid]["rows"].append(r)

    def _row_html(r):
        who = html.escape((r.get("full_name") or "").strip() or f"کاربر {r['user_id']}")
        status = ('<span class="text-emerald-600 text-xs font-medium">✅ اطلاع داده شد</span>' if r["notified"]
                   else '<span class="text-amber-600 text-xs font-medium">⏳ در انتظار</span>')
        return f"""
          <tr class="border-b">
            <td class="p-2 text-sm">{who}</td>
            <td class="p-2 text-xs text-gray-400">{fa_date(r.get('created_at') or '')}</td>
            <td class="p-2">{status}</td>
            <td class="p-2">
              <form method="post" action="/admin/stock-requests/{r['id']}/delete" onsubmit="return confirm('این درخواست حذف بشه؟')">
                <button class="text-red-500 text-xs">🗑 حذف</button>
              </form>
            </td>
          </tr>"""

    if not order:
        cards = '<div class="bg-white rounded-xl shadow-sm p-10 text-center text-gray-400 text-sm">فعلاً هیچ درخواست اطلاع‌رسانی‌ای ثبت نشده.</div>'
    else:
        cards = ""
        for pid in order:
            g = groups[pid]
            pending = sum(1 for r in g["rows"] if not r["notified"])
            notify_btn = ""
            if pending:
                notify_btn = f"""
                  <form method="post" action="/admin/stock-requests/{pid}/notify-now" onsubmit="return confirm('پیام موجود شدن به {pending} کاربر منتظر ارسال بشه؟')">
                    <button class="bg-indigo-600 text-white text-xs px-3 py-1.5 rounded-lg">🔔 اطلاع‌رسانی الان</button>
                  </form>"""
            cards += f"""
            <div class="bg-white rounded-xl shadow-sm overflow-x-auto mb-4">
              <div class="px-4 py-3 border-b bg-gray-50 flex items-center justify-between flex-wrap gap-2">
                <div>
                  <h2 class="font-bold text-gray-700 text-sm">{html.escape(g['title'])}</h2>
                  <div class="text-xs text-gray-400 mt-0.5">موجودی فعلی: {g['stock']} — {len(g['rows'])} درخواست، {pending} در انتظار</div>
                </div>
                {notify_btn}
              </div>
              <table class="w-full text-right">
                <thead><tr class="text-xs text-gray-400 border-b">
                  <th class="p-2">کاربر</th><th class="p-2">تاریخ درخواست</th><th class="p-2">وضعیت</th><th class="p-2">عملیات</th>
                </tr></thead>
                <tbody>{"".join(_row_html(r) for r in g["rows"])}</tbody>
              </table>
            </div>"""

    body = f"""
    <div class="text-xs text-gray-400 mb-4">کاربرانی که روی محصول ناموجود دکمهٔ «اطلاع بده وقتی موجود شد» رو زدن (از ربات یا مینی‌اپ). با اضافه‌شدن موجودی از صفحهٔ «موجودی»، این کاربران خودکار مطلع می‌شن؛ از اینجا هم می‌شه دستی همین الان اطلاع‌رسانی کرد یا یک درخواست رو حذف کرد.</div>
    {cards}"""
    return _layout("درخواست‌های اطلاع‌رسانی موجودی", body, adm, flash=flash)


@router.post("/stock-requests/{rid}/delete")
async def stock_requests_delete(request: Request, rid: int):
    adm = _get_admin(request)
    guard = _require(adm, "notifications")
    if guard: return guard
    from db import delete_stock_request
    delete_stock_request(rid)
    _log(request, "حذف درخواست اطلاع‌رسانی موجودی", "درخواست‌های اطلاع‌رسانی موجودی", f"request #{rid}", admin_info=adm)
    return _redir(f"/admin/stock-requests?flash={e('✅ حذف شد')}")


@router.post("/stock-requests/{pid}/notify-now")
async def stock_requests_notify_now(request: Request, pid: int, background_tasks: BackgroundTasks):
    adm = _get_admin(request)
    guard = _require(adm, "notifications")
    if guard: return guard
    background_tasks.add_task(_notify_restock_subscribers, pid, True)
    _log(request, "اطلاع‌رسانی دستی موجودی", "درخواست‌های اطلاع‌رسانی موجودی", f"product:{pid}", admin_info=adm)
    return _redir(f"/admin/stock-requests?flash={e('✅ در حال ارسال اطلاع‌رسانی…')}")


# ══════════════════════════════════════════════════════════════════════════
# ─── آموزش — CMS داخلی کامل (فاز ۲) ────────────────────────────────────────
# ══════════════════════════════════════════════════════════════════════════

_TUT_IMAGE_EXTS = (".jpg", ".jpeg", ".png", ".webp", ".gif")
_TUT_VIDEO_EXTS = (".mp4", ".webm", ".mov")
_TUT_DOC_EXTS = (".pdf", ".zip", ".rar", ".7z", ".doc", ".docx", ".xls", ".xlsx", ".ppt", ".pptx", ".txt", ".csv")


async def _save_tutorial_file(file, subdir: str, allowed_exts: tuple, prefix: str, force_ext: bool = True) -> str:
    """آپلود یک فایل آموزش (کاور/گالری/ویدیو/دانلودی) — مسیر عمومی برمی‌گردونه یا ''.
    force_ext=True یعنی پسوند نامعتبر به اولین پسوند مجاز تبدیل می‌شه (برای تصویر بی‌خطره)؛
    False یعنی فایل با پسوند غیرمجاز کلاً رد می‌شه (برای ویدیو/فایل دانلودی که تبدیل پسوند خرابش می‌کنه)."""
    if file is None or not getattr(file, "filename", ""):
        return ""
    try:
        ext = os.path.splitext(file.filename)[1].lower()
        if ext not in allowed_exts:
            if not force_ext:
                return ""
            ext = allowed_exts[0]
        target_dir = os.path.join(APP_MEDIA_DIR, "tutorials", subdir)
        os.makedirs(target_dir, exist_ok=True)
        fname = f"{prefix}{int(time.time()*1000)}{ext}"
        target_path = os.path.join(target_dir, fname)
        from image_utils import COMPRESSIBLE_EXTS, compress_image_bytes
        if ext in COMPRESSIBLE_EXTS:
            # عکس (نه ویدیو/فایل دانلودی) — کل فایل رو می‌خونیم (حجم عکس معمولاً کوچیکه،
            # برخلاف ویدیو) تا با Pillow فشرده/resize بشه قبل از نوشتن روی دیسک.
            raw = await file.read()
            size = len(raw)
            if size:
                compressed = compress_image_bytes(raw, ext)
                with open(target_path, "wb") as f:
                    f.write(compressed)
        else:
            # ویدیو/فایل دانلودی/gif — تکه‌تکه روی دیسک می‌نویسیم، نه یکجا در RAM —
            # وگرنه آپلود ویدیوهای حجیم می‌تونه به مصرف حافظهٔ زیاد/تایم‌اوت منجر بشه.
            size = 0
            with open(target_path, "wb") as f:
                while True:
                    chunk = await file.read(1024 * 1024)
                    if not chunk:
                        break
                    size += len(chunk)
                    f.write(chunk)
        if size == 0:
            try:
                os.remove(target_path)
            except Exception:
                pass
            return ""
        return f"/app-media/tutorials/{subdir}/{fname}"
    except Exception:
        return ""


@router.get("/tutorials", response_class=HTMLResponse)
async def tutorials_page(request: Request, status: str = "", category: str = "", q: str = "", flash: str = ""):
    adm = _get_admin(request)
    guard = _require(adm, "articles")
    if guard: return guard
    from db import get_tutorials, get_tutorial_categories
    cats = get_tutorial_categories()
    cat_names = {c["id"]: c["name"] for c in cats}
    items = get_tutorials(
        category_id=int(category) if category.isdigit() else None,
        status=status or None, q=q.strip() or None, sort="newest", limit=200,
    )

    status_tabs = ''.join(
        f'<a href="/admin/tutorials?status={s}" class="px-3 py-1.5 rounded-lg text-xs border '
        f'{"bg-indigo-600 text-white" if status==s else "bg-white text-gray-500"}">{l}</a>'
        for s, l in [("", "همه"), ("published", "✅ منتشرشده"), ("draft", "📝 پیش‌نویس")]
    )
    cat_opts = ''.join(f'<option value="{c["id"]}" {"selected" if category==str(c["id"]) else ""}>{html.escape(c["name"])}</option>' for c in cats)

    rows = ""
    for it in items:
        badge = ('<span class="text-green-600 text-xs">✅ منتشرشده</span>' if it.get("status") == "published"
                 else '<span class="text-gray-400 text-xs">📝 پیش‌نویس</span>')
        featured = '<span class="text-amber-500 text-xs">⭐️</span>' if it.get("featured") else ""
        img = (f'<img src="{html.escape(it.get("cover_image") or "")}" class="w-10 h-10 rounded-lg object-cover">'
               if it.get("cover_image") else '<div class="w-10 h-10 rounded-lg bg-gray-100"></div>')
        toggle_label = "غیرفعال کن" if it.get("status") == "published" else "منتشر کن"
        rows += f"""
        <tr class="border-b">
          <td class="p-2">{img}</td>
          <td class="p-2 text-sm font-medium">{featured} {html.escape(it.get('title') or '')}</td>
          <td class="p-2 text-xs">{html.escape(cat_names.get(it.get('category_id'), '—'))}</td>
          <td class="p-2">{badge}</td>
          <td class="p-2 text-xs text-gray-400">👁 {it.get('view_count', 0)}</td>
          <td class="p-2 text-xs text-gray-400">{fa_date(it.get('publish_date') or it.get('created_at') or '')}</td>
          <td class="p-2 whitespace-nowrap">
            <a href="/admin/tutorials/{it['id']}/preview" target="_blank" class="text-gray-500 text-xs ml-2">👁 پیش‌نمایش</a>
            <a href="/admin/tutorials/{it['id']}/edit" class="text-indigo-600 text-xs ml-2">✏️ ویرایش</a>
            <form method="post" action="/admin/tutorials/{it['id']}/toggle" class="inline">
              <button class="text-amber-600 text-xs ml-2">{toggle_label}</button>
            </form>
            <form method="post" action="/admin/tutorials/{it['id']}/delete" class="inline" onsubmit="return confirm('حذف شود؟')">
              <button class="text-red-500 text-xs">🗑 حذف</button>
            </form>
          </td>
        </tr>"""
    if not rows:
        rows = '<tr><td colspan="7" class="p-6 text-center text-gray-400 text-sm">آموزشی یافت نشد.</td></tr>'

    body = f"""
    <div class="flex items-center justify-between mb-4 flex-wrap gap-2">
      <div class="flex gap-2 flex-wrap">{status_tabs}</div>
      <div class="flex gap-2">
        <a href="/admin/tutorials/categories" class="border px-4 py-2 rounded-lg text-sm">🏷 دسته‌بندی‌ها</a>
        <a href="/admin/tutorials/new" class="bg-indigo-600 text-white px-4 py-2 rounded-lg text-sm">＋ افزودن آموزش</a>
      </div>
    </div>
    <form method="get" action="/admin/tutorials" class="flex gap-2 mb-4">
      <input type="hidden" name="status" value="{e(status)}">
      <input name="q" value="{html.escape(q)}" placeholder="جستجو در عنوان/توضیح…" class="flex-1 border rounded-lg p-2 text-sm">
      <select name="category" class="border rounded-lg p-2 text-sm" onchange="this.form.submit()">
        <option value="">همهٔ دسته‌ها</option>{cat_opts}
      </select>
      <button class="border px-4 py-2 rounded-lg text-sm">جستجو</button>
    </form>
    <div class="bg-white rounded-xl shadow-sm overflow-x-auto">
      <table class="w-full text-right">
        <thead><tr class="text-xs text-gray-400 border-b">
          <th class="p-2">کاور</th><th class="p-2">عنوان</th><th class="p-2">دسته</th>
          <th class="p-2">وضعیت</th><th class="p-2">بازدید</th><th class="p-2">تاریخ</th><th class="p-2">عملیات</th>
        </tr></thead>
        <tbody>{rows}</tbody>
      </table>
    </div>"""
    return _layout("آموزش", body, adm, flash=flash)


def _tutorial_form(it=None, categories=None):
    it = it or {}
    categories = categories or []
    cat_opts = ''.join(
        f'<option value="{c["id"]}" {"selected" if it.get("category_id")==c["id"] else ""}>{html.escape(c["name"])}</option>'
        for c in categories
    )
    tags_str = ", ".join(json.loads(it.get("tags") or "[]")) if it else ""
    gallery = json.loads(it.get("gallery") or "[]") if it else []
    gallery_preview = "".join(f"""
      <label class="relative inline-block">
        <img src="{html.escape(g)}" class="w-16 h-16 rounded-lg object-cover border">
        <input type="checkbox" name="gallery_remove" value="{html.escape(g)}" class="absolute top-1 right-1">
      </label>""" for g in gallery)
    checked_pub = "checked" if it.get("status") == "published" else ""
    checked_draft = "checked" if it.get("status") != "published" else ""
    checked_featured = "checked" if it.get("featured") else ""
    body_json = json.dumps(it.get("body") or "")

    from datetime import date as _dt_date
    pubdate_g = str(it.get("publish_date") or "").strip() or _dt_date.today().isoformat()
    pubdate_fa = _to_jalali(pubdate_g)  # فقط برای نمایش اولیه؛ ذخیره از publish_date_fa در سمت سرور انجام می‌شه

    return f"""
    <form method="post" action="/admin/tutorials/save" enctype="multipart/form-data" class="max-w-3xl space-y-4">
      <input type="hidden" name="tid" value="{it.get('id','')}">
      <div class="bg-white rounded-xl shadow-sm p-5 space-y-4">
        <div>
          <label class="block text-xs text-gray-500 mb-1">عنوان</label>
          <input name="title" required value="{html.escape(str(it.get('title') or ''))}" class="w-full border rounded-lg p-2 text-sm">
        </div>
        <div>
          <label class="block text-xs text-gray-500 mb-1">توضیح کوتاه</label>
          <textarea name="short_desc" rows="2" class="w-full border rounded-lg p-2 text-sm">{html.escape(str(it.get('short_desc') or ''))}</textarea>
        </div>
        <div>
          <label class="block text-xs text-gray-500 mb-1">تصویر کاور {f'<img src="{html.escape(it.get("cover_image",""))}" class="inline w-8 h-8 rounded object-cover align-middle mr-1">' if it.get('cover_image') else ''}</label>
          <div class="flex items-center gap-2">
            <input type="file" name="cover_image" id="cover_image_input" accept="image/*" class="flex-1 min-w-0 text-sm">
            <button type="button" onclick="document.getElementById('cover_image_input').value=''" class="shrink-0 text-xs text-red-500 border border-red-200 rounded-lg px-2 py-1.5 whitespace-nowrap">✕ پاک کردن</button>
          </div>
          {f'<label class="flex items-center gap-1 text-xs text-red-500 mt-1"><input type="checkbox" name="cover_image_remove" value="1"> 🗑 حذف تصویر فعلاً بارگذاری‌شده</label>' if it.get('cover_image') else ''}
        </div>
        <div>
          <label class="block text-xs text-gray-500 mb-1">متن کامل آموزش</label>
          <link href="https://cdn.jsdelivr.net/npm/quill@2.0.2/dist/quill.snow.css" rel="stylesheet">
          <div id="quill-editor" class="quill-editor-box rounded-lg border"></div>
          <textarea name="body" id="body-input" hidden></textarea>
        </div>
        <div>
          <label class="block text-xs text-gray-500 mb-1">گالری تصاویر</label>
          <div class="flex gap-2 flex-wrap mb-2">{gallery_preview}</div>
          <input type="file" name="gallery" accept="image/*" multiple class="w-full text-sm">
          <div class="text-xs text-gray-400 mt-1">تیک‌خورده‌ها موقع ذخیره حذف می‌شن؛ فایل‌های جدید اضافه می‌شن.</div>
        </div>
        <div class="grid grid-cols-1 md:grid-cols-2 gap-4">
          <div>
            <label class="block text-xs text-gray-500 mb-1">ویدئوی آپلودی {f'(فعلی: <a href="{html.escape(it.get("video_upload",""))}" target="_blank" class="text-indigo-600">مشاهده</a>)' if it.get('video_upload') else ''}</label>
            <div class="flex items-center gap-2">
              <input type="file" name="video_upload" id="video_upload_input" accept="video/*" class="flex-1 min-w-0 text-sm">
              <button type="button" onclick="document.getElementById('video_upload_input').value=''" class="shrink-0 text-xs text-red-500 border border-red-200 rounded-lg px-2 py-1.5 whitespace-nowrap">✕ پاک کردن</button>
            </div>
            {f'<label class="flex items-center gap-1 text-xs text-red-500 mt-1"><input type="checkbox" name="video_upload_remove" value="1"> 🗑 حذف ویدئوی فعلاً بارگذاری‌شده</label>' if it.get('video_upload') else ''}
          </div>
          <div>
            <label class="block text-xs text-gray-500 mb-1">لینک ویدیو (آپارات/یوتیوب/مستقیم)</label>
            <input name="video_link" value="{html.escape(str(it.get('video_link') or ''))}" dir="ltr" class="w-full border rounded-lg p-2 text-sm" placeholder="https://aparat.com/v/...">
          </div>
        </div>
        <div class="grid grid-cols-1 md:grid-cols-2 gap-4">
          <div>
            <label class="block text-xs text-gray-500 mb-1">فایل دانلودی {f'(فعلی: <a href="{html.escape(it.get("download_file",""))}" target="_blank" class="text-indigo-600">دانلود</a>)' if it.get('download_file') else ''}</label>
            <div class="flex items-center gap-2">
              <input type="file" name="download_file" id="download_file_input" class="flex-1 min-w-0 text-sm">
              <button type="button" onclick="document.getElementById('download_file_input').value=''" class="shrink-0 text-xs text-red-500 border border-red-200 rounded-lg px-2 py-1.5 whitespace-nowrap">✕ پاک کردن</button>
            </div>
            {f'<label class="flex items-center gap-1 text-xs text-red-500 mt-1"><input type="checkbox" name="download_file_remove" value="1"> 🗑 حذف فایل فعلاً بارگذاری‌شده</label>' if it.get('download_file') else ''}
          </div>
          <div>
            <label class="block text-xs text-gray-500 mb-1">برچسب دکمهٔ دانلود</label>
            <input name="download_label" value="{html.escape(str(it.get('download_label') or ''))}" class="w-full border rounded-lg p-2 text-sm" placeholder="مثلاً: دانلود PDF">
          </div>
        </div>
        <div class="grid grid-cols-1 md:grid-cols-2 gap-4">
          <div>
            <label class="block text-xs text-gray-500 mb-1">دسته‌بندی</label>
            <select name="category_id" class="w-full border rounded-lg p-2 text-sm">
              <option value="0">بدون دسته</option>{cat_opts}
            </select>
          </div>
          <div>
            <label class="block text-xs text-gray-500 mb-1">برچسب‌ها (با کاما جدا کنید)</label>
            <input name="tags" value="{html.escape(tags_str)}" class="w-full border rounded-lg p-2 text-sm" placeholder="اپل، آموزش، مبتدی">
          </div>
        </div>
        <div class="grid grid-cols-1 md:grid-cols-3 gap-4">
          <div>
            <label class="block text-xs text-gray-500 mb-1">تاریخ انتشار (شمسی)</label>
            <input type="text" name="publish_date_fa" id="pubdate_fa" value="{pubdate_fa}" placeholder="۱۴۰۴/۰۱/۰۱"
              class="w-full border rounded-lg p-2 text-sm" autocomplete="off">
          </div>
          <div>
            <label class="block text-xs text-gray-500 mb-1">ترتیب نمایش</label>
            <input type="number" name="sort_order" value="{it.get('sort_order', 0)}" class="w-full border rounded-lg p-2 text-sm">
          </div>
          <div class="flex items-end pb-2">
            <label class="flex items-center gap-2 text-sm"><input type="checkbox" name="featured" value="1" {checked_featured}> ⭐️ آموزش ویژه</label>
          </div>
        </div>
        <div class="flex gap-4 text-sm">
          <label class="flex items-center gap-2"><input type="radio" name="status" value="published" {checked_pub}> منتشر شده</label>
          <label class="flex items-center gap-2"><input type="radio" name="status" value="draft" {checked_draft}> پیش‌نویس</label>
        </div>
      </div>
      <div class="flex gap-2">
        <button id="tut-save-btn" class="bg-indigo-600 text-white px-5 py-2 rounded-lg text-sm">ذخیره</button>
        <a href="/admin/tutorials" class="px-5 py-2 rounded-lg text-sm border">انصراف</a>
      </div>
      <div id="tut-upload-note" hidden class="text-xs text-amber-600 bg-amber-50 border border-amber-200 rounded-lg p-2">
        ⏳ در حال آپلود… برای فایل‌های بزرگ (ویدیو) ممکنه چند دقیقه طول بکشه، لطفاً این صفحه رو نبند و دوباره روی دکمه نزن.
      </div>
    </form>
    <script src="https://cdn.jsdelivr.net/npm/quill@2.0.2/dist/quill.js"></script>
    <script>
    (function(){{
      var quill = new Quill('#quill-editor', {{theme:'snow', placeholder:'متن کامل آموزش را بنویسید...'}});
      quill.root.innerHTML = {body_json};
      document.querySelector('form[action="/admin/tutorials/save"]').addEventListener('submit', function(){{
        document.getElementById('body-input').value = quill.root.innerHTML;
        var btn=document.getElementById('tut-save-btn');
        btn.disabled=true;btn.textContent='⏳ در حال آپلود…';btn.classList.add('opacity-60');
        document.getElementById('tut-upload-note').hidden=false;
      }});
    }})();
    </script>"""


@router.get("/tutorials/new", response_class=HTMLResponse)
async def tutorials_new(request: Request):
    adm = _get_admin(request)
    guard = _require(adm, "articles")
    if guard: return guard
    from db import get_tutorial_categories
    return _layout("افزودن آموزش", _tutorial_form(categories=get_tutorial_categories()), adm)


@router.get("/tutorials/{tid}/edit", response_class=HTMLResponse)
async def tutorials_edit(request: Request, tid: int):
    adm = _get_admin(request)
    guard = _require(adm, "articles")
    if guard: return guard
    from db import get_tutorial, get_tutorial_categories
    it = get_tutorial(tid)
    if not it:
        return _redir("/admin/tutorials")
    return _layout("ویرایش آموزش", _tutorial_form(it, get_tutorial_categories()), adm)


@router.get("/tutorials/{tid}/preview", response_class=HTMLResponse)
async def tutorials_preview(request: Request, tid: int):
    adm = _get_admin(request)
    guard = _require(adm, "articles")
    if guard: return guard
    from db import get_tutorial, get_tutorial_categories
    it = get_tutorial(tid)
    if not it:
        return _redir("/admin/tutorials")
    cats = {c["id"]: c["name"] for c in get_tutorial_categories()}
    gallery = json.loads(it.get("gallery") or "[]")
    tags = json.loads(it.get("tags") or "[]")
    gallery_html = "".join(f'<img src="{html.escape(g)}" class="w-24 h-24 rounded-lg object-cover">' for g in gallery)
    tags_html = "".join(f'<span class="bg-gray-100 text-xs px-2 py-1 rounded-full">{html.escape(t)}</span>' for t in tags)
    body = f"""
    <div class="max-w-2xl bg-white rounded-xl shadow-sm p-6 space-y-4">
      {f'<img src="{html.escape(it.get("cover_image",""))}" class="w-full rounded-xl">' if it.get('cover_image') else ''}
      <div class="text-xs text-gray-400">{html.escape(cats.get(it.get('category_id'), '—'))} · {html.escape(str(it.get('publish_date') or ''))}</div>
      <h1 class="text-xl font-bold">{html.escape(it.get('title') or '')}</h1>
      <p class="text-gray-500">{html.escape(it.get('short_desc') or '')}</p>
      <div class="flex gap-2 flex-wrap">{tags_html}</div>
      <div class="prose max-w-none">{it.get('body') or ''}</div>
      <div class="flex gap-2 flex-wrap">{gallery_html}</div>
      {f'<video controls class="w-full rounded-xl" src="{html.escape(it.get("video_upload",""))}"></video>' if it.get('video_upload') else ''}
      {f'<a href="{html.escape(it.get("video_link",""))}" target="_blank" class="text-indigo-600 text-sm block">🎬 مشاهدهٔ ویدیو</a>' if it.get('video_link') else ''}
      {f'<a href="{html.escape(it.get("download_file",""))}" target="_blank" class="bg-gray-100 rounded-lg px-4 py-2 text-sm inline-block">⬇️ {html.escape(it.get("download_label") or "دانلود فایل")}</a>' if it.get('download_file') else ''}
    </div>"""
    return _layout("پیش‌نمایش آموزش", body, adm)


@router.post("/tutorials/save")
async def tutorials_save(request: Request):
    adm = _get_admin(request)
    guard = _require(adm, "articles")
    if guard: return guard
    from db import add_tutorial, update_tutorial, get_tutorial, jalali_str_to_gregorian_iso

    form = await request.form()
    tid = str(form.get("tid") or "").strip()
    title = str(form.get("title") or "").strip()
    if not title:
        return _redir("/admin/tutorials?flash=" + e("❌ عنوان الزامی است"))

    existing = get_tutorial(int(tid)) if tid.isdigit() else None

    cover_upload = form.get("cover_image")
    if cover_upload is not None and getattr(cover_upload, "filename", ""):
        cover_path = await _save_tutorial_file(cover_upload, "cover", _TUT_IMAGE_EXTS, "cv")
    elif form.get("cover_image_remove") == "1":
        cover_path = ""
    else:
        cover_path = existing.get("cover_image") if existing else ""

    video_upload_file = form.get("video_upload")
    if video_upload_file is not None and getattr(video_upload_file, "filename", ""):
        video_upload_path = await _save_tutorial_file(video_upload_file, "video", _TUT_VIDEO_EXTS, "vd", force_ext=False)
    elif form.get("video_upload_remove") == "1":
        video_upload_path = ""
    else:
        video_upload_path = existing.get("video_upload") if existing else ""

    download_upload_file = form.get("download_file")
    if download_upload_file is not None and getattr(download_upload_file, "filename", ""):
        download_path = await _save_tutorial_file(download_upload_file, "files", _TUT_DOC_EXTS, "dl", force_ext=False)
    elif form.get("download_file_remove") == "1":
        download_path = ""
    else:
        download_path = existing.get("download_file") if existing else ""

    # گالری: حذف تیک‌خورده‌ها + اضافه‌کردن فایل‌های جدید
    gallery = json.loads(existing.get("gallery") or "[]") if existing else []
    removed = set(form.getlist("gallery_remove"))
    gallery = [g for g in gallery if g not in removed]
    for gf in form.getlist("gallery"):
        if getattr(gf, "filename", ""):
            path = await _save_tutorial_file(gf, "gallery", _TUT_IMAGE_EXTS, "gl")
            if path:
                gallery.append(path)

    tags = [t.strip() for t in str(form.get("tags") or "").split(",") if t.strip()]

    try:
        category_id = int(form.get("category_id") or 0)
    except Exception:
        category_id = 0
    try:
        sort_order = int(form.get("sort_order") or 0)
    except Exception:
        sort_order = 0

    from datetime import date as _tut_date
    publish_date = jalali_str_to_gregorian_iso(str(form.get("publish_date_fa") or "").strip()) or _tut_date.today().isoformat()

    fields = dict(
        title=title,
        cover_image=cover_path or "",
        short_desc=str(form.get("short_desc") or "").strip(),
        body=str(form.get("body") or ""),
        gallery=json.dumps(gallery, ensure_ascii=False),
        video_upload=video_upload_path or "",
        video_link=str(form.get("video_link") or "").strip(),
        download_file=download_path or "",
        download_label=str(form.get("download_label") or "").strip(),
        category_id=category_id,
        tags=json.dumps(tags, ensure_ascii=False),
        status=str(form.get("status") or "draft"),
        publish_date=publish_date,
        sort_order=sort_order,
        featured=1 if form.get("featured") == "1" else 0,
    )

    if existing:
        update_tutorial(existing["id"], **fields)
        _log(request, "ویرایش آموزش", "آموزش", title, admin_info=adm)
    else:
        add_tutorial(**fields)
        _log(request, "افزودن آموزش", "آموزش", title, admin_info=adm)
    return _redir(f"/admin/tutorials?flash={e('✅ ذخیره شد')}")


@router.post("/tutorials/{tid}/toggle")
async def tutorials_toggle(request: Request, tid: int):
    adm = _get_admin(request)
    guard = _require(adm, "articles")
    if guard: return guard
    from db import get_tutorial, set_tutorial_status
    it = get_tutorial(tid)
    if it:
        new_status = "draft" if it.get("status") == "published" else "published"
        set_tutorial_status(tid, new_status)
        _log(request, f"سوییچ وضعیت آموزش → {new_status}", "آموزش", it.get("title") or "", admin_info=adm)
    return _redir("/admin/tutorials")


@router.post("/tutorials/{tid}/delete")
async def tutorials_delete(request: Request, tid: int):
    adm = _get_admin(request)
    guard = _require(adm, "articles")
    if guard: return guard
    from db import delete_tutorial
    delete_tutorial(tid)
    _log(request, "حذف آموزش", "آموزش", str(tid), admin_info=adm)
    return _redir("/admin/tutorials")


@router.get("/tutorials/categories", response_class=HTMLResponse)
async def tutorial_categories_page(request: Request, flash: str = ""):
    adm = _get_admin(request)
    guard = _require(adm, "articles")
    if guard: return guard
    from db import get_tutorial_categories
    cats = get_tutorial_categories()
    rows = "".join(f"""
    <tr class="border-b">
      <td class="p-2 text-sm">{html.escape(c['name'])}</td>
      <td class="p-2 text-xs text-gray-400" dir="ltr">{html.escape(c.get('slug') or '')}</td>
      <td class="p-2">
        <form method="post" action="/admin/tutorials/categories/{c['id']}/delete" onsubmit="return confirm('حذف شود؟ آموزش‌های این دسته بدون‌دسته می‌شن.')">
          <button class="text-red-500 text-xs">🗑 حذف</button>
        </form>
      </td>
    </tr>""" for c in cats) or '<tr><td colspan="3" class="p-6 text-center text-gray-400 text-sm">دسته‌ای ثبت نشده.</td></tr>'

    body = f"""
    <a href="/admin/tutorials" class="text-indigo-600 text-sm mb-4 inline-block">← بازگشت به آموزش‌ها</a>
    <div class="bg-white rounded-xl shadow-sm p-5 max-w-lg mb-4">
      <div class="text-sm font-medium mb-2">افزودن دستهٔ جدید</div>
      <form method="post" action="/admin/tutorials/categories/add" class="flex gap-2">
        <input name="name" required placeholder="نام دسته" class="flex-1 border rounded-lg p-2 text-sm">
        <button class="bg-indigo-600 text-white px-4 py-2 rounded-lg text-sm">افزودن</button>
      </form>
    </div>
    <div class="bg-white rounded-xl shadow-sm overflow-x-auto max-w-lg">
      <table class="w-full text-right">
        <thead><tr class="text-xs text-gray-400 border-b"><th class="p-2">نام</th><th class="p-2">اسلاگ</th><th class="p-2">عملیات</th></tr></thead>
        <tbody>{rows}</tbody>
      </table>
    </div>"""
    return _layout("دسته‌بندی‌های آموزش", body, adm, flash=flash)


@router.post("/tutorials/categories/add")
async def tutorial_categories_add(request: Request, name: str = Form(...)):
    adm = _get_admin(request)
    guard = _require(adm, "articles")
    if guard: return guard
    from db import add_tutorial_category
    if name.strip():
        add_tutorial_category(name.strip())
        _log(request, "افزودن دستهٔ آموزش", "آموزش", name.strip(), admin_info=adm)
    return _redir("/admin/tutorials/categories")


@router.post("/tutorials/categories/{cid}/delete")
async def tutorial_categories_delete(request: Request, cid: int):
    adm = _get_admin(request)
    guard = _require(adm, "articles")
    if guard: return guard
    from db import delete_tutorial_category
    delete_tutorial_category(cid)
    _log(request, "حذف دستهٔ آموزش", "آموزش", str(cid), admin_info=adm)
    return _redir("/admin/tutorials/categories")


# ══════════════════════════════════════════════════════════════════════════
# ─── کارشناسی هوشمند قیمت آیفون ────────────────────────────────────────────
# ══════════════════════════════════════════════════════════════════════════

_IV_CATEGORY_LABELS = {
    "grade": "🏷 سری اصالت دستگاه (M/N/F/P/۳/۴/۵) — پیش‌فرض قیمت روی M است",
    "condition": "وضعیت کلی دستگاه",
    "battery": "سلامت باتری",
    "repair": "تعمیرات / بازشدگی",
    "registry": "وضعیت مالکیت",
    "box": "جعبه و لوازم",
    "cosmetic": "وضعیت ظاهری",
    "cable": "کابل",
    "component": "قطعات خراب (فقط وقتی «نیازمند تعمیر» انتخاب بشه پرسیده می‌شه)",
}


def _iv_num(v):
    return f"{v:,}" if v else "—"


_IV_DIGIT_MAP = str.maketrans("۰۱۲۳۴۵۶۷۸۹٠١٢٣٤٥٦٧٨٩", "01234567890123456789")


def _iv_parse_num(raw, cast=int, default=0):
    """اعداد فارسی/عربی رو به انگلیسی تبدیل می‌کنه و به int/float پارس می‌کنه —
    چون input type=number روی کیبورد فارسی موبایل، رقم فارسی رو اصلاً قبول نمی‌کنه
    (کلاً وارد نمی‌شه)، این فیلدها به text+inputmode تبدیل شدن و پارس اینجا انجام می‌شه."""
    if raw is None:
        return default
    s = str(raw).translate(_IV_DIGIT_MAP).strip()
    s = s.replace(",", "").replace("٬", "").replace(" ", "").replace("٫", ".")
    if not s or s in ("-", "."):
        return default
    try:
        f = float(s)
    except ValueError:
        return default
    return cast(f)


@router.get("/iphone", response_class=HTMLResponse)
async def iphone_dashboard(request: Request, flash: str = ""):
    adm = _get_admin(request)
    guard = _require(adm, "ai_pricing")
    if guard: return guard
    from ui_texts import is_main_button_enabled
    import iphone_valuation.db as ivdb
    stats = ivdb.get_stats()
    enabled = is_main_button_enabled("MAIN_BTN_IPHONE_VALUATION")

    popular_rows = "".join(
        f'<tr class="border-b"><td class="p-2 text-sm whitespace-nowrap">{html.escape(p["name"] or "—")}</td>'
        f'<td class="p-2 text-sm text-gray-500 whitespace-nowrap">{p["cnt"]}</td></tr>'
        for p in stats["popular_models"]
    ) or '<tr><td colspan="2" class="p-4 text-center text-gray-400 text-sm">هنوز کارشناسی‌ای ثبت نشده.</td></tr>'

    body = f"""
    <div class="bg-white rounded-xl shadow-sm p-5 mb-4 flex items-center justify-between flex-wrap gap-3">
      <div>
        <div class="text-sm font-medium mb-1">📱 کارشناس هوشمند قیمت آیفون</div>
        <div class="text-xs text-gray-400">فعال/غیرفعال بودن دکمهٔ منوی اصلی ربات</div>
      </div>
      <form method="post" action="/admin/iphone/toggle">
        <button class="px-4 py-2 rounded-lg text-sm font-medium {'bg-red-50 text-red-600' if enabled else 'bg-emerald-50 text-emerald-600'}">
          {'⛔️ غیرفعال کردن' if enabled else '✅ فعال کردن'}
        </button>
      </form>
    </div>

    <div class="grid grid-cols-2 md:grid-cols-4 gap-3 mb-4">
      <div class="bg-white rounded-xl shadow-sm p-4"><div class="text-xs text-gray-400 mb-1">کارشناسی امروز</div><div class="text-xl font-bold">{stats['today']}</div></div>
      <div class="bg-white rounded-xl shadow-sm p-4"><div class="text-xs text-gray-400 mb-1">کل کارشناسی‌ها</div><div class="text-xl font-bold">{stats['total']}</div></div>
      <div class="bg-white rounded-xl shadow-sm p-4"><div class="text-xs text-gray-400 mb-1">میانگین قیمت منصفانه</div><div class="text-xl font-bold">{_iv_num(stats['avg_fair_price'])}</div></div>
      <div class="bg-white rounded-xl shadow-sm p-4"><div class="text-xs text-gray-400 mb-1">میانگین اختلاف با پیشنهاد کاربر</div><div class="text-xl font-bold">{stats['avg_price_gap_pct']}٪</div></div>
    </div>

    <div class="grid md:grid-cols-2 gap-4 mb-4">
      <a href="/admin/iphone/prices" class="bg-white rounded-xl shadow-sm p-4 hover:shadow-md transition block">
        <div class="font-medium text-sm mb-1">💰 قیمت‌ها</div>
        <div class="text-xs text-gray-400">لیست کامل قیمت‌های ثبت‌شده + ثبت سریع قیمت تازه</div>
      </a>
      <a href="/admin/iphone/coefficients" class="bg-white rounded-xl shadow-sm p-4 hover:shadow-md transition block">
        <div class="font-medium text-sm mb-1">⚖️ ضرایب قیمت و امتیاز</div>
        <div class="text-xs text-gray-400">باتری، رجیستری، پک، ظاهر، کابل + وزن امتیازدهی</div>
      </a>
      <a href="/admin/iphone/repairs" class="bg-white rounded-xl shadow-sm p-4 hover:shadow-md transition block">
        <div class="font-medium text-sm mb-1">🩹 مدیریت تعمیرات</div>
        <div class="text-xs text-gray-400">قطعات معیوب و تعویض‌شده — یک لیست، دو درصد جدا</div>
      </a>
      <a href="/admin/iphone/series" class="bg-white rounded-xl shadow-sm p-4 hover:shadow-md transition block">
        <div class="font-medium text-sm mb-1">🗂 گروه‌بندی نسل‌ها</div>
        <div class="text-xs text-gray-400">دسته‌بندی مدل‌ها برای مرحلهٔ اول ویزارد ربات</div>
      </a>
      <a href="/admin/iphone/colors" class="bg-white rounded-xl shadow-sm p-4 hover:shadow-md transition block">
        <div class="font-medium text-sm mb-1">🎨 مدیریت رنگ‌ها</div>
        <div class="text-xs text-gray-400">تغییر نام نمایشی رنگ + درصد اثر روی قیمت</div>
      </a>
      <a href="/admin/iphone/fx" class="bg-white rounded-xl shadow-sm p-4 hover:shadow-md transition block">
        <div class="font-medium text-sm mb-1">💵 نرخ ارز</div>
        <div class="text-xs text-gray-400">منابع خودکار یا نرخ دستی + حساسیت نوسان</div>
      </a>
      <a href="/admin/iphone/history" class="bg-white rounded-xl shadow-sm p-4 hover:shadow-md transition block">
        <div class="font-medium text-sm mb-1">🕓 تاریخچهٔ کارشناسی‌ها</div>
        <div class="text-xs text-gray-400">لیست کامل کارشناسی‌های انجام‌شده</div>
      </a>
      <a href="/admin/iphone/ai" class="bg-white rounded-xl shadow-sm p-4 hover:shadow-md transition block">
        <div class="font-medium text-sm mb-1">🤖 کارشناس مکمل هوش مصنوعی</div>
        <div class="text-xs text-gray-400">اتصال AI، بازهٔ مجاز تعدیل قیمت، آستانهٔ اطمینان</div>
      </a>
    </div>

    <div class="bg-white rounded-xl shadow-sm w-full">
      <div class="px-4 py-3 border-b bg-gray-50"><h2 class="font-bold text-gray-700 text-sm">🔥 محبوب‌ترین مدل‌ها</h2></div>
      <div class="overflow-x-auto">
        <table class="w-full text-right">
          <thead><tr class="text-xs text-gray-400 border-b"><th class="p-2">مدل</th><th class="p-2">تعداد کارشناسی</th></tr></thead>
          <tbody>{popular_rows}</tbody>
        </table>
      </div>
    </div>"""
    return _layout("کارشناسی آیفون", body, adm, flash=flash)


@router.post("/iphone/toggle")
async def iphone_toggle(request: Request):
    adm = _get_admin(request)
    guard = _require(adm, "ai_pricing")
    if guard: return guard
    from ui_texts import is_main_button_enabled, set_main_button_enabled
    cur = is_main_button_enabled("MAIN_BTN_IPHONE_VALUATION")
    set_main_button_enabled("MAIN_BTN_IPHONE_VALUATION", not cur)
    _log(request, "فعال/غیرفعال‌سازی کارشناسی آیفون", "کارشناسی آیفون", "غیرفعال" if cur else "فعال", admin_info=adm)
    return _redir("/admin/iphone")


@router.get("/iphone/prices", response_class=HTMLResponse)
async def iphone_prices_page(request: Request, flash: str = "", edit: str = ""):
    """صفحهٔ مستقل و ساده فقط برای ثبت/ویرایش/حذف قیمت — یک فرم واحد با دراپ‌داون‌های
    آبشاری (مدل → ظرفیت → رنگ/پارت/سری اصالت چند‌انتخابی) بالای صفحه، و لیست تخت همهٔ
    قیمت‌های ثبت‌شده (شبیه لیست تیکت‌ها) پایینش — هر ردیف یعنی یک رکورد قیمت مستقل.
    ادیت جدای صفحه نداره — طبق درخواست صریح مالک پروژه، زدن «✏️ ویرایش» همین فرم بالا رو
    با تمام تنظیمات اون ردیف (مدل/ظرفیت/رنگ/پارت/سری اصالت/قیمت) از پیش پر می‌کنه، دقیقاً
    مثل زمان تعریف اولیه — با همون قابلیت چند‌انتخابی."""
    adm = _get_admin(request)
    guard = _require(adm, "ai_pricing")
    if guard: return guard
    import iphone_valuation.db as ivdb
    models = ivdb.list_models(active_only=False)
    parts = ivdb.list_parts(active_only=True)
    series_list = ivdb.list_series(active_only=False)
    # سری‌های P/۳/۴ مسیر محاسبهٔ جدا (توقف قیمت‌گذاری، بخش ۲۲.۷ CLAUDE.md) دارن — تعریف
    # ردیف قیمت دقیق براشون بی‌معنیه، پس اصلاً توی این فرم پیشنهاد نمی‌شن.
    grades = [g for g in ivdb.list_coefficients(category="grade", active_only=True)
              if g["option_key"] not in ivdb.GRADE_STOP_CALC_KEYS]

    edit_cap = ivdb.get_capacity(int(edit)) if edit.isdigit() else None

    model_data = {}
    for m in models:
        storages = ivdb.list_storages(m["id"], active_only=True)
        colors = ivdb.list_colors(model_id=m["id"], active_only=True)
        part_relevant = bool((m.get("dual_sim_parts") or "").strip())
        # ترکیب‌های (ظرفیت+پارت+رنگ+سری اصالت) که از قبل برای این مدل ثبت شدن — سمت کلاینت
        # برای غیرفعال/خط‌خورده کردن چک‌باکسی که انتخابش دقیقاً یه ترکیب تکراری می‌سازه
        # استفاده می‌شه (ردیفی که خودِ ادمین الان داره ادیت می‌کنه از قبل تیک‌خورده می‌مونه،
        # پس این منطق مزاحمش نمی‌شه — چون چک‌باکس‌های تیک‌خورده هیچ‌وقت غیرفعال نمی‌شن).
        existing_combos = [
            {"storage_id": c["storage_id"], "part_id": c["part_id"],
             "color_id": c["color_id"], "grade_id": c["grade_id"]}
            for c in ivdb.list_capacities(model_id=m["id"], active_only=True)
        ]
        model_data[str(m["id"])] = {
            # colors/parts/grades همیشه کامل می‌رن (نه فقط وقتی toggle روشنه) — چون خودِ
            # toggle هم همینجاست و اگه ادمین روشنش کنه، دراپ‌داون باید بی‌نیاز از رفرش صفحه پر بشه.
            "storages": [{"id": s["id"], "label": s["label"]} for s in storages],
            "colors": [{"id": c["id"], "name": c["name"]} for c in colors],
            "color_pricing": bool(m.get("color_pricing")),
            "part_relevant": part_relevant,
            "part_pricing": bool(m.get("part_pricing")),
            "parts": [{"id": p["id"], "label": p["label"]} for p in parts] if part_relevant else [],
            "grade_pricing": bool(m.get("grade_pricing")),
            "grades": [{"id": g["id"], "label": g["option_label"]} for g in grades],
            # capacity_pricing برخلاف سه توگل بالا پیش‌فرضش روشنه (۱) — اکثر مدل‌ها واقعاً
            # با ظرفیت قیمت‌شون فرق می‌کنه، بی‌اثر بودنش استثناست نه قاعده.
            "capacity_pricing": bool(m.get("capacity_pricing", 1)),
            "existing_combos": existing_combos,
        }

    # مدل‌ها گروه‌بندی‌شده بر اساس سری/نسل (قدیمی‌ترین تا جدیدترین) — همون ترتیبی که
    # /admin/iphone/series نگه می‌داره؛ مدل‌های بدون سری آخر، زیر یک گروه جدا.
    models_by_series: dict = {}
    unassigned_models = []
    for m in models:
        sid = m.get("series_id")
        if sid:
            models_by_series.setdefault(sid, []).append(m)
        else:
            unassigned_models.append(m)
    model_opts = ""
    for s in series_list:
        group = models_by_series.get(s["id"]) or []
        if not group:
            continue
        opts = "".join(f'<option value="{gm["id"]}">{html.escape(gm["name"])}</option>' for gm in group)
        model_opts += f'<optgroup label="{html.escape(s["name"])}">{opts}</optgroup>'
    if unassigned_models:
        opts = "".join(f'<option value="{gm["id"]}">{html.escape(gm["name"])} ({html.escape(gm["series"])})</option>'
                        for gm in unassigned_models)
        model_opts += f'<optgroup label="بدون سری">{opts}</optgroup>'

    add_form = f"""
    <div class="bg-white rounded-xl shadow-sm p-4 mb-4" id="iv-p-editform">
      <div class="text-sm font-medium mb-3">📝 ثبت قیمت تازه</div>
      <div id="iv-p-editbanner" class="text-xs bg-amber-50 text-amber-700 border border-amber-200 rounded-lg px-3 py-2 mb-3" style="display:none">
        ✏️ در حال ویرایش یک ردیف قیمت ثبت‌شده — <a href="/admin/iphone/prices" class="underline">لغو ویرایش</a>
      </div>
      <form method="post" action="/admin/iphone/prices/upsert" class="grid grid-cols-2 sm:grid-cols-4 gap-2 items-end">
        <input type="hidden" name="edit_cap_id" id="iv-p-editcapid" value="">
        <div class="flex flex-col gap-0.5 col-span-2 sm:col-span-1">
          <span class="text-[9px] text-gray-400">مدل</span>
          <select name="model_id" id="iv-p-model" class="border rounded p-1.5 text-xs" required>
            <option value="">— انتخاب مدل —</option>
            {model_opts}
          </select>
        </div>
        <div class="flex flex-col gap-0.5" id="iv-p-storage-wrap">
          <span class="text-[9px] text-gray-400">ظرفیت</span>
          <select name="storage_id" id="iv-p-storage" class="border rounded p-1.5 text-xs" required disabled>
            <option value="">ابتدا مدل رو انتخاب کن</option>
          </select>
        </div>
        <div class="flex flex-col gap-0.5 col-span-2 sm:col-span-1" id="iv-p-color-wrap" style="display:none">
          <span class="text-[9px] text-gray-400">رنگ (چند انتخابی — خالی=همهٔ رنگ‌ها، یک قیمت مشترک)</span>
          <div id="iv-p-color-list" class="flex flex-wrap gap-1 border rounded p-1.5 text-xs max-h-20 overflow-y-auto"></div>
        </div>
        <div class="flex flex-col gap-0.5 col-span-2 sm:col-span-1" id="iv-p-grade-wrap" style="display:none">
          <span class="text-[9px] text-gray-400">سری اصالت (چند انتخابی — خالی=همهٔ سری‌ها، یک قیمت مشترک)</span>
          <div id="iv-p-grade-list" class="flex flex-wrap gap-1 border rounded p-1.5 text-xs max-h-20 overflow-y-auto"></div>
        </div>
        <div class="flex flex-col gap-0.5 col-span-2 sm:col-span-1" id="iv-p-part-wrap" style="display:none">
          <span class="text-[9px] text-gray-400">پارت (چند انتخابی — خالی=همهٔ پارت‌ها، یک قیمت مشترک)</span>
          <div id="iv-p-part-list" class="flex flex-wrap gap-1 border rounded p-1.5 text-xs max-h-20 overflow-y-auto"></div>
        </div>
        <input type="text" inputmode="numeric" dir="ltr" name="base_price" class="border rounded p-1.5 text-xs" placeholder="قیمت پایه" required>
        <input type="text" inputmode="numeric" dir="ltr" name="buy_price_ref" class="border rounded p-1.5 text-xs" placeholder="قیمت خرید" required>
        <input type="text" inputmode="numeric" dir="ltr" name="sell_price_ref" class="border rounded p-1.5 text-xs" placeholder="قیمت فروش" required>
        <input type="text" inputmode="numeric" dir="ltr" name="fx_ref_rate" class="border rounded p-1.5 text-xs" placeholder="نرخ ارز مرجع (خالی=فعلی)">
        <button class="bg-indigo-600 text-white px-4 py-2 rounded-lg text-xs">💾 ذخیره</button>
      </form>
      <p class="text-[10px] text-gray-400 mt-2">چک‌باکس‌های خاکستری/خط‌خورده یعنی این ترکیب دقیق قبلاً برای همین ظرفیت ثبت شده — برای ویرایشش از لیست پایین «✏️ ویرایش» بزن.</p>
      <div class="flex flex-wrap gap-3 mt-4 pt-4 border-t" id="iv-p-toggles-row" style="display:none">
        <div class="flex items-center gap-3 bg-gray-50 rounded-lg px-3 py-2 flex-1 min-w-[200px]" id="iv-p-colorpricing-wrap">
          <span class="text-xs text-gray-600 flex-1">🎨 اثر رنگ روی قیمت این مدل</span>
          <label class="relative inline-flex items-center cursor-pointer shrink-0">
            <input type="checkbox" id="iv-p-colorpricing-cb" class="sr-only peer">
            <div class="w-11 h-6 bg-gray-200 rounded-full peer transition-colors peer-checked:bg-indigo-600
                        after:content-[''] after:absolute after:top-[2px] after:start-[2px] after:bg-white
                        after:rounded-full after:h-5 after:w-5 after:transition-all after:shadow
                        peer-checked:after:translate-x-full rtl:peer-checked:after:-translate-x-full"></div>
          </label>
        </div>
        <div class="flex items-center gap-3 bg-gray-50 rounded-lg px-3 py-2 flex-1 min-w-[200px]" id="iv-p-partpricing-wrap" style="display:none">
          <span class="text-xs text-gray-600 flex-1">🔠 اثر پارت روی قیمت این مدل</span>
          <label class="relative inline-flex items-center cursor-pointer shrink-0">
            <input type="checkbox" id="iv-p-partpricing-cb" class="sr-only peer">
            <div class="w-11 h-6 bg-gray-200 rounded-full peer transition-colors peer-checked:bg-indigo-600
                        after:content-[''] after:absolute after:top-[2px] after:start-[2px] after:bg-white
                        after:rounded-full after:h-5 after:w-5 after:transition-all after:shadow
                        peer-checked:after:translate-x-full rtl:peer-checked:after:-translate-x-full"></div>
          </label>
        </div>
        <div class="flex items-center gap-3 bg-gray-50 rounded-lg px-3 py-2 flex-1 min-w-[200px]" id="iv-p-gradepricing-wrap">
          <span class="text-xs text-gray-600 flex-1">🏷 اثر سری اصالت روی قیمت این مدل</span>
          <label class="relative inline-flex items-center cursor-pointer shrink-0">
            <input type="checkbox" id="iv-p-gradepricing-cb" class="sr-only peer">
            <div class="w-11 h-6 bg-gray-200 rounded-full peer transition-colors peer-checked:bg-indigo-600
                        after:content-[''] after:absolute after:top-[2px] after:start-[2px] after:bg-white
                        after:rounded-full after:h-5 after:w-5 after:transition-all after:shadow
                        peer-checked:after:translate-x-full rtl:peer-checked:after:-translate-x-full"></div>
          </label>
        </div>
        <div class="flex items-center gap-3 bg-gray-50 rounded-lg px-3 py-2 flex-1 min-w-[200px]" id="iv-p-capacitypricing-wrap">
          <span class="text-xs text-gray-600 flex-1">🗄 اثر ظرفیت روی قیمت این مدل</span>
          <label class="relative inline-flex items-center cursor-pointer shrink-0">
            <input type="checkbox" id="iv-p-capacitypricing-cb" class="sr-only peer">
            <div class="w-11 h-6 bg-gray-200 rounded-full peer transition-colors peer-checked:bg-indigo-600
                        after:content-[''] after:absolute after:top-[2px] after:start-[2px] after:bg-white
                        after:rounded-full after:h-5 after:w-5 after:transition-all after:shadow
                        peer-checked:after:translate-x-full rtl:peer-checked:after:-translate-x-full"></div>
          </label>
        </div>
      </div>
    </div>
    <script>
    var IV_MODEL_DATA = {json.dumps(model_data, ensure_ascii=False)};
    var IV_EDIT_CAP = {json.dumps({
        "id": edit_cap["id"], "model_id": edit_cap["model_id"], "storage_id": edit_cap.get("storage_id"),
        "color_id": edit_cap.get("color_id"), "part_id": edit_cap.get("part_id"), "grade_id": edit_cap.get("grade_id"),
        "base_price": edit_cap["base_price"], "buy_price_ref": edit_cap["buy_price_ref"],
        "sell_price_ref": edit_cap["sell_price_ref"],
    }, ensure_ascii=False) if edit_cap else "null"};
    (function(){{
      var modelSel = document.getElementById('iv-p-model');
      var storWrap = document.getElementById('iv-p-storage-wrap');
      var storSel = document.getElementById('iv-p-storage');
      var colorWrap = document.getElementById('iv-p-color-wrap');
      var colorList = document.getElementById('iv-p-color-list');
      var gradeWrap = document.getElementById('iv-p-grade-wrap');
      var gradeList = document.getElementById('iv-p-grade-list');
      var partWrap = document.getElementById('iv-p-part-wrap');
      var partList = document.getElementById('iv-p-part-list');
      var togglesRow = document.getElementById('iv-p-toggles-row');
      var cpWrap = document.getElementById('iv-p-colorpricing-wrap');
      var cpCb = document.getElementById('iv-p-colorpricing-cb');
      var ppWrap = document.getElementById('iv-p-partpricing-wrap');
      var ppCb = document.getElementById('iv-p-partpricing-cb');
      var gpCb = document.getElementById('iv-p-gradepricing-cb');
      var capCb = document.getElementById('iv-p-capacitypricing-cb');
      function refreshColorVisibility(d){{
        colorWrap.style.display = (d.color_pricing && d.colors.length) ? '' : 'none';
      }}
      function refreshPartVisibility(d){{
        partWrap.style.display = (d.part_pricing && d.parts.length) ? '' : 'none';
      }}
      function refreshGradeVisibility(d){{
        gradeWrap.style.display = (d.grade_pricing && d.grades.length) ? '' : 'none';
      }}
      function refreshStorageVisibility(d){{
        // capacity_pricing برخلاف بقیه پیش‌فرضش روشنه — یعنی وقتی خاموشه، فیلد ظرفیت
        // مخفی و غیرالزامی می‌شه (قیمت برای همهٔ ظرفیت‌های این مدل یکسانه).
        var on = !!d.capacity_pricing;
        storWrap.style.display = on ? '' : 'none';
        storSel.required = on;
      }}
      function mkCheckbox(container, fieldName, id, label){{
        var lbl = document.createElement('label');
        lbl.className = 'flex items-center gap-1 bg-gray-50 border rounded px-1.5 py-0.5 cursor-pointer';
        var cb = document.createElement('input');
        cb.type = 'checkbox'; cb.name = fieldName; cb.value = id; cb.className = 'ml-0.5';
        lbl.appendChild(cb);
        lbl.appendChild(document.createTextNode(label));
        container.appendChild(lbl);
      }}
      function comboKey(s, p, c, g){{
        return [s===null?'':s, p===null?'':p, c===null?'':c, g===null?'':g].join('|');
      }}
      function firstChecked(container){{
        var cb = container.querySelector('input[type=checkbox]:checked');
        return cb ? parseInt(cb.value, 10) : null;
      }}
      // ترکیب‌های از قبل ثبت‌شده رو نشون می‌ده — برای هر چک‌باکس، ترکیب کاندید رو با
      // *اولین* گزینهٔ تیک‌خوردهٔ دو دستهٔ دیگه می‌سازه (نه کارتزین کامل — برای حالت رایج
      // «یه پارت/سری ثابت + چند رنگ متغیر» دقیقه، برای موارد خیلی چندبعدی فقط یه راهنماست،
      // بلاک واقعی سمت سرور اتفاق می‌افته). چک‌باکس‌های از قبل تیک‌خورده (مثلاً موقع ادیت)
      // هیچ‌وقت غیرفعال نمی‌شن — کاربر همیشه می‌تونه انتخاب فعلیش رو تغییر بده.
      function recomputeDuplicates(){{
        var d = IV_MODEL_DATA[modelSel.value];
        if(!d) return;
        var existingSet = {{}};
        (d.existing_combos || []).forEach(function(row){{
          existingSet[comboKey(row.storage_id, row.part_id, row.color_id, row.grade_id)] = true;
        }});
        var curStorage = (d.capacity_pricing && storSel.value) ? parseInt(storSel.value, 10) : null;
        function markDim(container, dim){{
          if(!container) return;
          container.querySelectorAll('label').forEach(function(lbl){{
            var cb = lbl.querySelector('input');
            var val = parseInt(cb.value, 10);
            var p = dim === 'part' ? val : firstChecked(partList);
            var c = dim === 'color' ? val : firstChecked(colorList);
            var g = dim === 'grade' ? val : firstChecked(gradeList);
            var dup = !!existingSet[comboKey(curStorage, p, c, g)];
            if(dup && !cb.checked){{
              cb.disabled = true;
              lbl.classList.add('opacity-40', 'line-through');
              lbl.title = 'قبلاً برای این ترکیب ثبت شده';
            }} else {{
              cb.disabled = false;
              lbl.classList.remove('opacity-40', 'line-through');
              lbl.title = '';
            }}
          }});
        }}
        markDim(colorList, 'color');
        markDim(partList, 'part');
        markDim(gradeList, 'grade');
      }}
      modelSel.addEventListener('change', function(){{
        var d = IV_MODEL_DATA[modelSel.value];
        storSel.innerHTML = '';
        partList.innerHTML = '';
        colorList.innerHTML = '';
        gradeList.innerHTML = '';
        if(!d){{
          storSel.disabled = true;
          storSel.innerHTML = '<option value="">ابتدا مدل رو انتخاب کن</option>';
          colorWrap.style.display = 'none'; partWrap.style.display = 'none'; gradeWrap.style.display = 'none'; togglesRow.style.display = 'none';
          return;
        }}
        storSel.disabled = false;
        if(!d.storages.length){{ storSel.innerHTML = '<option value="">این مدل ظرفیتی نداره</option>'; }}
        d.storages.forEach(function(s){{ var o=document.createElement('option'); o.value=s.id; o.textContent=s.label; storSel.appendChild(o); }});
        d.colors.forEach(function(c){{ mkCheckbox(colorList, 'color_ids', c.id, c.name); }});
        d.parts.forEach(function(p){{ mkCheckbox(partList, 'part_ids', p.id, p.label); }});
        d.grades.forEach(function(g){{ mkCheckbox(gradeList, 'grade_ids', g.id, g.label); }});
        refreshColorVisibility(d);
        refreshPartVisibility(d);
        refreshGradeVisibility(d);
        refreshStorageVisibility(d);
        togglesRow.style.display = '';
        cpCb.checked = !!d.color_pricing;
        ppWrap.style.display = d.part_relevant ? '' : 'none';
        ppCb.checked = !!d.part_pricing;
        gpCb.checked = !!d.grade_pricing;
        capCb.checked = !!d.capacity_pricing;
        recomputeDuplicates();
      }});
      storSel.addEventListener('change', recomputeDuplicates);
      colorList.addEventListener('change', recomputeDuplicates);
      partList.addEventListener('change', recomputeDuplicates);
      gradeList.addEventListener('change', recomputeDuplicates);
      cpCb.addEventListener('change', function(){{
        var mid = modelSel.value;
        var d = IV_MODEL_DATA[mid];
        if(!mid || !d) return;
        var wanted = cpCb.checked;
        fetch('/admin/iphone/prices/color-pricing', {{
          method: 'POST',
          headers: {{'Content-Type': 'application/x-www-form-urlencoded'}},
          body: 'model_id=' + encodeURIComponent(mid) + '&enabled=' + (wanted ? '1' : '0')
        }}).then(function(r){{ return r.json(); }}).then(function(res){{
          if(!res.ok){{ cpCb.checked = !wanted; alert(res.error || 'خطا در ذخیره'); return; }}
          d.color_pricing = wanted;
          refreshColorVisibility(d);
        }}).catch(function(){{ cpCb.checked = !wanted; alert('خطا در ارتباط با سرور'); }});
      }});
      ppCb.addEventListener('change', function(){{
        var mid = modelSel.value;
        var d = IV_MODEL_DATA[mid];
        if(!mid || !d) return;
        var wanted = ppCb.checked;
        fetch('/admin/iphone/prices/part-pricing', {{
          method: 'POST',
          headers: {{'Content-Type': 'application/x-www-form-urlencoded'}},
          body: 'model_id=' + encodeURIComponent(mid) + '&enabled=' + (wanted ? '1' : '0')
        }}).then(function(r){{ return r.json(); }}).then(function(res){{
          if(!res.ok){{ ppCb.checked = !wanted; alert(res.error || 'خطا در ذخیره'); return; }}
          d.part_pricing = wanted;
          refreshPartVisibility(d);
        }}).catch(function(){{ ppCb.checked = !wanted; alert('خطا در ارتباط با سرور'); }});
      }});
      gpCb.addEventListener('change', function(){{
        var mid = modelSel.value;
        var d = IV_MODEL_DATA[mid];
        if(!mid || !d) return;
        var wanted = gpCb.checked;
        fetch('/admin/iphone/prices/grade-pricing', {{
          method: 'POST',
          headers: {{'Content-Type': 'application/x-www-form-urlencoded'}},
          body: 'model_id=' + encodeURIComponent(mid) + '&enabled=' + (wanted ? '1' : '0')
        }}).then(function(r){{ return r.json(); }}).then(function(res){{
          if(!res.ok){{ gpCb.checked = !wanted; alert(res.error || 'خطا در ذخیره'); return; }}
          d.grade_pricing = wanted;
          refreshGradeVisibility(d);
        }}).catch(function(){{ gpCb.checked = !wanted; alert('خطا در ارتباط با سرور'); }});
      }});
      capCb.addEventListener('change', function(){{
        var mid = modelSel.value;
        var d = IV_MODEL_DATA[mid];
        if(!mid || !d) return;
        var wanted = capCb.checked;
        fetch('/admin/iphone/prices/capacity-pricing', {{
          method: 'POST',
          headers: {{'Content-Type': 'application/x-www-form-urlencoded'}},
          body: 'model_id=' + encodeURIComponent(mid) + '&enabled=' + (wanted ? '1' : '0')
        }}).then(function(r){{ return r.json(); }}).then(function(res){{
          if(!res.ok){{ capCb.checked = !wanted; alert(res.error || 'خطا در ذخیره'); return; }}
          d.capacity_pricing = wanted;
          refreshStorageVisibility(d);
        }}).catch(function(){{ capCb.checked = !wanted; alert('خطا در ارتباط با سرور'); }});
      }});
      // حالت ادیت: طبق درخواست صریح مالک پروژه، ادیت صفحهٔ جدا نداره — همین فرم بالا با
      // تمام تنظیمات ردیف انتخابی (مدل/ظرفیت/رنگ/پارت/سری اصالت/قیمت) از پیش پر می‌شه،
      // دقیقاً مثل زمان تعریف اولیه، با همون چک‌باکس‌های چند‌انتخابی.
      if(IV_EDIT_CAP){{
        var ec = IV_EDIT_CAP;
        modelSel.value = String(ec.model_id);
        modelSel.dispatchEvent(new Event('change'));
        if(ec.storage_id){{ storSel.value = String(ec.storage_id); }}
        function checkValue(container, val){{
          if(val === null || val === undefined) return;
          var cb = container.querySelector('input[value="' + val + '"]');
          if(cb) cb.checked = true;
        }}
        checkValue(colorList, ec.color_id);
        checkValue(partList, ec.part_id);
        checkValue(gradeList, ec.grade_id);
        document.querySelector('input[name=base_price]').value = ec.base_price;
        document.querySelector('input[name=buy_price_ref]').value = ec.buy_price_ref;
        document.querySelector('input[name=sell_price_ref]').value = ec.sell_price_ref;
        document.getElementById('iv-p-editcapid').value = String(ec.id);
        document.getElementById('iv-p-editbanner').style.display = '';
        recomputeDuplicates();
        setTimeout(function(){{
          document.getElementById('iv-p-editform').scrollIntoView({{behavior: 'smooth', block: 'start'}});
        }}, 50);
      }}
    }})();
    </script>
    """

    caps = ivdb.list_capacities(active_only=False)
    model_names = {m["id"]: f"{m['name']} ({m['series']})" for m in models}
    model_sort_order = {m["id"]: m["sort_order"] for m in models}
    caps.sort(key=lambda c: (model_sort_order.get(c["model_id"], 9999), ivdb.capacity_sort_key(c["capacity_label"])))

    rows_html = ""
    hidden_forms = ""
    for c in caps:
        mname = model_names.get(c["model_id"], "—")
        rows_html += f"""
        <tr class="border-b hover:bg-gray-50 {'opacity-40' if not c['active'] else ''}" data-model="{e(mname.lower())}">
          <td class="px-3 py-2 text-xs font-medium whitespace-nowrap">{html.escape(mname)}</td>
          <td class="px-3 py-2 text-xs whitespace-nowrap">{html.escape(c['capacity_label']) or '—'}</td>
          <td class="px-3 py-2 text-xs whitespace-nowrap">{html.escape(c['color']) or '—'}</td>
          <td class="px-3 py-2 text-xs whitespace-nowrap">{html.escape(c['part_number']) or '—'}</td>
          <td class="px-3 py-2 text-xs whitespace-nowrap">{html.escape(c.get('grade_label') or '') or '—'}</td>
          <td class="px-3 py-2 text-xs text-gray-400 whitespace-nowrap">{fa_date(c['updated_at'] or '', with_time=True)}</td>
          <td class="px-3 py-2 whitespace-nowrap">
            <a href="/admin/iphone/prices?edit={c['id']}#iv-p-editform" class="px-2 py-1 text-xs bg-indigo-50 text-indigo-700 border border-indigo-200 rounded inline-block">✏️ ویرایش</a>
            <button type="submit" form="iv-pd-{c['id']}" class="px-2 py-1 text-xs bg-red-50 text-red-500 border border-red-200 rounded mr-1">🗑 حذف</button>
          </td>
        </tr>"""
        hidden_forms += (
            f'<form id="iv-pd-{c["id"]}" method="post" action="/admin/iphone/prices/{c["id"]}/delete" '
            f'onsubmit="return confirm(\'این ردیف قیمت حذف بشه؟\')"></form>')

    body = f"""
    <a href="/admin/iphone" class="text-indigo-600 text-sm mb-4 inline-block">← بازگشت به کارشناسی آیفون</a>
    <div class="flex items-center justify-between mb-4 flex-wrap gap-2">
      <h1 class="text-xl font-bold text-gray-800">💰 قیمت‌های آیفون</h1>
    </div>
    {add_form}
    <div class="bg-white rounded-xl shadow-sm p-3 mb-3">
      <input type="text" id="iv-price-search" placeholder="🔍 جست‌وجو بر اساس نام مدل…" class="border rounded-lg p-2 text-sm w-full">
    </div>
    <div class="card overflow-hidden">
      <div class="overflow-x-auto">
        <table class="w-full text-right min-w-max" id="iv-price-table">
          <thead><tr class="text-xs text-gray-500 border-b bg-gray-50">
            <th class="px-3 py-3">مدل</th><th class="px-3 py-3">ظرفیت</th><th class="px-3 py-3">رنگ</th>
            <th class="px-3 py-3">پارت</th><th class="px-3 py-3">سری اصالت</th><th class="px-3 py-3">بروزرسانی</th><th class="px-3 py-3"></th>
          </tr></thead>
          <tbody>{rows_html or '<tr><td colspan="7" class="text-center py-8 text-gray-400">هنوز قیمتی ثبت نشده.</td></tr>'}</tbody>
        </table>
      </div>
    </div>
    {hidden_forms}
    <script>
    document.getElementById('iv-price-search').addEventListener('input', function(){{
      var q = this.value.trim().toLowerCase();
      document.querySelectorAll('#iv-price-table tbody tr[data-model]').forEach(function(tr){{
        tr.style.display = tr.dataset.model.indexOf(q) === -1 ? 'none' : '';
      }});
    }});
    </script>
    """
    return _layout("قیمت‌های آیفون", body, adm, flash=flash)


@router.post("/iphone/prices/upsert")
async def iphone_prices_upsert(request: Request):
    """فرم واحد ثبت/ویرایش قیمت — طبق درخواست صریح مالک پروژه، ادیت صفحهٔ جدا نداره،
    همین روت هم برای ثبت تازه هم برای ویرایش (وقتی `edit_cap_id` توی فرم باشه) استفاده
    می‌شه. رنگ/پارت/سری اصالت هرسه **چند‌انتخابی**‌ان (چک‌باکس) — روی کارتزین کامل سه بعد
    حلقه می‌زنه و به‌ازای هر ترکیب upsert می‌کنه (مثلاً «۱۳ نرمال، پارت ZA/A و CH/A با هم،
    همهٔ رنگ‌ها به‌جز آبی» با یک بار ثبت ممکنه).

    قانون تازهٔ ضدتکرار: برخلاف رفتار قبلی (upsert بی‌صدا روی هر ترکیب تکراری)، حالا اگه
    یه ترکیب دقیقاً از قبل برای یه ردیف *دیگه* (نه همون ردیفی که با edit_cap_id داریم
    ویرایشش می‌کنیم) ثبت شده باشه، **رد می‌شه و توی پیام نهایی گزارش می‌شه** — تا ادمین
    سهواً قیمت یه ترکیب موجود رو بدون اطلاع بازنویسی نکنه. برای عوض‌کردن قیمت یه ترکیب
    از قبل ثبت‌شده، باید از دکمهٔ «✏️ ویرایش» همون ردیف استفاده کرد (که edit_cap_id رو ست
    می‌کنه و دقیقاً همون ترکیب رو مجاز به آپدیت می‌کنه، نه رد)."""
    adm = _get_admin(request)
    guard = _require(adm, "ai_pricing")
    if guard: return guard
    import iphone_valuation.db as ivdb
    import iphone_valuation.fx as ivfx
    form = await request.form()
    model_id_raw = (form.get("model_id") or "").strip()
    storage_id_raw = (form.get("storage_id") or "").strip()
    if not model_id_raw.isdigit():
        return _redir(f"/admin/iphone/prices?flash={e('مدل الزامی است')}")
    model_id = int(model_id_raw)
    m0 = ivdb.get_model(model_id)
    # ظرفیت فقط وقتی الزامیه که «اثر ظرفیت روی قیمت» برای این مدل روشن باشه (پیش‌فرض)؛
    # وقتی خاموشه، فیلد توی فرم اصلاً نمایش داده نمی‌شه (JS)، پس اینجا هم اجباری نیست —
    # storage_id=None یعنی این ردیف قیمت برای همهٔ ظرفیت‌های این مدل مشترکه.
    if m0 and not m0.get("capacity_pricing", 1):
        storage_id = None
    elif storage_id_raw.isdigit():
        storage_id = int(storage_id_raw)
    else:
        return _redir(f"/admin/iphone/prices?flash={e('ظرفیت الزامی است')}")
    edit_cap_id_raw = (form.get("edit_cap_id") or "").strip()
    edit_cap_id = int(edit_cap_id_raw) if edit_cap_id_raw.isdigit() else None
    part_ids_raw = [v.strip() for v in form.getlist("part_ids") if v.strip()]
    color_ids_raw = [v.strip() for v in form.getlist("color_ids") if v.strip()]
    grade_ids_raw = [v.strip() for v in form.getlist("grade_ids") if v.strip()]
    base_price = form.get("base_price") or ""
    buy_price_ref = form.get("buy_price_ref") or ""
    sell_price_ref = form.get("sell_price_ref") or ""
    fx_ref_rate = form.get("fx_ref_rate") or ""
    # حتی اگه فرم دستکاری بشه، اگه قیمت‌گذاری رنگ/پارت/سری‌اصالت برای این مدل خاموش باشه،
    # همیشه روی ردیف عمومی (color_id/part_id/grade_id=NULL) ذخیره می‌شه — همون‌طور که
    # resolve_capacity هم طبق همین سه فلگ رفتار می‌کنه، تا هیچ ردیف یتیمی که هیچ‌وقت
    # استفاده نمی‌شه ساخته نشه.
    m = m0
    # دفاع در عمق: حتی اگه فرم دستکاری بشه، شناسهٔ سری‌های توقف‌محاسبه (P/۳/۴) هیچ‌وقت
    # به‌عنوان بعد قیمت‌گذاری ذخیره نمی‌شه — این سه مسیر محاسبهٔ جدا دارن (بخش ۲۲.۷ CLAUDE.md).
    stop_grade_ids = {g["id"] for g in ivdb.list_coefficients(category="grade", active_only=False)
                       if g["option_key"] in ivdb.GRADE_STOP_CALC_KEYS}
    part_ids = [int(v) for v in part_ids_raw] if (m and m.get("part_pricing")) else []
    color_ids = [int(v) for v in color_ids_raw] if (m and m.get("color_pricing")) else []
    grade_ids = ([int(v) for v in grade_ids_raw if int(v) not in stop_grade_ids]
                 if (m and m.get("grade_pricing")) else [])
    part_ids = part_ids or [None]
    color_ids = color_ids or [None]
    grade_ids = grade_ids or [None]
    bp, bpr, spr = _iv_parse_num(base_price), _iv_parse_num(buy_price_ref), _iv_parse_num(sell_price_ref)
    fx_rate = _iv_parse_num(fx_ref_rate) or ivfx.get_current_rate()

    color_map = {c["id"]: c["name"] for c in ivdb.list_colors(model_id=model_id, active_only=False)}
    part_map = {p["id"]: p["label"] for p in ivdb.list_parts(active_only=False)}
    grade_map = {g["id"]: g["option_label"] for g in ivdb.list_coefficients(category="grade", active_only=False)}

    def _combo_label(pid, cid_, gid):
        bits = [x for x in (color_map.get(cid_), part_map.get(pid), grade_map.get(gid)) if x]
        return "/".join(bits) if bits else "عمومی"

    count, skipped = 0, []
    for pid_ in part_ids:
        for cid_ in color_ids:
            for gid_ in grade_ids:
                existing = ivdb.get_capacity_exact(model_id, storage_id, pid_, cid_, gid_)
                if existing and existing["id"] != edit_cap_id:
                    skipped.append(_combo_label(pid_, cid_, gid_))
                    continue
                ivdb.upsert_capacity(model_id, storage_id, bp, bpr, spr, fx_rate,
                                      part_id=pid_, color_id=cid_, grade_id=gid_)
                count += 1

    _log(request, "ثبت/به‌روزرسانی قیمت آیفون", "کارشناسی آیفون",
         f"model={model_id} storage={storage_id} parts={part_ids} colors={color_ids} grades={grade_ids} "
         f"({count} ردیف ذخیره، {len(skipped)} تکراری رد شد)", admin_info=adm)
    msg = "✅ قیمت ذخیره شد" if count == 1 else f"✅ {count} ردیف قیمت ذخیره شد"
    if count == 0 and skipped:
        msg = "⚠️ هیچ ردیف تازه‌ای ذخیره نشد — همهٔ ترکیب‌های انتخابی قبلاً ثبت شده بودن"
    if skipped:
        shown = "، ".join(skipped[:5]) + ("، …" if len(skipped) > 5 else "")
        msg += f" — ⚠️ {len(skipped)} ترکیب قبلاً ثبت شده بود و رد شد: {shown}"
    return _redir(f"/admin/iphone/prices?flash={e(msg)}")


@router.post("/iphone/prices/{cid}/delete")
async def iphone_prices_delete(request: Request, cid: int):
    adm = _get_admin(request)
    guard = _require(adm, "ai_pricing")
    if guard: return guard
    import iphone_valuation.db as ivdb
    ivdb.delete_capacity(cid)
    return _redir(f"/admin/iphone/prices?flash={e('✅ ردیف قیمت حذف شد')}")


@router.post("/iphone/prices/color-pricing")
async def iphone_prices_color_pricing_toggle(request: Request):
    """توگل «رنگ روی قیمت این مدل اثر داره یا نه» — تنها بازماندهٔ صفحهٔ حذف‌شدهٔ
    مدل‌ها، طبق درخواست مالک پروژه به همین‌جا (بخش قیمت‌ها) منتقل شد."""
    from fastapi.responses import JSONResponse
    adm = _get_admin(request)
    guard = _require(adm, "ai_pricing")
    if guard: return JSONResponse({"error": "unauthorized"})
    form = await request.form()
    model_id = (form.get("model_id") or "").strip()
    if not model_id.isdigit():
        return JSONResponse({"error": "مدل نامعتبر"})
    enabled = form.get("enabled") == "1"
    import iphone_valuation.db as ivdb
    ivdb.update_model(int(model_id), color_pricing=1 if enabled else 0)
    _log(request, "تغییر اثر رنگ روی قیمت", "قیمت‌های آیفون",
         f"model:{model_id} -> {'روشن' if enabled else 'خاموش'}", admin_info=adm)
    return JSONResponse({"ok": True})


@router.post("/iphone/prices/part-pricing")
async def iphone_prices_part_pricing_toggle(request: Request):
    """توگل «پارت روی قیمت این مدل اثر داره یا نه» — دقیقاً همون الگوی توگل رنگ بالا،
    فقط تا این نشست هیچ UI ای نداشت (فقط فلگ دیتابیسش بود)."""
    from fastapi.responses import JSONResponse
    adm = _get_admin(request)
    guard = _require(adm, "ai_pricing")
    if guard: return JSONResponse({"error": "unauthorized"})
    form = await request.form()
    model_id = (form.get("model_id") or "").strip()
    if not model_id.isdigit():
        return JSONResponse({"error": "مدل نامعتبر"})
    enabled = form.get("enabled") == "1"
    import iphone_valuation.db as ivdb
    ivdb.update_model(int(model_id), part_pricing=1 if enabled else 0)
    _log(request, "تغییر اثر پارت روی قیمت", "قیمت‌های آیفون",
         f"model:{model_id} -> {'روشن' if enabled else 'خاموش'}", admin_info=adm)
    return JSONResponse({"ok": True})


@router.post("/iphone/prices/grade-pricing")
async def iphone_prices_grade_pricing_toggle(request: Request):
    """توگل «سری اصالت روی قیمت این مدل اثر داره یا نه» — دقیقاً همون الگوی توگل رنگ/پارت
    بالا؛ این فلگ فقط تعیین‌کنندهٔ *بعد سوم ردیف قیمت دقیق* است، کاملاً جدا از دستهٔ سراسری
    ضریب درصدی grade در iv_coefficients (بخش ۲۲.۷ CLAUDE.md)."""
    from fastapi.responses import JSONResponse
    adm = _get_admin(request)
    guard = _require(adm, "ai_pricing")
    if guard: return JSONResponse({"error": "unauthorized"})
    form = await request.form()
    model_id = (form.get("model_id") or "").strip()
    if not model_id.isdigit():
        return JSONResponse({"error": "مدل نامعتبر"})
    enabled = form.get("enabled") == "1"
    import iphone_valuation.db as ivdb
    ivdb.update_model(int(model_id), grade_pricing=1 if enabled else 0)
    _log(request, "تغییر اثر سری اصالت روی قیمت", "قیمت‌های آیفون",
         f"model:{model_id} -> {'روشن' if enabled else 'خاموش'}", admin_info=adm)
    return JSONResponse({"ok": True})


@router.post("/iphone/prices/capacity-pricing")
async def iphone_prices_capacity_pricing_toggle(request: Request):
    """توگل «ظرفیت روی قیمت این مدل اثر داره یا نه» — دقیقاً همون الگوی سه توگل بالا،
    فقط پیش‌فرضش برعکسه (روشن) چون ظرفیت معمولاً واقعاً روی قیمت اثر داره؛ برای مدل‌های
    خیلی قدیمی/ارزون که دیگه فرقی نمی‌کنه، ادمین آگاهانه خاموشش می‌کنه."""
    from fastapi.responses import JSONResponse
    adm = _get_admin(request)
    guard = _require(adm, "ai_pricing")
    if guard: return JSONResponse({"error": "unauthorized"})
    form = await request.form()
    model_id = (form.get("model_id") or "").strip()
    if not model_id.isdigit():
        return JSONResponse({"error": "مدل نامعتبر"})
    enabled = form.get("enabled") == "1"
    import iphone_valuation.db as ivdb
    ivdb.update_model(int(model_id), capacity_pricing=1 if enabled else 0)
    _log(request, "تغییر اثر ظرفیت روی قیمت", "قیمت‌های آیفون",
         f"model:{model_id} -> {'روشن' if enabled else 'خاموش'}", admin_info=adm)
    return JSONResponse({"ok": True})


def _iv_series_options(series_list, selected_id):
    opts = ['<option value="">— بدون سری —</option>']
    for s in series_list:
        sel = " selected" if selected_id == s["id"] else ""
        opts.append(f'<option value="{s["id"]}"{sel}>{html.escape(s["name"])}</option>')
    return "".join(opts)


@router.get("/iphone/series", response_class=HTMLResponse)
async def iphone_series_page(request: Request, flash: str = ""):
    """گروه‌بندی نسل/سری مدل‌ها — مرحلهٔ اول ویزارد ربات از همین‌جا تغذیه می‌شه
    (list_bot_visible_series/list_bot_visible_models در iphone_valuation/db.py).
    CRUD خود سری‌ها + تخصیص هر مدل به یک سری (دراپ‌داون auto-submit)."""
    adm = _get_admin(request)
    guard = _require(adm, "ai_pricing")
    if guard: return guard
    import iphone_valuation.db as ivdb
    series_list = ivdb.list_series(active_only=False)
    models = ivdb.list_models(active_only=False)

    series_rows = "".join(f"""
      <div class="flex flex-wrap gap-2 items-center border-b py-1.5 text-sm">
        <input type="text" name="name_{s['id']}" value="{html.escape(s['name'])}" class="border rounded p-1 flex-1 min-w-[140px] text-xs">
        <input type="number" step="1" name="sort_order_{s['id']}" value="{s['sort_order']}" class="border rounded p-1 w-16 text-xs">
        <label class="text-xs flex items-center gap-1"><input type="checkbox" name="active_{s['id']}" {"checked" if s['active'] else ""}> فعال</label>
        <button type="submit" form="iv-series-del-{s['id']}" class="text-red-500 text-xs">حذف</button>
      </div>""" for s in series_list)
    del_forms = "".join(
        f'<form id="iv-series-del-{s["id"]}" method="post" action="/admin/iphone/series/{s["id"]}/delete" '
        f'onsubmit="return confirm(\'حذف بشه؟ مدل‌های این سری بدون‌سری می‌مونن، حذف نمی‌شن.\')"></form>'
        for s in series_list)

    assign_rows = "".join(f"""
      <div class="flex items-center gap-2 border-b py-1.5 text-sm">
        <span class="flex-1">{html.escape(m['name'])}</span>
        <select name="series_{m['id']}" class="border rounded p-1 text-xs">
          {_iv_series_options(series_list, m.get("series_id"))}
        </select>
      </div>""" for m in models)

    body = f"""
    <a href="/admin/iphone" class="text-indigo-600 text-sm mb-4 inline-block">← بازگشت به کارشناسی آیفون</a>
    <form method="post" action="/admin/iphone/series/bulk-save">
    <div class="bg-white rounded-xl shadow-sm p-4 mb-3">
      <div class="font-medium text-sm mb-2">🗂 سری‌ها/نسل‌ها</div>
      {series_rows or '<div class="text-xs text-gray-400">سری‌ای ثبت نشده.</div>'}
    </div>
    <div class="bg-white rounded-xl shadow-sm p-4 mb-3">
      <div class="font-medium text-sm mb-2">تخصیص مدل به سری</div>
      <input type="text" id="iv-series-search" placeholder="🔍 جست‌وجوی مدل…" class="border rounded-lg p-2 text-xs w-full mb-2">
      <div id="iv-series-assign-list">{assign_rows}</div>
    </div>
    <button class="bg-indigo-600 text-white px-5 py-2.5 rounded-lg text-sm font-medium mb-4">💾 ذخیرهٔ همهٔ تغییرات</button>
    </form>

    <div class="bg-white rounded-xl shadow-sm p-4">
      <div class="font-medium text-sm mb-2">+ افزودن سری تازه</div>
      <form method="post" action="/admin/iphone/series/add" class="flex flex-wrap gap-2 items-center text-sm">
        <input type="text" name="name" class="border rounded p-1 flex-1 min-w-[160px] text-xs" placeholder="نام سری، مثلاً iPhone 11" required>
        <input type="number" step="1" name="sort_order" value="0" class="border rounded p-1 w-16 text-xs" placeholder="ترتیب">
        <button class="bg-indigo-600 text-white px-3 py-1.5 rounded-lg text-xs">+ افزودن سری</button>
      </form>
    </div>
    {del_forms}
    <script>
    document.getElementById('iv-series-search').addEventListener('input', function(){{
      var q = this.value.trim().toLowerCase();
      document.querySelectorAll('#iv-series-assign-list > div').forEach(function(row){{
        row.style.display = row.textContent.toLowerCase().indexOf(q) === -1 ? 'none' : '';
      }});
    }});
    </script>"""
    return _layout("گروه‌بندی نسل‌های آیفون", body, adm, flash=flash)


@router.post("/iphone/series/add")
async def iphone_series_add(request: Request, name: str = Form(...), sort_order: int = Form(0)):
    adm = _get_admin(request)
    guard = _require(adm, "ai_pricing")
    if guard: return guard
    import iphone_valuation.db as ivdb
    ivdb.create_series(name.strip(), sort_order)
    _log(request, "افزودن سری آیفون", "کارشناسی آیفون", name.strip(), admin_info=adm)
    return _redir("/admin/iphone/series")


@router.post("/iphone/series/bulk-save")
async def iphone_series_bulk_save(request: Request):
    """ذخیرهٔ یک‌جای همهٔ ردیف‌های سری + تخصیص مدل→سری با یک دکمهٔ واحد پایین صفحه —
    جایگزین دکمه‌های ذخیرهٔ جداگانهٔ قبلی طبق درخواست صریح مالک پروژه."""
    adm = _get_admin(request)
    guard = _require(adm, "ai_pricing")
    if guard: return guard
    import iphone_valuation.db as ivdb
    form = await request.form()

    for s in ivdb.list_series(active_only=False):
        sid = s["id"]
        name_key = f"name_{sid}"
        if name_key not in form:
            continue
        name = (form.get(name_key) or "").strip() or s["name"]
        try:
            sort_order = int(form.get(f"sort_order_{sid}") or 0)
        except (TypeError, ValueError):
            sort_order = s["sort_order"]
        active = 1 if form.get(f"active_{sid}") else 0
        ivdb.update_series(sid, name=name, sort_order=sort_order, active=active)

    for m in ivdb.list_models(active_only=False):
        key = f"series_{m['id']}"
        if key not in form:
            continue
        val = (form.get(key) or "").strip()
        ivdb.update_model(m["id"], series_id=int(val) if val.isdigit() else None)

    _log(request, "ذخیرهٔ گروهی سری‌ها/تخصیص مدل‌ها", "کارشناسی آیفون", "", admin_info=adm)
    return _redir(f"/admin/iphone/series?flash={e('✅ تغییرات ذخیره شد')}")


@router.post("/iphone/series/{sid}/delete")
async def iphone_series_delete(request: Request, sid: int):
    adm = _get_admin(request)
    guard = _require(adm, "ai_pricing")
    if guard: return guard
    import iphone_valuation.db as ivdb
    ivdb.delete_series(sid)
    _log(request, "حذف سری آیفون", "کارشناسی آیفون", str(sid), admin_info=adm)
    return _redir("/admin/iphone/series")


@router.get("/iphone/repairs", response_class=HTMLResponse)
async def iphone_repairs_page(request: Request, flash: str = ""):
    """مدیریت تعمیرات — طبق درخواست صریح پروژه، یک صفحهٔ واحد برای هر دو بعد («قطعات
    معیوب»/component و «قطعات تعویض‌شده»/replaced) که توی ویزارد ربات دو مولتی‌سلکت
    جدا هستن (محاسبهٔ جدا در pricing_engine) ولی این‌جا برای ادمین یک لیست/فرم مشترکه —
    هر قطعه یک ردیف، با دو درصد کنار هم."""
    adm = _get_admin(request)
    guard = _require(adm, "ai_pricing")
    if guard: return guard
    import iphone_valuation.db as ivdb
    parts = ivdb.list_repair_parts(active_only=False)
    comp_by_key = {c["option_key"]: c for c in ivdb.list_coefficients(category="component", active_only=False)}
    repl_by_key = {c["option_key"]: c for c in ivdb.list_coefficients(category="replaced", active_only=False)}
    weights = ivdb.list_score_weights()

    rows = ""
    del_forms = ""
    for p in parts:
        comp = comp_by_key.get(p["code"])
        repl = repl_by_key.get(p["code"])
        rows += f"""
        <tr class="border-b {'opacity-40' if not p['active'] else ''}">
          <td class="px-2 py-2"><input type="text" name="label_{p['id']}" value="{html.escape(p['label'])}" class="border rounded p-1 w-40 text-xs"></td>
          <td class="px-2 py-2"><input type="number" step="0.1" name="defective_percent_{p['id']}" value="{comp['percent'] if comp else 0}" class="border rounded p-1 w-20 text-xs"></td>
          <td class="px-2 py-2"><input type="number" step="0.1" name="replaced_percent_{p['id']}" value="{repl['percent'] if repl else 0}" class="border rounded p-1 w-20 text-xs"></td>
          <td class="px-2 py-2"><label class="text-xs flex items-center gap-1"><input type="checkbox" name="active_{p['id']}" {"checked" if p['active'] else ""}> فعال</label></td>
          <td class="px-2 py-2 whitespace-nowrap">
            <button type="submit" form="iv-rpd-{p['id']}" class="px-2 py-1 text-xs bg-red-50 text-red-500 border border-red-200 rounded">🗑 حذف</button>
          </td>
        </tr>"""
        del_forms += (
            f'<form id="iv-rpd-{p["id"]}" method="post" action="/admin/iphone/repairs/{p["id"]}/delete" '
            f'onsubmit="return confirm(\'این قطعه حذف بشه؟\')"></form>')

    comp_w = weights.get("component", 0)
    repl_w = weights.get("replaced", 0)
    body = f"""
    <a href="/admin/iphone" class="text-indigo-600 text-sm mb-4 inline-block">← بازگشت به کارشناسی آیفون</a>
    <form method="post" action="/admin/iphone/repairs/bulk-save">
    <div class="bg-white rounded-xl shadow-sm p-4 mb-3">
      <div class="font-medium text-sm mb-2">⚖️ وزن امتیازدهی</div>
      <div class="flex flex-wrap gap-4 text-xs">
        <div class="flex items-center gap-1">🩹 قطعات معیوب: <input type="number" step="1" name="weight_component" value="{comp_w}" class="border rounded p-1 w-14 text-xs"></div>
        <div class="flex items-center gap-1">🔁 قطعات تعویض‌شده: <input type="number" step="1" name="weight_replaced" value="{repl_w}" class="border rounded p-1 w-14 text-xs"></div>
      </div>
    </div>

    <div class="card overflow-hidden mb-4">
      <div class="overflow-x-auto">
        <table class="w-full text-right min-w-max">
          <thead><tr class="text-xs text-gray-500 border-b bg-gray-50">
            <th class="px-2 py-3">قطعه</th><th class="px-2 py-3">٪ معیوب</th><th class="px-2 py-3">٪ تعویض‌شده</th>
            <th class="px-2 py-3">فعال</th><th class="px-2 py-3"></th>
          </tr></thead>
          <tbody>{rows or '<tr><td colspan="5" class="text-center py-8 text-gray-400">هنوز قطعه‌ای ثبت نشده.</td></tr>'}</tbody>
        </table>
      </div>
    </div>
    <button class="bg-indigo-600 text-white px-5 py-2.5 rounded-lg text-sm font-medium mb-4">💾 ذخیرهٔ همهٔ تغییرات</button>
    </form>

    <div class="bg-white rounded-xl shadow-sm p-4 mb-3">
      <div class="font-medium text-sm mb-2">📝 افزودن قطعهٔ تازه</div>
      <form method="post" action="/admin/iphone/repairs/add" class="grid grid-cols-2 sm:grid-cols-5 gap-2 items-end">
        <input type="text" name="code" class="border rounded p-1.5 text-xs" placeholder="کد یکتا مثلاً comp_battery" required>
        <input type="text" name="label" class="border rounded p-1.5 text-xs" placeholder="برچسب فارسی" required>
        <input type="number" step="0.1" name="defective_percent" class="border rounded p-1.5 text-xs" placeholder="٪ معیوب">
        <input type="number" step="0.1" name="replaced_percent" class="border rounded p-1.5 text-xs" placeholder="٪ تعویض‌شده">
        <button class="bg-indigo-600 text-white px-3 py-1.5 rounded-lg text-xs">+ افزودن</button>
      </form>
    </div>
    {del_forms}
    <div class="text-xs text-gray-400 mt-2">این لیست منبع دو مرحلهٔ مولتی‌سلکت جدای ویزارد رباته: «قطعات معیوب» (فقط وقتی وضعیت کلی = نیازمند تعمیر) و «قطعات تعویض‌شده» (همیشه پرسیده می‌شه) — هر قطعه یک درصد جدا برای هرکدوم داره.</div>
    """
    return _layout("مدیریت تعمیرات آیفون", body, adm, flash=flash)


@router.post("/iphone/repairs/add")
async def iphone_repairs_add(request: Request, code: str = Form(...), label: str = Form(...),
                              defective_percent: float = Form(0), replaced_percent: float = Form(0)):
    adm = _get_admin(request)
    guard = _require(adm, "ai_pricing")
    if guard: return guard
    import iphone_valuation.db as ivdb
    ivdb.create_repair_part(code.strip(), label.strip(), defective_percent, replaced_percent)
    _log(request, "افزودن قطعهٔ تعمیر آیفون", "کارشناسی آیفون", code.strip(), admin_info=adm)
    return _redir("/admin/iphone/repairs")


@router.post("/iphone/repairs/bulk-save")
async def iphone_repairs_bulk_save(request: Request):
    """ذخیرهٔ یک‌جای وزن امتیازدهی + همهٔ ردیف‌های قطعات با یک دکمهٔ واحد پایین صفحه."""
    adm = _get_admin(request)
    guard = _require(adm, "ai_pricing")
    if guard: return guard
    import iphone_valuation.db as ivdb
    form = await request.form()

    for cat, field in (("component", "weight_component"), ("replaced", "weight_replaced")):
        if field in form:
            try:
                ivdb.set_score_weight(cat, float(form.get(field) or 0))
            except (TypeError, ValueError):
                pass

    for p in ivdb.list_repair_parts(active_only=False):
        pid = p["id"]
        label_key = f"label_{pid}"
        if label_key not in form:
            continue
        label = (form.get(label_key) or "").strip() or p["label"]
        try:
            defective_percent = float(form.get(f"defective_percent_{pid}") or 0)
        except (TypeError, ValueError):
            defective_percent = 0.0
        try:
            replaced_percent = float(form.get(f"replaced_percent_{pid}") or 0)
        except (TypeError, ValueError):
            replaced_percent = 0.0
        active = 1 if form.get(f"active_{pid}") else 0
        ivdb.update_repair_part(pid, label=label, defective_percent=defective_percent,
                                 replaced_percent=replaced_percent, active=active)

    _log(request, "ذخیرهٔ گروهی تعمیرات آیفون", "کارشناسی آیفون", "", admin_info=adm)
    return _redir(f"/admin/iphone/repairs?flash={e('✅ تغییرات ذخیره شد')}")


@router.post("/iphone/repairs/{pid}/delete")
async def iphone_repairs_delete(request: Request, pid: int):
    adm = _get_admin(request)
    guard = _require(adm, "ai_pricing")
    if guard: return guard
    import iphone_valuation.db as ivdb
    ivdb.delete_repair_part(pid)
    _log(request, "حذف قطعهٔ تعمیر آیفون", "کارشناسی آیفون", str(pid), admin_info=adm)
    return _redir("/admin/iphone/repairs")


@router.get("/iphone/colors", response_class=HTMLResponse)
async def iphone_colors_page(request: Request, flash: str = ""):
    """مدیریت رنگ‌ها — اسم نمایشیِ رنگ پیش‌فرض هر مدل قابل ویرایشه (نه تعریف رنگ تازه؛
    مثلاً «نور ستاره‌ای» → «سفید» چون بازار ایران این‌جوری صداش می‌کنه) + درصد اثر
    این رنگ روی قیمت پایه (مستقل از مکانیزم قدیمی‌تر «ردیف قیمت اختصاصی هر رنگ»)."""
    adm = _get_admin(request)
    guard = _require(adm, "ai_pricing")
    if guard: return guard
    import iphone_valuation.db as ivdb
    models = ivdb.list_models(active_only=False)
    rows = ""
    del_forms = ""
    for m in models:
        colors = ivdb.list_colors(m["id"], active_only=False)
        for c in colors:
            rows += f"""
            <tr class="border-b hover:bg-gray-50 {'opacity-40' if not c['active'] else ''}" data-model="{e(m['name'].lower())}">
              <td class="px-3 py-2 text-xs whitespace-nowrap">{html.escape(m['name'])}</td>
              <td class="px-2 py-2"><input type="text" name="name_{c['id']}" value="{html.escape(c['name'])}" class="border rounded p-1 w-32 text-xs"></td>
              <td class="px-2 py-2"><input type="number" step="0.1" name="price_percent_{c['id']}" value="{c['price_percent']}" class="border rounded p-1 w-20 text-xs"></td>
              <td class="px-2 py-2"><label class="text-xs flex items-center gap-1"><input type="checkbox" name="active_{c['id']}" {"checked" if c['active'] else ""}> فعال</label></td>
              <td class="px-2 py-2 whitespace-nowrap">
                <button type="submit" form="iv-cold-{c['id']}" class="px-2 py-1 text-xs bg-red-50 text-red-500 border border-red-200 rounded">🗑 حذف</button>
              </td>
            </tr>"""
            del_forms += (
                f'<form id="iv-cold-{c["id"]}" method="post" action="/admin/iphone/colors/{c["id"]}/delete" '
                f'onsubmit="return confirm(\'این رنگ حذف بشه؟\')"></form>')

    body = f"""
    <a href="/admin/iphone" class="text-indigo-600 text-sm mb-4 inline-block">← بازگشت به کارشناسی آیفون</a>
    <h1 class="text-xl font-bold text-gray-800 mb-4">🎨 مدیریت رنگ‌ها</h1>
    <div class="bg-white rounded-xl shadow-sm p-3 mb-3">
      <input type="text" id="iv-color-search" placeholder="🔍 جست‌وجو بر اساس نام مدل…" class="border rounded-lg p-2 text-sm w-full">
    </div>
    <form method="post" action="/admin/iphone/colors/bulk-save">
    <div class="card overflow-hidden mb-4">
      <div class="overflow-x-auto">
        <table class="w-full text-right min-w-max" id="iv-color-table">
          <thead><tr class="text-xs text-gray-500 border-b bg-gray-50">
            <th class="px-3 py-3">مدل</th><th class="px-3 py-3">نام رنگ</th>
            <th class="px-3 py-3">٪ اثر روی قیمت</th><th class="px-3 py-3">فعال</th><th class="px-3 py-3"></th>
          </tr></thead>
          <tbody>{rows or '<tr><td colspan="5" class="text-center py-8 text-gray-400">رنگی ثبت نشده.</td></tr>'}</tbody>
        </table>
      </div>
    </div>
    <button class="bg-indigo-600 text-white px-5 py-2.5 rounded-lg text-sm font-medium">💾 ذخیرهٔ همهٔ تغییرات</button>
    </form>
    {del_forms}
    <div class="text-xs text-gray-400 mt-2">این درصد جدا از قیمت اختصاصی هر رنگ (اگه برای مدلی ثبت شده) اضافه می‌شه — اگه برای این مدل قیمت دقیق جداگانه به‌ازای رنگ ثبت کردی، این درصد رو صفر نگه دار تا دوبار حساب نشه.</div>
    <script>
    document.getElementById('iv-color-search').addEventListener('input', function(){{
      var q = this.value.trim().toLowerCase();
      document.querySelectorAll('#iv-color-table tbody tr[data-model]').forEach(function(tr){{
        tr.style.display = tr.dataset.model.indexOf(q) === -1 ? 'none' : '';
      }});
    }});
    </script>
    """
    return _layout("مدیریت رنگ‌ها", body, adm, flash=flash)


@router.post("/iphone/colors/bulk-save")
async def iphone_colors_bulk_save(request: Request):
    """ذخیرهٔ یک‌جای همهٔ ردیف‌های رنگ (هر مدل) با یک دکمهٔ واحد پایین صفحه."""
    adm = _get_admin(request)
    guard = _require(adm, "ai_pricing")
    if guard: return guard
    import iphone_valuation.db as ivdb
    form = await request.form()
    for m in ivdb.list_models(active_only=False):
        for c in ivdb.list_colors(m["id"], active_only=False):
            cid = c["id"]
            name_key = f"name_{cid}"
            if name_key not in form:
                continue
            name = (form.get(name_key) or "").strip() or c["name"]
            price_percent = _iv_parse_num(form.get(f"price_percent_{cid}") or "0", cast=float)
            active = 1 if form.get(f"active_{cid}") else 0
            ivdb.update_color(cid, name=name, price_percent=price_percent, active=active)
    _log(request, "ذخیرهٔ گروهی رنگ‌های آیفون", "کارشناسی آیفون", "", admin_info=adm)
    return _redir(f"/admin/iphone/colors?flash={e('✅ تغییرات ذخیره شد')}")


@router.post("/iphone/colors/{cid}/delete")
async def iphone_colors_delete(request: Request, cid: int):
    adm = _get_admin(request)
    guard = _require(adm, "ai_pricing")
    if guard: return guard
    import iphone_valuation.db as ivdb
    ivdb.delete_color(cid)
    _log(request, "حذف رنگ آیفون", "کارشناسی آیفون", str(cid), admin_info=adm)
    return _redir(f"/admin/iphone/colors?flash={e('✅ رنگ حذف شد')}")


@router.get("/iphone/coefficients", response_class=HTMLResponse)
async def iphone_coefficients_page(request: Request, flash: str = ""):
    adm = _get_admin(request)
    guard = _require(adm, "ai_pricing")
    if guard: return guard
    import iphone_valuation.db as ivdb
    coeffs = ivdb.list_coefficients(active_only=False)
    weights = ivdb.list_score_weights()

    by_cat = {}
    for c in coeffs:
        by_cat.setdefault(c["category"], []).append(c)

    sections = ""
    add_hidden_forms = ""
    del_hidden_forms = ""
    for cat in ivdb.COEFFICIENT_CATEGORIES:
        if cat in ("component", "replaced"):
            # این دو دسته (قطعات معیوب/تعویض‌شده) یک صفحهٔ اختصاصی مشترک دارن
            # (/admin/iphone/repairs) که هر دو درصد رو کنار هم نشون می‌ده — نه این‌جا.
            continue
        rows = "".join(f"""
          <div class="flex flex-wrap gap-2 items-center border-b py-1.5 text-sm">
            <span class="flex-1 min-w-[140px] {'text-gray-400 line-through' if not c['active'] else ''}">{html.escape(c['option_label'])}</span>
            <input type="number" step="0.1" name="percent_{c['id']}" value="{c['percent']}" class="border rounded p-1 w-20 text-xs">
            <span class="text-xs text-gray-400">٪</span>
            <label class="text-xs flex items-center gap-1"><input type="checkbox" name="active_{c['id']}" {"checked" if c['active'] else ""}> فعال</label>
            <button type="submit" form="iv-coef-del-{c['id']}" class="text-red-500 text-xs">حذف</button>
          </div>""" for c in by_cat.get(cat, []))
        for c in by_cat.get(cat, []):
            del_hidden_forms += (
                f'<form id="iv-coef-del-{c["id"]}" method="post" '
                f'action="/admin/iphone/coefficients/{c["id"]}/delete" onsubmit="return confirm(\'حذف بشه؟\')"></form>')
        weight = weights.get(cat, 0)
        sections += f"""
        <div class="bg-white rounded-xl shadow-sm p-4 mb-3">
          <div class="flex items-center justify-between mb-2">
            <div class="font-medium text-sm">{_IV_CATEGORY_LABELS.get(cat, cat)}</div>
            <div class="flex items-center gap-1 text-xs text-gray-400">
              وزن امتیاز: <input type="number" step="1" name="weight_{cat}" value="{weight}" class="border rounded p-1 w-14 text-xs">
            </div>
          </div>
          {rows or '<div class="text-xs text-gray-400">گزینه‌ای ثبت نشده.</div>'}
          <div class="flex flex-wrap gap-2 items-center pt-2 text-sm">
            <input type="hidden" name="category" value="{cat}" form="iv-coef-add-{cat}">
            <input type="text" name="option_key" form="iv-coef-add-{cat}" class="border rounded p-1 w-32 text-xs" placeholder="کلید یکتا مثلاً opt_1" required>
            <input type="text" name="option_label" form="iv-coef-add-{cat}" class="border rounded p-1 flex-1 min-w-[160px] text-xs" placeholder="برچسب فارسی" required>
            <input type="number" step="0.1" name="percent" form="iv-coef-add-{cat}" class="border rounded p-1 w-20 text-xs" placeholder="درصد" required>
            <button type="submit" form="iv-coef-add-{cat}" class="bg-indigo-600 text-white px-3 py-1.5 rounded-lg text-xs">+ افزودن گزینه</button>
          </div>
        </div>"""
        add_hidden_forms += f'<form id="iv-coef-add-{cat}" method="post" action="/admin/iphone/coefficients/add"></form>'

    feat_weight = weights.get("features", 0)
    body = f"""
    <a href="/admin/iphone" class="text-indigo-600 text-sm mb-4 inline-block">← بازگشت به کارشناسی آیفون</a>
    <form method="post" action="/admin/iphone/coefficients/bulk-save">
    <div class="bg-white rounded-xl shadow-sm p-4 mb-3">
      <div class="flex items-center justify-between">
        <div class="font-medium text-sm">تست امکانات دستگاه (سؤال کوتاه ربات)</div>
        <div class="flex items-center gap-1 text-xs text-gray-400">
          وزن امتیاز: <input type="number" step="1" name="weight_features" value="{feat_weight}" class="border rounded p-1 w-14 text-xs">
        </div>
      </div>
      <div class="text-xs text-gray-400 mt-1">اگه کاربر بگه یکی از امکانات (Face ID، دوربین، شارژ بی‌سیم و ...) مشکل داره، نصف این وزن از امتیاز کم می‌شه.</div>
    </div>
    {sections}
    <button class="bg-indigo-600 text-white px-5 py-2.5 rounded-lg text-sm font-medium mb-4">💾 ذخیرهٔ همهٔ تغییرات</button>
    </form>
    {add_hidden_forms}
    {del_hidden_forms}"""
    return _layout("ضرایب قیمت و امتیاز آیفون", body, adm, flash=flash)


@router.post("/iphone/coefficients/add")
async def iphone_coefficients_add(request: Request, category: str = Form(...), option_key: str = Form(...),
                                   option_label: str = Form(...), percent: float = Form(...)):
    adm = _get_admin(request)
    guard = _require(adm, "ai_pricing")
    if guard: return guard
    import iphone_valuation.db as ivdb
    ivdb.create_coefficient(category, option_key.strip(), option_label.strip(), percent)
    _log(request, "افزودن ضریب آیفون", "کارشناسی آیفون", f"{category}:{option_key}", admin_info=adm)
    return _redir("/admin/iphone/coefficients")


@router.post("/iphone/coefficients/bulk-save")
async def iphone_coefficients_bulk_save(request: Request):
    """ذخیرهٔ یک‌جای وزن‌های امتیازدهی + همهٔ ردیف‌های ضریب همهٔ دسته‌ها با یک دکمهٔ
    واحد پایین صفحه."""
    adm = _get_admin(request)
    guard = _require(adm, "ai_pricing")
    if guard: return guard
    import iphone_valuation.db as ivdb
    form = await request.form()

    for cat in list(ivdb.COEFFICIENT_CATEGORIES) + ["features"]:
        field = f"weight_{cat}"
        if field in form:
            try:
                ivdb.set_score_weight(cat, float(form.get(field) or 0))
            except (TypeError, ValueError):
                pass

    for c in ivdb.list_coefficients(active_only=False):
        cid = c["id"]
        percent_key = f"percent_{cid}"
        if percent_key not in form:
            continue
        try:
            percent = float(form.get(percent_key) or 0)
        except (TypeError, ValueError):
            percent = c["percent"]
        active = 1 if form.get(f"active_{cid}") else 0
        ivdb.update_coefficient(cid, percent=percent, active=active)

    _log(request, "ذخیرهٔ گروهی ضرایب آیفون", "کارشناسی آیفون", "", admin_info=adm)
    return _redir(f"/admin/iphone/coefficients?flash={e('✅ تغییرات ذخیره شد')}")


@router.post("/iphone/coefficients/{cid}/delete")
async def iphone_coefficients_delete(request: Request, cid: int):
    adm = _get_admin(request)
    guard = _require(adm, "ai_pricing")
    if guard: return guard
    import iphone_valuation.db as ivdb
    ivdb.delete_coefficient(cid)
    _log(request, "حذف ضریب آیفون", "کارشناسی آیفون", str(cid), admin_info=adm)
    return _redir("/admin/iphone/coefficients")


@router.get("/iphone/fx", response_class=HTMLResponse)
async def iphone_fx_page(request: Request, flash: str = ""):
    adm = _get_admin(request)
    guard = _require(adm, "ai_pricing")
    if guard: return guard
    from db import get_cfg
    import iphone_valuation.db as ivdb
    sources = ivdb.list_fx_sources()
    mode = get_cfg("IV_FX_MODE", "auto")
    manual_rate = get_cfg("IV_FX_MANUAL_RATE", "0")
    sensitivity = get_cfg("IV_FX_SENSITIVITY", "0.5")
    market_weight = get_cfg("IV_MARKET_DATA_WEIGHT", "0.15")
    last_good = get_cfg("IV_FX_LAST_GOOD", "0")

    src_rows = "".join(f"""
      <div class="border-b py-2">
        <div class="flex flex-wrap gap-2 items-center text-sm">
          <input type="text" name="name_{s['id']}" value="{e(s['name'])}" class="border rounded p-1 w-32 text-xs">
          <input type="text" name="url_{s['id']}" value="{e(s['url'])}" class="border rounded p-1 flex-1 min-w-[220px] text-xs" dir="ltr">
          <input type="text" name="json_path_{s['id']}" value="{e(s['json_path'])}" class="border rounded p-1 w-28 text-xs" dir="ltr" placeholder="usd.sell">
          <input type="number" name="priority_{s['id']}" value="{s['priority']}" class="border rounded p-1 w-14 text-xs">
          <label class="text-xs flex items-center gap-1"><input type="checkbox" name="active_{s['id']}" {"checked" if s['active'] else ""}> فعال</label>
        </div>
        <div class="text-xs text-gray-400 mt-1 flex items-center gap-3 flex-wrap">
          <span>آخرین مقدار: {_iv_num(s['last_value'])}</span>
          <span>آخرین فچ: {fa_date(s['last_fetched_at'], with_time=True) if s['last_fetched_at'] else '—'}</span>
          {f'<span class="text-red-500">خطا: {html.escape(s["last_error"])}</span>' if s['last_error'] else ''}
          <button type="submit" form="iv-fx-test-{s['id']}" class="text-indigo-600">تست فچ</button>
          <button type="submit" form="iv-fx-del-{s['id']}" class="text-red-500">حذف</button>
        </div>
      </div>""" for s in sources) or '<div class="text-xs text-gray-400 py-2">هنوز منبعی ثبت نشده.</div>'
    test_del_forms = "".join(
        f'<form id="iv-fx-test-{s["id"]}" method="post" action="/admin/iphone/fx/{s["id"]}/test"></form>'
        f'<form id="iv-fx-del-{s["id"]}" method="post" action="/admin/iphone/fx/{s["id"]}/delete" '
        f'onsubmit="return confirm(\'حذف بشه؟\')"></form>' for s in sources)

    body = f"""
    <a href="/admin/iphone" class="text-indigo-600 text-sm mb-4 inline-block">← بازگشت به کارشناسی آیفون</a>
    <div class="bg-white rounded-xl shadow-sm p-4 mb-4 max-w-2xl">
      <div class="text-sm font-medium mb-2">تنظیمات نرخ ارز</div>
      <form method="post" action="/admin/iphone/fx/settings" class="space-y-2 text-sm">
        <div class="flex items-center gap-3">
          <label class="flex items-center gap-1"><input type="radio" name="mode" value="auto" {"checked" if mode!="manual" else ""}> خودکار از منابع</label>
          <label class="flex items-center gap-1"><input type="radio" name="mode" value="manual" {"checked" if mode=="manual" else ""}> دستی</label>
        </div>
        <div class="flex items-center gap-2">
          <span class="text-xs text-gray-400 w-40">نرخ دستی (تومان)</span>
          <input type="number" name="manual_rate" value="{manual_rate}" class="border rounded p-1.5 text-xs w-32">
        </div>
        <div class="flex items-center gap-2">
          <span class="text-xs text-gray-400 w-40">حساسیت نوسان ارز</span>
          <input type="number" step="0.05" name="sensitivity" value="{sensitivity}" class="border rounded p-1.5 text-xs w-32">
          <span class="text-xs text-gray-400">۰ تا ۱ — چقدر قیمت آیفون از نوسان دلار تاثیر بگیره</span>
        </div>
        <div class="flex items-center gap-2">
          <span class="text-xs text-gray-400 w-40">وزن دادهٔ بازار استوک‌لند</span>
          <input type="number" step="0.05" name="market_weight" value="{market_weight}" class="border rounded p-1.5 text-xs w-32">
        </div>
        <button class="bg-indigo-600 text-white px-4 py-2 rounded-lg text-xs">ذخیره تنظیمات</button>
      </form>
      <div class="text-xs text-gray-400 mt-3">آخرین نرخ معتبر ثبت‌شده (فال‌بک اگه همهٔ منابع قطع باشن): {_iv_num(int(last_good) if str(last_good).isdigit() else 0)} تومان</div>
    </div>

    <div class="bg-white rounded-xl shadow-sm p-4 mb-3 max-w-2xl">
      <div class="text-sm font-medium mb-2">+ افزودن منبع نرخ ارز</div>
      <form method="post" action="/admin/iphone/fx/add" class="flex flex-wrap gap-2 text-sm">
        <input type="text" name="name" class="border rounded p-1.5 w-32 text-xs" placeholder="نام منبع" required>
        <input type="text" name="url" class="border rounded p-1.5 flex-1 min-w-[220px] text-xs" dir="ltr" placeholder="https://..." required>
        <input type="text" name="json_path" class="border rounded p-1.5 w-28 text-xs" dir="ltr" placeholder="مسیر فیلد مثلاً usd.sell">
        <input type="number" name="priority" value="1" class="border rounded p-1.5 w-16 text-xs" placeholder="اولویت">
        <button class="bg-indigo-600 text-white px-3 py-1.5 rounded-lg text-xs">+ افزودن</button>
      </form>
      <div class="text-xs text-gray-400 mt-2">مسیر فیلد یعنی کجای پاسخ JSON منبع، عدد نرخ رو نگه داشته — مثلاً اگه پاسخ شبیه usd → sell باشه، مسیر را به‌صورت usd.sell بنویس.</div>
    </div>

    <form method="post" action="/admin/iphone/fx/bulk-save">
    <div class="bg-white rounded-xl shadow-sm p-4 max-w-2xl mb-4">
      <div class="text-sm font-medium mb-2">منابع ثبت‌شده</div>
      {src_rows}
    </div>
    <button class="bg-indigo-600 text-white px-5 py-2.5 rounded-lg text-sm font-medium max-w-2xl">💾 ذخیرهٔ همهٔ تغییرات</button>
    </form>
    {test_del_forms}"""
    return _layout("نرخ ارز کارشناسی آیفون", body, adm, flash=flash)


@router.post("/iphone/fx/settings")
async def iphone_fx_settings(request: Request, mode: str = Form("auto"), manual_rate: int = Form(0),
                              sensitivity: float = Form(0.5), market_weight: float = Form(0.15)):
    adm = _get_admin(request)
    guard = _require(adm, "ai_pricing")
    if guard: return guard
    from db import set_cfg
    set_cfg("IV_FX_MODE", mode if mode in ("auto", "manual") else "auto")
    set_cfg("IV_FX_MANUAL_RATE", str(max(0, manual_rate)))
    set_cfg("IV_FX_SENSITIVITY", str(sensitivity))
    set_cfg("IV_MARKET_DATA_WEIGHT", str(market_weight))
    _log(request, "تنظیمات نرخ ارز کارشناسی آیفون", "کارشناسی آیفون", mode, admin_info=adm)
    return _redir("/admin/iphone/fx")


@router.post("/iphone/fx/add")
async def iphone_fx_add(request: Request, name: str = Form(...), url: str = Form(...),
                         json_path: str = Form("price"), priority: int = Form(1)):
    adm = _get_admin(request)
    guard = _require(adm, "ai_pricing")
    if guard: return guard
    import iphone_valuation.db as ivdb
    ivdb.create_fx_source(name, url, json_path or "price", priority)
    _log(request, "افزودن منبع نرخ ارز", "کارشناسی آیفون", name, admin_info=adm)
    return _redir("/admin/iphone/fx")


@router.post("/iphone/fx/bulk-save")
async def iphone_fx_bulk_save(request: Request):
    """ذخیرهٔ یک‌جای همهٔ ردیف‌های منابع نرخ ارز با یک دکمهٔ واحد پایین لیست."""
    adm = _get_admin(request)
    guard = _require(adm, "ai_pricing")
    if guard: return guard
    import iphone_valuation.db as ivdb
    form = await request.form()
    for s in ivdb.list_fx_sources():
        sid = s["id"]
        name_key = f"name_{sid}"
        if name_key not in form:
            continue
        name = (form.get(name_key) or "").strip() or s["name"]
        url = (form.get(f"url_{sid}") or "").strip() or s["url"]
        json_path = (form.get(f"json_path_{sid}") or "").strip() or "price"
        try:
            priority = int(form.get(f"priority_{sid}") or s["priority"])
        except (TypeError, ValueError):
            priority = s["priority"]
        active = 1 if form.get(f"active_{sid}") else 0
        ivdb.update_fx_source(sid, name=name, url=url, json_path=json_path, priority=priority, active=active)
    _log(request, "ذخیرهٔ گروهی منابع نرخ ارز", "کارشناسی آیفون", "", admin_info=adm)
    return _redir(f"/admin/iphone/fx?flash={e('✅ تغییرات ذخیره شد')}")


@router.post("/iphone/fx/{sid}/delete")
async def iphone_fx_delete(request: Request, sid: int):
    adm = _get_admin(request)
    guard = _require(adm, "ai_pricing")
    if guard: return guard
    import iphone_valuation.db as ivdb
    ivdb.delete_fx_source(sid)
    _log(request, "حذف منبع نرخ ارز", "کارشناسی آیفون", str(sid), admin_info=adm)
    return _redir("/admin/iphone/fx")


@router.post("/iphone/fx/{sid}/test")
async def iphone_fx_test(request: Request, sid: int):
    adm = _get_admin(request)
    guard = _require(adm, "ai_pricing")
    if guard: return guard
    import iphone_valuation.db as ivdb
    import iphone_valuation.fx as ivfx
    sources = {s["id"]: s for s in ivdb.list_fx_sources()}
    src = sources.get(sid)
    if src:
        val = ivfx.fetch_source(src)
        msg = ('✅ نتیجه: ' + str(val)) if val else '❌ فچ ناموفق بود، خطا رو زیر همون منبع ببین'
        return _redir(f"/admin/iphone/fx?flash={e(msg)}")
    return _redir("/admin/iphone/fx")


@router.get("/iphone/ai", response_class=HTMLResponse)
async def iphone_ai_page(request: Request, flash: str = ""):
    """پنل کارشناس مکمل هوش مصنوعی (بخش ۲۲.۸ CLAUDE.md) — توگل روشن/خاموش، انتخاب
    provider (فعلاً فقط Claude، ولی دراپ‌داون آمادهٔ افزودن بعدیه)، کلید API (فقط نوشتنی،
    هیچ‌وقت مقدار واقعی نمایش داده نمی‌شه)، بازهٔ مجاز تعدیل و آستانهٔ اطمینان، و لاگ
    آخرین تحلیل‌ها برای ممیزی/پیگیری هزینه."""
    adm = _get_admin(request)
    guard = _require(adm, "ai_pricing")
    if guard: return guard
    from db import get_cfg
    import iphone_valuation.db as ivdb
    import iphone_valuation.ai_crypto as ai_crypto
    from iphone_valuation.ai_providers import AI_PROVIDERS

    enabled = get_cfg("IV_AI_ENABLED", "0") == "1"
    provider = get_cfg("IV_AI_PROVIDER", "claude") or "claude"
    model_name = get_cfg("IV_AI_MODEL", "")
    max_adjust = get_cfg("IV_AI_MAX_ADJUST_PCT", "5") or "5"
    min_confidence = get_cfg("IV_AI_MIN_CONFIDENCE", "60") or "60"
    has_key = bool(get_cfg("IV_AI_API_KEY_ENC", ""))
    crypto_ready = ai_crypto.crypto_available()

    from iphone_valuation.ai_providers import PROVIDER_LABELS
    provider_opts = "".join(
        f'<option value="{e(p)}" {"selected" if p == provider else ""}>{e(PROVIDER_LABELS.get(p, p))}</option>'
        for p in AI_PROVIDERS.keys())

    warn_banner = "" if crypto_ready else """
    <div class="bg-red-50 border border-red-200 text-red-700 rounded-xl p-3 mb-4 text-sm">
      ⚠️ متغیر محیطی <code>IV_AI_ENC_KEY</code> روی سرور تنظیم نشده — بدون این متغیر، ذخیرهٔ
      رمزنگاری‌شدهٔ کلید API ممکن نیست و کل قابلیت AI غیرفعال می‌مونه (fail closed).
    </div>"""

    key_status = ('✅ کلید API ثبت شده (مقدار واقعی هیچ‌وقت نمایش داده نمی‌شه)' if has_key
                  else '❌ هنوز کلیدی ثبت نشده')

    logs = ivdb.list_ai_analyses(limit=50)
    log_rows = "".join(f"""
      <tr class="border-b {'bg-red-50' if r['error'] else ''}">
        <td class="p-2 text-xs whitespace-nowrap">{html.escape((r['model_name_join'] or '—') + ' ' + (r['capacity_label'] or ''))}</td>
        <td class="p-2 text-xs">{r['confidence']}</td>
        <td class="p-2 text-xs">{r['adjustment_percent']}٪</td>
        <td class="p-2 text-xs">{_iv_num(r['final_price']) if r['final_price'] else '—'}</td>
        <td class="p-2 text-xs text-amber-600">{html.escape(', '.join(json.loads(r['warnings'] or '[]')))}</td>
        <td class="p-2 text-xs text-red-500">{html.escape(r['error'] or '')}</td>
        <td class="p-2 text-xs text-gray-400 whitespace-nowrap">{fa_date(r['created_at'], with_time=True)}</td>
      </tr>""" for r in logs) or '<tr><td colspan="7" class="p-6 text-center text-gray-400 text-sm">هنوز تحلیلی ثبت نشده.</td></tr>'

    body = f"""
    <a href="/admin/iphone" class="text-indigo-600 text-sm mb-4 inline-block">← بازگشت به کارشناسی آیفون</a>
    <h1 class="text-xl font-bold text-gray-800 mb-4">🤖 کارشناس مکمل هوش مصنوعی</h1>
    {warn_banner}
    <div class="bg-white rounded-xl shadow-sm p-4 mb-4 max-w-2xl">
      <div class="flex items-center gap-3 bg-gray-50 rounded-lg px-3 py-3 mb-4">
        <span class="text-sm text-gray-700 flex-1">فعال‌سازی کارشناس مکمل هوش مصنوعی — وقتی خاموشه،
          سیستم قیمت‌گذاری دستی (موتور دیتامحور) دقیقاً مثل قبل و بدون هیچ تغییری کار می‌کنه.</span>
        <form method="post" action="/admin/iphone/ai/toggle">
          <button class="px-4 py-2 rounded-lg text-sm font-medium {'bg-red-50 text-red-600' if enabled else 'bg-emerald-50 text-emerald-600'}">
            {'⛔️ غیرفعال کردن' if enabled else '✅ فعال کردن'}
          </button>
        </form>
      </div>

      <form method="post" action="/admin/iphone/ai/settings" class="space-y-3 text-sm">
        <div class="flex items-center gap-2">
          <span class="text-xs text-gray-400 w-40">Provider</span>
          <select name="provider" class="border rounded p-1.5 text-xs">{provider_opts}</select>
        </div>
        <div class="flex items-center gap-2">
          <span class="text-xs text-gray-400 w-40">مدل</span>
          <input type="text" name="model_name" value="{e(model_name)}" dir="ltr" class="border rounded p-1.5 text-xs w-52" placeholder="خالی = پیش‌فرض همون provider">
        </div>
        <div class="flex items-center gap-2">
          <span class="text-xs text-gray-400 w-40">حداکثر بازهٔ مجاز تعدیل قیمت</span>
          <input type="text" inputmode="decimal" dir="ltr" name="max_adjust_pct" value="{e(str(max_adjust))}" class="border rounded p-1.5 text-xs w-24">
          <span class="text-xs text-gray-400">٪ (مثلاً ۵ یعنی حداکثر ±۵٪)</span>
        </div>
        <div class="flex items-center gap-2">
          <span class="text-xs text-gray-400 w-40">حداقل آستانهٔ اطمینان</span>
          <input type="text" inputmode="numeric" dir="ltr" name="min_confidence" value="{e(str(min_confidence))}" class="border rounded p-1.5 text-xs w-24">
          <span class="text-xs text-gray-400">٪ — پایین‌تر از این یعنی هیچ تعدیلی روی قیمت اعمال نشه</span>
        </div>
        <button class="bg-indigo-600 text-white px-4 py-2 rounded-lg text-xs">ذخیره تنظیمات</button>
      </form>
    </div>

    <div class="bg-white rounded-xl shadow-sm p-4 mb-4 max-w-2xl">
      <div class="text-sm font-medium mb-1">کلید API</div>
      <div class="text-xs text-gray-500 mb-3">{key_status}</div>
      <form method="post" action="/admin/iphone/ai/api-key" class="flex flex-wrap gap-2 text-sm">
        <input type="password" name="api_key" dir="ltr" class="border rounded p-1.5 text-xs flex-1 min-w-[240px]" placeholder="sk-ant-api03-...">
        <button class="bg-indigo-600 text-white px-3 py-1.5 rounded-lg text-xs">ذخیرهٔ رمزنگاری‌شده</button>
      </form>
      {'<form method="post" action="/admin/iphone/ai/api-key/clear" class="mt-2" onsubmit="return confirm(\'کلید حذف بشه؟\')"><button class="text-red-500 text-xs">🗑 حذف کلید</button></form>' if has_key else ''}
    </div>

    <div class="bg-white rounded-xl shadow-sm w-full">
      <div class="px-4 py-3 border-b bg-gray-50"><h2 class="font-bold text-gray-700 text-sm">📋 آخرین تحلیل‌ها</h2></div>
      <div class="overflow-x-auto">
        <table class="w-full text-right min-w-max">
          <thead><tr class="text-xs text-gray-400 border-b bg-gray-50">
            <th class="p-2">مدل/ظرفیت</th><th class="p-2">اطمینان</th><th class="p-2">تعدیل</th>
            <th class="p-2">قیمت نهایی</th><th class="p-2">هشدارها</th><th class="p-2">خطا</th><th class="p-2">تاریخ</th>
          </tr></thead>
          <tbody>{log_rows}</tbody>
        </table>
      </div>
    </div>"""
    return _layout("کارشناس مکمل هوش مصنوعی", body, adm, flash=flash)


@router.post("/iphone/ai/toggle")
async def iphone_ai_toggle(request: Request):
    adm = _get_admin(request)
    guard = _require(adm, "ai_pricing")
    if guard: return guard
    from db import get_cfg, set_cfg
    import iphone_valuation.ai_crypto as ai_crypto
    cur = get_cfg("IV_AI_ENABLED", "0") == "1"
    if not cur and not ai_crypto.crypto_available():
        return _redir(f"/admin/iphone/ai?flash={e('⚠️ اول IV_AI_ENC_KEY رو روی سرور تنظیم کن')}")
    set_cfg("IV_AI_ENABLED", "0" if cur else "1")
    _log(request, "فعال/غیرفعال‌سازی کارشناس AI", "کارشناسی آیفون", "غیرفعال" if cur else "فعال", admin_info=adm)
    return _redir("/admin/iphone/ai")


@router.post("/iphone/ai/settings")
async def iphone_ai_settings(request: Request, provider: str = Form("claude"), model_name: str = Form(""),
                              max_adjust_pct: str = Form("5"), min_confidence: str = Form("60")):
    adm = _get_admin(request)
    guard = _require(adm, "ai_pricing")
    if guard: return guard
    from db import set_cfg
    from iphone_valuation.ai_providers import AI_PROVIDERS
    set_cfg("IV_AI_PROVIDER", provider if provider in AI_PROVIDERS else "claude")
    set_cfg("IV_AI_MODEL", (model_name or "").strip())
    try:
        max_adjust = abs(float(max_adjust_pct))
    except (TypeError, ValueError):
        max_adjust = 5.0
    try:
        min_conf = max(0, min(100, int(float(min_confidence))))
    except (TypeError, ValueError):
        min_conf = 60
    set_cfg("IV_AI_MAX_ADJUST_PCT", str(max_adjust))
    set_cfg("IV_AI_MIN_CONFIDENCE", str(min_conf))
    _log(request, "تنظیمات کارشناس AI", "کارشناسی آیفون",
         f"provider={provider} max_adjust={max_adjust} min_conf={min_conf}", admin_info=adm)
    return _redir("/admin/iphone/ai")


@router.post("/iphone/ai/api-key")
async def iphone_ai_api_key(request: Request, api_key: str = Form("")):
    adm = _get_admin(request)
    guard = _require(adm, "ai_pricing")
    if guard: return guard
    import iphone_valuation.ai_crypto as ai_crypto
    if not ai_crypto.crypto_available():
        return _redir(f"/admin/iphone/ai?flash={e('⚠️ اول IV_AI_ENC_KEY رو روی سرور تنظیم کن')}")
    api_key = (api_key or "").strip()
    if not api_key:
        return _redir(f"/admin/iphone/ai?flash={e('کلید خالی بود، ذخیره نشد')}")
    from db import set_cfg
    set_cfg("IV_AI_API_KEY_ENC", ai_crypto.encrypt_api_key(api_key))
    _log(request, "ثبت کلید API کارشناس AI", "کارشناسی آیفون", "•••• (مقدار در لاگ ثبت نمی‌شه)", admin_info=adm)
    return _redir(f"/admin/iphone/ai?flash={e('✅ کلید ذخیره شد')}")


@router.post("/iphone/ai/api-key/clear")
async def iphone_ai_api_key_clear(request: Request):
    adm = _get_admin(request)
    guard = _require(adm, "ai_pricing")
    if guard: return guard
    from db import set_cfg
    set_cfg("IV_AI_API_KEY_ENC", "")
    _log(request, "حذف کلید API کارشناس AI", "کارشناسی آیفون", "", admin_info=adm)
    return _redir(f"/admin/iphone/ai?flash={e('🗑 کلید حذف شد')}")


@router.get("/iphone/history", response_class=HTMLResponse)
async def iphone_history_page(request: Request, flash: str = ""):
    adm = _get_admin(request)
    guard = _require(adm, "ai_pricing")
    if guard: return guard
    import iphone_valuation.db as ivdb
    items = ivdb.list_valuations(limit=100)

    rows = "".join(f"""
      <tr class="border-b">
        <td class="p-2 text-sm">{html.escape(((r['model_name'] or '—') + ' ' + (r['capacity_label'] or '')).strip())}</td>
        <td class="p-2 text-sm">{_iv_num(r['fair_price'])} تومان</td>
        <td class="p-2 text-sm">{_iv_num(r['seller_price'])}</td>
        <td class="p-2 text-sm">{r['score']}</td>
        <td class="p-2 text-sm">{html.escape(r['verdict'] or '')}</td>
        <td class="p-2 text-xs text-gray-400">{fa_date(r['created_at'], with_time=True)}</td>
      </tr>""" for r in items) or '<tr><td colspan="6" class="p-6 text-center text-gray-400 text-sm">هنوز کارشناسی‌ای ثبت نشده.</td></tr>'

    body = f"""
    <a href="/admin/iphone" class="text-indigo-600 text-sm mb-4 inline-block">← بازگشت به کارشناسی آیفون</a>
    <div class="bg-white rounded-xl shadow-sm overflow-x-auto">
      <div class="px-4 py-3 border-b bg-gray-50"><h2 class="font-bold text-gray-700 text-sm">🕓 تاریخچهٔ کارشناسی‌ها ({len(items)})</h2></div>
      <table class="w-full text-right">
        <thead><tr class="text-xs text-gray-400 border-b">
          <th class="p-2">دستگاه</th><th class="p-2">قیمت منصفانه</th><th class="p-2">پیشنهاد فروشنده</th>
          <th class="p-2">امتیاز</th><th class="p-2">نتیجه</th><th class="p-2">تاریخ</th>
        </tr></thead>
        <tbody>{rows}</tbody>
      </table>
    </div>"""
    return _layout("تاریخچهٔ کارشناسی آیفون", body, adm, flash=flash)


# ═══════════════════════════════════════════════════════════════════════════
# درگاه‌های پرداخت آنلاین (چند‌درگاهی، مدیریت از پنل، failover خودکار)
# منطق هر درگاه در پکیج payment_gateways/ است. این‌جا فقط UI مدیریت + تست.
# ═══════════════════════════════════════════════════════════════════════════
@router.get("/payment-gateways", response_class=HTMLResponse)
async def payment_gateways_page(request: Request, flash: str = ""):
    adm = _get_admin(request)
    guard = _require(adm, "payment_gateways")
    if guard: return guard
    import db as _db
    from payment_gateways import PAYMENT_GATEWAYS
    saved = {g["gateway"]: g for g in _db.list_payment_gateways()}

    rows = ""
    hidden_forms = ""
    for code, meta in PAYMENT_GATEWAYS.items():
        cfg = saved.get(code) or {"enabled": 0, "priority": 100, "credentials": {}, "sandbox": 0}
        checked = "checked" if cfg["enabled"] else ""
        sb_checked = "checked" if cfg["sandbox"] else ""
        creds = cfg.get("credentials") or {}
        cred_html = ""
        for fkey, flabel in meta["fields"]:
            has = bool(str(creds.get(fkey) or "").strip())
            ph = "•••• ثبت شده (برای تغییر مقدار تازه وارد کن)" if has else "وارد کن"
            cred_html += (
                f'<div class="flex flex-col gap-1">'
                f'<span class="text-xs text-gray-500">{html.escape(flabel)}</span>'
                f'<input type="password" autocomplete="new-password" name="cred_{code}_{fkey}" '
                f'placeholder="{ph}" class="border rounded p-2 text-sm" dir="ltr"></div>')
        sandbox_html = ""
        if meta.get("supports_sandbox"):
            sandbox_html = (
                f'<label class="flex items-center gap-2 text-xs text-gray-600 mt-2">'
                f'<input type="checkbox" name="sandbox_{code}" {sb_checked}> حالت تست (Sandbox)</label>')
        rows += f"""
        <div class="bg-white rounded-xl shadow-sm p-4 border">
          <div class="flex items-center justify-between mb-3">
            <div class="font-bold text-gray-800">{html.escape(meta['label'])}</div>
            <label class="relative inline-flex items-center cursor-pointer shrink-0">
              <input type="checkbox" name="enabled_{code}" class="sr-only peer" {checked}>
              <div class="w-11 h-6 bg-gray-200 rounded-full peer transition-colors peer-checked:bg-indigo-600
                          after:content-[''] after:absolute after:top-[2px] after:start-[2px] after:bg-white
                          after:rounded-full after:h-5 after:w-5 after:transition-all after:shadow
                          peer-checked:after:translate-x-full rtl:peer-checked:after:-translate-x-full"></div>
            </label>
          </div>
          <div class="grid grid-cols-1 sm:grid-cols-2 gap-3">
            <div class="flex flex-col gap-1">
              <span class="text-xs text-gray-500">اولویت (کوچک‌تر = زودتر امتحان می‌شه)</span>
              <input type="text" inputmode="numeric" dir="ltr" name="priority_{code}" value="{cfg['priority']}" class="border rounded p-2 text-sm">
            </div>
            {cred_html}
          </div>
          {sandbox_html}
          <div class="mt-3">
            <button type="submit" form="pgtest-{code}" class="px-3 py-1.5 text-xs bg-emerald-50 text-emerald-700 border border-emerald-200 rounded">🔌 تست اتصال</button>
          </div>
        </div>"""
        hidden_forms += (
            f'<form id="pgtest-{code}" method="post" action="/admin/payment-gateways/test">'
            f'<input type="hidden" name="gateway" value="{code}"></form>')

    body = f"""
    <div class="flex items-center justify-between mb-4 flex-wrap gap-2">
      <h1 class="text-xl font-bold text-gray-800">💳 درگاه‌های پرداخت</h1>
    </div>
    <div class="bg-blue-50 border border-blue-200 rounded-lg p-3 mb-4 text-xs text-blue-800 leading-6">
      درخواست پرداخت اول به درگاهِ فعالِ با <b>کمترین عدد اولویت</b> می‌ره؛ اگه خطا داد، خودکار سراغ درگاه بعدی می‌ره (failover).
      برای failover واقعی حداقل دو درگاه فعال لازمه. کلیدها فقط ذخیره می‌شن و هیچ‌وقت دوباره نمایش داده نمی‌شن.
      <br>⚠️ درگاه‌های غیر از زرین‌پال قبل از فعال‌سازی حتماً با دکمهٔ «تست اتصال» با کلید واقعی بررسی بشن.
      اگه هیچ درگاهی این‌جا فعال نباشه، سیستم به‌صورت خودکار از زرین‌پالِ تنظیم‌شده در فایل env استفاده می‌کنه (رفتار قبلی).
    </div>
    <form method="post" action="/admin/payment-gateways/save" class="space-y-4">
      {rows}
      <button class="bg-indigo-600 text-white px-5 py-2.5 rounded-lg text-sm font-medium">💾 ذخیرهٔ همهٔ تغییرات</button>
    </form>
    {hidden_forms}
    """
    return _layout("درگاه‌های پرداخت", body, adm, flash=flash)


@router.post("/payment-gateways/save")
async def payment_gateways_save(request: Request):
    adm = _get_admin(request)
    guard = _require(adm, "payment_gateways")
    if guard: return guard
    import db as _db
    from payment_gateways import PAYMENT_GATEWAYS
    form = await request.form()
    saved = {g["gateway"]: g for g in _db.list_payment_gateways()}
    for code, meta in PAYMENT_GATEWAYS.items():
        enabled = 1 if form.get(f"enabled_{code}") else 0
        try:
            priority = int(str(form.get(f"priority_{code}") or "100").strip() or "100")
        except ValueError:
            priority = 100
        sandbox = 1 if (meta.get("supports_sandbox") and form.get(f"sandbox_{code}")) else 0
        # کلیدها: فقط وقتی مقدار تازهٔ غیرخالی وارد شده باشه بازنویسی می‌شن — یعنی ادمین
        # لازم نیست هر بار ذخیره، همهٔ کلیدها رو دوباره تایپ کنه (وگرنه با هر save پاک می‌شدن).
        existing_creds = dict((saved.get(code) or {}).get("credentials") or {})
        for fkey, _ in meta["fields"]:
            val = str(form.get(f"cred_{code}_{fkey}") or "").strip()
            if val:
                existing_creds[fkey] = val
        _db.save_payment_gateway(code, enabled, priority, existing_creds, sandbox)
    _log(request, "ذخیرهٔ تنظیمات درگاه‌های پرداخت", "درگاه‌های پرداخت", "", admin_info=adm)
    return _redir(f"/admin/payment-gateways?flash={e('✅ تنظیمات درگاه‌ها ذخیره شد')}")


@router.post("/payment-gateways/test")
async def payment_gateways_test(request: Request):
    """تست اتصال یک درگاه — یه درخواست ساخت پرداخت آزمایشی می‌زنه (تراکنش واقعی در دیتابیس
    ما ثبت نمی‌شه؛ فقط چک می‌کنه که با کلید فعلی، درگاه لینک پرداخت می‌سازه یا خطا می‌ده)."""
    import os
    adm = _get_admin(request)
    guard = _require(adm, "payment_gateways")
    if guard: return guard
    import db as _db
    from payment_gateways import get_gateway_module, gateway_label
    form = await request.form()
    code = str(form.get("gateway") or "").strip()
    mod = get_gateway_module(code)
    row = _db.get_payment_gateway(code)
    if not mod or not row:
        return _redir(f"/admin/payment-gateways?flash={e('ابتدا تنظیمات این درگاه را ذخیره کن')}")
    cfg = dict(row.get("credentials") or {})
    cfg["sandbox"] = bool(row.get("sandbox"))
    base_cb = (os.getenv("BASE_CALLBACK_URL") or "https://panel.stland.ir/payment/callback").rstrip("/")
    try:
        res = mod.create_payment(10000, base_cb + "/" + code, "تست اتصال درگاه استوک‌لند", cfg)
    except Exception as exc:
        res = {"ok": False, "error": str(exc)}
    if res.get("ok"):
        msg = f"✅ {gateway_label(code)}: اتصال موفق بود (لینک پرداخت آزمایشی ساخته شد)"
    else:
        msg = f"❌ {gateway_label(code)}: {str(res.get('error') or 'خطای نامشخص')[:200]}"
    _log(request, "تست اتصال درگاه پرداخت", "درگاه‌های پرداخت", f"{code}: {'ok' if res.get('ok') else 'fail'}", admin_info=adm)
    return _redir(f"/admin/payment-gateways?flash={e(msg)}")
